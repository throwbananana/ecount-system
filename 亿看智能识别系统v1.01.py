# -*- coding: utf-8 -*-
"""
一般凭证 Excel 模板转换工具（带 GUI + 自动识别列匹配）

功能：
1. 读取同目录下的 Template.xlsx（一般凭证模板），自动获取表头和批注。
2. 选择任意原始 Excel 文件 + 工作表。
3. 【自动识别匹配列】：
   - 表头规范化（去空格/特殊符号/大小写）
   - 包含/被包含判断
   - difflib 相似度打分
   - 同义词映射（如“科目代码”≈“科目编码”）
4. 用户可以在 GUI 中微调映射。
5. 根据每个模板表头的规则（日期、金额、小数位、长度等）进行转换。
6. 将转换结果填入模板，保留原有表头与批注。

依赖：
    pip install pandas openpyxl
"""

import os
import math
import difflib
import re
import json
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from typing import Optional, List, Dict

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from treeview_tools import attach_treeview_tools

from base_data_manager import BaseDataManager
from summary_intelligence import SummaryIntelligence, DEFAULT_API_KEY

TEMPLATE_FILE = "Template.xlsx"

# ======== 字段规则：根据模板表头控制格式 ========
FIELD_RULES = {
    "凭证日期": {"type": "date"},
    "日期": {"type": "date"},
    "序号": {"type": "text", "max_len": 4},
    "会计凭证No.": {"type": "text", "max_len": 30},
    "摘要编码": {"type": "text", "max_len": 2},
    "摘要": {"type": "text", "max_len": 200},
    "类型": {"type": "text", "max_len": 1},
    "科目编码": {"type": "text", "max_len": 8},
    "往来单位编码": {"type": "text", "max_len": 30},
    "往来单位名": {"type": "text", "max_len": 100},
    "金额": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
    "外币金额": {"type": "number", "max_int_len": 15, "max_decimal_len": 4},
    "汇率": {"type": "number", "max_int_len": 14, "max_decimal_len": 4},
    "部门": {"type": "text", "max_len": 50},
    # 销售出库模板新增字段
    "职员": {"type": "text", "max_len": 50},
    "发货仓库": {"type": "text", "max_len": 50},
    "交易类型": {"type": "text", "max_len": 2},
    "货币": {"type": "text", "max_len": 10},
    "收货公司": {"type": "text", "max_len": 200},
    "PEDIDO No": {"type": "text", "max_len": 200},
    "No.FACTURA": {"type": "text", "max_len": 200},
    "CONSIGNADO": {"type": "text", "max_len": 200},
    "CONDI. PAGO": {"type": "text", "max_len": 200},
    "CON DESTINO A": {"type": "text", "max_len": 200},
    "EMBARCADO VIA": {"type": "text", "max_len": 200},
    "MARCA": {"type": "text", "max_len": 200},
    "CARGOS": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
    "品目编码": {"type": "text", "max_len": 20},
    "品目名": {"type": "text", "max_len": 100},
    "规格": {"type": "text", "max_len": 100},
    "Mo.TAMA": {"type": "text", "max_len": 200},
    "COLOR": {"type": "text", "max_len": 200},
    "装数": {"type": "number", "max_int_len": 15, "max_decimal_len": 0},
    "CAJA": {"type": "number", "max_int_len": 12, "max_decimal_len": 0},
    "数量": {"type": "number", "max_int_len": 12, "max_decimal_len": 0},
    "单价": {"type": "number", "max_int_len": 12, "max_decimal_len": 6},
    "金额1": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
    "供应价": {"type": "number", "max_int_len": 12, "max_decimal_len": 2},
    "附带费用": {"type": "number", "max_int_len": 12, "max_decimal_len": 2},
    "增值税": {"type": "number", "max_int_len": 12, "max_decimal_len": 2},
    "生成生产入库": {"type": "text", "max_len": 1},
    "体积": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
    "总体积": {"type": "number", "max_int_len": 15, "max_decimal_len": 1},
    "重量": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
    "总重量": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
}

# ======== 同义词（自动识别用，可以按需要继续补充） ========
# key 为模板字段，value 为一组在原始表中可能出现的叫法
FIELD_SYNONYMS = {
    "凭证日期": ["日期", "记账日期", "制单日期"],
    "日期": ["日期", "fecha", "fecha日期", "日期fecha", "日期 fecha", "fecha 日期"],
    "序号": ["行号", "行次", "序号", "订单号", "单据号", "no.pedi", "pedido", "No.PEDI."],
    "会计凭证No.": ["凭证号", "凭证编号", "会计凭证号"],
    "摘要编码": ["摘要代码", "摘要编码"],
    "摘要": ["摘要", "摘要说明", "说明", "备注", "descrip", "内容", "descripción", "DESCRIPCION"],
    "类型": ["借贷标志", "借贷", "方向", "类型"],
    "科目编码": ["科目代码", "科目编号", "会计科目", "科目"],
    "往来单位编码": ["往来单位代码", "客户编码", "供应商编码", "往来编码", "客戶 cliente", "cliente", "客户"],
    "往来单位名": ["往来单位名称", "客户名称", "供应商名称", "单位名称", "cliente", "客戶"],
    "金额": ["金额", "本币金额", "发生额"],
    "外币金额": ["外币金额", "原币金额", "合计", "total", "合计 total"],
    "汇率": ["汇率", "折算汇率"],
    "部门": ["部门", "部门名称", "成本中心"],
    # 销售出库模板补充
    "发货仓库": ["发货仓库", "出货仓库", "仓库"],
    "品目编码": ["品目编码", "物料编码", "商品编码", "代号", "代號", "codig", "代號 codigo", "代號 codig"],
    "品目名": ["品目名", "品目名称", "商品名称", "内容", "descrip", "descripcion", "descripción"],
    "数量": ["数量", "数量 canti", "canti", "qty"],
    "单价": ["单价", "价格", "precio", "价格 precio"],
    "供应价": ["供应价", "合计", "total", "合计 total", "importe"],
    "金额1": ["金额1", "金额", "合计", "total"],
    "交易类型": ["交易类型", "交易编码", "单据类型"],
    "货币": ["货币", "币别", "currency", "货币代码"],
    "收货公司": ["收货公司", "收货方", "cliente", "客户"],
    "PEDIDO No": ["pedido", "No.PEDI.", "订单号", "pedido no"],
    "No.FACTURA": ["factura", "发票号"],
    "CONSIGNADO": ["consignado"],
    "CONDI. PAGO": ["支付条件", "付款条件", "pago"],
    "CON DESTINO A": ["destino"],
    "EMBARCADO VIA": ["embarcado", "via"],
    "MARCA": ["marca", "品牌"],
    "装数": ["装数", "装箱数"],
    "CAJA": ["caja", "箱数"],
    "附带费用": ["附加费用", "费用", "运费"],
    "增值税": ["增值税", "税额", "营业税"],
}

# 转换模式
MODE_GENERAL_VOUCHER = "通用凭证模式"
MODE_SALES_OUTBOUND = "销售出库模式"

# 针对销售出库格式的快捷映射候选（用于缺省映射兜底）
SALES_OUTBOUND_MAPPING_CANDIDATES = {
    "日期": ["日期 FECHA", "fecha", "日期"],
    "序号": ["No.PEDI.", "pedido", "订单号", "序号"],
    "PEDIDO No": ["No.PEDI.", "pedido", "订单号"],
    "往来单位编码": ["客戶 CLIENTE", "cliente", "客户编码", "客户"],
    "往来单位名": ["客戶 CLIENTE", "cliente", "客户名称"],
    "收货公司": ["客戶 CLIENTE", "cliente"],
    "品目编码": ["代號 CODIGO", "codig", "代号", "品目编码"],
    "品目名": ["内容 DESCRIPCION", "descrip", "内容", "商品名称"],
    "数量": ["数量 CANTI", "canti", "数量"],
    "单价": ["价格 PRECIO", "precio", "价格", "单价"],
    "供应价": ["合计 TOTAL", "total", "合计"],
    "外币金额": ["合计 TOTAL", "total", "合计"],
    "金额1": ["合计 TOTAL", "total", "合计"],
    "摘要": ["内容 DESCRIPCION", "descrip", "摘要"],
}


class TemplateHeader:
    def __init__(self, name, col_idx, comment=""):
        self.name = name
        self.col_idx = col_idx
        self.comment = comment or ""


def load_template_headers(template_path=TEMPLATE_FILE):
    """从模板文件中读取表头和批注。"""
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"未找到模板文件：{template_path}")

    wb = load_workbook(template_path, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(
                TemplateHeader(
                    str(cell.value).strip(),
                    cell.column,
                    cell.comment.text if cell.comment else "",
                )
            )
    return headers, wb, ws


def safe_str(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def format_date(value):
    """转成 YYYYMMDD，空则用今天。"""
    if value is None or (isinstance(value, float) and math.isnan(value)) or safe_str(value) == "":
        return datetime.today().strftime("%Y%m%d")

    if isinstance(value, (datetime, date)):
        return value.strftime("%Y%m%d")

    s = safe_str(value)

    # 增加 %d/%m/%Y 格式支持 (例如 01/11/2025)
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%Y%m%d")
        except ValueError:
            continue

    # Excel 序列号
    try:
        serial = float(s)
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=serial)
        return dt.strftime("%Y%m%d")
    except Exception:
        pass

    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) >= 8:
        return digits[:8]

    return datetime.today().strftime("%Y%m%d")


def format_number(value, max_int_len=15, max_decimal_len=2):
    """根据整数位/小数位限制格式化数字。"""
    if value is None or (isinstance(value, float) and math.isnan(value)) or safe_str(value) == "":
        return ""

    s = safe_str(value).replace(",", "")

    try:
        d = Decimal(s)
    except InvalidOperation:
        return s

    if max_decimal_len > 0:
        quantize_str = "1." + ("0" * max_decimal_len)
        d = d.quantize(Decimal(quantize_str))
    else:
        d = d.quantize(Decimal("1"))

    s_full = format(d, "f")
    if "." in s_full:
        int_part, dec_part = s_full.split(".")
    else:
        int_part, dec_part = s_full, ""

    negative = int_part.startswith("-")
    int_digits = int_part.lstrip("-")

    if len(int_digits) > max_int_len:
        int_digits = int_digits[-max_int_len:]

    int_part = ("-" if negative else "") + int_digits

    if max_decimal_len == 0:
        return int_part

    if len(dec_part) > max_decimal_len:
        dec_part = dec_part[:max_decimal_len]

    return int_part + "." + dec_part if dec_part else int_part


def apply_text_length(value, max_len=None):
    s = safe_str(value)
    if max_len and len(s) > max_len:
        return s[:max_len]
    return s


def convert_value(header_name, src_value):
    """根据字段规则把原始值转换成模板要求格式。"""
    rule = FIELD_RULES.get(header_name, {"type": "text"})
    t = rule.get("type", "text")

    if t == "date":
        return format_date(src_value)
    if t == "number":
        return format_number(
            src_value,
            max_int_len=rule.get("max_int_len", 15),
            max_decimal_len=rule.get("max_decimal_len", 2),
        )
    return apply_text_length(src_value, rule.get("max_len"))


# ========== 自动匹配相关工具函数 ==========

EMPTY_OPTION = "<留空>"


def normalize_header(s: str) -> str:
    """表头规范化：去空格、符号、大小写、全角半角等。"""
    if s is None:
        return ""
    s = str(s)
    # 全角转半角
    def _dbc_to_sbc(ch):
        code = ord(ch)
        if code == 0x3000:
            return " "
        if 0xFF01 <= code <= 0xFF5E:
            return chr(code - 0xFEE0)
        return ch

    s = "".join(_dbc_to_sbc(ch) for ch in s)
    # 去掉常见的分隔符号
    s = re.sub(r"[\s\-\_/（）\(\)【】\[\]：:，,\.。]", "", s)
    return s.lower()


def score_similarity(template_name: str, src_name: str, template_key: str) -> float:
    """
    给某个模板表头 & 源表头 的匹配打分：
    1. 规范名是否完全相等
    2. 包含关系
    3. 同义词命中
    4. difflib 模糊匹配
    综合出一个 0~1 分数。
    """
    raw_t = template_name
    raw_s = src_name
    n_t = normalize_header(raw_t)
    n_s = normalize_header(raw_s)

    # 完全相等
    if n_t and n_t == n_s:
        return 1.0

    score = 0.0

    # 包含关系
    if n_t and n_s:
        if n_t in n_s or n_s in n_t:
            score = max(score, 0.85)

    # 同义词匹配
    syns = FIELD_SYNONYMS.get(template_key, [])
    for syn in syns:
        n_syn = normalize_header(syn)
        if n_syn and (n_syn == n_s or n_syn in n_s or n_s in n_syn):
            score = max(score, 0.9)

    # difflib 相似度
    ratio = difflib.SequenceMatcher(None, n_t, n_s).ratio()
    score = max(score, ratio * 0.9)  # 稍微降一点权重

    return score


class ExcelConverterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("一般凭证 Excel 格式转换工具（含自动识别列匹配）")

        self.template_headers = []
        self.template_workbook = None
        self.template_sheet = None

        self.input_path = ""
        self.input_df = None
        self.input_columns = []
        self.mapping_vars = {}

        self.sheet_var = tk.StringVar()
        self.input_path_var = tk.StringVar()
        self.convert_mode_var = tk.StringVar(value=MODE_SALES_OUTBOUND)
        self.search_field_var = tk.StringVar(value="")
        self.current_columns = []
        self.sort_states = {}
        
        # 基础数据管理器 (先初始化，以便读取配置)
        self.base_data_mgr = None
        self._init_base_data()

        # 读取持久化配置
        db_configs = {}
        if self.base_data_mgr:
            db_configs = self.base_data_mgr.get_all_configs()

        # AI 设置 (优先使用数据库中的 Key)
        self.api_key = db_configs.get("api_key", DEFAULT_API_KEY)

        # 默认值 (从数据库加载)
        # 过滤掉 api_key，剩下的视为 default_values
        self.default_values = {k: v for k, v in db_configs.items() if k != "api_key"}

        # 摘要智能识别器
        self.summary_recognizer = None
        self._init_summary_recognizer()

        self._build_ui()
        self._load_template()
        self._on_mode_changed()

    def _show_ai_settings(self):
        """显示 AI 设置对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("AI 智能识别设置")
        dialog.geometry("500x200")
        
        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill="both", expand=True)
        
        ttk.Label(frame, text="智谱 AI API Key:").pack(anchor="w", pady=(0, 5))
        
        key_var = tk.StringVar(value=self.api_key)
        entry = ttk.Entry(frame, textvariable=key_var, width=60)
        entry.pack(fill="x", pady=(0, 10))
        
        ttk.Label(frame, text="提示：启用 AI 识别需要联网，可能会产生少量费用。\n默认使用免费或试用 Key。", foreground="gray").pack(anchor="w")
        
        def save():
            new_key = key_var.get().strip()
            if new_key:
                self.api_key = new_key
                # 持久化保存 Key
                if self.base_data_mgr:
                    self.base_data_mgr.set_config("api_key", new_key)
                
                # 更新识别器中的 Key
                if self.summary_recognizer:
                    self.summary_recognizer.update_api_key(new_key)
                messagebox.showinfo("成功", "API Key 已更新")
                dialog.destroy()
            else:
                messagebox.showwarning("提示", "API Key 不能为空")
        
        ttk.Button(frame, text="保存", command=save).pack(pady=10)

    def _init_base_data(self):
        """初始化基础数据"""
        try:
            self.base_data_mgr = BaseDataManager()
            stats = self.base_data_mgr.get_statistics()
            total_records = sum(stats.values())
            if total_records == 0:
                # 数据库为空，尝试自动导入
                result = self.base_data_mgr.import_all_data()
                if result["success"]:
                    print(f"基础数据自动导入成功: {result['message']}")
                else:
                    print(f"基础数据自动导入提示: {result['message']}")
        except Exception as e:
            print(f"初始化基础数据失败: {e}")
            self.base_data_mgr = None

    def _init_summary_recognizer(self):
        """初始化摘要智能识别器"""
        try:
            if self.base_data_mgr:
                self.summary_recognizer = SummaryIntelligence(self.base_data_mgr, self.default_values)
                print("摘要智能识别器初始化成功")
            else:
                print("摘要智能识别器初始化失败：基础数据管理器未就绪")
        except Exception as e:
            print(f"初始化摘要识别器失败: {e}")
            self.summary_recognizer = None

    def _update_recognizer_defaults(self):
        """更新识别器的默认值"""
        if self.summary_recognizer:
            self.summary_recognizer.default_values = self.default_values

    def _build_ui(self):
        # 菜单栏
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # 文件菜单
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="文件", menu=file_menu)
        file_menu.add_command(label="导入基础数据", command=self.import_base_data)
        file_menu.add_command(label="查看统计信息", command=self.show_base_data_stats)
        file_menu.add_separator()
        file_menu.add_command(label="AI 设置", command=self._show_ai_settings)  # 新增 AI 设置
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.root.quit)

        # 创建标签页控件
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)

        # 标签页1: Excel转换
        self._build_excel_converter_tab()

        # 标签页2: 基础数据管理
        self._build_base_data_tab()

    def _build_excel_converter_tab(self):
        """构建Excel转换标签页"""
        excel_frame = ttk.Frame(self.notebook)
        self.notebook.add(excel_frame, text="Excel凭证转换")

        # 顶部：文件选择
        top = ttk.Frame(excel_frame, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="原始 Excel：").grid(row=0, column=0, sticky="e")
        entry = ttk.Entry(top, textvariable=self.input_path_var, width=50)
        entry.grid(row=0, column=1, sticky="we", padx=5)
        ttk.Button(top, text="浏览...", command=self.select_input_file).grid(
            row=0, column=2, padx=5
        )

        ttk.Label(top, text="工作表：").grid(row=1, column=0, sticky="e", pady=(5, 0))
        self.sheet_combo = ttk.Combobox(top, textvariable=self.sheet_var, state="disabled", width=20)
        self.sheet_combo.grid(row=1, column=1, sticky="w", padx=5, pady=(5, 0))
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_changed)

        # 转换模式
        ttk.Label(top, text="转换模式：").grid(row=2, column=0, sticky="e", pady=(5, 0))
        self.mode_combo = ttk.Combobox(
            top,
            textvariable=self.convert_mode_var,
            values=[MODE_GENERAL_VOUCHER, MODE_SALES_OUTBOUND],
            state="readonly",
            width=20
        )
        self.mode_combo.grid(row=2, column=1, sticky="w", padx=5, pady=(5, 0))
        self.mode_combo.bind("<<ComboboxSelected>>", self._on_mode_changed)

        # 智能识别选项
        self.enable_smart_recognition = tk.BooleanVar(value=True)
        smart_check = ttk.Checkbutton(
            top,
            text="启用摘要智能识别（自动填充字段）",
            variable=self.enable_smart_recognition
        )
        smart_check.grid(row=3, column=1, sticky="w", padx=5, pady=(5, 0))

        # AI 深度识别选项 (新增)
        self.use_ai_var = tk.BooleanVar(value=False)
        ai_check = ttk.Checkbutton(
            top,
            text="启用 AI 深度识别 (科目判断)",
            variable=self.use_ai_var
        )
        ai_check.grid(row=3, column=1, sticky="w", padx=(250, 5), pady=(5, 0)) # 放在同一行右侧

        # 外币模式选项 (新增)
        self.use_foreign_currency_var = tk.BooleanVar(value=False)
        foreign_check = ttk.Checkbutton(
            top,
            text="启用外币模式 (自动计算本币)",
            variable=self.use_foreign_currency_var
        )
        foreign_check.grid(row=4, column=1, sticky="w", padx=(250, 5), pady=(2, 0)) # 放在下一行右侧

        # 自动生成对方分录选项
        self.auto_balance_var = tk.BooleanVar(value=False)
        balance_check = ttk.Checkbutton(
            top,
            text="自动生成对方分录（借贷平衡）",
            variable=self.auto_balance_var
        )
        balance_check.grid(row=4, column=1, sticky="w", padx=5, pady=(2, 0))

        # 提示信息
        ttk.Label(top, text="注：对方科目和往来将使用'基础数据管理'中设置的默认值", foreground="gray").grid(row=5, column=1, sticky="w", padx=5)

        # 默认值设置按钮
        ttk.Button(
            top,
            text="设置默认值...",
            command=self._show_default_values_dialog
        ).grid(row=3, column=2, padx=5, pady=(5, 0))

        top.grid_columnconfigure(1, weight=1)

        # 中部：字段映射
        mid = ttk.LabelFrame(excel_frame, text="字段映射（模板列 -> 原始列）", padding=10)
        mid.pack(fill="both", expand=True, padx=10, pady=5)
        # 可滚动区域，避免表头过多时看不到后续行
        self.mapping_canvas = tk.Canvas(mid, highlightthickness=0)
        mapping_scrollbar = ttk.Scrollbar(mid, orient="vertical", command=self.mapping_canvas.yview)
        mapping_scrollbar_x = ttk.Scrollbar(mid, orient="horizontal", command=self.mapping_canvas.xview)
        self.mapping_container = ttk.Frame(self.mapping_canvas)
        self.mapping_container.bind(
            "<Configure>",
            lambda e: self.mapping_canvas.configure(scrollregion=self.mapping_canvas.bbox("all"))
        )
        self.mapping_canvas.create_window((0, 0), window=self.mapping_container, anchor="nw")
        self.mapping_canvas.configure(yscrollcommand=mapping_scrollbar.set, xscrollcommand=mapping_scrollbar_x.set)
        self.mapping_canvas.pack(side="left", fill="both", expand=True)
        mapping_scrollbar.pack(side="right", fill="y")
        mapping_scrollbar_x.pack(side="bottom", fill="x")
        self.mapping_frame = self.mapping_container

        # 底部：操作按钮
        bottom = ttk.Frame(excel_frame, padding=10)
        bottom.pack(fill="x")

        ttk.Button(bottom, text="自动识别匹配", command=self.auto_match).pack(side="left")
        ttk.Button(bottom, text="开始转换并导出", command=self.do_convert).pack(side="right")

    def _build_base_data_tab(self):
        """构建基础数据管理标签页"""
        base_frame = ttk.Frame(self.notebook)
        self.notebook.add(base_frame, text="基础数据管理")

        # 左侧：数据类型列表
        left_frame = ttk.Frame(base_frame, width=200)
        left_frame.pack(side="left", fill="y", padx=5, pady=5)

        ttk.Label(left_frame, text="数据类型", font=("", 10, "bold")).pack(pady=5)

        # 数据类型按钮
        self.data_types = {
            "币种": "currency",
            "部门": "department",
            "仓库": "warehouse",
            "科目编码": "account_subject",
            "品目信息": "product",
            "往来单位": "business_partner",
            "账户": "bank_account",
            "智能识别缓存": "smart_recognition_cache" # 新增缓存管理入口
        }

        self.current_table = tk.StringVar(value="currency")

        for label, table_name in self.data_types.items():
            btn = ttk.Radiobutton(
                left_frame,
                text=label,
                variable=self.current_table,
                value=table_name,
                command=self._load_base_data_table
            )
            btn.pack(fill="x", padx=5, pady=2)

        # 右侧：数据显示和编辑区域
        right_frame = ttk.Frame(base_frame)
        right_frame.pack(side="right", fill="both", expand=True, padx=5, pady=5)

        # 顶部：搜索和操作按钮
        top_bar = ttk.Frame(right_frame)
        top_bar.pack(fill="x", pady=(0, 5))

        ttk.Label(top_bar, text="搜索:").pack(side="left", padx=5)
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(top_bar, textvariable=self.search_var, width=30)
        self.search_entry.pack(side="left", padx=5)
        ttk.Label(top_bar, text="字段:").pack(side="left")
        self.search_field_combo = ttk.Combobox(top_bar, textvariable=self.search_field_var, width=12, state="readonly")
        self.search_field_combo.pack(side="left", padx=(0, 5))
        ttk.Button(top_bar, text="搜索", command=self._search_base_data).pack(side="left", padx=2)
        ttk.Button(top_bar, text="显示全部", command=self._load_base_data_table).pack(side="left", padx=2)

        # 编辑按钮
        ttk.Button(top_bar, text="新增", command=self._add_base_data_record).pack(side="right", padx=2)
        ttk.Button(top_bar, text="编辑", command=self._edit_base_data_record).pack(side="right", padx=2)
        ttk.Button(top_bar, text="删除", command=self._delete_base_data_record).pack(side="right", padx=2)
        ttk.Button(top_bar, text="设置为默认值", command=self._set_as_default_base_data).pack(side="right", padx=2)
        ttk.Button(top_bar, text="默认往来单位", command=self._set_default_partner_via_subject).pack(side="right", padx=2)

        # 中部：数据表格
        table_frame = ttk.Frame(right_frame)
        table_frame.pack(fill="both", expand=True)

        # 滚动条
        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical")
        scrollbar_y.pack(side="right", fill="y")

        scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal")
        scrollbar_x.pack(side="bottom", fill="x")

        # 树形视图
        self.base_data_tree = ttk.Treeview(
            table_frame,
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set,
            selectmode="extended"  # 启用多选
        )
        self.base_data_tree.pack(fill="both", expand=True)
        self.base_data_tree_tools = attach_treeview_tools(self.base_data_tree)

        scrollbar_y.config(command=self.base_data_tree.yview)
        scrollbar_x.config(command=self.base_data_tree.xview)

        # 双击编辑 (支持单元格直接编辑)
        self.base_data_tree.bind("<Double-1>", self._on_tree_double_click)

        # 底部：状态栏
        self.status_label = ttk.Label(right_frame, text="就绪", relief="sunken")
        self.status_label.pack(fill="x", pady=(5, 0))

        # 初始化加载第一个表
        self._load_base_data_table()

    def _on_tree_double_click(self, event):
        """处理双击事件：尝试单元格编辑，否则弹出对话框"""
        region = self.base_data_tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = self.base_data_tree.identify_column(event.x)
        item_id = self.base_data_tree.identify_row(event.y)
        
        if not item_id:
            return

        table_name = self.current_table.get()
        
        # 获取列索引 (例如 #1, #2...)
        col_index = int(column.replace("#", "")) - 1
        columns = self.base_data_tree["columns"]
        
        # 防止索引越界（虽然理论上不应该）
        if col_index >= len(columns):
            return
            
        col_name = columns[col_index]

        # 1. 禁止编辑 ID 列 (通常是第0列)
        if col_index == 0 or col_name == "id":
            return

        # 2. 智能识别缓存表：禁止编辑摘要和创建时间，只允许编辑科目编码
        if table_name == "smart_recognition_cache":
            if col_name == "account_code":
                self._show_entry_editor(item_id, col_index, col_name)
            else:
                # 其他列(摘要)太长或只读，弹窗显示更合适，或者直接忽略
                self._edit_base_data_record()
            return

        # 3. 其他所有基础数据表：允许直接编辑所有非ID列
        self._show_entry_editor(item_id, col_index, col_name)

    def _show_entry_editor(self, item_id, col_index, col_name):
        """在Treeview单元格上显示输入框"""
        # 获取单元格坐标
        bbox = self.base_data_tree.bbox(item_id, column=col_name)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # 获取当前值
        current_values = self.base_data_tree.item(item_id, "values")
        current_val = current_values[col_index]

        # 创建输入框
        entry = ttk.Entry(self.base_data_tree, width=width)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, current_val)
        entry.select_range(0, tk.END)
        entry.focus()

        def save_edit(event=None):
            new_val = entry.get().strip()
            # 销毁输入框
            entry.destroy()
            
            # 如果值没变，不进行数据库操作
            if new_val == str(current_val):
                return

            table_name = self.current_table.get()
            record_id = int(current_values[0]) # 假设第一列总是ID
            
            result = None
            
            # 分发更新逻辑
            if table_name == "smart_recognition_cache":
                if col_name == "account_code":
                    result = self.base_data_mgr.update_cached_recognition(record_id, new_val)
            else:
                # 通用表更新
                # 注意：对于数字类型的列，可能需要转换类型，这里暂且都当字符串存，sqlite是弱类型的，通常没问题
                # 但为了严谨，如果base_data_manager有类型转换逻辑更好。
                # 这里为了通用性，直接传递。
                data = {col_name: new_val}
                result = self.base_data_mgr.update_record(table_name, record_id, data)
            
            if result and result["success"]:
                # 更新界面显示
                new_values = list(current_values)
                new_values[col_index] = new_val
                self.base_data_tree.item(item_id, values=new_values)
                self.status_label.config(text=f"更新成功: ID {record_id} [{col_name}] -> {new_val}")
            else:
                msg = result["message"] if result else "未知错误"
                messagebox.showerror("错误", msg)

        def cancel_edit(event=None):
            entry.destroy()

        # 绑定事件
        entry.bind("<Return>", save_edit)
        # entry.bind("<FocusOut>", lambda e: entry.destroy()) # 失去焦点暂不保存
        # 为了体验更像Excel，失去焦点通常意味着保存。但为了防止误触，我们这里设定：
        # 如果点击了表格其他地方，会触发Treeview的点击，导致FocusOut。
        # 暂且设为：失去焦点 = 取消 (您之前似乎倾向于安全)
        # 如果您希望失去焦点自动保存，请告诉我。
        entry.bind("<FocusOut>", lambda e: entry.destroy()) 
        entry.bind("<Escape>", cancel_edit)

    def _load_template(self):
        try:
            headers, wb, ws = load_template_headers(TEMPLATE_FILE)
        except FileNotFoundError as e:
            messagebox.showerror("错误", f"读取模板失败：\n{e}\n\n请确认 {TEMPLATE_FILE} 与本脚本在同一目录。")
            return
        except Exception as e:
            messagebox.showerror("错误", f"读取模板失败：\n{e}")
            return

        self.template_headers = headers
        self.template_workbook = wb
        self.template_sheet = ws

    def _on_mode_changed(self, event=None):
        """切换转换模式"""
        mode = self.convert_mode_var.get()
        base_title = "一般凭证 Excel 格式转换工具（含自动识别列匹配）"
        if mode == MODE_SALES_OUTBOUND:
            base_title += " - 销售出库模式"
        self.root.title(base_title)

        # 重新加载模板（当前两个模式共用同一模板文件，但预留扩展）
        if event is not None:  # 用户手动切换时才重新加载，初始化阶段避免重复读取
            self._load_template()
        # 切换模式后重建映射区域，确保下拉框重置
        if self.input_df is not None:
            self._create_mapping_widgets()

    def select_input_file(self):
        path = filedialog.askopenfilename(
            title="选择原始 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx;*.xls"), ("所有文件", "*.*")],
        )
        if not path:
            return

        self.input_path = path
        self.input_path_var.set(path)

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
        except Exception as e:
            messagebox.showerror("错误", f"读取工作表列表失败：\n{e}")
            return

        self.sheet_combo["values"] = sheet_names
        self.sheet_combo.configure(state="readonly")
        if sheet_names:
            self.sheet_var.set(sheet_names[0])
            self._load_input_sheet()

    def _on_sheet_changed(self, event=None):
        self._load_input_sheet()

    def _load_input_sheet(self):
        if not self.input_path or not self.sheet_var.get():
            return
        try:
            self.input_df = pd.read_excel(self.input_path, sheet_name=self.sheet_var.get())
        except Exception as e:
            messagebox.showerror("错误", f"读取 Excel 数据失败：\n{e}")
            return

        self.input_columns = [str(c) for c in self.input_df.columns]
        self._create_mapping_widgets()

    def _create_mapping_widgets(self):
        for w in self.mapping_frame.winfo_children():
            w.destroy()

        ttk.Label(self.mapping_frame, text="模板列名", width=20).grid(
            row=0, column=0, padx=3, pady=3, sticky="w"
        )
        ttk.Label(self.mapping_frame, text="来源列（原始 Excel）", width=30).grid(
            row=0, column=1, padx=3, pady=3, sticky="w"
        )
        ttk.Label(self.mapping_frame, text="批注说明", width=10).grid(
            row=0, column=2, padx=3, pady=3, sticky="w"
        )

        self.mapping_vars = {}

        row_idx = 1
        for header in self.template_headers:
            ttk.Label(self.mapping_frame, text=header.name).grid(
                row=row_idx, column=0, padx=3, pady=3, sticky="w"
            )

            var = tk.StringVar(value=EMPTY_OPTION)
            combo_values = [EMPTY_OPTION] + self.input_columns
            combo = ttk.Combobox(
                self.mapping_frame,
                textvariable=var,
                values=combo_values,
                state="readonly",
                width=30,
            )
            combo.grid(row=row_idx, column=1, padx=3, pady=3, sticky="w")

            self.mapping_vars[header.name] = var

            btn = ttk.Button(
                self.mapping_frame,
                text="查看",
                command=lambda h=header: self.show_comment(h),
                width=6,
            )
            btn.grid(row=row_idx, column=2, padx=3, pady=3, sticky="w")

            row_idx += 1

    # ---------- 自动识别匹配 ----------
    def auto_match(self):
        if not self.input_columns:
            messagebox.showwarning("提示", "请先选择原始 Excel 文件并加载工作表。")
            return

        # 预取少量示例值供 AI 参考
        sample_values = {}
        if self.input_df is not None and not self.input_df.empty:
            head = self.input_df.head(3)
            for col in self.input_columns:
                vals = head[col].dropna().astype(str).tolist()
                sample_values[col] = vals[:3]

        for header in self.template_headers:
            best_col = None
            best_score = 0.0

            for src_col in self.input_columns:
                score = score_similarity(header.name, src_col, header.name)
                if score > best_score:
                    best_score = score
                    best_col = src_col

            # 阈值可以根据实际情况微调
            if best_score >= 0.6 and best_col:
                self.mapping_vars[header.name].set(best_col)
            else:
                # 尝试 AI 辅助匹配（需勾选 AI 开关且客户端可用）
                ai_col = None
                if self.use_ai_var.get() and self.summary_recognizer and self.summary_recognizer.zhipu_client:
                    try:
                        ai_col = self._ai_match_column(header.name, self.input_columns, sample_values)
                    except Exception as e:
                        print(f"AI 列匹配失败: {e}")
                        ai_col = None

                if ai_col:
                    self.mapping_vars[header.name].set(ai_col)
                else:
                    self.mapping_vars[header.name].set(EMPTY_OPTION)

        messagebox.showinfo("自动识别完成", "已根据表头名称自动匹配列，请检查映射结果，如有需要可手动调整。")

    def _ai_match_column(self, template_header: str, src_columns: list, samples: dict) -> str:
        """调用 AI 辅助选择最合适的源列"""
        if not src_columns:
            return ""

        # 构建提示
        column_desc = []
        for col in src_columns:
            vals = samples.get(col, [])
            sample_str = " | ".join(vals) if vals else ""
            column_desc.append(f"{col}: {sample_str}")

        prompt = f"""你是表头映射助手。现在要把源表的列映射到模板字段。
模板字段: {template_header}
源表列及样例值:
{chr(10).join(column_desc)}
请选择最匹配模板字段的源表列名，只返回列名；如果不确定，返回 None。"""

        resp = self.summary_recognizer.zhipu_client.chat.completions.create(
            model="glm-4-flash",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
        )
        result = resp.choices[0].message.content.strip()
        # 直接返回匹配到的列名
        for col in src_columns:
            if col == result:
                return col
        # 简单包含匹配
        for col in src_columns:
            if result in col or col in result:
                return col
        return ""

    def show_comment(self, header: TemplateHeader):
        text = header.comment.strip() if header.comment else "该列暂无批注说明。"
        messagebox.showinfo(f"{header.name} 批注", text)

    def _show_default_values_dialog(self):
        """显示默认值设置对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("设置默认值")
        dialog.geometry("500x400")

        # 创建主框架
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill="both", expand=True)

        ttk.Label(
            main_frame,
            text="为未映射或未识别的字段设置默认值：",
            font=("", 10, "bold")
        ).pack(anchor="w", pady=(0, 10))

        # 创建滚动区域
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 常用字段的默认值设置
        default_fields = [
            ("部门", "部门编码或名称，例如: 10001"),
            ("科目编码", "默认科目编码，例如: 1002"),
            ("默认账户往来", "默认账户往来编码 (用于银行对账)"),
            ("往来单位编码", "通用默认往来单位 (低优先级)"),
            ("外币金额", "外币金额，例如: 0"),
            ("汇率", "汇率，例如: 1 或 6.5"),
            ("序号", "序号，例如: 1"),
            ("会计凭证No.", "会计凭证号"),
            ("发货仓库", "出库仓库编码或名称（销售出库）"),
            ("默认仓库", "备用默认仓库编码/名称（为空时回退到发货仓库）"),
            ("交易类型", "销售出库交易类型编码，例如: 11 或 12"),
            ("货币", "默认货币编码，例如: USD 或 CNY"),
            ("职员", "经办人编码或名称"),
            ("收货公司", "默认收货公司/客户名称"),
            ("生成生产入库", "Y/N，留空默认为 N"),
        ]

        entry_vars = {}
        for idx, (field_name, hint) in enumerate(default_fields):
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill="x", pady=5)

            ttk.Label(frame, text=f"{field_name}:", width=15).pack(side="left")

            var = tk.StringVar(value=self.default_values.get(field_name, ""))
            entry = ttk.Entry(frame, textvariable=var, width=30)
            entry.pack(side="left", padx=5)
            entry_vars[field_name] = var

            ttk.Label(frame, text=hint, foreground="gray", font=("", 8)).pack(side="left")

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 底部按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))

        def save_defaults():
            # 保存默认值
            for field_name, var in entry_vars.items():
                value = var.get().strip()
                if value:
                    self.default_values[field_name] = value
                    if self.base_data_mgr:
                        self.base_data_mgr.set_config(field_name, value)
                else:
                    # 移除空值
                    self.default_values.pop(field_name, None)
                    # 数据库中也应该移除（或者设为空），这里暂且设为空字符串覆盖
                    if self.base_data_mgr:
                        self.base_data_mgr.set_config(field_name, "")

            # 更新识别器
            self._update_recognizer_defaults()

            messagebox.showinfo("成功", f"已保存默认值设置\n共设置 {len(self.default_values)} 个默认值")
            dialog.destroy()

        def clear_defaults():
            # 清空所有默认值
            for var in entry_vars.values():
                var.set("")
            self.default_values.clear()
            self._update_recognizer_defaults()
            messagebox.showinfo("成功", "已清空所有默认值")

        ttk.Button(button_frame, text="保存", command=save_defaults, width=12).pack(side="right", padx=5)
        ttk.Button(button_frame, text="清空", command=clear_defaults, width=12).pack(side="right")
        ttk.Button(button_frame, text="取消", command=dialog.destroy, width=12).pack(side="right")

        # 使窗口模态
        dialog.transient(self.root)
        dialog.grab_set()

    # ---------- 转换与导出 ----------
    def do_convert(self):
        # 强制刷新配置（以防万一）
        if self.base_data_mgr:
            db_configs = self.base_data_mgr.get_all_configs()
            # 更新 self.default_values，保留 api_key 不动（虽然它也在 configs 里）
            for k, v in db_configs.items():
                if k != "api_key":
                    self.default_values[k] = v
        
        print(f"调试: 当前所有默认值 keys: {list(self.default_values.keys())}")
        print(f"调试: 科目编码={self.default_values.get('科目编码')}, 默认账户往来={self.default_values.get('默认账户往来')}")

        if self.input_df is None:
            messagebox.showwarning("提示", "请先选择原始 Excel 文件。")
            return
        if not self.template_headers:
            messagebox.showerror("错误", "未加载模板表头，无法转换。")
            return

        mapping = {}
        for header in self.template_headers:
            var = self.mapping_vars.get(header.name)
            if not var:
                continue
            sel = var.get()
            mapping[header.name] = None if sel == EMPTY_OPTION else sel

        # 销售出库模式走专用转换逻辑
        if self.convert_mode_var.get() == MODE_SALES_OUTBOUND:
            output_rows, recognition_info, unmatched_info = self._convert_sales_outbound(mapping)
            if not output_rows:
                messagebox.showwarning("提示", "原始 Excel 中没有数据行。")
                return
            need_preview = True  # 销售出库模式强制预处理预览
            if need_preview:
                preview_result = self._show_preprocessing_preview(output_rows, recognition_info, unmatched_info)
                if not preview_result:
                    return
            self._export_to_excel(output_rows, unmatched_info)
            return

        # 准备转换数据
        output_rows = []
        recognition_info = []  # 存储智能识别信息
        unmatched_info = []

        print(f"调试: 智能识别开关={self.enable_smart_recognition.get()}, AI开关={self.use_ai_var.get()}")

        # 序号生成器状态
        serial_map = {}
        next_serial_id = 1

        for idx, src_row in self.input_df.iterrows():
            # 准备行数据字典（用于智能识别）
            row_dict = {}
            for col_name in self.input_df.columns:
                row_dict[str(col_name)] = src_row[col_name]

            # 预先识别借/贷金额列（支持两个独立列）
            def _norm_header(name: str) -> str:
                return normalize_header(name)

            debit_syn = {"借", "借方", "借方金额", "借方本币", "借方金额(本币)", "debit", "dr"}
            credit_syn = {"贷", "贷方", "贷方金额", "贷方本币", "贷方金额(本币)", "credit", "cr"}

            debit_cols = [c for c in self.input_df.columns if _norm_header(c) in debit_syn]
            credit_cols = [c for c in self.input_df.columns if _norm_header(c) in credit_syn]

            def _first_non_empty(cols):
                for c in cols:
                    v = src_row.get(c)
                    if v is not None and not (isinstance(v, float) and pd.isna(v)) and str(v).strip() != "":
                        return v
                return None

            debit_val = _first_non_empty(debit_cols)
            credit_val = _first_non_empty(credit_cols)
            derived_amount = None
            derived_type = None
            if debit_val is not None:
                derived_amount = debit_val
                derived_type = "3"
            elif credit_val is not None:
                derived_amount = credit_val
                derived_type = "4"

            # 如果启用智能识别，尝试从摘要中提取信息
            smart_data = {}
            if self.enable_smart_recognition.get() and self.summary_recognizer:
                # 获取摘要内容
                summary_col = mapping.get("摘要")
                if summary_col:
                    summary_value = src_row[summary_col] if summary_col in src_row.index else None
                    if summary_value and not pd.isna(summary_value):
                        # 智能识别
                        # 传递 AI 识别开关和外币模式开关
                        use_ai = self.use_ai_var.get()
                        use_foreign = self.use_foreign_currency_var.get()
                        smart_data = self.summary_recognizer.recognize(str(summary_value), row_dict, use_ai=use_ai, use_foreign_currency=use_foreign)
                        # 记录识别信息
                        if smart_data:
                            recognition_info.append({
                                "row_num": idx + 2,  # Excel行号（从2开始）
                                "summary": str(summary_value),
                                "recognized": smart_data
                            })
        
            # 构建输出行
            out_row = []
            for header in self.template_headers:
                # 1. 获取智能识别结果（作为备选）
                smart_value = smart_data.get(header.name, None)
                
                # 2. 尝试获取手动映射的值
                src_col = mapping.get(header.name)
                src_value = None
                if src_col:
                    src_value = src_row[src_col]

                # 3. 【核心修复】空值回退逻辑
                # 判断 src_value 是否为"有效空"（None, NaN, 或清除空白后的空字符串）
                is_src_empty = (src_value is None) or \
                               (isinstance(src_value, float) and pd.isna(src_value)) or \
                               (str(src_value).strip() == "")
                
                # 如果手动映射为空，且有智能识别结果，则使用智能识别结果
                if is_src_empty and smart_value is not None:
                    src_value = smart_value
                # 如果既没有手动映射，也没有智能识别，但没有手动映射列（src_col is None），则直接用智能识别（哪怕是None）
                elif src_col is None:
                    src_value = smart_value

                # --- 特殊处理：序号自动生成 (解决4字限制，保持单据捆绑) ---
                if header.name == "序号":
                    # 如果有源值（通常是原始单号），则通过映射生成短序号
                    s_val = str(src_value).strip() if src_value is not None and not pd.isna(src_value) else ""
                    if s_val:
                        if s_val not in serial_map:
                            serial_map[s_val] = str(next_serial_id)
                            next_serial_id += 1
                        src_value = serial_map[s_val]
                    else:
                        # 如果没有源值（未映射列），则默认自动递增（假设每一行是一张新凭证，除非后续有合并逻辑）
                        # 或者是为了避免空值报错
                        src_value = str(next_serial_id)
                        next_serial_id += 1
                # -----------------------------------------------------

                # --- 特殊处理：类型自动转换 (借/贷 -> 3/4) ---
                # 预先判断并存储推断的类型，只在处理"类型"列时才赋值给src_value
                inferred_type_for_row = None
                s_val_for_type_inference = str(src_value).strip() if src_value is not None and not pd.isna(src_value) else ""

                # 1. 优先尝试从当前列值（关键字）推断
                if "借" in s_val_for_type_inference: inferred_type_for_row = "3"
                elif "贷" in s_val_for_type_inference: inferred_type_for_row = "4"
                elif "出" in s_val_for_type_inference: inferred_type_for_row = "1"
                elif "入" in s_val_for_type_inference: inferred_type_for_row = "2"
                elif derived_type:
                    inferred_type_for_row = derived_type
                
                # 2. 如果关键字推断失败，且当前列为空，则尝试从金额推断
                # (金额正数为借-3，负数为贷-4)
                if inferred_type_for_row is None:
                     # 无论当前列是否为空（只要没匹配到关键字），都尝试从金额反推
                     # 这样即使"类型"列填了奇怪的东西，也能纠正。但为了安全，还是限制在空值或无法识别的情况
                     
                     if s_val_for_type_inference == "" or (header.name == "类型"):
                        amt_col = mapping.get("金额")
                        raw_amt = None
                        
                        # (A) 从原始映射列获取金额
                        if amt_col:
                            raw_amt = src_row[amt_col]
                        
                        # (B) 如果没有，从智能识别结果获取金额
                        if (raw_amt is None or pd.isna(raw_amt) or str(raw_amt).strip() == "") and smart_data:
                            raw_amt = smart_data.get("金额")

                        # (C) 推断
                        if raw_amt is not None:
                            try:
                                f_amt = float(raw_amt)
                                if f_amt < 0:
                                    inferred_type_for_row = "4" # 贷
                                else:
                                    inferred_type_for_row = "3" # 借
                            except (ValueError, TypeError):
                                pass
                
                if header.name == "类型":
                    if inferred_type_for_row is not None:
                        src_value = inferred_type_for_row
                    elif smart_value is not None and is_src_empty:
                         # 允许使用智能识别的类型，只要它是合法的 3 或 4
                         s_smart = str(smart_value).strip()
                         if s_smart in ["3", "4", "借", "贷"]:
                             if s_smart == "借": src_value = "3"
                             elif s_smart == "贷": src_value = "4"
                             else: src_value = s_smart
                         else:
                             # 忽略 "1", "2" 等其他干扰
                             src_value = None 
                # ---------------------------------------------

                # --- 特殊处理：外币模式 (新增) ---
                # 如果启用外币模式，且当前列是金额/外币/汇率，我们需要进行联动计算
                # 逻辑：识别到的数值 -> 视为外币金额 -> * 默认汇率 -> 得到本币金额
                
                if self.use_foreign_currency_var.get():
                    # 获取默认汇率
                    default_rate_str = self.default_values.get("汇率", "1")
                    try:
                        default_rate = float(default_rate_str)
                    except:
                        default_rate = 1.0

                    # 找到原始识别到的"金额"数值（无论是从映射还是智能识别）
                    # 注意：这里的 src_value 是当前 header 的值。我们需要一个统一的地方处理金额逻辑。
                    # 最好的方式是在处理 "金额" 列时统一计算。
                    
                    if header.name == "金额":
                        # 获取原始数值（这通常是用户眼中的"交易金额"，在外币模式下即为外币）
                        raw_amount = src_value 

                        # 计算本币
                        local_amount = 0
                        if raw_amount:
                            try:
                                local_amount = float(raw_amount) * default_rate
                            except:
                                pass
                        src_value = local_amount # 将本币金额填入"金额"列
                    
                    elif header.name == "外币金额":
                        # 在外币模式下，"外币金额"列应该填入原始识别到的金额
                        # 我们需要重新获取一下原始金额（因为 src_value 此时可能是空的，或者我们想覆盖它）
                        
                        # 尝试从映射获取原始金额
                        amt_col = mapping.get("金额") # 注意：用户通常把Excel里的金额列映射给"金额"
                        raw_val = None
                        if amt_col:
                            raw_val = src_row[amt_col]
                        
                        # 回退到智能识别的金额
                        if (raw_val is None or pd.isna(raw_val) or str(raw_val).strip() == "") and smart_data:
                            raw_val = smart_data.get("金额")
                            
                        src_value = raw_val # 将原始金额填入"外币金额"列

                    elif header.name == "汇率":
                        # 强制填入默认汇率
                        src_value = default_rate
                # ---------------------------------------------

                # --- 特殊处理：金额取绝对值 ---
                if header.name in ["金额", "外币金额"]:
                     if src_value is not None:
                         try:
                             # 尝试转为浮点数取绝对值，再转回字符串给后续 format_number 处理
                             f_val = float(src_value)
                             src_value = abs(f_val)
                         except (ValueError, TypeError):
                             pass

                # --- 借贷分列兜底：如果金额列为空，但有借/贷分列，优先使用 ---
                if header.name == "金额" and (src_value is None or (isinstance(src_value, float) and pd.isna(src_value)) or str(src_value).strip() == ""):
                    if derived_amount is not None:
                        try:
                            src_value = abs(float(derived_amount))
                        except (ValueError, TypeError):
                            src_value = derived_amount

                if header.name == "类型" and (src_value is None or str(src_value).strip() == ""):
                    if derived_type:
                        src_value = derived_type
                # ---------------------------------------------

                converted = convert_value(header.name, src_value)
                out_row.append(converted)
            
            output_rows.append(out_row)

            # --- 自动生成对方分录 (借贷平衡) ---
            if self.auto_balance_var.get():
                # 获取当前行的关键信息
                current_type = ""
                current_amount = ""
                current_serial = ""
                current_subject = ""
                current_partner = ""
                
                # 需要找到对应列的索引
                header_map = {h.name: i for i, h in enumerate(self.template_headers)}
                
                if "类型" in header_map: current_type = out_row[header_map["类型"]]
                if "金额" in header_map: current_amount = out_row[header_map["金额"]]
                if "序号" in header_map: current_serial = out_row[header_map["序号"]]
                if "科目编码" in header_map: current_subject = str(out_row[header_map["科目编码"]]).strip()
                if "往来单位编码" in header_map: current_partner = str(out_row[header_map["往来单位编码"]]).strip()
                if "部门" in header_map: current_department = str(out_row[header_map["部门"]]).strip()

                # 只有当类型明确为3或4时才生成对方分录
                if current_type in ["3", "4"]:
                    balance_row = list(out_row) # 复制原行数据
                    partner_override = None
                    if self.base_data_mgr:
                        partner_override = self.base_data_mgr.get_partner_for_subject(current_subject)

                    # 1. 反转类型
                    new_type = "4" if current_type == "3" else "3"
                    if "类型" in header_map:
                        balance_row[header_map["类型"]] = new_type
                    
                    # 2. 设置对方科目
                    # 从默认值中获取：科目编码 -> 对方科目，默认账户往来 -> 对方往来
                    target_subject = str(self.default_values.get("科目编码", "")).strip()
                    target_partner = str(self.default_values.get("默认账户往来", "")).strip()
                    
                    # 【安全检查】如果默认科目与当前行的往来单位编码相同，说明用户可能在设置中填错了（把往来填到了科目里）
                    # 此时强制清空默认科目，避免生成的对方科目变成客户编码
                    if target_subject and target_subject == current_partner:
                        print(f"警告: 默认科目({target_subject})与当前往来({current_partner})相同，疑似配置错误，已忽略默认科目。")
                        target_subject = ""

                    print(f"调试: 默认对方科目={target_subject}, 默认对方往来={target_partner}") # 调试日志

                    final_subject = target_subject
                    final_partner = "" # 最终决定的对方往来
                    
                    # --- 智能推断逻辑 ---
                    # 场景：如果原分录是银行类科目且有往来单位，对方科目优先 1122，避免 1002/1002 平账
                    curr_sub_str = str(current_subject).strip().replace(".0", "")
                    tgt_sub_str = str(target_subject).strip().replace(".0", "")

                    is_bank = curr_sub_str.startswith("100")
                    if is_bank and current_partner:
                        final_subject = "1122"
                        final_partner = current_partner
                    elif curr_sub_str == tgt_sub_str:
                        final_subject = "" # 默认先清空，避免借贷同科目
                        if current_type == "3" and current_partner:
                            final_subject = "1122"
                            final_partner = current_partner
                    else:
                        if final_subject and target_partner:
                            final_partner = target_partner

                    # 科目级默认往来优先填充
                    if not final_partner and partner_override:
                        final_partner = partner_override

                    # 如果仍未确定对方科目，回退到默认科目或银行科目，确保不为空
                    if not final_subject:
                        final_subject = target_subject or "1002"
                        if not final_partner and target_partner:
                            final_partner = target_partner

                    # 填入最终决定的科目
                    if "科目编码" in header_map:
                        balance_row[header_map["科目编码"]] = final_subject
                        
                    # 3. 处理往来单位
                    if "往来单位编码" in header_map:
                        balance_row[header_map["往来单位编码"]] = final_partner
                    if "往来单位名" in header_map:
                        balance_row[header_map["往来单位名"]] = "" # 清空名称，让系统自动带出或避免混淆
                    
                    # 部门保持一致（不能空着）
                    if "部门" in header_map:
                        balance_row[header_map["部门"]] = current_department

                    output_rows.append(balance_row)
            # ---------------------------------------------

        if not output_rows:
            messagebox.showwarning("提示", "原始 Excel 中没有数据行。")
            return

        print(f"调试: 识别到 {len(recognition_info)} 行信息")

        # 显示预处理预览窗口
        need_preview = (
            self.convert_mode_var.get() == MODE_SALES_OUTBOUND
            or (self.enable_smart_recognition.get() and recognition_info)
            or bool(unmatched_info)
        )
        if need_preview:
            preview_result = self._show_preprocessing_preview(output_rows, recognition_info, unmatched_info)
            if not preview_result:
                return  # 用户取消了转换

        # 执行导出
        self._export_to_excel(output_rows, unmatched_info)

    def _build_sales_mapping(self, mapping: Dict[str, str]) -> Dict[str, str]:
        """根据候选列名为销售出库模式补全映射"""
        effective = dict(mapping)
        for target, candidates in SALES_OUTBOUND_MAPPING_CANDIDATES.items():
            if effective.get(target):
                continue
            # 直接命中同名列
            for col in self.input_columns:
                if normalize_header(col) == normalize_header(target):
                    effective[target] = col
                    break
            if effective.get(target):
                continue
            # 候选模糊匹配
            for col in self.input_columns:
                norm_col = normalize_header(col)
                for cand in candidates:
                    norm_cand = normalize_header(cand)
                    if norm_cand and (norm_cand == norm_col or norm_cand in norm_col or norm_col in norm_cand):
                        effective[target] = col
                        break
                if effective.get(target):
                    break
        return effective

    def _convert_sales_outbound(self, mapping):
        """销售出库导入模板转换"""
        output_rows = []
        recognition_info = []  # 销售出库暂不使用智能识别预览
        unmatched_info = []  # 未能匹配到基础资料的行
        effective_mapping = self._build_sales_mapping(mapping)

        # 预加载基础资料，便于模糊匹配
        product_rows = []
        partner_rows = []
        if self.base_data_mgr:
            try:
                product_rows = self.base_data_mgr.query("product")
            except Exception:
                product_rows = []
            try:
                partner_rows = self.base_data_mgr.query("business_partner")
            except Exception:
                partner_rows = []

        def _norm(val: str) -> str:
            if val is None:
                return ""
            # 去掉空格/横线/下划线，方便匹配 3WF-600-20L、3WF 600 20L 等
            return re.sub(r"[\s\-_]+", "", str(val)).lower()

        def _code_key(val: str) -> str:
            """生成忽略前导0的编码键，形如 L0021 -> l:21，BC520 -> bc:520"""
            if not val:
                return ""
            s = _norm(val)
            # 前缀字母 + 数字 + 可选尾巴
            m = re.match(r"([a-z]+)(0*)(\d+)(.*)", s)
            if m:
                prefix, _, num, tail = m.groups()
                return f"{prefix}:{num.lstrip('0') or '0'}{tail}"
            # 纯数字前缀（少见）
            m2 = re.match(r"(\d+)(.*)", s)
            if m2:
                num, tail = m2.groups()
                return f":{num.lstrip('0') or '0'}{tail}"
            return s

        def _fuzzy_find(rows, value):
            """在基础资料中模糊匹配 code/name"""
            if not value:
                return None
            n_val = _norm(value)
            if not n_val:
                return None
            key_val = _code_key(value)
            def _parse_match_items(raw):
                if not raw:
                    return []
                if isinstance(raw, list):
                    items = raw
                else:
                    try:
                        items = json.loads(raw)
                    except Exception:
                        items = []
                cleaned = []
                for i in items:
                    if i is None:
                        continue
                    s = str(i).strip()
                    if s:
                        cleaned.append(s)
                return cleaned
            # 1) 编码去0匹配 (如 L0021 vs L021)，优先返回编码长度更长的候选
            key_matches = [r for r in rows if _code_key(r.get("code")) == key_val]
            if key_matches:
                key_matches.sort(key=lambda x: len(str(x.get("code", ""))), reverse=True)
                return key_matches[0]
            # 2) 精确 code
            for r in rows:
                if _norm(r.get("code")) == n_val:
                    return r
            # 3) 精确 name
            for r in rows:
                if _norm(r.get("name")) == n_val:
                    return r
            # 4) 包含匹配
            for r in rows:
                # 4.0) 先尝试映射匹配项
                for alias in _parse_match_items(r.get("match_items")):
                    if n_val == _norm(alias) or n_val in _norm(alias) or _norm(alias) in n_val:
                        return r
                if n_val in _norm(r.get("code")) or _norm(r.get("code")) in n_val:
                    return r
                if n_val in _norm(r.get("name")) or _norm(r.get("name")) in n_val:
                    return r
            return None

        # 默认值
        default_rate = self.default_values.get("汇率", "1")
        try:
            default_rate_val = float(default_rate)
        except Exception:
            default_rate_val = 1.0
        default_currency = self.default_values.get("货币", "") or self.default_values.get("货币编码", "")
        default_department = self.default_values.get("部门", "")
        default_wh = self.default_values.get("发货仓库", "") or self.default_values.get("默认仓库", "")
        default_tran_type = self.default_values.get("交易类型", "11")
        default_staff = self.default_values.get("职员", "")
        default_receiver = self.default_values.get("收货公司", "")
        default_inbound_flag = self.default_values.get("生成生产入库", "N") or "N"

        def _is_empty(val):
            return val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == ""

        def pick_value(row, target_name, fallbacks=None):
            """按映射取值"""
            fallbacks = fallbacks or []
            mapped_col = effective_mapping.get(target_name)
            candidates = [mapped_col] + [effective_mapping.get(fb) for fb in fallbacks]
            for col in candidates:
                if not col:
                    continue
                if col in row:
                    val = row[col]
                    if val is not None and not (isinstance(val, float) and pd.isna(val)):
                        return val
            return None

        order_serial_map = {}
        next_order_serial = 1

        for idx, src_row in self.input_df.iterrows():
            unmatched_msgs = []
            unmatched_fields = []

            order_no = pick_value(src_row, "序号", fallbacks=["PEDIDO No"])
            if order_no is None:
                order_no = pick_value(src_row, "PEDIDO No")
            order_raw = str(order_no).strip() if order_no not in (None, "") else ""
            if order_raw:
                if order_raw not in order_serial_map:
                    order_serial_map[order_raw] = str(next_order_serial)
                    next_order_serial += 1
                order_serial = order_serial_map[order_raw]
            else:
                order_serial = str(next_order_serial)
                next_order_serial += 1

            customer_code = pick_value(src_row, "往来单位编码")
            customer_name = pick_value(src_row, "往来单位名") or customer_code
            receiver = pick_value(src_row, "收货公司") or customer_name or default_receiver
            product_code = pick_value(src_row, "品目编码")
            product_name = pick_value(src_row, "品目名")
            quantity = pick_value(src_row, "数量")
            unit_price = pick_value(src_row, "单价")
            total_amount = pick_value(src_row, "供应价", fallbacks=["外币金额", "金额1"])
            if total_amount is None:
                total_amount = pick_value(src_row, "外币金额")
            if total_amount is None and quantity is not None and unit_price is not None:
                try:
                    total_amount = float(quantity) * float(unit_price)
                except Exception:
                    pass
            raw_partner_value = customer_code or customer_name
            raw_product_value = product_code or product_name

            # 基础资料匹配（往来单位）
            partner_match = _fuzzy_find(partner_rows, customer_code or customer_name)
            if partner_match:
                customer_code = partner_match.get("code") or customer_code
                customer_name = partner_match.get("name") or customer_name
            elif raw_partner_value:
                unmatched_msgs.append(f"往来单位未匹配: {raw_partner_value}")
                unmatched_fields.extend(["往来单位编码", "往来单位名"])

            # 基础资料匹配（品目）
            product_match = _fuzzy_find(product_rows, product_code or product_name)
            if product_match:
                product_code = product_match.get("code") or product_code
                product_name = product_match.get("name") or product_name
                spec_val = product_match.get("specification") or product_match.get("spec_info")
                mo_tama_val = product_match.get("size_range") or product_match.get("spec_info")
                color_val = product_match.get("color")
                pack_qty_val = product_match.get("pack_qty")
                volume_val = product_match.get("volume")
                weight_val = product_match.get("weight")
            else:
                spec_val = mo_tama_val = color_val = pack_qty_val = volume_val = weight_val = None
                if raw_product_value:
                    unmatched_msgs.append(f"品目未匹配: {raw_product_value}")
                    unmatched_fields.extend(["品目编码", "品目名"])

            # 计算外币金额/本币供应价
            foreign_amount = total_amount
            if foreign_amount is None and quantity is not None and unit_price is not None:
                try:
                    foreign_amount = float(quantity) * float(unit_price)
                except Exception:
                    pass
            local_amount = None
            if foreign_amount is not None:
                try:
                    local_amount = float(foreign_amount) * default_rate_val
                except Exception:
                    local_amount = foreign_amount

            # 体积/重量汇总
            total_volume_val = None
            total_weight_val = None
            if volume_val is not None and quantity not in (None, ""):
                try:
                    total_volume_val = float(volume_val) * float(quantity if quantity is not None else 0)
                except Exception:
                    total_volume_val = volume_val
            if weight_val is not None and quantity not in (None, ""):
                try:
                    total_weight_val = float(weight_val) * float(quantity if quantity is not None else 0)
                except Exception:
                    total_weight_val = weight_val

            summary_text = pick_value(src_row, "摘要") or f"{customer_code or ''} {product_name or product_code or ''}".strip()

            out_row = []
            for header in self.template_headers:
                # 体积/重量类字段强制使用系统计算，不从映射取值
                if header.name in ["体积", "总体积", "重量", "总重量"]:
                    src_col = None
                else:
                    src_col = effective_mapping.get(header.name)
                value = src_row[src_col] if src_col else None

                # 针对关键字段的兜底赋值
                if header.name == "日期":
                    if value is None:
                        value = pick_value(src_row, "日期")
                elif header.name == "序号":
                    value = order_serial
                elif header.name == "PEDIDO No":
                    value = value if value not in (None, "") else order_no
                elif header.name == "往来单位编码":
                    value = customer_code
                elif header.name == "往来单位名":
                    value = customer_name
                elif header.name == "规格":
                    value = spec_val
                elif header.name == "Mo.TAMA":
                    value = mo_tama_val
                elif header.name == "COLOR":
                    value = color_val
                elif header.name == "装数":
                    value = pack_qty_val
                elif header.name == "CAJA":
                    value = pack_qty_val
                elif header.name == "收货公司":
                    value = value if value not in (None, "") else receiver
                elif header.name == "品目编码":
                    value = product_code
                elif header.name == "品目名":
                    value = product_name
                elif header.name == "数量":
                    value = quantity
                elif header.name == "单价":
                    value = unit_price
                elif header.name == "外币金额":
                    value = foreign_amount
                elif header.name in ["供应价", "金额1"]:
                    value = local_amount
                elif header.name == "货币":
                    value = value if not _is_empty(value) else default_currency
                elif header.name == "汇率":
                    value = value if not _is_empty(value) else default_rate_val
                elif header.name == "交易类型":
                    value = value if not _is_empty(value) else default_tran_type
                elif header.name == "部门":
                    value = value if not _is_empty(value) else default_department
                elif header.name == "发货仓库":
                    value = value if not _is_empty(value) else default_wh
                elif header.name == "职员":
                    value = value if not _is_empty(value) else default_staff
                elif header.name == "摘要":
                    value = summary_text
                elif header.name == "生成生产入库":
                    value = value if not _is_empty(value) else default_inbound_flag
                elif header.name == "体积":
                    value = volume_val
                elif header.name == "总体积":
                    value = total_volume_val
                elif header.name == "重量":
                    value = weight_val
                elif header.name == "总重量":
                    value = total_weight_val

                converted = convert_value(header.name, value)
                out_row.append(converted)

            output_rows.append(out_row)
            current_output_idx = len(output_rows)

            if unmatched_msgs:
                unique_fields = list(dict.fromkeys(unmatched_fields))
                unmatched_info.append({
                    "row_num": idx + 2,  # Excel 行号（包含表头）
                    "output_index": current_output_idx,
                    "fields": unique_fields,
                    "messages": unmatched_msgs
                })

        return output_rows, recognition_info, unmatched_info

    def _show_preprocessing_preview(self, output_rows, recognition_info, unmatched_info=None):
        """显示预处理预览窗口"""
        unmatched_info = unmatched_info or []
        preview_window = tk.Toplevel(self.root)
        preview_window.title("预处理预览 - 智能识别结果")
        preview_window.geometry("1200x700")

        # 创建主框架
        main_frame = ttk.Frame(preview_window, padding="10")
        main_frame.pack(fill="both", expand=True)

        # 顶部信息
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(
            info_frame,
            text=f"智能识别统计：共处理 {len(output_rows)} 行数据，识别到 {len(recognition_info)} 行包含可识别信息",
            font=("Arial", 10, "bold")
        ).pack(anchor="w")
        if unmatched_info:
            ttk.Label(
                info_frame,
                text=f"未匹配提醒：{len(unmatched_info)} 行存在基础资料未命中，已在下方标红，请手工核对",
                foreground="red"
            ).pack(anchor="w", pady=(2, 0))

        # 创建Notebook分页
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill="both", expand=True)

        # 第一页：识别详情
        detail_frame = ttk.Frame(notebook)
        notebook.add(detail_frame, text="识别详情")

        # 创建滚动区域
        detail_canvas = tk.Canvas(detail_frame)
        detail_scrollbar_y = ttk.Scrollbar(detail_frame, orient="vertical", command=detail_canvas.yview)
        detail_scrollbar_x = ttk.Scrollbar(detail_frame, orient="horizontal", command=detail_canvas.xview)
        detail_scrollable = ttk.Frame(detail_canvas)

        detail_scrollable.bind(
            "<Configure>",
            lambda e: detail_canvas.configure(scrollregion=detail_canvas.bbox("all"))
        )

        detail_canvas.create_window((0, 0), window=detail_scrollable, anchor="nw")
        detail_canvas.configure(yscrollcommand=detail_scrollbar_y.set, xscrollcommand=detail_scrollbar_x.set)

        detail_canvas.pack(side="left", fill="both", expand=True)
        detail_scrollbar_y.pack(side="right", fill="y")
        detail_scrollbar_x.pack(side="bottom", fill="x")

        # 显示识别详情
        for info in recognition_info:
            item_frame = ttk.LabelFrame(
                detail_scrollable,
                text=f"第 {info['row_num']} 行",
                padding="10"
            )
            item_frame.pack(fill="x", padx=5, pady=5)

            # 摘要
            summary_frame = ttk.Frame(item_frame)
            summary_frame.pack(fill="x", pady=(0, 5))
            ttk.Label(summary_frame, text="摘要:", font=("Arial", 9, "bold")).pack(side="left")
            ttk.Label(summary_frame, text=info['summary'], wraplength=900).pack(side="left", padx=(5, 0))

            # 识别字段
            recognized_frame = ttk.Frame(item_frame)
            recognized_frame.pack(fill="x")
            ttk.Label(recognized_frame, text="识别结果:", font=("Arial", 9, "bold")).pack(anchor="w")

            rec_text = ttk.Frame(recognized_frame)
            rec_text.pack(fill="x", padx=(20, 0))

            for key, value in info['recognized'].items():
                if value:  # 只显示非空值
                    ttk.Label(
                        rec_text,
                        text=f"• {key}: {value}",
                        foreground="blue"
                    ).pack(anchor="w")

        if unmatched_info:
            ttk.Separator(detail_scrollable, orient="horizontal").pack(fill="x", pady=5)
            ttk.Label(
                detail_scrollable,
                text="未匹配提醒（基础资料未命中/需人工确认）",
                foreground="red",
                font=("Arial", 10, "bold")
            ).pack(anchor="w", pady=(0, 5))

            for warn in unmatched_info:
                warn_frame = ttk.LabelFrame(
                    detail_scrollable,
                    text=f"第 {warn.get('row_num')} 行",
                    padding="10"
                )
                warn_frame.pack(fill="x", padx=5, pady=3)
                for msg in warn.get("messages", []):
                    ttk.Label(
                        warn_frame,
                        text=f"! {msg}",
                        foreground="red"
                    ).pack(anchor="w")

        # 第二页：转换结果预览
        preview_frame = ttk.Frame(notebook)
        notebook.add(preview_frame, text="转换结果预览")

        unmatched_output_indices = {
            item.get("output_index") for item in unmatched_info if item.get("output_index")
        }

        if unmatched_output_indices:
            ttk.Label(
                preview_frame,
                text="标红行包含未匹配基础资料，请核对后再导出",
                foreground="red"
            ).pack(anchor="w", padx=5)

        # 创建树形视图显示前20行数据
        tree_frame = ttk.Frame(preview_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # 创建滚动条
        tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")

        # 创建Treeview
        preview_tree = ttk.Treeview(
            tree_frame,
            columns=[h.name for h in self.template_headers],
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set
        )

        tree_scroll_y.config(command=preview_tree.yview)
        tree_scroll_x.config(command=preview_tree.xview)
        preview_tree.tag_configure("unmatched", background="#fff1f0")

        # 设置列标题
        for header in self.template_headers:
            preview_tree.heading(header.name, text=header.name)
            preview_tree.column(header.name, width=100, minwidth=80)

        # 插入数据（只显示前20行）
        for idx, row in enumerate(output_rows[:20], 1):
            tags = [f"row{idx}"]
            if idx in unmatched_output_indices:
                tags.append("unmatched")
            preview_tree.insert("", "end", values=row, tags=tuple(tags))

        preview_tree.pack(side="left", fill="both", expand=True)
        tree_scroll_y.pack(side="right", fill="y")
        tree_scroll_x.pack(side="bottom", fill="x")
        attach_treeview_tools(preview_tree)

        if len(output_rows) > 20:
            ttk.Label(
                preview_frame,
                text=f"注：仅显示前20行，共 {len(output_rows)} 行",
                foreground="gray"
            ).pack(pady=5)

        # 底部按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))

        result = {"confirmed": False}

        def export_preview():
            """导出完整预览数据为 Excel（包含所有行）"""
            path = filedialog.asksaveasfilename(
                title="导出预览为 Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            )
            if not path:
                return
            try:
                df = pd.DataFrame(output_rows, columns=[h.name for h in self.template_headers])
                df.to_excel(path, index=False)
                messagebox.showinfo("完成", f"已导出预览数据：\n{path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败：{e}")

        def confirm():
            result["confirmed"] = True
            preview_window.destroy()

        def cancel():
            result["confirmed"] = False
            preview_window.destroy()

        ttk.Button(button_frame, text="导出预览为Excel", command=export_preview, width=18).pack(side="left")
        ttk.Button(button_frame, text="确认并导出", command=confirm, width=15).pack(side="right", padx=5)
        ttk.Button(button_frame, text="取消", command=cancel, width=15).pack(side="right")

        # 使窗口模态
        preview_window.transient(self.root)
        preview_window.grab_set()
        self.root.wait_window(preview_window)

        return result["confirmed"]

    def _export_to_excel(self, output_rows, unmatched_info=None):
        """导出到Excel文件"""
        unmatched_info = unmatched_info or []
        save_path = filedialog.asksaveasfilename(
            title="保存转换结果",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not save_path:
            return

        try:
            wb = load_workbook(TEMPLATE_FILE, data_only=False)
            ws = wb.active

            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            header_index = {h.name: idx for idx, h in enumerate(self.template_headers, start=1)}
            start_row = 2
            for r_idx, row_values in enumerate(output_rows, start=start_row):
                for c_idx, v in enumerate(row_values, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=v)

            if unmatched_info:
                highlight_fill = PatternFill(start_color="FFFFE6E6", end_color="FFFFE6E6", fill_type="solid")
                for warn in unmatched_info:
                    output_idx = warn.get("output_index")
                    if not output_idx:
                        continue
                    excel_row = start_row + output_idx - 1
                    for field in warn.get("fields", []):
                        col_idx = header_index.get(field)
                        if not col_idx:
                            continue
                        ws.cell(row=excel_row, column=col_idx).fill = highlight_fill

            wb.save(save_path)
        except Exception as e:
            messagebox.showerror("错误", f"保存结果失败：\n{e}")
            return

        messagebox.showinfo("完成", f"转换完成！\n已保存到：\n{save_path}")

    # ========== 基础数据管理标签页功能 ==========

    def _load_base_data_table(self):
        """加载基础数据表"""
        if not self.base_data_mgr:
            self.status_label.config(text="错误：基础数据管理器未初始化")
            return

        table_name = self.current_table.get()

        try:
            # 获取数据
            if table_name == "smart_recognition_cache":
                results = self.base_data_mgr.get_all_cached_recognitions()
            else:
                results = self.base_data_mgr.query(table_name)

            # 如果是科目编码表，为当前默认科目附上“默认往来单位编码”列
            if table_name == "account_subject":
                partner_map = self.base_data_mgr.get_subject_partner_map()

                def _extract_subject_code(row):
                    code_name = row.get("code_name", "") or row.get("code", "")
                    m = re.search(r'\[(\d+)\]', str(code_name))
                    if m:
                        return m.group(1)
                    return str(code_name).strip()

                for row in results:
                    subj_code = _extract_subject_code(row)
                    row["默认往来单位编码"] = partner_map.get(subj_code, "")

            self._populate_base_data_table(results, table_name=table_name)

        except Exception as e:
            messagebox.showerror("错误", f"加载数据失败：\n{e}")
            self.status_label.config(text=f"加载失败: {str(e)}")

    def _search_base_data(self):
        """搜索基础数据"""
        if not self.base_data_mgr:
            return

        keyword = self.search_var.get().strip()
        table_name = self.current_table.get()
        search_field = self.search_field_var.get().strip()

        try:
            if table_name == "smart_recognition_cache":
                # 对缓存表进行搜索 (目前没有直接的 search_by_name，这里简单过滤)
                all_results = self.base_data_mgr.get_all_cached_recognitions()
                results = [r for r in all_results if keyword.lower() in str(r.get("summary", "")).lower() or keyword.lower() in str(r.get("account_code", "")).lower()]
            else:
                results = self.base_data_mgr.query(table_name)
                if keyword and results:
                    # 按指定字段过滤，若字段不存在则对整行做模糊过滤
                    k_lower = keyword.lower()
                    if search_field and search_field in results[0].keys():
                        results = [r for r in results if k_lower in str(r.get(search_field, "")).lower()]
                    else:
                        results = [
                            r for r in results
                            if any(k_lower in str(v).lower() for v in r.values())
                        ]

            # 科目表补充“默认往来单位编码”
            if table_name == "account_subject":
                partner_map = self.base_data_mgr.get_subject_partner_map()

                def _extract_subject_code(row):
                    code_name = row.get("code_name", "") or row.get("code", "")
                    m = re.search(r'\[(\d+)\]', str(code_name))
                    if m:
                        return m.group(1)
                    return str(code_name).strip()

                for row in results:
                    subj_code = _extract_subject_code(row)
                    row["默认往来单位编码"] = partner_map.get(subj_code, "")

            self._populate_base_data_table(results, table_name=table_name)

            if keyword:
                self.status_label.config(text=f"找到 {len(results)} 条包含 '{keyword}' 的记录")
            else:
                self.status_label.config(text=f"共 {len(results)} 条记录")

        except Exception as e:
            messagebox.showerror("错误", f"搜索失败：\n{e}")

    def _populate_base_data_table(self, results, table_name=None):
        """渲染基础数据表格（支持搜索和加载复用）"""
        # 清空现有内容
        for item in self.base_data_tree.get_children():
            self.base_data_tree.delete(item)

        table_name = table_name or self.current_table.get()
        if not results:
            self.status_label.config(text=f"{table_name} - 无数据")
            # 为空时也要设置列，防止Treeview报错
            if table_name == "smart_recognition_cache":
                columns = ["id", "summary", "match_items", "account_code", "created_at"]
                self.base_data_tree["columns"] = columns
                self.base_data_tree["show"] = "headings"
                for col in columns:
                    heading_text = "映射匹配项" if col == "match_items" else col
                    self.base_data_tree.heading(col, text=heading_text)
                    self.base_data_tree.column(col, width=100)
                    if col == "summary": self.base_data_tree.column(col, width=300)
                    if col == "match_items": self.base_data_tree.column(col, width=220)
                    if col == "created_at": self.base_data_tree.column(col, width=150)
            else:
                # 对于非缓存表，如果无数据，先清空列头，避免显示旧的
                self.base_data_tree["columns"] = []
            return

        # 将匹配项展示为易读的逗号分隔形式
        tables_with_match_items = {"smart_recognition_cache", "account_subject", "product", "business_partner"}
        if table_name in tables_with_match_items:
            normalized_results = []
            for row in results:
                row_copy = dict(row)
                raw_items = row_copy.get("match_items", "")
                display_items = ""
                try:
                    parsed = json.loads(raw_items) if isinstance(raw_items, str) else (raw_items or [])
                    if isinstance(parsed, list):
                        display_items = ", ".join(str(i).strip() for i in parsed if str(i).strip())
                    else:
                        display_items = str(parsed)
                except Exception:
                    # 如果不是有效的 JSON，直接展示原始字符串
                    display_items = str(raw_items) if raw_items is not None else ""
                row_copy["match_items"] = display_items
                normalized_results.append(row_copy)
            results = normalized_results

        # 获取列名
        columns = list(results[0].keys())
        self.current_columns = columns
        # 更新搜索字段下拉
        self.search_field_combo["values"] = columns
        if self.search_field_var.get() not in columns:
            self.search_field_var.set(columns[0] if columns else "")

        # 配置树形视图列
        self.base_data_tree["columns"] = columns
        self.base_data_tree["show"] = "headings"

        # 设置列标题和宽度，并绑定排序
        for col in columns:
            heading_text = "映射匹配项" if col == "match_items" else col
            self.base_data_tree.heading(col, text=heading_text, command=lambda c=col: self._sort_treeview(c))
            width = 100
            if col == "summary": width = 300
            elif col == "created_at": width = 150
            elif col == "account_code": width = 80
            elif col == "match_items": width = 220
            self.base_data_tree.column(col, width=width, stretch=True)

        # 插入数据
        for row in results:
            values = [row.get(col, "") for col in columns]
            self.base_data_tree.insert("", "end", values=values)

        self.status_label.config(text=f"{table_name} - 共 {len(results)} 条记录")

    def _sort_treeview(self, column):
        """点击表头排序，数字优先按数值排序"""
        # 读取当前数据
        items = self.base_data_tree.get_children("")
        data = []
        for iid in items:
            values = self.base_data_tree.item(iid, "values")
            data.append((iid, values))

        if not data:
            return

        col_idx = self.base_data_tree["columns"].index(column)

        def _as_number(val):
            try:
                return float(val)
            except (TypeError, ValueError):
                return val

        # 切换排序方向
        ascending = not self.sort_states.get(column, True)
        self.sort_states[column] = ascending

        sorted_data = sorted(
            data,
            key=lambda item: _as_number(item[1][col_idx]),
            reverse=not ascending
        )

        # 重新排列
        for idx, (iid, _) in enumerate(sorted_data):
            self.base_data_tree.move(iid, "", idx)

    def _add_base_data_record(self):
        """新增基础数据记录"""
        table_name = self.current_table.get()
        if table_name == "smart_recognition_cache":
            # 打开缓存专用新增对话框
            self._add_cached_recognition_dialog(self._load_base_data_table)
            return

        if not self.base_data_mgr:
            return

        columns = self.base_data_mgr.get_table_columns(table_name)

        self._edit_record_dialog(table_name, columns, None, self._load_base_data_table)

    def _add_cached_recognition_dialog(self, refresh_callback):
        """新增智能识别缓存记录（手动）"""
        dialog = tk.Toplevel(self.root)
        dialog.title("新增智能识别缓存")
        dialog.geometry("650x460")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="摘要（原文）:").grid(row=0, column=0, sticky="ne", padx=8, pady=8)
        summary_text = tk.Text(dialog, height=4, width=55, wrap="word")
        summary_text.grid(row=0, column=1, columnspan=3, sticky="we", padx=5, pady=8)

        normalize_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            dialog,
            text="保存前自动标准化摘要/匹配项（去日期/金额噪声，提高命中率）",
            variable=normalize_var
        ).grid(row=1, column=1, columnspan=3, sticky="w", padx=5)

        ttk.Label(dialog, text="科目编码:").grid(row=2, column=0, sticky="e", padx=8, pady=8)
        account_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=account_var, width=30).grid(row=2, column=1, sticky="w", padx=5, pady=8)

        ttk.Label(dialog, text="映射匹配项（多对一，逗号/换行分隔）:").grid(row=3, column=0, sticky="ne", padx=8, pady=4)
        match_items_text = tk.Text(dialog, height=3, width=55, wrap="word")
        match_items_text.grid(row=3, column=1, columnspan=3, sticky="we", padx=5, pady=4)

        preview_var = tk.StringVar()
        ttk.Label(dialog, text="将被缓存的摘要:").grid(row=4, column=0, sticky="ne", padx=8)
        preview_label = ttk.Label(dialog, textvariable=preview_var, wraplength=420, foreground="gray")
        preview_label.grid(row=4, column=1, columnspan=3, sticky="w", padx=5)

        def _normalize(text: str) -> str:
            if not text:
                return ""
            if normalize_var.get() and self.summary_recognizer:
                try:
                    return self.summary_recognizer._normalize_summary(text)
                except Exception:
                    pass
            return text.strip()

        def _parse_match_items() -> List[str]:
            raw = match_items_text.get("1.0", "end").strip()
            if not raw:
                return []
            parts = re.split(r"[\\n,，;；、]+", raw)
            items = []
            for part in parts:
                p = part.strip()
                if not p:
                    continue
                items.append(_normalize(p))
            return items

        def refresh_preview():
            raw = summary_text.get("1.0", "end").strip()
            preview_var.set(_normalize(raw))

        ttk.Button(dialog, text="刷新预览", command=refresh_preview, width=10).grid(row=0, column=4, padx=5, pady=8, sticky="n")

        def save():
            raw_summary = summary_text.get("1.0", "end").strip()
            account_code = account_var.get().strip()

            if not raw_summary:
                messagebox.showwarning("警告", "摘要不能为空！")
                return
            if not account_code:
                messagebox.showwarning("警告", "科目编码不能为空！")
                return

            summary_to_save = _normalize(raw_summary)
            match_items = _parse_match_items()
            try:
                self.base_data_mgr.save_cached_recognition(summary_to_save, account_code, match_items)
                messagebox.showinfo("成功", f"已保存缓存：\n摘要= {summary_to_save}\n科目= {account_code}")
                dialog.destroy()
                refresh_callback()
            except Exception as e:
                messagebox.showerror("错误", f"保存失败：{e}")

        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=5, column=0, columnspan=5, pady=12)
        ttk.Button(button_frame, text="保存", command=save, width=12).pack(side="right", padx=6)
        ttk.Button(button_frame, text="取消", command=dialog.destroy, width=12).pack(side="right")

    def _edit_base_data_record(self):
        """编辑基础数据记录（支持缓存批量编辑）"""
        if not self.base_data_mgr:
            return

        selections = self.base_data_tree.selection()
        if not selections:
            messagebox.showwarning("提示", "请先选择要编辑的记录")
            return

        table_name = self.current_table.get()
        
        # === 批量编辑逻辑 (仅限智能识别缓存) ===
        if len(selections) > 1:
            if table_name == "smart_recognition_cache":
                self._batch_edit_cached_recognition_dialog(selections, self._load_base_data_table)
                return
            else:
                messagebox.showwarning("提示", "当前数据类型不支持批量编辑，请只选择一条记录。")
                return

        # === 单条编辑逻辑 ===
        item = selections[0]
        values = self.base_data_tree.item(item, "values")
        
        if table_name == "smart_recognition_cache":
            # 针对缓存表，只允许编辑科目编码
            columns = ["id", "summary", "match_items", "account_code", "created_at"]
            record_data = dict(zip(columns, values))
            self._edit_cached_recognition_dialog(record_data, self._load_base_data_table)
            return

        columns = self.base_data_mgr.get_table_columns(table_name)
        record_data = dict(zip(columns, values))

        self._edit_record_dialog(table_name, columns, record_data, self._load_base_data_table)

    def _batch_edit_cached_recognition_dialog(self, selections, refresh_callback):
        """批量编辑智能识别缓存对话框"""
        count = len(selections)
        dialog = tk.Toplevel(self.root)
        dialog.title(f"批量编辑 ({count} 条记录)")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text=f"正在批量修改 {count} 条缓存记录", font=("", 10, "bold")).pack(pady=10)
        
        frame = ttk.Frame(dialog)
        frame.pack(pady=5)
        ttk.Label(frame, text="统一修改科目编码为:").pack(side="left")
        new_code_var = tk.StringVar()
        ttk.Entry(frame, textvariable=new_code_var, width=20).pack(side="left", padx=5)

        def save_batch():
            new_code = new_code_var.get().strip()
            if not new_code:
                messagebox.showwarning("警告", "科目编码不能为空！")
                return
            
            success_count = 0
            
            # 循环更新
            for item in selections:
                values = self.base_data_tree.item(item, "values")
                record_id = int(values[0])
                result = self.base_data_mgr.update_cached_recognition(record_id, new_code)
                if result["success"]:
                    success_count += 1
            
            messagebox.showinfo("完成", f"成功更新 {success_count} 条记录")
            dialog.destroy()
            refresh_callback()

        ttk.Button(dialog, text="保存全部", command=save_batch).pack(pady=15)

    def _delete_base_data_record(self):
        """删除基础数据记录（支持批量）"""
        if not self.base_data_mgr:
            return

        selections = self.base_data_tree.selection()
        if not selections:
            messagebox.showwarning("提示", "请先选择要删除的记录")
            return

        count = len(selections)
        table_name = self.current_table.get()
        
        # 确认删除
        confirm = messagebox.askyesno("确认删除", f"确定要删除选中的 {count} 条记录吗？\n此操作不可恢复！")
        if not confirm:
            return
        
        success_count = 0
        fail_count = 0
        
        for item in selections:
            values = self.base_data_tree.item(item, "values")
            record_id = values[0]  # 第一列是id
            
            result = None
            if table_name == "smart_recognition_cache":
                result = self.base_data_mgr.delete_cached_recognition(int(record_id))
            else:
                result = self.base_data_mgr.delete_record(table_name, int(record_id))
            
            if result and result["success"]:
                success_count += 1
            else:
                fail_count += 1

        if success_count > 0:
            messagebox.showinfo("操作完成", f"成功删除 {success_count} 条记录" + (f"，失败 {fail_count} 条" if fail_count > 0 else ""))
            self._load_base_data_table()  # 刷新列表
        else:
            messagebox.showerror("错误", "删除失败")

    def _edit_cached_recognition_dialog(self, record_data: Dict, refresh_callback):
        """编辑智能识别缓存对话框 (科目编码 + 映射匹配项)"""
        dialog = tk.Toplevel(self.root)
        dialog.title("编辑智能识别缓存")
        dialog.geometry("560x320")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="摘要:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        ttk.Label(dialog, text=record_data.get("summary", ""), wraplength=350, justify="left").grid(row=0, column=1, sticky="w", padx=5, pady=5)

        ttk.Label(dialog, text="当前科目编码:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        current_code_label = ttk.Label(dialog, text=record_data.get("account_code", ""), font=("", 10, "bold"))
        current_code_label.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        ttk.Label(dialog, text="新科目编码:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
        new_code_var = tk.StringVar(value=record_data.get("account_code", ""))
        new_code_entry = ttk.Entry(dialog, textvariable=new_code_var, width=30)
        new_code_entry.grid(row=2, column=1, sticky="we", padx=5, pady=5)

        ttk.Label(dialog, text="映射匹配项:").grid(row=3, column=0, sticky="ne", padx=5, pady=5)
        match_items_text = tk.Text(dialog, height=3, width=45, wrap="word")
        match_items_text.grid(row=3, column=1, columnspan=2, sticky="we", padx=5, pady=5)

        # 预填充匹配项
        existing_items = record_data.get("match_items", "")
        preset = ""
        try:
            parsed = json.loads(existing_items) if isinstance(existing_items, str) else (existing_items or [])
            if isinstance(parsed, list):
                preset = "\n".join(str(i).strip() for i in parsed if str(i).strip())
            elif parsed:
                preset = str(parsed)
        except Exception:
            preset = str(existing_items) if existing_items else ""
        if preset:
            match_items_text.insert("1.0", preset)

        def _parse_match_items() -> List[str]:
            raw = match_items_text.get("1.0", "end").strip()
            if not raw:
                return []
            parts = re.split(r"[\n,，;；、]+", raw)
            return [p.strip() for p in parts if p and p.strip()]

        # AI 重新识别功能
        def retry_ai():
            if not self.summary_recognizer:
                messagebox.showerror("错误", "智能识别器未初始化")
                return
            
            summary = record_data.get("summary", "")
            if not summary:
                return

            # 禁用按钮防止重复点击
            retry_btn.config(state="disabled", text="识别中...")
            dialog.update()

            try:
                # 调用核心 AI 逻辑（不走缓存）
                code = self.summary_recognizer.recognize_account_with_ai_core(summary)
                if code:
                    new_code_var.set(code)
                    messagebox.showinfo("识别成功", f"AI 建议科目: {code}")
                else:
                    messagebox.showwarning("无结果", "AI 未能识别出有效科目")
            except Exception as e:
                messagebox.showerror("错误", f"识别失败: {e}")
            finally:
                retry_btn.config(state="normal", text="AI 重新识别")

        retry_btn = ttk.Button(dialog, text="AI 重新识别", command=retry_ai, width=12)
        retry_btn.grid(row=2, column=2, padx=5, pady=5)

        def save():
            new_code = new_code_var.get().strip()
            if not new_code:
                messagebox.showwarning("警告", "科目编码不能为空！")
                return
            
            match_items = _parse_match_items()
            result = self.base_data_mgr.update_cached_recognition(record_data["id"], new_code, match_items)
            if result["success"]:
                messagebox.showinfo("成功", result["message"])
                dialog.destroy()
                refresh_callback()
            else:
                messagebox.showerror("错误", result["message"])

        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=4, column=0, columnspan=3, pady=10)
        ttk.Button(button_frame, text="保存", command=save, width=10).pack(side="left", padx=5)
        ttk.Button(button_frame, text="取消", command=dialog.destroy, width=10).pack(side="left", padx=5)


    def _set_as_default_base_data(self):
        """将选中的基础数据设为默认值"""
        if not self.base_data_mgr:
            return

        selection = self.base_data_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要设为默认值的记录")
            return

        table_name = self.current_table.get()

        # 映射表名到模板字段名
        table_to_field = {
            "department": "部门",
            "account_subject": "科目编码",
            "business_partner": "往来单位编码",
            "currency": "汇率",
            # 账户默认值：存账户编码，同时自动从账户关联科目里提取默认对方科目
            "bank_account": "默认账户往来",
            "warehouse": "发货仓库",
        }

        target_field = table_to_field.get(table_name)
        
        # 特殊处理币种：从选中记录中获取汇率
        val = None
        if table_name == "currency":
            # 获取选中行数据
            item = selection[0]
            tree_columns = self.base_data_tree["columns"]
            values = self.base_data_tree.item(item, "values")
            if tree_columns and values:
                row_data = dict(zip(tree_columns, values))
                rate_val = row_data.get("exchange_rate") # 尝试获取汇率
                code_val = row_data.get("code")
                if not rate_val and not code_val:
                    messagebox.showwarning("提示", "选中的币种记录无效（无汇率或编码）。")
                    return
                # 同时写入汇率与货币编码默认值
                if rate_val is not None:
                    self.default_values["汇率"] = rate_val
                    if self.base_data_mgr:
                        self.base_data_mgr.set_config("汇率", rate_val)
                if code_val:
                    self.default_values["货币"] = code_val
                    if self.base_data_mgr:
                        self.base_data_mgr.set_config("货币", code_val)
                self._update_recognizer_defaults()
                messagebox.showinfo("成功", f"已将币种默认设置为 货币={code_val or ''} 汇率={rate_val or ''}")
                return
        
        elif not target_field:
            messagebox.showinfo("提示", f"当前数据类型 '{table_name}' 暂不支持设为转换默认值。\n目前支持：部门、科目编码、往来单位、币种、账户。")
            return
        
        # 对于其他支持的数据类型，优先使用 code，其次使用 name
        if not val: # If val was not set by currency specific logic
            item = selection[0]
            tree_columns = self.base_data_tree["columns"]
            values = self.base_data_tree.item(item, "values")
            
            if not tree_columns or not values:
                messagebox.showerror("错误", "无法获取记录数据")
                return

            row_data = dict(zip(tree_columns, values))

            # 特殊处理科目编码：从 code_name 中提取
            if table_name == "account_subject":
                code_name = row_data.get("code_name", "")
                match = re.search(r'\[(\d+)\]', code_name)
                if match:
                    val = match.group(1)
                else:
                    val = code_name # Fallback to full string if format doesn't match
            else:
                val = row_data.get("code")
                if not val:
                    val = row_data.get("name")
            
            if not val:
                messagebox.showwarning("提示", "选中的记录无效（无代码或名称）。")
                return
        
        val = str(val).strip()

        # 设置默认值 (内存 + 数据库持久化)
        self.default_values[target_field] = val
        if self.base_data_mgr:
            self.base_data_mgr.set_config(target_field, val)
            # 兼容销售出库默认仓库兜底
            if table_name == "warehouse":
                self.default_values["默认仓库"] = val
                self.base_data_mgr.set_config("默认仓库", val)

        self._update_recognizer_defaults()

        messagebox.showinfo("成功", f"已将 '{val}' 设为 '{target_field}' 的默认值。")

    def _set_default_partner_via_subject(self):
        """在科目编码页签，设置默认往来单位"""
        if self.current_table.get() != "account_subject":
            messagebox.showinfo("提示", "请切换到“科目编码”页签后再设置默认往来单位。")
            return

        if not self.base_data_mgr:
            messagebox.showerror("错误", "基础数据管理器未初始化")
            return

        # 获取所有往来单位供选择
        partners = self.base_data_mgr.query("business_partner")
        if not partners:
            messagebox.showwarning("提示", "未找到往来单位数据，请先导入基础数据。")
            return

        # 需选中一个科目
        selection = self.base_data_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先在科目编码列表中选中一个科目。")
            return
        item = selection[0]
        tree_columns = self.base_data_tree["columns"]
        values = self.base_data_tree.item(item, "values")
        row_data = dict(zip(tree_columns, values))

        # 获取科目编码
        subject_display = row_data.get("code_name", "") or row_data.get("code", "")
        subject_code = ""
        match = re.search(r'\[(\d+)\]', subject_display)
        if match:
            subject_code = match.group(1)
        else:
            subject_code = str(subject_display).strip()

        dialog = tk.Toplevel(self.root)
        dialog.title("设置默认往来单位")
        dialog.geometry("480x260")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text=f"科目: {subject_display}", font=("", 10, "bold")).pack(pady=(10, 5))
        ttk.Label(dialog, text="选择该科目的默认往来单位（用于自动生成对方分录）").pack()

        options = [f"{p.get('code','')} - {p.get('name','')}" for p in partners]
        partner_var = tk.StringVar()
        combo = ttk.Combobox(dialog, textvariable=partner_var, values=options, width=50)
        combo.pack(padx=10, pady=5)
        combo.state(["readonly"])

        # 允许手动输入编码
        manual_var = tk.StringVar()
        ttk.Label(dialog, text="或手动输入往来单位编码：").pack()
        ttk.Entry(dialog, textvariable=manual_var, width=30).pack(pady=5)

        def save_partner():
            sel = partner_var.get().strip()
            manual = manual_var.get().strip()
            code = ""
            if manual:
                code = manual
            elif sel:
                code = sel.split(" - ")[0]

            if not code:
                messagebox.showwarning("提示", "请先选择或输入往来单位编码。")
                return

            # 保存科目-往来映射（每个科目独立存储）
            self.base_data_mgr.set_partner_for_subject(subject_code, code)
            messagebox.showinfo("成功", f"已将往来单位 {code} 设为科目 {subject_code} 的默认。")
            dialog.destroy()
            self._load_base_data_table()  # 刷新列表展示“默认往来单位编码”

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="保存", command=save_partner, width=10).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy, width=10).pack(side="left", padx=5)

    # ========== 基础数据管理菜单功能 ==========
    def import_base_data(self):
        """导入基础数据"""
        if not self.base_data_mgr:
            messagebox.showerror("错误", "基础数据管理器未初始化")
            return

        # 询问是否确认导入
        confirm = messagebox.askyesno(
            "确认导入",
            "将从 '基础数据/基础数据' 目录导入所有基础数据。\n"
            "这将覆盖现有数据，是否继续？"
        )
        if not confirm:
            return

        # 显示进度提示
        progress_window = tk.Toplevel(self.root)
        progress_window.title("导入中")
        progress_window.geometry("300x100")
        progress_label = ttk.Label(progress_window, text="正在导入基础数据，请稍候...")
        progress_label.pack(expand=True)
        progress_window.update()

        try:
            result = self.base_data_mgr.import_all_data()
            progress_window.destroy()

            # 显示详细结果
            details_text = "\n".join([
                f"{file}: {info['message']}"
                for file, info in result["details"].items()
            ])

            messagebox.showinfo(
                "导入结果",
                f"{result['message']}\n\n详细信息：\n{details_text}"
            )

        except Exception as e:
            progress_window.destroy()
            messagebox.showerror("错误", f"导入失败：\n{e}")

    def show_base_data_stats(self):
        """显示基础数据统计信息"""
        if not self.base_data_mgr:
            messagebox.showerror("错误", "基础数据管理器未初始化")
            return

        try:
            stats = self.base_data_mgr.get_statistics()

            # 表名中文映射
            table_names = {
                "currency": "币种",
                "department": "部门",
                "warehouse": "仓库",
                "account_subject": "科目编码",
                "product": "品目信息",
                "business_partner": "往来单位",
                "bank_account": "账户"
            }

            stats_text = "\n".join([
                f"{table_names.get(table, table)}: {count} 条"
                for table, count in stats.items()
            ])

            total = sum(stats.values())

            messagebox.showinfo(
                "基础数据统计",
                f"数据库文件: {self.base_data_mgr.db_path}\n\n"
                f"{stats_text}\n\n"
                f"总计: {total} 条记录"
            )

        except Exception as e:
            messagebox.showerror("错误", f"获取统计信息失败：\n{e}")

    def query_base_data(self, table_name: str):
        """查询基础数据"""
        if not self.base_data_mgr:
            messagebox.showerror("错误", "基础数据管理器未初始化")
            return

        # 创建查询窗口
        query_window = tk.Toplevel(self.root)
        query_window.title(f"查询 - {table_name}")
        query_window.geometry("800x600")

        # 搜索框
        search_frame = ttk.Frame(query_window, padding=10)
        search_frame.pack(fill="x")

        ttk.Label(search_frame, text="搜索（名称）:").pack(side="left")
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side="left", padx=5)

        def do_search():
            keyword = search_var.get().strip()
            if keyword:
                results = self.base_data_mgr.search_by_name(table_name, keyword)
            else:
                results = self.base_data_mgr.query(table_name)

            # 清空树形视图
            for item in tree.get_children():
                tree.delete(item)

            # 显示结果
            if results:
                for row in results:
                    values = [row.get(col, "") for col in columns]
                    tree.insert("", "end", values=values)
                status_label.config(text=f"找到 {len(results)} 条记录")
            else:
                status_label.config(text="未找到记录")

        ttk.Button(search_frame, text="搜索", command=do_search).pack(side="left", padx=5)
        ttk.Button(search_frame, text="显示全部", command=do_search).pack(side="left")

        # 编辑按钮框
        def add_record():
            self._edit_record_dialog(table_name, columns, None, do_search)

        def edit_record():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("提示", "请先选择要编辑的记录")
                return

            # 获取选中行的数据
            item = selection[0]
            values = tree.item(item, "values")
            record_data = dict(zip(columns, values))

            self._edit_record_dialog(table_name, columns, record_data, do_search)

        def delete_record():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("提示", "请先选择要删除的记录")
                return

            # 获取选中行的ID
            item = selection[0]
            values = tree.item(item, "values")
            record_id = values[0]  # 第一列是id

            # 确认删除
            confirm = messagebox.askyesno("确认删除", f"确定要删除ID为 {record_id} 的记录吗？")
            if not confirm:
                return

            result = self.base_data_mgr.delete_record(table_name, int(record_id))
            if result["success"]:
                messagebox.showinfo("成功", result["message"])
                do_search()  # 刷新列表
            else:
                messagebox.showerror("错误", result["message"])

        edit_button_frame = ttk.Frame(search_frame)
        edit_button_frame.pack(side="right")

        ttk.Button(edit_button_frame, text="新增", command=add_record).pack(side="left", padx=2)
        ttk.Button(edit_button_frame, text="编辑", command=edit_record).pack(side="left", padx=2)
        ttk.Button(edit_button_frame, text="删除", command=delete_record).pack(side="left", padx=2)

        # 结果显示区域
        result_frame = ttk.Frame(query_window)
        result_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # 创建滚动条
        scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
        scrollbar_y.pack(side="right", fill="y")

        scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")
        scrollbar_x.pack(side="bottom", fill="x")

        # 获取列信息
        sample_data = self.base_data_mgr.query(table_name)
        if sample_data:
            columns = list(sample_data[0].keys())
        else:
            columns = ["id", "code", "name"]

        # 创建树形视图
        tree = ttk.Treeview(
            result_frame,
            columns=columns,
            show="headings",
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )

        scrollbar_y.config(command=tree.yview)
        scrollbar_x.config(command=tree.xview)

        # 设置列标题
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)

        tree.pack(fill="both", expand=True)
        attach_treeview_tools(tree)

        # 状态栏
        status_label = ttk.Label(query_window, text="", relief="sunken")
        status_label.pack(fill="x", side="bottom")

        # 自动加载全部数据
        do_search()

    def _edit_record_dialog(self, table_name: str, columns: List[str], record_data: Optional[Dict], refresh_callback):
        """编辑/新增记录对话框"""
        is_new = record_data is None

        # 创建对话框
        dialog = tk.Toplevel(self.root)
        dialog.title("新增记录" if is_new else "编辑记录")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()

        # 字段中文名映射
        field_labels = {
            "id": "ID",
            "code": "编码",
            "name": "名称",
            "exchange_rate": "汇率",
            "use_type": "使用类型",
            "is_active": "是否启用",
            "type": "类型",
            "production_process": "生产流程",
            "outsource_partner": "外包单位",
            "code_name": "科目编码名称",
            "is_subject": "是否科目",
            "debit_credit_type": "借贷类型",
            "subject_type": "科目类型",
            "contra_type": "备抵类型",
            "summary": "摘要",
            "parent_subject": "上级科目",
            "display_name": "显示名称",
            "product_type": "品目类型",
            "spec_info": "规格信息",
            "unit": "单位",
            "search_keyword": "搜索关键词",
            "pack_qty": "装数",
            "unit_conversion_denominator": "单位转换分母",
            "unit_conversion_numerator": "单位转换分子",
            "specification": "规格",
            "in_price": "入库单价",
            "out_price": "出库单价",
            "price_a": "单价A",
            "price_b": "单价B",
            "price_c": "单价C",
            "length": "长度",
            "width": "宽度",
            "height": "高度",
            "volume": "体积",
            "weight": "重量",
            "color": "颜色",
            "size_range": "尺码范围",
            "contact_person": "联系人",
            "mobile": "手机",
            "phone": "电话",
            "email": "邮箱",
            "category": "分类",
            "file_management": "文件管理",
            "tax_number": "税号",
            "bank_name": "开户行",
            "bank_account": "银行账号",
            "account_subject": "会计科目",
            "foreign_currency": "外币存折",
            "match_items": "映射匹配项",
        }

        # 创建滚动框架
        canvas = tk.Canvas(dialog)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 输入字段
        entry_vars = {}
        text_widgets = {}
        row = 0

        for col in columns:
            if col == "id" and is_new:
                continue  # 新增时跳过ID字段

            label_text = field_labels.get(col, col)
            ttk.Label(scrollable_frame, text=f"{label_text}:").grid(
                row=row, column=0, sticky="e", padx=5, pady=5
            )

            # 映射匹配项使用多行输入
            if col == "match_items":
                preset = ""
                if record_data and col in record_data and record_data[col]:
                    raw_val = record_data[col]
                    try:
                        parsed = json.loads(raw_val) if isinstance(raw_val, str) else (raw_val or [])
                        if isinstance(parsed, list):
                            preset = "\n".join(str(i).strip() for i in parsed if str(i).strip())
                        elif parsed:
                            preset = str(parsed)
                    except Exception:
                        preset = str(raw_val)

                text = tk.Text(scrollable_frame, height=3, width=40, wrap="word")
                text.grid(row=row, column=1, sticky="we", padx=5, pady=5)
                if preset:
                    text.insert("1.0", preset)
                text_widgets[col] = text
                row += 1
                continue

            var = tk.StringVar()
            if record_data and col in record_data:
                value = record_data[col]
                var.set(str(value) if value is not None else "")

            entry = ttk.Entry(scrollable_frame, textvariable=var, width=40)
            entry.grid(row=row, column=1, sticky="we", padx=5, pady=5)

            if col == "id" and not is_new:
                entry.config(state="readonly")  # ID字段只读

            entry_vars[col] = var
            row += 1

        scrollable_frame.grid_columnconfigure(1, weight=1)

        # 布局滚动框架
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

        # 按钮框
        button_frame = ttk.Frame(dialog)
        button_frame.pack(fill="x", padx=10, pady=10)

        def save_record():
            # 收集数据
            data = {}
            for col, var in entry_vars.items():
                value = var.get().strip()
                # 空字符串转为None，数字字段尝试转换
                if value == "":
                    data[col] = None
                else:
                    # 尝试转换数字字段
                    if col in ["exchange_rate", "pack_qty", "unit_conversion_denominator",
                               "unit_conversion_numerator", "in_price", "out_price",
                               "price_a", "price_b", "price_c", "length", "width",
                               "height", "volume", "weight"]:
                        try:
                            data[col] = float(value) if value else None
                        except ValueError:
                            data[col] = value
                    else:
                        data[col] = value

            # 处理 match_items
            for col, widget in text_widgets.items():
                raw = widget.get("1.0", "end").strip()
                if not raw:
                    data[col] = "[]"
                else:
                    parts = re.split(r"[\\n,，;；、]+", raw)
                    cleaned = [p.strip() for p in parts if p and p.strip()]
                    data[col] = json.dumps(cleaned, ensure_ascii=False)

            # 保存
            if is_new:
                result = self.base_data_mgr.add_record(table_name, data)
            else:
                record_id = int(entry_vars["id"].get())
                result = self.base_data_mgr.update_record(table_name, record_id, data)

            if result["success"]:
                messagebox.showinfo("成功", result["message"])
                dialog.destroy()
                refresh_callback()  # 刷新列表
            else:
                messagebox.showerror("错误", result["message"])

        ttk.Button(button_frame, text="保存", command=save_record).pack(side="right", padx=5)
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side="right")


def main():
    root = tk.Tk()
    app = ExcelConverterGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
