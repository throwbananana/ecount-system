# -*- coding: utf-8 -*-
"""
图片智能识别模块
支持从图片中提取表格数据，并转换为结构化Excel格式
"""

import os
import re
import sys
import json
import base64
import subprocess
from typing import Optional, Dict, List, Any, Tuple
from pathlib import Path
from export_format_manager import apply_export_format

# 尝试导入图像处理库
try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# 全局变量记录可用的OCR引擎
HAS_PADDLEOCR = False
HAS_EASYOCR = False
HAS_TESSERACT = False
HAS_ZHIPUAI = False
HAS_OPENAI = False

# Tesseract 可执行文件路径（Windows默认安装路径）
TESSERACT_CMD_PATHS = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    r"D:\Program Files\Tesseract-OCR\tesseract.exe",
    "/usr/bin/tesseract",
    "/usr/local/bin/tesseract",
]

def check_and_install_dependencies(auto_install: bool = True) -> Dict[str, bool]:
    """
    检查并自动安装依赖

    Args:
        auto_install: 是否自动安装缺失的依赖

    Returns:
        各依赖的可用状态字典
    """
    global HAS_PIL, HAS_PADDLEOCR, HAS_EASYOCR, HAS_TESSERACT, HAS_ZHIPUAI, HAS_OPENAI

    status = {
        "pillow": False,
        "pytesseract": False,
        "tesseract_exe": False,
        "paddleocr": False,
        "easyocr": False,
        "zhipuai": False,
        "openai": False
    }

    def install_package(package_name: str, pip_name: str = None) -> bool:
        """使用pip安装包"""
        pip_name = pip_name or package_name
        try:
            print(f"正在安装 {pip_name}...")
            result = subprocess.run(
                [sys.executable, "-m", "pip", "install", pip_name, "-q"],
                capture_output=True,
                text=True,
                timeout=300
            )
            if result.returncode == 0:
                print(f"  ✓ {pip_name} 安装成功")
                return True
            else:
                print(f"  ✗ {pip_name} 安装失败: {result.stderr}")
                return False
        except subprocess.TimeoutExpired:
            print(f"  ✗ {pip_name} 安装超时")
            return False
        except Exception as e:
            print(f"  ✗ {pip_name} 安装失败: {e}")
            return False

    # 检查 Pillow
    try:
        from PIL import Image
        HAS_PIL = True
        status["pillow"] = True
    except ImportError:
        if auto_install and install_package("pillow", "Pillow"):
            try:
                from PIL import Image
                HAS_PIL = True
                status["pillow"] = True
            except:
                pass

    # 检查 zhipuai (优先级最高，用于AI视觉识别)
    try:
        from zhipuai import ZhipuAI
        HAS_ZHIPUAI = True
        status["zhipuai"] = True
    except ImportError:
        if auto_install and install_package("zhipuai"):
            try:
                from zhipuai import ZhipuAI
                HAS_ZHIPUAI = True
                status["zhipuai"] = True
            except:
                pass

    # 检查 openai (用于LM Studio本地视觉模型)
    try:
        from openai import OpenAI
        HAS_OPENAI = True
        status["openai"] = True
    except ImportError:
        if auto_install and install_package("openai"):
            try:
                from openai import OpenAI
                HAS_OPENAI = True
                status["openai"] = True
            except:
                pass

    # 检查 PaddleOCR (可选，本地OCR备选方案)
    try:
        from paddleocr import PaddleOCR
        HAS_PADDLEOCR = True
        status["paddleocr"] = True
    except ImportError:
        # PaddleOCR 安装较复杂，不自动安装，仅提示
        pass

    # 检查 EasyOCR (可选，本地OCR备选方案)
    try:
        import easyocr
        HAS_EASYOCR = True
        status["easyocr"] = True
    except ImportError:
        # EasyOCR 安装较大，不自动安装
        pass

    # 检查 Tesseract OCR
    # 1. 检查 pytesseract 库
    try:
        import pytesseract
        status["pytesseract"] = True

        # 2. 检查 Tesseract 可执行文件
        tesseract_found = False

        # 尝试默认路径
        for path in TESSERACT_CMD_PATHS:
            if os.path.exists(path):
                pytesseract.pytesseract.tesseract_cmd = path
                tesseract_found = True
                break

        # 尝试从 PATH 环境变量查找
        if not tesseract_found:
            try:
                result = subprocess.run(
                    ["tesseract", "--version"],
                    capture_output=True,
                    text=True,
                    timeout=5
                )
                if result.returncode == 0:
                    tesseract_found = True
            except:
                pass

        if tesseract_found:
            HAS_TESSERACT = True
            status["tesseract_exe"] = True
        else:
            print("[WARN] pytesseract 已安装，但未找到 Tesseract 可执行文件")
            print("       请从 https://github.com/UB-Mannheim/tesseract/wiki 下载安装")

    except ImportError:
        if auto_install and install_package("pytesseract"):
            try:
                import pytesseract
                status["pytesseract"] = True
                # 安装后再检查可执行文件
                for path in TESSERACT_CMD_PATHS:
                    if os.path.exists(path):
                        pytesseract.pytesseract.tesseract_cmd = path
                        HAS_TESSERACT = True
                        status["tesseract_exe"] = True
                        break
            except:
                pass

    return status


class ImageIntelligence:
    """图片智能识别器"""

    # 支持的识别引擎
    ENGINE_ZHIPU = "zhipu"          # 智谱AI视觉模型
    ENGINE_LM_STUDIO = "lm_studio"  # LM Studio本地视觉模型
    ENGINE_TESSERACT = "tesseract"  # Tesseract OCR
    ENGINE_PADDLEOCR = "paddleocr"  # PaddleOCR
    ENGINE_EASYOCR = "easyocr"      # EasyOCR

    def __init__(self,
                 ai_provider: str = "zhipu",
                 api_key: str = "",
                 base_url: str = "http://localhost:1234/v1",
                 model_name: str = "local-model",
                 tesseract_cmd: str = "",
                 tesseract_lang: str = "chi_sim+eng+por",
                 default_engine: str = "auto",
                 auto_install: bool = False):
        """
        初始化图片识别器

        Args:
            ai_provider: AI提供商 ("zhipu" / "lm_studio")
            api_key: API密钥
            base_url: LM Studio的Base URL
            model_name: 模型名称
            tesseract_cmd: Tesseract可执行文件路径（留空则自动检测）
            tesseract_lang: Tesseract语言包（默认中英葡）
            default_engine: 默认识别引擎 ("auto" / "zhipu" / "lm_studio" / "tesseract" / "paddleocr" / "easyocr")
            auto_install: 是否自动安装缺失依赖
        """
        self.ai_provider = ai_provider
        self.api_key = api_key
        self.base_url = base_url
        self.model_name = model_name
        self.tesseract_cmd = tesseract_cmd
        self.tesseract_lang = tesseract_lang
        self.default_engine = default_engine
        self.ai_client = None
        self.ocr_engine = None
        self.tesseract_ocr = None

        # 检查并安装依赖
        self.dep_status = check_and_install_dependencies(auto_install)

        # 初始化各识别引擎
        self._init_ai_client()
        self._init_tesseract()

    def _init_ai_client(self):
        """初始化AI客户端"""
        global HAS_ZHIPUAI, HAS_OPENAI, HAS_PADDLEOCR, HAS_EASYOCR

        if self.ai_provider == "zhipu" and HAS_ZHIPUAI:
            try:
                from zhipuai import ZhipuAI
                self.ai_client = ZhipuAI(api_key=self.api_key)
                print("[OK] 智谱AI视觉客户端初始化成功")
            except Exception as e:
                print(f"[ERR] 智谱AI初始化失败: {e}")

        elif self.ai_provider == "lm_studio" and HAS_OPENAI:
            try:
                from openai import OpenAI
                effective_key = self.api_key if self.api_key else "lm-studio"
                self.ai_client = OpenAI(base_url=self.base_url, api_key=effective_key)
                print("[OK] LM Studio视觉客户端初始化成功")
            except Exception as e:
                print(f"[ERR] LM Studio初始化失败: {e}")

        elif self.ai_provider == "local_ocr":
            self._init_local_ocr()

    def _init_local_ocr(self):
        """初始化本地OCR引擎"""
        global HAS_PADDLEOCR, HAS_EASYOCR

        if HAS_PADDLEOCR:
            try:
                from paddleocr import PaddleOCR
                self.ocr_engine = PaddleOCR(
                    use_angle_cls=True,
                    lang='ch',
                    use_gpu=False,
                    show_log=False
                )
                print("[OK] PaddleOCR初始化成功")
                return
            except Exception as e:
                print(f"[WARN] PaddleOCR初始化失败: {e}")

        if HAS_EASYOCR:
            try:
                import easyocr
                self.ocr_engine = easyocr.Reader(['ch_sim', 'en', 'pt'], gpu=False)
                print("[OK] EasyOCR初始化成功")
                return
            except Exception as e:
                print(f"[WARN] EasyOCR初始化失败: {e}")

        print("[WARN] 没有可用的本地OCR引擎")

    def _init_tesseract(self):
        """初始化Tesseract OCR"""
        global HAS_TESSERACT

        if not HAS_TESSERACT:
            return

        try:
            import pytesseract

            # 如果指定了路径，使用指定路径
            if self.tesseract_cmd and os.path.exists(self.tesseract_cmd):
                pytesseract.pytesseract.tesseract_cmd = self.tesseract_cmd
            else:
                # 自动检测
                for path in TESSERACT_CMD_PATHS:
                    if os.path.exists(path):
                        pytesseract.pytesseract.tesseract_cmd = path
                        self.tesseract_cmd = path
                        break

            self.tesseract_ocr = pytesseract
            print(f"[OK] Tesseract OCR初始化成功 (语言: {self.tesseract_lang})")

        except Exception as e:
            print(f"[ERR] Tesseract初始化失败: {e}")

    def update_config(self, provider=None, api_key=None, base_url=None, model_name=None,
                      tesseract_cmd=None, tesseract_lang=None, default_engine=None):
        """更新配置"""
        if provider:
            self.ai_provider = provider
        if api_key:
            self.api_key = api_key
        if base_url:
            self.base_url = base_url
        if model_name:
            self.model_name = model_name
        if tesseract_cmd:
            self.tesseract_cmd = tesseract_cmd
        if tesseract_lang:
            self.tesseract_lang = tesseract_lang
        if default_engine:
            self.default_engine = default_engine
        self._init_ai_client()
        self._init_tesseract()

    def _encode_image_base64(self, image_path: str) -> str:
        """将图片编码为base64"""
        with open(image_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")

    def _get_image_mime_type(self, image_path: str) -> str:
        """获取图片MIME类型"""
        ext = Path(image_path).suffix.lower()
        mime_types = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".gif": "image/gif",
            ".webp": "image/webp",
            ".bmp": "image/bmp"
        }
        return mime_types.get(ext, "image/jpeg")

    def recognize_image_with_ai(self, image_path: str, prompt: str = None) -> Dict[str, Any]:
        """
        使用AI视觉模型识别图片

        Args:
            image_path: 图片路径
            prompt: 自定义提示词

        Returns:
            识别结果字典，包含 raw_text, table_data, status
        """
        if not os.path.exists(image_path):
            return {"status": "error", "message": f"文件不存在: {image_path}"}

        # 默认提示词 - 针对财务表格优化
        if not prompt:
            prompt = """请仔细分析这张图片，这是一份财务/银行转账记录表格。

请按以下要求提取数据：
1. 识别表格的所有列标题（如：银行账户、PIX、收款人、金额等）
2. 逐行提取每条记录的数据
3. 保持数据的原始格式（特别是金额、日期、账号等）

请以JSON格式返回结果，格式如下：
{
    "headers": ["列1", "列2", "列3", ...],
    "rows": [
        ["数据1", "数据2", "数据3", ...],
        ["数据1", "数据2", "数据3", ...],
        ...
    ],
    "summary": "表格简要描述"
}

注意：
- 金额请保持原格式（如 R$ 1.000,00）
- 如果某个单元格为空，用空字符串""表示
- 如果无法识别某个值，用"[?]"表示"""

        try:
            base64_image = self._encode_image_base64(image_path)
            mime_type = self._get_image_mime_type(image_path)

            if self.ai_provider == "zhipu" and self.ai_client:
                return self._recognize_with_zhipu(base64_image, mime_type, prompt)
            elif self.ai_provider == "lm_studio" and self.ai_client:
                return self._recognize_with_openai(base64_image, mime_type, prompt)
            else:
                return {"status": "error", "message": "没有可用的AI客户端"}

        except Exception as e:
            return {"status": "error", "message": f"识别失败: {str(e)}"}

    def _recognize_with_zhipu(self, base64_image: str, mime_type: str, prompt: str) -> Dict[str, Any]:
        """使用智谱AI进行图片识别"""
        try:
            response = self.ai_client.chat.completions.create(
                model="glm-4v-flash",  # 智谱视觉模型
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:{mime_type};base64,{base64_image}"
                                }
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ]
                    }
                ],
                temperature=0.1,
                max_tokens=4096
            )

            result_text = response.choices[0].message.content
            return self._parse_ai_response(result_text)

        except Exception as e:
            return {"status": "error", "message": f"智谱AI识别失败: {str(e)}"}

    def _recognize_with_openai(self, base64_image: str, mime_type: str, prompt: str) -> Dict[str, Any]:
        """使用OpenAI兼容接口进行图片识别"""
        try:
            response = self.ai_client.chat.completions.create(
                model=self.model_name,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:{mime_type};base64,{base64_image}"
                                }
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ]
                    }
                ],
                temperature=0.1,
                max_tokens=4096
            )

            result_text = response.choices[0].message.content
            return self._parse_ai_response(result_text)

        except Exception as e:
            return {"status": "error", "message": f"OpenAI识别失败: {str(e)}"}

    def _parse_ai_response(self, response_text: str) -> Dict[str, Any]:
        """解析AI响应"""
        try:
            # 尝试提取JSON部分
            json_match = re.search(r'\{[\s\S]*\}', response_text)
            if json_match:
                json_str = json_match.group(0)
                data = json.loads(json_str)
                return {
                    "status": "success",
                    "raw_text": response_text,
                    "headers": data.get("headers", []),
                    "rows": data.get("rows", []),
                    "summary": data.get("summary", "")
                }
            else:
                # 如果没有JSON，返回原始文本
                return {
                    "status": "partial",
                    "raw_text": response_text,
                    "headers": [],
                    "rows": [],
                    "message": "未能解析为表格格式，返回原始识别文本"
                }
        except json.JSONDecodeError as e:
            return {
                "status": "partial",
                "raw_text": response_text,
                "headers": [],
                "rows": [],
                "message": f"JSON解析失败: {e}"
            }

    def recognize_with_local_ocr(self, image_path: str) -> Dict[str, Any]:
        """使用本地OCR识别图片"""
        global HAS_PADDLEOCR, HAS_EASYOCR

        if not os.path.exists(image_path):
            return {"status": "error", "message": f"文件不存在: {image_path}"}

        if not self.ocr_engine:
            self._init_local_ocr()

        if not self.ocr_engine:
            return {"status": "error", "message": "没有可用的本地OCR引擎，请安装paddleocr或easyocr"}

        try:
            if HAS_PADDLEOCR and hasattr(self.ocr_engine, 'ocr'):
                # PaddleOCR
                result = self.ocr_engine.ocr(image_path, cls=True)
                lines = []
                if result and result[0]:
                    for line in result[0]:
                        text = line[1][0]
                        lines.append(text)
                return {
                    "status": "success",
                    "raw_text": "\n".join(lines),
                    "lines": lines,
                    "message": "PaddleOCR识别完成"
                }

            elif HAS_EASYOCR and hasattr(self.ocr_engine, 'readtext'):
                # EasyOCR
                result = self.ocr_engine.readtext(image_path)
                lines = [item[1] for item in result]
                return {
                    "status": "success",
                    "raw_text": "\n".join(lines),
                    "lines": lines,
                    "message": "EasyOCR识别完成"
                }

            else:
                return {"status": "error", "message": "OCR引擎不可用"}

        except Exception as e:
            return {"status": "error", "message": f"本地OCR识别失败: {str(e)}"}

    def recognize_with_tesseract(self, image_path: str) -> Dict[str, Any]:
        """使用Tesseract OCR识别图片"""
        global HAS_TESSERACT

        if not os.path.exists(image_path):
            return {"status": "error", "message": f"文件不存在: {image_path}"}

        if not HAS_TESSERACT or not self.tesseract_ocr:
            self._init_tesseract()

        if not self.tesseract_ocr:
            return {"status": "error", "message": "Tesseract OCR 不可用，请确保已安装 Tesseract"}

        try:
            from PIL import Image

            # 打开图片
            img = Image.open(image_path)

            # 使用 Tesseract 识别
            # 配置参数：PSM 6 = 假设为单一文本块
            config = f'--psm 6 -l {self.tesseract_lang}'

            # 识别文本
            text = self.tesseract_ocr.image_to_string(img, config=config)

            # 尝试识别表格数据（使用 image_to_data）
            try:
                data = self.tesseract_ocr.image_to_data(img, config=config, output_type=self.tesseract_ocr.Output.DICT)

                # 解析表格结构
                lines = []
                current_line = []
                prev_line_num = -1

                for i, word in enumerate(data['text']):
                    if word.strip():
                        line_num = data['line_num'][i]
                        if line_num != prev_line_num and current_line:
                            lines.append(current_line)
                            current_line = []
                        current_line.append(word.strip())
                        prev_line_num = line_num

                if current_line:
                    lines.append(current_line)

                # 尝试转换为表格格式
                if lines:
                    # 假设第一行是表头
                    headers = lines[0] if len(lines[0]) > 1 else []
                    rows = lines[1:] if len(lines) > 1 else []

                    return {
                        "status": "success",
                        "raw_text": text,
                        "headers": headers,
                        "rows": rows,
                        "lines": lines,
                        "message": f"Tesseract识别完成 (语言: {self.tesseract_lang})"
                    }

            except Exception:
                pass

            # 如果表格识别失败，返回纯文本
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            return {
                "status": "partial",
                "raw_text": text,
                "headers": [],
                "rows": [],
                "lines": lines,
                "message": "Tesseract识别完成（纯文本模式）"
            }

        except Exception as e:
            return {"status": "error", "message": f"Tesseract识别失败: {str(e)}"}

    def get_available_engines(self) -> Dict[str, bool]:
        """获取可用的识别引擎状态"""
        global HAS_TESSERACT, HAS_PADDLEOCR, HAS_EASYOCR, HAS_ZHIPUAI, HAS_OPENAI

        return {
            "zhipu": HAS_ZHIPUAI and bool(self.api_key),
            "lm_studio": HAS_OPENAI,
            "tesseract": HAS_TESSERACT,
            "paddleocr": HAS_PADDLEOCR,
            "easyocr": HAS_EASYOCR
        }

    def recognize_image(self, image_path: str, use_ai: bool = True, custom_prompt: str = None,
                        engine: str = None) -> Dict[str, Any]:
        """
        智能识别图片

        Args:
            image_path: 图片路径
            use_ai: 是否优先使用AI识别（当 engine=None 时有效）
            custom_prompt: 自定义提示词（仅AI识别时有效）
            engine: 指定识别引擎 (None=自动选择, "zhipu", "lm_studio", "tesseract", "paddleocr", "easyocr")

        Returns:
            识别结果
        """
        # 如果指定了引擎，直接使用
        if engine:
            return self._recognize_with_engine(image_path, engine, custom_prompt)

        # 使用默认引擎设置
        if self.default_engine and self.default_engine != "auto":
            result = self._recognize_with_engine(image_path, self.default_engine, custom_prompt)
            if result.get("status") in ["success", "partial"]:
                return result

        # 自动选择模式
        if use_ai and self.ai_client:
            result = self.recognize_image_with_ai(image_path, custom_prompt)
            if result.get("status") in ["success", "partial"]:
                return result

        # AI失败后尝试 Tesseract
        if HAS_TESSERACT and self.tesseract_ocr:
            result = self.recognize_with_tesseract(image_path)
            if result.get("status") in ["success", "partial"]:
                return result

        # 最后尝试其他本地OCR
        if self.ocr_engine:
            return self.recognize_with_local_ocr(image_path)

        return {"status": "error", "message": "没有可用的识别方案"}

    def _recognize_with_engine(self, image_path: str, engine: str, custom_prompt: str = None) -> Dict[str, Any]:
        """使用指定引擎识别"""
        if engine == self.ENGINE_ZHIPU:
            if self.ai_provider != "zhipu":
                # 临时切换
                old_provider = self.ai_provider
                self.ai_provider = "zhipu"
                self._init_ai_client()
            return self.recognize_image_with_ai(image_path, custom_prompt)

        elif engine == self.ENGINE_LM_STUDIO:
            if self.ai_provider != "lm_studio":
                old_provider = self.ai_provider
                self.ai_provider = "lm_studio"
                self._init_ai_client()
            return self.recognize_image_with_ai(image_path, custom_prompt)

        elif engine == self.ENGINE_TESSERACT:
            return self.recognize_with_tesseract(image_path)

        elif engine == self.ENGINE_PADDLEOCR:
            if not self.ocr_engine:
                self._init_local_ocr()
            if HAS_PADDLEOCR and self.ocr_engine:
                return self.recognize_with_local_ocr(image_path)
            return {"status": "error", "message": "PaddleOCR 不可用"}

        elif engine == self.ENGINE_EASYOCR:
            if not self.ocr_engine:
                self._init_local_ocr()
            if HAS_EASYOCR and self.ocr_engine:
                return self.recognize_with_local_ocr(image_path)
            return {"status": "error", "message": "EasyOCR 不可用"}

        else:
            return {"status": "error", "message": f"未知的识别引擎: {engine}"}

    def batch_recognize(self, image_paths: List[str], use_ai: bool = True) -> List[Dict[str, Any]]:
        """
        批量识别图片

        Args:
            image_paths: 图片路径列表
            use_ai: 是否使用AI识别

        Returns:
            识别结果列表
        """
        results = []
        for i, path in enumerate(image_paths):
            print(f"正在识别 ({i+1}/{len(image_paths)}): {os.path.basename(path)}")
            result = self.recognize_image(path, use_ai)
            result["file_path"] = path
            result["file_name"] = os.path.basename(path)
            results.append(result)
        return results

    def merge_results_to_table(self, results: List[Dict[str, Any]],
                                smart_merge: bool = True) -> Tuple[List[str], List[List[str]]]:
        """
        合并多个识别结果为统一表格

        Args:
            results: 识别结果列表
            smart_merge: 是否启用智能合并（处理不同表头的情况）

        Returns:
            (headers, rows) 元组
        """
        all_headers = []
        all_rows = []
        header_mapping = {}  # 用于映射不同表头

        for result in results:
            if result.get("status") in ["success", "partial"]:
                headers = result.get("headers", [])
                rows = result.get("rows", [])

                if not headers or not rows:
                    continue

                # 第一个有效结果的headers作为基准
                if not all_headers:
                    all_headers = headers
                    header_mapping = {h: i for i, h in enumerate(headers)}
                elif smart_merge:
                    # 智能合并：处理不同表头
                    for h in headers:
                        if h not in header_mapping:
                            # 尝试模糊匹配
                            matched = False
                            for existing_h in all_headers:
                                if self._is_similar_header(h, existing_h):
                                    header_mapping[h] = header_mapping[existing_h]
                                    matched = True
                                    break
                            if not matched:
                                # 新表头，添加到末尾
                                header_mapping[h] = len(all_headers)
                                all_headers.append(h)

                # 合并数据行
                for row in rows:
                    if smart_merge and headers != all_headers:
                        # 需要重新映射列
                        new_row = [""] * len(all_headers)
                        for i, val in enumerate(row):
                            if i < len(headers):
                                col_name = headers[i]
                                if col_name in header_mapping:
                                    new_row[header_mapping[col_name]] = val
                        all_rows.append(new_row)
                    else:
                        # 直接添加，补齐列数
                        padded_row = list(row) + [""] * (len(all_headers) - len(row))
                        all_rows.append(padded_row[:len(all_headers)])

        return all_headers, all_rows

    def batch_recognize_and_merge(self, image_paths: List[str], use_ai: bool = True,
                                   output_path: str = None) -> Dict[str, Any]:
        """
        一键批量识别并合并导出

        Args:
            image_paths: 图片路径列表
            use_ai: 是否使用AI识别
            output_path: 输出Excel路径（可选）

        Returns:
            包含合并结果和统计信息的字典
        """
        # 批量识别
        results = self.batch_recognize(image_paths, use_ai)

        # 合并结果
        headers, rows = self.merge_results_to_table(results, smart_merge=True)

        # 统计
        success_count = sum(1 for r in results if r.get("status") == "success")
        partial_count = sum(1 for r in results if r.get("status") == "partial")
        error_count = sum(1 for r in results if r.get("status") == "error")

        output = {
            "status": "success" if rows else "no_data",
            "headers": headers,
            "rows": rows,
            "total_images": len(image_paths),
            "success_count": success_count,
            "partial_count": partial_count,
            "error_count": error_count,
            "total_rows": len(rows),
            "results": results  # 保留原始结果以便查看详情
        }

        # 如果指定了输出路径，自动导出
        if output_path and rows:
            export_success = self.export_to_excel(headers, rows, output_path)
            output["exported"] = export_success
            output["export_path"] = output_path if export_success else None

        return output

    def export_to_excel(self, headers: List[str], rows: List[List[str]], output_path: str,
                       template_path: str = None) -> bool:
        """
        导出识别结果到Excel

        Args:
            headers: 表头列表
            rows: 数据行列表
            output_path: 输出文件路径
            template_path: 可选的模板文件路径

        Returns:
            是否成功
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook, Workbook

            headers, rows, _ = apply_export_format("image_recognition", headers or [], rows or [])

            if template_path and os.path.exists(template_path):
                # 使用模板
                wb = load_workbook(template_path)
                ws = wb.active

                # 获取模板表头
                template_headers = []
                for cell in ws[1]:
                    if cell.value:
                        template_headers.append(str(cell.value).strip())

                # 清除模板数据（保留表头）
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row)

                # 建立列映射
                col_mapping = {}
                for i, th in enumerate(template_headers):
                    for j, sh in enumerate(headers):
                        if self._is_similar_header(th, sh):
                            col_mapping[j] = i
                            break

                # 写入数据
                for row_idx, row_data in enumerate(rows, start=2):
                    for src_col, tgt_col in col_mapping.items():
                        if src_col < len(row_data):
                            ws.cell(row_idx, tgt_col + 1).value = row_data[src_col]

                wb.save(output_path)

            else:
                # 直接创建新文件
                df = pd.DataFrame(rows, columns=headers if headers else None)
                df.to_excel(output_path, index=False)

            return True

        except Exception as e:
            print(f"导出Excel失败: {e}")
            return False

    def _is_similar_header(self, h1: str, h2: str) -> bool:
        """判断两个表头是否相似"""
        h1 = h1.lower().strip()
        h2 = h2.lower().strip()

        if h1 == h2:
            return True
        if h1 in h2 or h2 in h1:
            return True

        # 常见同义词映射
        synonyms = {
            "银行": ["banco", "bank", "conta"],
            "金额": ["valor", "amount", "total", "vale"],
            "日期": ["data", "date", "fecha"],
            "收款人": ["beneficiario", "destinatario", "nome"],
            "pix": ["pix", "chave pix"],
        }

        for key, values in synonyms.items():
            if key in h1 or key in h2:
                for v in values:
                    if v in h1 or v in h2:
                        return True

        return False


def get_available_engines() -> Dict[str, bool]:
    """获取可用的识别引擎状态"""
    return check_and_install_dependencies(auto_install=False)


# 便捷函数
def recognize_image_simple(image_path: str, api_key: str = "") -> Dict[str, Any]:
    """
    简单的图片识别函数

    Args:
        image_path: 图片路径
        api_key: 智谱AI API Key

    Returns:
        识别结果
    """
    recognizer = ImageIntelligence(
        ai_provider="zhipu",
        api_key=api_key,
        auto_install=True
    )
    return recognizer.recognize_image(image_path)


if __name__ == "__main__":
    # 测试代码
    print("检查依赖状态...")
    status = check_and_install_dependencies(auto_install=False)
    print(f"依赖状态: {status}")

    print("\n可用的识别引擎:")
    for engine, available in status.items():
        print(f"  {engine}: {'✓' if available else '✗'}")
