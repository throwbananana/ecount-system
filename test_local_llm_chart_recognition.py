import tempfile
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, Reference

from local_llm_analyzer import LocalLLMAnalyzer


def _build_workbook_with_chart(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "趋势分析"
    ws.append(["月份", "销售额"])
    ws.append(["2025-10", 120])
    ws.append(["2025-11", 180])
    ws.append(["2025-12", 240])

    chart = BarChart()
    chart.title = "销售趋势"
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E2")
    wb.save(path)
    wb.close()


def _build_chart_only_workbook(path: Path):
    wb = openpyxl.Workbook()
    ws_data = wb.active
    ws_data.title = "数据源"
    ws_data.append(["月份", "毛利"])
    ws_data.append(["2025-10", 12])
    ws_data.append(["2025-11", 15])
    ws_data.append(["2025-12", 18])

    ws_chart = wb.create_sheet("图表页")
    chart = BarChart()
    chart.title = "毛利趋势"
    data = Reference(ws_data, min_col=2, min_row=1, max_row=4)
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_chart.add_chart(chart, "A1")
    wb.save(path)
    wb.close()


def test_collect_chart_context_map():
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "chart.xlsx"
        _build_workbook_with_chart(file_path)

        analyzer = LocalLLMAnalyzer(enable_chart_recognition=True)
        context_map = analyzer._collect_sheet_chart_context_map(str(file_path), ["趋势分析"])

        assert "趋势分析" in context_map
        text = context_map["趋势分析"]
        assert "图表识别" in text
        assert "销售趋势" in text
        assert "数值范围" in text


def test_collect_sheet_analyses_includes_chart_context():
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "chart.xlsx"
        _build_workbook_with_chart(file_path)

        analyzer = LocalLLMAnalyzer(enable_chart_recognition=True)
        captured = {}

        def _fake_analyze_dataframe(df, sheet_name, context=""):
            captured[sheet_name] = context
            return "【分析】\n- ok\n【结论】\n- done"

        analyzer.analyze_dataframe = _fake_analyze_dataframe

        analyses, _ = analyzer._collect_sheet_analyses(
            str(file_path),
            {"趋势分析": "基础上下文"},
            include_related_context=False,
        )

        assert "趋势分析" in analyses
        assert "图表识别" in captured.get("趋势分析", "")


def test_chart_only_sheet_uses_chart_mode():
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "chart_only.xlsx"
        _build_chart_only_workbook(file_path)

        analyzer = LocalLLMAnalyzer(enable_chart_recognition=True)
        called = {"count": 0}

        def _fake_chart_analyze(sheet_name, context=""):
            called["count"] += 1
            assert sheet_name == "图表页"
            assert "图表识别" in context
            return "【分析】\n- chart only\n【结论】\n- check"

        analyzer.analyze_chart_context = _fake_chart_analyze
        analyses, _ = analyzer._collect_sheet_analyses(
            str(file_path),
            {"图表页": "图表页上下文"},
            include_related_context=False,
        )

        assert called["count"] == 1
        assert "图表页" in analyses
        assert "chart only" in analyses["图表页"]["analysis"]


if __name__ == "__main__":
    test_collect_chart_context_map()
    test_collect_sheet_analyses_includes_chart_context()
    test_chart_only_sheet_uses_chart_mode()
    print("PASS: test_local_llm_chart_recognition")
