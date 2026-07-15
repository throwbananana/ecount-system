# -*- coding: utf-8 -*-
"""
经营报告生成模块
用于从基础数据生成标准化的经营分析报告。
"""

import os
import re
import copy
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from openpyxl.chart import LineChart, Reference, BarChart, ScatterChart, Series
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from openpyxl.cell.cell import MergedCell
from datetime import datetime, date
from collections import Counter
import warnings

# Suppress warnings for cleaner output
warnings.filterwarnings('ignore')

class ReportGenerator:
    DETAIL_ROW_MONTH_CATEGORIES = {"expense", "sales", "ar", "ap", "cash"}

    def __init__(self, base_data_dir):
        self.base_data_dir = base_data_dir
        # Data structure: self.data[category][month_str] = dataframe
        # categories: 'profit', 'cost', 'expense', 'asset', 'sales', 'ar', 'ap', 'cash'
        self.data = {
            'profit': {},
            'cost': {},
            'expense': {},
            'asset': {},
            'sales': {},
            'ar': {},
            'ap': {},
            'cash': {},
        }
        # current: only target year; recent_two_years: target year + previous year; all: full history
        self.year_scope = "current"
        self.sales_df = None
        self.ar_detail_df = None
        self.ap_detail_df = None
        self.cash_detail_df = None
        self.audit_logs = []
        self.data_quality_issues = []
        self.report_params = {
            "replenishment": {
                "manual": True,
                "lead_days": 30,
                "safety_days": 20,
                "window_months": 3,
            },
            "cashflow": {
                "manual": True,
                "dso_threshold": 90,
                "dio_threshold": 180,
                "ccc_threshold": 120,
                "cash_coverage_threshold": 1.5,
            },
            "ai": {
                # 非 AI 模式默认不在报表中预留 AI 占位文本。
                "include_placeholders": False,
            },
        }
        self.expense_detail_sheet_name = "费用异常明细"
        self.expense_analysis_sheet_name = "费用分析"
        self.expense_detail_key_row_map = {}

    def _normalize_year_scope(self, year_scope=None):
        raw = self.year_scope if year_scope is None else year_scope
        text = str(raw or "current").strip()
        lowered = text.lower()
        if lowered in {"current", "this_year", "same_year"} or "当前" in text or "不跨" in text:
            return "current"
        if (
            lowered in {"recent_two_years", "max_two_years", "two_years", "2_years", "two-year", "2-year"}
            or "最多2年" in text
            or "最多两年" in text
            or "跨年" in text
        ):
            return "recent_two_years"
        if lowered in {"all", "history", "historical", "full_history"} or "历年" in text or "全部" in text:
            return "all"
        return lowered if lowered else "current"

    def _scope_start_key(self, target_year, year_scope=None):
        scope = self._normalize_year_scope(year_scope)
        if scope == "current":
            return f"{int(target_year):04d}-01"
        if scope == "recent_two_years":
            return f"{int(target_year) - 1:04d}-01"
        return None

    def _scope_uses_history_chart_labels(self, year_scope=None):
        return self._normalize_year_scope(year_scope) in {"recent_two_years", "all"}

    def _should_include_ai_placeholders(self):
        return bool((self.report_params or {}).get("ai", {}).get("include_placeholders", False))

    def _remove_ai_placeholder_texts(self, wb):
        tokens = (
            "AI 智能分析",
            "AI分析",
            "AI结论",
            "此处预留用于 AI",
            "深度洞察与建议",
        )
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    text = val.strip()
                    if not text:
                        continue
                    if any(token in text for token in tokens):
                        cell.value = None

    def _add_pareto_chart(self, ws, cat_col, data_col, cum_col, header_row, data_start, data_end, title, anchor):
        """
        创建帕累托图 (Pareto Chart)
        data_col: 绝对值 (柱状图)
        cum_col: 累计百分比 (折线图, 次轴)
        """
        # 1. Bar Chart (Data)
        bar_chart = BarChart()
        bar_chart.title = title
        bar_chart.y_axis.title = "金额"
        
        cats = self._chart_category_reference(ws, cat_col, data_start, data_end)
        data = Reference(ws, min_col=data_col, min_row=header_row, max_row=data_end)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)

        # 2. Line Chart (Cumulative %)
        line_chart = LineChart()
        line_chart.y_axis.title = "累计占比"
        line_chart.y_axis.axId = 200
        line_chart.y_axis.crosses = "max"
        
        cum_data = Reference(ws, min_col=cum_col, min_row=header_row, max_row=data_end)
        line_chart.add_data(cum_data, titles_from_data=True)

        # 3. Combine
        bar_chart.y_axis.crosses = "min"
        bar_chart += line_chart
        
        bar_chart.height = 10
        bar_chart.width = 18
        ws.add_chart(bar_chart, anchor)
        return True

    def _add_scatter_chart(self, ws, x_col, y_col, header_row, data_start, data_end, title, anchor, x_title=None, y_title=None):
        chart = ScatterChart()
        chart.title = title
        chart.style = 13
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title

        xvalues = Reference(ws, min_col=x_col, min_row=data_start, max_row=data_end)
        yvalues = Reference(ws, min_col=y_col, min_row=data_start, max_row=data_end)
        series_title = ws.cell(row=header_row, column=y_col).value
        if series_title is None or str(series_title).strip() == "":
            series_title = y_title or f"Y列{y_col}"
        series = Series(yvalues, xvalues, title=str(series_title))
        series.marker.symbol = "circle"
        series.marker.graphicalProperties.solidFill = "4472C4"
        series.marker.graphicalProperties.line.noFill = True
        
        chart.series.append(series)
        chart.height = 10
        chart.width = 18
        ws.add_chart(chart, anchor)
        return True

    def _apply_header_style(self, ws, row_idx, start_col=1, max_col=None):
        """Applies a unified header style to a specific row."""
        if max_col is None:
            max_col = ws.max_column
        if start_col < 1:
            start_col = 1

        # Style definition mimicking "按产品汇总_含合计"
        # Blue background, White Bold Text, Center Aligned, Thin Border
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Side(border_style="thin", color="000000")
        header_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        for col in range(start_col, max_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            if cell.value is not None: # Only style cells with content
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = header_border

    def _ensure_header_column(self, ws, header, header_row=1):
        target = str(header).strip()
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val is None:
                continue
            if str(val).strip() == target:
                return col
        col = ws.max_column + 1
        ws.cell(row=header_row, column=col).value = target
        return col

    def _get_expense_mom_column_maps(self, ws, header_row=1):
        value_cols = {}
        rate_cols = {}
        delta_cols = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if not isinstance(header, str):
                continue
            name = header.strip()
            if '_本期值' in name:
                month_label = name.split('_')[0]
                value_cols[month_label.replace('/', '-')] = col
            elif '_环比增速' in name:
                month_label = name.split('_')[0]
                rate_cols[month_label.replace('/', '-')] = col
            elif '_环比增量' in name:
                month_label = name.split('_')[0]
                delta_cols[month_label.replace('/', '-')] = col
        return value_cols, rate_cols, delta_cols

    def _select_expense_display_flags(self, flags, target_year=None, target_month=None, target_only=False):
        selected_by_pair = {}
        if not flags:
            return selected_by_pair

        target_key = None
        if target_year and target_month:
            target_key = f"{target_year}-{int(target_month):02d}"

        pair_map = {}
        for f in flags:
            pair = (f.get("Category"), f.get("Subcategory"))
            pair_map.setdefault(pair, []).append(f)

        for pair, pair_flags in pair_map.items():
            chosen = None
            if target_key:
                target_flags = [x for x in pair_flags if x.get("MonthStr") == target_key]
                if target_flags:
                    target_flags.sort(
                        key=lambda x: (
                            x.get("AnomalyScore") or 0,
                            abs(x.get("Delta") or 0),
                            abs(x.get("Amount") or 0),
                        ),
                        reverse=True,
                    )
                    chosen = target_flags[0]

            if chosen is None and target_only:
                continue

            if chosen is None:
                pair_flags.sort(
                    key=lambda x: (
                        x.get("AnomalyScore") or 0,
                        abs(x.get("Delta") or 0),
                        abs(x.get("Amount") or 0),
                        x.get("MonthStr") or "",
                    ),
                    reverse=True,
                )
                chosen = pair_flags[0] if pair_flags else None

            if chosen is not None:
                selected_by_pair[pair] = chosen

        return selected_by_pair

    def _refresh_expense_mom_conditional_formatting(
        self,
        ws,
        data_start_row=2,
        data_end_row=None,
        rate_threshold=0.5,
        delta_threshold=10000,
    ):
        if data_end_row is None:
            data_end_row = ws.max_row
        if data_end_row < data_start_row:
            return

        try:
            ws.conditional_formatting._cf_rules.clear()
        except Exception:
            pass

        value_cols, rate_cols, delta_cols = self._get_expense_mom_column_maps(ws, header_row=1)
        common_months = sorted(set(value_cols) & set(rate_cols) & set(delta_cols), reverse=True)
        if not common_months:
            return

        red_font = Font(color="00FF0000")
        for m_key in common_months:
            value_col = value_cols[m_key]
            rate_col = rate_cols[m_key]
            delta_col = delta_cols[m_key]
            rate_letter = get_column_letter(rate_col)
            delta_letter = get_column_letter(delta_col)
            formula = f"AND(ABS(${rate_letter}{data_start_row})>{rate_threshold},ABS(${delta_letter}{data_start_row})>{delta_threshold})"

            for col_idx in (value_col, rate_col, delta_col):
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(formula=[formula], font=red_font),
                )

    def _log_audit(self, message):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.audit_logs.append((timestamp, message))

    def get_data_quality_summary(self):
        """返回数据质量问题统计。"""
        counter = Counter()
        for item in self.data_quality_issues:
            severity = str(item.get("severity") or "WARN").upper()
            counter[severity] += 1
        return {
            "ERROR": counter.get("ERROR", 0),
            "WARN": counter.get("WARN", 0),
            "INFO": counter.get("INFO", 0),
            "TOTAL": len(self.data_quality_issues),
        }

    def get_data_quality_issues(self, target_year=None, target_month=None, year_scope=None):
        """返回数据质量问题明细，可按报告目标期间过滤。"""
        issues = []
        for item in self.data_quality_issues:
            period = item.get("period")
            if (
                target_year
                and target_month
                and isinstance(period, str)
                and re.match(r"^20\d{2}-\d{2}$", period)
            ):
                if not self._month_key_in_scope(period, target_year, target_month, year_scope):
                    continue
            issues.append(item)
        return issues

    def format_data_quality_details(
        self,
        target_year=None,
        target_month=None,
        year_scope=None,
        max_per_severity=None,
    ):
        """将数据质量问题格式化为可读明细。"""
        issues = self.get_data_quality_issues(target_year, target_month, year_scope)
        if not issues:
            return "未发现数据质量问题。"

        lines = []
        severity_order = ["ERROR", "WARN", "INFO"]
        for severity in severity_order:
            group = [
                item for item in issues
                if str(item.get("severity") or "WARN").upper() == severity
            ]
            if not group:
                continue

            lines.append(f"{severity} ({len(group)}):")
            shown = group
            if max_per_severity is not None:
                shown = group[:max(0, int(max_per_severity))]

            for idx, item in enumerate(shown, start=1):
                category = item.get("category") or "未知类别"
                period = item.get("period") or "全局"
                issue_type = item.get("issue_type") or "未知问题"
                detail = item.get("detail") or ""
                lines.append(f"  {idx}. [{category}] {period} - {issue_type}: {detail}")

            if len(shown) < len(group):
                lines.append(f"  ... 还有 {len(group) - len(shown)} 条 {severity} 未显示，请查看生成日志或“数据质量检查”Sheet。")

        return "\n".join(lines)

    def _get_data_quality_summary_for_scope(self, target_year=None, target_month=None, year_scope=None):
        counter = Counter()
        total = 0
        for item in self.get_data_quality_issues(target_year, target_month, year_scope):
            severity = str(item.get("severity") or "WARN").upper()
            counter[severity] += 1
            total += 1
        return {
            "ERROR": counter.get("ERROR", 0),
            "WARN": counter.get("WARN", 0),
            "INFO": counter.get("INFO", 0),
            "TOTAL": total,
        }

    def _detect_template_risk(self, wb, template_path):
        """
        检测模板是否疑似“已生成报告文件”，避免把输出文件继续当模板导致口径污染。
        返回: (has_risk: bool, reasons: list[str])
        """
        reasons = []
        base_name = os.path.basename(template_path or "")
        name_l = base_name.lower()
        if (
            "经营分析报告" in base_name
            and "汇总结果" not in base_name
            and "模板" not in base_name
            and "template" not in name_l
        ):
            reasons.append(f"文件名疑似已生成报告: {base_name}")

        if "审计日志" in wb.sheetnames:
            ws = wb["审计日志"]
            audit_tokens = ("开始生成报告", "报告生成成功", "自动校验发现", "自动校验通过")
            max_row = min(ws.max_row, 200)
            for r in range(1, max_row + 1):
                c1 = ws.cell(row=r, column=1).value
                c2 = ws.cell(row=r, column=2).value
                text = f"{c1 or ''} {c2 or ''}"
                if any(token in str(text) for token in audit_tokens):
                    reasons.append("检测到历史生成审计记录，模板可能是已输出文件")
                    break

        return bool(reasons), reasons
        
    def load_all_data(self):
        """加载所有基础资料"""
        self._log_audit("开始加载基础资料")
        print("开始加载基础资料...")
        if not os.path.exists(self.base_data_dir):
            print(f"错误: 目录不存在 {self.base_data_dir}")
            return

        files = os.listdir(self.base_data_dir)
        for f in files:
            if not (f.endswith('.xlsx') or f.endswith('.xls')):
                continue
            if str(f).startswith('~$'):
                continue
            
            path = os.path.join(self.base_data_dir, f)
            print(f"正在处理文件: {f}")
            
            file_type = self._classify_source_file(f, path)
            # Identify type by filename and workbook header content.
            if file_type == "cost":
                self._load_cost_data(path, f)
            elif file_type == "expense":
                self._load_expense_data(path, f)
            elif file_type == "profit":
                self._load_profit_data(path, f)
            elif file_type == "sales":
                self._load_sales_data(path, f)
            elif file_type == "ar":
                self._load_ar_data(path, f)
            elif file_type == "ap":
                self._load_ap_data(path, f)
            elif file_type == "cash":
                self._load_cash_data(path, f)
            elif file_type == "asset":
                self._load_asset_data(path, f)
        
        print("基础资料加载完成。")
        self._log_audit("基础资料加载完成")
        self._print_stats()
        self._run_data_quality_checks()

    def _source_file_text(self, path=None, df_peek=None):
        parts = []
        if df_peek is not None:
            try:
                for value in df_peek.to_numpy().ravel():
                    if value is not None and not pd.isna(value):
                        parts.append(str(value))
            except Exception:
                pass
        elif path:
            try:
                peek = pd.read_excel(path, header=None, nrows=3)
                for value in peek.to_numpy().ravel():
                    if value is not None and not pd.isna(value):
                        parts.append(str(value))
            except Exception:
                pass

        if path:
            try:
                with pd.ExcelFile(path) as xl:
                    parts.extend([str(s) for s in xl.sheet_names])
            except Exception:
                pass

        return " ".join(parts)

    def _classify_source_file(self, filename, path=None, df_peek=None):
        name = str(filename or "")
        text = f"{name} {self._source_file_text(path, df_peek)}"

        if "利润表" in text or "利润" in name:
            return "profit"
        if "资产负债表" in text or "资产" in name:
            return "asset"
        if "销售出库" in text or "销售" in name or "出库" in name:
            return "sales"
        if "成本" in name:
            return "cost"

        # 明细账常见文件名不包含“应收/应付/银行”，需要优先看首行账簿描述。
        if "应付账款" in text or "应付" in name:
            return "ap"
        if "应收账款" in text or "应收" in name:
            return "ar"
        if "银行存款" in text or "货币资金" in text:
            return "cash"

        # 科目账簿类费用文件可能只在首行出现 6601/6602/6603 或“办公费”等科目名。
        expense_terms = (
            "销售费用", "管理费用", "财务费用", "办公费", "工资", "福利费", "房租",
            "水电费", "差旅费", "保险费", "手续费", "汇兑损益",
        )
        if "费用" in name or re.search(r"\b660[123]\d*\b", text) or any(term in text for term in expense_terms):
            return "expense"
        return None

    def list_available_months(self, ready_only=True):
        """返回可识别月份键列表 (YYYY-MM)。ready_only=True 时仅返回报表可用月份。"""
        # 月份候选以核心三表为准，避免被跨年明细文件(如应收明细)误导。
        core_required = ['profit', 'cost', 'asset']

        # 优先使用已加载的数据，避免仅靠文件名/表头推断导致年份误判。
        has_loaded_data = any(bool(months) for months in self.data.values())
        if has_loaded_data:
            if ready_only:
                month_sets = [set(self.data[cat].keys()) for cat in core_required if self.data.get(cat)]
                if not month_sets:
                    return []
                return sorted(set.intersection(*month_sets))
            all_months = set()
            for months in self.data.values():
                all_months.update(months.keys())
            return sorted(all_months)

        if not os.path.exists(self.base_data_dir):
            return []
        try:
            files = os.listdir(self.base_data_dir)
        except Exception:
            return []

        month_by_cat = {cat: set() for cat in core_required}
        for f in files:
            if not (f.endswith('.xlsx') or f.endswith('.xls')):
                continue
            if str(f).startswith('~$'):
                continue
            path = os.path.join(self.base_data_dir, f)
            df_peek = None
            try:
                df_peek = pd.read_excel(path, header=None, nrows=1)
            except Exception:
                df_peek = None
            cat = self._classify_source_file(f, path, df_peek)
            if cat not in month_by_cat:
                continue
            month_key = self._determine_period_key(path, f, df_peek)
            if month_key and month_key != "Unknown":
                month_by_cat[cat].add(month_key)

        if ready_only:
            month_sets = [month_by_cat[cat] for cat in core_required if month_by_cat[cat]]
            if not month_sets:
                return []
            months = set.intersection(*month_sets)
            if not months:
                months = set().union(*(month_by_cat[cat] for cat in core_required))
            return sorted(months)

        months = set()
        for values in month_by_cat.values():
            months.update(values)
        return sorted(months)

    def list_available_years(self):
        """根据可识别月份返回年份列表 (int)。默认返回报表可用年份。"""
        years = set()
        for m in self.list_available_months(ready_only=True):
            if isinstance(m, str) and len(m) >= 7 and m[:4].isdigit():
                years.add(int(m[:4]))
        return sorted(years)

    @staticmethod
    def normalize_month_key(year, month):
        """将年月参数标准化为 YYYY-MM。"""
        year_int = int(str(year).strip())
        month_int = int(str(month).strip())
        if month_int < 1 or month_int > 12:
            raise ValueError(f"月份必须在 1-12 之间: {month}")
        return f"{year_int:04d}-{month_int:02d}"

    @classmethod
    def build_month_range(cls, start_year, start_month, end_year, end_month):
        """返回起止年月之间的连续月份键列表，包含首尾。"""
        start_key = cls.normalize_month_key(start_year, start_month)
        end_key = cls.normalize_month_key(end_year, end_month)
        if start_key > end_key:
            raise ValueError("开始月份不能晚于结束月份")

        year = int(start_key[:4])
        month = int(start_key[-2:])
        end_year_int = int(end_key[:4])
        end_month_int = int(end_key[-2:])
        month_keys = []
        while (year, month) <= (end_year_int, end_month_int):
            month_keys.append(f"{year:04d}-{month:02d}")
            month += 1
            if month > 12:
                year += 1
                month = 1
        return month_keys

    @staticmethod
    def _category_label(category):
        labels = {
            "profit": "利润表",
            "cost": "成本合计表",
            "expense": "费用明细",
            "asset": "资产负债表",
            "sales": "销售出库",
            "ar": "应收账款",
            "ap": "应付账款",
            "cash": "货币资金",
        }
        return labels.get(category, category or "未识别")

    @staticmethod
    def _is_valid_month_key(value):
        if not isinstance(value, str) or not re.match(r"^20\d{2}-\d{2}$", value):
            return False
        month = int(value[-2:])
        return 1 <= month <= 12

    @classmethod
    def _expand_month_key_range(cls, start_key, end_key):
        if start_key > end_key:
            start_key, end_key = end_key, start_key
        return cls.build_month_range(start_key[:4], start_key[-2:], end_key[:4], end_key[-2:])

    @classmethod
    def _extract_month_keys_from_text(cls, text):
        """从文件名/文本中提取单月或连续月份范围。"""
        raw = str(text or "")
        hits = []

        def add_hit(pos, year, month):
            try:
                year_int = int(year)
                month_int = int(month)
            except Exception:
                return
            if 1 <= month_int <= 12:
                hits.append((pos, f"{year_int:04d}-{month_int:02d}"))

        for match in re.finditer(r"(20\d{2})\s*年\s*(\d{1,2})\s*月", raw):
            add_hit(match.start(), match.group(1), match.group(2))
        for match in re.finditer(r"(20\d{2})[./-](\d{1,2})(?!\d)", raw):
            add_hit(match.start(), match.group(1), match.group(2))
        for match in re.finditer(r"(?<!\d)(20\d{2})(\d{2})(?:\d{2})?(?!\d)", raw):
            add_hit(match.start(), match.group(1), match.group(2))

        if not hits:
            return []

        hits.sort(key=lambda item: item[0])
        ordered = []
        for _, key in hits:
            if key not in ordered:
                ordered.append(key)
        if len(ordered) >= 2:
            return cls._expand_month_key_range(ordered[0], ordered[-1])
        return ordered

    def _parse_export_datetime_value(self, value):
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass

        if isinstance(value, pd.Timestamp):
            if pd.isna(value):
                return None
            return value.to_pydatetime()
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        if isinstance(value, (int, float)):
            return None

        text = str(value).strip()
        if not text or "20" not in text:
            return None

        patterns = [
            r"(20\d{2})[./-](\d{1,2})[./-](\d{1,2})(?:\s+(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?)?",
            r"(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?(?:\s+(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?)?",
        ]
        for pattern in patterns:
            match = re.search(pattern, text)
            if not match:
                continue
            try:
                year = int(match.group(1))
                month = int(match.group(2))
                day = int(match.group(3))
                hour = int(match.group(4) or 0)
                minute = int(match.group(5) or 0)
                second = int(match.group(6) or 0)
                return datetime(year, month, day, hour, minute, second)
            except Exception:
                continue

        parsed = pd.to_datetime(text, errors="coerce")
        if pd.isna(parsed):
            return None
        try:
            return parsed.to_pydatetime()
        except Exception:
            return None

    def _extract_export_datetime_from_values(self, rows):
        recent_rows = []
        for row in rows or []:
            values = []
            for value in row:
                if value is None:
                    continue
                try:
                    if pd.isna(value):
                        continue
                except Exception:
                    pass
                text = str(value).strip()
                if text:
                    values.append(value)
            if values:
                recent_rows.append(values)
                if len(recent_rows) > 80:
                    recent_rows.pop(0)

        labeled_keywords = ("导出", "生成", "制表", "打印")
        for values in reversed(recent_rows):
            row_text = " ".join(str(value) for value in values)
            if any(keyword in row_text for keyword in labeled_keywords):
                for value in values:
                    dt = self._parse_export_datetime_value(value)
                    if dt:
                        return dt
                dt = self._parse_export_datetime_value(row_text)
                if dt:
                    return dt

        for values in reversed(recent_rows):
            # 导出时间常见为最后一行单独一个日期时间。避免把明细行中的业务日期当成导出日期。
            if len(values) > 2:
                continue
            dt_values = [self._parse_export_datetime_value(value) for value in values]
            dt_values = [dt for dt in dt_values if dt]
            if len(dt_values) == 1:
                return dt_values[0]
        return None

    def _extract_export_datetime_from_workbook(self, path):
        """读取导出文件底部的导出日期；读不到时返回 None。"""
        ext = str(path).lower()
        if ext.endswith((".xlsx", ".xlsm")):
            wb = None
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                all_rows = []
                for ws in wb.worksheets:
                    try:
                        if hasattr(ws, "reset_dimensions"):
                            ws.reset_dimensions()
                    except Exception:
                        pass
                    for row in ws.iter_rows(values_only=True):
                        all_rows.append(row)
                        if len(all_rows) > 120:
                            all_rows.pop(0)
                return self._extract_export_datetime_from_values(all_rows)
            except Exception:
                return None
            finally:
                if wb is not None:
                    try:
                        wb.close()
                    except Exception:
                        pass

        try:
            df = pd.read_excel(path, header=None)
            rows = [tuple(row) for row in df.tail(120).itertuples(index=False, name=None)]
            return self._extract_export_datetime_from_values(rows)
        except Exception:
            return None

    def _month_key_from_date(self, value):
        try:
            return f"{int(value.year):04d}-{int(value.month):02d}"
        except Exception:
            return None

    def _month_key_from_excel_cell(self, cell):
        """仅从看起来像日期的单元格中提取月份，避免把金额/标题误判为月份。"""
        value = getattr(cell, "value", cell)
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass

        if isinstance(value, pd.Timestamp):
            return self._month_key_from_date(value)
        if isinstance(value, datetime):
            return self._month_key_from_date(value)
        if isinstance(value, date):
            return f"{value.year:04d}-{value.month:02d}"

        try:
            is_excel_date = bool(getattr(cell, "is_date", False))
        except Exception:
            is_excel_date = False
        if is_excel_date:
            parsed = self._parse_date_value_flexible(value)
            if pd.notna(parsed):
                return self._month_key_from_date(parsed)

        # 普通数字很可能是金额、数量或单号，不按 Excel 序列日期解析。
        if isinstance(value, (int, float)):
            return None

        text = str(value).strip()
        if not text or "20" not in text:
            return None

        patterns = [
            r"^\s*(20\d{2})[./-](\d{1,2})[./-]\d{1,2}",
            r"^\s*(20\d{2})\s*年\s*(\d{1,2})\s*月\s*\d{1,2}\s*日?",
            r"^\s*(20\d{2})(\d{2})(\d{2})(?!\d)",
        ]
        for pattern in patterns:
            match = re.search(pattern, text)
            if not match:
                continue
            try:
                year = int(match.group(1))
                month = int(match.group(2))
            except Exception:
                continue
            key = f"{year:04d}-{month:02d}"
            if self._is_valid_month_key(key):
                return key
        return None

    @staticmethod
    def _excel_row_nonempty_values(cells):
        values = []
        for cell in cells:
            value = getattr(cell, "value", cell)
            if value is None:
                continue
            try:
                if pd.isna(value):
                    continue
            except Exception:
                pass
            if str(value).strip() == "":
                continue
            values.append(value)
        return values

    def _excel_row_month_key(self, cells):
        """
        从一行明细数据中识别月份。
        这里故意只认行内日期，不用文件名/表头范围，保证“覆盖”按整月数据处理。
        """
        cells = list(cells or [])
        nonempty = self._excel_row_nonempty_values(cells)
        if len(nonempty) < 2:
            return None

        row_text = " ".join(str(value) for value in nonempty[:12])
        if len(nonempty) <= 2 and re.search(r"20\d{2}.*[~～\-].*20\d{2}", row_text):
            return None
        if len(nonempty) <= 3 and any(token in row_text for token in ("期间", "账页", "科目名称", "单位名称")):
            return None

        month_keys = []
        for cell in cells:
            key = self._month_key_from_excel_cell(cell)
            if key:
                month_keys.append(key)
        if month_keys:
            # 明细账常同时包含“日期-号码”和“最终修改日期”；业务月份应取靠前的业务日期。
            return month_keys[0]
        return None

    def _extract_data_month_keys_from_workbook(self, path):
        """从工作簿明细行日期识别实际存在的月份。"""
        if not str(path).lower().endswith((".xlsx", ".xlsm")):
            return []

        month_keys = set()
        wb = None
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                try:
                    if hasattr(ws, "reset_dimensions"):
                        ws.reset_dimensions()
                except Exception:
                    pass
                for row in ws.iter_rows():
                    month_key = self._excel_row_month_key(row)
                    if month_key:
                        month_keys.add(month_key)
        except Exception:
            return []
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
        return sorted(month_keys)

    def _determine_file_period_info(self, path, filename, df_peek=None, category=None):
        """识别文件覆盖的月份列表，并标记识别来源。范围文件返回范围内所有月份。"""
        if df_peek is None:
            try:
                df_peek = pd.read_excel(path, header=None, nrows=1)
            except Exception:
                df_peek = None

        header_keys = []
        if df_peek is not None:
            start_date, end_date = self._parse_header_date(df_peek)
            if start_date and end_date:
                header_keys = self._expand_month_key_range(
                    f"{start_date.year:04d}-{start_date.month:02d}",
                    f"{end_date.year:04d}-{end_date.month:02d}",
                )

        filename_keys = self._extract_month_keys_from_text(filename)

        if category in self.DETAIL_ROW_MONTH_CATEGORIES:
            data_keys = self._extract_data_month_keys_from_workbook(path)
            if data_keys:
                return {"periods": data_keys, "source": "data_rows"}

        if category in {"profit", "asset", "cost"}:
            if header_keys:
                source = "header"
                if filename_keys and filename_keys[-1] != header_keys[-1]:
                    source = "header_filename_mismatch"
                return {"periods": [header_keys[-1]], "source": source}
            if filename_keys:
                return {"periods": [filename_keys[-1]], "source": "filename"}

        if len(filename_keys) > len(header_keys):
            return {"periods": filename_keys, "source": "filename"}
        if header_keys:
            return {"periods": header_keys, "source": "header"}

        month = self._determine_period_key(path, filename, df_peek)
        if self._is_valid_month_key(month):
            return {"periods": [month], "source": "fallback"}
        return {"periods": [], "source": ""}

    def _determine_file_period_keys(self, path, filename, df_peek=None, category=None):
        """识别文件覆盖的月份列表。范围文件返回范围内所有月份。"""
        return self._determine_file_period_info(path, filename, df_peek, category).get("periods", [])

    def inspect_data_folder(self, required_categories=None, expected_months=None):
        """
        检视基础资料目录中的数据文件。
        返回文件识别、重复月份、缺失月份、可整理重复数据候选等信息。
        """
        required_categories = list(required_categories or ["profit", "cost", "expense", "asset", "sales", "ar"])
        result = {
            "base_dir": self.base_data_dir,
            "files": [],
            "duplicates": [],
            "cleanup_candidates": [],
            "manual_review": [],
            "missing": [],
            "unclassified": [],
            "unreadable": [],
            "expected_months": [],
            "category_periods": {},
        }

        if not os.path.isdir(self.base_data_dir):
            result["error"] = f"基础资料目录不存在: {self.base_data_dir}"
            return result

        try:
            filenames = sorted(os.listdir(self.base_data_dir))
        except Exception as exc:
            result["error"] = f"读取基础资料目录失败: {exc}"
            return result

        excel_exts = (".xlsx", ".xls", ".xlsm")
        for filename in filenames:
            if str(filename).startswith("~$"):
                continue
            if not str(filename).lower().endswith(excel_exts):
                continue

            path = os.path.join(self.base_data_dir, filename)
            meta = {
                "filename": filename,
                "path": os.path.abspath(path),
                "category": None,
                "category_label": "未识别",
                "periods": [],
                "periods_source": "",
                "mtime": None,
                "mtime_text": "",
                "export_datetime": None,
                "export_time": None,
                "export_time_text": "",
                "recency_time": None,
                "recency_time_text": "",
                "recency_source": "",
                "size": None,
                "error": "",
            }

            try:
                stat = os.stat(path)
                meta["mtime"] = stat.st_mtime
                meta["mtime_text"] = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                meta["size"] = stat.st_size
            except Exception as exc:
                meta["error"] = f"读取文件属性失败: {exc}"

            try:
                export_dt = self._extract_export_datetime_from_workbook(path)
                if export_dt:
                    meta["export_datetime"] = export_dt
                    meta["export_time"] = export_dt.timestamp()
                    meta["export_time_text"] = export_dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception as exc:
                meta["error"] = (meta["error"] + "; " if meta["error"] else "") + f"读取导出日期失败: {exc}"

            if meta.get("export_time") is not None:
                meta["recency_time"] = meta["export_time"]
                meta["recency_time_text"] = meta["export_time_text"]
                meta["recency_source"] = "导出日期"
            else:
                meta["recency_time"] = meta.get("mtime")
                meta["recency_time_text"] = meta.get("mtime_text", "")
                meta["recency_source"] = "文件修改时间"

            df_peek = None
            try:
                df_peek = pd.read_excel(path, header=None, nrows=1)
            except Exception as exc:
                meta["error"] = (meta["error"] + "; " if meta["error"] else "") + f"读取表头失败: {exc}"

            try:
                category = self._classify_source_file(filename, path, df_peek)
                meta["category"] = category
                meta["category_label"] = self._category_label(category)
            except Exception as exc:
                meta["error"] = (meta["error"] + "; " if meta["error"] else "") + f"识别类型失败: {exc}"

            try:
                period_info = self._determine_file_period_info(path, filename, df_peek, meta.get("category"))
                meta["periods"] = period_info.get("periods", [])
                meta["periods_source"] = period_info.get("source", "")
                if meta["periods_source"] == "header_filename_mismatch":
                    filename_keys = self._extract_month_keys_from_text(filename)
                    meta["period_warning"] = (
                        f"文件名月份 {','.join(filename_keys)} 与表头月份 "
                        f"{','.join(meta['periods'])} 不一致，按表头月份处理"
                    )
            except Exception as exc:
                meta["error"] = (meta["error"] + "; " if meta["error"] else "") + f"识别月份失败: {exc}"

            result["files"].append(meta)
            if meta["error"]:
                result["unreadable"].append(meta)
            elif not meta["category"]:
                result["unclassified"].append(meta)

        by_key = {}
        category_periods = {}
        for meta in result["files"]:
            category = meta.get("category")
            if not category:
                continue
            for period in meta.get("periods") or []:
                if not self._is_valid_month_key(period):
                    continue
                by_key.setdefault((category, period), []).append(meta)
                category_periods.setdefault(category, set()).add(period)

        result["category_periods"] = {
            category: sorted(periods) for category, periods in category_periods.items()
        }

        latest_by_key = {}
        for (category, period), metas in by_key.items():
            if not metas:
                continue
            ordered = sorted(
                metas,
                key=lambda item: (
                    item.get("recency_time") or 0,
                    item.get("mtime") or 0,
                    item.get("filename") or "",
                ),
                reverse=True,
            )
            latest = ordered[0]
            latest_by_key[(category, period)] = latest
            if len(ordered) > 1:
                result["duplicates"].append({
                    "category": category,
                    "category_label": self._category_label(category),
                    "period": period,
                    "latest": latest,
                    "latest_time_text": latest.get("recency_time_text", ""),
                    "latest_time_source": latest.get("recency_source", ""),
                    "files": ordered,
                })

        seen_candidate_paths = set()
        for meta in result["files"]:
            category = meta.get("category")
            periods = [p for p in (meta.get("periods") or []) if self._is_valid_month_key(p)]
            if not category or not periods:
                continue

            superseded_periods = []
            blocking_periods = []
            keeper_files = []
            keeper_filenames = []
            for period in periods:
                group = by_key.get((category, period), [])
                latest = latest_by_key.get((category, period))
                if len(group) > 1 and latest and latest.get("path") != meta.get("path"):
                    superseded_periods.append(period)
                    keeper_name = latest.get("filename")
                    if keeper_name and keeper_name not in keeper_filenames:
                        keeper_filenames.append(keeper_name)
                    keeper_label = keeper_name or ""
                    if latest.get("recency_time_text"):
                        keeper_label = (
                            f"{keeper_label}"
                            f"({latest.get('recency_source')}: {latest.get('recency_time_text')})"
                        )
                    if keeper_label and keeper_label not in keeper_files:
                        keeper_files.append(keeper_label)
                else:
                    blocking_periods.append(period)

            if not superseded_periods:
                continue

            item = dict(meta)
            item["superseded_periods"] = superseded_periods
            item["blocking_periods"] = blocking_periods
            item["periods_to_remove"] = superseded_periods
            item["periods_to_keep"] = blocking_periods
            item["keeper_filenames"] = keeper_filenames
            item["keeper_filename"] = "；".join(keeper_filenames)
            item["keeper_files"] = keeper_files
            item["keeper_file"] = "；".join(keeper_files)
            path_key = item.get("path")
            if path_key in seen_candidate_paths:
                continue
            if not blocking_periods:
                item["cleanup_action"] = "delete_file"
                item["reason"] = "整份文件的月份均已在保留文件中重复"
                result["cleanup_candidates"].append(item)
                seen_candidate_paths.add(path_key)
            else:
                item["cleanup_action"] = "trim_months"
                if category in self.DETAIL_ROW_MONTH_CATEGORIES and item.get("periods_source") == "data_rows":
                    item["reason"] = "删除本文件中的重复月份整月数据，其他月份保留"
                    result["cleanup_candidates"].append(item)
                    seen_candidate_paths.add(path_key)
                else:
                    item["reason"] = "部分月份重复，但未能按行内日期安全定位需删除的整月行"
                    result["manual_review"].append(item)

        result["duplicates"].sort(key=lambda item: (item["category"], item["period"]))
        result["cleanup_candidates"].sort(key=lambda item: (item.get("category") or "", item.get("filename") or ""))
        result["manual_review"].sort(key=lambda item: (item.get("category") or "", item.get("filename") or ""))

        if expected_months is None:
            scope_months = set()
            for category in ("profit", "cost", "asset"):
                scope_months.update(category_periods.get(category, set()))
            if not scope_months:
                for category in required_categories:
                    scope_months.update(category_periods.get(category, set()))
            expected_months = self._expand_month_key_range(min(scope_months), max(scope_months)) if scope_months else []
        else:
            expected_months = [
                self.normalize_month_key(str(month)[:4], str(month)[-2:])
                for month in expected_months
                if self._is_valid_month_key(str(month))
            ]

        result["expected_months"] = expected_months
        for period in expected_months:
            missing_categories = [
                category for category in required_categories
                if period not in category_periods.get(category, set())
            ]
            if missing_categories:
                result["missing"].append({
                    "period": period,
                    "missing_categories": missing_categories,
                    "missing_labels": [self._category_label(cat) for cat in missing_categories],
                })

        return result

    def format_data_folder_inspection(self, inspection, max_items=80):
        """将数据文件夹检视结果格式化为日志文本。"""
        if not inspection:
            return "未生成检视结果。"
        if inspection.get("error"):
            return inspection["error"]

        files = inspection.get("files") or []
        duplicates = inspection.get("duplicates") or []
        candidates = inspection.get("cleanup_candidates") or []
        missing = inspection.get("missing") or []
        unclassified = inspection.get("unclassified") or []
        unreadable = inspection.get("unreadable") or []
        manual_review = inspection.get("manual_review") or []
        period_warnings = [item for item in files if item.get("period_warning")]
        expected_months = inspection.get("expected_months") or []
        delete_count = sum(1 for item in candidates if item.get("cleanup_action") == "delete_file")
        trim_count = sum(1 for item in candidates if item.get("cleanup_action") == "trim_months")

        lines = [
            "数据检视结果:",
            f"- 数据目录: {inspection.get('base_dir')}",
            (
                f"- Excel文件: {len(files)} 个；重复月份: {len(duplicates)} 组；"
                f"可处理重复数据: {len(candidates)} 项(删重复文件 {delete_count}，清重复月份 {trim_count})；"
                f"缺失月份: {len(missing)} 个"
            ),
        ]
        if expected_months:
            lines.append(f"- 检视月份范围: {expected_months[0]} 至 {expected_months[-1]}")
        if unreadable:
            lines.append(f"- 读取异常文件: {len(unreadable)} 个")
        if unclassified:
            lines.append(f"- 未识别类型文件: {len(unclassified)} 个")
        if manual_review:
            lines.append(f"- 部分重叠无法自动整理: {len(manual_review)} 个")
        if period_warnings:
            lines.append(f"- 文件名与表头月份不一致: {len(period_warnings)} 个")

        shown = 0
        if missing:
            lines.append("")
            lines.append("缺失数据:")
            for item in missing[:max_items]:
                lines.append(f"- {item['period']}: 缺少 {'、'.join(item.get('missing_labels') or [])}")
                shown += 1
            if len(missing) > max_items:
                lines.append(f"- ... 还有 {len(missing) - max_items} 条缺失记录未显示")

        if duplicates:
            lines.append("")
            lines.append("重复数据:")
            remaining = max(0, max_items - shown)
            for item in duplicates[:remaining or max_items]:
                file_names = "；".join(meta.get("filename", "") for meta in item.get("files", [])[:4])
                if len(item.get("files", [])) > 4:
                    file_names += f"；... 共{len(item.get('files', []))}个"
                latest_name = (item.get("latest") or {}).get("filename", "")
                latest_time = item.get("latest_time_text") or ""
                latest_source = item.get("latest_time_source") or ""
                latest_note = f" ({latest_source}: {latest_time})" if latest_time else ""
                lines.append(f"- {item['period']} {item['category_label']}: 保留 {latest_name}{latest_note}；重复文件 {file_names}")
            if len(duplicates) > (remaining or max_items):
                lines.append(f"- ... 还有 {len(duplicates) - (remaining or max_items)} 组重复记录未显示")

        if candidates:
            lines.append("")
            lines.append("可处理重复数据候选:")
            for item in candidates[:max_items]:
                periods = ",".join(item.get("superseded_periods") or item.get("periods") or [])
                keeper = item.get("keeper_file") or ""
                if item.get("cleanup_action") == "trim_months":
                    keep = ",".join(item.get("blocking_periods") or [])
                    lines.append(
                        f"- {item.get('category_label')} 清理重复月份 {periods}: "
                        f"{item.get('filename')}；重复月份保留在 {keeper}；本文件保留月份 {keep}"
                    )
                else:
                    lines.append(
                        f"- {item.get('category_label')} 删除整份重复文件 {periods}: "
                        f"{item.get('filename')}；保留在 {keeper} ({item.get('mtime_text')})"
                    )
            if len(candidates) > max_items:
                lines.append(f"- ... 还有 {len(candidates) - max_items} 个候选项未显示")

        if manual_review:
            lines.append("")
            lines.append("无法自动整理的部分重叠文件:")
            for item in manual_review[:max_items]:
                sup = ",".join(item.get("superseded_periods") or [])
                keep = ",".join(item.get("blocking_periods") or [])
                reason = item.get("reason") or "需人工判断"
                lines.append(f"- {item.get('filename')}: 重复月份 {sup}；仍需保留月份 {keep}；{reason}")
            if len(manual_review) > max_items:
                lines.append(f"- ... 还有 {len(manual_review) - max_items} 个文件未显示")

        if period_warnings:
            lines.append("")
            lines.append("文件名与表头月份不一致:")
            for item in period_warnings[:max_items]:
                lines.append(f"- {item.get('filename')}: {item.get('period_warning')}")
            if len(period_warnings) > max_items:
                lines.append(f"- ... 还有 {len(period_warnings) - max_items} 个文件未显示")

        return "\n".join(lines)

    def _validate_base_data_excel_path(self, path):
        base_abs = os.path.abspath(self.base_data_dir)
        abs_path = os.path.abspath(path)
        try:
            if os.path.commonpath([base_abs, abs_path]) != base_abs:
                raise ValueError("文件不在当前基础资料目录内")
        except ValueError:
            raise ValueError("文件不在当前基础资料目录内")
        if not os.path.isfile(abs_path):
            raise FileNotFoundError("文件不存在")
        if not abs_path.lower().endswith((".xlsx", ".xls", ".xlsm")):
            raise ValueError("只允许处理Excel数据文件")
        return abs_path

    @staticmethod
    def _delete_rows_grouped(ws, row_numbers):
        rows = sorted(set(int(row) for row in row_numbers), reverse=True)
        if not rows:
            return
        group_start = rows[0]
        group_end = rows[0]
        for row in rows[1:]:
            if row == group_end - 1:
                group_end = row
                continue
            ws.delete_rows(group_end, group_start - group_end + 1)
            group_start = row
            group_end = row
        ws.delete_rows(group_end, group_start - group_end + 1)

    def trim_data_file_months(self, path, month_keys):
        """删除一个明细文件中指定月份的整月数据行，不删除文件本身。"""
        abs_path = self._validate_base_data_excel_path(path)
        if not abs_path.lower().endswith((".xlsx", ".xlsm")):
            raise ValueError("仅支持清理 .xlsx/.xlsm 文件内的旧月份数据；.xls 请先另存为 .xlsx")

        target_months = sorted({
            str(month).strip()
            for month in (month_keys or [])
            if self._is_valid_month_key(str(month).strip())
        })
        if not target_months:
            raise ValueError("未指定有效的待清理月份")

        wb = None
        deleted_rows_by_sheet = {}
        try:
            wb = openpyxl.load_workbook(abs_path, keep_vba=abs_path.lower().endswith(".xlsm"))
            target_set = set(target_months)
            for ws in wb.worksheets:
                rows_to_delete = []
                for row in ws.iter_rows():
                    month_key = self._excel_row_month_key(row)
                    if month_key in target_set:
                        rows_to_delete.append(row[0].row)
                if rows_to_delete:
                    self._delete_rows_grouped(ws, rows_to_delete)
                    deleted_rows_by_sheet[ws.title] = len(rows_to_delete)

            deleted_total = sum(deleted_rows_by_sheet.values())
            if deleted_total <= 0:
                raise ValueError("未找到可按月份清理的明细行")
            wb.save(abs_path)
            return {
                "path": abs_path,
                "months": target_months,
                "deleted_rows": deleted_total,
                "deleted_rows_by_sheet": deleted_rows_by_sheet,
            }
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def delete_data_files(self, file_paths):
        """删除基础资料目录内指定 Excel 文件，返回删除结果。"""
        deleted = []
        failed = []
        for path in file_paths or []:
            try:
                abs_path = self._validate_base_data_excel_path(path)
                os.remove(abs_path)
                deleted.append(abs_path)
            except Exception as exc:
                failed.append({"path": path, "error": str(exc)})
        return {"deleted": deleted, "failed": failed}

    def cleanup_superseded_data(self, cleanup_items):
        """按检测结果处理重复数据：每个类型+月份只保留一个来源。"""
        deleted = []
        trimmed = []
        failed = []

        for item in cleanup_items or []:
            path = item.get("path") if isinstance(item, dict) else item
            try:
                if isinstance(item, dict):
                    action = item.get("cleanup_action") or (
                        "trim_months" if item.get("blocking_periods") else "delete_file"
                    )
                else:
                    action = "delete_file"

                if action == "trim_months":
                    months = (
                        item.get("periods_to_remove")
                        or item.get("superseded_periods")
                        or []
                    )
                    trimmed.append(self.trim_data_file_months(path, months))
                elif action == "delete_file":
                    result = self.delete_data_files([path])
                    deleted.extend(result.get("deleted", []))
                    failed.extend(result.get("failed", []))
                else:
                    raise ValueError(f"未知整理动作: {action}")
            except Exception as exc:
                failed.append({"path": path, "error": str(exc)})

        return {"deleted": deleted, "trimmed": trimmed, "failed": failed}

    def _extract_month_from_filename(self, filename):
        """从文件名提取月份 (e.g., '成本1月.xlsx' -> '2025-01')"""
        year = datetime.now().year # Default
        year_match = re.search(r'(20\d{2})', filename)
        if year_match:
            year = int(year_match.group(1))
            
        month_match = re.search(r'(?<![\d\-])(\d{1,2})月', filename)
        if month_match:
            month = int(month_match.group(1))
            return f"{year}-{month:02d}"
            
        month_matches = re.findall(r'(\d{1,2})月', filename)
        if month_matches:
            month = int(month_matches[-1])
            return f"{year}-{month:02d}"

        return "Unknown"

    def _parse_header_date(self, df):
        """从表头文本解析日期范围或单月，返回(start_date, end_date)。"""
        if df is None or df.empty:
            return None, None

        candidates = []
        try:
            first_cell = df.iloc[0, 0]
            if first_cell is not None and str(first_cell).strip():
                candidates.append(str(first_cell))
        except Exception:
            pass
        try:
            first_col = df.columns[0]
            if first_col is not None and str(first_col).strip():
                candidates.append(str(first_col))
        except Exception:
            pass

        sep_pattern = r'\s*[~～\-]\s*'
        full_date = r'(20\d{2}[./-]\d{1,2}[./-]\d{1,2})'
        month_date = r'(20\d{2}[./-]\d{1,2})'

        def _parse_full(token):
            token = token.replace('.', '/').replace('-', '/').strip()
            return datetime.strptime(token, '%Y/%m/%d')

        def _parse_month(token):
            token = token.replace('.', '/').replace('-', '/').strip()
            return datetime.strptime(token + '/01', '%Y/%m/%d')

        for raw in candidates:
            text = str(raw)
            text = text.replace('年', '/').replace('月', '/').replace('日', '')
            text = re.sub(r'\s+', ' ', text).strip()

            match = re.search(full_date + sep_pattern + full_date, text)
            if match:
                try:
                    return _parse_full(match.group(1)), _parse_full(match.group(2))
                except Exception:
                    pass

            match = re.search(month_date + sep_pattern + month_date, text)
            if match:
                try:
                    return _parse_month(match.group(1)), _parse_month(match.group(2))
                except Exception:
                    pass

            # 单月格式: 2025/12、2025-12、202512
            match = re.search(r'(20\d{2})[./-](\d{1,2})(?![./-]\d)', text)
            if not match:
                match = re.search(r'(20\d{2})(\d{2})(?!\d)', text)
            if match:
                try:
                    year = int(match.group(1))
                    month = int(match.group(2))
                    if 1 <= month <= 12:
                        dt = datetime(year, month, 1)
                        return dt, dt
                except Exception:
                    pass

        return None, None

    def _determine_period_key(self, path, filename, df_peek=None):
        if df_peek is None:
            try:
                df_peek = pd.read_excel(path, header=None, nrows=1)
            except Exception:
                df_peek = None
        
        if df_peek is not None:
            s_date, e_date = self._parse_header_date(df_peek)
            if s_date and e_date:
                if s_date.year == e_date.year and s_date.month == e_date.month:
                    return f"{s_date.year}-{s_date.month:02d}"
                return f"{e_date.year}-{e_date.month:02d}"

        return self._extract_month_from_filename(filename)

    def _load_cost_data(self, path, filename):
        try:
            df_peek = pd.read_excel(path, header=None, nrows=1)
            month = self._determine_period_key(path, filename, df_peek)
            df = pd.read_excel(path, header=[0, 1])
            new_cols = []
            for col in df.columns:
                c0, c1 = str(col[0]).strip(), str(col[1]).strip()
                if 'Unnamed' in c0: c0 = ''
                if 'Unnamed' in c1: c1 = ''
                new_cols.append(f"{c0}_{c1}" if c0 and c1 else (c0 or c1))
            df.columns = new_cols
            self.data['cost'][month] = df
        except Exception as e:
            print(f"加载成本数据失败 {filename}: {e}")

    def _load_expense_data(self, path, filename):
        try:
            def store_monthly_expense(month_key, data_frame):
                if not month_key or data_frame is None:
                    return
                incoming = data_frame.copy()
                existing = self.data['expense'].get(month_key)
                if existing is not None and not existing.empty:
                    self.data['expense'][month_key] = pd.concat(
                        [existing, incoming],
                        ignore_index=True,
                        sort=False,
                    )
                else:
                    self.data['expense'][month_key] = incoming

            date_col, df, header_row = None, None, 0
            for h_row in [0, 1]:
                temp_df = pd.read_excel(path, header=h_row)
                for c in temp_df.columns:
                    if '日期' in str(c) or 'Date' in str(c):
                        date_col, df, header_row = c, temp_df, h_row
                        break
                if date_col: break
            
            if not date_col:
                for h_row in [0, 1]:
                    temp_df = pd.read_excel(path, header=h_row)
                    if not temp_df.empty:
                        first_col = temp_df.columns[0]
                        sample_val = str(temp_df.iloc[0, 0])
                        if sample_val.startswith('202') and ('/' in sample_val or '-' in sample_val):
                            date_col, df, header_row = first_col, temp_df, h_row
                            break
                    if date_col: break

            if date_col:
                def parse_date(val):
                    if pd.isna(val):
                        return None
                    s = str(val).strip()
                    # Try explicit YYYY-MM-DD or YYYY/MM/DD extraction
                    m = re.search(r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})', s)
                    if m:
                        s = m.group(1)
                    try:
                        return pd.to_datetime(s)
                    except Exception:
                        try:
                            # Excel serial date
                            return pd.to_datetime(float(s), unit="d", origin="1899-12-30")
                        except Exception:
                            # Try YYYYMMDD
                            if re.match(r'^\d{8}$', s):
                                try:
                                    return pd.to_datetime(s, format='%Y%m%d')
                                except Exception:
                                    return None
                            return None
                df['ParsedDate'] = df[date_col].apply(parse_date)
                df.dropna(subset=['ParsedDate'], inplace=True)
                non_business_cols = [
                    c for c in df.columns
                    if c not in {date_col, 'ParsedDate', 'MonthStr'}
                ]
                if non_business_cols:
                    content = df[non_business_cols].apply(
                        lambda col: col.notna() & (col.astype(str).str.strip() != "")
                    )
                    df = df[content.any(axis=1)].copy()
                df['MonthStr'] = df['ParsedDate'].dt.strftime('%Y-%m')
                for m, group in df.groupby('MonthStr'):
                    store_monthly_expense(m, group)
            else:
                df_peek = pd.read_excel(path, header=None, nrows=1)
                header_date_range = self._parse_header_date(df_peek)
                month = f"{header_date_range[1].year}-{header_date_range[1].month:02d}" if header_date_range[1] else self._extract_month_from_filename(filename)
                store_monthly_expense(month, pd.read_excel(path, header=0))
        except Exception as e:
            print(f"加载费用数据失败 {filename}: {e}")

    def _select_statement_sheet(self, path, keywords):
        """优先选择名称包含指定关键词的工作表；找不到时使用首个工作表。"""
        try:
            with pd.ExcelFile(path) as xl:
                for sheet_name in xl.sheet_names:
                    sheet_text = str(sheet_name)
                    if any(keyword in sheet_text for keyword in keywords):
                        return sheet_name
                if xl.sheet_names:
                    return xl.sheet_names[0]
        except Exception:
            pass
        return 0

    def _detect_statement_header_row(self, preview_df, default_header_row=0):
        """从前几行中定位报表表头行，兼容“标题行 + 财务报表显示名”格式。"""
        if preview_df is None or preview_df.empty:
            return default_header_row

        header_tokens = ("财务报表显示名", "项目", "科目名称", "科目")
        for idx, row in preview_df.iterrows():
            values = []
            for value in row.tolist():
                if value is None:
                    continue
                try:
                    if pd.isna(value):
                        continue
                except Exception:
                    pass
                text = str(value).strip()
                if text:
                    values.append(text)
            if not values:
                continue

            first_text = values[0].replace(" ", "").replace("\u3000", "")
            row_text = " ".join(values)
            if any(token in first_text for token in header_tokens):
                return int(idx)
            if any(token in row_text for token in header_tokens) and (
                len(values) > 1
                and (
                    any(self._extract_month_key_from_text(value) for value in values[1:])
                    or any(value in ("合计", "本期金额", "本月数") for value in values[1:])
                )
            ):
                return int(idx)

        return default_header_row

    @staticmethod
    def _row_has_nonempty_values(row, columns):
        for col in columns:
            value = row.get(col)
            if value is None:
                continue
            try:
                if pd.isna(value):
                    continue
            except Exception:
                pass
            if str(value).strip() != "":
                return True
        return False

    def _clean_profit_dataframe(self, df):
        """清理利润表读取结果：删除全空列、空行和仅含导出时间的尾行。"""
        if df is None or df.empty:
            return df

        df = df.dropna(axis=1, how='all')
        if df.empty:
            return df

        name_col = df.columns[0]
        value_cols = list(df.columns[1:])
        keep_indices = []
        for idx, row in df.iterrows():
            label = row.get(name_col)
            has_values = self._row_has_nonempty_values(row, value_cols)

            if label is None:
                if has_values:
                    keep_indices.append(idx)
                continue
            try:
                if pd.isna(label):
                    if has_values:
                        keep_indices.append(idx)
                    continue
            except Exception:
                pass

            text = str(label).strip()
            if not text and not has_values:
                continue

            if not has_values and self._parse_export_datetime_value(text):
                continue

            keep_indices.append(idx)

        if len(keep_indices) != len(df.index):
            df = df.loc[keep_indices].reset_index(drop=True)
        return df

    def _load_profit_data(self, path, filename):
        try:
            sheet_name = self._select_statement_sheet(path, ("利润表", "损益"))
            df_peek = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=8)
            month = self._determine_period_key(path, filename, df_peek.head(1))
            default_header_row = 1 if self._parse_header_date(df_peek.head(1))[0] else 0
            header_row = self._detect_statement_header_row(df_peek, default_header_row)
            df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
            df = self._clean_profit_dataframe(df)
            self.data['profit'][month] = df
        except Exception as e:
            print(f"加载利润数据失败 {filename}: {e}")

    def _load_sales_data(self, path, filename):
        try:
            df_peek = pd.read_excel(path, header=None, nrows=1)
            header_row = 1 if df_peek.iloc[0].count() == 1 and isinstance(df_peek.iloc[0,0], str) else 0
            df = pd.read_excel(path, header=header_row)

            date_col = None
            for c in df.columns:
                if '日期' in str(c) or 'Date' in str(c):
                    date_col = c
                    break
            if not date_col and not df.empty:
                date_col = df.columns[0]

            def parse_sales_date(val):
                if pd.isna(val):
                    return None
                s = str(val).strip()
                m = re.search(r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})', s)
                if m:
                    s = m.group(1)
                try:
                    return pd.to_datetime(s)
                except Exception:
                    return None

            if date_col:
                df['ParsedDate'] = df[date_col].apply(parse_sales_date)
                df = df.dropna(subset=['ParsedDate'])
                df['MonthStr'] = df['ParsedDate'].dt.strftime('%Y-%m')
                for m_key, group in df.groupby('MonthStr'):
                    self.data['sales'][m_key] = group.copy()
                if self.sales_df is None:
                    self.sales_df = df
                else:
                    self.sales_df = pd.concat([self.sales_df, df], ignore_index=True)
            else:
                month = self._determine_period_key(path, filename, df_peek)
                self.data['sales'][month] = df
        except Exception as e:
            print(f"加载销售数据失败 {filename}: {e}")

    def _load_ar_data(self, path, filename):
        try:
            df_peek = pd.read_excel(path, header=None, nrows=1)
            header_row = 1 if self._parse_header_date(df_peek)[0] else 0
            df = pd.read_excel(path, header=header_row)
            try:
                detail_df = self._read_multiblock_ledger_df(path, filename)
                if detail_df is None or detail_df.empty:
                    detail_df = pd.read_excel(path, header=1)
                    detail_df.columns = [str(c).strip().rstrip('\t') for c in detail_df.columns]
                    detail_df["SourceFile"] = filename
                date_col, parsed = self._pick_best_date_column(
                    detail_df, ["日期", "date", "业务日期", "单据日期", "日期-号码"]
                )
                if date_col and parsed is not None and parsed.notna().sum() > 0:
                    detail_df = detail_df.copy()
                    detail_df["ParsedDate"] = parsed
                    detail_df = detail_df.dropna(subset=["ParsedDate"])
                    detail_df["MonthStr"] = detail_df["ParsedDate"].dt.strftime("%Y-%m")
                    for m_key, group in detail_df.groupby("MonthStr"):
                        existing = self.data['ar'].get(m_key)
                        if existing is None or existing.empty:
                            self.data['ar'][m_key] = group.copy()
                        else:
                            self.data['ar'][m_key] = pd.concat([existing, group.copy()], ignore_index=True)
                    if self.ar_detail_df is None or self.ar_detail_df.empty:
                        self.ar_detail_df = detail_df
                    else:
                        self.ar_detail_df = pd.concat([self.ar_detail_df, detail_df], ignore_index=True)
                else:
                    month = self._determine_period_key(path, filename, df_peek)
                    self.data['ar'][month] = df
                    if self.ar_detail_df is None or self.ar_detail_df.empty:
                        self.ar_detail_df = detail_df
                    else:
                        self.ar_detail_df = pd.concat([self.ar_detail_df, detail_df], ignore_index=True)
            except Exception as e:
                month = self._determine_period_key(path, filename, df_peek)
                self.data['ar'][month] = df
                msg = f"加载应收明细失败 {filename}: {e}"
                print(msg)
                self._log_audit(msg)
        except Exception as e:
            print(f"加载应收数据失败 {filename}: {e}")

    def _append_detail_frame(self, attr_name, df):
        if df is None or df.empty:
            return
        existing = getattr(self, attr_name, None)
        if existing is None or existing.empty:
            setattr(self, attr_name, df.copy())
        else:
            setattr(self, attr_name, pd.concat([existing, df.copy()], ignore_index=True))

    def _load_account_detail_data(self, path, filename, category, detail_attr, label):
        try:
            df_peek = pd.read_excel(path, header=None, nrows=1)
            detail_df = self._read_multiblock_ledger_df(path, filename)
            if detail_df is None or detail_df.empty:
                detail_df = pd.read_excel(path, header=1)
                detail_df.columns = [str(c).strip().rstrip('\t') for c in detail_df.columns]
                detail_df["SourceFile"] = filename

            date_col, parsed = self._pick_best_date_column(
                detail_df, ["日期", "date", "业务日期", "单据日期", "日期-号码"]
            )
            if date_col and parsed is not None and parsed.notna().sum() > 0:
                detail_df = detail_df.copy()
                detail_df["ParsedDate"] = parsed
                detail_df = detail_df.dropna(subset=["ParsedDate"])
                detail_df["MonthStr"] = detail_df["ParsedDate"].dt.strftime("%Y-%m")
                for m_key, group in detail_df.groupby("MonthStr"):
                    existing = self.data[category].get(m_key)
                    if existing is None or existing.empty:
                        self.data[category][m_key] = group.copy()
                    else:
                        self.data[category][m_key] = pd.concat([existing, group.copy()], ignore_index=True)
                self._append_detail_frame(detail_attr, detail_df)
            else:
                month = self._determine_period_key(path, filename, df_peek)
                self.data[category][month] = detail_df
                self._append_detail_frame(detail_attr, detail_df)
        except Exception as e:
            print(f"加载{label}数据失败 {filename}: {e}")

    def _load_ap_data(self, path, filename):
        self._load_account_detail_data(path, filename, "ap", "ap_detail_df", "应付账款")

    def _load_cash_data(self, path, filename):
        self._load_account_detail_data(path, filename, "cash", "cash_detail_df", "货币资金")

    def _load_asset_data(self, path, filename):
        try:
            df_peek = pd.read_excel(path, header=None, nrows=1)
            month = self._determine_period_key(path, filename, df_peek)
            header_row = 1 if self._parse_header_date(df_peek)[0] else 0
            df = pd.read_excel(path, header=header_row)
            self.data['asset'][month] = df
        except Exception as e:
            print(f"加载资产数据失败 {filename}: {e}")

    def _print_stats(self):
        print("\n--- 数据加载统计 ---")
        for cat, months in self.data.items():
            print(f"[{cat.upper()}]: {len(months)} 个月份数据 ({', '.join(sorted(months.keys()))})")

    def _add_quality_issue(self, category, period, issue_type, detail, severity="WARN"):
        self.data_quality_issues.append({
            "category": category,
            "period": period,
            "issue_type": issue_type,
            "detail": detail,
            "severity": severity,
        })

    def _check_required_columns(self, df, category, period, required_groups):
        for group in required_groups:
            group_hit = False
            for col in df.columns:
                col_norm = self._normalize_column_name(col)
                if not col_norm:
                    continue
                for alias in group:
                    alias_norm = self._normalize_label(alias).replace("_", "")
                    if not alias_norm:
                        continue
                    if col_norm == alias_norm or alias_norm in col_norm:
                        group_hit = True
                        break
                if group_hit:
                    break
            if not group_hit:
                self._add_quality_issue(
                    category,
                    period,
                    "缺少字段",
                    f"缺少字段组: {', '.join(group)}",
                    severity="ERROR",
                )

    def _normalize_column_name(self, col):
        name = self._normalize_label(col)
        return name.replace("_", "")

    def _find_columns_by_keywords(self, df, include_keywords, exclude_keywords=None):
        include = [self._normalize_label(k) for k in include_keywords]
        exclude = [self._normalize_label(k) for k in exclude_keywords] if exclude_keywords else []
        matches = []
        for col in df.columns:
            name = self._normalize_column_name(col)
            if any(k in name for k in include) and not any(k in name for k in exclude):
                matches.append(col)
        return matches

    def _pick_first_column(self, df, include_keywords, exclude_keywords=None):
        cols = self._find_columns_by_keywords(df, include_keywords, exclude_keywords)
        return cols[0] if cols else None

    def _parse_date_value_flexible(self, value):
        if value is None or pd.isna(value):
            return pd.NaT
        if isinstance(value, (datetime, pd.Timestamp)):
            return pd.to_datetime(value, errors="coerce")

        # Excel serial date support.
        if isinstance(value, (int, float)):
            try:
                return pd.to_datetime(float(value), unit="D", origin="1899-12-30", errors="coerce")
            except Exception:
                return pd.NaT

        s = str(value).strip()
        if not s:
            return pd.NaT

        # 2025/12/31, 2025-12-31, 2025.12.31
        m = re.search(r"(20\d{2})[./-](\d{1,2})[./-](\d{1,2})", s)
        if m:
            try:
                return pd.Timestamp(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except Exception:
                pass

        # 2025年12月31日
        m = re.search(r"(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?", s)
        if m:
            try:
                return pd.Timestamp(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except Exception:
                pass

        # 20251231
        if re.fullmatch(r"20\d{6}", s):
            try:
                return pd.to_datetime(s, format="%Y%m%d", errors="coerce")
            except Exception:
                pass

        # Fallback parser.
        try:
            return pd.to_datetime(s, errors="coerce")
        except Exception:
            return pd.NaT

    def _parse_date_series_flexible(self, series):
        if series is None:
            return pd.Series(dtype="datetime64[ns]")
        return series.apply(self._parse_date_value_flexible)

    def _pick_best_date_column(self, df, include_keywords=None):
        if include_keywords is None:
            include_keywords = ["日期", "date", "业务日期", "单据日期"]
        candidates = self._find_columns_by_keywords(df, include_keywords)
        if not candidates:
            return None, None

        exact_candidates = []
        for keyword in include_keywords:
            for col in candidates:
                if str(col).strip().lower() == str(keyword).strip().lower() and col not in exact_candidates:
                    exact_candidates.append(col)
        non_metadata_candidates = [
            col for col in candidates
            if col not in exact_candidates and not str(col).strip().startswith("账页")
        ]
        metadata_candidates = [
            col for col in candidates
            if col not in exact_candidates and col not in non_metadata_candidates
        ]

        for candidate_group in [exact_candidates, non_metadata_candidates, metadata_candidates]:
            best_col = None
            best_parsed = None
            best_valid = -1
            for col in candidate_group:
                try:
                    parsed = self._parse_date_series_flexible(df[col])
                except Exception:
                    continue
                valid = int(parsed.notna().sum())
                if valid > best_valid:
                    best_valid = valid
                    best_col = col
                    best_parsed = parsed
            if best_valid > 0:
                return best_col, best_parsed

        return None, None

    def _dedupe_excel_headers(self, headers):
        result = []
        seen = Counter()
        for idx, header in enumerate(headers):
            if header is None or pd.isna(header) or str(header).strip() == "":
                base = f"Unnamed:{idx + 1}"
            else:
                base = str(header).strip().rstrip('\t')
            count = seen[base]
            seen[base] += 1
            result.append(base if count == 0 else f"{base}.{count}")
        return result

    def _parse_ledger_block_title(self, title):
        text = str(title or "").strip()
        info = {
            "账页标题": text,
            "账页公司名称": None,
            "账页科目名": None,
            "账页起始日期": None,
            "账页结束日期": None,
            "账页往来单位编码": None,
            "账页往来单位名": None,
        }
        m = re.search(r"公司名称\s*[:：]\s*(.*?)\s+/\s+", text)
        if m:
            info["账页公司名称"] = m.group(1).strip()
        m = re.search(r"公司名称\s*[:：]\s*.*?\s+/\s+(.+?)\s+/\s+20\d{2}[/-]", text)
        if m:
            info["账页科目名"] = m.group(1).strip()
        m = re.search(
            r"(20\d{2}[/-]\d{1,2}[/-]\d{1,2})\s*~\s*(20\d{2}[/-]\d{1,2}[/-]\d{1,2})",
            text,
        )
        if m:
            info["账页起始日期"] = m.group(1)
            info["账页结束日期"] = m.group(2)
        m = re.search(r"/\s*([A-Za-z0-9_.-]+)\s*[(（](.*)[)）]\s*$", text)
        if m:
            info["账页往来单位编码"] = m.group(1).strip()
            info["账页往来单位名"] = m.group(2).strip()
        return info

    def _row_matches_headers(self, row):
        total = 0
        match = 0
        for col, val in row.items():
            if val is None or pd.isna(val) or str(val).strip() == "":
                continue
            total += 1
            if str(val).strip() == str(col).strip():
                match += 1
        return total > 0 and match / total >= 0.6

    def _read_multiblock_ledger_df(self, path, filename):
        try:
            raw = pd.read_excel(path, sheet_name=0, header=None)
        except Exception:
            return None
        if raw is None or raw.empty:
            return None

        first_col = raw.iloc[:, 0]
        title_rows = [
            idx for idx, value in first_col.items()
            if "公司名称" in str(value) and ("明细账" in str(value) or "科目" in str(value))
        ]
        if not title_rows:
            return None

        def header_score(row):
            values = [str(v).strip() for v in row.tolist() if v is not None and not pd.isna(v)]
            joined = "|".join(values)
            return sum(
                1
                for key in ["摘要", "日期-号码", "借方金额", "贷方金额", "余额", "相对科目编码", "对应往来单位"]
                if key in joined
            )

        frames = []
        for pos, title_idx in enumerate(title_rows):
            next_title = title_rows[pos + 1] if pos + 1 < len(title_rows) else len(raw)
            header_idx = None
            for candidate in range(title_idx + 1, min(title_idx + 6, next_title)):
                if header_score(raw.iloc[candidate]) >= 3:
                    header_idx = candidate
                    break
            if header_idx is None:
                continue

            header_values = raw.iloc[header_idx].tolist()
            non_empty_header_cols = [
                i for i, value in enumerate(header_values)
                if value is not None and not pd.isna(value) and str(value).strip() != ""
            ]
            if not non_empty_header_cols:
                continue
            last_col = max(non_empty_header_cols)
            headers = self._dedupe_excel_headers(header_values[:last_col + 1])
            block = raw.iloc[header_idx + 1:next_title, :last_col + 1].copy()
            block.columns = headers
            block = block.dropna(how="all")
            if block.empty:
                continue
            repeated = block.apply(self._row_matches_headers, axis=1)
            block = block.loc[~repeated].copy()
            if block.empty:
                continue

            title_info = self._parse_ledger_block_title(raw.iloc[title_idx, 0])
            party_code = title_info.get("账页往来单位编码")
            party_name = title_info.get("账页往来单位名")

            insert_at = 0
            if "对应往来单位编码" in block.columns:
                insert_at = block.columns.get_loc("对应往来单位编码")
            if "往来单位编码" not in block.columns:
                block.insert(insert_at, "往来单位编码", party_code)
                insert_at += 1
            else:
                missing = block["往来单位编码"].isna() | (block["往来单位编码"].astype(str).str.strip() == "")
                block.loc[missing, "往来单位编码"] = party_code
            if "往来单位名" not in block.columns:
                block.insert(insert_at, "往来单位名", party_name)
            else:
                missing = block["往来单位名"].isna() | (block["往来单位名"].astype(str).str.strip() == "")
                block.loc[missing, "往来单位名"] = party_name

            for col, value in title_info.items():
                block[col] = value
            block["SourceFile"] = filename
            frames.append(block)

        if not frames:
            return None
        return pd.concat(frames, ignore_index=True)

    def _ledger_duplicate_key_frame(self, df):
        """Build a source-independent transaction key for ledger overlap checks."""
        if df is None or df.empty:
            return None, None

        def normalize_text(value):
            if value is None or pd.isna(value):
                return ""
            text = str(value).strip().replace("\t", "")
            if text.lower() in {"nan", "none", "nat"}:
                return ""
            return re.sub(r"\s+", " ", text)

        def normalize_amount(value):
            num = self._to_float(value)
            if num is None:
                return ""
            return f"{round(float(num), 6):.6f}"

        date_col = self._pick_ledger_column(
            df,
            ["日期-号码", "日期号码", "单据号码", "业务日期", "单据日期", "日期", "ParsedDate"],
        )
        party_code_col = self._pick_ledger_column(
            df,
            ["往来单位编码", "账页往来单位编码", "对应往来单位编码", "客户编码", "供应商编码"],
        )
        party_name_col = self._pick_ledger_column(
            df,
            ["往来单位名", "账页往来单位名", "对应往来单位名", "客户名", "客户", "供应商名", "供应商"],
        )
        summary_col = self._pick_ledger_column(df, ["摘要", "说明", "用途"])
        subject_code_col = self._pick_ledger_column(df, ["相对科目编码", "科目编码", "对方科目编码"])
        subject_name_col = self._pick_ledger_column(df, ["相对科目编码名", "科目名", "对方科目"])
        debit_col = self._pick_ledger_amount_column(df, "借方金额") or self._pick_ledger_amount_column(df, "借方")
        credit_col = self._pick_ledger_amount_column(df, "贷方金额") or self._pick_ledger_amount_column(df, "贷方")

        required = [date_col, party_name_col or party_code_col, debit_col or credit_col]
        if any(col is None for col in required):
            return None, None

        key_df = pd.DataFrame(index=df.index)
        for key, col in [
            ("date", date_col),
            ("party_code", party_code_col),
            ("party_name", party_name_col),
            ("summary", summary_col),
            ("subject_code", subject_code_col),
            ("subject_name", subject_name_col),
        ]:
            if col is not None and col in df.columns:
                key_df[key] = df[col].apply(normalize_text)
            else:
                key_df[key] = ""

        for key, col in [("debit", debit_col), ("credit", credit_col)]:
            if col is not None and col in df.columns:
                key_df[key] = df[col].apply(normalize_amount)
            else:
                key_df[key] = ""

        valid = (
            key_df["date"].ne("")
            & (key_df["party_name"].ne("") | key_df["party_code"].ne(""))
            & (key_df["debit"].ne("") | key_df["credit"].ne(""))
        )
        return key_df, valid

    def _source_file_priority(self, source_file, row_position):
        source_name = str(source_file or "").strip()
        mtime = None
        if source_name:
            path = source_name
            if not os.path.isabs(path):
                path = os.path.join(self.base_data_dir, source_name)
            try:
                mtime = os.path.getmtime(path)
            except Exception:
                mtime = None
        # If the physical file is unavailable, later loaded rows are treated as newer.
        return (mtime if mtime is not None else -1.0, row_position)

    def _drop_older_source_duplicate_rows(self, df):
        """
        Remove duplicate ledger rows from older source files.

        When the same transaction appears in multiple imported AR ledgers, the
        newest source file wins. This lets later corrections or month-specific
        exports override older period files instead of being suppressed by them.
        """
        self._last_ledger_duplicate_summary = {
            "rows": 0,
            "debit": 0.0,
            "credit": 0.0,
            "net": 0.0,
        }
        if df is None or df.empty:
            return df
        if "SourceFile" not in df.columns:
            return df

        key_df, valid = self._ledger_duplicate_key_frame(df)
        if key_df is None or valid is None:
            return df

        duplicate_mask = valid & key_df.duplicated(keep=False)
        if not duplicate_mask.any():
            return df

        key_cols = list(key_df.columns)
        grouped_keys = key_df.loc[duplicate_mask].groupby(key_cols, dropna=False).groups
        position_by_index = {idx: pos for pos, idx in enumerate(df.index)}
        drop_indices = set()

        for indices in grouped_keys.values():
            idx_list = list(indices)
            sources = []
            for idx in idx_list:
                source = df.at[idx, "SourceFile"]
                source_text = "" if source is None or pd.isna(source) else str(source).strip()
                if not source_text:
                    continue
                sources.append((
                    idx,
                    source_text,
                    self._source_file_priority(source_text, position_by_index.get(idx, 0)),
                ))
            distinct_sources = {source for _, source, _ in sources}
            if len(distinct_sources) < 2:
                continue

            newest_priority = max(priority for _, _, priority in sources)
            for idx, _, priority in sources:
                if priority < newest_priority:
                    drop_indices.add(idx)

        if not drop_indices:
            return df

        dropped = df.loc[list(drop_indices)]
        debit_col = self._pick_ledger_amount_column(dropped, "借方金额") or self._pick_ledger_amount_column(dropped, "借方")
        credit_col = self._pick_ledger_amount_column(dropped, "贷方金额") or self._pick_ledger_amount_column(dropped, "贷方")

        def amount_sum(col):
            if not col or col not in dropped.columns:
                return 0.0
            return float(sum((self._to_float(v) or 0.0) for v in dropped[col]))

        debit_sum = amount_sum(debit_col)
        credit_sum = amount_sum(credit_col)
        self._last_ledger_duplicate_summary = {
            "rows": len(drop_indices),
            "debit": debit_sum,
            "credit": credit_sum,
            "net": debit_sum - credit_sum,
        }
        self._log_audit(f"应收账款账龄分析已忽略旧来源重复明细 {len(drop_indices)} 行")
        return df.drop(index=list(drop_indices)).copy()

    def _ratio_str(self, count, total):
        if not total:
            return "0/0"
        return f"{count}/{total} ({count/total:.1%})"

    def _extract_issue_ratio(self, detail):
        if detail is None:
            return None
        m = re.search(r"\((\d+(?:\.\d+)?)%\)", str(detail))
        if not m:
            return None
        try:
            return float(m.group(1)) / 100.0
        except Exception:
            return None

    def _promote_critical_quality_warnings(self):
        critical_types = {
            "日期解析失败",
            "月份不匹配",
            "科目缺失",
            "客户/单位缺失",
            "数值缺失",
        }
        for item in self.data_quality_issues:
            severity = str(item.get("severity") or "WARN").upper()
            if severity != "WARN":
                continue
            issue_type = str(item.get("issue_type") or "")
            if issue_type not in critical_types:
                continue
            ratio = self._extract_issue_ratio(item.get("detail"))
            if ratio is not None and ratio >= 0.8:
                item["severity"] = "ERROR"

    def _header_row_repeated(self, df):
        if df is None or df.empty:
            return False
        try:
            row0 = df.iloc[0]
        except Exception:
            return False
        total = 0
        match = 0
        for col in df.columns:
            if col not in row0:
                continue
            val = row0[col]
            if pd.isna(val):
                continue
            total += 1
            if str(val).strip() == str(col).strip():
                match += 1
        return total > 0 and match / total >= 0.6

    def _coerce_numeric(self, series):
        return pd.to_numeric(series, errors='coerce')

    def _is_optional_empty_numeric_column(self, category, col_name):
        name = str(col_name or "")
        # 外币列在本报表中只是辅助信息；很多账簿某个月只有本位币或只有借/贷一边，
        # 整列为空不代表业务数据缺失。
        if "外币" in name:
            return True
        return False

    def _run_data_quality_checks(self):
        self.data_quality_issues = []
        required_categories = {"profit", "cost", "expense", "asset", "sales", "ar"}
        for category, months in self.data.items():
            if not months:
                if category not in required_categories:
                    continue
                self._add_quality_issue(category, None, "缺失数据", "未加载到数据", severity="ERROR")
                continue
            for period, df in months.items():
                if df is None or df.empty:
                    self._add_quality_issue(category, period, "空表", "数据为空", severity="ERROR")
                    continue

                # 基础结构检查
                if period in (None, "Unknown") or not re.match(r"^20\d{2}-\d{2}$", str(period)):
                    self._add_quality_issue(
                        category, period, "期间识别异常",
                        f"月份标识异常: {period}", severity="WARN"
                    )
                if self._header_row_repeated(df):
                    self._add_quality_issue(
                        category, period, "表头重复",
                        "首行与表头高度相似，可能未正确跳过表头行", severity="WARN"
                    )
                normalized_cols = [self._normalize_column_name(c) for c in df.columns]
                dup_cols = [name for name, cnt in Counter(normalized_cols).items() if cnt > 1 and name]
                if dup_cols:
                    self._add_quality_issue(
                        category, period, "字段重复",
                        f"存在重复字段: {', '.join(dup_cols[:5])}", severity="WARN"
                    )

                if category == "sales":
                    required = [
                        ["日期-号码", "日期", "Date"],
                        ["品目编码"],
                        ["数量"],
                        ["销售金额合计", "合计"],
                        ["销售出库供应价合计", "供应价"],
                        ["往来单位名", "收货公司"],
                    ]
                    self._check_required_columns(df, category, period, required)
                elif category == "expense":
                    required = [
                        ["日期", "Date"],
                        ["科目名"],
                        ["借方金额", "金额", "外币借方金额"],
                        ["贷方金额", "外币贷方金额"],
                    ]
                    self._check_required_columns(df, category, period, required)
                elif category == "cost":
                    required = [
                        ["品目编码"],
                        # 三层表头成本表常被扁平化为: 期末 / 期末.1 / 期末.2
                        ["期末数量", "数量", "期末", "期末.1", "期末.2"],
                        ["期末金额", "金额", "期末.2"],
                    ]
                    self._check_required_columns(df, category, period, required)
                elif category == "profit":
                    if df.columns.size < 2:
                        self._add_quality_issue(category, period, "字段不足", "利润表字段过少", severity="ERROR")
                elif category == "asset":
                    if df.columns.size < 2:
                        self._add_quality_issue(category, period, "字段不足", "资产表字段过少", severity="ERROR")
                elif category == "ar":
                    required = [
                        ["往来单位名", "客户名", "客户"],
                        ["借方金额", "金额", "外币借方金额"],
                        ["贷方金额", "外币贷方金额"],
                    ]
                    self._check_required_columns(df, category, period, required)
                elif category == "ap":
                    required = [
                        ["往来单位名", "供应商", "对应往来单位名"],
                        ["借方金额", "金额", "外币借方金额"],
                        ["贷方金额", "外币贷方金额"],
                    ]
                    self._check_required_columns(df, category, period, required)
                elif category == "cash":
                    required = [
                        ["日期-号码", "日期", "Date"],
                        ["借方金额", "金额", "外币借方金额"],
                        ["贷方金额", "外币贷方金额"],
                    ]
                    self._check_required_columns(df, category, period, required)

                # 通用字段质量检查
                amount_cols = self._find_columns_by_keywords(
                    df,
                    ["金额", "合计", "余额", "收入", "成本", "费用", "利润", "税额", "发生额", "本期", "期末", "借方", "贷方"],
                    ["率", "比例", "%", "汇率", "单价", "均价"]
                )
                qty_cols = self._find_columns_by_keywords(
                    df,
                    ["数量", "件数", "重量", "箱数", "台数"],
                    ["率", "比例", "%"]
                )
                key_cols = self._find_columns_by_keywords(
                    df,
                    ["单号", "凭证号", "发票", "订单", "单据号", "编号", "流水号"],
                    ["科目编码", "品目编码", "客户编码", "供应商编码"]
                )

                # 日期解析检查
                date_col = None
                parsed = None
                if "ParsedDate" in df.columns:
                    parsed = pd.to_datetime(df["ParsedDate"], errors="coerce")
                    if parsed.notna().sum() > 0:
                        date_col = "ParsedDate"
                if date_col is None:
                    date_col, parsed = self._pick_best_date_column(
                        df, ["日期", "date", "业务日期", "单据日期", "日期-号码"]
                    )
                if date_col and parsed is not None:
                    total = len(parsed)
                    valid_count = int(parsed.notna().sum())
                    invalid = total - valid_count
                    if total and invalid / total > 0.05:
                        self._add_quality_issue(
                            category, period, "日期解析失败",
                            f"{date_col} 无法解析 {self._ratio_str(invalid, total)}", severity="WARN"
                        )
                    if valid_count > 0:
                        too_future = (parsed > (datetime.now() + pd.Timedelta(days=2))).sum()
                        if too_future > 0:
                            self._add_quality_issue(
                                category, period, "日期异常",
                                f"{date_col} 存在未来日期 {too_future} 行", severity="WARN"
                            )
                        # 若单表覆盖多个自然月(如跨期明细台账)，不做“月份不匹配”告警，避免误报。
                        month_span = parsed.dt.strftime("%Y-%m").dropna().nunique()
                        skip_period_match = month_span > 3
                        if (
                            not skip_period_match
                            and period
                            and re.match(r"^20\d{2}-\d{2}$", str(period))
                        ):
                            month_mask = parsed.dt.strftime('%Y-%m') == str(period)
                            mismatched = parsed.notna() & (~month_mask)
                            if mismatched.any():
                                ratio = mismatched.sum() / max(parsed.notna().sum(), 1)
                                if ratio > 0.2:
                                    self._add_quality_issue(
                                        category, period, "月份不匹配",
                                        f"{date_col} 与期间 {period} 不匹配 {self._ratio_str(mismatched.sum(), parsed.notna().sum())}",
                                        severity="WARN"
                                    )

                # 关键字段缺失率
                if category == "expense":
                    subj_col = self._pick_first_column(df, ["科目名", "科目"])
                    if subj_col:
                        missing = df[subj_col].isna() | (df[subj_col].astype(str).str.strip() == "")
                        if missing.mean() > 0.05:
                            self._add_quality_issue(
                                category, period, "科目缺失",
                                f"{subj_col} 为空 {self._ratio_str(missing.sum(), len(df))}", severity="WARN"
                            )
                if category in {"sales", "cost", "ar", "ap"}:
                    customer_col = self._pick_ledger_column(
                        df,
                        ["往来单位名", "账页往来单位名", "客户名", "客户", "收货公司", "供应商", "对应往来单位名"],
                    )
                    if customer_col:
                        missing = df[customer_col].isna() | (df[customer_col].astype(str).str.strip() == "")
                        if missing.mean() > 0.1:
                            self._add_quality_issue(
                                category, period, "客户/单位缺失",
                                f"{customer_col} 为空 {self._ratio_str(missing.sum(), len(df))}", severity="WARN"
                            )

                # 借贷金额合理性
                if category in {"expense", "ar", "ap", "cash"}:
                    debit_col = self._pick_first_column(df, ["借方金额", "借方"])
                    credit_col = self._pick_first_column(df, ["贷方金额", "贷方"])
                    if debit_col and credit_col:
                        debit = self._coerce_numeric(df[debit_col]).fillna(0)
                        credit = self._coerce_numeric(df[credit_col]).fillna(0)
                        both = (debit != 0) & (credit != 0)
                        if both.any():
                            self._add_quality_issue(
                                category, period, "借贷同填",
                                f"{debit_col}/{credit_col} 同时有值 {self._ratio_str(both.sum(), len(df))}", severity="WARN"
                            )
                        neg_debit = (debit < 0).sum()
                        neg_credit = (credit < 0).sum()
                        if neg_debit:
                            self._add_quality_issue(
                                category, period, "借方负数",
                                f"{debit_col} 存在负数 {self._ratio_str(neg_debit, len(df))}", severity="WARN"
                            )
                        if neg_credit:
                            self._add_quality_issue(
                                category, period, "贷方负数",
                                f"{credit_col} 存在负数 {self._ratio_str(neg_credit, len(df))}", severity="WARN"
                            )

                # 数值字段合理性
                for col_name in amount_cols + qty_cols:
                    values = self._coerce_numeric(df[col_name])
                    total = len(values)
                    valid = values.notna().sum()
                    if total == 0:
                        continue
                    invalid = (df[col_name].notna() & values.isna()).sum()
                    if invalid / total > 0.1:
                        self._add_quality_issue(
                            category, period, "数值解析失败",
                            f"{col_name} 非数值 {self._ratio_str(invalid, total)}", severity="WARN"
                        )
                    if valid == 0:
                        if self._is_optional_empty_numeric_column(category, col_name):
                            continue
                        self._add_quality_issue(
                            category, period, "数值缺失",
                            f"{col_name} 无有效数值", severity="WARN"
                        )
                        continue
                    zeros = (values == 0).sum()
                    if zeros / valid > 0.5:
                        self._add_quality_issue(
                            category, period, "零值占比高",
                            f"{col_name} 为零 {self._ratio_str(zeros, valid)}", severity="WARN"
                        )
                    if category in {"sales", "cost", "asset", "ar", "ap", "cash"} or (category == "expense"):
                        neg = (values < 0).sum()
                        if neg / valid > 0.02:
                            self._add_quality_issue(
                                category, period, "负数异常",
                                f"{col_name} 负数 {self._ratio_str(neg, valid)}", severity="WARN"
                            )
                    # IQR 异常值检测
                    valid_values = values.dropna()
                    if len(valid_values) >= 20:
                        q1 = valid_values.quantile(0.25)
                        q3 = valid_values.quantile(0.75)
                        iqr = q3 - q1
                        if iqr > 0:
                            lower = q1 - 3 * iqr
                            upper = q3 + 3 * iqr
                            outliers = ((valid_values < lower) | (valid_values > upper)).sum()
                            if outliers / len(valid_values) > 0.02:
                                self._add_quality_issue(
                                    category, period, "金额/数量异常值",
                                    f"{col_name} 异常值 {self._ratio_str(outliers, len(valid_values))}", severity="INFO"
                                )

                # 数量与金额配比
                if qty_cols and amount_cols:
                    qty_col = qty_cols[0]
                    amt_col = amount_cols[0]
                    qty = self._coerce_numeric(df[qty_col])
                    amt = self._coerce_numeric(df[amt_col])
                    valid = qty.notna() & amt.notna()
                    if valid.any():
                        zero_qty_nonzero_amt = (qty == 0) & (amt != 0)
                        if zero_qty_nonzero_amt.any():
                            self._add_quality_issue(
                                category, period, "数量金额不匹配",
                                f"{qty_col} 为 0 但 {amt_col} 有值 {self._ratio_str(zero_qty_nonzero_amt.sum(), len(df))}",
                                severity="WARN"
                            )
                        nonzero_qty = qty != 0
                        unit_price = (amt / qty).where(nonzero_qty)
                        if unit_price.notna().sum() >= 20:
                            up_valid = unit_price.dropna()
                            q1 = up_valid.quantile(0.25)
                            q3 = up_valid.quantile(0.75)
                            iqr = q3 - q1
                            if iqr > 0:
                                lower = q1 - 3 * iqr
                                upper = q3 + 3 * iqr
                                outliers = ((up_valid < lower) | (up_valid > upper)).sum()
                                if outliers / len(up_valid) > 0.02:
                                    self._add_quality_issue(
                                        category, period, "单价异常",
                                        f"{amt_col}/{qty_col} 单价异常 {self._ratio_str(outliers, len(up_valid))}",
                                        severity="INFO"
                                    )

                # 关键字段重复
                if key_cols:
                    key_col = key_cols[0]
                    key_series = df[key_col].fillna("").astype(str).str.strip()
                    valid_keys = key_series[
                        (~key_series.str.lower().isin(["", "nan", "none", "nat"]))
                    ]
                    if not valid_keys.empty:
                        dup = valid_keys.duplicated().sum()
                        ratio = dup / len(valid_keys)
                        if ratio > 0.2:
                            severity = "WARN"
                            # 销售订单经常一单多行，重复更常是业务明细粒度而非脏数据。
                            if category == "sales":
                                severity = "INFO"
                            self._add_quality_issue(
                                category, period, "单号重复",
                                f"{key_col} 重复 {self._ratio_str(dup, len(valid_keys))}", severity=severity
                            )

        self._promote_critical_quality_warnings()
        self._log_audit(f"数据质量检查完成，发现 {len(self.data_quality_issues)} 条问题")
    def _to_float(self, value):
        if value is None:
            return None
        if isinstance(value, float) and pd.isna(value):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s:
            return None
        s = s.replace(',', '')
        try:
            return float(s)
        except Exception:
            m = re.search(r'-?\d+(?:\.\d+)?', s)
            if m:
                try:
                    return float(m.group(0))
                except Exception:
                    return None
        return None

    def _safe_sum(self, values):
        vals = [v for v in values if v is not None]
        return sum(vals) if vals else None

    def _normalize_label(self, text):
        if text is None:
            return ''
        s = str(text).strip()
        s = s.replace(" ", "").replace("\u3000", "")
        s = re.sub(r'^[一二三四五六七八九十0-9]+[、\.．]', '', s)
        s = s.lstrip('()（）')
        return s

    def _normalize_profit_label(self, text):
        s = self._normalize_label(text)
        if not s:
            return ''
        s = s.replace('减：', '').replace('减:', '').replace('加：', '').replace('加:', '')
        s = s.replace('其中：', '').replace('其中:', '')
        # 统一利润表常见别名，避免“营业收入/营业成本”与“主营业务收入/主营业务成本”分散写入
        if s == '营业收入':
            s = '主营业务收入'
        elif s == '营业成本':
            s = '主营业务成本'
        elif s == '营业税金及附加':
            s = '税金及附加'
        elif s.startswith('营业收入-'):
            s = '主营业务收入-' + s.split('-', 1)[1]
        elif s.startswith('营业成本-'):
            s = '主营业务成本-' + s.split('-', 1)[1]
        return s

    def _normalize_category(self, text):
        if text is None:
            return ''
        s = str(text).strip()
        if not s:
            return ''
        if s in ('合计', '总计'):
            return '合计'
        s = s.replace(" ", "").replace("\u3000", "")
        s = self._strip_category_share_suffix(s)
        s = re.sub(r'(品类|类别|类目)$', '', s)
        s = re.sub(r'类$', '', s)
        return s

    def _strip_category_share_suffix(self, text):
        if text is None:
            return ''
        s = str(text).strip()
        if not s:
            return ''
        return re.sub(
            r'[（(]\s*(?:收入占比|占比)?\s*[-+]?\d+(?:\.\d+)?[％%]\s*[）)]\s*$',
            '',
            s,
        ).strip()

    def _extract_name_spec_tokens(self, name_spec):
        if name_spec is None:
            return []
        text = str(name_spec).strip()
        if not text:
            return []
        tokens = []
        tokens.extend(re.findall(r'\[([^\[\]]+)\]', text))
        tokens.extend(re.findall(r'【([^【】]+)】', text))
        cleaned = [str(t).strip() for t in tokens if str(t).strip()]
        return cleaned or [text]

    def _resolve_uncategorized_product(self, category, name_spec):
        if category is not None:
            c = str(category).strip()
            if c and c.lower() != 'nan':
                return c
        tokens = self._extract_name_spec_tokens(name_spec)
        if not tokens:
            return None
        joined = " ".join(tokens)
        lowered = joined.lower()
        if "鞋" in joined or "shoe" in lowered:
            return "鞋类"
        return "电器类"

    def _extract_sales_revenue(self, df):
        if df is None:
            return pd.Series(dtype='float64')
        if len(df.index) == 0:
            return pd.Series(index=df.index, dtype='float64')

        # 行级金额优先，避免使用单据级“销售金额合计”导致重复累计。
        preferred_cols = ['合计', '销售合计', '销售金额', '销售收入']
        for col in preferred_cols:
            if col in df.columns:
                series = pd.to_numeric(df.get(col), errors='coerce')
                if series.notna().any():
                    return series

        for col in df.columns:
            name = str(col).strip()
            if (
                '合计' in name
                and '销售金额合计' not in name
                and '销售出库供应价合计' not in name
                and '外币' not in name
            ):
                series = pd.to_numeric(df.get(col), errors='coerce')
                if series.notna().any():
                    return series

        for col in ['销售金额合计', '销售出库供应价合计']:
            if col in df.columns:
                series = pd.to_numeric(df.get(col), errors='coerce')
                if series.notna().any():
                    return series

        return pd.Series([None] * len(df), index=df.index, dtype='float64')

    def _sum_numeric_or_none(self, series):
        """Return numeric sum when there is at least one valid value; otherwise None."""
        if series is None:
            return None
        s = pd.to_numeric(series, errors='coerce')
        if s.notna().sum() == 0:
            return None
        total = s.sum(min_count=1)
        return None if pd.isna(total) else total

    def _calc_sales_metrics_from_group(self, group):
        revenue = self._sum_numeric_or_none(group.get('Revenue'))
        cost = self._sum_numeric_or_none(group.get('Cost'))
        qty = self._sum_numeric_or_none(group.get('Qty'))
        profit = (revenue - cost) if (revenue is not None and cost is not None) else None
        margin = (profit / revenue) if (profit is not None and revenue not in (None, 0)) else None
        return {
            'qty': qty,
            'revenue': revenue,
            'cost': cost,
            'profit': profit,
            'margin': margin,
        }

    def _is_repeated_document_total(self, df, value_series):
        if df is None or value_series is None:
            return False
        if len(df.index) != len(value_series):
            return False

        key_col = None
        for c in ['日期-号码', '单据号', '单据编号', '销售订单号', 'PEDIDO No', '号码']:
            if c in df.columns:
                key_col = c
                break
        if not key_col:
            for c in df.columns:
                name = str(c)
                if any(k in name for k in ['日期-号码', '单据', '订单号', '号码']):
                    key_col = c
                    break
        if not key_col:
            return False

        key_series = df[key_col].astype(str).str.strip()
        val_series = pd.to_numeric(value_series, errors='coerce')
        probe = pd.DataFrame({'k': key_series, 'v': val_series}, index=df.index)
        probe = probe[(probe['k'] != '') & (probe['k'].str.lower() != 'nan') & probe['v'].notna()]
        if probe.empty:
            return False

        probe['v_round'] = probe['v'].round(2)
        grp = probe.groupby('k')['v_round'].agg(['size', 'nunique'])
        grp = grp[grp['size'] >= 2]
        if grp.empty:
            return False

        repeated_ratio = (grp['nunique'] <= 1).mean()
        return repeated_ratio >= 0.8

    def _can_use_raw_cost_fallback(self, df, raw_cost_series, revenue_series):
        raw = pd.to_numeric(raw_cost_series, errors='coerce')
        if raw.notna().sum() == 0:
            return False

        rev = pd.to_numeric(revenue_series, errors='coerce')
        comparable = raw.notna() & rev.notna()
        if comparable.any():
            same_ratio = ((raw[comparable] - rev[comparable]).abs() <= 0.01).mean()
            if same_ratio >= 0.9:
                return False

        if self._is_repeated_document_total(df, raw):
            return False
        return True

    def _dashboard_has_template_formula_layout(self, ws):
        for coord in ("A5", "A7", "A9", "A11", "B13"):
            val = ws[coord].value
            if isinstance(val, str) and val.startswith("="):
                return True
        return False

    def _extend_template_dashboard_metrics(self, ws):
        label_row = 4
        value_row = 5
        value_row_end = 6
        delta_row = 7
        block_width = 4

        def has_label(label):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=label_row, column=col).value
                if isinstance(val, str) and label in val:
                    return True
            return False

        metrics = []
        if not has_label("营业利润"):
            metrics.append({
                "label": "营业利润（元）",
                "value_formula": "=INDEX('经营指标'!$G$2:$G$12, MATCH($B$3,'经营指标'!$A$2:$A$12,0))",
                "delta_formula": None,
                "style_src_col": 5,  # E block
                "delta_src_col": 1,  # A block for MoM pattern
            })
        if not has_label("营业利润率"):
            metrics.append({
                "label": "营业利润率",
                "value_formula": "=INDEX('经营指标'!$N$2:$N$12, MATCH($B$3,'经营指标'!$A$2:$A$12,0))",
                "delta_formula": None,
                "style_src_col": 9,  # I block (rate)
                "delta_src_col": 9,  # I block for pp change pattern
            })
        if not metrics:
            return False

        def range_overlaps(min_row, max_row, min_col, max_col):
            for merged in ws.merged_cells.ranges:
                if merged.max_row < min_row or merged.min_row > max_row:
                    continue
                if merged.max_col < min_col or merged.min_col > max_col:
                    continue
                return True
            return False

        def next_slot(start_col=1):
            max_col = max(ws.max_column, 16) + block_width * len(metrics)
            col = start_col
            while col <= max_col:
                if not ws.cell(row=label_row, column=col).value:
                    if not range_overlaps(label_row, delta_row, col, col + block_width - 1):
                        return col
                col += block_width
            return None

        def copy_block_style(src_col, dst_col):
            for r in range(label_row, delta_row + 1):
                if value_row < r < value_row_end:
                    continue
                for offset in range(block_width):
                    self._copy_cell_style(
                        ws.cell(row=r, column=src_col + offset),
                        ws.cell(row=r, column=dst_col + offset),
                    )
            for r in range(value_row, value_row_end + 1):
                for offset in range(block_width):
                    self._copy_cell_style(
                        ws.cell(row=r, column=src_col + offset),
                        ws.cell(row=r, column=dst_col + offset),
                    )
            for offset in range(block_width):
                src_letter = get_column_letter(src_col + offset)
                dst_letter = get_column_letter(dst_col + offset)
                if src_letter in ws.column_dimensions:
                    ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width

        def merge_block(col):
            ws.merge_cells(start_row=label_row, start_column=col, end_row=label_row, end_column=col + block_width - 1)
            ws.merge_cells(start_row=value_row, start_column=col, end_row=value_row_end, end_column=col + block_width - 1)
            ws.merge_cells(start_row=delta_row, start_column=col, end_row=delta_row, end_column=col + block_width - 1)

        for metric in metrics:
            dest_col = next_slot()
            if not dest_col:
                continue

            src_col = metric["style_src_col"]
            copy_block_style(src_col, dest_col)
            merge_block(dest_col)

            ws.cell(row=label_row, column=dest_col).value = metric["label"]
            ws.cell(row=value_row, column=dest_col).value = metric["value_formula"]

            delta_cell = ws.cell(row=delta_row, column=metric["delta_src_col"]).value
            if isinstance(delta_cell, str) and delta_cell.startswith("="):
                if metric["delta_src_col"] == 1:
                    delta_formula = delta_cell.replace("$C$2:$C$12", "$G$2:$G$12")
                else:
                    delta_formula = delta_cell
                ws.cell(row=delta_row, column=dest_col).value = delta_formula
            elif metric["delta_formula"]:
                ws.cell(row=delta_row, column=dest_col).value = metric["delta_formula"]
        return True

    def _dashboard_summary_card_col(self, ws, label, label_row=8):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=label_row, column=col).value
            if not (isinstance(val, str) and label in val):
                continue
            for merged in ws.merged_cells.ranges:
                if (
                    merged.min_row <= label_row <= merged.max_row
                    and merged.min_col <= col <= merged.max_col
                ):
                    return merged.min_col
            return col
        return None

    def _dashboard_summary_range_overlaps(self, ws, min_row, max_row, min_col, max_col):
        for merged in ws.merged_cells.ranges:
            if merged.max_row < min_row or merged.min_row > max_row:
                continue
            if merged.max_col < min_col or merged.min_col > max_col:
                continue
            return True
        return False

    def _write_dashboard_financial_rate_summary_card_formulas(self, ws):
        col = self._dashboard_summary_card_col(ws, "财务费用率", label_row=8)
        if not col:
            return False

        col_letter = get_column_letter(col)
        prev_month_expr = 'TEXT(DATE(LEFT($B$3,4),RIGHT($B$3,2),1)-1,"yyyy/mm")'
        curr_expr = (
            f'INDEX(\'经营指标\'!$A:$ZZ, MATCH($B$3, \'经营指标\'!$A:$A, 0), '
            f'MATCH("财务费用率", \'经营指标\'!$1:$1, 0))'
        )
        prev_expr = (
            f'INDEX(\'经营指标\'!$A:$ZZ, MATCH({prev_month_expr}, \'经营指标\'!$A:$A, 0), '
            f'MATCH("财务费用率", \'经营指标\'!$1:$1, 0))'
        )
        ws.cell(row=9, column=col).value = f'=IFERROR({curr_expr},"")'
        ws.cell(row=11, column=col).value = (
            f'=IFERROR("较上月："&TEXT(({col_letter}9-{prev_expr})*100,'
            f'"+0.00;-0.00")&"pp","较上月：—")'
        )
        return True

    def _ensure_dashboard_financial_rate_summary_card(self, ws):
        if self._dashboard_summary_card_col(ws, "财务费用率", label_row=8):
            return self._write_dashboard_financial_rate_summary_card_formulas(ws)

        source_col = (
            self._dashboard_summary_card_col(ws, "管理费用率", label_row=8)
            or self._dashboard_summary_card_col(ws, "销售费用率", label_row=8)
        )
        if not source_col:
            return False

        label_row = 8
        value_row = 9
        value_row_end = 10
        delta_row = 11
        block_width = 4

        dest_col = None
        max_scan_col = max(ws.max_column, 16) + block_width
        for col in range(1, max_scan_col + 1, block_width):
            if ws.cell(row=label_row, column=col).value:
                continue
            if self._dashboard_summary_range_overlaps(
                ws,
                label_row,
                delta_row,
                col,
                col + block_width - 1,
            ):
                continue
            dest_col = col
            break
        if not dest_col:
            return False

        for row in range(label_row, delta_row + 1):
            for offset in range(block_width):
                self._copy_cell_style(
                    ws.cell(row=row, column=source_col + offset),
                    ws.cell(row=row, column=dest_col + offset),
                )

        for offset in range(block_width):
            src_letter = get_column_letter(source_col + offset)
            dst_letter = get_column_letter(dest_col + offset)
            if src_letter in ws.column_dimensions:
                ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width

        ws.merge_cells(
            start_row=label_row,
            start_column=dest_col,
            end_row=label_row,
            end_column=dest_col + block_width - 1,
        )
        ws.merge_cells(
            start_row=value_row,
            start_column=dest_col,
            end_row=value_row_end,
            end_column=dest_col + block_width - 1,
        )
        ws.merge_cells(
            start_row=delta_row,
            start_column=dest_col,
            end_row=delta_row,
            end_column=dest_col + block_width - 1,
        )

        ws.cell(row=label_row, column=dest_col).value = "财务费用率"
        return self._write_dashboard_financial_rate_summary_card_formulas(ws)

    def _dashboard_budget_header_row(self, wb):
        if wb is None or "目标_预算" not in wb.sheetnames:
            return 10
        ws = wb["目标_预算"]
        for r in range(1, min(ws.max_row, 50) + 1):
            if str(ws.cell(row=r, column=1).value).strip() == "月份":
                return r
        return 10

    def _dashboard_row_contains(self, ws, text, start_row=1, end_row=None):
        end_row = end_row or ws.max_row
        for r in range(start_row, end_row + 1):
            val = ws.cell(row=r, column=1).value
            if isinstance(val, str) and text in val:
                return r
        return None

    def _ensure_dashboard_financial_rate_table_row(self, ws):
        header_row = self._find_header_row_by_keyword(ws, "指标", max_row=80)
        if not header_row:
            return None
        search_end = min(ws.max_row, header_row + 20)
        existing = self._dashboard_row_contains(ws, "财务费用率", header_row + 1, search_end)
        if existing:
            return existing

        admin_row = self._dashboard_row_contains(ws, "管理费用率", header_row + 1, search_end)
        if not admin_row:
            return None
        insert_row = admin_row + 1
        self._insert_rows_preserve_merges(ws, insert_row, 1)
        self._copy_row_style(ws, admin_row, insert_row, max_col=max(ws.max_column, 8))
        ws.cell(row=insert_row, column=1).value = "财务费用率"
        for col in range(2, 9):
            ws.cell(row=insert_row, column=col).value = None
        return insert_row

    def _repair_dashboard_target_table_formulas(self, ws, budget_header_row=10):
        header_row = self._find_header_row_by_keyword(ws, "指标", max_row=80)
        if not header_row:
            return False

        self._ensure_dashboard_financial_rate_table_row(ws)
        search_end = min(ws.max_row, header_row + 20)

        prev_month_expr = 'TEXT(DATE(LEFT($B$3,4),RIGHT($B$3,2),1)-1,"yyyy/mm")'

        def metric_curr(col_letter):
            return f"INDEX('经营指标'!${col_letter}:${col_letter}, MATCH($B$3, '经营指标'!$A:$A, 0))"

        def metric_prev(col_letter):
            return f"INDEX('经营指标'!${col_letter}:${col_letter}, MATCH({prev_month_expr}, '经营指标'!$A:$A, 0))"

        def metric_curr_by_header(header):
            return (
                f'INDEX(\'经营指标\'!$A:$ZZ, MATCH($B$3, \'经营指标\'!$A:$A, 0), '
                f'MATCH("{header}", \'经营指标\'!$1:$1, 0))'
            )

        def metric_prev_by_header(header):
            return (
                f'INDEX(\'经营指标\'!$A:$ZZ, MATCH({prev_month_expr}, \'经营指标\'!$A:$A, 0), '
                f'MATCH("{header}", \'经营指标\'!$1:$1, 0))'
            )

        def budget_target(header):
            return (
                f'INDEX(\'目标_预算\'!$A:$ZZ, MATCH($B$3, \'目标_预算\'!$A:$A, 0), '
                f'MATCH("{header}", \'目标_预算\'!${budget_header_row}:${budget_header_row}, 0))'
            )

        net_curr = "INDEX('利润表'!$C:$N, MATCH(\"*净利润*\", '利润表'!$A:$A, 0), MATCH($B$3, '利润表'!$C$1:$N$1, 0))"
        net_prev = f"INDEX('利润表'!$C:$N, MATCH(\"*净利润*\", '利润表'!$A:$A, 0), MATCH({prev_month_expr}, '利润表'!$C$1:$N$1, 0))"

        def row_for(label_part):
            return self._dashboard_row_contains(ws, label_part, header_row + 1, search_end)

        def set_target_row(label_part, actual_expr, target_header, prev_expr, compare_kind):
            r = row_for(label_part)
            if not r:
                return
            ws.cell(row=r, column=2).value = f'=IFERROR({actual_expr},"")'
            ws.cell(row=r, column=3).value = f'=IFERROR({budget_target(target_header)},"")'
            ws.cell(row=r, column=4).value = f'=IF(OR(C{r}="",C{r}=0),"",B{r}/C{r})'
            ws.cell(row=r, column=5).value = f'=IF(C{r}="","",B{r}-C{r})'

            if compare_kind == "ratio":
                ws.cell(row=r, column=6).value = f'=IFERROR(IF({prev_expr}=0,"",B{r}/{prev_expr}-1),"")'
                ws.cell(row=r, column=7).value = (
                    f'=IF(F{r}="","",IF(ABS(F{r})>目标_预算!$B$7,'
                    f'"⚠ 环比"&IF(F{r}>0,"上升","下降")&TEXT(ABS(F{r}),"0.0%"),""))'
                )
            elif compare_kind == "rate":
                ws.cell(row=r, column=6).value = f'=IFERROR(B{r}-{prev_expr},"")'
                ws.cell(row=r, column=7).value = (
                    f'=IF(F{r}="","",IF(ABS(F{r})>目标_预算!$C$7,'
                    f'"⚠ 环比"&IF(F{r}>0,"上升","下降")&TEXT(ABS(F{r}),"0.00%")&"pp",""))'
                )
            elif compare_kind == "days":
                ws.cell(row=r, column=6).value = f'=IFERROR(B{r}-{prev_expr},"")'
                ws.cell(row=r, column=7).value = (
                    f'=IF(F{r}="","",IF(ABS(F{r})>目标_预算!$D$7,'
                    f'"⚠ 环比"&IF(F{r}>0,"增加","减少")&TEXT(ABS(F{r}),"0")&"天",""))'
                )

        set_target_row("主营业务收入", metric_curr("C"), "主营业务收入目标", metric_prev("C"), "ratio")
        set_target_row("净利润（元）", net_curr, "营业利润目标", net_prev, "ratio")
        set_target_row("净利润率", f"({net_curr}/{metric_curr('C')})", "营业利润率目标", f"({net_prev}/{metric_prev('C')})", "rate")
        set_target_row("成本率", metric_curr("K"), "成本率目标", metric_prev("K"), "rate")
        set_target_row("销售费用率", metric_curr("L"), "销售费用率目标", metric_prev("L"), "rate")
        set_target_row("管理费用率", metric_curr("M"), "管理费用率目标", metric_prev("M"), "rate")
        set_target_row(
            "财务费用率",
            metric_curr_by_header("财务费用率"),
            "财务费用率目标",
            metric_prev_by_header("财务费用率"),
            "rate",
        )
        set_target_row("应收账款余额", metric_curr("H"), "应收账款余额目标", metric_prev("H"), "ratio")
        set_target_row("存货期末余额", metric_curr("I"), "存货期末余额目标", metric_prev("I"), "ratio")
        set_target_row("存货周转天数", metric_curr("O"), "存货周转天数目标", metric_prev("O"), "days")
        return True

    def _ensure_dashboard_financial_rate_chart_series(self, wb):
        if "仪表盘" not in wb.sheetnames or "经营指标" not in wb.sheetnames:
            return False

        dash_ws = wb["仪表盘"]
        metric_ws = wb["经营指标"]
        header_map = self._get_header_map(metric_ws, 1)
        month_col = header_map.get("月份")
        financial_col = header_map.get("财务费用率")
        if not month_col or not financial_col:
            return False

        required_cols = [
            header_map.get("成本率") or header_map.get("主营业务成本成本率"),
            header_map.get("销售费用率"),
            header_map.get("管理费用率"),
        ]
        required_cols = [c for c in required_cols if c]
        if len(required_cols) < 2:
            return False

        data_start, data_end = self._find_month_data_bounds(metric_ws, month_col, 2)
        if data_end < data_start:
            return False

        for chart in dash_ws._charts:
            value_cols = set()
            for series in chart.series:
                val_formula = None
                try:
                    val_formula = series.val.numRef.f if series.val and series.val.numRef else None
                except Exception:
                    val_formula = None
                parsed = self._parse_chart_range_formula(val_formula)
                if not parsed:
                    continue
                sheet_name, min_col, _, max_col, _ = parsed
                if sheet_name == "经营指标" and min_col == max_col:
                    value_cols.add(min_col)

            if financial_col in value_cols:
                return False
            if not all(col in value_cols for col in required_cols):
                continue

            data_ref = Reference(
                metric_ws,
                min_col=financial_col,
                max_col=financial_col,
                min_row=1,
                max_row=data_end,
            )
            chart.add_data(data_ref, titles_from_data=True)
            cats_ref = self._chart_category_reference(metric_ws, month_col, data_start, data_end)
            chart.set_categories(cats_ref)
            return True
        return False

    def _ensure_dashboard_financial_expense_rate(self, wb):
        if "仪表盘" not in wb.sheetnames:
            return False
        changed = False
        dash_ws = wb["仪表盘"]
        if self._ensure_dashboard_financial_rate_summary_card(dash_ws):
            changed = True
        row = self._ensure_dashboard_financial_rate_table_row(dash_ws)
        if row:
            changed = True
        if self._ensure_dashboard_financial_rate_chart_series(wb):
            changed = True
        return changed

    def _repair_template_dashboard_formulas(self, ws, budget_header_row=10):
        # 模板中原有“环比/较上月”大量采用 MATCH-1 取上月行，在月份倒序时会失效。
        prev_month_expr = 'TEXT(DATE(LEFT($B$3,4),RIGHT($B$3,2),1)-1,"yyyy/mm")'

        def metric_curr(col_letter):
            return f"INDEX('经营指标'!${col_letter}:${col_letter}, MATCH($B$3, '经营指标'!$A:$A, 0))"

        def metric_prev(col_letter):
            return f"INDEX('经营指标'!${col_letter}:${col_letter}, MATCH({prev_month_expr}, '经营指标'!$A:$A, 0))"

        net_curr = "INDEX('利润表'!$C:$N, MATCH(\"*净利润*\", '利润表'!$A:$A, 0), MATCH($B$3, '利润表'!$C$1:$N$1, 0))"
        net_prev = f"INDEX('利润表'!$C:$N, MATCH(\"*净利润*\", '利润表'!$A:$A, 0), MATCH({prev_month_expr}, '利润表'!$C$1:$N$1, 0))"

        def mom_text(curr_expr, prev_expr):
            return (
                f'=IFERROR(IF({prev_expr}=0,"环比：—",'
                f'"环比："&TEXT(({curr_expr}-{prev_expr})/{prev_expr},"+0.00%;-0.00%")),'
                f'"环比：—")'
            )

        def pp_text(curr_expr, prev_expr):
            return (
                # 比率类指标以百分点(pp)展示，需将小数差值乘以 100。
                f'=IFERROR("较上月："&TEXT(({curr_expr}-{prev_expr})*100,"+0.00;-0.00")&"pp",'
                f'"较上月：—")'
            )

        revenue_curr = metric_curr("C")
        revenue_prev = metric_prev("C")
        op_curr = metric_curr("G")
        op_prev = metric_prev("G")
        op_rate_curr = metric_curr("N")
        op_rate_prev = metric_prev("N")
        cost_rate_curr = metric_curr("K")
        cost_rate_prev = metric_prev("K")
        sales_rate_curr = metric_curr("L")
        sales_rate_prev = metric_prev("L")
        admin_rate_curr = metric_curr("M")
        admin_rate_prev = metric_prev("M")
        ar_curr = metric_curr("H")
        ar_prev = metric_prev("H")
        inv_curr = metric_curr("I")
        inv_prev = metric_prev("I")
        days_curr = metric_curr("O")
        days_prev = metric_prev("O")

        ws["A5"].value = f'=IFERROR({revenue_curr},"")'
        ws["E5"].value = f'=IFERROR({net_curr},"")'
        ws["I5"].value = '=IFERROR(E5/A5,"")'
        ws["M5"].value = f'=IFERROR({cost_rate_curr},"")'
        ws["Q5"].value = f'=IFERROR({op_curr},"")'

        ws["A7"].value = mom_text("A5", revenue_prev)
        ws["E7"].value = mom_text("E5", net_prev)
        ws["I7"].value = pp_text("(E5/A5)", f"({net_prev}/{revenue_prev})")
        ws["M7"].value = pp_text("M5", cost_rate_prev)
        ws["Q7"].value = mom_text("Q5", op_prev)

        # 若模板补充了营业利润率区块(U列)，同步修复“较上月”公式。
        if ws.max_column >= 21:
            ws["U5"].value = f'=IFERROR({op_rate_curr},"")'
            ws["U7"].value = pp_text("U5", op_rate_prev)

        ws["A9"].value = f'=IFERROR({sales_rate_curr},"")'
        ws["E9"].value = f'=IFERROR({admin_rate_curr},"")'
        ws["I9"].value = f'=IFERROR({ar_curr},"")'
        ws["M9"].value = f'=IFERROR({inv_curr},"")'

        ws["A11"].value = pp_text("A9", sales_rate_prev)
        ws["E11"].value = pp_text("E9", admin_rate_prev)
        ws["I11"].value = mom_text("I9", ar_prev)
        ws["M11"].value = mom_text("M9", inv_prev)
        self._write_dashboard_financial_rate_summary_card_formulas(ws)

        ws["B13"].value = f'=IFERROR({days_curr},"")'
        ws["D13"].value = f'=IFERROR(B13-({days_prev}),"")'
        self._repair_dashboard_target_table_formulas(ws, budget_header_row=budget_header_row)

        # 仪表盘比较口径统一为“较上月”，避免模板残留“较上一年/较上年”文案误导。
        for r in range(1, min(ws.max_row, 80) + 1):
            for c in range(1, min(ws.max_column, 24) + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if not isinstance(val, str):
                    continue
                new_val = val.replace("较上一年", "较上月").replace("较上年", "较上月")
                if new_val != val:
                    cell.value = new_val

    def _normalize_product_family(self, category, name_hint=None):
        resolved = self._resolve_uncategorized_product(category, name_hint)
        if resolved is None:
            return None
        s = str(resolved).strip().replace(" ", "").replace("\u3000", "")
        if not s or s.lower() == "nan":
            return None
        if "鞋" in s or "shoe" in s.lower():
            return "鞋类"
        if ("机电" in s) or ("电器" in s) or ("工具" in s):
            return "电器类"
        return s

    def _normalize_profit_family_label(self, family_label):
        if family_label is None:
            return None
        s = str(family_label).strip().replace(" ", "").replace("\u3000", "")
        if not s or s.lower() == "nan":
            return None
        if "鞋" in s:
            return "鞋类"
        if ("机电" in s) or ("电器" in s) or ("工具" in s):
            return "电器类"
        return s

    def _build_profit_category_totals(self, month_key):
        sales_df = self.data['sales'].get(month_key)
        if sales_df is None or sales_df.empty:
            return {}
        df = sales_df.copy()
        df['Revenue'] = self._extract_sales_revenue(df)
        df['Qty'] = pd.to_numeric(df.get('数量'), errors='coerce')
        try:
            y, m = month_key.split('-', 1)
            df = self._attach_sales_cost(df, y, m, year_scope="current")
        except Exception:
            df['Cost'] = pd.to_numeric(df.get('销售出库供应价合计'), errors='coerce')
        if 'Cost' not in df.columns:
            df['Cost'] = pd.to_numeric(df.get('销售出库供应价合计'), errors='coerce')

        category_col = '品目组合1名' if '品目组合1名' in df.columns else None
        name_col = None
        for c in ['品目名', '产品名称', '品目名规格']:
            if c in df.columns:
                name_col = c
                break

        totals = {}
        for _, row in df.iterrows():
            family = self._normalize_product_family(
                row.get(category_col) if category_col else None,
                row.get(name_col) if name_col else None,
            )
            if not family:
                continue
            item = totals.setdefault(family, {'revenue': 0.0, 'cost': 0.0})
            rev = self._to_float(row.get('Revenue'))
            cost = self._to_float(row.get('Cost'))
            if rev is not None:
                item['revenue'] += rev
            if cost is not None:
                item['cost'] += cost
        return totals

    def _fill_profit_category_rows(self, ws, month_col, month_key, header_row):
        if not month_col:
            return
        category_totals = self._build_profit_category_totals(month_key)
        if not category_totals:
            return

        top_rev_row = self._find_row_by_label(ws, "主营业务收入", col=1, start_row=2)
        top_cost_row = self._find_row_by_label(ws, "主营业务成本", col=1, start_row=2)
        top_rev_total = self._to_float(ws.cell(row=top_rev_row, column=month_col).value) if top_rev_row else None
        top_cost_total = self._to_float(ws.cell(row=top_cost_row, column=month_col).value) if top_cost_row else None
        raw_rev_total = sum(v.get('revenue', 0) for v in category_totals.values() if v.get('revenue') is not None)
        raw_cost_total = sum(v.get('cost', 0) for v in category_totals.values() if v.get('cost') is not None)

        for r in range((header_row or 1) + 1, ws.max_row + 1):
            raw_label = ws.cell(row=r, column=1).value
            if not isinstance(raw_label, str):
                continue
            label = raw_label.strip().replace(" ", "").replace("\u3000", "")
            if '-' not in label:
                continue
            prefix, family_part = label.split('-', 1)
            prefix_norm = self._normalize_profit_label(prefix)
            metric = None
            if prefix_norm in ('主营业务收入', '营业收入'):
                metric = 'revenue'
            elif prefix_norm in ('主营业务成本', '营业成本'):
                metric = 'cost'
            if metric is None:
                continue

            family_key = self._normalize_profit_family_label(family_part)
            if not family_key:
                continue
            item = category_totals.get(family_key)
            if not item:
                continue
            value = item.get(metric)
            if metric == 'revenue' and value is not None and top_rev_total is not None and raw_rev_total:
                value = top_rev_total * value / raw_rev_total
            elif metric == 'cost' and value is not None and top_cost_total is not None and raw_cost_total:
                value = top_cost_total * value / raw_cost_total
            if value is None:
                continue

            current = ws.cell(row=r, column=month_col).value
            if current is None or (isinstance(current, str) and not current.strip()):
                ws.cell(row=r, column=month_col).value = value

    def _sync_profit_duplicate_rows(self, ws, month_col):
        if not month_col:
            return
        pairs = [
            ("二、营业利润", "营业利润"),
            ("三、利润总额", "利润总额"),
            ("四、净利润", "净利润"),
        ]
        for primary, secondary in pairs:
            primary_row = self._find_row_by_label(ws, primary, col=1, start_row=2)
            secondary_row = self._find_row_by_label(ws, secondary, col=1, start_row=2)
            if not primary_row or not secondary_row or primary_row == secondary_row:
                continue
            p_val = self._to_float(ws.cell(row=primary_row, column=month_col).value)
            s_val = self._to_float(ws.cell(row=secondary_row, column=month_col).value)
            if p_val is not None and s_val is None:
                ws.cell(row=secondary_row, column=month_col).value = p_val
            elif s_val is not None and p_val is None:
                ws.cell(row=primary_row, column=month_col).value = s_val

    def _first_numeric_in_row(self, row):
        for v in row:
            if pd.isna(v):
                continue
            if isinstance(v, (datetime, pd.Timestamp)):
                continue
            if isinstance(v, (int, float)):
                return float(v)
            val = self._to_float(v)
            if val is not None:
                return val
        return None

    def _numeric_values_in_row(self, row):
        vals = []
        for v in row:
            if pd.isna(v):
                continue
            if isinstance(v, (int, float)):
                vals.append(float(v))
                continue
            val = self._to_float(v)
            if val is not None:
                vals.append(val)
        return vals

    def _month_key_to_label(self, month_key):
        return month_key.replace('-', '/')

    def _month_key_to_history_label(self, month_key):
        key = self._month_label_exact_to_key(month_key) or self._extract_month_key_from_text(month_key)
        if not key:
            return month_key
        m = re.match(r'^(20\d{2})-(\d{1,2})$', str(key))
        if not m:
            return month_key
        return f"{m.group(1)}.{int(m.group(2))}月"

    def _history_chart_category_col(self, ws, cat_col, data_start, data_end, year_scope=None):
        if year_scope is None:
            year_scope = self.year_scope
        if not self._scope_uses_history_chart_labels(year_scope) or not cat_col or data_end < data_start:
            return cat_col
        labels = []
        already_history = True
        for row in range(data_start, data_end + 1):
            value = ws.cell(row=row, column=cat_col).value
            if value is None or str(value).strip() == "":
                labels.append(None)
                continue
            month_key = self._month_label_exact_to_key(value)
            if not month_key:
                return cat_col
            history_label = self._month_key_to_history_label(month_key)
            labels.append(history_label)
            if str(value).strip() != history_label:
                already_history = False
        if not any(label for label in labels):
            return cat_col
        if already_history:
            return cat_col

        helper_col = ws.max_column + 1
        header_row = max(1, data_start - 1)
        ws.cell(row=header_row, column=helper_col).value = "图表月份"
        for offset, label in enumerate(labels):
            cell = ws.cell(row=data_start + offset, column=helper_col)
            cell.value = label
            cell.number_format = "@"
        return helper_col

    def _chart_category_reference(self, ws, cat_col, data_start, data_end, year_scope=None):
        chart_cat_col = self._history_chart_category_col(ws, cat_col, data_start, data_end, year_scope)
        return Reference(ws, min_col=chart_cat_col, min_row=data_start, max_row=data_end)

    def _history_chart_category_reference_from_formula(self, wb, formula, year_scope=None):
        if year_scope is None:
            year_scope = self.year_scope
        if not self._scope_uses_history_chart_labels(year_scope):
            return None
        parsed = self._parse_chart_range_formula(formula)
        if not parsed:
            return None
        sheet_name, min_col, min_row, max_col, max_row = parsed
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        if min_col == max_col:
            helper_col = self._history_chart_category_col(ws, min_col, min_row, max_row, year_scope)
            if helper_col == min_col:
                return None
            return Reference(ws, min_col=helper_col, min_row=min_row, max_row=max_row)
        if min_row != max_row:
            return None

        labels = []
        already_history = True
        for col in range(min_col, max_col + 1):
            value = ws.cell(row=min_row, column=col).value
            if value is None or str(value).strip() == "":
                labels.append(None)
                continue
            month_key = self._month_label_exact_to_key(value)
            if not month_key:
                return None
            history_label = self._month_key_to_history_label(month_key)
            labels.append(history_label)
            if str(value).strip() != history_label:
                already_history = False
        if not any(label for label in labels) or already_history:
            return None

        helper_row = ws.max_row + 1
        for offset, label in enumerate(labels):
            cell = ws.cell(row=helper_row, column=min_col + offset)
            cell.value = label
            cell.number_format = "@"
        return Reference(ws, min_col=min_col, max_col=max_col, min_row=helper_row, max_row=helper_row)

    def _rewrite_chart_month_categories_for_history(self, wb, year_scope=None):
        if year_scope is None:
            year_scope = self.year_scope
        if not self._scope_uses_history_chart_labels(year_scope):
            return 0
        changed = 0
        for ws in wb.worksheets:
            for chart in getattr(ws, "_charts", []):
                formula = self._chart_category_formula(chart)
                ref = self._history_chart_category_reference_from_formula(wb, formula, year_scope)
                if ref is None:
                    continue
                try:
                    chart.set_categories(ref)
                    changed += 1
                except Exception:
                    continue
        if changed:
            print(f"已转换历年图表月份标签: {changed}")
        return changed

    def _month_key_to_period_label(self, month_key):
        return month_key.replace('-', '/') + '/01-' + month_key.replace('-', '/') + '/01'

    def _label_to_month_key(self, label):
        if not label:
            return None
        m = re.search(r'(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?', str(label))
        if not m:
            return None
        return f"{m.group(1)}-{int(m.group(2)):02d}"

    def _extract_month_key_from_text(self, text):
        if text is None:
            return None
        if isinstance(text, (datetime, pd.Timestamp)):
            return text.strftime('%Y-%m')
        s = str(text).strip()
        if not s:
            return None
        m = re.search(r'(20\d{2})\s*年\s*(\d{1,2})\s*月', s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
        m = re.search(r'(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?', s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
        return None

    def _chart_category_formula(self, chart):
        if not chart or not getattr(chart, "series", None):
            return None
        if not chart.series:
            return None
        cat = getattr(chart.series[0], "cat", None)
        if not cat:
            return None
        if getattr(cat, "strRef", None) is not None and getattr(cat.strRef, "f", None):
            return cat.strRef.f
        if getattr(cat, "numRef", None) is not None and getattr(cat.numRef, "f", None):
            return cat.numRef.f
        return None

    def _parse_chart_range_formula(self, formula):
        if not formula or "!" not in formula:
            return None
        if formula.startswith("="):
            formula = formula[1:]
        if "[" in formula:
            return None
        if "," in formula:
            return None
        sheet_part, cell_range = formula.rsplit("!", 1)
        sheet_name = sheet_part.strip("'").replace("''", "'")
        if ":" not in cell_range:
            cell_range = f"{cell_range}:{cell_range}"
        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        except Exception:
            return None
        return sheet_name, min_col, min_row, max_col, max_row

    def _extract_month_keys_from_range(self, ws, min_col, min_row, max_col, max_row):
        keys = []
        if min_col == max_col:
            for r in range(min_row, max_row + 1):
                val = ws.cell(row=r, column=min_col).value
                if val is None or str(val).strip() == "":
                    continue
                key = self._month_label_exact_to_key(val) or self._extract_month_key_from_text(val)
                if not key:
                    return None
                keys.append(key)
        elif min_row == max_row:
            for c in range(min_col, max_col + 1):
                val = ws.cell(row=min_row, column=c).value
                if val is None or str(val).strip() == "":
                    continue
                key = self._month_label_exact_to_key(val) or self._extract_month_key_from_text(val)
                if not key:
                    return None
                keys.append(key)
        else:
            return None
        return keys if len(keys) >= 2 else None

    def _apply_chart_month_axis_order(self, wb, chart):
        formula = self._chart_category_formula(chart)
        parsed = self._parse_chart_range_formula(formula)
        if not parsed:
            return False
        sheet_name, min_col, min_row, max_col, max_row = parsed
        if sheet_name not in wb.sheetnames:
            return False
        keys = self._extract_month_keys_from_range(
            wb[sheet_name], min_col, min_row, max_col, max_row
        )
        if not keys:
            return False
        is_desc = all(keys[i] >= keys[i + 1] for i in range(len(keys) - 1))
        is_asc = all(keys[i] <= keys[i + 1] for i in range(len(keys) - 1))
        axis = getattr(chart, "x_axis", None)
        if not axis or not getattr(axis, "scaling", None):
            return False
        if is_desc and not is_asc:
            axis.scaling.orientation = "maxMin"
            return True
        if is_asc and not is_desc:
            axis.scaling.orientation = "minMax"
            return True
        return False

    def _ensure_chart_month_axis_order(self, wb):
        adjusted = 0
        for ws in wb.worksheets:
            for chart in getattr(ws, "_charts", []):
                if self._apply_chart_month_axis_order(wb, chart):
                    adjusted += 1
        if adjusted:
            print(f"已调整图表日期顺序: {adjusted}")

    def _build_range_formula(self, sheet_name, min_col, min_row, max_col, max_row):
        sheet_ref = f"'{sheet_name}'" if any(ch in str(sheet_name) for ch in " -()") else str(sheet_name)
        return (
            f"{sheet_ref}!${get_column_letter(min_col)}${min_row}:"
            f"${get_column_letter(max_col)}${max_row}"
        )

    def _trim_chart_data_ranges(self, wb):
        trimmed = 0
        for ws in wb.worksheets:
            for chart in getattr(ws, "_charts", []):
                if self._trim_single_chart_data_range(wb, chart):
                    trimmed += 1
        if trimmed:
            print(f"已裁剪图表空白范围: {trimmed}")

    def _trim_single_chart_data_range(self, wb, chart):
        if not chart or not getattr(chart, "series", None):
            return False
        if not chart.series:
            return False

        cat_formula = self._chart_category_formula(chart)
        parsed_cat = self._parse_chart_range_formula(cat_formula)
        if not parsed_cat:
            return False
        cat_sheet, cat_min_col, cat_min_row, cat_max_col, cat_max_row = parsed_cat
        if cat_sheet not in wb.sheetnames:
            return False
        if cat_min_col != cat_max_col:
            return False

        val_ranges = []
        for series in chart.series:
            try:
                val_formula = series.val.numRef.f if series.val and series.val.numRef else None
            except Exception:
                val_formula = None
            parsed_val = self._parse_chart_range_formula(val_formula)
            if not parsed_val:
                return False
            val_sheet, val_min_col, val_min_row, val_max_col, val_max_row = parsed_val
            if val_sheet != cat_sheet or val_min_col != val_max_col:
                return False
            if (val_min_row, val_max_row) != (cat_min_row, cat_max_row):
                return False
            val_ranges.append((series, val_min_col, val_max_col))

        data_ws = wb[cat_sheet]
        active_rows = []
        for row in range(cat_min_row, cat_max_row + 1):
            cat_val = data_ws.cell(row=row, column=cat_min_col).value
            has_cat = cat_val not in (None, "")
            has_val = False
            for _, val_col, _ in val_ranges:
                value = data_ws.cell(row=row, column=val_col).value
                if value not in (None, ""):
                    has_val = True
                    break
            if has_cat and has_val:
                active_rows.append(row)

        if not active_rows:
            return False

        new_min_row = min(active_rows)
        new_max_row = max(active_rows)
        if (new_min_row, new_max_row) == (cat_min_row, cat_max_row):
            return False

        new_cat_formula = self._build_range_formula(cat_sheet, cat_min_col, new_min_row, cat_max_col, new_max_row)
        for series, val_col, _ in val_ranges:
            try:
                if series.cat and series.cat.strRef and series.cat.strRef.f:
                    series.cat.strRef.f = new_cat_formula
                elif series.cat and series.cat.numRef and series.cat.numRef.f:
                    series.cat.numRef.f = new_cat_formula
            except Exception:
                pass
            new_val_formula = self._build_range_formula(cat_sheet, val_col, new_min_row, val_col, new_max_row)
            try:
                if series.val and series.val.numRef and series.val.numRef.f:
                    series.val.numRef.f = new_val_formula
            except Exception:
                pass
        return True

    def _row_value_for_month(self, row, candidate_cols, month_key=None):
        cols = list(candidate_cols)
        if not cols:
            return None

        if month_key:
            tagged_cols = []
            matched_col = None
            for col in cols:
                col_key = self._extract_month_key_from_text(col)
                if not col_key:
                    continue
                tagged_cols.append(col)
                if col_key == month_key and matched_col is None:
                    matched_col = col
            if matched_col is not None:
                return self._to_float(row.get(matched_col))
            if tagged_cols:
                # If monthly columns are explicit but target month value is missing,
                # avoid falling back to prior-year / comparison columns.
                return None

        return self._first_numeric_in_row([row.get(col) for col in cols])

    def _month_key_in_scope(self, month_key, target_year, target_month, year_scope=None):
        if not month_key:
            return False
        if year_scope is None:
            year_scope = self.year_scope
        if not target_year or not target_month:
            return True
        year_scope = self._normalize_year_scope(year_scope)
        limit_key = f"{target_year}-{int(target_month):02d}"
        month_key = str(month_key)
        if month_key > limit_key:
            return False
        start_key = self._scope_start_key(target_year, year_scope)
        if start_key and month_key < start_key:
            return False
        return True

    def _apply_current_scope_visibility(self, wb, target_year, target_month, year_scope=None):
        if year_scope is None:
            year_scope = self.year_scope
        year_scope = self._normalize_year_scope(year_scope)
        if year_scope not in {"current", "recent_two_years"} or not target_year or not target_month:
            return

        if "利润表" in wb.sheetnames:
            self._hide_out_of_scope_month_columns_simple(wb["利润表"], target_year, target_month, header_row=1, year_scope=year_scope)
        if "资产负债表" in wb.sheetnames:
            self._hide_out_of_scope_month_columns_simple(wb["资产负债表"], target_year, target_month, header_row=1, year_scope=year_scope)
        if "按品类汇总(按月)" in wb.sheetnames:
            self._hide_out_of_scope_month_columns_grouped_suffix(wb["按品类汇总(按月)"], target_year, target_month, year_scope=year_scope)
        if "按产品汇总_含合计" in wb.sheetnames:
            self._hide_out_of_scope_month_columns_grouped_suffix(wb["按产品汇总_含合计"], target_year, target_month, year_scope=year_scope)
        for sheet_name in ["本量利分析", "目标_预算", "经营指标"]:
            if sheet_name in wb.sheetnames:
                self._hide_out_of_scope_month_rows(wb[sheet_name], target_year, target_month, year_scope=year_scope)
        for sheet_name in ["经营指标", "费用对比"]:
            if sheet_name in wb.sheetnames:
                self._hide_rows_before_first_month(wb[sheet_name], header_keyword="月份", month_col=1)

    def _hide_out_of_scope_month_columns_simple(self, ws, target_year, target_month, header_row=1, year_scope="current"):
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            month_key = self._label_to_month_key(header)
            if not month_key:
                continue
            ws.column_dimensions[get_column_letter(col)].hidden = not self._month_key_in_scope(
                month_key, target_year, target_month, year_scope
            )

    def _hide_out_of_scope_month_columns_grouped_suffix(self, ws, target_year, target_month, header_row=1, year_scope="current"):
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if not header:
                continue
            m = re.match(r'(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?_', str(header).strip())
            if not m:
                continue
            month_key = f"{m.group(1)}-{int(m.group(2)):02d}"
            ws.column_dimensions[get_column_letter(col)].hidden = not self._month_key_in_scope(
                month_key, target_year, target_month, year_scope
            )

    def _hide_out_of_scope_month_rows(self, ws, target_year, target_month, header_keyword="月份", month_col=1, year_scope="current"):
        header_row = self._find_header_row_by_keyword(ws, header_keyword, max_row=50)
        if not header_row:
            return
        for r in range(header_row + 1, ws.max_row + 1):
            month_key = self._label_to_month_key(ws.cell(row=r, column=month_col).value)
            if not month_key:
                continue
            ws.row_dimensions[r].hidden = not self._month_key_in_scope(
                month_key, target_year, target_month, year_scope
            )

    def _hide_leading_blank_rows(self, ws, header_row=1):
        first_content_row = None
        for r in range(header_row + 1, ws.max_row + 1):
            if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
                first_content_row = r
                break
        if first_content_row is None:
            return
        for r in range(header_row + 1, first_content_row):
            ws.row_dimensions[r].hidden = True

    def _hide_rows_before_first_month(self, ws, header_keyword="月份", month_col=1):
        header_row = self._find_header_row_by_keyword(ws, header_keyword, max_row=50)
        if not header_row:
            return
        first_month_row = None
        for r in range(header_row + 1, ws.max_row + 1):
            month_key = self._label_to_month_key(ws.cell(row=r, column=month_col).value)
            if month_key:
                first_month_row = r
                break
        if first_month_row is None:
            return
        for r in range(header_row + 1, first_month_row):
            ws.row_dimensions[r].hidden = True

    def _filter_month_keys(self, month_keys, target_year, target_month, year_scope=None):
        if year_scope is None:
            year_scope = self.year_scope
        if not target_year or not target_month:
            return sorted(month_keys)
        year_scope = self._normalize_year_scope(year_scope)
        limit_key = f"{target_year}-{int(target_month):02d}"
        filtered = [str(m) for m in month_keys if str(m) <= limit_key]
        start_key = self._scope_start_key(target_year, year_scope)
        if start_key:
            filtered = [m for m in filtered if m >= start_key]
        return sorted(filtered)

    def _filter_df_by_scope(self, df, target_year, target_month, year_scope=None):
        if df is None or df.empty or 'MonthStr' not in df.columns:
            return df
        if year_scope is None:
            year_scope = self.year_scope
        if not target_year or not target_month:
            return df
        year_scope = self._normalize_year_scope(year_scope)
        limit_key = f"{target_year}-{int(target_month):02d}"
        month_values = df['MonthStr'].astype(str)
        scoped = df[month_values <= limit_key]
        start_key = self._scope_start_key(target_year, year_scope)
        if start_key:
            scoped = scoped[scoped['MonthStr'].astype(str) >= start_key]
        return scoped

    def _extract_profit_metrics(self, df, month_key=None):
        if df is None or df.empty:
            return {}
        name_col = df.columns[0]
        metrics = {}
        value_cols = list(df.columns[1:])

        def assign_metric(key, value):
            if key not in metrics or metrics.get(key) is None:
                metrics[key] = value

        def fallback_rows(match_df, key):
            rows = [match_row for _, match_row in match_df.iterrows()]
            summary_rows = []
            for match_row in rows:
                label = self._normalize_profit_label(match_row[name_col])
                if label and '-' not in label:
                    summary_rows.append(match_row)
            if summary_rows:
                return summary_rows
            return [] if key in metrics else rows

        for _, row in df.iterrows():
            label = self._normalize_profit_label(row[name_col])
            if not label:
                continue
            value = self._row_value_for_month(row, value_cols, month_key)
            if label == "主营业务收入":
                assign_metric('revenue', value)
            elif label == "主营业务成本":
                assign_metric('cost', value)
            elif label == "税金及附加":
                assign_metric('tax_surcharges', value)
            elif label == "销售费用":
                assign_metric('sales_expense', value)
            elif label == "管理费用":
                assign_metric('admin_expense', value)
            elif label == "研发费用":
                assign_metric('rd_expense', value)
            elif label == "财务费用":
                assign_metric('financial_expense', value)
            elif label == "其他收益":
                assign_metric('other_income', value)
            elif label == "投资收益":
                assign_metric('investment_income', value)
            elif label == "信用减值损失":
                assign_metric('credit_impairment_loss', value)
            elif label == "资产减值损失":
                assign_metric('asset_impairment_loss', value)
            elif label in ["资产处置收益", "资产处置损益"]:
                assign_metric('asset_disposal_gain', value)
            elif label == "营业利润":
                assign_metric('operating_profit', value)
            elif label == "营业外收入":
                assign_metric('non_operating_income', value)
            elif label == "营业外支出":
                assign_metric('non_operating_expense', value)
            elif label == "利润总额":
                assign_metric('total_profit', value)
            elif label in ["所得税费用", "所得税"]:
                assign_metric('income_tax', value)
            elif label == "净利润":
                assign_metric('net_profit', value)
        # Fallback if exact matches were not found
        if metrics.get('sales_expense') is None:
            match = df[df[name_col].astype(str).str.contains('销售费用', na=False)]
            if not match.empty:
                for match_row in fallback_rows(match, 'sales_expense'):
                    value = self._row_value_for_month(match_row, value_cols, month_key)
                    if value is not None:
                        assign_metric('sales_expense', value)
                        break
        if metrics.get('admin_expense') is None:
            match = df[df[name_col].astype(str).str.contains('管理费用', na=False)]
            if not match.empty:
                for match_row in fallback_rows(match, 'admin_expense'):
                    value = self._row_value_for_month(match_row, value_cols, month_key)
                    if value is not None:
                        assign_metric('admin_expense', value)
                        break
        fallback_pairs = [
            ("税金及附加", "tax_surcharges"),
            ("研发费用", "rd_expense"),
            ("财务费用", "financial_expense"),
            ("其他收益", "other_income"),
            ("投资收益", "investment_income"),
            ("信用减值", "credit_impairment_loss"),
            ("资产减值", "asset_impairment_loss"),
            ("资产处置", "asset_disposal_gain"),
            ("营业外收入", "non_operating_income"),
            ("营业外支出", "non_operating_expense"),
            ("利润总额", "total_profit"),
            ("所得税", "income_tax"),
        ]
        for keyword, key in fallback_pairs:
            if metrics.get(key) is not None:
                continue
            match = df[df[name_col].astype(str).str.contains(keyword, na=False)]
            if not match.empty:
                for match_row in fallback_rows(match, key):
                    value = self._row_value_for_month(match_row, value_cols, month_key)
                    if value is not None:
                        assign_metric(key, value)
                        break
        return metrics

    def _extract_asset_metrics(self, df, month_key=None):
        if df is None or df.empty:
            return {}
        metrics = {}
        def split_suffix(col):
            s = str(col)
            m = re.match(r'^(.*?)(\.\d+)$', s)
            if m:
                return m.group(1), m.group(2)
            return s, ""

        def looks_like_date_col(name):
            if not name:
                return False
            text = str(name)
            if re.search(r'20\d{2}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日', text):
                return True
            if re.search(r'20\d{2}[/-]\d{1,2}[/-]\d{1,2}', text):
                return True
            if re.search(r'20\d{2}[/-]\d{1,2}$', text):
                return True
            return False

        # Build label/value column groups by suffix (e.g., "", ".1")
        suffix_groups = {}
        for col in df.columns:
            base, suffix = split_suffix(col)
            suffix_groups.setdefault(suffix, {})[base] = col

        label_groups = []
        for _, base_map in suffix_groups.items():
            label_col = None
            for base in ["财务报表显示名", "科目名", "科目名称"]:
                if base in base_map:
                    label_col = base_map[base]
                    break
            if not label_col:
                continue
            value_cols = [col for base, col in base_map.items() if looks_like_date_col(base)]
            if not value_cols:
                exclude = {label_col}
                if "科目编码" in base_map:
                    exclude.add(base_map["科目编码"])
                value_cols = [col for col in base_map.values() if col not in exclude]
            label_groups.append((label_col, value_cols))

        if not label_groups:
            label_groups = [(df.columns[0], list(df.columns[1:]))]

        def assign_metric(key, value):
            if value is None:
                return
            if metrics.get(key) is None:
                metrics[key] = value

        for _, row in df.iterrows():
            for label_col, value_cols in label_groups:
                raw_label = row.get(label_col)
                if raw_label is None or str(raw_label).strip() == "":
                    continue
                label = self._normalize_label(raw_label)
                if not label:
                    continue
                value = self._row_value_for_month(row, value_cols, month_key)
                if label in ["货币资金", "现金", "银行存款", "现金及现金等价物"]:
                    assign_metric('cash', value)
                elif label in ["应收账款", "应收账款净额", "应收账款净值", "应收账款余额"]:
                    assign_metric('ar_balance', value)
                elif label == "存货":
                    if value is not None:
                        assign_metric('inventory_end', value)
                    nums = self._numeric_values_in_row([row.get(c) for c in value_cols])
                    if len(nums) > 1 and metrics.get('inventory_start') is None:
                        metrics['inventory_start'] = nums[1]
                elif label in ["固定资产", "固定资产净值"]:
                    assign_metric('fixed_assets', value)
                elif label in ["应付账款"]:
                    assign_metric('ap_balance', value)
                elif label in ["短期借款"]:
                    assign_metric('short_debt', value)
                elif label in ["长期借款"]:
                    assign_metric('long_debt', value)
                elif label in ["应付票据"]:
                    assign_metric('notes_payable', value)
                elif label in ["应交税费"]:
                    assign_metric('taxes_payable', value)
                elif label in ["其他应付款"]:
                    assign_metric('other_payables', value)
                elif label in ["预收账款"]:
                    assign_metric('advances_from_customers', value)
                elif label in ["实收资本", "股本"]:
                    assign_metric('paid_in_capital', value)
                elif label in ["本年利润"]:
                    assign_metric('current_year_profit', value)
                elif label in ["利润分配"]:
                    assign_metric('profit_distribution', value)
                elif label in ["未分配利润"]:
                    assign_metric('retained_earnings', value)
                elif label in ["流动资产合计", "流动资产"] or label.startswith("流动资产"):
                    assign_metric('current_assets', value)
                elif label in ["非流动资产合计", "非流动资产", "长期资产合计"] or label.startswith("非流动资产") or label.startswith("长期资产"):
                    assign_metric('non_current_assets', value)
                elif label in ["流动负债合计", "流动负债"] or label.startswith("流动负债"):
                    assign_metric('current_liabilities', value)
                elif label in ["非流动负债合计", "非流动负债"] or label.startswith("非流动负债"):
                    assign_metric('non_current_liabilities', value)
                elif label in ["资产总计", "资产合计", "总资产"] or label.startswith("资产总计"):
                    assign_metric('total_assets', value)
                elif label in ["负债合计", "负债总计", "总负债"] or label.startswith("负债合计"):
                    assign_metric('total_liabilities', value)
                elif label in ["所有者权益合计", "股东权益合计", "所有者权益(或股东权益)合计", "净资产"] or label.startswith("所有者权益合计") or label.startswith("股东权益合计") or (label.startswith("所有者权益") and "合计" in label):
                    assign_metric('total_equity', value)
                elif label.startswith("负债和所有者权益") and metrics.get("total_assets") is None:
                    assign_metric('total_assets', value)
        return metrics

    def _build_monthly_metrics(self, target_year=None, target_month=None, year_scope=None):
        keys = set(self.data['profit'].keys()) | set(self.data['asset'].keys())
        month_keys = self._filter_month_keys(keys, target_year, target_month, year_scope)
        metrics_by_month = {}
        for m in month_keys:
            profit_metrics = self._extract_profit_metrics(self.data['profit'].get(m), m)
            asset_metrics = self._extract_asset_metrics(self.data['asset'].get(m), m)
            combined = {}
            combined.update(profit_metrics)
            combined.update(asset_metrics)
            revenue = combined.get('revenue')
            cost = combined.get('cost')
            sales_expense = combined.get('sales_expense')
            admin_expense = combined.get('admin_expense')
            financial_expense = combined.get('financial_expense')
            operating_profit = combined.get('operating_profit')
            inventory_start = combined.get('inventory_start')
            inventory_end = combined.get('inventory_end')

            if combined.get('gross_profit') is None and revenue is not None and cost is not None:
                combined['gross_profit'] = revenue - cost
            if combined.get('total_profit') is None:
                non_op_income = combined.get('non_operating_income')
                non_op_expense = combined.get('non_operating_expense')
                if operating_profit is not None and non_op_income is not None and non_op_expense is not None:
                    combined['total_profit'] = operating_profit + non_op_income - non_op_expense
            if combined.get('net_profit') is None:
                total_profit = combined.get('total_profit')
                income_tax = combined.get('income_tax')
                if total_profit is not None and income_tax is not None:
                    combined['net_profit'] = total_profit - income_tax

            if revenue is not None and revenue != 0:
                if cost is not None:
                    combined['cost_rate'] = cost / revenue
                if sales_expense is not None:
                    combined['sales_expense_rate'] = sales_expense / revenue
                if admin_expense is not None:
                    combined['admin_expense_rate'] = admin_expense / revenue
                if financial_expense is not None:
                    combined['financial_expense_rate'] = financial_expense / revenue
                if operating_profit is not None:
                    combined['operating_profit_rate'] = operating_profit / revenue
                if combined.get('net_profit') is not None:
                    combined['net_profit_rate'] = combined['net_profit'] / revenue

            avg_inventory = None
            if inventory_start is not None and inventory_end is not None:
                avg_inventory = (inventory_start + inventory_end) / 2
            elif inventory_end is not None:
                avg_inventory = inventory_end
            elif inventory_start is not None:
                avg_inventory = inventory_start
            if avg_inventory is not None and cost is not None and cost != 0:
                combined['inventory_turnover_days'] = avg_inventory / cost * 365
            metrics_by_month[m] = combined
        return metrics_by_month

    def _update_management_metrics_sheet(
        self,
        ws,
        metrics_by_month,
        target_year,
        target_month,
        year_scope=None,
        metrics_by_month_all=None,
    ):
        if ws.max_row < 2:
            return
        header_map = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                header_map[str(header).strip()] = col
        for required_header in ("财务费用", "财务费用率"):
            if required_header not in header_map:
                header_map[required_header] = self._ensure_header_column(ws, required_header)

        month_rows = {}
        non_month_rows = []
        for r in range(2, ws.max_row + 1):
            m_val = ws.cell(row=r, column=1).value
            if m_val:
                label = str(m_val).strip()
                if self._label_to_month_key(label):
                    month_rows[label] = r
                else:
                    non_month_rows.append(r)

        month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
        for m_key in month_keys:
            m_label = self._month_key_to_label(m_key)
            if m_label not in month_rows:
                if non_month_rows:
                    row = min(non_month_rows)
                    style_row = row - 1 if row > 2 else 2
                    self._insert_rows_preserve_merges(ws, row)
                    if 1 <= style_row <= ws.max_row:
                        self._copy_row_style(ws, style_row, row)
                    month_rows = {
                        label: (idx + 1 if idx >= row else idx)
                        for label, idx in month_rows.items()
                    }
                    non_month_rows = [idx + 1 if idx >= row else idx for idx in non_month_rows]
                else:
                    row = ws.max_row + 1
                    style_row = row - 1 if row > 2 else None
                    if style_row and 1 <= style_row <= ws.max_row:
                        self._copy_row_style(ws, style_row, row)
                month_rows[m_label] = row
                ws.cell(row=row, column=1).value = m_label
            r = month_rows[m_label]
            ws.cell(row=r, column=1).value = m_label
            if '部门' in header_map:
                ws.cell(row=r, column=header_map['部门']).value = '合计'
            data = metrics_by_month.get(m_key, {})
            mapping = {
                '主营业务收入': 'revenue',
                '主营业务成本': 'cost',
                '销售费用': 'sales_expense',
                '管理费用': 'admin_expense',
                '财务费用': 'financial_expense',
                '营业利润': 'operating_profit',
                '净利润': 'net_profit',
                '应收账款余额': 'ar_balance',
                '存货期末余额': 'inventory_end',
                '存货期初余额': 'inventory_start',
                '主营业务成本成本率': 'cost_rate',
                '主营业务成本率': 'cost_rate',
                '销售费用率': 'sales_expense_rate',
                '管理费用率': 'admin_expense_rate',
                '财务费用率': 'financial_expense_rate',
                '营业利润率': 'operating_profit_rate',
                '存货周转天数': 'inventory_turnover_days',
            }
            for header, key in mapping.items():
                col = header_map.get(header)
                if not col:
                    continue
                ws.cell(row=r, column=col).value = data.get(key)

        if metrics_by_month_all is None:
            metrics_by_month_all = metrics_by_month or {}

        compare_metrics = [
            ("主营业务收入", "revenue"),
            ("主营业务成本", "cost"),
            ("销售费用", "sales_expense"),
            ("管理费用", "admin_expense"),
            ("财务费用", "financial_expense"),
            ("营业利润", "operating_profit"),
        ]
        compare_cols = {}
        for label, key in compare_metrics:
            compare_cols[(key, "yoy_delta")] = self._ensure_header_column(ws, f"{label}_同比增量")
            compare_cols[(key, "yoy_rate")] = self._ensure_header_column(ws, f"{label}_同比增速")
            compare_cols[(key, "mom_delta")] = self._ensure_header_column(ws, f"{label}_环比增量")
            compare_cols[(key, "mom_rate")] = self._ensure_header_column(ws, f"{label}_环比增速")

        self._apply_header_style(ws, 1, max_col=ws.max_column)

        def _month_shift(month_key, offset):
            try:
                p = pd.Period(month_key, freq='M') + offset
                return f"{p.year}-{p.month:02d}"
            except Exception:
                return None

        def _delta_rate(curr, prev):
            if curr is None or prev is None:
                return None, None
            delta = curr - prev
            rate = (delta / prev) if prev else None
            return delta, rate

        for m_label, r in month_rows.items():
            m_key = self._label_to_month_key(m_label)
            if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                for col in compare_cols.values():
                    self._safe_set_cell_value(ws, r, col, None)
                continue

            curr_data_scoped = metrics_by_month.get(m_key, {})
            prev_mom_key = _month_shift(m_key, -1)
            prev_yoy_key = _month_shift(m_key, -12)
            prev_mom_data = metrics_by_month.get(prev_mom_key, {}) if prev_mom_key else {}
            prev_yoy_data = metrics_by_month_all.get(prev_yoy_key, {}) if prev_yoy_key else {}

            for _, key in compare_metrics:
                curr = curr_data_scoped.get(key)
                yoy_delta, yoy_rate = _delta_rate(curr, prev_yoy_data.get(key))
                mom_delta, mom_rate = _delta_rate(curr, prev_mom_data.get(key))
                self._safe_set_cell_value(ws, r, compare_cols[(key, "yoy_delta")], yoy_delta)
                self._safe_set_cell_value(ws, r, compare_cols[(key, "yoy_rate")], yoy_rate)
                self._safe_set_cell_value(ws, r, compare_cols[(key, "mom_delta")], mom_delta)
                self._safe_set_cell_value(ws, r, compare_cols[(key, "mom_rate")], mom_rate)

        # 刷新“合计/全年合计”行，避免模板残留历史值造成口径不一致。
        summary_labels = {"合计", "全年合计", "全年汇总", "本年累计", "年累计"}
        summary_row = None
        for r in range(2, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if isinstance(label, str) and label.strip() in summary_labels:
                summary_row = r
                break
        if summary_row and month_keys:
            month_metrics = [metrics_by_month.get(m, {}) for m in month_keys]
            revenue_total = self._safe_sum([m.get("revenue") for m in month_metrics])
            cost_total = self._safe_sum([m.get("cost") for m in month_metrics])
            sales_total = self._safe_sum([m.get("sales_expense") for m in month_metrics])
            admin_total = self._safe_sum([m.get("admin_expense") for m in month_metrics])
            financial_total = self._safe_sum([m.get("financial_expense") for m in month_metrics])
            op_total = self._safe_sum([m.get("operating_profit") for m in month_metrics])
            net_total = self._safe_sum([m.get("net_profit") for m in month_metrics])

            first_month = month_keys[0]
            last_month = month_keys[-1]
            first_data = metrics_by_month.get(first_month, {})
            last_data = metrics_by_month.get(last_month, {})

            ar_last = last_data.get("ar_balance")
            inventory_end_last = last_data.get("inventory_end")
            inventory_start_first = first_data.get("inventory_start")
            cost_rate = (cost_total / revenue_total) if (revenue_total not in (None, 0) and cost_total is not None) else None
            sales_rate = (sales_total / revenue_total) if (revenue_total not in (None, 0) and sales_total is not None) else None
            admin_rate = (admin_total / revenue_total) if (revenue_total not in (None, 0) and admin_total is not None) else None
            financial_rate = (financial_total / revenue_total) if (revenue_total not in (None, 0) and financial_total is not None) else None
            op_rate = (op_total / revenue_total) if (revenue_total not in (None, 0) and op_total is not None) else None

            inv_days = None
            if (
                cost_total not in (None, 0)
                and inventory_start_first is not None
                and inventory_end_last is not None
            ):
                avg_inventory = (inventory_start_first + inventory_end_last) / 2
                if avg_inventory not in (None, 0):
                    inv_days = 365 * avg_inventory / cost_total

            write_map = {
                "主营业务收入": revenue_total,
                "主营业务成本": cost_total,
                "销售费用": sales_total,
                "管理费用": admin_total,
                "财务费用": financial_total,
                "营业利润": op_total,
                "净利润": net_total,
                "应收账款余额": ar_last,
                "存货期末余额": inventory_end_last,
                "存货期初余额": inventory_start_first,
                "主营业务成本成本率": cost_rate,
                "主营业务成本率": cost_rate,
                "销售费用率": sales_rate,
                "管理费用率": admin_rate,
                "财务费用率": financial_rate,
                "营业利润率": op_rate,
                "存货周转天数": inv_days,
            }
            for header, value in write_map.items():
                col = header_map.get(header)
                if col:
                    self._safe_set_cell_value(ws, summary_row, col, value)
            for col in compare_cols.values():
                self._safe_set_cell_value(ws, summary_row, col, None)

        if target_year and target_month:
            for m_label, r in month_rows.items():
                m_key = self._label_to_month_key(m_label)
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    for col in range(1, ws.max_column + 1):
                        self._safe_set_cell_value(ws, r, col, None)

        self._reorder_month_rows_desc(ws)

    def _update_expense_compare_sheet(self, ws, metrics_by_month, target_year, target_month, year_scope=None):
        header_map = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                header_map[str(header).strip()] = col
        for required_header in ("财务费用", "财务费用占比"):
            if required_header not in header_map:
                header_map[required_header] = self._ensure_header_column(ws, required_header)
        month_rows = {}
        non_month_rows = []
        for r in range(2, ws.max_row + 1):
            m_val = ws.cell(row=r, column=1).value
            if m_val:
                label = str(m_val).strip()
                if self._label_to_month_key(label):
                    month_rows[label] = r
                else:
                    non_month_rows.append(r)
        for m_key in self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope):
            m_label = self._month_key_to_label(m_key)
            if m_label not in month_rows:
                if non_month_rows:
                    row = min(non_month_rows)
                    style_row = row - 1 if row > 2 else 2
                    self._insert_rows_preserve_merges(ws, row)
                    if 1 <= style_row <= ws.max_row:
                        self._copy_row_style(ws, style_row, row)
                    month_rows = {
                        label: (idx + 1 if idx >= row else idx)
                        for label, idx in month_rows.items()
                    }
                    non_month_rows = [idx + 1 if idx >= row else idx for idx in non_month_rows]
                else:
                    row = ws.max_row + 1
                    style_row = row - 1 if row > 2 else None
                    if style_row and 1 <= style_row <= ws.max_row:
                        self._copy_row_style(ws, style_row, row)
                month_rows[m_label] = row
                ws.cell(row=row, column=1).value = m_label
            r = month_rows[m_label]
            ws.cell(row=r, column=1).value = m_label
            if '部门' in header_map:
                ws.cell(row=r, column=header_map['部门']).value = '合计'
            data = metrics_by_month.get(m_key, {})
            revenue = data.get('revenue')
            cost = data.get('cost')
            sales_expense = data.get('sales_expense')
            admin_expense = data.get('admin_expense')
            financial_expense = data.get('financial_expense')
            operating_profit = data.get('operating_profit')
            if '主营业务收入' in header_map:
                ws.cell(row=r, column=header_map['主营业务收入']).value = revenue
            if '主营业务成本' in header_map:
                ws.cell(row=r, column=header_map['主营业务成本']).value = cost
            if '销售费用' in header_map:
                ws.cell(row=r, column=header_map['销售费用']).value = sales_expense
            if '管理费用' in header_map:
                ws.cell(row=r, column=header_map['管理费用']).value = admin_expense
            if '财务费用' in header_map:
                ws.cell(row=r, column=header_map['财务费用']).value = financial_expense
            if '营业利润' in header_map:
                ws.cell(row=r, column=header_map['营业利润']).value = operating_profit
            if revenue is not None and revenue != 0:
                if '主营业务成本占比' in header_map and cost is not None:
                    ws.cell(row=r, column=header_map['主营业务成本占比']).value = cost / revenue
                if '销售费用占比' in header_map and sales_expense is not None:
                    ws.cell(row=r, column=header_map['销售费用占比']).value = sales_expense / revenue
                if '管理费用占比' in header_map and admin_expense is not None:
                    ws.cell(row=r, column=header_map['管理费用占比']).value = admin_expense / revenue
                if '财务费用占比' in header_map and financial_expense is not None:
                    ws.cell(row=r, column=header_map['财务费用占比']).value = financial_expense / revenue

        if target_year and target_month:
            for m_label, r in month_rows.items():
                m_key = self._label_to_month_key(m_label)
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    for col in header_map.values():
                        self._safe_set_cell_value(ws, r, col, None)

        # 刷新“合计/全年合计”行，确保与月度明细口径一致。
        summary_labels = {"合计", "全年合计", "全年汇总", "本年累计", "年累计"}
        summary_row = None
        for r in range(2, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if isinstance(label, str) and label.strip() in summary_labels:
                summary_row = r
                break
        if summary_row:
            month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
            month_metrics = [metrics_by_month.get(m, {}) for m in month_keys]
            revenue_total = self._safe_sum([m.get("revenue") for m in month_metrics])
            cost_total = self._safe_sum([m.get("cost") for m in month_metrics])
            sales_total = self._safe_sum([m.get("sales_expense") for m in month_metrics])
            admin_total = self._safe_sum([m.get("admin_expense") for m in month_metrics])
            financial_total = self._safe_sum([m.get("financial_expense") for m in month_metrics])
            op_total = self._safe_sum([m.get("operating_profit") for m in month_metrics])

            write_map = {
                "主营业务收入": revenue_total,
                "主营业务成本": cost_total,
                "销售费用": sales_total,
                "管理费用": admin_total,
                "财务费用": financial_total,
                "营业利润": op_total,
                "主营业务成本占比": (cost_total / revenue_total) if (revenue_total not in (None, 0) and cost_total is not None) else None,
                "销售费用占比": (sales_total / revenue_total) if (revenue_total not in (None, 0) and sales_total is not None) else None,
                "管理费用占比": (admin_total / revenue_total) if (revenue_total not in (None, 0) and admin_total is not None) else None,
                "财务费用占比": (financial_total / revenue_total) if (revenue_total not in (None, 0) and financial_total is not None) else None,
            }
            if '部门' in header_map:
                self._safe_set_cell_value(ws, summary_row, header_map['部门'], '合计')
            for header, value in write_map.items():
                col = header_map.get(header)
                if col:
                    self._safe_set_cell_value(ws, summary_row, col, value)

    def _update_compare_sheet(self, ws, metrics_by_month, compare_type, target_year=None, target_month=None, year_scope=None):
        # compare_type: 'yoy' or 'mom'
        if ws.max_row < 2:
            return
        header_map = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                header_map[str(header).strip()] = col

        key_by_indicator = {
            '主营业务收入': 'revenue',
            '主营业务成本': 'cost',
            '销售费用': 'sales_expense',
            '管理费用': 'admin_expense',
            '财务费用': 'financial_expense',
            '营业利润': 'operating_profit',
        }

        month_col = header_map.get("月份")
        indicator_col = header_map.get("指标")
        dept_col = header_map.get("部门")
        if month_col and indicator_col:
            scoped_months = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
            existing_pairs = set()
            for r in range(2, ws.max_row + 1):
                m_key = self._label_to_month_key(ws.cell(row=r, column=month_col).value)
                indicator = ws.cell(row=r, column=indicator_col).value
                if not m_key or not indicator:
                    continue
                existing_pairs.add((m_key, self._normalize_label(indicator)))
            for label in ("财务费用",):
                norm_label = self._normalize_label(label)
                for m_key in scoped_months:
                    if (m_key, norm_label) in existing_pairs:
                        continue
                    row = ws.max_row + 1
                    src_row = row - 1 if row > 2 else 2
                    if 1 <= src_row <= ws.max_row:
                        self._copy_row_style(ws, src_row, row)
                    for c in range(1, ws.max_column + 1):
                        self._safe_set_cell_value(ws, row, c, None)
                    ws.cell(row=row, column=month_col).value = self._month_key_to_label(m_key)
                    ws.cell(row=row, column=indicator_col).value = label
                    if dept_col:
                        ws.cell(row=row, column=dept_col).value = "合计"
                    existing_pairs.add((m_key, norm_label))

        for r in range(2, ws.max_row + 1):
            m_val = ws.cell(row=r, column=1).value
            indicator = ws.cell(row=r, column=3).value
            if not m_val or not indicator:
                continue
            m_label = str(m_val).strip()
            m_key = self._label_to_month_key(m_label)
            if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                for key in ['本期值', '同比增量', '同比增速', '环比增量', '环比增速']:
                    col = header_map.get(key)
                    if col:
                        self._safe_set_cell_value(ws, r, col, None)
                continue
            ind_norm = self._normalize_label(indicator)
            metric_key = key_by_indicator.get(ind_norm)
            if not metric_key or m_key not in metrics_by_month:
                for key in ['本期值', '同比增量', '同比增速', '环比增量', '环比增速']:
                    col = header_map.get(key)
                    if col:
                        self._safe_set_cell_value(ws, r, col, None)
                continue
            current_val = metrics_by_month[m_key].get(metric_key)
            if '本期值' in header_map:
                ws.cell(row=r, column=header_map['本期值']).value = current_val
            # compare
            prev_val = None
            if compare_type == 'mom':
                try:
                    year, month = m_key.split('-')
                    month = int(month)
                    if month > 1:
                        prev_key = f"{year}-{month-1:02d}"
                        prev_val = metrics_by_month.get(prev_key, {}).get(metric_key)
                except Exception:
                    prev_val = None
            elif compare_type == 'yoy':
                try:
                    year, month = m_key.split('-')
                    prev_key = f"{int(year) - 1}-{month}"
                    prev_val = metrics_by_month.get(prev_key, {}).get(metric_key)
                except Exception:
                    prev_val = None
            delta = None
            rate = None
            if current_val is not None and prev_val is not None:
                delta = current_val - prev_val
                if prev_val:
                    rate = delta / prev_val
            if '同比增量' in header_map:
                ws.cell(row=r, column=header_map['同比增量']).value = delta
            if '同比增速' in header_map:
                ws.cell(row=r, column=header_map['同比增速']).value = rate
            if '环比增量' in header_map:
                ws.cell(row=r, column=header_map['环比增量']).value = delta
            if '环比增速' in header_map:
                ws.cell(row=r, column=header_map['环比增速']).value = rate

        indicator_col = header_map.get("指标")
        self._reorder_month_rows_desc(ws, group_by_cols=[indicator_col] if indicator_col else None)

    def _update_yoy_multi_year_sheet(self, wb, metrics_by_month, target_year, target_month, year_scope=None):
        if not target_year or not target_month:
            return
        month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
        years = sorted({int(m.split('-')[0]) for m in month_keys})
        if len(years) < 2:
            return

        sheet_name = "同比经营分析_多年度"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row:
                ws.delete_rows(1, ws.max_row)
            ws._charts = []
        else:
            insert_idx = wb.sheetnames.index("同比经营分析") + 1 if "同比经营分析" in wb.sheetnames else None
            ws = wb.create_sheet(sheet_name, insert_idx)

        indicator_map = [
            ("主营业务收入", "revenue"),
            ("主营业务成本", "cost"),
            ("销售费用", "sales_expense"),
            ("管理费用", "admin_expense"),
            ("财务费用", "financial_expense"),
            ("营业利润", "operating_profit"),
        ]

        month_end = int(target_month)
        row = 1
        for label, metric_key in indicator_map:
            ws.cell(row=row, column=1).value = label
            header_row = row + 1
            ws.cell(row=header_row, column=1).value = "月份"
            for idx, year in enumerate(years):
                ws.cell(row=header_row, column=2 + idx).value = str(year)
            self._apply_header_style(ws, header_row)

            data_start = header_row + 1
            for m in range(1, month_end + 1):
                ws.cell(row=data_start + m - 1, column=1).value = f"{m:02d}"
                for idx, year in enumerate(years):
                    m_key = f"{year}-{m:02d}"
                    val = metrics_by_month.get(m_key, {}).get(metric_key)
                    ws.cell(row=data_start + m - 1, column=2 + idx).value = val

            data_ref = Reference(
                ws,
                min_col=2,
                max_col=1 + len(years),
                min_row=header_row,
                max_row=data_start + month_end - 1,
            )
            cats_ref = Reference(
                ws,
                min_col=1,
                min_row=data_start,
                max_row=data_start + month_end - 1,
            )
            chart = LineChart()
            chart.title = f"{label}历年对比"
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height = 8
            chart.width = 16
            anchor = f"{get_column_letter(len(years) + 4)}{row}"
            ws.add_chart(chart, anchor)

            row = data_start + month_end + 2

    def _ensure_template_charts(self, wb, template_path):
        try:
            template_wb = openpyxl.load_workbook(template_path)
        except Exception as e:
            print(f"模板图表检查失败: {e}")
            return

        for name in template_wb.sheetnames:
            tmpl_ws = template_wb[name]
            if not getattr(tmpl_ws, "_charts", None):
                continue
            if name not in wb.sheetnames:
                continue
            out_ws = wb[name]
            tmpl_count = len(tmpl_ws._charts)
            out_count = len(out_ws._charts)
            if out_count >= tmpl_count:
                continue
            out_ws._charts = []
            for ch in tmpl_ws._charts:
                out_ws.add_chart(copy.deepcopy(ch))
            print(f"已修复图表: {name} ({out_count} -> {tmpl_count})")

    def _audit_chart_counts(self, wb):
        for ws in wb.worksheets:
            count = len(ws._charts)
            print(f"图表检查: {ws.title} = {count}")

    def _chart_title_text(self, chart):
        try:
            return chart.title.tx.rich.p[0].r[0].t
        except Exception:
            return ""

    def _worksheet_has_chart_title(self, ws, title_part):
        for chart in getattr(ws, "_charts", []):
            if title_part in self._chart_title_text(chart):
                return True
        return False

    def _chart_first_value_formula(self, chart):
        for series in getattr(chart, "series", []) or []:
            if series.val is not None and series.val.numRef is not None and series.val.numRef.f:
                return series.val.numRef.f
        return None

    def _infer_chart_category_reference(self, wb, chart):
        value_formula = self._chart_first_value_formula(chart)
        parsed = self._parse_chart_range_formula(value_formula)
        if not parsed:
            return None
        sheet_name, min_col, min_row, _max_col, max_row = parsed
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        header_row = max(1, min_row - 1)
        header_map = self._get_header_map(ws, header_row)

        preferred_headers = [
            "月份",
            "客户",
            "往来单位名",
            "产品",
            "产品名称",
            "品类",
            "供应商/对象",
            "科目/对象",
            "主体",
            "职员",
        ]
        for header in preferred_headers:
            col = header_map.get(header)
            if not col or col == min_col:
                continue
            values = [ws.cell(row=r, column=col).value for r in range(min_row, max_row + 1)]
            if any(v is not None and str(v).strip() != "" for v in values):
                return Reference(ws, min_col=col, min_row=min_row, max_row=max_row)

        # Most helper chart tables use the column immediately before the first series as labels.
        if min_col > 1:
            prev_values = [ws.cell(row=r, column=min_col - 1).value for r in range(min_row, max_row + 1)]
            if any(v is not None and str(v).strip() != "" for v in prev_values):
                return Reference(ws, min_col=min_col - 1, min_row=min_row, max_row=max_row)

        first_values = [ws.cell(row=r, column=1).value for r in range(min_row, max_row + 1)]
        if min_col != 1 and any(v is not None and str(v).strip() != "" for v in first_values):
            return Reference(ws, min_col=1, min_row=min_row, max_row=max_row)
        return None

    def _repair_missing_chart_categories(self, wb):
        repaired = 0
        for ws in wb.worksheets:
            for chart in getattr(ws, "_charts", []) or []:
                if isinstance(chart, ScatterChart):
                    continue
                if self._chart_category_formula(chart):
                    continue
                ref = self._infer_chart_category_reference(wb, chart)
                if ref is None:
                    continue
                chart.set_categories(ref)
                for sub_chart in getattr(chart, "_charts", []) or []:
                    if not isinstance(sub_chart, ScatterChart):
                        sub_chart.set_categories(ref)
                repaired += 1
        if repaired:
            print(f"已修复图表分类轴: {repaired}")
        return repaired

    def _get_header_map(self, ws, header_row=1):
        header_map = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if header is None:
                continue
            header_map[str(header).strip()] = col
        return header_map

    def _find_header_row_by_keyword(self, ws, keyword, max_row=30):
        max_scan = min(ws.max_row, max_row)
        for r in range(1, max_scan + 1):
            for c in range(1, ws.max_column + 1):
                value = ws.cell(row=r, column=c).value
                if isinstance(value, str) and value.strip() == keyword:
                    return r
        return None

    def _find_last_row_by_column(self, ws, col, start_row):
        last = start_row - 1
        for r in range(start_row, ws.max_row + 1):
            value = ws.cell(row=r, column=col).value
            if value is not None and str(value).strip() != "":
                last = r
        return last

    def _find_month_data_bounds(self, ws, col, start_row):
        first = None
        last = start_row - 1
        for r in range(start_row, ws.max_row + 1):
            value = ws.cell(row=r, column=col).value
            if self._month_label_exact_to_key(value):
                if first is None:
                    first = r
                last = r
                continue
            if first is not None:
                break
        return (first or start_row), last

    def _copy_cell_style(self, src, dst):
        if not src or not dst:
            return
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.border = copy.copy(src.border)
            dst.fill = copy.copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy.copy(src.protection)
            dst.alignment = copy.copy(src.alignment)

    def _clone_hyperlink(self, link, dest_cell):
        if not link or not dest_cell:
            return None
        if isinstance(link, str):
            if link.startswith("#"):
                return Hyperlink(ref=dest_cell.coordinate, location=link[1:])
            return Hyperlink(ref=dest_cell.coordinate, target=link)
        new_link = copy.copy(link)
        new_link.ref = dest_cell.coordinate
        if new_link.target and str(new_link.target).startswith("#") and not new_link.location:
            new_link.location = str(new_link.target)[1:]
            new_link.target = None
            new_link.id = None
        return new_link

    def _apply_hyperlink(self, dest_cell, link):
        dest_cell.hyperlink = self._clone_hyperlink(link, dest_cell)

    def _normalize_internal_hyperlinks(self, wb):
        """
        统一将 '#Sheet!A1' 形式的目标链接转换为内部 location，
        避免 Excel 将其识别为“外部文件链接”导致无法打开。
        """
        if wb is None:
            return
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    link = cell.hyperlink
                    if not link:
                        continue
                    if isinstance(link, str):
                        if link.startswith("#"):
                            cell.hyperlink = Hyperlink(ref=cell.coordinate, location=link[1:])
                        continue
                    target = getattr(link, "target", None)
                    location = getattr(link, "location", None)
                    if target and str(target).startswith("#") and not location:
                        new_link = copy.copy(link)
                        new_link.ref = cell.coordinate
                        new_link.location = str(target)[1:]
                        new_link.target = None
                        new_link.id = None
                        cell.hyperlink = new_link

    def _last_n_month_keys(self, target_year, target_month, n=3):
        if not target_year or not target_month:
            return []
        keys = []
        dt = datetime(int(target_year), int(target_month), 1)
        for i in range(max(n, 1)):
            m = dt - pd.DateOffset(months=i)
            keys.append(f"{m.year}-{m.month:02d}")
        return sorted(keys)

    def _month_days(self, month_key):
        try:
            return pd.Period(month_key, freq='M').days_in_month
        except Exception:
            return 30

    def _pick_expense_amount_cols(self, df):
        if df is None:
            return None, None, None

        def pick_amount_col(keyword):
            if keyword in df.columns:
                return keyword
            for c in df.columns:
                if keyword in str(c) and '外币' not in str(c):
                    return c
            for c in df.columns:
                if keyword in str(c):
                    return c
            return None

        debit_col = pick_amount_col('借方金额')
        credit_col = pick_amount_col('贷方金额')
        amount_col = pick_amount_col('金额')
        return debit_col, credit_col, amount_col

    def _derive_expense_category(self, name, code=None):
        category_map = {'6601': '销售费用', '6602': '管理费用', '6603': '财务费用'}
        name_text = str(name).strip() if name is not None else ""
        code_text = str(code).strip() if code is not None else ""

        category = None
        subcategory = name_text if name_text else None

        if code_text:
            prefix = code_text[:4]
            category = category_map.get(prefix)

        if not category and name_text:
            for k in ["销售费用", "管理费用", "财务费用", "研发费用"]:
                if name_text.startswith(k):
                    category = k
                    rest = name_text.replace(k, "", 1).strip(" -_/—")
                    if rest:
                        subcategory = rest
                    break

        if not category and name_text:
            for sep in ["-", "—", "/", "\\", "_"]:
                if sep in name_text:
                    left, right = name_text.split(sep, 1)
                    left = left.strip()
                    right = right.strip()
                    if left and right:
                        category = left
                        subcategory = right
                        break

        if not category:
            category = name_text if name_text else None

        return category, subcategory

    def _prepare_expense_analysis_df(self, df, target_year=None, target_month=None, year_scope=None):
        if df is None or df.empty:
            return None

        scoped = df.copy()
        if not scoped.columns.is_unique:
            scoped = scoped.loc[:, ~scoped.columns.duplicated()].copy()

        if 'MonthStr' not in scoped.columns:
            date_col = next((c for c in scoped.columns if '日期' in str(c) or 'Date' in str(c)), None)
            if date_col:
                scoped['ParsedDate'] = pd.to_datetime(scoped[date_col], errors='coerce')
                scoped['MonthStr'] = scoped['ParsedDate'].dt.strftime('%Y-%m')
        if 'MonthStr' not in scoped.columns:
            return None

        scoped = scoped[scoped['MonthStr'].notna()]
        scoped = self._filter_df_by_scope(scoped, target_year, target_month, year_scope)
        if scoped is None or scoped.empty:
            return None

        if '科目名' in scoped.columns:
            scoped = scoped[scoped['科目名'].notna()]
            scoped = scoped[scoped['科目名'] != '科目名']

        debit_col, credit_col, amount_col = self._pick_expense_amount_cols(scoped)
        if debit_col:
            scoped['DebitAmount'] = scoped[debit_col].apply(self._to_float)
        else:
            scoped['DebitAmount'] = None
        if credit_col:
            scoped['CreditAmount'] = scoped[credit_col].apply(self._to_float)
        else:
            scoped['CreditAmount'] = None
        if debit_col and credit_col:
            scoped['Amount'] = scoped[debit_col].apply(self._to_float).fillna(0) - scoped[credit_col].apply(self._to_float).fillna(0)
        elif debit_col:
            scoped['Amount'] = scoped[debit_col].apply(self._to_float)
        elif amount_col:
            scoped['Amount'] = scoped[amount_col].apply(self._to_float)
        else:
            return None

        scoped = scoped[scoped['Amount'].notna()]
        scoped = scoped[scoped['Amount'] != 0]
        if scoped.empty:
            return None

        code_col = next((c for c in scoped.columns if '科目编码' in str(c)), None)
        name_col = next((c for c in scoped.columns if '科目名' in str(c)), None)

        if name_col:
            def derive(row):
                return self._derive_expense_category(row.get(name_col), row.get(code_col) if code_col else None)

            derived = scoped.apply(lambda r: pd.Series(derive(r), index=['Category', 'Subcategory']), axis=1)
            scoped['Category'] = derived['Category']
            scoped['Subcategory'] = derived['Subcategory']
        else:
            scoped['Category'] = None
            scoped['Subcategory'] = None

        dept_col = next((c for c in scoped.columns if str(c) in ['部门名', '部门']), None)
        summary_col = next((c for c in scoped.columns if '摘要' in str(c)), None)
        scoped['Department'] = scoped[dept_col] if dept_col else '合计'
        scoped['Summary'] = scoped[summary_col] if summary_col else None
        scoped['AmountAbs'] = scoped['Amount'].apply(lambda x: abs(x) if x is not None else 0)
        scoped['MonthLabel'] = scoped['MonthStr'].apply(self._month_key_to_label)
        return scoped

    def _safe_set_cell_value(self, ws, row, col, value):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            return False
        cell.value = self._sanitize_excel_text(value)
        return True

    def _sanitize_excel_text(self, value):
        """
        规避 Excel 公式注入:
        当文本以公式起始符开头时，转义为纯文本。
        """
        if not isinstance(value, str):
            return value
        stripped = value.lstrip()
        if not stripped:
            return value
        starts_as_formula = stripped[0] in ("=", "+", "@")
        starts_as_dash_formula = stripped.startswith("-=") or stripped.startswith("-+") or stripped.startswith("-@")
        if starts_as_formula or starts_as_dash_formula:
            return "'" + value
        return value

    def _copy_row_style(self, ws, src_row, dst_row, max_col=None):
        if max_col is None:
            max_col = ws.max_column
        for col in range(1, max_col + 1):
            self._copy_cell_style(ws.cell(row=src_row, column=col), ws.cell(row=dst_row, column=col))
        if src_row in ws.row_dimensions:
            ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    def _move_row_preserve(self, ws, src_row, dest_row):
        if not src_row or not dest_row or src_row == dest_row:
            return False
        max_col = ws.max_column
        snapshot = self._snapshot_row(ws, src_row, max_col)
        ws.delete_rows(src_row, 1)
        if src_row < dest_row:
            dest_row -= 1
        self._insert_rows_preserve_merges(ws, dest_row, 1)
        self._restore_row(ws, dest_row, snapshot, max_col)
        return True

    def _insert_rows_preserve_merges(self, ws, insert_row, amount=1):
        if amount <= 0:
            return

        to_adjust = []
        for merged in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged.bounds
            if max_row < insert_row:
                continue
            to_adjust.append((str(merged), min_col, min_row, max_col, max_row))

        for merged_ref, _, _, _, _ in to_adjust:
            ws.unmerge_cells(merged_ref)

        ws.insert_rows(insert_row, amount)

        for _, min_col, min_row, max_col, max_row in to_adjust:
            if min_row >= insert_row:
                min_row += amount
                max_row += amount
            elif min_row < insert_row <= max_row:
                max_row += amount
            merged_ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{max_row}"
            )
            ws.merge_cells(merged_ref)

    def _copy_column_style(self, ws, src_col, dst_col, max_row=None):
        if max_row is None:
            max_row = ws.max_row
        for row in range(1, max_row + 1):
            self._copy_cell_style(ws.cell(row=row, column=src_col), ws.cell(row=row, column=dst_col))
        src_letter = get_column_letter(src_col)
        dst_letter = get_column_letter(dst_col)
        if src_letter in ws.column_dimensions:
            ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width

    def _ensure_metric_columns(self, ws, header_row, required_headers):
        header_map = self._get_header_map(ws, header_row)
        last_col = 0
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val is not None and str(val).strip() != "":
                last_col = col
        if last_col == 0:
            last_col = 1
        missing = [h for h in required_headers if h not in header_map]
        if not missing:
            return header_map
        src_col = last_col
        for header in missing:
            last_col += 1
            ws.cell(row=header_row, column=last_col).value = header
            self._copy_column_style(ws, src_col, last_col, max_row=header_row)
            header_map[header] = last_col
        return header_map

    def _normalize_month_label(self, label):
        if not label:
            return None
        m = re.search(r'(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?', str(label))
        if not m:
            return None
        return f"{m.group(1)}/{int(m.group(2)):02d}"

    def _iter_missing_month_labels(self, existing_labels, target_label):
        existing_keys = {
            lbl.replace('/', '-')
            for lbl in (existing_labels or [])
            if isinstance(lbl, str) and self._normalize_month_label(lbl)
        }
        normalized_target = self._normalize_month_label(target_label)
        if not normalized_target:
            return []
        target_key = normalized_target.replace('/', '-')
        if not existing_keys:
            return [normalized_target]

        start_key = min(existing_keys)
        if start_key > target_key:
            start_key = target_key

        missing = []
        current = datetime.strptime(start_key + "-01", "%Y-%m-%d")
        target_dt = datetime.strptime(target_key + "-01", "%Y-%m-%d")
        while current <= target_dt:
            key = current.strftime("%Y-%m")
            label = key.replace('-', '/')
            if key not in existing_keys:
                missing.append(label)
            year = current.year + (1 if current.month == 12 else 0)
            month = 1 if current.month == 12 else current.month + 1
            current = current.replace(year=year, month=month)
        return missing

    def _ensure_month_columns_simple(self, ws, target_year, target_month, header_row=1):
        target_label = f"{target_year}/{int(target_month):02d}"
        changed = False
        for label_to_insert in self._iter_missing_month_labels(
            [
                self._normalize_month_label(ws.cell(row=header_row, column=col).value)
                for col in range(1, ws.max_column + 1)
            ],
            target_label,
        ):
            month_cols = []
            for col in range(1, ws.max_column + 1):
                label = self._normalize_month_label(ws.cell(row=header_row, column=col).value)
                if label:
                    month_cols.append((col, label))
            if not month_cols:
                return changed

            keys = [lbl.replace('/', '-') for _, lbl in month_cols]
            is_asc = keys == sorted(keys)
            target_key = label_to_insert.replace('/', '-')
            insert_col = None
            if is_asc:
                for col, lbl in month_cols:
                    if lbl.replace('/', '-') > target_key:
                        insert_col = col
                        break
                if insert_col is None:
                    insert_col = month_cols[-1][0] + 1
                src_col = month_cols[0][0]
                for col, _ in month_cols:
                    if col < insert_col:
                        src_col = col
            else:
                for col, lbl in month_cols:
                    if lbl.replace('/', '-') < target_key:
                        insert_col = col
                        break
                if insert_col is None:
                    insert_col = month_cols[-1][0] + 1
                src_col = month_cols[0][0]
                for col, _ in month_cols:
                    if col < insert_col:
                        src_col = col

            ws.insert_cols(insert_col)
            if src_col >= insert_col:
                src_col += 1
            self._copy_column_style(ws, src_col, insert_col)
            ws.cell(row=header_row, column=insert_col).value = label_to_insert
            changed = True
        return changed

    def _ensure_month_rows_simple(self, ws, target_year, target_month, header_keyword="月份", total_label="合计"):
        header_row = self._find_header_row_by_keyword(ws, header_keyword, max_row=50)
        if not header_row:
            return False
        target_label = f"{target_year}/{int(target_month):02d}"
        changed = False
        for label_to_insert in self._iter_missing_month_labels(
            [
                self._normalize_month_label(ws.cell(row=r, column=1).value)
                for r in range(header_row + 1, ws.max_row + 1)
            ],
            target_label,
        ):
            month_rows = []
            for r in range(header_row + 1, ws.max_row + 1):
                label = self._normalize_month_label(ws.cell(row=r, column=1).value)
                if label:
                    month_rows.append((r, label))

            insert_at = None
            src_row = None
            for r in range(header_row + 1, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if isinstance(v, str) and v.strip() == total_label:
                    insert_at = r
                    src_row = r - 1 if r > header_row + 1 else header_row + 1
                    break

            if insert_at is None:
                if month_rows:
                    src_row = month_rows[-1][0]
                    insert_at = src_row + 1
                else:
                    src_row = header_row
                    insert_at = header_row + 1

            ws.insert_rows(insert_at)
            self._copy_row_style(ws, src_row, insert_at)
            for c in range(1, ws.max_column + 1):
                self._safe_set_cell_value(ws, insert_at, c, None)
            ws.cell(row=insert_at, column=1).value = label_to_insert

            header_map = self._get_header_map(ws, header_row)
            dept_col = header_map.get("部门")
            if dept_col:
                dept_val = ws.cell(row=src_row, column=dept_col).value or "合计"
                ws.cell(row=insert_at, column=dept_col).value = dept_val
            changed = True
        return changed

    def _ensure_compare_month_row(self, ws, target_year, target_month):
        header_row = self._find_header_row_by_keyword(ws, "月份", max_row=10) or 1
        header_map = self._get_header_map(ws, header_row)
        month_col = header_map.get("月份")
        indicator_col = header_map.get("指标")
        if not month_col or not indicator_col:
            return False
        target_label = f"{target_year}/{int(target_month):02d}"

        groups = {}
        order = []
        for r in range(header_row + 1, ws.max_row + 1):
            indicator = ws.cell(row=r, column=indicator_col).value
            if indicator is None:
                continue
            ind = str(indicator).strip()
            if ind not in groups:
                groups[ind] = []
                order.append(ind)
            groups[ind].append(r)

        if not order:
            return False

        dept_col = header_map.get("部门")
        changed = False
        for ind in reversed(order):
            rows = groups[ind]
            months = [self._normalize_month_label(ws.cell(row=r, column=month_col).value) for r in rows]
            if target_label in months:
                continue
            insert_at = max(rows) + 1
            src_row = max(rows)
            ws.insert_rows(insert_at)
            self._copy_row_style(ws, src_row, insert_at)
            for c in range(1, ws.max_column + 1):
                self._safe_set_cell_value(ws, insert_at, c, None)
            ws.cell(row=insert_at, column=month_col).value = target_label
            ws.cell(row=insert_at, column=indicator_col).value = ind
            if dept_col:
                dept_val = ws.cell(row=src_row, column=dept_col).value or "合计"
                ws.cell(row=insert_at, column=dept_col).value = dept_val
            changed = True

        return changed

    def _ensure_month_columns_grouped_by_suffix(self, ws, target_year, target_month):
        target_label = f"{target_year}/{int(target_month):02d}"
        changed = False
        for label_to_insert in self._iter_missing_month_labels(
            [
                self._normalize_month_label(ws.cell(row=1, column=col).value)
                for col in range(1, ws.max_column + 1)
            ],
            target_label,
        ):
            groups = {}
            order = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if not header:
                    continue
                m = re.match(r'(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?_(.+)', str(header))
                if not m:
                    continue
                label = f"{m.group(1)}/{int(m.group(2)):02d}"
                suffix = m.group(3)
                if suffix not in groups:
                    groups[suffix] = []
                    order.append(suffix)
                groups[suffix].append((col, label))

            if not order:
                return changed

            for suffix in sorted(order, key=lambda s: max(c for c, _ in groups[s]), reverse=True):
                cols = groups[suffix]
                labels = [lbl for _, lbl in cols]
                if label_to_insert in labels:
                    continue
                insert_col = max(c for c, _ in cols) + 1
                src_col = max(c for c, _ in cols)
                ws.insert_cols(insert_col)
                self._copy_column_style(ws, src_col, insert_col)
                ws.cell(row=1, column=insert_col).value = f"{label_to_insert}_{suffix}"
                changed = True
        return changed

    def _ensure_month_columns_grouped_by_month(self, ws, target_year, target_month):
        target_label = f"{target_year}/{int(target_month):02d}"
        changed = False
        for label_to_insert in self._iter_missing_month_labels(
            [
                self._normalize_month_label(ws.cell(row=1, column=col).value)
                for col in range(1, ws.max_column + 1)
            ],
            target_label,
        ):
            month_groups = {}
            order = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if not header:
                    continue
                m = re.match(r'(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?_(.+)', str(header))
                if not m:
                    continue
                label = f"{m.group(1)}/{int(m.group(2)):02d}"
                if label not in month_groups:
                    month_groups[label] = []
                    order.append(label)
                month_groups[label].append(col)

            if not order or label_to_insert in month_groups:
                continue

            source_label = order[-1]
            source_cols = sorted(month_groups[source_label])
            insert_col = max(source_cols) + 1
            for offset, src_col in enumerate(source_cols):
                dest_col = insert_col + offset
                ws.insert_cols(dest_col)
                self._copy_column_style(ws, src_col, dest_col)
                src_header = ws.cell(row=1, column=src_col).value
                suffix = str(src_header).split('_', 1)[1] if '_' in str(src_header) else ""
                ws.cell(row=1, column=dest_col).value = f"{label_to_insert}_{suffix}" if suffix else label_to_insert
                changed = True
        return changed

    def _reorder_columns_by_order(self, ws, ordered_cols):
        if not ordered_cols:
            return False
        start_col = min(ordered_cols)
        end_col = max(ordered_cols)
        if (end_col - start_col + 1) != len(ordered_cols):
            return False

        insert_at = end_col + 1
        ws.insert_cols(insert_at, amount=len(ordered_cols))
        max_row = ws.max_row
        for idx, src_col in enumerate(ordered_cols):
            dest_col = insert_at + idx
            self._copy_column_style(ws, src_col, dest_col, max_row=max_row)
            for r in range(1, max_row + 1):
                src_cell = ws.cell(row=r, column=src_col)
                dest_cell = ws.cell(row=r, column=dest_col)
                dest_cell.value = src_cell.value
                if src_cell.comment:
                    dest_cell.comment = copy.copy(src_cell.comment)
                if src_cell.hyperlink:
                    self._apply_hyperlink(dest_cell, src_cell.hyperlink)

        ws.delete_cols(start_col, amount=len(ordered_cols))
        return True

    def _reorder_month_columns_desc(self, ws, header_row=1):
        month_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if not header:
                continue
            m = re.match(r'(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?_(.+)', str(header))
            if not m:
                continue
            label = f"{m.group(1)}/{int(m.group(2)):02d}"
            month_cols.append((col, label))

        if not month_cols:
            return False

        order = []
        groups = {}
        for col, label in month_cols:
            if label not in groups:
                groups[label] = []
                order.append(label)
            groups[label].append(col)

        start_col = min(col for col, _ in month_cols)
        end_col = max(col for col, _ in month_cols)
        if (end_col - start_col + 1) != len(month_cols):
            # Non-contiguous month columns, skip to avoid unexpected shifts.
            return False

        order_desc = sorted(order, key=lambda x: x.replace('/', '-'), reverse=True)
        if order == order_desc:
            return False

        desired_cols = []
        for label in order_desc:
            desired_cols.extend(groups[label])

        return self._reorder_columns_by_order(ws, desired_cols)

    def _reorder_month_columns_simple_desc(self, ws, header_row=1):
        month_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if not header:
                continue
            m = re.match(r'^(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?$', str(header))
            if not m:
                continue
            label = f"{m.group(1)}/{int(m.group(2)):02d}"
            month_cols.append((col, label))

        if not month_cols:
            return False

        start_col = min(col for col, _ in month_cols)
        end_col = max(col for col, _ in month_cols)
        if (end_col - start_col + 1) != len(month_cols):
            return False

        labels = [lbl for _, lbl in month_cols]
        order_desc = sorted(labels, key=lambda x: x.replace('/', '-'), reverse=True)
        if labels == order_desc:
            return False

        desired_cols = [
            col for col, _ in sorted(month_cols, key=lambda x: x[1].replace('/', '-'), reverse=True)
        ]
        return self._reorder_columns_by_order(ws, desired_cols)

    def _reorder_month_columns_grouped_by_suffix_desc(self, ws, header_row=1):
        groups = {}
        order = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if not header:
                continue
            m = re.match(r'(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?_(.+)', str(header))
            if not m:
                continue
            label = f"{m.group(1)}/{int(m.group(2)):02d}"
            suffix = m.group(3)
            if suffix not in groups:
                groups[suffix] = []
                order.append(suffix)
            groups[suffix].append((col, label))

        if not order:
            return False

        changed = False
        for suffix in order:
            cols = groups[suffix]
            labels = [lbl for _, lbl in cols]
            order_desc = sorted(labels, key=lambda x: x.replace('/', '-'), reverse=True)
            if labels == order_desc:
                continue
            start_col = min(col for col, _ in cols)
            end_col = max(col for col, _ in cols)
            if (end_col - start_col + 1) != len(cols):
                continue
            desired_cols = [
                col for col, _ in sorted(cols, key=lambda x: x[1].replace('/', '-'), reverse=True)
            ]
            if self._reorder_columns_by_order(ws, desired_cols):
                changed = True
        return changed

    def _month_label_exact_to_key(self, label):
        if label is None:
            return None
        if isinstance(label, (datetime, pd.Timestamp)):
            return label.strftime('%Y-%m')
        s = str(label).strip()
        if not s:
            return None
        m = re.match(r'^(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?$', s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
        m = re.match(r'^(20\d{2})\s*年\s*(\d{1,2})\s*月$', s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
        return None

    def _snapshot_row(self, ws, row_idx, max_col):
        cells = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=c)
            cells.append({
                "value": cell.value,
                "font": copy.copy(cell.font),
                "border": copy.copy(cell.border),
                "fill": copy.copy(cell.fill),
                "number_format": cell.number_format,
                "protection": copy.copy(cell.protection),
                "alignment": copy.copy(cell.alignment),
                "comment": copy.copy(cell.comment) if cell.comment else None,
                "hyperlink": cell.hyperlink,
            })
        height = ws.row_dimensions[row_idx].height
        return {"height": height, "cells": cells}

    def _restore_row(self, ws, row_idx, snapshot, max_col):
        if snapshot.get("height") is not None:
            ws.row_dimensions[row_idx].height = snapshot.get("height")
        for c in range(1, max_col + 1):
            data = snapshot["cells"][c - 1]
            cell = ws.cell(row=row_idx, column=c)
            cell.value = data["value"]
            cell.font = data["font"]
            cell.border = data["border"]
            cell.fill = data["fill"]
            cell.number_format = data["number_format"]
            cell.protection = data["protection"]
            cell.alignment = data["alignment"]
            cell.comment = data["comment"]
            if data["hyperlink"]:
                self._apply_hyperlink(cell, data["hyperlink"])

    def _snapshot_column(self, ws, col_idx, max_row):
        cells = []
        for r in range(1, max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            cells.append({
                "value": cell.value,
                "font": copy.copy(cell.font),
                "border": copy.copy(cell.border),
                "fill": copy.copy(cell.fill),
                "number_format": cell.number_format,
                "protection": copy.copy(cell.protection),
                "alignment": copy.copy(cell.alignment),
                "comment": copy.copy(cell.comment) if cell.comment else None,
                "hyperlink": cell.hyperlink,
            })
        letter = get_column_letter(col_idx)
        dimension = ws.column_dimensions[letter]
        return {
            "width": dimension.width,
            "hidden": dimension.hidden,
            "cells": cells,
        }

    def _restore_column(self, ws, col_idx, snapshot, max_row):
        letter = get_column_letter(col_idx)
        dimension = ws.column_dimensions[letter]
        dimension.width = snapshot.get("width")
        dimension.hidden = bool(snapshot.get("hidden"))
        for r in range(1, max_row + 1):
            data = snapshot["cells"][r - 1]
            cell = ws.cell(row=r, column=col_idx)
            cell.value = data["value"]
            cell.font = data["font"]
            cell.border = data["border"]
            cell.fill = data["fill"]
            cell.number_format = data["number_format"]
            cell.protection = data["protection"]
            cell.alignment = data["alignment"]
            cell.comment = data["comment"]
            if data["hyperlink"]:
                self._apply_hyperlink(cell, data["hyperlink"])
            else:
                cell.hyperlink = None

    def _reorder_month_rows_desc(self, ws, header_keyword="月份", month_col=None, header_row=None, group_by_cols=None, max_col=None):
        if header_row is None or month_col is None:
            for r in range(1, min(ws.max_row, 50) + 1):
                for c in range(1, min(ws.max_column, 12) + 1):
                    if str(ws.cell(row=r, column=c).value).strip() == header_keyword:
                        header_row = r
                        month_col = c
                        break
                if header_row:
                    break
        if not header_row or not month_col:
            return False

        start_row = None
        for r in range(header_row + 1, ws.max_row + 1):
            key = self._month_label_exact_to_key(ws.cell(row=r, column=month_col).value)
            if key:
                start_row = r
                break
        if start_row is None:
            return False

        month_rows = []
        for r in range(start_row, ws.max_row + 1):
            key = self._month_label_exact_to_key(ws.cell(row=r, column=month_col).value)
            if not key:
                break
            month_rows.append((r, key))
        if not month_rows:
            return False

        group_map = {}
        group_order = []
        for r, key in month_rows:
            group_key = None
            if group_by_cols:
                values = []
                for c in group_by_cols:
                    if not c:
                        continue
                    values.append(ws.cell(row=r, column=c).value)
                group_key = tuple(values) if values else None
            if group_key not in group_map:
                group_map[group_key] = []
                group_order.append(group_key)
            group_map[group_key].append((r, key))

        desired_rows = []
        for gkey in group_order:
            rows = group_map[gkey]
            rows_sorted = sorted(rows, key=lambda x: x[1], reverse=True)
            desired_rows.extend([r for r, _ in rows_sorted])

        original_rows = [r for r, _ in month_rows]
        if desired_rows == original_rows:
            return False

        max_col = max_col or ws.max_column
        snapshots = {r: self._snapshot_row(ws, r, max_col) for r in original_rows}
        for idx, target_row in enumerate(range(start_row, start_row + len(original_rows))):
            src_row = desired_rows[idx]
            self._restore_row(ws, target_row, snapshots[src_row], max_col)
        return True

    def _ensure_report_month_structure(self, wb, target_year, target_month):
        changed = set()
        if "利润表" in wb.sheetnames:
            if self._ensure_month_columns_simple(wb["利润表"], target_year, target_month, header_row=1):
                changed.add("利润表")
        if "资产负债表" in wb.sheetnames:
            if self._ensure_month_columns_simple(wb["资产负债表"], target_year, target_month, header_row=1):
                changed.add("资产负债表")
        if "同比经营分析" in wb.sheetnames:
            if self._ensure_compare_month_row(wb["同比经营分析"], target_year, target_month):
                changed.add("同比经营分析")
        if "环比经营分析" in wb.sheetnames:
            if self._ensure_compare_month_row(wb["环比经营分析"], target_year, target_month):
                changed.add("环比经营分析")
        if "本量利分析" in wb.sheetnames:
            if self._ensure_month_rows_simple(wb["本量利分析"], target_year, target_month):
                changed.add("本量利分析")
        if "目标_预算" in wb.sheetnames:
            if self._ensure_month_rows_simple(wb["目标_预算"], target_year, target_month, total_label=""):
                changed.add("目标_预算")
        if "按品类汇总(按月)" in wb.sheetnames:
            if self._ensure_month_columns_grouped_by_suffix(wb["按品类汇总(按月)"], target_year, target_month):
                changed.add("按品类汇总(按月)")
        if "费用明细环比分析" in wb.sheetnames:
            if self._ensure_month_columns_grouped_by_suffix(wb["费用明细环比分析"], target_year, target_month):
                changed.add("费用明细环比分析")
        if "按产品汇总_含合计" in wb.sheetnames:
            if self._ensure_month_columns_grouped_by_month(wb["按产品汇总_含合计"], target_year, target_month):
                changed.add("按产品汇总_含合计")
        return changed

    def _infer_numeric_columns(self, ws, start_row, exclude_cols=None):
        exclude_cols = {c for c in (exclude_cols or []) if c}
        numeric_cols = []
        scan_end = min(ws.max_row, start_row + 5)
        for col in range(1, ws.max_column + 1):
            if col in exclude_cols:
                continue
            for r in range(start_row, scan_end + 1):
                value = ws.cell(row=r, column=col).value
                if isinstance(value, (int, float)):
                    numeric_cols.append(col)
                    break
        return numeric_cols

    def _pick_existing_columns(self, header_map, preferred):
        cols = []
        for key in preferred:
            col = header_map.get(key)
            if col:
                cols.append(col)
        return cols

    def _write_table(self, ws, start_row, start_col, headers, rows):
        for idx, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + idx)
            if not isinstance(cell, MergedCell):
                cell.value = self._sanitize_excel_text(header)
        
        # Apply standard header style
        self._apply_header_style(
            ws,
            start_row,
            start_col=start_col,
            max_col=start_col + len(headers) - 1,
        )

        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = self._sanitize_excel_text(value)

    def _add_line_chart_by_columns(
        self,
        ws,
        cat_col,
        series_cols,
        header_row,
        data_start_row,
        data_end_row,
        title,
        anchor,
    ):
        if not series_cols or data_end_row < data_start_row:
            return False
        chart = LineChart()
        chart.title = title
        for col in series_cols:
            data_ref = Reference(
                ws,
                min_col=col,
                max_col=col,
                min_row=header_row,
                max_row=data_end_row,
            )
            chart.add_data(data_ref, titles_from_data=True)
        cats_ref = self._chart_category_reference(ws, cat_col, data_start_row, data_end_row)
        chart.set_categories(cats_ref)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, anchor)
        return True

    def _add_bar_chart_from_table(self, ws, start_row, start_col, end_row, title, anchor):
        chart = BarChart()
        chart.title = title
        data_ref = Reference(
            ws,
            min_col=start_col + 1,
            max_col=start_col + 1,
            min_row=start_row,
            max_row=end_row,
        )
        cats_ref = self._chart_category_reference(ws, start_col, start_row + 1, end_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, anchor)
        return True

    def _add_bar_chart_by_columns(self, ws, cat_col, data_cols, header_row, data_start, data_end, title, anchor):
        if not data_cols or data_end < data_start:
            return False
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        cats_ref = self._chart_category_reference(ws, cat_col, data_start, data_end)
        for col in data_cols:
            data_ref = Reference(ws, min_col=col, max_col=col, min_row=header_row, max_row=data_end)
            chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 10
        chart.width = 18
        ws.add_chart(chart, anchor)
        return True

    def _add_combo_chart(self, ws, cat_col, bar_cols, line_cols, header_row, data_start, data_end, title, anchor):
        """
        创建一个组合图表：柱状图(主轴) + 折线图(次轴)
        bar_cols: list of column indices for Bar chart (Primary Axis)
        line_cols: list of column indices for Line chart (Secondary Axis)
        """
        # 1. Create Bar Chart (Primary)
        bar_chart = BarChart()
        bar_chart.title = title
        bar_chart.y_axis.title = "金额"
        bar_chart.x_axis.title = "月份"
        
        cats_ref = self._chart_category_reference(ws, cat_col, data_start, data_end)

        for col in bar_cols:
            data_ref = Reference(ws, min_col=col, max_col=col, min_row=header_row, max_row=data_end)
            bar_chart.add_data(data_ref, titles_from_data=True)
        if bar_cols:
            bar_chart.set_categories(cats_ref)

        # 2. Create Line Chart (Secondary)
        line_chart = LineChart()
        line_chart.y_axis.title = "利润/比率"
        line_chart.y_axis.axId = 200
        line_chart.y_axis.crosses = "max" # Put secondary axis on right
        
        # Ensure data exists for lines
        for col in line_cols:
            data_ref = Reference(ws, min_col=col, max_col=col, min_row=header_row, max_row=data_end)
            line_chart.add_data(data_ref, titles_from_data=True)
        if line_cols:
            line_chart.set_categories(cats_ref)

        # 3. Combine
        bar_chart.y_axis.crosses = "min"
        bar_chart += line_chart
        
        bar_chart.height = 10
        bar_chart.width = 20
        ws.add_chart(bar_chart, anchor)
        return True

    def _add_stacked_bar_chart(self, ws, cat_col, data_cols, header_row, data_start, data_end, title, anchor, percent=False):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.grouping = "percentStacked" if percent else "stacked"
        chart.overlap = 100
        chart.title = title
        
        cats_ref = self._chart_category_reference(ws, cat_col, data_start, data_end)

        for col in data_cols:
            data_ref = Reference(ws, min_col=col, max_col=col, min_row=header_row, max_row=data_end)
            chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        chart.height = 10
        chart.width = 18
        ws.add_chart(chart, anchor)
        return True

    def _short_chart_label(self, value, max_len=24):
        text = "" if value is None else str(value).strip()
        if len(text) <= max_len:
            return text
        return text[: max(1, max_len - 3)] + "..."

    def _add_doughnut_chart(self, ws, label_col, data_col, header_row, data_start, data_end, title, anchor):
        from openpyxl.chart import DoughnutChart
        chart = DoughnutChart()
        chart.title = title
        chart.style = 26
        
        labels = Reference(ws, min_col=label_col, min_row=data_start, max_row=data_end)
        data = Reference(ws, min_col=data_col, min_row=header_row, max_row=data_end)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        chart.height = 10
        chart.width = 10
        ws.add_chart(chart, anchor)
        return True

    def _write_chart_note(self, ws, col, row, text):
        if not text:
            return
        cell = ws.cell(row=row, column=col)
        cell.value = text
        if isinstance(text, str) and ("图表说明" in text or text.startswith("说明：")):
            cell.font = Font(name="微软雅黑", size=9, color="666666")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        self._normalize_note_row_height(ws, row)

    def _normalize_note_row_height(self, ws, row_idx):
        """
        Keep note rows visually consistent with neighboring rows.
        This avoids oversized note rows from template remnants or Excel auto-adjust side effects.
        """
        if row_idx <= 1:
            return

        default_height = float(ws.sheet_format.defaultRowHeight or 15.0)
        prev_height = ws.row_dimensions[row_idx - 1].height if row_idx > 1 else None
        next_height = ws.row_dimensions[row_idx + 1].height if row_idx < ws.max_row else None

        neighbor_heights = []
        for h in (prev_height, next_height):
            if isinstance(h, (int, float)) and h > 0:
                neighbor_heights.append(float(h))

        if neighbor_heights:
            target_height = sum(neighbor_heights) / len(neighbor_heights)
        else:
            target_height = default_height

        current_height = ws.row_dimensions[row_idx].height
        if current_height is None:
            # Only materialize row height when neighbors are explicitly styled.
            if neighbor_heights:
                ws.row_dimensions[row_idx].height = target_height
            return

        if abs(float(current_height) - target_height) > 0.1:
            ws.row_dimensions[row_idx].height = target_height

    def _row_height_emu(self, ws, row_idx):
        # Prefer sheet default row height when unspecified.
        h = ws.row_dimensions[row_idx].height
        if h is None:
            h = ws.sheet_format.defaultRowHeight or 15.0
        return float(h) * 12700

    def _col_width_emu(self, ws, col_idx):
        # Excel default column width is 8.43 chars when unspecified.
        letter = get_column_letter(col_idx)
        width_chars = ws.column_dimensions[letter].width
        if width_chars is None:
            width_chars = 8.43
        pixels = int(float(width_chars) * 7 + 5)
        return float(pixels) * 9525

    def _chart_extent_emu(self, chart, anchor=None):
        # openpyxl chart width/height unit is cm (1 cm = 360000 EMU).
        cx = 0
        cy = 0
        if anchor is not None:
            ext = getattr(anchor, "ext", None)
            if ext is not None:
                cx = int(getattr(ext, "cx", 0) or 0)
                cy = int(getattr(ext, "cy", 0) or 0)

        if cx <= 0:
            width_cm = getattr(chart, "width", None)
            if isinstance(width_cm, (int, float)) and width_cm > 0:
                cx = int(float(width_cm) * 360000)
        if cy <= 0:
            height_cm = getattr(chart, "height", None)
            if isinstance(height_cm, (int, float)) and height_cm > 0:
                cy = int(float(height_cm) * 360000)

        if cx <= 0:
            cx = 6480000
        if cy <= 0:
            cy = 2594250
        return cx, cy

    def _extract_chart_anchor_bbox(self, ws, chart):
        anchor = getattr(chart, "anchor", None)
        if anchor is None:
            return 1, 1, 8, 14

        if isinstance(anchor, str):
            m = re.match(r"^([A-Z]+)(\d+)$", anchor.strip())
            if m:
                col = column_index_from_string(m.group(1))
                row = int(m.group(2))
                cx, cy = self._chart_extent_emu(chart, anchor=None)

                remx = cx
                col_end = col
                while remx > 0 and col_end < 2000:
                    remx -= self._col_width_emu(ws, col_end)
                    if remx > 0:
                        col_end += 1

                remy = cy
                row_end = row
                while remy > 0 and row_end < 2000:
                    remy -= self._row_height_emu(ws, row_end)
                    if remy > 0:
                        row_end += 1

                return col, row, col_end, row_end
            return 1, 1, 8, 14

        if not hasattr(anchor, "_from"):
            return 1, 1, 8, 14

        col_start = int(getattr(anchor._from, "col", 0)) + 1
        row_start = int(getattr(anchor._from, "row", 0)) + 1

        # TwoCellAnchor provides explicit end marker.
        if hasattr(anchor, "_to") and getattr(anchor, "_to", None) is not None:
            col_end = int(getattr(anchor._to, "col", 0)) + 1
            row_end = int(getattr(anchor._to, "row", 0)) + 1
            return col_start, row_start, max(col_start, col_end), max(row_start, row_end)

        col_off = int(getattr(anchor._from, "colOff", 0) or 0)
        row_off = int(getattr(anchor._from, "rowOff", 0) or 0)
        cx, cy = self._chart_extent_emu(chart, anchor=anchor)

        remain_x = cx + col_off
        col_end = col_start
        while remain_x > 0 and col_end < 2000:
            remain_x -= self._col_width_emu(ws, col_end)
            if remain_x > 0:
                col_end += 1

        remain_y = cy + row_off
        row_end = row_start
        while remain_y > 0 and row_end < 2000:
            remain_y -= self._row_height_emu(ws, row_end)
            if remain_y > 0:
                row_end += 1

        return col_start, row_start, col_end, row_end

    def _manager_chart_focus_text(self, sheet_name):
        mapping = {
            "仪表盘": "先看收入、净利润与现金流是否同向；若背离，优先排查回款、库存与一次性费用。",
            "目标_预算": "关注预算偏差绝对值与连续超阈值月份，优先处理偏差最大的两项指标。",
            "预算执行与偏差": "按“偏差金额+偏差率”双维度排序，区分规模问题与效率问题。",
            "经营指标": "先看收入/成本/利润趋势，再看费用率与周转天数，判断增长质量。",
            "利润表": "优先判断毛利率与净利率变化，再回溯费用与营业外项目对利润的影响。",
            "资产负债表": "关注应收、存货、负债结构变化，评估资金占用与偿债压力。",
            "费用对比": "聚焦异常上升科目，结合业务动作区分结构性上升与短期波动。",
            "按品类汇总(按月)": "先看品类贡献变化，再看占比迁移，识别结构升级或单品依赖风险。",
            "按产品汇总_含合计": "按收入/毛利贡献排序，识别头部产品依赖与尾部低效占用。",
            "明细_销售与库存": "联动销量、库存与金额，优先关注高库存低周转品目。",
            "本量利分析": "重点看安全边际与盈亏平衡点，评估利润抗压能力。",
            "应收账款账龄分析": "优先看超期账龄占比与Top客户集中度，制定回款优先级。",
            "客户贡献与回款": "比较客户利润贡献与回款效率，优先优化高贡献低回款客户。",
            "现金流量表(估算)": "对比净利润与经营现金流方向，识别利润现金化问题。",
        }
        return mapping.get(sheet_name, "先看趋势方向，再看异常波动，最后下钻明细定位原因并给出动作。")

    def _append_chart_notes_below(self, wb, target_year=None, target_month=None):
        period = ""
        if target_year and target_month:
            period = f"{target_year}年{int(target_month):02d}月"

        for ws in wb.worksheets:
            charts = list(getattr(ws, "_charts", []) or [])
            if not charts:
                continue

            # Normalize pre-existing management notes (from template or previous runs).
            existing_notes = []
            for cell in getattr(ws, "_cells", {}).values():
                val = cell.value
                if not isinstance(val, str):
                    continue
                if "图表说明（" in val and "管理者看点：" in val:
                    existing_notes.append(cell)
            for cell in existing_notes:
                cell.font = Font(name="微软雅黑", size=9, color="666666")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                self._normalize_note_row_height(ws, cell.row)

            # Remove accidental adjacent duplicates in same column.
            seen_row_by_key = {}
            for cell in sorted(existing_notes, key=lambda c: (c.column, c.row)):
                key = (cell.column, cell.value)
                prev_row = seen_row_by_key.get(key)
                if prev_row is not None and abs(cell.row - prev_row) <= 1:
                    cell.value = None
                    continue
                seen_row_by_key[key] = cell.row

            bboxes = [self._extract_chart_anchor_bbox(ws, ch) for ch in charts]
            occupied = set()
            placed_notes = [
                (cell.row, cell.column, cell.value)
                for cell in existing_notes
                if isinstance(cell.value, str) and cell.value.strip()
            ]

            for idx, bbox in enumerate(bboxes, start=1):
                col_start, row_start, col_end, row_end = bbox
                note_col = col_start
                note_row = row_end + 2

                # Keep moving down until the note cell no longer falls into any chart area.
                moved = True
                while moved:
                    moved = False
                    for b_col_start, b_row_start, b_col_end, b_row_end in bboxes:
                        if not (b_col_start <= note_col <= b_col_end):
                            continue
                        if (b_row_start - 1) <= note_row <= (b_row_end + 1):
                            note_row = b_row_end + 2
                            moved = True

                # Avoid merged-cell targets and duplicate note positions.
                while isinstance(ws.cell(row=note_row, column=note_col), MergedCell) or (note_row, note_col) in occupied:
                    note_row += 1
                while (note_row, note_col) in occupied:
                    note_row += 1
                occupied.add((note_row, note_col))

                focus = self._manager_chart_focus_text(ws.title)
                prefix = f"图表说明（{period}）" if period else "图表说明"
                note_text = (
                    f"{prefix}：先看趋势方向，再看异常点，再下钻明细核因。"
                    f"管理者看点：{focus}"
                )
                # Skip near-duplicate notes in the same column (often caused by adjacent charts).
                if any(c == note_col and txt == note_text and abs(r - note_row) <= 1 for r, c, txt in placed_notes):
                    continue
                cell = ws.cell(row=note_row, column=note_col)
                cell.value = note_text
                cell.font = Font(name="微软雅黑", size=9, color="666666")
                # Disable wrap to avoid Excel auto-expanding this row on open.
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                self._normalize_note_row_height(ws, note_row)
                placed_notes.append((note_row, note_col, note_text))

        self._log_audit("已在图表下补充管理解读说明")

    def _update_manager_guide_sheet(self, wb, target_year=None, target_month=None):
        insert_after = "计算说明" if "计算说明" in wb.sheetnames else "目录"
        ws = self._prepare_sheet(wb, "管理者解读", insert_after=insert_after)
        period_text = ""
        if target_year and target_month:
            period_text = f"（{target_year}年{int(target_month):02d}月）"

        lines = [
            f"经营分析报告管理者阅读指南{period_text}",
            "",
            "一、建议阅读顺序",
            "1. 先看「仪表盘」：判断收入、净利润、现金流是否同向。",
            "2. 再看「经营指标/利润表」：确认增长质量（规模、毛利、净利）。",
            "3. 然后看「预算执行与偏差/费用对比」：定位偏差来源和费用异常。",
            "4. 最后看「按品类汇总(按月)/按产品汇总_含合计/存货健康度/应收账款账龄分析」：落实到品类、产品、库存、回款动作。",
            "",
            "二、管理层重点判断框架",
            "1. 是否“增收增利增现”：收入增长是否同时带来利润与经营现金流改善。",
            "2. 结构是否优化：高毛利品类/产品占比是否提升，低效项目是否收缩。",
            "3. 资金是否安全：应收与存货周转是否恶化，现金覆盖月数是否下降。",
            "4. 偏差是否可控：预算偏差是否集中在少数指标，是否存在连续超阈值。",
            "",
            "三、建议动作模板（管理例会可直接使用）",
            "1. 本月最重要的3个异常：说明指标、原因、影响金额、责任人、完成时点。",
            "2. 下月追踪的3个先行指标：回款率、库存周转、费用率（按部门/品类拆解）。",
            "3. 决策结论：继续投入、结构调整、风险控制三类动作分别列出。",
        ]

        for idx, text in enumerate(lines, start=1):
            cell = ws.cell(row=idx, column=1)
            cell.value = text
            if idx == 1:
                cell.font = Font(name="微软雅黑", size=13, bold=True, color="1F4E79")
            elif text.startswith(("一、", "二、", "三、")):
                cell.font = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
            else:
                cell.font = Font(name="微软雅黑", size=10, color="333333")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions["A"].width = 120

    def _add_chart_management_metrics(self, ws):
        header_row = 1
        header_map = self._get_header_map(ws, header_row)
        cat_col = header_map.get("月份")
        if not cat_col:
            return False
        data_start = header_row + 1
        data_start, data_end = self._find_month_data_bounds(ws, cat_col, data_start)
        if data_end < data_start:
            return False
        
        # Split columns for Dual Axis
        bar_keys = ["主营业务收入", "主营业务成本"]
        line_keys = ["营业利润", "净利润"]
        
        bar_cols = self._pick_existing_columns(header_map, bar_keys)
        line_cols = self._pick_existing_columns(header_map, line_keys)
        
        if not bar_cols and not line_cols:
            return False
            
        anchor_col = ws.max_column + 2
        anchor = f"{get_column_letter(anchor_col)}2"
        
        # Use Combo Chart
        added = self._add_combo_chart(
            ws,
            cat_col,
            bar_cols,
            line_cols,
            header_row,
            data_start,
            data_end,
            "经营指标趋势 (收入成本vs利润)",
            anchor,
        )
        
        if added:
            self._write_chart_note(ws, anchor_col, 1, "图表说明：柱状图=收入/成本(左轴)，折线图=利润(右轴)。")
        return added

    def _add_chart_expense_compare(self, ws):
        header_row = 1
        header_map = self._get_header_map(ws, header_row)
        cat_col = header_map.get("月份")
        if not cat_col:
            return False
        data_start = header_row + 1
        data_start, data_end = self._find_month_data_bounds(ws, cat_col, data_start)
        if data_end < data_start:
            return False
        preferred = ["主营业务成本占比", "销售费用占比", "管理费用占比", "财务费用占比"]
        series_cols = self._pick_existing_columns(header_map, preferred)
        if not series_cols:
            preferred = ["主营业务成本", "销售费用", "管理费用", "财务费用"]
            series_cols = self._pick_existing_columns(header_map, preferred)
        if not series_cols:
            return False
        anchor_col = ws.max_column + 2
        anchor = f"{get_column_letter(anchor_col)}2"
        added = self._add_line_chart_by_columns(
            ws,
            cat_col,
            series_cols,
            header_row,
            data_start,
            data_end,
            "费用对比趋势",
            anchor,
        )
        if added:
            self._write_chart_note(ws, anchor_col, 1, "图表说明：成本、销售、管理、财务费用占比的月度变化。")
        return added

    def _add_chart_cvp(self, ws):
        header_row = 1
        header_map = self._get_header_map(ws, header_row)
        cat_col = header_map.get("月份")
        if not cat_col:
            return False
        data_start = header_row + 1
        data_start, data_end = self._find_month_data_bounds(ws, cat_col, data_start)
        if data_end < data_start:
            return False

        anchor_col = ws.max_column + 2
        added = False

        amount_cols = self._pick_existing_columns(
            header_map,
            ["销售收入", "变动成本", "固定成本", "总成本", "盈亏平衡点"],
        )
        if amount_cols:
            added = self._add_line_chart_by_columns(
                ws,
                cat_col,
                amount_cols,
                header_row,
                data_start,
                data_end,
                "本量利金额趋势",
                f"{get_column_letter(anchor_col)}2",
            ) or added

        margin_series_cols = self._pick_existing_columns(
            header_map,
            ["销售收入", "盈亏平衡点", "安全边际"],
        )
        if margin_series_cols:
            margin_chart = LineChart()
            margin_chart.title = "盈亏平衡与安全边际"
            margin_chart.y_axis.title = "金额"
            margin_chart.x_axis.title = "月份"
            for col in margin_series_cols:
                data_ref = Reference(ws, min_col=col, max_col=col, min_row=header_row, max_row=data_end)
                margin_chart.add_data(data_ref, titles_from_data=True)
            cats_ref = self._chart_category_reference(ws, cat_col, data_start, data_end)
            margin_chart.set_categories(cats_ref)
            margin_chart.height = 10
            margin_chart.width = 18
            ws.add_chart(margin_chart, f"{get_column_letter(anchor_col)}18")
            added = True

        rate_cols = self._pick_existing_columns(
            header_map,
            ["贡献毛利率", "安全边际率"],
        )
        if rate_cols:
            rate_chart = LineChart()
            rate_chart.title = "贡献毛利率/安全边际率趋势"
            rate_chart.y_axis.title = "比率"
            rate_chart.x_axis.title = "月份"
            for col in rate_cols:
                data_ref = Reference(ws, min_col=col, max_col=col, min_row=header_row, max_row=data_end)
                rate_chart.add_data(data_ref, titles_from_data=True)
            cats_ref = self._chart_category_reference(ws, cat_col, data_start, data_end)
            rate_chart.set_categories(cats_ref)
            rate_chart.height = 10
            rate_chart.width = 18
            ws.add_chart(rate_chart, f"{get_column_letter(anchor_col)}34")
            added = True

        if added:
            self._write_chart_note(
                ws,
                anchor_col,
                1,
                "图表说明：图表不含合计行；依次展示金额趋势、盈亏平衡/安全边际、贡献毛利率与安全边际率。",
            )
        return added

    def _add_chart_budget(self, ws):
        header_row = self._find_header_row_by_keyword(ws, "月份", max_row=30)
        if not header_row:
            return False
        header_map = self._get_header_map(ws, header_row)
        cat_col = header_map.get("月份")
        if not cat_col:
            return False
        data_start = header_row + 1
        data_start, data_end = self._find_month_data_bounds(ws, cat_col, data_start)
        if data_end < data_start:
            return False
        preferred = [
            "主营业务收入目标",
            "营业利润目标",
            "营业利润率目标",
            "成本率目标",
            "销售费用率目标",
            "管理费用率目标",
            "财务费用率目标",
            "应收账款余额目标",
            "存货期末余额目标",
        ]
        series_cols = self._pick_existing_columns(header_map, preferred)
        if not series_cols:
            exclude = [cat_col]
            series_cols = self._infer_numeric_columns(ws, data_start, exclude)[:4]
        else:
            series_cols = series_cols[:4]
        if not series_cols:
            return False
        anchor_col = ws.max_column + 2
        anchor = f"{get_column_letter(anchor_col)}2"
        added = self._add_line_chart_by_columns(
            ws,
            cat_col,
            series_cols,
            header_row,
            data_start,
            data_end,
            "目标/预算趋势",
            anchor,
        )
        if added:
            self._write_chart_note(ws, anchor_col, 1, "图表说明：关键目标指标的月度变化趋势。")
        return added

    def _add_chart_compare_sheet(self, ws, label):
        header_row = 1
        header_map = self._get_header_map(ws, header_row)
        month_col = header_map.get("月份")
        indicator_col = header_map.get("指标")
        value_col = header_map.get("本期值")
        if not month_col or not indicator_col or not value_col:
            return False
        data = {}
        months = []
        for r in range(header_row + 1, ws.max_row + 1):
            month = ws.cell(row=r, column=month_col).value
            indicator = ws.cell(row=r, column=indicator_col).value
            value = self._to_float(ws.cell(row=r, column=value_col).value)
            if month is None or indicator is None:
                continue
            if month not in months:
                months.append(month)
            indicator = str(indicator).strip()
            data.setdefault(indicator, {})[month] = value
        if not months or not data:
            return False
        preferred = ["主营业务收入", "主营业务成本", "销售费用", "管理费用", "财务费用", "营业利润"]
        indicators = [name for name in preferred if name in data]
        if not indicators:
            indicators = list(data.keys())[:3]
        if not indicators:
            return False
        start_col = ws.max_column + 2
        headers = ["月份"] + indicators
        rows = []
        for month in months:
            rows.append([month] + [data.get(ind, {}).get(month) for ind in indicators])
        self._write_table(ws, 1, start_col, headers, rows)
        anchor_col = start_col + len(headers) + 2
        anchor = f"{get_column_letter(anchor_col)}2"
        added = self._add_line_chart_by_columns(
            ws,
            start_col,
            list(range(start_col + 1, start_col + 1 + len(indicators))),
            1,
            2,
            1 + len(rows),
            f"{label}趋势",
            anchor,
        )
        if added:
            self._write_chart_note(ws, anchor_col, 1, f"图表说明：主要指标本期值的{label}趋势。")
        return added

    def _add_chart_expense_detail(self, ws):
        header_row = 1
        header_map = self._get_header_map(ws, header_row)
        sub_col = header_map.get("子科目")
        cat_col = header_map.get("费用类别")
        dim_col = sub_col or cat_col
        dim_label = "子科目" if sub_col else "费用类别"
        if not dim_col:
            return False
        amount_col = None
        for key in ["金额", "本币金额", "本位币金额", "金额(本位币)"]:
            if key in header_map:
                amount_col = header_map[key]
                break
        if not amount_col:
            for name, col in header_map.items():
                if "金额" in name and "外币" not in name:
                    amount_col = col
                    break
        if not amount_col:
            return False
        totals = {}
        for r in range(header_row + 1, ws.max_row + 1):
            dim_val = ws.cell(row=r, column=dim_col).value
            amount = self._to_float(ws.cell(row=r, column=amount_col).value)
            if not dim_val or amount is None:
                continue
            dim_val = str(dim_val).strip()
            totals[dim_val] = totals.get(dim_val, 0) + amount
        if not totals:
            return False
        top_items = sorted(totals.items(), key=lambda x: abs(x[1]), reverse=True)[:8]
        start_col = ws.max_column + 2
        headers = [dim_label, "金额"]
        rows = [[name, value] for name, value in top_items]
        self._write_table(ws, 1, start_col, headers, rows)
        anchor_col = start_col + len(headers) + 2
        anchor = f"{get_column_letter(anchor_col)}2"
        added = self._add_bar_chart_from_table(
            ws,
            1,
            start_col,
            1 + len(rows),
            f"{dim_label}Top (柱状)",
            anchor,
        )
        
        # Add Doughnut Chart
        anchor_pie = f"{get_column_letter(anchor_col)}18"
        self._add_doughnut_chart(
            ws,
            start_col,     # Labels (Category Name)
            start_col + 1, # Data (Amount)
            1,
            2,             # Data Start Row
            1 + len(rows),
            f"费用构成（按{dim_label}）",
            anchor_pie
        )
        
        if added:
            self._write_chart_note(ws, anchor_col, 1, f"图表说明：按{dim_label}汇总金额的Top分布及构成。")
        return added

    def _add_chart_expense_mom(self, ws):
        header_row = 1
        header_map = self._get_header_map(ws, header_row)
        cat_col = header_map.get("费用类别")
        sub_col = header_map.get("子科目")
        if not cat_col:
            return False
        month_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if isinstance(header, str) and header.endswith("_本期值"):
                month_label = header.split("_")[0]
                month_cols.append((month_label, col))
        if not month_cols:
            return False
        items = []
        for r in range(header_row + 1, ws.max_row + 1):
            cat = ws.cell(row=r, column=cat_col).value
            if not cat:
                continue
            sub = ws.cell(row=r, column=sub_col).value if sub_col else None
            label = f"{cat}-{sub}" if sub else str(cat)
            values = {}
            total = 0
            for month_label, col in month_cols:
                val = self._to_float(ws.cell(row=r, column=col).value)
                if val is not None:
                    values[month_label] = val
                    total += abs(val)
            if total > 0:
                items.append({"label": label, "values": values, "total": total})
        if not items:
            return False
        items = sorted(items, key=lambda x: x["total"], reverse=True)[:5]
        start_col = ws.max_column + 2
        headers = ["月份"] + [item["label"] for item in items]
        rows = []
        for month_label, _ in month_cols:
            rows.append([month_label] + [item["values"].get(month_label) for item in items])
        self._write_table(ws, 1, start_col, headers, rows)
        anchor_col = start_col + len(headers) + 2
        anchor = f"{get_column_letter(anchor_col)}2"
        added = self._add_line_chart_by_columns(
            ws,
            start_col,
            list(range(start_col + 1, start_col + 1 + len(items))),
            1,
            2,
            1 + len(rows),
            "费用明细环比趋势",
            anchor,
        )
        if added:
            self._write_chart_note(ws, anchor_col, 1, "图表说明：主要费用子科目的月度变化趋势。")
        return added

    def _add_chart_category_month(self, ws):
        header_row = 1
        header_map = self._get_header_map(ws, header_row)
        cat_col = header_map.get("产品大类")
        if not cat_col:
            return False
        month_cols = []
        for header, col in header_map.items():
            if isinstance(header, str) and header.endswith("_毛利润"):
                month_cols.append((header.replace("_毛利润", ""), col))
        if not month_cols:
            return False
        items = []
        for r in range(header_row + 1, ws.max_row + 1):
            cat = ws.cell(row=r, column=cat_col).value
            if not cat:
                continue
            cat = str(cat).strip()
            if cat == "合计" or cat.endswith("占比"):
                continue
            values = {}
            for month_label, col in month_cols:
                values[month_label] = self._to_float(ws.cell(row=r, column=col).value)
            items.append({"label": cat, "values": values})
        if not items:
            return False
        start_col = ws.max_column + 2
        headers = ["月份"] + [item["label"] for item in items]
        rows = []
        for month_label, _ in month_cols:
            rows.append([month_label] + [item["values"].get(month_label) for item in items])
        self._write_table(ws, 1, start_col, headers, rows)
        anchor_col = start_col + len(headers) + 2
        anchor = f"{get_column_letter(anchor_col)}2"
        added = self._add_line_chart_by_columns(
            ws,
            start_col,
            list(range(start_col + 1, start_col + 1 + len(items))),
            1,
            2,
            1 + len(rows),
            "品类毛利润趋势",
            anchor,
        )
        if added:
            self._write_chart_note(ws, anchor_col, 1, "图表说明：各品类毛利润的月度趋势。")
        return added

    def _add_chart_product_summary(self, ws):
        header_row = 1
        header_map = self._get_header_map(ws, header_row)
        name_col = header_map.get("产品名称") or header_map.get("品目名规格")
        if not name_col:
            return False

        revenue_col = header_map.get("年销售收入合计")
        profit_col = header_map.get("年销售利润合计")
        margin_col = header_map.get("年毛利率平均")
        inventory_col = (
            header_map.get("当前库存金额")
            or header_map.get("年末存货金额")
            or header_map.get("期末库存金额")
        )
        turnover_days_col = header_map.get("存货周转天数")
        qty_col = header_map.get("年销售数量合计")

        products = []
        total_row = None
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=name_col).value
            if not name:
                continue
            name_text = str(name).strip()
            if not name_text:
                continue
            if "合计" in name_text or name_text.upper() == "TOTAL":
                if total_row is None:
                    total_row = r
                continue

            item = {
                "name": name_text,
                "label": self._short_chart_label(name_text, 18),
                "revenue": self._to_float(ws.cell(row=r, column=revenue_col).value) if revenue_col else None,
                "profit": self._to_float(ws.cell(row=r, column=profit_col).value) if profit_col else None,
                "margin": self._to_float(ws.cell(row=r, column=margin_col).value) if margin_col else None,
                "inventory": self._to_float(ws.cell(row=r, column=inventory_col).value) if inventory_col else None,
                "turnover_days": self._to_float(ws.cell(row=r, column=turnover_days_col).value) if turnover_days_col else None,
                "qty": self._to_float(ws.cell(row=r, column=qty_col).value) if qty_col else None,
            }
            if any(item.get(key) is not None for key in ("revenue", "profit", "margin", "inventory", "turnover_days", "qty")):
                products.append(item)
        if not products:
            return False

        start_col = ws.max_column + 2
        anchor_col = start_col + 6
        row_cursor = 1
        added = False

        def has_series_value(rows, index):
            return any(row[index] is not None for row in rows)

        def advance(rows):
            return max(len(rows) + 4, 22)

        top_revenue = [
            p for p in products
            if p.get("revenue") is not None or p.get("profit") is not None or p.get("margin") is not None
        ]
        top_revenue = sorted(top_revenue, key=lambda p: abs(p.get("revenue") or 0), reverse=True)[:10]
        if top_revenue and not self._worksheet_has_chart_title(ws, "产品收入/毛利Top"):
            rows = [[p["label"], p.get("revenue"), p.get("profit"), p.get("margin")] for p in top_revenue]
            self._write_table(ws, row_cursor, start_col, ["产品", "销售收入", "毛利润", "毛利率"], rows)
            for rr in range(row_cursor + 1, row_cursor + 1 + len(rows)):
                ws.cell(row=rr, column=start_col + 3).number_format = "0.00%"
            bar_cols = [start_col + idx for idx in (1, 2) if has_series_value(rows, idx)]
            line_cols = [start_col + 3] if has_series_value(rows, 3) else []
            anchor = f"{get_column_letter(anchor_col)}{row_cursor + 1}"
            if bar_cols and line_cols:
                chart_added = self._add_combo_chart(
                    ws, start_col, bar_cols, line_cols, row_cursor,
                    row_cursor + 1, row_cursor + len(rows), "产品收入/毛利Top", anchor
                )
            else:
                chart_added = self._add_bar_chart_by_columns(
                    ws, start_col, bar_cols or line_cols, row_cursor,
                    row_cursor + 1, row_cursor + len(rows), "产品收入/毛利Top", anchor
                )
            if chart_added:
                added = True
                self._write_chart_note(ws, anchor_col, row_cursor, "图表说明：头部产品收入、毛利润与毛利率联动。")
            row_cursor += advance(rows)

        top_inventory = [p for p in products if p.get("inventory") is not None]
        top_inventory = sorted(top_inventory, key=lambda p: abs(p.get("inventory") or 0), reverse=True)[:10]
        if top_inventory and not self._worksheet_has_chart_title(ws, "产品库存占用与周转Top"):
            rows = [[p["label"], p.get("inventory"), p.get("turnover_days")] for p in top_inventory]
            self._write_table(ws, row_cursor, start_col, ["产品", "库存金额", "周转天数"], rows)
            bar_cols = [start_col + 1] if has_series_value(rows, 1) else []
            line_cols = [start_col + 2] if has_series_value(rows, 2) else []
            anchor = f"{get_column_letter(anchor_col)}{row_cursor + 1}"
            if bar_cols and line_cols:
                chart_added = self._add_combo_chart(
                    ws, start_col, bar_cols, line_cols, row_cursor,
                    row_cursor + 1, row_cursor + len(rows), "产品库存占用与周转Top", anchor
                )
            else:
                chart_added = self._add_bar_chart_by_columns(
                    ws, start_col, bar_cols or line_cols, row_cursor,
                    row_cursor + 1, row_cursor + len(rows), "产品库存占用与周转Top", anchor
                )
            if chart_added:
                added = True
                self._write_chart_note(ws, anchor_col, row_cursor, "图表说明：库存占用大的产品及其周转天数。")
            row_cursor += advance(rows)

        low_margin = [p for p in products if p.get("margin") is not None and (p.get("revenue") or 0) > 0]
        low_margin = sorted(low_margin, key=lambda p: p.get("margin"))[:10]
        if low_margin and not self._worksheet_has_chart_title(ws, "低毛利产品Top"):
            rows = [[p["label"], p.get("margin")] for p in low_margin]
            self._write_table(ws, row_cursor, start_col, ["产品", "毛利率"], rows)
            for rr in range(row_cursor + 1, row_cursor + 1 + len(rows)):
                ws.cell(row=rr, column=start_col + 1).number_format = "0.00%"
            anchor = f"{get_column_letter(anchor_col)}{row_cursor + 1}"
            chart_added = self._add_bar_chart_from_table(
                ws, row_cursor, start_col, row_cursor + len(rows), "低毛利产品Top", anchor
            )
            if chart_added:
                added = True
                self._write_chart_note(ws, anchor_col, row_cursor, "图表说明：优先复核低毛利或负毛利产品的定价与成本。")
            row_cursor += advance(rows)

        matrix_rows = [
            p for p in products
            if p.get("revenue") is not None and p.get("margin") is not None
        ]
        matrix_rows = sorted(matrix_rows, key=lambda p: abs(p.get("revenue") or 0), reverse=True)[:30]
        if len(matrix_rows) >= 2 and not self._worksheet_has_chart_title(ws, "产品收入 vs 毛利率矩阵"):
            rows = [[p["label"], p.get("revenue"), p.get("margin"), p.get("inventory")] for p in matrix_rows]
            self._write_table(ws, row_cursor, start_col, ["产品", "销售收入", "毛利率", "库存金额"], rows)
            for rr in range(row_cursor + 1, row_cursor + 1 + len(rows)):
                ws.cell(row=rr, column=start_col + 2).number_format = "0.00%"
            anchor = f"{get_column_letter(anchor_col)}{row_cursor + 1}"
            chart_added = self._add_scatter_chart(
                ws,
                start_col + 1,
                start_col + 2,
                row_cursor,
                row_cursor + 1,
                row_cursor + len(rows),
                "产品收入 vs 毛利率矩阵",
                anchor,
                x_title="销售收入",
                y_title="毛利率",
            )
            if chart_added:
                added = True
                self._write_chart_note(ws, anchor_col, row_cursor, "图表说明：识别高收入低毛利、低收入高毛利等产品组合。")
            row_cursor += advance(rows)

        month_metric_cols = {}
        for header, col in header_map.items():
            m = re.match(r"(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?_(.+)", str(header))
            if not m:
                continue
            month_key = f"{m.group(1)}-{int(m.group(2)):02d}"
            month_label = f"{m.group(1)}/{int(m.group(2)):02d}"
            suffix = m.group(3)
            metric = None
            if "销售收入" in suffix:
                metric = "revenue"
            elif "销售利润" in suffix or "毛利润" in suffix:
                metric = "profit"
            elif "期末库存金额" in suffix:
                metric = "inventory"
            if metric:
                month_metric_cols.setdefault(month_key, {"label": month_label})[metric] = col

        trend_rows = []
        for month_key in sorted(month_metric_cols.keys()):
            cols = month_metric_cols[month_key]
            row = [cols["label"]]
            for metric in ("revenue", "profit", "inventory"):
                col = cols.get(metric)
                if not col:
                    row.append(None)
                elif total_row:
                    row.append(self._to_float(ws.cell(row=total_row, column=col).value))
                else:
                    row.append(self._safe_sum([
                        self._to_float(ws.cell(row=r, column=col).value)
                        for r in range(header_row + 1, ws.max_row + 1)
                        if ws.cell(row=r, column=name_col).value
                        and "合计" not in str(ws.cell(row=r, column=name_col).value)
                    ]))
            if any(v is not None for v in row[1:]):
                trend_rows.append(row)
        if len(trend_rows) >= 2 and not self._worksheet_has_chart_title(ws, "产品销售/库存月度趋势"):
            self._write_table(ws, row_cursor, start_col, ["月份", "销售收入", "销售利润", "期末库存金额"], trend_rows)
            series_cols = [
                start_col + idx
                for idx in (1, 2, 3)
                if any(row[idx] is not None for row in trend_rows)
            ]
            anchor = f"{get_column_letter(anchor_col)}{row_cursor + 1}"
            chart_added = self._add_line_chart_by_columns(
                ws,
                start_col,
                series_cols,
                row_cursor,
                row_cursor + 1,
                row_cursor + len(trend_rows),
                "产品销售/库存月度趋势",
                anchor,
            )
            if chart_added:
                added = True
                self._write_chart_note(ws, anchor_col, row_cursor, "图表说明：合计口径下销售、毛利与期末库存金额趋势。")
        return added

    def _add_chart_sales_inventory_detail(self, ws):
        header_row = 1
        header_map = self._get_header_map(ws, header_row)
        cat_col = header_map.get("产品大类") or header_map.get("品目类型")
        if not cat_col:
            return False

        revenue_col = header_map.get("销售收入")
        inventory_col = header_map.get("期末金额")
        qty_col = header_map.get("销售数量")
        product_col = header_map.get("产品名称") or header_map.get("品目名规格") or header_map.get("品目编码")
        profit_col = header_map.get("毛利润") or header_map.get("销售利润")
        margin_col = header_map.get("毛利率")

        records = []
        cat_totals = {}

        def add_total(cat, key, value):
            if value is None:
                return
            bucket = cat_totals.setdefault(cat, {})
            bucket[key] = bucket.get(key, 0) + value
            bucket[f"{key}_has"] = True

        for r in range(header_row + 1, ws.max_row + 1):
            cat = ws.cell(row=r, column=cat_col).value
            if not cat:
                continue
            cat = str(cat).strip()
            if not cat:
                continue
            product = ws.cell(row=r, column=product_col).value if product_col else cat
            product_label = self._short_chart_label(product or cat, 18)
            revenue = self._to_float(ws.cell(row=r, column=revenue_col).value) if revenue_col else None
            inventory = self._to_float(ws.cell(row=r, column=inventory_col).value) if inventory_col else None
            qty = self._to_float(ws.cell(row=r, column=qty_col).value) if qty_col else None
            profit = self._to_float(ws.cell(row=r, column=profit_col).value) if profit_col else None
            margin = self._to_float(ws.cell(row=r, column=margin_col).value) if margin_col else None
            if profit is None and revenue is not None:
                cost_col = header_map.get("销售成本")
                cost = self._to_float(ws.cell(row=r, column=cost_col).value) if cost_col else None
                if cost is not None:
                    profit = revenue - cost
            if margin is None and revenue not in (None, 0) and profit is not None:
                margin = profit / revenue

            add_total(cat, "revenue", revenue)
            add_total(cat, "inventory", inventory)
            add_total(cat, "qty", qty)
            add_total(cat, "profit", profit)

            if any(v is not None for v in (revenue, inventory, qty, profit, margin)):
                records.append({
                    "category": cat,
                    "product": product_label,
                    "revenue": revenue,
                    "inventory": inventory,
                    "qty": qty,
                    "profit": profit,
                    "margin": margin,
                })

        if not records and not cat_totals:
            return False

        def total_value(bucket, key):
            return bucket.get(key) if bucket.get(f"{key}_has") else None

        # Keep a stable helper range so chart references do not drift right across regenerations.
        start_col = 28  # AB
        helper_end_col = start_col + 8
        for r in range(1, ws.max_row + 1):
            for c in range(start_col, helper_end_col + 1):
                self._safe_set_cell_value(ws, r, c, None)

        anchor_col = start_col + 4
        row_cursor = 1
        added = False

        def has_series_value(rows, index):
            return any(row[index] is not None for row in rows)

        def advance(rows):
            return max(len(rows) + 4, 22)

        cat_rows = []
        has_revenue = any(bucket.get("revenue_has") for bucket in cat_totals.values())
        amount_key = "revenue" if has_revenue else "inventory"
        for cat, bucket in cat_totals.items():
            value = total_value(bucket, amount_key)
            if value is not None:
                cat_rows.append([cat, value])
        top_items = sorted(cat_rows, key=lambda x: abs(x[1]), reverse=True)[:8]
        if top_items:
            self._write_table(ws, row_cursor, start_col, ["品类", "金额"], top_items)
            anchor = f"{get_column_letter(anchor_col)}{row_cursor + 1}"
            chart_added = self._add_bar_chart_from_table(
                ws,
                row_cursor,
                start_col,
                row_cursor + len(top_items),
                "销售/库存Top",
                anchor,
            )
            if chart_added:
                added = True
                self._write_chart_note(ws, anchor_col, row_cursor, "图表说明：按品类汇总的销售收入/期末金额。")
            row_cursor += advance(top_items)

        compare_rows = []
        for cat, bucket in cat_totals.items():
            revenue = total_value(bucket, "revenue")
            inventory = total_value(bucket, "inventory")
            profit = total_value(bucket, "profit")
            if any(v is not None for v in (revenue, inventory, profit)):
                compare_rows.append([cat, revenue, inventory, profit])
        compare_rows = sorted(
            compare_rows,
            key=lambda row: max(abs(row[1] or 0), abs(row[2] or 0), abs(row[3] or 0)),
            reverse=True,
        )[:10]
        if compare_rows:
            self._write_table(ws, row_cursor, start_col, ["品类", "销售收入", "期末库存金额", "毛利润"], compare_rows)
            series_cols = [start_col + idx for idx in (1, 2, 3) if has_series_value(compare_rows, idx)]
            anchor = f"{get_column_letter(anchor_col)}{row_cursor + 1}"
            chart_added = self._add_bar_chart_by_columns(
                ws,
                start_col,
                series_cols,
                row_cursor,
                row_cursor + 1,
                row_cursor + len(compare_rows),
                "品类销售/库存/毛利对比",
                anchor,
            )
            if chart_added:
                added = True
                self._write_chart_note(ws, anchor_col, row_cursor, "图表说明：比较品类销售收入、库存占用与毛利润。")
            row_cursor += advance(compare_rows)

        inventory_rows = [
            [r["product"], r.get("inventory")]
            for r in records
            if r.get("inventory") is not None
        ]
        inventory_rows = sorted(inventory_rows, key=lambda row: abs(row[1] or 0), reverse=True)[:10]
        if inventory_rows:
            self._write_table(ws, row_cursor, start_col, ["产品", "期末库存金额"], inventory_rows)
            anchor = f"{get_column_letter(anchor_col)}{row_cursor + 1}"
            chart_added = self._add_bar_chart_from_table(
                ws,
                row_cursor,
                start_col,
                row_cursor + len(inventory_rows),
                "产品库存金额Top",
                anchor,
            )
            if chart_added:
                added = True
                self._write_chart_note(ws, anchor_col, row_cursor, "图表说明：库存金额最高的产品，优先检查周转与备货策略。")
            row_cursor += advance(inventory_rows)

        risk_candidates = []
        for r in records:
            inventory = r.get("inventory")
            revenue = r.get("revenue")
            if inventory is None or inventory <= 0:
                continue
            ratio = inventory / revenue if revenue and revenue > 0 else None
            risk_score = ratio if ratio is not None else 1e9
            risk_candidates.append((risk_score, [r["product"], inventory, revenue, ratio]))
        risk_rows = [row for _, row in sorted(risk_candidates, key=lambda item: item[0], reverse=True)[:10]]
        if risk_rows:
            self._write_table(ws, row_cursor, start_col, ["产品", "期末库存金额", "销售收入", "库存收入比"], risk_rows)
            for rr in range(row_cursor + 1, row_cursor + 1 + len(risk_rows)):
                ws.cell(row=rr, column=start_col + 3).number_format = "0.00%"
            bar_cols = [start_col + idx for idx in (1, 2) if has_series_value(risk_rows, idx)]
            line_cols = [start_col + 3] if has_series_value(risk_rows, 3) else []
            anchor = f"{get_column_letter(anchor_col)}{row_cursor + 1}"
            if bar_cols and line_cols:
                chart_added = self._add_combo_chart(
                    ws,
                    start_col,
                    bar_cols,
                    line_cols,
                    row_cursor,
                    row_cursor + 1,
                    row_cursor + len(risk_rows),
                    "高库存低销售风险Top",
                    anchor,
                )
            else:
                chart_added = self._add_bar_chart_by_columns(
                    ws,
                    start_col,
                    bar_cols,
                    row_cursor,
                    row_cursor + 1,
                    row_cursor + len(risk_rows),
                    "高库存低销售风险Top",
                    anchor,
                )
            if chart_added:
                added = True
                self._write_chart_note(ws, anchor_col, row_cursor, "图表说明：库存金额高但销售收入低的产品存在滞销或备货过量风险。")
            row_cursor += advance(risk_rows)

        matrix_records = [
            r for r in records
            if r.get("revenue") is not None and r.get("margin") is not None
        ]
        matrix_records = sorted(matrix_records, key=lambda r: abs(r.get("revenue") or 0), reverse=True)[:30]
        if len(matrix_records) >= 2:
            matrix_rows = [[r["product"], r.get("revenue"), r.get("margin")] for r in matrix_records]
            self._write_table(ws, row_cursor, start_col, ["产品", "销售收入", "毛利率"], matrix_rows)
            for rr in range(row_cursor + 1, row_cursor + 1 + len(matrix_rows)):
                ws.cell(row=rr, column=start_col + 2).number_format = "0.00%"
            anchor = f"{get_column_letter(anchor_col)}{row_cursor + 1}"
            chart_added = self._add_scatter_chart(
                ws,
                start_col + 1,
                start_col + 2,
                row_cursor,
                row_cursor + 1,
                row_cursor + len(matrix_rows),
                "产品销售收入 vs 毛利率矩阵",
                anchor,
                x_title="销售收入",
                y_title="毛利率",
            )
            if chart_added:
                added = True
                self._write_chart_note(ws, anchor_col, row_cursor, "图表说明：识别高收入低毛利产品，辅助定价和成本复盘。")
        return added

    def _add_chart_wide_sheet(self, ws, preferred_labels, title, note):
        header_row = 1
        month_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if isinstance(header, str) and re.match(r"20\d{2}[./\-]\d{1,2}\s*月?", header):
                month_cols.append((col, header))
        if len(month_cols) < 2:
            return False
        label_rows = {}
        for r in range(header_row + 1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if isinstance(label, str) and label.strip():
                label_rows[label.strip()] = r
        selected = [label for label in preferred_labels if label in label_rows]
        if not selected:
            for label, row in label_rows.items():
                has_value = False
                for col, _ in month_cols:
                    if isinstance(ws.cell(row=row, column=col).value, (int, float)):
                        has_value = True
                        break
                if has_value:
                    selected.append(label)
                if len(selected) >= 3:
                    break
        if not selected:
            return False
        start_col = ws.max_column + 2
        headers = ["月份"] + selected
        rows = []
        for col, month_label in month_cols:
            row = [month_label]
            for label in selected:
                value = self._to_float(ws.cell(row=label_rows[label], column=col).value)
                row.append(value)
            rows.append(row)
        self._write_table(ws, 1, start_col, headers, rows)
        anchor_col = start_col + len(headers) + 2
        anchor = f"{get_column_letter(anchor_col)}2"
        added = self._add_line_chart_by_columns(
            ws,
            start_col,
            list(range(start_col + 1, start_col + 1 + len(selected))),
            1,
            2,
            1 + len(rows),
            title,
            anchor,
        )
        if added:
            self._write_chart_note(ws, anchor_col, 1, note)
        return added

    def _add_chart_balance_sheet(self, ws):
        preferred = ["流动资产", "货币资金", "应收账款", "存货"]
        return self._add_chart_wide_sheet(ws, preferred, "资产负债表趋势", "图表说明：关键资产项的月度变化。")

    def _add_chart_profit_sheet(self, ws):
        preferred = ["主营业务收入", "主营业务成本", "营业利润", "净利润"]
        return self._add_chart_wide_sheet(ws, preferred, "利润表趋势", "图表说明：主营业务收入/成本/利润的月度变化。")

    def _add_chart_ar_aging(self, ws):
        header_row = 1
        header_map = self._get_header_map(ws, header_row)
        name_col = header_map.get("往来单位名")
        value_col = header_map.get("总欠款")
        if not name_col or not value_col:
            return False
        items = []
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=name_col).value
            value = self._to_float(ws.cell(row=r, column=value_col).value)
            if not name or value is None:
                continue
            items.append((str(name), value))
        if not items:
            return False
        top_items = sorted(items, key=lambda x: abs(x[1]), reverse=True)[:10]
        start_col = ws.max_column + 2
        headers = ["客户", "总欠款"]
        rows = [[name, value] for name, value in top_items]
        self._write_table(ws, 1, start_col, headers, rows)
        anchor_col = start_col + len(headers) + 2
        anchor = f"{get_column_letter(anchor_col)}2"
        added = self._add_bar_chart_from_table(
            ws,
            1,
            start_col,
            1 + len(rows),
            "应收账款Top",
            anchor,
        )
        if added:
            self._write_chart_note(ws, anchor_col, 1, "图表说明：应收账款余额Top客户分布。")
        return added

    def _ensure_report_charts(self, wb):
        skip = {
            "目录",
            "计算说明",
            "图表数据源_隐藏",
            "仪表盘",
            "产品对比(动态图表)",
            "数据质量检查",
            "审计日志",
            "异常预警",
            # 以下页面以明细/说明为主，默认不强制图表。
            "年度经营指标",
            "年度利润表",
            "年度资产负债表",
            "费用异常明细",
            "费用分析",
            "多主体汇总",
            "币种汇总",
            "补货预警",
            "资金链预警",
        }
        for ws in wb.worksheets:
            if ws.title in skip:
                continue
            force_rebuild = ws.title in {"明细_销售与库存", "本量利分析", "经营指标"}
            if ws.title == "按产品汇总_含合计" and len(ws._charts) > 0:
                required_titles = (
                    "产品收入/毛利Top",
                    "产品库存占用与周转Top",
                    "产品收入 vs 毛利率矩阵",
                )
                if any(not self._worksheet_has_chart_title(ws, title) for title in required_titles):
                    if self._add_chart_product_summary(ws):
                        print(f"已自动补充图表: {ws.title}")
                continue
            if ws.title == "按品类汇总(按月)" and len(ws._charts) > 0:
                if not self._worksheet_has_chart_title(ws, "品类毛利润趋势"):
                    if self._add_chart_category_month(ws):
                        print(f"已自动补充图表: {ws.title}")
                continue
            if len(ws._charts) > 0 and not force_rebuild:
                continue
            if force_rebuild and len(ws._charts) > 0:
                ws._charts = []
            added = False
            if ws.title == "经营指标":
                added = self._add_chart_management_metrics(ws)
            elif ws.title == "费用对比":
                added = self._add_chart_expense_compare(ws)
            elif ws.title == "本量利分析":
                added = self._add_chart_cvp(ws)
            elif ws.title == "目标_预算":
                added = self._add_chart_budget(ws)
            elif ws.title == "同比经营分析":
                added = self._add_chart_compare_sheet(ws, "同比")
            elif ws.title == "环比经营分析":
                added = self._add_chart_compare_sheet(ws, "环比")
            elif ws.title == "费用明细":
                added = self._add_chart_expense_detail(ws)
            elif ws.title == "费用明细环比分析":
                added = self._add_chart_expense_mom(ws)
            elif ws.title == "按品类汇总(按月)":
                added = self._add_chart_category_month(ws)
            elif ws.title == "按产品汇总_含合计":
                added = self._add_chart_product_summary(ws)
            elif ws.title == "明细_销售与库存":
                added = self._add_chart_sales_inventory_detail(ws)
            elif ws.title == "资产负债表":
                added = self._add_chart_balance_sheet(ws)
            elif ws.title == "利润表":
                added = self._add_chart_profit_sheet(ws)
            elif ws.title == "应收账款账龄分析":
                added = self._add_chart_ar_aging(ws)
            if added:
                print(f"已自动生成图表: {ws.title}")
            else:
                print(f"未生成图表: {ws.title}")

    def _prepare_sheet(self, wb, name, insert_after=None):
        if name in wb.sheetnames:
            ws = wb[name]
            ws._charts = []
            for merged in list(ws.merged_cells.ranges):
                try:
                    ws.unmerge_cells(str(merged))
                except KeyError:
                    # Some templates have inconsistent merge maps; skip if unmerge fails.
                    continue
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    cell.value = None
        else:
            idx = None
            if insert_after in wb.sheetnames:
                idx = wb.sheetnames.index(insert_after) + 1
            ws = wb.create_sheet(name, idx)
        return ws

    def _write_sheet_redirect(self, wb, source_name, target_name, message=None):
        if source_name not in wb.sheetnames or target_name not in wb.sheetnames:
            return
        ws = self._prepare_sheet(wb, source_name, insert_after=target_name)
        text = message or f"本表已合并至「{target_name}」"
        cell = ws.cell(row=1, column=1)
        cell.value = text
        self._apply_hyperlink(cell, f"#'{target_name}'!A1")
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="0563C1", underline="single")

    def _delete_sheets_if_exist(self, wb, sheet_names):
        for name in sheet_names:
            if name in wb.sheetnames:
                ws = wb[name]
                wb.remove(ws)

    def _remove_directory_entries(self, wb, sheet_names):
        if "目录" not in wb.sheetnames or not sheet_names:
            return
        ws = wb["目录"]
        targets = {str(name).strip() for name in sheet_names if name}
        for r in range(ws.max_row, 0, -1):
            name = ws.cell(row=r, column=1).value
            if name is None:
                continue
            if str(name).strip() in targets:
                ws.delete_rows(r, 1)

    def _write_rows(self, ws, start_row, start_col, rows):
        for r_idx, row in enumerate(rows, start=0):
            for c_idx, value in enumerate(row, start=0):
                cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = self._sanitize_excel_text(value)

    def _write_data_quality_sheet(self, wb):
        ws = self._prepare_sheet(wb, "数据质量检查", insert_after="目录")
        headers = ["级别", "类别", "月份", "问题类型", "说明", "建议"]
        rows = []
        for item in self.data_quality_issues:
            rows.append([
                item.get("severity"),
                item.get("category"),
                item.get("period"),
                item.get("issue_type"),
                item.get("detail"),
                "请检查原始数据或字段映射",
            ])
        if not rows:
            rows = [["INFO", "系统", None, "无异常", "未发现明显数据质量问题", "无"]]
        self._write_table(ws, 1, 1, headers, rows)
        self._write_chart_note(ws, 1, len(rows) + 3, "说明：此表为自动化数据质量检查结果。")

    def _write_audit_log_sheet(self, wb):
        ws = self._prepare_sheet(wb, "审计日志", insert_after="数据质量检查")
        headers = ["时间", "操作"]
        rows = self.audit_logs or [(datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "无审计记录")]
        self._write_table(ws, 1, 1, headers, rows)
        self._write_chart_note(ws, 1, len(rows) + 3, "说明：此表记录报告生成过程关键步骤。")

    def _read_budget_targets(self, wb):
        targets = {}
        attain_green = 1.0
        attain_yellow = 0.9
        ratio_threshold = 0.2
        pp_threshold = 0.01
        days_threshold = 10
        if "目标_预算" not in wb.sheetnames:
            return targets, (attain_green, attain_yellow), (ratio_threshold, pp_threshold, days_threshold)

        ws = wb["目标_预算"]
        header_row = None
        for r in range(1, min(ws.max_row, 50) + 1):
            if str(ws.cell(row=r, column=1).value).strip() == "月份":
                header_row = r
                break
        if header_row:
            header_map = self._get_header_map(ws, header_row)

            def read_target_value(row_idx, header_name):
                col_idx = header_map.get(header_name)
                if not col_idx or col_idx < 1:
                    return None
                return self._to_float(ws.cell(row=row_idx, column=col_idx).value)

            for r in range(header_row + 1, ws.max_row + 1):
                label = ws.cell(row=r, column=1).value
                if not label:
                    continue
                month_key = self._label_to_month_key(label)
                if not month_key:
                    continue
                targets[month_key] = {
                    "revenue": read_target_value(r, "主营业务收入目标"),
                    "profit": read_target_value(r, "营业利润目标"),
                    "profit_rate": read_target_value(r, "营业利润率目标"),
                    "cost_rate": read_target_value(r, "成本率目标"),
                    "sales_rate": read_target_value(r, "销售费用率目标"),
                    "admin_rate": read_target_value(r, "管理费用率目标"),
                    "financial_rate": read_target_value(r, "财务费用率目标"),
                    "ar_balance": read_target_value(r, "应收账款余额目标"),
                    "inventory_end": read_target_value(r, "存货期末余额目标"),
                }

        for r in range(1, min(ws.max_row, 50) + 1):
            label = ws.cell(row=r, column=1).value
            if isinstance(label, str) and "达成率阈值" in label:
                attain_green = self._to_float(ws.cell(row=r + 1, column=2).value) or attain_green
                attain_yellow = self._to_float(ws.cell(row=r + 1, column=3).value) or attain_yellow
            if isinstance(label, str) and "异常阈值" in label:
                ratio_threshold = self._to_float(ws.cell(row=r + 1, column=2).value) or ratio_threshold
                pp_threshold = self._to_float(ws.cell(row=r + 1, column=3).value) or pp_threshold
                days_threshold = self._to_float(ws.cell(row=r + 1, column=4).value) or days_threshold

        return targets, (attain_green, attain_yellow), (ratio_threshold, pp_threshold, days_threshold)

    def _add_months(self, year, month, offset):
        total = (year * 12 + (month - 1)) + offset
        new_year = total // 12
        new_month = total % 12 + 1
        return new_year, new_month

    def _build_month_sequence(self, target_year, target_month, count):
        months = []
        for i in range(count):
            year, month = self._add_months(int(target_year), int(target_month), i)
            months.append(f"{year}-{month:02d}")
        return months

    def _normalize_balance_label(self, text):
        label = self._normalize_label(text)
        if not label:
            return ''
        label = re.sub(r'[（(].*?[）)]', '', label)
        return label.rstrip('：:')

    def _extract_balance_value_map(self, df, target_year=None, target_month=None):
        """从资产负债原始DataFrame中提取 科目名 -> 金额 的映射（支持左右两栏结构）。"""
        if df is None or df.empty:
            return {}

        work = df.copy().dropna(axis=1, how='all')
        if work.empty:
            return {}

        # 某些文件以首行作为“伪表头”，这里自动提升为列名
        if len(work) > 1 and any(str(c).startswith('Unnamed') for c in work.columns):
            first_row = ["" if pd.isna(v) else str(v).strip() for v in work.iloc[0].tolist()]
            if any('财务报表显示名' in v for v in first_row):
                seen = {}
                new_cols = []
                for idx, raw in enumerate(first_row):
                    base = raw or f"col_{idx}"
                    count = seen.get(base, 0)
                    col_name = f"{base}.{count}" if count else base
                    seen[base] = count + 1
                    new_cols.append(col_name)
                work = work.iloc[1:].copy()
                work.columns = new_cols

        label_cols = [c for c in work.columns if '财务报表显示名' in str(c)]
        if not label_cols:
            label_cols = [work.columns[0]]

        target_tokens = []
        if target_year and target_month:
            month_num = int(target_month)
            target_tokens = [
                f"{target_year}年{month_num}月",
                f"{target_year}年{month_num:02d}月",
                f"{target_year}/{month_num:02d}",
                f"{target_year}-{month_num:02d}",
            ]

        def _col_suffix(col_name):
            m = re.search(r'\.(\d+)$', str(col_name))
            return f".{m.group(1)}" if m else ""

        col_pairs = []
        for label_col in label_cols:
            suffix = _col_suffix(label_col)
            value_candidates = []
            for c in work.columns:
                c_str = str(c)
                if c == label_col:
                    continue
                if '财务报表显示名' in c_str or '科目编码' in c_str or '科目名' in c_str:
                    continue
                if suffix and not c_str.endswith(suffix):
                    continue
                if not suffix and re.search(r'\.\d+$', c_str):
                    continue
                value_candidates.append(c)

            value_col = None
            if target_tokens:
                for c in value_candidates:
                    c_str = str(c)
                    if any(token in c_str for token in target_tokens):
                        value_col = c
                        break

            if value_col is None:
                for c in value_candidates:
                    if pd.to_numeric(work[c], errors='coerce').notna().any():
                        value_col = c
                        break

            if value_col is not None:
                col_pairs.append((label_col, value_col))

        if not col_pairs:
            label_col = work.columns[0]
            numeric_cols = [c for c in work.columns[1:] if pd.to_numeric(work[c], errors='coerce').notna().any()]
            if numeric_cols:
                col_pairs.append((label_col, numeric_cols[0]))

        value_map = {}
        for label_col, value_col in col_pairs:
            for _, row in work.iterrows():
                key = self._normalize_balance_label(row.get(label_col))
                if not key:
                    continue
                val = self._to_float(row.get(value_col))
                if val is None:
                    continue
                value_map[key] = val
                value_map[f"{key}："] = val

        return value_map

    def _extract_balance_snapshot(self, df, month_key=None):
        if df is None or df.empty:
            return {}
        year = None
        month = None
        if month_key and re.match(r'^20\d{2}-\d{2}$', str(month_key)):
            year, month = month_key.split('-')
        balance_map = self._extract_balance_value_map(df, year, month)
        if not balance_map:
            return {}
        snapshot = {}
        key_map = {
            "cash": ["货币资金", "现金", "银行存款", "现金及现金等价物"],
            "ar": ["应收账款"],
            "inventory": ["存货"],
            "ap": ["应付账款"],
            "other_payable": ["其他应付款"],
            "fixed_assets": ["固定资产"],
            "short_debt": ["短期借款"],
            "long_debt": ["长期借款"],
        }
        for key, keywords in key_map.items():
            if key in snapshot:
                continue
            for label, value in balance_map.items():
                if value is None:
                    continue
                if any(k in label for k in keywords):
                    snapshot[key] = value
                    break
        return snapshot

    def _update_dashboard(self, wb, metrics_by_month, target_year, target_month):
        if "仪表盘" in wb.sheetnames:
            existing = wb["仪表盘"]
            if self._dashboard_has_template_formula_layout(existing):
                if self._extend_template_dashboard_metrics(existing):
                    self._log_audit("模板仪表盘已补充营业利润/营业利润率。")
                else:
                    self._log_audit("检测到模板仪表盘公式，跳过自动重建仪表盘。")
                if self._ensure_dashboard_financial_expense_rate(wb):
                    self._log_audit("模板仪表盘已补充财务费用率表格/图表。")
                self._repair_template_dashboard_formulas(
                    existing,
                    budget_header_row=self._dashboard_budget_header_row(wb),
                )
                return
        ws = self._prepare_sheet(wb, "仪表盘", insert_after="目录")
        
        # 1. Key Metrics Table (Current Month)
        key_metrics_headers = ["指标", "本月数值", "环比", "同比"]
        current_key = f"{target_year}-{int(target_month):02d}"
        
        # Calculate Previous Month Key
        prev_month_dt = datetime(int(target_year), int(target_month), 1) - pd.Timedelta(days=1)
        prev_month_key = f"{prev_month_dt.year}-{prev_month_dt.month:02d}"
        
        # Calculate Same Month Last Year Key
        last_year_key = f"{int(target_year)-1}-{int(target_month):02d}"
        
        current_data = metrics_by_month.get(current_key, {})
        prev_data = metrics_by_month.get(prev_month_key, {})
        last_year_data = metrics_by_month.get(last_year_key, {})
        
        metrics_list = [
            ("主营业务收入", "revenue"),
            ("主营业务成本", "cost"),
            ("毛利润", "gross_profit"), # Special handling
            ("营业利润", "operating_profit"),
            ("净利润", "net_profit"),
        ]
        
        rows = []
        for label, key in metrics_list:
            curr = current_data.get(key)
            prev = prev_data.get(key)
            last = last_year_data.get(key)
            
            # Special handling for gross profit if not pre-calculated
            if key == "gross_profit":
                if curr is None and current_data.get("revenue") is not None and current_data.get("cost") is not None:
                    curr = current_data["revenue"] - current_data["cost"]
                if prev is None and prev_data.get("revenue") is not None and prev_data.get("cost") is not None:
                    prev = prev_data["revenue"] - prev_data["cost"]
                if last is None and last_year_data.get("revenue") is not None and last_year_data.get("cost") is not None:
                    last = last_year_data["revenue"] - last_year_data["cost"]

            mom = (curr - prev) / prev if (curr is not None and prev) else None
            yoy = (curr - last) / last if (curr is not None and last) else None
            
            rows.append([label, curr, mom, yoy])
            
        self._write_table(ws, 2, 2, key_metrics_headers, rows)
        self._apply_header_style(ws, 2)
        
        # 2. Add Trend Chart (Last 12 Months Revenue & Profit)
        start_col = 8
        trend_headers = ["月份", "收入", "利润"]
        trend_rows = []
        
        # Sort months
        sorted_keys = sorted(metrics_by_month.keys())
        # Filter for last 12 months ending at target date
        relevant_keys = [k for k in sorted_keys if k <= current_key][-12:]
        
        for k in relevant_keys:
            d = metrics_by_month.get(k, {})
            trend_rows.append([k, d.get("revenue"), d.get("operating_profit")])
            
        self._write_table(ws, 2, start_col, trend_headers, trend_rows)
        
        anchor = f"{get_column_letter(start_col + 4)}3"
        self._add_combo_chart(
            ws, 
            start_col, # Cat col (Month)
            [start_col+1], # Bar cols (Revenue)
            [start_col+2], # Line cols (Profit)
            2, # Header row
            3, # Data start
            2 + len(trend_rows), # Data end
            "近12个月营收与利润趋势",
            anchor
        )

    def _update_cashflow_sheet(self, wb, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "现金流量表(估算)", insert_after="利润表")
        headers = [
            "月份",
            "经营利润",
            "应收账款变动",
            "存货变动",
            "应付账款变动",
            "经营活动现金流(估算)",
            "投资现金流(估算)",
            "融资现金流(估算)",
            "净现金流",
            "期末现金余额",
        ]
        rows = []
        month_keys = self._filter_month_keys(
            set(self.data['profit'].keys()) | set(self.data['asset'].keys()),
            target_year,
            target_month,
            year_scope,
        )
        prev_snapshot = None
        for m_key in month_keys:
            profit_metrics = self._extract_profit_metrics(self.data['profit'].get(m_key), m_key)
            snapshot = self._extract_balance_snapshot(self.data['asset'].get(m_key), m_key)
            profit = profit_metrics.get("net_profit") or profit_metrics.get("operating_profit")
            delta_ar = None
            delta_inv = None
            delta_ap = None
            if prev_snapshot:
                if snapshot.get("ar") is not None and prev_snapshot.get("ar") is not None:
                    delta_ar = snapshot.get("ar") - prev_snapshot.get("ar")
                if snapshot.get("inventory") is not None and prev_snapshot.get("inventory") is not None:
                    delta_inv = snapshot.get("inventory") - prev_snapshot.get("inventory")
                if snapshot.get("ap") is not None and prev_snapshot.get("ap") is not None:
                    delta_ap = snapshot.get("ap") - prev_snapshot.get("ap")
            cfo = None
            if profit is not None:
                cfo = profit
                if delta_ar is not None:
                    cfo -= delta_ar
                if delta_inv is not None:
                    cfo -= delta_inv
                if delta_ap is not None:
                    cfo += delta_ap

            cfi = None
            if prev_snapshot and snapshot.get("fixed_assets") is not None and prev_snapshot.get("fixed_assets") is not None:
                cfi = -(snapshot.get("fixed_assets") - prev_snapshot.get("fixed_assets"))
            cff = None
            if prev_snapshot and snapshot.get("short_debt") is not None and prev_snapshot.get("short_debt") is not None:
                cff = (snapshot.get("short_debt") - prev_snapshot.get("short_debt"))
            if prev_snapshot and snapshot.get("long_debt") is not None and prev_snapshot.get("long_debt") is not None:
                cff = (cff or 0) + (snapshot.get("long_debt") - prev_snapshot.get("long_debt"))

            net_cash = None
            if cfo is not None or cfi is not None or cff is not None:
                net_cash = (cfo or 0) + (cfi or 0) + (cff or 0)
            rows.append([
                self._month_key_to_label(m_key),
                profit,
                delta_ar,
                delta_inv,
                delta_ap,
                cfo,
                cfi,
                cff,
                net_cash,
                snapshot.get("cash"),
            ])
            prev_snapshot = snapshot

        if not rows:
            rows = [[None] * len(headers)]
        self._write_table(ws, 1, 1, headers, rows)
        anchor = f"{get_column_letter(len(headers) + 2)}2"
        self._add_line_chart_by_columns(ws, 1, [6, 9], 1, 2, 1 + len(rows), "现金流趋势", anchor)
        self._write_chart_note(ws, len(headers) + 2, 1, "图表说明：经营活动现金流与净现金流的估算趋势。")

        self._reorder_month_rows_desc(ws)

    def _pick_ledger_column(self, df, preferred, exclude_keywords=None):
        if df is None or df.empty:
            return None
        for name in preferred:
            for col in df.columns:
                if str(col).strip() == name:
                    return col
        return self._pick_first_column(df, preferred, exclude_keywords)

    def _pick_ledger_amount_column(self, df, keyword):
        if df is None or df.empty:
            return None
        for col in df.columns:
            if str(col).strip() == keyword:
                return col
        for col in df.columns:
            name = str(col).strip()
            if keyword in name and "外币" not in name:
                return col
        for col in df.columns:
            if keyword in str(col):
                return col
        return None

    def _last_numeric_or_none(self, series):
        if series is None:
            return None
        values = pd.to_numeric(series, errors="coerce").dropna()
        if values.empty:
            return None
        return float(values.iloc[-1])

    def _prepare_ledger_detail_df(self, df):
        if df is None or df.empty:
            return pd.DataFrame()

        work = df.copy()
        work.columns = [str(c).strip().rstrip('\t') for c in work.columns]
        for col in work.columns:
            if work[col].dtype == object:
                work[col] = work[col].apply(lambda v: v.strip().replace('\t', '') if isinstance(v, str) else v)

        if "ParsedDate" in work.columns:
            parsed = pd.to_datetime(work["ParsedDate"], errors="coerce")
        else:
            _, parsed = self._pick_best_date_column(
                work, ["日期", "date", "业务日期", "单据日期", "日期-号码"]
            )
        if parsed is None:
            return pd.DataFrame()

        work["ParsedDate"] = parsed
        work = work.dropna(subset=["ParsedDate"]).copy()
        if work.empty:
            return work
        work["MonthStr"] = work["ParsedDate"].dt.strftime("%Y-%m")

        debit_col = self._pick_ledger_amount_column(work, "借方金额") or self._pick_ledger_amount_column(work, "借方")
        credit_col = self._pick_ledger_amount_column(work, "贷方金额") or self._pick_ledger_amount_column(work, "贷方")
        balance_col = self._pick_ledger_amount_column(work, "余额")
        summary_col = self._pick_ledger_column(work, ["摘要", "说明", "用途"])
        party_col = self._pick_ledger_column(
            work,
            ["往来单位名", "对应往来单位名", "供应商名", "客户名", "供应商", "客户", "单位名", "收货公司"],
        )
        subject_col = self._pick_ledger_column(work, ["相对科目编码名", "科目名", "对方科目", "账户名"])

        work["Debit"] = pd.to_numeric(work[debit_col], errors="coerce") if debit_col else 0.0
        work["Credit"] = pd.to_numeric(work[credit_col], errors="coerce") if credit_col else 0.0
        work["Balance"] = pd.to_numeric(work[balance_col], errors="coerce") if balance_col else None
        work["SummaryText"] = work[summary_col].astype(str).str.strip() if summary_col else ""
        work["Counterparty"] = work[party_col].astype(str).str.strip() if party_col else ""
        work["SubjectName"] = work[subject_col].astype(str).str.strip() if subject_col else ""
        work["Counterparty"] = work["Counterparty"].replace({"nan": "", "None": ""})
        work["SubjectName"] = work["SubjectName"].replace({"nan": "", "None": ""})
        work["SummaryText"] = work["SummaryText"].replace({"nan": "", "None": ""})
        return work

    def _ledger_group_label(self, row, prefer_counterparty=True):
        labels = []
        if prefer_counterparty:
            labels = [row.get("Counterparty"), row.get("SubjectName"), row.get("SummaryText")]
        else:
            labels = [row.get("SubjectName"), row.get("Counterparty"), row.get("SummaryText")]
        for label in labels:
            if label is not None and str(label).strip() and str(label).strip().lower() != "nan":
                return str(label).strip()
        return "未识别"

    def _update_ap_analysis_sheet(self, wb, metrics_by_month, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "应付账款分析", insert_after="资金链预警")
        df = self._prepare_ledger_detail_df(self.ap_detail_df)
        if df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["未检测到应付账款明细文件"]])
            return

        df = self._filter_df_by_scope(df, target_year, target_month, year_scope)
        if df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["当前口径下无应付账款明细"]])
            return

        month_keys = self._filter_month_keys(
            set(metrics_by_month.keys()) | set(df["MonthStr"].dropna().astype(str).unique()),
            target_year,
            target_month,
            year_scope,
        )
        headers = ["月份", "借方发生额", "贷方发生额", "应付净增加", "明细期末余额", "余额绝对值", "报表应付账款余额", "明细覆盖率"]
        rows = []
        for m_key in month_keys:
            group = df[df["MonthStr"] == m_key].sort_values("ParsedDate")
            debit = self._sum_numeric_or_none(group.get("Debit")) if not group.empty else None
            credit = self._sum_numeric_or_none(group.get("Credit")) if not group.empty else None
            net_increase = ((credit or 0) - (debit or 0)) if (debit is not None or credit is not None) else None
            ending = self._last_numeric_or_none(group.get("Balance")) if not group.empty else None
            ending_abs = abs(ending) if ending is not None else None
            statement_ap = (metrics_by_month.get(m_key) or {}).get("ap_balance")
            coverage = (ending_abs / statement_ap) if (ending_abs is not None and statement_ap not in (None, 0)) else None
            rows.append([
                self._month_key_to_label(m_key),
                debit,
                credit,
                net_increase,
                ending,
                ending_abs,
                statement_ap,
                coverage,
            ])

        self._write_table(ws, 1, 1, headers, rows or [[None] * len(headers)])
        if rows:
            self._add_line_chart_by_columns(ws, 1, [2, 3, 6], 1, 2, 1 + len(rows), "应付账款发生额与余额", f"{get_column_letter(len(headers) + 2)}2")

        summary = []
        df = df.copy()
        df["_GroupLabel"] = df.apply(lambda row: self._ledger_group_label(row, prefer_counterparty=True), axis=1)
        for label, group in df.groupby("_GroupLabel"):
            debit = self._sum_numeric_or_none(group.get("Debit"))
            credit = self._sum_numeric_or_none(group.get("Credit"))
            net_increase = ((credit or 0) - (debit or 0)) if (debit is not None or credit is not None) else None
            ending = self._last_numeric_or_none(group.sort_values("ParsedDate").get("Balance"))
            latest_date = group["ParsedDate"].max()
            summary.append([
                label,
                int(len(group)),
                debit,
                credit,
                net_increase,
                ending,
                abs(ending) if ending is not None else None,
                latest_date.strftime("%Y-%m-%d") if pd.notna(latest_date) else None,
            ])
        summary.sort(key=lambda r: abs(r[6] if r[6] is not None else (r[4] or 0)), reverse=True)
        start_row = len(rows) + 4
        summary_headers = ["供应商/对象", "交易笔数", "借方发生额", "贷方发生额", "应付净增加", "期末余额", "余额绝对值", "最近交易日"]
        self._write_table(ws, start_row, 1, summary_headers, summary[:50] or [[None] * len(summary_headers)])
        if summary:
            chart_col = len(summary_headers) + 2
            chart_rows = [[r[0], r[6]] for r in summary[:10]]
            self._write_table(ws, start_row, chart_col, ["供应商/对象", "余额绝对值"], chart_rows)
            self._add_bar_chart_from_table(ws, start_row, chart_col, start_row + len(chart_rows), "应付余额Top", f"{get_column_letter(chart_col + 3)}{start_row + 1}")
        self._write_chart_note(ws, 1, start_row + min(len(summary), 50) + 3, "说明：明细覆盖率=已导入应付明细余额绝对值/资产负债表应付账款余额；未导入全部供应商明细时该比例可能低于100%。")

    def _update_cash_balance_analysis_sheet(self, wb, metrics_by_month, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "货币资金分析", insert_after="现金流量表(估算)")
        df = self._prepare_ledger_detail_df(self.cash_detail_df)
        month_keys = set(metrics_by_month.keys())
        if not df.empty:
            df = self._filter_df_by_scope(df, target_year, target_month, year_scope)
            month_keys.update(df["MonthStr"].dropna().astype(str).unique())
        month_keys = self._filter_month_keys(month_keys, target_year, target_month, year_scope)

        if not month_keys:
            self._write_table(ws, 1, 1, ["提示"], [["未检测到货币资金明细或资产负债表货币资金数据"]])
            return

        headers = ["月份", "明细借方/流入", "明细贷方/流出", "明细净流入", "明细期末余额", "报表货币资金", "明细覆盖率", "差异"]
        rows = []
        for m_key in month_keys:
            group = df[df["MonthStr"] == m_key].sort_values("ParsedDate") if not df.empty else pd.DataFrame()
            debit = self._sum_numeric_or_none(group.get("Debit")) if not group.empty else None
            credit = self._sum_numeric_or_none(group.get("Credit")) if not group.empty else None
            net = ((debit or 0) - (credit or 0)) if (debit is not None or credit is not None) else None
            ending = self._last_numeric_or_none(group.get("Balance")) if not group.empty else None
            statement_cash = (metrics_by_month.get(m_key) or {}).get("cash")
            coverage = (ending / statement_cash) if (ending is not None and statement_cash not in (None, 0)) else None
            diff = (statement_cash - ending) if (statement_cash is not None and ending is not None) else None
            rows.append([
                self._month_key_to_label(m_key),
                debit,
                credit,
                net,
                ending,
                statement_cash,
                coverage,
                diff,
            ])

        self._write_table(ws, 1, 1, headers, rows)
        self._add_line_chart_by_columns(ws, 1, [4, 5, 6], 1, 2, 1 + len(rows), "货币资金余额与净流入", f"{get_column_letter(len(headers) + 2)}2")

        if not df.empty:
            df = df.copy()
            df["_GroupLabel"] = df.apply(lambda row: self._ledger_group_label(row, prefer_counterparty=False), axis=1)
            groups = []
            for label, group in df.groupby("_GroupLabel"):
                inflow = self._sum_numeric_or_none(group.get("Debit"))
                outflow = self._sum_numeric_or_none(group.get("Credit"))
                net = ((inflow or 0) - (outflow or 0)) if (inflow is not None or outflow is not None) else None
                groups.append([label, int(len(group)), inflow, outflow, net])
            groups.sort(key=lambda r: abs(r[3] if r[3] is not None else (r[4] or 0)), reverse=True)
            start_row = len(rows) + 4
            group_headers = ["科目/对象", "交易笔数", "流入", "流出", "净流入"]
            self._write_table(ws, start_row, 1, group_headers, groups[:50] or [[None] * len(group_headers)])
            if groups:
                chart_col = len(group_headers) + 2
                chart_rows = [[r[0], r[3]] for r in groups[:10]]
                self._write_table(ws, start_row, chart_col, ["科目/对象", "流出"], chart_rows)
                self._add_bar_chart_from_table(ws, start_row, chart_col, start_row + len(chart_rows), "货币资金流出Top", f"{get_column_letter(chart_col + 3)}{start_row + 1}")
            self._write_chart_note(ws, 1, start_row + min(len(groups), 50) + 3, "说明：货币资金明细覆盖率=已导入银行/现金明细期末余额/资产负债表货币资金；只导入部分账户时该比例用于识别覆盖缺口。")
        else:
            self._write_chart_note(ws, 1, len(rows) + 3, "说明：未导入银行存款/现金明细，本页仅展示资产负债表货币资金余额。")

    def _update_budget_variance_sheet(self, wb, metrics_by_month, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "预算执行与偏差", insert_after="目标_预算")
        targets, (attain_green, attain_yellow), _ = self._read_budget_targets(wb)
        headers = ["月份", "指标", "目标值", "实际值", "偏差", "偏差率", "达成率", "预警级别"]
        rows = []
        month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
        
        variance_data = []
        current_month_key = f"{target_year}-{int(target_month):02d}"

        for m_key in month_keys:
            data = metrics_by_month.get(m_key, {})
            revenue = data.get("revenue")
            profit = data.get("operating_profit")
            profit_rate = profit / revenue if (profit is not None and revenue not in (None, 0)) else None
            cost_rate = data.get("cost_rate")
            sales_expense = data.get("sales_expense")
            admin_expense = data.get("admin_expense")
            financial_expense = data.get("financial_expense")
            sales_rate = sales_expense / revenue if (sales_expense is not None and revenue not in (None, 0)) else None
            admin_rate = admin_expense / revenue if (admin_expense is not None and revenue not in (None, 0)) else None
            financial_rate = financial_expense / revenue if (financial_expense is not None and revenue not in (None, 0)) else None
            
            if m_key == current_month_key:
                tgt_rev = targets.get(m_key, {}).get("revenue")
                tgt_prof = targets.get(m_key, {}).get("profit")
                if tgt_rev is not None and tgt_prof is not None:
                    rev_act = revenue or 0
                    total_var = (profit or 0) - tgt_prof
                    rev_var = rev_act - tgt_rev
                    cost_exp_var = total_var - rev_var
                    variance_data = [
                        ["项目", "金额"],
                        ["收入影响", rev_var],
                        ["成本/费用影响", cost_exp_var],
                        ["总利润偏差", total_var]
                    ]

            metrics = [
                ("主营业务收入", "revenue", revenue),
                ("营业利润", "profit", profit),
                ("营业利润率", "profit_rate", profit_rate),
                ("成本率", "cost_rate", cost_rate),
                ("销售费用率", "sales_rate", sales_rate),
                ("管理费用率", "admin_rate", admin_rate),
                ("财务费用率", "financial_rate", financial_rate),
                ("应收账款余额", "ar_balance", data.get("ar_balance")),
                ("存货期末余额", "inventory_end", data.get("inventory_end")),
            ]
            for label, key, actual in metrics:
                target = targets.get(m_key, {}).get(key)
                if actual is None and target is None:
                    continue
                variance = None
                variance_rate = None
                attain = None
                if actual is not None and target:
                    variance = actual - target
                    variance_rate = variance / target if target else None
                    attain = actual / target if target else None
                level = "N/A"
                if attain is not None:
                    if attain >= attain_green:
                        level = "绿色"
                    elif attain >= attain_yellow:
                        level = "黄色"
                    else:
                        level = "红色"
                rows.append([
                    self._month_key_to_label(m_key),
                    label,
                    target,
                    actual,
                    variance,
                    variance_rate,
                    attain,
                    level,
                ])

        if not rows:
            rows = [[None] * len(headers)]
        self._write_table(ws, 1, 1, headers, rows)

        start_col = 10
        chart_rows = []
        for m_key in month_keys:
            t = targets.get(m_key, {}).get("revenue")
            a = metrics_by_month.get(m_key, {}).get("revenue")
            if t is None and a is None:
                continue
            chart_rows.append([self._month_key_to_label(m_key), t, a])
        if chart_rows:
            self._write_table(ws, 1, start_col, ["月份", "收入目标", "收入实际"], chart_rows)
            anchor = f"{get_column_letter(start_col + 4)}2"
            self._add_line_chart_by_columns(
                ws,
                start_col,
                [start_col + 1, start_col + 2],
                1,
                2,
                1 + len(chart_rows),
                "收入目标 vs 实际",
                anchor,
            )
            self._add_detailed_analysis_box(ws, start_col + 4, 14, "收入趋势图", [
                "折线对比了每月的主营业务收入目标(Target)与实际达成(Actual)。",
                "若实际线长期低于目标线，需检查市场需求或销售策略。",
                "趋势的一致性反映了业务的可预测性。"
            ])
            
        if variance_data:
            v_start_col = start_col
            v_start_row = 24
            self._write_table(ws, v_start_row, v_start_col, ["项目", "金额"], variance_data[1:])
            
            anchor_v = f"{get_column_letter(v_start_col + 4)}{v_start_row}"
            chart_v = BarChart()
            chart_v.title = "本月利润偏差归因 (收入 vs 成本费用)"
            chart_v.y_axis.title = "金额"
            data_v = Reference(ws, min_col=v_start_col+1, min_row=v_start_row, max_row=v_start_row+3)
            cats_v = Reference(ws, min_col=v_start_col, min_row=v_start_row+1, max_row=v_start_row+3)
            chart_v.add_data(data_v, titles_from_data=True)
            chart_v.set_categories(cats_v)
            chart_v.height = 10
            chart_v.width = 15
            ws.add_chart(chart_v, anchor_v)
            
            self._add_detailed_analysis_box(ws, v_start_col + 4, v_start_row + 12, "利润偏差归因图", [
                "展示了导致本月利润偏离预算的主要驱动因素。",
                "收入影响：(实际收入-目标收入)，正值代表增收带来的利润增加。",
                "成本/费用影响：(目标成本-实际成本)，正值代表节约带来的利润增加。",
                "关注负值(红色)条柱，这是利润未达标的主要缺口。"
            ])

        self._reorder_month_rows_desc(ws, max_col=len(headers))

    def _update_product_contribution_sheet(self, wb, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "产品贡献毛利", insert_after="按产品汇总_含合计")
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无销售数据"]])
            return
        df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, year_scope)
        if df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无有效销售数据"]])
            return
        df['Revenue'] = self._extract_sales_revenue(df)
        df['Qty'] = pd.to_numeric(df.get('数量'), errors='coerce')
        df = self._attach_sales_cost(df, target_year, target_month, year_scope)
        df = df.dropna(subset=['品目编码'])
        if df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["品目编码缺失"]])
            return
        total_revenue = self._sum_numeric_or_none(df.get('Revenue'))
        total_cost = self._sum_numeric_or_none(df.get('Cost'))
        total_profit = (total_revenue - total_cost) if (total_revenue is not None and total_cost is not None) else None
        inventory_map = {}
        cost_month_keys = self._filter_month_keys(self.data.get('cost', {}).keys(), target_year, target_month, year_scope)
        if cost_month_keys:
            last_cost_key = cost_month_keys[-1]
            inventory_map = self._build_cost_inventory_map(self.data['cost'].get(last_cost_key), last_cost_key)

        def product_risk_label(qty, revenue, profit, margin, inv_qty, inv_amt):
            if profit is not None and profit < 0:
                return "负毛利"
            if inv_qty not in (None, 0) and (qty is None or qty == 0):
                return "有库存无销量"
            inventory_revenue_ratio = inv_amt / revenue if (inv_amt is not None and revenue not in (None, 0)) else None
            if inventory_revenue_ratio is not None and inventory_revenue_ratio >= 1 and (margin is None or margin < 0.1):
                return "高库存低毛利"
            if inventory_revenue_ratio is not None and inventory_revenue_ratio >= 2:
                return "高库存"
            if margin is not None and margin < 0.05:
                return "低毛利"
            return "正常"

        rows = []
        for code, group in df.groupby('品目编码'):
            metrics = self._calc_sales_metrics_from_group(group)
            revenue = metrics.get('revenue')
            cost = metrics.get('cost')
            qty = metrics.get('qty')
            profit = metrics.get('profit')
            margin = metrics.get('margin')
            name = None
            if '品目名' in group and not group.get('品目名').dropna().empty:
                name = group.get('品目名').dropna().iloc[0]
            category = None
            if '品目组合1名' in group and not group.get('品目组合1名').dropna().empty:
                category = group.get('品目组合1名').dropna().iloc[0]
            avg_price = revenue / qty if (revenue is not None and qty not in (None, 0)) else None
            avg_cost = cost / qty if (cost is not None and qty not in (None, 0)) else None
            code_text = str(code).strip()
            inv_snapshot = inventory_map.get(code_text, {})
            inv_qty = inv_snapshot.get('qty_end')
            inv_amt = inv_snapshot.get('amt_end')
            inventory_revenue_ratio = inv_amt / revenue if (inv_amt is not None and revenue not in (None, 0)) else None
            risk_label = product_risk_label(qty, revenue, profit, margin, inv_qty, inv_amt)
            rows.append([
                code_text,
                name,
                category,
                qty,
                revenue,
                cost,
                profit,
                margin,
                (revenue / total_revenue) if (revenue is not None and total_revenue not in (None, 0)) else None,
                (profit / total_profit) if (profit is not None and total_profit not in (None, 0)) else None,
                avg_price,
                avg_cost,
                inv_qty,
                inv_amt,
                inventory_revenue_ratio,
                risk_label,
            ])

        headers = [
            "品目编码", "产品名称", "品类", "销量", "销售收入", "销售成本", "毛利润",
            "毛利率", "收入占比", "利润占比", "平均售价", "平均成本",
            "期末库存数量", "期末库存金额", "库存收入比", "风险标签"
        ]
        rows = sorted(rows, key=lambda x: x[6] if x[6] is not None else 0, reverse=True)
        self._write_table(ws, 1, 1, headers, rows)

        top_rows = rows[:10]
        if top_rows:
            start_col = len(headers) + 2
            chart_rows = [[r[1] or r[0], r[6]] for r in top_rows]
            self._write_table(ws, 1, start_col, ["产品", "毛利润"], chart_rows)
            anchor = f"{get_column_letter(start_col + 3)}2"
            self._add_bar_chart_from_table(ws, 1, start_col, 1 + len(chart_rows), "产品毛利润Top", anchor)

            low_margin_rows = []
            for r in sorted(rows, key=lambda x: x[6] if x[6] is not None else 0):
                if r[6] is None:
                    continue
                if r[6] < 0 or (r[7] is not None and r[7] < 0.05):
                    low_margin_rows.append([r[1] or r[0], r[6]])
                if len(low_margin_rows) >= 10:
                    break
            if low_margin_rows:
                low_start_col = start_col
                low_start_row = 14
                self._write_table(ws, low_start_row, low_start_col, ["低毛利/负毛利产品", "毛利润"], low_margin_rows)
                self._add_bar_chart_from_table(
                    ws,
                    low_start_row,
                    low_start_col,
                    low_start_row + len(low_margin_rows),
                    "低毛利/负毛利产品Top",
                    f"{get_column_letter(low_start_col + 3)}{low_start_row}",
                )

            inventory_risk_rows = []
            for r in sorted(rows, key=lambda x: abs(x[13] or 0), reverse=True):
                if r[15] == "正常" or r[13] is None:
                    continue
                inventory_risk_rows.append([r[1] or r[0], r[13]])
                if len(inventory_risk_rows) >= 10:
                    break
            if inventory_risk_rows:
                risk_start_col = start_col
                risk_start_row = 27
                self._write_table(ws, risk_start_row, risk_start_col, ["库存风险产品", "期末库存金额"], inventory_risk_rows)
                self._add_bar_chart_from_table(
                    ws,
                    risk_start_row,
                    risk_start_col,
                    risk_start_row + len(inventory_risk_rows),
                    "高库存风险产品Top",
                    f"{get_column_letter(risk_start_col + 3)}{risk_start_row}",
                )

            # Add Product Portfolio Matrix (Scatter)
            anchor_scatter = f"{get_column_letter(start_col + 8)}2"
            self._add_scatter_chart(
                ws,
                5, # X: Revenue
                8, # Y: Margin %
                1,
                2,
                1 + len(rows),
                "产品矩阵：收入 vs 毛利率",
                anchor_scatter,
                x_title="销售收入",
                y_title="毛利率"
            )
            if inventory_map:
                self._add_scatter_chart(
                    ws,
                    14, # X: Ending inventory amount
                    8,  # Y: Margin %
                    1,
                    2,
                    1 + len(rows),
                    "产品矩阵：库存金额 vs 毛利率",
                    f"{get_column_letter(start_col + 8)}20",
                    x_title="期末库存金额",
                    y_title="毛利率",
                )

            self._write_chart_note(ws, start_col + 3, 1, "图表说明：产品毛利、低毛利、库存风险和产品矩阵联动分析。")

    def _update_category_contribution_sheet(self, wb, target_year, target_month, year_scope=None):
        merged_into_month_sheet = "按品类汇总(按月)" in wb.sheetnames
        ws = wb["按品类汇总(按月)"] if merged_into_month_sheet else self._prepare_sheet(wb, "品类贡献毛利", insert_after="按品类汇总(按月)")

        merge_header = "品类贡献分析(合并视图)"
        merge_start_col = None
        if merged_into_month_sheet:
            for col in range(1, ws.max_column + 1):
                if str(ws.cell(row=1, column=col).value).strip() == merge_header:
                    merge_start_col = col
                    break
            if merge_start_col is None:
                merge_start_col = max(ws.max_column + 2, 70)

            block_width = 14
            clear_end_row = max(ws.max_row, 300)
            for r in range(1, clear_end_row + 1):
                for c in range(merge_start_col, merge_start_col + block_width):
                    self._safe_set_cell_value(ws, r, c, None)

            ws.cell(row=1, column=merge_start_col).value = merge_header
            ws.cell(row=2, column=merge_start_col).value = "与【按品类汇总(按月)】合并展示的品类贡献分析区"

        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            if merged_into_month_sheet:
                self._write_table(ws, 4, merge_start_col, ["提示"], [["无销售数据"]])
            else:
                self._write_table(ws, 1, 1, ["提示"], [["无销售数据"]])
            return
        df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, year_scope)
        if df.empty:
            if merged_into_month_sheet:
                self._write_table(ws, 4, merge_start_col, ["提示"], [["无有效销售数据"]])
            else:
                self._write_table(ws, 1, 1, ["提示"], [["无有效销售数据"]])
            return
        if '品目组合1名' not in df.columns:
            if merged_into_month_sheet:
                self._write_table(ws, 4, merge_start_col, ["提示"], [["缺少品类字段"]])
            else:
                self._write_table(ws, 1, 1, ["提示"], [["缺少品类字段"]])
            return
        df['Revenue'] = self._extract_sales_revenue(df)
        df['Qty'] = pd.to_numeric(df.get('数量'), errors='coerce')
        df = self._attach_sales_cost(df, target_year, target_month, year_scope)
        total_revenue = self._sum_numeric_or_none(df.get('Revenue'))
        total_cost = self._sum_numeric_or_none(df.get('Cost'))
        total_profit = (total_revenue - total_cost) if (total_revenue is not None and total_cost is not None) else None
        rows = []
        for category, group in df.groupby('品目组合1名'):
            metrics = self._calc_sales_metrics_from_group(group)
            revenue = metrics.get('revenue')
            cost = metrics.get('cost')
            qty = metrics.get('qty')
            profit = metrics.get('profit')
            margin = metrics.get('margin')
            avg_price = revenue / qty if (revenue is not None and qty not in (None, 0)) else None
            rows.append([
                category,
                qty,
                revenue,
                cost,
                profit,
                margin,
                (revenue / total_revenue) if (revenue is not None and total_revenue not in (None, 0)) else None,
                (profit / total_profit) if (profit is not None and total_profit not in (None, 0)) else None,
                avg_price,
            ])
        headers = ["品类", "销量", "销售收入", "销售成本", "毛利润", "毛利率", "收入占比", "利润占比", "平均售价"]
        rows = sorted(rows, key=lambda x: x[4] if x[4] is not None else 0, reverse=True)

        if merged_into_month_sheet:
            table_row = 4
            self._write_table(ws, table_row, merge_start_col, headers, rows)
            top_rows = rows[:10]
            helper_rows_used = len(top_rows) if top_rows else 0
            if top_rows:
                start_col = merge_start_col + len(headers) + 2
                chart_rows = [[r[0], r[4]] for r in top_rows]
                self._write_table(ws, table_row, start_col, ["品类", "毛利润"], chart_rows)
                self._add_bar_chart_from_table(
                    ws,
                    table_row,
                    start_col,
                    table_row + len(chart_rows),
                    "品类毛利润Top",
                    f"{get_column_letter(start_col + 3)}{table_row}",
                )
                share_rows = [[r[0], r[6]] for r in top_rows if r[6] is not None]
                if share_rows:
                    share_row = table_row + len(chart_rows) + 3
                    self._write_table(ws, share_row, start_col, ["品类", "收入占比"], share_rows)
                    self._add_doughnut_chart(
                        ws,
                        start_col,
                        start_col + 1,
                        share_row,
                        share_row + 1,
                        share_row + len(share_rows),
                        "品类收入占比",
                        f"{get_column_letter(start_col + 3)}{share_row}",
                    )
                    helper_rows_used = max(helper_rows_used, len(chart_rows) + len(share_rows) + 3)
            note_row = table_row + max(len(rows), helper_rows_used) + 2
            self._write_chart_note(ws, merge_start_col, note_row, "说明：本区与左侧“按品类汇总(按月)”共享同一销售数据源。")
        else:
            self._write_table(ws, 1, 1, headers, rows)
            top_rows = rows[:10]
            if top_rows:
                start_col = len(headers) + 2
                chart_rows = [[r[0], r[4]] for r in top_rows]
                self._write_table(ws, 1, start_col, ["品类", "毛利润"], chart_rows)
                anchor = f"{get_column_letter(start_col + 3)}2"
                self._add_bar_chart_from_table(ws, 1, start_col, 1 + len(chart_rows), "品类毛利润Top", anchor)
                share_rows = [[r[0], r[6]] for r in top_rows if r[6] is not None]
                if share_rows:
                    share_row = len(chart_rows) + 4
                    self._write_table(ws, share_row, start_col, ["品类", "收入占比"], share_rows)
                    self._add_doughnut_chart(
                        ws,
                        start_col,
                        start_col + 1,
                        share_row,
                        share_row + 1,
                        share_row + len(share_rows),
                        "品类收入占比",
                        f"{get_column_letter(start_col + 3)}{share_row}",
                    )
                self._write_chart_note(ws, start_col + 3, 1, "图表说明：品类毛利润Top分布。")

    def _get_ar_balance_by_customer(self, target_year, target_month):
        if self.ar_detail_df is None or self.ar_detail_df.empty:
            return {}
        df = self.ar_detail_df.copy()
        date_col = next((c for c in df.columns if '日期' in str(c)), None)
        cust_col = next((c for c in df.columns if '往来单位名' in str(c)), None)
        debit_col = next((c for c in df.columns if '借方金额' in str(c)), None)
        credit_col = next((c for c in df.columns if '贷方金额' in str(c)), None)
        if not date_col or not cust_col or not debit_col or not credit_col:
            return {}

        df['ParsedDate'] = pd.to_datetime(df[date_col], errors='coerce')
        if target_year and target_month:
            end_dt = datetime(int(target_year), int(target_month), 1)
            last_day = (end_dt.replace(day=28) + pd.Timedelta(days=4)).replace(day=1) - pd.Timedelta(days=1)
            df = df[df['ParsedDate'] <= last_day]
        df['Debit'] = pd.to_numeric(df[debit_col], errors='coerce').fillna(0)
        df['Credit'] = pd.to_numeric(df[credit_col], errors='coerce').fillna(0)
        df['Amount'] = df['Debit'] - df['Credit']
        df = df.dropna(subset=[cust_col])
        balances = df.groupby(cust_col)['Amount'].sum().to_dict()
        return balances

    def _update_customer_profit_sheet(self, wb, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "客户贡献与回款", insert_after="应收账款账龄分析")
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无销售数据"]])
            return
        df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, year_scope)
        if df.empty or '往来单位名' not in df.columns:
            self._write_table(ws, 1, 1, ["提示"], [["缺少客户字段或无数据"]])
            return
        df['Revenue'] = self._extract_sales_revenue(df)
        df['Qty'] = pd.to_numeric(df.get('数量'), errors='coerce')
        df = self._attach_sales_cost(df, target_year, target_month, year_scope)
        balances = self._get_ar_balance_by_customer(target_year, target_month)
        rows = []
        for cust, group in df.groupby('往来单位名'):
            metrics = self._calc_sales_metrics_from_group(group)
            revenue = metrics.get('revenue')
            cost = metrics.get('cost')
            qty = metrics.get('qty')
            profit = metrics.get('profit')
            margin = metrics.get('margin')
            ar_balance = balances.get(cust)
            dso = (ar_balance / revenue * 365) if (ar_balance is not None and revenue) else None
            rows.append([cust, qty, revenue, cost, profit, margin, ar_balance, dso])
        
        rows = sorted(rows, key=lambda x: x[2] if x[2] is not None else 0, reverse=True)
        
        total_rev = sum(r[2] or 0 for r in rows)
        cum_rev = 0
        for r in rows:
            rev = r[2] or 0
            cum_rev += rev
            r.append(cum_rev / total_rev if total_rev else 0)
            
        headers = ["客户", "销量", "销售收入", "销售成本", "毛利润", "毛利率", "应收余额", "DSO(天)", "累计收入占比"]
        self._write_table(ws, 1, 1, headers, rows)

        start_col = len(headers) + 2
        top_rows = rows[:10]
        if top_rows:
            chart_rows = [[r[0], r[2], r[8]] for r in top_rows]
            self._write_table(ws, 1, start_col, ["客户", "销售收入", "累计占比"], chart_rows)
            anchor = f"{get_column_letter(start_col + 4)}2"
            
            self._add_pareto_chart(
                ws, 
                start_col, 
                start_col + 1, 
                start_col + 2, 
                1, 
                2, 
                1 + len(chart_rows), 
                "客户收入集中度 (Pareto)", 
                anchor
            )
            self._add_detailed_analysis_box(ws, start_col + 4, 14, "客户集中度分析", [
                "二八法则：观察是否20%的客户贡献了80%的收入。",
                "曲线越陡峭，说明对头部客户的依赖度越高（风险越高）。",
                "若曲线平缓，说明客户结构较为分散，抗风险能力较强。"
            ])
        
        anchor_scatter = f"{get_column_letter(start_col + 4)}24"
        self._add_scatter_chart(
            ws, 
            3, 
            6, 
            1, 
            2, 
            1 + len(rows), 
            "客户价值矩阵 (收入 vs 毛利率)",
            anchor_scatter,
            x_title="销售收入",
            y_title="毛利率"
        )
        self._add_detailed_analysis_box(ws, start_col + 4, 36, "客户价值矩阵", [
            "第一象限(右上)：明星客户(高收入高毛利)，需重点维护。",
            "第二象限(左上)：潜力客户(低收入高毛利)，需提升销量。",
            "第三象限(左下)：长尾客户(低收入低毛利)，需控制服务成本。",
            "第四象限(右下)：瘦狗客户(高收入低毛利)，需优化定价或降低成本。"
        ])

    def _update_channel_profit_sheet(self, wb, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "渠道贡献", insert_after="客户贡献与回款")
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无销售数据"]])
            return
        df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, year_scope)
        channel_col = None
        for col in ['职员(负责)名', '收货公司', '职员', '渠道']:
            if col in df.columns:
                channel_col = col
                break
        if not channel_col:
            self._write_table(ws, 1, 1, ["提示"], [["缺少渠道字段"]])
            return
        df['Revenue'] = self._extract_sales_revenue(df)
        df['Qty'] = pd.to_numeric(df.get('数量'), errors='coerce')
        df = self._attach_sales_cost(df, target_year, target_month, year_scope)
        rows = []
        for channel, group in df.groupby(channel_col):
            metrics = self._calc_sales_metrics_from_group(group)
            revenue = metrics.get('revenue')
            cost = metrics.get('cost')
            qty = metrics.get('qty')
            profit = metrics.get('profit')
            margin = metrics.get('margin')
            rows.append([channel, qty, revenue, cost, profit, margin])
        headers = ["渠道", "销量", "销售收入", "销售成本", "毛利润", "毛利率"]
        rows = sorted(rows, key=lambda x: x[4] if x[4] is not None else 0, reverse=True)
        self._write_table(ws, 1, 1, headers, rows)

        top_rows = rows[:10]
        if top_rows:
            start_col = len(headers) + 2
            chart_rows = [[r[0], r[2]] for r in top_rows]
            self._write_table(ws, 1, start_col, ["渠道", "销售收入"], chart_rows)
            anchor = f"{get_column_letter(start_col + 3)}2"
            self._add_bar_chart_from_table(ws, 1, start_col, 1 + len(chart_rows), "渠道收入Top", anchor)
            self._write_chart_note(ws, start_col + 3, 1, "图表说明：渠道销售收入Top分布。")

    def _update_inventory_health_sheet(self, wb, metrics_by_month, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "存货健康度", insert_after="资产负债表")
        headers = ["月份", "主营业务成本", "期初存货", "期末存货", "平均存货", "周转率", "周转天数", "DSO(应收)", "DPO(应付)", "现金周期(CCC)"]
        rows = []
        month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
        for m_key in month_keys:
            data = metrics_by_month.get(m_key, {})
            cost = data.get("cost")
            revenue = data.get("revenue")
            inv_start = data.get("inventory_start")
            inv_end = data.get("inventory_end")
            ar_balance = data.get("ar_balance")
            ap_balance = data.get("ap_balance")
            
            avg_inv = None
            if inv_start is not None and inv_end is not None:
                avg_inv = (inv_start + inv_end) / 2
            elif inv_end is not None:
                avg_inv = inv_end
                
            turnover = cost / avg_inv if (cost and avg_inv) else None
            dio = avg_inv / cost * 30 if (avg_inv and cost) else None
            dso = ar_balance / revenue * 30 if (ar_balance and revenue) else None
            dpo = ap_balance / cost * 30 if (ap_balance and cost) else None
            
            ccc = None
            if dio is not None and dso is not None and dpo is not None:
                ccc = dio + dso - dpo
                
            rows.append([self._month_key_to_label(m_key), cost, inv_start, inv_end, avg_inv, turnover, dio, dso, dpo, ccc])
            
        if not rows:
            rows = [[None] * len(headers)]
        self._write_table(ws, 1, 1, headers, rows)
        
        anchor = f"{get_column_letter(len(headers) + 2)}2"
        self._add_line_chart_by_columns(
            ws, 
            1, 
            [7, 8, 9, 10],
            1, 
            2, 
            1 + len(rows), 
            "现金循环周期 (CCC) 分解", 
            anchor
        )
        self._add_detailed_analysis_box(ws, len(headers) + 2, 14, "现金循环周期(CCC)分析", [
            "CCC (Cash Conversion Cycle) 衡量公司将投入资源的现金转化为销售收入现金所需的天数。",
            "公式：CCC = DSO (应收天数) + DIO (库存天数) - DPO (应付天数)。",
            "CCC 越短越好，甚至可以为负（利用供应商资金运营）。",
            "DSO 上升：回款变慢；DIO 上升：库存积压；DPO 下降：付款变快（资金压力大）。"
        ])

        self._reorder_month_rows_desc(ws)

    def _bool_param(self, value, default=True):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "y", "手动", "manual"}:
            return True
        if text in {"0", "false", "no", "n", "自动", "auto"}:
            return False
        return default

    def _clamp_int_param(self, value, default, min_value, max_value):
        numeric = self._to_float(value)
        if numeric is None:
            numeric = default
        numeric = int(round(numeric))
        return max(min_value, min(max_value, numeric))

    def _clamp_float_param(self, value, default, min_value, max_value, digits=1):
        numeric = self._to_float(value)
        if numeric is None:
            numeric = default
        numeric = max(min_value, min(max_value, float(numeric)))
        return round(numeric, digits)

    def _percentile_or_none(self, values, percentile):
        valid = [self._to_float(v) for v in values]
        valid = [v for v in valid if v is not None]
        if not valid:
            return None
        return float(pd.Series(valid, dtype="float64").quantile(percentile))

    def _latest_cost_month_key(self, target_year=None, target_month=None, year_scope=None):
        month_keys = self._filter_month_keys(self.data.get("cost", {}).keys(), target_year, target_month, year_scope)
        return month_keys[-1] if month_keys else None

    def _cash_detail_balance_by_month(self, target_year=None, target_month=None, year_scope=None):
        df = self._prepare_ledger_detail_df(self.cash_detail_df)
        if df.empty:
            return {}
        df = self._filter_df_by_scope(df, target_year, target_month, year_scope)
        if df.empty:
            return {}
        result = {}
        for m_key, group in df.groupby("MonthStr"):
            ending = self._last_numeric_or_none(group.sort_values("ParsedDate").get("Balance"))
            if ending is not None:
                result[str(m_key)] = ending
        return result

    def _cost_inventory_amount_by_month(self, target_year=None, target_month=None, year_scope=None):
        result = {}
        month_keys = self._filter_month_keys(self.data.get("cost", {}).keys(), target_year, target_month, year_scope)
        for m_key in month_keys:
            cost_map = self._build_cost_inventory_map(self.data.get("cost", {}).get(m_key), m_key)
            amount = self._safe_sum([item.get("amt_end") for item in cost_map.values()])
            if amount is not None:
                result[str(m_key)] = amount
        return result

    def _infer_replenishment_warning_params(self, target_year, target_month, year_scope=None, fallback=None):
        fallback = fallback or {}
        params = {
            "lead_days": self._clamp_int_param(fallback.get("lead_days"), 30, 1, 365),
            "safety_days": self._clamp_int_param(fallback.get("safety_days"), 20, 0, 180),
            "window_months": self._clamp_int_param(fallback.get("window_months"), 3, 1, 12),
        }
        notes = []

        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            return params, "自动分析缺少销售数据，沿用默认补货参数"

        sales_df = sales_df.copy()
        if "MonthStr" not in sales_df.columns:
            date_col = next((c for c in sales_df.columns if "日期" in str(c) or "Date" in str(c)), None)
            if date_col:
                sales_df["ParsedDate"] = pd.to_datetime(sales_df[date_col], errors="coerce")
                sales_df["MonthStr"] = sales_df["ParsedDate"].dt.strftime("%Y-%m")
        if "MonthStr" not in sales_df.columns or "品目编码" not in sales_df.columns or "数量" not in sales_df.columns:
            return params, "自动分析缺少销售月份、品目编码或数量字段，沿用默认补货参数"

        sales_df["Qty"] = pd.to_numeric(sales_df.get("数量"), errors="coerce")
        sales_df = sales_df[sales_df["Qty"].notna()].copy()
        sales_df["品目编码"] = sales_df["品目编码"].astype(str).str.strip()
        sales_df = sales_df[
            (sales_df["品目编码"] != "")
            & (sales_df["品目编码"].str.lower() != "nan")
        ].copy()
        sales_df = self._filter_df_by_scope(sales_df, target_year, target_month, year_scope)
        if sales_df.empty:
            return params, "自动分析当前口径下无销售数据，沿用默认补货参数"

        available_months = sorted(sales_df["MonthStr"].dropna().astype(str).unique())
        if available_months:
            params["window_months"] = max(1, min(3, len(available_months)))
            notes.append(f"销量窗口取最近{params['window_months']}个月")

        cost_key = f"{target_year}-{int(target_month):02d}" if target_year and target_month else None
        cost_df = self.data.get("cost", {}).get(cost_key)
        if cost_df is None or cost_df.empty:
            cost_key = self._latest_cost_month_key(target_year, target_month, year_scope)
            cost_df = self.data.get("cost", {}).get(cost_key) if cost_key else None

        coverage_days = []
        volatility_values = []
        if cost_df is not None and not cost_df.empty:
            code_col = self._find_cost_col_contains(cost_df, "品目编码")
            q_end_col = self._find_cost_col_suffix(cost_df, "期末")
            if code_col and q_end_col:
                window_keys = [
                    m for m in self._last_n_month_keys(target_year, target_month, params["window_months"])
                    if m in available_months
                ]
                if not window_keys:
                    window_keys = available_months[-params["window_months"]:]
                total_days = sum(self._month_days(m) for m in window_keys) or 30 * len(window_keys)
                recent_sales = sales_df[sales_df["MonthStr"].isin(window_keys)].copy()
                qty_by_code = recent_sales.groupby("品目编码")["Qty"].sum().to_dict()
                for _, row in cost_df.iterrows():
                    code = row.get(code_col)
                    if code is None or pd.isna(code):
                        continue
                    code = str(code).strip()
                    qty_sold = qty_by_code.get(code)
                    inv_qty = self._to_float(row.get(q_end_col))
                    if qty_sold is None or inv_qty is None or qty_sold <= 0 or total_days <= 0:
                        continue
                    avg_daily = qty_sold / total_days
                    if avg_daily > 0:
                        coverage_days.append(inv_qty / avg_daily)

                monthly_qty = (
                    recent_sales.groupby(["品目编码", "MonthStr"])["Qty"]
                    .sum()
                    .unstack(fill_value=0)
                )
                for _, qty_row in monthly_qty.iterrows():
                    mean_qty = qty_row.mean()
                    if mean_qty and mean_qty > 0 and len(qty_row) > 1:
                        volatility_values.append(float(qty_row.std(ddof=0) / mean_qty))

        median_coverage = self._percentile_or_none(coverage_days, 0.5)
        if median_coverage is not None:
            params["lead_days"] = self._clamp_int_param(median_coverage * 0.2, params["lead_days"], 15, 60)
            notes.append("采购周期按库存覆盖天数中位数折算")

        median_volatility = self._percentile_or_none(volatility_values, 0.5)
        if median_volatility is not None:
            params["safety_days"] = self._clamp_int_param(10 + median_volatility * 20, params["safety_days"], 10, 45)
            notes.append("安全库存按销量波动系数折算")
        elif median_coverage is not None:
            params["safety_days"] = self._clamp_int_param(median_coverage * 0.15, params["safety_days"], 10, 45)
            notes.append("安全库存按库存覆盖天数折算")

        if not notes:
            notes.append("自动分析未取得有效库存覆盖样本，沿用默认补货参数")
        return params, "；".join(notes)

    def _cost_quantity_history_by_code(self, target_year, target_month, year_scope=None, sales_qty_by_month=None):
        """按品目编码整理成本表中的期初、入库/增加、减少、期末数量历史。"""
        history = {}
        month_keys = self._filter_month_keys(
            self.data.get("cost", {}).keys(),
            target_year,
            target_month,
            year_scope,
        )
        for m_key in month_keys:
            cost_df = self.data.get("cost", {}).get(m_key)
            if cost_df is None or cost_df.empty:
                continue
            code_col = self._find_cost_col_contains(cost_df, "品目编码")
            q_start_col = self._find_cost_col_suffix(cost_df, "期初")
            q_in_col = (
                self._find_cost_col_suffix(cost_df, "增加")
                or self._find_cost_col_suffix(cost_df, "入库")
                or self._find_cost_col_contains(cost_df, "入库")
            )
            q_out_col = self._find_cost_col_suffix(cost_df, "减少")
            q_end_col = self._find_cost_col_suffix(cost_df, "期末")
            if not code_col:
                continue

            for _, row in cost_df.iterrows():
                raw_code = row.get(code_col)
                if raw_code is None or pd.isna(raw_code):
                    continue
                code = str(raw_code).strip()
                if (
                    not code
                    or code == "品目编码"
                    or code.lower() == "nan"
                    or self._is_summary_or_footer_label(code)
                ):
                    continue

                qty_start = self._to_float(row.get(q_start_col)) if q_start_col else None
                qty_in = self._to_float(row.get(q_in_col)) if q_in_col else None
                qty_out = self._to_float(row.get(q_out_col)) if q_out_col else None
                qty_end = self._to_float(row.get(q_end_col)) if q_end_col else None
                effective_in = qty_in
                if effective_in is None and qty_start is not None and qty_end is not None:
                    sold_qty = (sales_qty_by_month or {}).get((code, m_key))
                    if sold_qty is not None:
                        effective_in = max(0, qty_end - qty_start + sold_qty)

                history.setdefault(code, []).append({
                    "month": m_key,
                    "qty_start": qty_start,
                    "qty_in": qty_in,
                    "qty_in_effective": effective_in,
                    "qty_out": qty_out,
                    "qty_end": qty_end,
                })

        for code in history:
            history[code] = sorted(history[code], key=lambda item: item.get("month") or "")
        return history

    def _month_gap_days(self, month_keys):
        periods = []
        for m_key in sorted(set(str(m) for m in month_keys if m)):
            try:
                periods.append(pd.Period(m_key, freq="M"))
            except Exception:
                continue
        gaps = []
        for idx in range(1, len(periods)):
            gaps.append((periods[idx].start_time - periods[idx - 1].start_time).days)
        return gaps

    def _infer_replenishment_params_by_product(
        self,
        sales_df,
        target_year,
        target_month,
        year_scope=None,
        fallback=None,
    ):
        """为每个品目单独推导采购/生产周期和安全库存天数。"""
        fallback = fallback or {}
        base_lead = self._clamp_int_param(fallback.get("lead_days"), 30, 1, 365)
        base_safety = self._clamp_int_param(fallback.get("safety_days"), 20, 0, 180)
        window_months = self._clamp_int_param(fallback.get("window_months"), 3, 1, 12)
        if sales_df is None or sales_df.empty or "品目编码" not in sales_df.columns:
            return {}

        work = sales_df.copy()
        if "Qty" not in work.columns:
            work["Qty"] = pd.to_numeric(work.get("数量"), errors="coerce")
        work["品目编码"] = work["品目编码"].astype(str).str.strip()
        work = work[
            work["Qty"].notna()
            & (work["品目编码"] != "")
            & (work["品目编码"].str.lower() != "nan")
        ].copy()
        if work.empty or "MonthStr" not in work.columns:
            return {}

        all_months = sorted(work["MonthStr"].dropna().astype(str).unique())
        if target_year and target_month:
            recent_months = [
                m for m in self._last_n_month_keys(target_year, target_month, window_months)
                if m in all_months
            ]
        else:
            recent_months = all_months[-window_months:]
        if not recent_months:
            recent_months = all_months[-window_months:]

        sales_qty_by_month = (
            work.groupby(["品目编码", "MonthStr"])["Qty"]
            .sum()
            .to_dict()
        )
        cost_history = self._cost_quantity_history_by_code(
            target_year,
            target_month,
            year_scope,
            sales_qty_by_month=sales_qty_by_month,
        )
        codes = sorted(set(work["品目编码"].unique()) | set(cost_history.keys()))
        result = {}

        for code in codes:
            monthly_qty = [self._to_float(sales_qty_by_month.get((code, m), 0)) or 0 for m in recent_months]
            mean_qty = (sum(monthly_qty) / len(monthly_qty)) if monthly_qty else None
            sales_cv = None
            if mean_qty and mean_qty > 0 and len(monthly_qty) > 1:
                sales_cv = float(pd.Series(monthly_qty, dtype="float64").std(ddof=0) / mean_qty)

            lead_days = base_lead
            safety_days = base_safety
            history_rows = cost_history.get(code, [])
            receipt_months = [
                row.get("month") for row in history_rows
                if (row.get("qty_in_effective") is not None and row.get("qty_in_effective") > 0)
            ]
            receipt_gaps = self._month_gap_days(receipt_months)

            if receipt_gaps:
                lead_days = self._clamp_int_param(
                    self._percentile_or_none(receipt_gaps, 0.5),
                    base_lead,
                    15,
                    120,
                )
            elif recent_months and mean_qty and mean_qty > 0 and history_rows:
                latest_inventory = None
                for row in reversed(history_rows):
                    if row.get("qty_end") is not None:
                        latest_inventory = row.get("qty_end")
                        break
                if latest_inventory is not None:
                    total_days = sum(self._month_days(m) for m in recent_months) or 30 * len(recent_months)
                    total_qty = sum(monthly_qty)
                    avg_daily = total_qty / total_days if total_days else None
                    coverage_days = latest_inventory / avg_daily if avg_daily else None
                    if coverage_days is not None:
                        lead_days = self._clamp_int_param(coverage_days * 0.2, base_lead, 15, 90)

            if sales_cv is not None:
                safety_days = self._clamp_int_param(10 + sales_cv * 20, base_safety, 7, 60)

            if len(receipt_gaps) > 1:
                mean_gap = sum(receipt_gaps) / len(receipt_gaps)
                if mean_gap:
                    receipt_cv = float(pd.Series(receipt_gaps, dtype="float64").std(ddof=0) / mean_gap)
                    if receipt_cv >= 0.1:
                        safety_days = self._clamp_int_param(
                            safety_days + min(15, receipt_cv * 10),
                            safety_days,
                            7,
                            75,
                        )
            elif history_rows and not receipt_months and mean_qty and mean_qty > 0:
                safety_days = self._clamp_int_param(safety_days + 5, safety_days, 7, 75)

            result[code] = {
                "lead_days": lead_days,
                "safety_days": safety_days,
            }

        return result

    def _cashflow_indicator_history(self, metrics_by_month, target_year, target_month, year_scope=None):
        month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
        cash_detail_balances = self._cash_detail_balance_by_month(target_year, target_month, year_scope)
        cost_inventory_amounts = self._cost_inventory_amount_by_month(target_year, target_month, year_scope)
        history = {
            "dso": [],
            "dio": [],
            "ccc": [],
            "cash_cover": [],
        }

        for m_key in month_keys:
            data = metrics_by_month.get(m_key, {})
            revenue = data.get("revenue")
            cost = data.get("cost")
            cash = data.get("cash")
            if cash is None:
                cash = cash_detail_balances.get(m_key)
            ar_balance = data.get("ar_balance")
            inv_end = data.get("inventory_end")
            if inv_end is None:
                inv_end = cost_inventory_amounts.get(m_key)
            ap_balance = data.get("ap_balance")
            sales_exp = data.get("sales_expense")
            admin_exp = data.get("admin_expense")
            fin_exp = data.get("financial_expense")

            dso = ar_balance / revenue * 30 if (ar_balance and revenue) else None
            dio = inv_end / cost * 30 if (inv_end and cost) else None
            dpo = ap_balance / cost * 30 if (ap_balance and cost) else None
            ccc = (dso + dio - dpo) if (dso is not None and dio is not None and dpo is not None) else None
            opex = None
            if sales_exp is not None or admin_exp is not None or fin_exp is not None:
                opex = (sales_exp or 0) + (admin_exp or 0) + (fin_exp or 0)
                if opex < 0:
                    opex = abs(opex)
            cash_cover = (cash / opex) if (cash is not None and opex) else None

            if dso is not None:
                history["dso"].append(dso)
            if dio is not None:
                history["dio"].append(dio)
            if ccc is not None:
                history["ccc"].append(ccc)
            if cash_cover is not None:
                history["cash_cover"].append(cash_cover)
        return history

    def _infer_cashflow_warning_params(self, metrics_by_month, target_year, target_month, year_scope=None, fallback=None):
        fallback = fallback or {}
        params = {
            "dso_threshold": self._clamp_int_param(fallback.get("dso_threshold"), 90, 1, 365),
            "dio_threshold": self._clamp_int_param(fallback.get("dio_threshold"), 180, 1, 720),
            "ccc_threshold": self._clamp_int_param(fallback.get("ccc_threshold"), 120, 1, 720),
            "cash_coverage_threshold": self._clamp_float_param(fallback.get("cash_coverage_threshold"), 1.5, 0.1, 12.0, 1),
        }
        if fallback.get("negative_cfo_streak_threshold") is not None:
            params["negative_cfo_streak_threshold"] = self._clamp_int_param(
                fallback.get("negative_cfo_streak_threshold"), 2, 1, 12
            )

        history = self._cashflow_indicator_history(metrics_by_month, target_year, target_month, year_scope)
        notes = []

        dso_p75 = self._percentile_or_none(history["dso"], 0.75)
        if dso_p75 is not None:
            params["dso_threshold"] = self._clamp_int_param(dso_p75 * 1.2, params["dso_threshold"], 30, 180)
            notes.append(f"DSO阈值基于{len(history['dso'])}个月历史75分位")

        dio_p75 = self._percentile_or_none(history["dio"], 0.75)
        if dio_p75 is not None:
            params["dio_threshold"] = self._clamp_int_param(dio_p75 * 1.2, params["dio_threshold"], 45, 360)
            notes.append(f"DIO阈值基于{len(history['dio'])}个月历史75分位")

        ccc_p75 = self._percentile_or_none(history["ccc"], 0.75)
        if ccc_p75 is not None:
            params["ccc_threshold"] = self._clamp_int_param(ccc_p75 * 1.2, params["ccc_threshold"], 45, 360)
            notes.append(f"CCC阈值基于{len(history['ccc'])}个月历史75分位")

        cash_cover_p25 = self._percentile_or_none(history["cash_cover"], 0.25)
        if cash_cover_p25 is not None:
            params["cash_coverage_threshold"] = self._clamp_float_param(
                cash_cover_p25 * 0.8,
                params["cash_coverage_threshold"],
                0.5,
                6.0,
                1,
            )
            notes.append(f"现金覆盖阈值基于{len(history['cash_cover'])}个月历史25分位")

        if not notes:
            notes.append("自动分析缺少应收、存货、货币资金或费用样本，沿用默认资金链参数")
        return params, "；".join(notes)

    def _resolve_warning_params(self, metrics_by_month, target_year, target_month, year_scope=None):
        repl_config = dict((self.report_params or {}).get("replenishment", {}))
        cash_config = dict((self.report_params or {}).get("cashflow", {}))
        repl_manual = self._bool_param(repl_config.pop("manual", True), default=True)
        cash_manual = self._bool_param(cash_config.pop("manual", True), default=True)

        repl_params = {
            "lead_days": self._clamp_int_param(repl_config.get("lead_days"), 30, 1, 365),
            "safety_days": self._clamp_int_param(repl_config.get("safety_days"), 20, 0, 180),
            "window_months": self._clamp_int_param(repl_config.get("window_months"), 3, 1, 12),
        }
        cash_params = {
            "dso_threshold": self._clamp_int_param(cash_config.get("dso_threshold"), 90, 1, 365),
            "dio_threshold": self._clamp_int_param(cash_config.get("dio_threshold"), 180, 1, 720),
            "ccc_threshold": self._clamp_int_param(cash_config.get("ccc_threshold"), 120, 1, 720),
            "cash_coverage_threshold": self._clamp_float_param(cash_config.get("cash_coverage_threshold"), 1.5, 0.1, 12.0, 1),
        }
        if cash_config.get("negative_cfo_streak_threshold") is not None:
            cash_params["negative_cfo_streak_threshold"] = self._clamp_int_param(
                cash_config.get("negative_cfo_streak_threshold"), 2, 1, 12
            )

        notes = {
            "replenishment": "参数来源：手动填入",
            "cashflow": "参数来源：手动填入",
        }
        if not repl_manual:
            repl_params, repl_note = self._infer_replenishment_warning_params(
                target_year, target_month, year_scope, fallback=repl_params
            )
            notes["replenishment"] = f"参数来源：自动分析（{repl_note}）"
        if not cash_manual:
            cash_params, cash_note = self._infer_cashflow_warning_params(
                metrics_by_month, target_year, target_month, year_scope, fallback=cash_params
            )
            notes["cashflow"] = f"参数来源：自动分析（{cash_note}）"

        self._log_audit(
            "预警参数已解析: "
            f"补货={repl_params}({notes['replenishment']}), "
            f"资金链={cash_params}({notes['cashflow']})"
        )
        return repl_params, cash_params, notes

    def _update_cashflow_alert_sheet(
        self,
        wb,
        metrics_by_month,
        target_year,
        target_month,
        year_scope=None,
        dso_threshold=90,
        dio_threshold=180,
        ccc_threshold=120,
        cash_coverage_threshold=1.5,
        negative_cfo_streak_threshold=2,
        param_source_note=None,
    ):
        ws = self._prepare_sheet(wb, "资金链预警", insert_after="现金流量表(估算)")
        month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
        if not month_keys:
            self._write_table(ws, 1, 1, ["提示"], [["无可用月份数据"]])
            return

        headers = ["月份", "经营现金流(估算)", "现金余额", "现金覆盖(月)", "DSO(应收)", "DIO(存货)", "DPO(应付)", "CCC", "风险提示"]
        rows = []

        prev = None
        negative_cfo_streak = 0
        for m_key in month_keys:
            data = metrics_by_month.get(m_key, {})
            revenue = data.get("revenue")
            cost = data.get("cost")
            sales_exp = data.get("sales_expense")
            admin_exp = data.get("admin_expense")
            fin_exp = data.get("financial_expense")
            profit = data.get("net_profit") if data.get("net_profit") is not None else data.get("operating_profit")
            cash = data.get("cash")
            ar_balance = data.get("ar_balance")
            inv_end = data.get("inventory_end")
            ap_balance = data.get("ap_balance")

            delta_ar = delta_inv = delta_ap = None
            if prev:
                if ar_balance is not None and prev.get("ar_balance") is not None:
                    delta_ar = ar_balance - prev.get("ar_balance")
                if inv_end is not None and prev.get("inventory_end") is not None:
                    delta_inv = inv_end - prev.get("inventory_end")
                if ap_balance is not None and prev.get("ap_balance") is not None:
                    delta_ap = ap_balance - prev.get("ap_balance")

            cfo = None
            if profit is not None:
                cfo = profit
                if delta_ar is not None:
                    cfo -= delta_ar
                if delta_inv is not None:
                    cfo -= delta_inv
                if delta_ap is not None:
                    cfo += delta_ap

            dso = ar_balance / revenue * 30 if (ar_balance and revenue) else None
            dio = inv_end / cost * 30 if (inv_end and cost) else None
            dpo = ap_balance / cost * 30 if (ap_balance and cost) else None
            ccc = (dso + dio - dpo) if (dso is not None and dio is not None and dpo is not None) else None

            opex = None
            if sales_exp is not None or admin_exp is not None or fin_exp is not None:
                opex = (sales_exp or 0) + (admin_exp or 0) + (fin_exp or 0)
                if opex < 0:
                    opex = abs(opex)
            cash_cover = (cash / opex) if (cash is not None and opex) else None

            risk_notes = []
            if cfo is not None and cfo < 0:
                negative_cfo_streak += 1
            else:
                negative_cfo_streak = 0
            if negative_cfo_streak >= max(1, int(negative_cfo_streak_threshold)):
                risk_notes.append("经营现金流连续为负")
            if dso is not None and dso > dso_threshold:
                risk_notes.append(f"DSO>{dso_threshold}")
            if dio is not None and dio > dio_threshold:
                risk_notes.append(f"DIO>{dio_threshold}")
            if ccc is not None and ccc > ccc_threshold:
                risk_notes.append(f"CCC>{ccc_threshold}")
            if cash_cover is not None and cash_cover < cash_coverage_threshold:
                risk_notes.append(f"现金覆盖<{cash_coverage_threshold}月")

            rows.append([
                self._month_key_to_label(m_key),
                cfo,
                cash,
                cash_cover,
                dso,
                dio,
                dpo,
                ccc,
                "；".join(risk_notes) if risk_notes else "正常",
            ])

            prev = data

        self._write_table(ws, 1, 1, headers, rows)
        self._write_chart_note(
            ws,
            1,
            len(rows) + 3,
            f"说明：阈值 DSO>{dso_threshold}天、DIO>{dio_threshold}天、CCC>{ccc_threshold}天、现金覆盖<{cash_coverage_threshold}月、经营现金流连续为负。"
            + (f" {param_source_note}" if param_source_note else ""),
        )
        self._reorder_month_rows_desc(ws)

    def _update_replenishment_alert_sheet(
        self,
        wb,
        target_year,
        target_month,
        year_scope=None,
        lead_days=30,
        safety_days=20,
        window_months=3,
        param_source_note=None,
        use_product_params=False,
    ):
        ws = self._prepare_sheet(wb, "补货预警", insert_after="存货健康度")
        if not target_year or not target_month:
            self._write_table(ws, 1, 1, ["提示"], [["缺少期间参数"]])
            return

        month_key = f"{target_year}-{int(target_month):02d}"
        cost_df = self.data['cost'].get(month_key)
        if cost_df is None or cost_df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无成本数据，无法计算期末库存"]])
            return

        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无销售数据，无法计算补货预警"]])
            return

        sales_df = sales_df.copy()
        if 'MonthStr' not in sales_df.columns:
            date_col = next((c for c in sales_df.columns if '日期' in str(c) or 'Date' in str(c)), None)
            if date_col:
                sales_df['ParsedDate'] = pd.to_datetime(sales_df[date_col], errors='coerce')
                sales_df['MonthStr'] = sales_df['ParsedDate'].dt.strftime('%Y-%m')

        if 'MonthStr' not in sales_df.columns:
            self._write_table(ws, 1, 1, ["提示"], [["销售数据缺少月份字段"]])
            return

        code_col = self._find_cost_col_contains(cost_df, '品目编码')
        name_col = self._find_cost_col_contains(cost_df, '品目名')
        q_end_col = self._find_cost_col_suffix(cost_df, '期末')
        cost_category_col = None
        for c in cost_df.columns:
            name = str(c)
            if (
                any(k in name for k in ('产品大类', '品目组合1名', '品类', '类别', '类目'))
                and '编码' not in name
                and '品目名' not in name
                and '名称' not in name
            ):
                cost_category_col = c
                break
        if not code_col or not q_end_col:
            self._write_table(ws, 1, 1, ["提示"], [["成本数据缺少品目编码或期末数量"]])
            return

        if '品目编码' not in sales_df.columns:
            self._write_table(ws, 1, 1, ["提示"], [["销售数据缺少品目编码字段"]])
            return

        if '数量' not in sales_df.columns:
            self._write_table(ws, 1, 1, ["提示"], [["销售数据缺少数量字段"]])
            return

        sales_df['Qty'] = pd.to_numeric(sales_df.get('数量'), errors='coerce')
        sales_df = sales_df[sales_df['Qty'].notna()]
        sales_df['品目编码'] = sales_df['品目编码'].astype(str).str.strip()
        sales_df = sales_df[
            (sales_df['品目编码'] != '')
            & (sales_df['品目编码'].str.lower() != 'nan')
        ].copy()
        sales_df = self._filter_df_by_scope(sales_df, target_year, target_month, year_scope)
        if sales_df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["当前口径下无销售数据"]])
            return

        available_months = sorted(sales_df['MonthStr'].dropna().unique())
        if target_year and target_month:
            window_keys = [m for m in self._last_n_month_keys(target_year, target_month, window_months) if m in available_months]
        else:
            window_keys = available_months[-window_months:]
        if not window_keys:
            self._write_table(ws, 1, 1, ["提示"], [["无有效月份数据"]])
            return

        product_params = {}
        if use_product_params:
            product_params = self._infer_replenishment_params_by_product(
                sales_df,
                target_year,
                target_month,
                year_scope,
                fallback={
                    "lead_days": lead_days,
                    "safety_days": safety_days,
                    "window_months": window_months,
                },
            )

        total_days = sum(self._month_days(m) for m in window_keys)
        if total_days <= 0:
            total_days = 30 * len(window_keys)

        sales_recent = sales_df[sales_df['MonthStr'].isin(window_keys)]
        sales_qty = sales_recent.groupby('品目编码')['Qty'].sum().to_dict()
        name_map = {}
        sales_category_by_code = {}
        if '品目名' in sales_recent.columns:
            name_map = (
                sales_recent.groupby('品目编码')
                .agg({'品目名': 'first'})
                .to_dict()
                .get('品目名', {})
            )
        sales_category_cols = [
            c for c in sales_recent.columns
            if (
                any(k in str(c) for k in ('产品大类', '品目组合1名', '品类', '类别', '类目'))
                and '编码' not in str(c)
                and '品目名' not in str(c)
                and '名称' not in str(c)
            )
        ]
        if sales_category_cols:
            for _, srow in sales_recent.iterrows():
                code = srow.get('品目编码')
                if code is None or pd.isna(code):
                    continue
                code = str(code).strip()
                if not code or code in sales_category_by_code:
                    continue
                for cat_col in sales_category_cols:
                    raw = srow.get(cat_col)
                    if raw is None or pd.isna(raw):
                        continue
                    text = str(raw).strip()
                    if text and text.lower() != 'nan':
                        sales_category_by_code[code] = text
                        break

        headers = [
            "品目编码",
            "品目名称",
            "类别",
            "期末库存数量",
            f"近{len(window_keys)}个月日均销量",
            "库存覆盖天数",
            "采购/生产周期(天)",
            "安全库存天数",
            "建议补货量",
            "风险等级",
        ]
        rows = []

        for _, row in cost_df.iterrows():
            code = row.get(code_col)
            if code is None or pd.isna(code):
                continue
            code = str(code).strip()
            if not code:
                continue

            inv_qty = self._to_float(row.get(q_end_col))
            if inv_qty is None:
                continue

            qty_sold = sales_qty.get(code)
            if qty_sold is None or total_days <= 0:
                continue

            avg_daily = qty_sold / total_days if total_days else None
            if not avg_daily or avg_daily <= 0:
                continue

            coverage_days = inv_qty / avg_daily if avg_daily else None
            if coverage_days is None:
                continue

            row_params = product_params.get(code, {}) if product_params else {}
            row_lead_days = row_params.get("lead_days", lead_days)
            row_safety_days = row_params.get("safety_days", safety_days)
            need_days = row_lead_days + row_safety_days
            if coverage_days >= need_days:
                continue

            reorder_qty = max(0, avg_daily * need_days - inv_qty)
            risk = "高" if coverage_days < row_safety_days else "中"
            name_spec = row.get(name_col) if name_col else None
            resolved_name = name_map.get(code) or name_spec
            category_hint = row.get(cost_category_col) if cost_category_col else None
            if category_hint is None or (isinstance(category_hint, float) and pd.isna(category_hint)):
                category_hint = sales_category_by_code.get(code)
            category = self._resolve_uncategorized_product(category_hint, resolved_name)

            rows.append([
                code,
                resolved_name,
                category,
                inv_qty,
                avg_daily,
                coverage_days,
                row_lead_days,
                row_safety_days,
                reorder_qty,
                risk,
            ])

        if not rows:
            self._write_table(ws, 1, 1, ["提示"], [["未发现需要补货的产品"]])
            return

        rows = sorted(rows, key=lambda x: x[5] if x[5] is not None else 999999)
        self._write_table(ws, 1, 1, headers, rows)
        self._write_chart_note(
            ws,
            1,
            len(rows) + 3,
            (
                f"说明：覆盖天数 < 每个品目的采购/生产周期+安全库存天数列为需补货。"
                f" 自动模式下按品目历史销售、库存覆盖和入库节奏单独计算，"
                f"缺少历史时使用全局兜底({lead_days}+{safety_days}天)。"
                if use_product_params else
                f"说明：覆盖天数 < 采购周期({lead_days})+安全库存({safety_days}) 列为需补货。"
            )
            + (f" {param_source_note}" if param_source_note else ""),
        )

    def _update_slow_moving_inventory_sheet(self, wb, target_year, target_month):
        ws = self._prepare_sheet(wb, "滞销与风险存货", insert_after="存货健康度")
        if not target_year or not target_month:
            self._write_table(ws, 1, 1, ["提示"], [["缺少期间参数"]])
            return
        month_key = f"{target_year}-{int(target_month):02d}"
        cost_df = self.data['cost'].get(month_key)
        if cost_df is None or cost_df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无成本数据"]])
            return

        def find_col_contains(keyword):
            for c in cost_df.columns:
                if keyword in str(c):
                    return c
            return None

        def find_col_suffix(suffix):
            for c in cost_df.columns:
                if str(c).endswith(suffix):
                    return c
            return None

        code_col = find_col_contains('品目编码')
        name_col = find_col_contains('品目名')
        qty_end_col = find_col_suffix('期末')
        amt_end_col = find_col_suffix('期末.2')
        if not code_col or not qty_end_col:
            self._write_table(ws, 1, 1, ["提示"], [["缺少库存关键字段"]])
            return

        sales_df = self._get_sales_df()
        sales_map = {}
        if sales_df is not None and not sales_df.empty:
            df = sales_df.copy()
            df['Revenue'] = self._extract_sales_revenue(df)
            df['Qty'] = pd.to_numeric(df.get('数量'), errors='coerce')
            start_year, start_month = self._add_months(int(target_year), int(target_month), -2)
            start_key = f"{start_year}-{start_month:02d}"
            df = df[df.get('MonthStr') >= start_key]
            df = df[df.get('MonthStr') <= month_key]
            for code, group in df.groupby('品目编码'):
                if code is None or pd.isna(code):
                    continue
                sales_map[str(code).strip()] = {
                    "qty": group['Qty'].sum(),
                    "revenue": group['Revenue'].sum(),
                }

        headers = ["品目编码", "品目名", "期末数量", "期末金额", "近3月销量", "近3月销售额", "风险等级"]
        rows = []
        risk_counts = {"高": 0, "中": 0, "低": 0}
        
        for _, row in cost_df.iterrows():
            code = row.get(code_col)
            if not code or pd.isna(code):
                continue
            code = str(code).strip()
            qty_end = self._to_float(row.get(qty_end_col))
            amt_end = self._to_float(row.get(amt_end_col)) if amt_end_col else None
            if not qty_end or qty_end <= 0:
                continue
            sales = sales_map.get(code, {})
            qty_sales = sales.get("qty") or 0
            rev_sales = sales.get("revenue") or 0
            risk = "高" if qty_sales == 0 else ("中" if qty_sales < qty_end * 0.2 else "低")
            risk_counts[risk] += 1
            rows.append([
                code,
                row.get(name_col),
                qty_end,
                amt_end,
                qty_sales,
                rev_sales,
                risk,
            ])

        rows = sorted(rows, key=lambda x: (x[6], x[2] if x[2] is not None else 0), reverse=True)
        if not rows:
            rows = [[None] * len(headers)]
        self._write_table(ws, 1, 1, headers, rows)
        
        start_col = len(headers) + 2
        chart_data = [["风险等级", "数量"], ["高", risk_counts["高"]], ["中", risk_counts["中"]], ["低", risk_counts["低"]]]
        self._write_table(ws, 1, start_col, ["风险等级", "数量"], chart_data[1:])
        
        anchor = f"{get_column_letter(start_col + 3)}2"
        self._add_doughnut_chart(ws, start_col, start_col + 1, 1, 2, 4, "库存风险等级分布", anchor)
        
        self._add_detailed_analysis_box(ws, start_col + 3, 14, "库存风险分布", [
            "风险等级依据近3月销量占库存比例计算。",
            "高风险：近3月无销量 (死库存)。",
            "中风险：近3月销量 < 库存的20% (周转极慢)。",
            "低风险：正常周转。",
            "建议优先处理高风险库存，释放现金流。"
        ])

    def _update_expense_structure_sheet(self, wb, metrics_by_month, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "费用结构与弹性", insert_after="费用明细环比分析")
        behavior_totals = self._calculate_expense_behavior_totals(target_year, target_month, year_scope)
        var_by_month = behavior_totals.get("variable", {})
        fixed_by_month = behavior_totals.get("fixed", {})
        finance_by_month = behavior_totals.get("financial_adjustment", {})
        inventory_by_month = behavior_totals.get("inventory_adjustment", {})
        other_by_month = behavior_totals.get("unclassified", {})
        total_by_month = behavior_totals.get("total", {})
        month_keys = (
            set(total_by_month.keys())
            | set(var_by_month.keys())
            | set(fixed_by_month.keys())
            | set(finance_by_month.keys())
            | set(inventory_by_month.keys())
            | set(other_by_month.keys())
            | set(metrics_by_month.keys())
        )
        month_keys = self._filter_month_keys(month_keys, target_year, target_month, year_scope)
        headers = [
            "月份", "变动费用", "固定费用", "财务/汇兑调整", "库存/盘点调整", "未分类费用",
            "费用合计", "分类差异", "变动占比", "固定占比", "库存调整占比", "收入", "费用率", "费用弹性"
        ]
        rows = []
        prev_total = None
        prev_revenue = None
        for m_key in month_keys:
            var_cost = var_by_month.get(m_key) or 0
            fixed_cost = fixed_by_month.get(m_key) or 0
            finance_cost = finance_by_month.get(m_key) or 0
            inventory_cost = inventory_by_month.get(m_key) or 0
            other_cost = other_by_month.get(m_key) or 0
            classified_total = var_cost + fixed_cost + finance_cost + inventory_cost + other_cost
            total = total_by_month.get(m_key)
            if total is None:
                total = classified_total
            classify_gap = total - classified_total if total is not None else None
            revenue = metrics_by_month.get(m_key, {}).get("revenue") if metrics_by_month.get(m_key) else None
            var_ratio = var_cost / total if total else None
            fixed_ratio = fixed_cost / total if total else None
            inventory_ratio = inventory_cost / total if total else None
            expense_rate = total / revenue if revenue else None
            elasticity = None
            if prev_total and prev_revenue and revenue:
                total_change = (total - prev_total) / prev_total if prev_total else None
                revenue_change = (revenue - prev_revenue) / prev_revenue if prev_revenue else None
                if total_change is not None and revenue_change:
                    elasticity = total_change / revenue_change
            rows.append([
                self._month_key_to_label(m_key),
                var_cost,
                fixed_cost,
                finance_cost,
                inventory_cost,
                other_cost,
                total,
                classify_gap,
                var_ratio,
                fixed_ratio,
                inventory_ratio,
                revenue,
                expense_rate,
                elasticity,
            ])
            prev_total = total
            if revenue:
                prev_revenue = revenue
        if not rows:
            rows = [[None] * len(headers)]
        self._write_table(ws, 1, 1, headers, rows)
        
        # 1. Trend Line Chart (Total Expense & Revenue)
        anchor1 = f"{get_column_letter(len(headers) + 2)}2"
        self._add_line_chart_by_columns(ws, 1, [7, 12], 1, 2, 1 + len(rows), "费用总额与收入趋势", anchor1)
        
        # 2. Stacked Bar Chart (Structure %)
        anchor2 = f"{get_column_letter(len(headers) + 2)}18"
        self._add_stacked_bar_chart(ws, 1, [2, 3, 4, 5, 6], 1, 2, 1 + len(rows), "费用结构占比趋势", anchor2, percent=True)
        
        self._write_chart_note(ws, len(headers) + 2, 1, "图表说明：费用总额趋势及变动/固定/财务/库存调整/未分类结构占比。")

        self._reorder_month_rows_desc(ws)

    def _update_anomaly_sheet(self, wb, metrics_by_month, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "异常预警", insert_after="费用结构与弹性")
        _, _, (ratio_threshold, pp_threshold, days_threshold) = self._read_budget_targets(wb)
        headers = ["月份", "指标", "本期值", "上期值", "变化值", "变化率/差值", "阈值", "级别", "建议"]
        rows = []
        month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
        prev_data = None
        for m_key in month_keys:
            data = metrics_by_month.get(m_key, {})
            revenue = data.get("revenue")
            profit = data.get("operating_profit")
            cost_rate = data.get("cost_rate")
            sales_expense = data.get("sales_expense")
            admin_expense = data.get("admin_expense")
            financial_expense = data.get("financial_expense")
            sales_rate = sales_expense / revenue if (sales_expense is not None and revenue not in (None, 0)) else None
            admin_rate = admin_expense / revenue if (admin_expense is not None and revenue not in (None, 0)) else None
            financial_rate = financial_expense / revenue if (financial_expense is not None and revenue not in (None, 0)) else None
            inv_start = data.get("inventory_start")
            inv_end = data.get("inventory_end")
            avg_inv = (inv_start + inv_end) / 2 if (inv_start is not None and inv_end is not None) else inv_end
            inv_days = avg_inv / data.get("cost") * 365 if (avg_inv and data.get("cost")) else None
            if prev_data:
                def check_ratio(label, current, previous):
                    if current is None or previous is None:
                        return
                    delta = current - previous
                    rate = delta / previous if previous else None
                    if rate is not None and abs(rate) >= ratio_threshold:
                        rows.append([
                            self._month_key_to_label(m_key), label, current, previous, delta,
                            rate, ratio_threshold, "高", "关注业务波动原因",
                        ])

                def check_pp(label, current, previous):
                    if current is None or previous is None:
                        return
                    delta = current - previous
                    if abs(delta) >= pp_threshold:
                        rows.append([
                            self._month_key_to_label(m_key), label, current, previous, delta,
                            delta, pp_threshold, "中", "检查费用率或成本率变化",
                        ])

                def check_days(label, current, previous):
                    if current is None or previous is None:
                        return
                    delta = current - previous
                    if abs(delta) >= days_threshold:
                        rows.append([
                            self._month_key_to_label(m_key), label, current, previous, delta,
                            delta, days_threshold, "中", "检查周转效率",
                        ])

                check_ratio("主营业务收入", revenue, prev_data.get("revenue"))
                check_ratio("营业利润", profit, prev_data.get("operating_profit"))
                check_ratio("应收账款余额", data.get("ar_balance"), prev_data.get("ar_balance"))
                check_ratio("存货期末余额", inv_end, prev_data.get("inventory_end"))
                check_pp("成本率", cost_rate, prev_data.get("cost_rate"))
                check_pp("销售费用率", sales_rate, prev_data.get("sales_rate"))
                check_pp("管理费用率", admin_rate, prev_data.get("admin_rate"))
                check_pp("财务费用率", financial_rate, prev_data.get("financial_rate"))
                check_days("存货周转天数", inv_days, prev_data.get("inv_days"))

            prev_data = {
                "revenue": revenue,
                "operating_profit": profit,
                "cost_rate": cost_rate,
                "sales_rate": sales_rate,
                "admin_rate": admin_rate,
                "financial_rate": financial_rate,
                "ar_balance": data.get("ar_balance"),
                "inventory_end": inv_end,
                "inv_days": inv_days,
            }

        if not rows:
            rows = [["-", "无异常", None, None, None, None, None, "低", ""]]
        self._write_table(ws, 1, 1, headers, rows)
        self._write_chart_note(ws, 1, len(rows) + 3, "说明：阈值来自目标_预算的异常阈值设置。")

        self._reorder_month_rows_desc(ws)

    def _add_expense_diagnostic_charts(
        self,
        ws,
        rows_a,
        top_rows,
        dept_rows,
        finance_rows,
        matrix_rows=None,
        subject_compare_rows=None,
    ):
        chart_col = max(ws.max_column + 2, 15)
        anchor_col = chart_col + 6

        def has_label(label):
            return label not in (None, "", "-")

        trend_rows = []
        for row in rows_a or []:
            if len(row) < 13 or not has_label(row[0]):
                continue
            total = self._to_float(row[6])
            revenue = self._to_float(row[11])
            rate = self._to_float(row[12])
            if total is None and revenue is None and rate is None:
                continue
            trend_rows.append([row[0], total, revenue, rate])
        if trend_rows:
            start_row = 1
            self._write_table(ws, start_row, chart_col, ["月份", "费用合计", "收入", "费用率"], trend_rows)
            self._add_combo_chart(
                ws,
                chart_col,
                [chart_col + 1, chart_col + 2],
                [chart_col + 3],
                start_row,
                start_row + 1,
                start_row + len(trend_rows),
                "费用合计、收入与费用率趋势",
                f"{get_column_letter(anchor_col)}{start_row}",
            )

        structure_rows = []
        for row in rows_a or []:
            if len(row) < 6 or not has_label(row[0]):
                continue
            values = [self._to_float(row[i]) for i in range(1, 6)]
            if not any(v is not None and v != 0 for v in values):
                continue
            structure_rows.append([row[0]] + values)
        if structure_rows:
            start_row = 18
            self._write_table(
                ws,
                start_row,
                chart_col,
                ["月份", "变动费用", "固定费用", "财务/汇兑调整", "库存/盘点调整", "未分类费用"],
                structure_rows,
            )
            self._add_stacked_bar_chart(
                ws,
                chart_col,
                [chart_col + 1, chart_col + 2, chart_col + 3, chart_col + 4, chart_col + 5],
                start_row,
                start_row + 1,
                start_row + len(structure_rows),
                "费用结构占比趋势",
                f"{get_column_letter(anchor_col)}{start_row}",
                percent=True,
            )

        anomaly_rows = []
        for row in top_rows or []:
            if len(row) < 10 or not has_label(row[1]):
                continue
            score = self._to_float(row[9])
            if score is None:
                delta = self._to_float(row[5])
                amount = self._to_float(row[3])
                score = abs(delta if delta is not None else (amount or 0))
            label = f"{row[1]}-{row[2]}" if row[2] else row[1]
            anomaly_rows.append([self._short_chart_label(label), score])
            if len(anomaly_rows) >= 12:
                break
        if not anomaly_rows:
            for row in matrix_rows or []:
                if len(row) < 7 or not has_label(row[0]):
                    continue
                score = self._to_float(row[6])
                if score is None:
                    delta = self._to_float(row[4])
                    amount = self._to_float(row[2])
                    score = abs(delta if delta is not None else (amount or 0))
                if score is None:
                    continue
                label = f"{row[0]}-{row[1]}" if row[1] else row[0]
                anomaly_rows.append([self._short_chart_label(label), score])
                if len(anomaly_rows) >= 12:
                    break
        if anomaly_rows:
            start_row = 35
            self._write_table(ws, start_row, chart_col, ["异常项目", "异常评分"], anomaly_rows)
            self._add_bar_chart_from_table(
                ws,
                start_row,
                chart_col,
                start_row + len(anomaly_rows),
                "费用异常评分Top",
                f"{get_column_letter(anchor_col)}{start_row}",
            )

        def build_subject_change_rows(delta_idx, rate_idx):
            rows = []
            source_rows = subject_compare_rows or []
            for row in source_rows:
                if len(row) <= max(delta_idx, rate_idx) or not has_label(row[0]):
                    continue
                delta = self._to_float(row[delta_idx])
                rate = self._to_float(row[rate_idx])
                if delta is None and rate is None:
                    continue
                label = f"{row[0]}-{row[1]}" if row[1] else row[0]
                rows.append([self._short_chart_label(label, 22), delta, rate])
            rows.sort(key=lambda x: abs(x[1] or 0), reverse=True)
            return rows[:12]

        mom_change_rows = build_subject_change_rows(4, 5)
        if mom_change_rows:
            start_row = 52
            self._write_table(ws, start_row, chart_col, ["费用科目", "环比增量", "环比增速"], mom_change_rows)
            for rr in range(start_row + 1, start_row + 1 + len(mom_change_rows)):
                ws.cell(row=rr, column=chart_col + 2).number_format = "0.00%"
            self._add_combo_chart(
                ws,
                chart_col,
                [chart_col + 1],
                [chart_col + 2],
                start_row,
                start_row + 1,
                start_row + len(mom_change_rows),
                "费用科目环比变动Top",
                f"{get_column_letter(anchor_col)}{start_row}",
            )

        yoy_change_rows = build_subject_change_rows(7, 8)
        if yoy_change_rows:
            start_row = 69
            self._write_table(ws, start_row, chart_col, ["费用科目", "同比增量", "同比增速"], yoy_change_rows)
            for rr in range(start_row + 1, start_row + 1 + len(yoy_change_rows)):
                ws.cell(row=rr, column=chart_col + 2).number_format = "0.00%"
            self._add_combo_chart(
                ws,
                chart_col,
                [chart_col + 1],
                [chart_col + 2],
                start_row,
                start_row + 1,
                start_row + len(yoy_change_rows),
                "费用科目同比变动Top",
                f"{get_column_letter(anchor_col)}{start_row}",
            )

        dept_chart_rows = []
        for row in dept_rows or []:
            if len(row) < 5 or not has_label(row[0]):
                continue
            values = [self._to_float(row[i]) for i in range(1, 5)]
            if not any(v is not None and v != 0 for v in values):
                continue
            dept_chart_rows.append([self._short_chart_label(row[0], 18)] + values)
            if len(dept_chart_rows) >= 10:
                break
        if dept_chart_rows:
            start_row = 86
            self._write_table(
                ws,
                start_row,
                chart_col,
                ["部门", "销售费用", "管理费用", "财务费用", "其他费用"],
                dept_chart_rows,
            )
            self._add_stacked_bar_chart(
                ws,
                chart_col,
                [chart_col + 1, chart_col + 2, chart_col + 3, chart_col + 4],
                start_row,
                start_row + 1,
                start_row + len(dept_chart_rows),
                "目标月部门费用构成",
                f"{get_column_letter(anchor_col)}{start_row}",
            )

        finance_chart_rows = []
        for row in finance_rows or []:
            if len(row) < 4 or not has_label(row[0]):
                continue
            net_amount = self._to_float(row[3])
            if net_amount is None:
                continue
            finance_chart_rows.append([self._short_chart_label(row[0], 18), net_amount])
            if len(finance_chart_rows) >= 10:
                break
        if finance_chart_rows:
            start_row = 103
            self._write_table(ws, start_row, chart_col, ["财务费用子科目", "净额"], finance_chart_rows)
            self._add_bar_chart_from_table(
                ws,
                start_row,
                chart_col,
                start_row + len(finance_chart_rows),
                "财务费用净额拆解",
                f"{get_column_letter(anchor_col)}{start_row}",
            )

        self._write_chart_note(ws, chart_col, 118, "图表说明：右侧集中展示费用趋势、结构、异常、科目环比/同比、部门贡献与财务费用拆解。")

    def _update_expense_diagnostic_center(
        self,
        wb,
        metrics_by_month,
        target_year,
        target_month,
        year_scope=None,
        anomaly_top_n=60,
        matrix_top_n=120,
        detail_lines_per_key=3,
    ):
        legacy_sheet_name = "费用诊断中心"
        if self.expense_analysis_sheet_name not in wb.sheetnames and legacy_sheet_name in wb.sheetnames:
            wb[legacy_sheet_name].title = self.expense_analysis_sheet_name
        ws = self._prepare_sheet(wb, self.expense_analysis_sheet_name, insert_after="费用明细环比分析")
        month_title = f"{target_year}/{int(target_month):02d}" if target_year and target_month else "当前"
        ws.cell(row=1, column=1).value = f"{month_title} {self.expense_analysis_sheet_name}"
        ws.cell(row=1, column=1).font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
        ws.cell(row=2, column=1).value = "整合原【费用明细/费用明细环比分析/费用结构与弹性/异常预警/年度费用异常Top】"

        raw_df = self._get_expense_df()
        df = self._prepare_expense_analysis_df(raw_df, target_year, target_month, year_scope)
        if df is None or df.empty:
            self._write_table(ws, 4, 1, ["提示"], [["无可用费用数据"]])
            return

        def month_shift(month_key, offset):
            try:
                p = pd.Period(month_key, freq='M') + offset
                return f"{p.year}-{p.month:02d}"
            except Exception:
                return None

        def delta_rate(curr, prev):
            if curr is None or prev is None:
                return None, None
            delta = curr - prev
            rate = (delta / prev) if prev else None
            return delta, rate

        row_cursor = 4

        # A) 月度费用结构与弹性
        ws.cell(row=row_cursor, column=1).value = "A. 月度费用结构与弹性"
        ws.cell(row=row_cursor, column=1).font = Font(name="微软雅黑", bold=True, color="1F4E79")
        row_cursor += 1

        behavior_totals = self._calculate_expense_behavior_totals(target_year, target_month, year_scope)
        var_by_month = behavior_totals.get("variable", {})
        fixed_by_month = behavior_totals.get("fixed", {})
        finance_by_month = behavior_totals.get("financial_adjustment", {})
        inventory_by_month = behavior_totals.get("inventory_adjustment", {})
        other_by_month = behavior_totals.get("unclassified", {})
        total_by_month = behavior_totals.get("total", {})
        month_keys = (
            set(total_by_month.keys())
            | set(var_by_month.keys())
            | set(fixed_by_month.keys())
            | set(finance_by_month.keys())
            | set(inventory_by_month.keys())
            | set(other_by_month.keys())
            | set(metrics_by_month.keys())
        )
        month_keys = self._filter_month_keys(month_keys, target_year, target_month, year_scope)
        headers_a = [
            "月份", "变动费用", "固定费用", "财务/汇兑调整", "库存/盘点调整", "未分类费用",
            "费用合计", "分类差异", "变动占比", "固定占比", "库存调整占比", "收入", "费用率", "费用弹性"
        ]
        rows_a = []
        prev_total = None
        prev_revenue = None
        for m_key in month_keys:
            var_cost = var_by_month.get(m_key) or 0
            fixed_cost = fixed_by_month.get(m_key) or 0
            finance_cost = finance_by_month.get(m_key) or 0
            inventory_cost = inventory_by_month.get(m_key) or 0
            other_cost = other_by_month.get(m_key) or 0
            classified_total = var_cost + fixed_cost + finance_cost + inventory_cost + other_cost
            total = total_by_month.get(m_key)
            if total is None:
                total = classified_total
            classify_gap = total - classified_total if total is not None else None
            revenue = metrics_by_month.get(m_key, {}).get("revenue") if metrics_by_month.get(m_key) else None
            var_ratio = var_cost / total if total else None
            fixed_ratio = fixed_cost / total if total else None
            inventory_ratio = inventory_cost / total if total else None
            expense_rate = total / revenue if revenue else None
            elasticity = None
            if prev_total and prev_revenue and revenue:
                total_change = (total - prev_total) / prev_total if prev_total else None
                revenue_change = (revenue - prev_revenue) / prev_revenue if prev_revenue else None
                if total_change is not None and revenue_change:
                    elasticity = total_change / revenue_change
            rows_a.append([
                self._month_key_to_label(m_key),
                var_cost,
                fixed_cost,
                finance_cost,
                inventory_cost,
                other_cost,
                total,
                classify_gap,
                var_ratio,
                fixed_ratio,
                inventory_ratio,
                revenue,
                expense_rate,
                elasticity,
            ])
            prev_total = total
            if revenue:
                prev_revenue = revenue
        if not rows_a:
            rows_a = [[None] * len(headers_a)]
        self._write_table(ws, row_cursor, 1, headers_a, rows_a)
        row_cursor += len(rows_a) + 2

        # B) 科目环比矩阵（重点）
        ws.cell(row=row_cursor, column=1).value = "B. 费用科目环比矩阵（重点）"
        ws.cell(row=row_cursor, column=1).font = Font(name="微软雅黑", bold=True, color="1F4E79")
        row_cursor += 1

        summary = (
            df.groupby(['Category', 'Subcategory', 'MonthStr'])['Amount']
            .sum()
            .reset_index()
        )
        comparison_df = self._prepare_expense_analysis_df(raw_df, target_year, target_month, "all")
        if comparison_df is None or comparison_df.empty:
            comparison_df = df
        comparison_summary = (
            comparison_df.groupby(['Category', 'Subcategory', 'MonthStr'])['Amount']
            .sum()
            .reset_index()
        )
        detail_count_map = (
            df.groupby(['MonthStr', 'Category', 'Subcategory'])['Amount']
            .size()
            .to_dict()
        )

        flags = self._collect_expense_mom_flags(df, target_year, target_month, year_scope)
        flags = sorted(
            flags,
            key=lambda x: (
                x.get("AnomalyScore") or 0,
                abs(x.get("Delta") or 0),
                abs(x.get("Amount") or 0),
            ),
            reverse=True,
        )
        target_key = f"{target_year}-{int(target_month):02d}" if target_year and target_month else None
        if not target_key:
            all_keys = sorted(df['MonthStr'].dropna().astype(str).unique().tolist()) if 'MonthStr' in df.columns else []
            target_key = all_keys[-1] if all_keys else None
        prev_key = month_shift(target_key, -1) if target_key else None
        yoy_key = month_shift(target_key, -12) if target_key else None
        selected_by_pair = self._select_expense_display_flags(
            flags,
            target_year,
            target_month,
            target_only=bool(target_key),
        )
        target_df = df[df['MonthStr'] == target_key].copy() if target_key else df.copy()

        subject_compare_rows = []
        for (cat, sub), g in comparison_summary.groupby(['Category', 'Subcategory']):
            month_map = {row['MonthStr']: row['Amount'] for _, row in g.iterrows()}
            curr = month_map.get(target_key) if target_key else None
            if curr is None:
                continue
            prev = month_map.get(prev_key) if prev_key else None
            yoy = month_map.get(yoy_key) if yoy_key else None
            mom_delta, mom_rate = delta_rate(curr, prev)
            yoy_delta, yoy_rate = delta_rate(curr, yoy)
            if mom_delta is None and yoy_delta is None:
                continue
            subject_compare_rows.append([
                cat,
                sub,
                curr,
                prev,
                mom_delta,
                mom_rate,
                yoy,
                yoy_delta,
                yoy_rate,
            ])
        subject_compare_rows = sorted(
            subject_compare_rows,
            key=lambda x: max(abs(x[4] or 0), abs(x[7] or 0), abs(x[2] or 0)),
            reverse=True,
        )

        matrix_headers = ["费用类别", "子科目", "本期金额", "上期金额", "环比增量", "环比增速", "异常评分", "异常标签", "明细笔数", "明细键"]
        matrix_rows = []
        for (cat, sub), g in summary.groupby(['Category', 'Subcategory']):
            month_map = {row['MonthStr']: row['Amount'] for _, row in g.iterrows()}
            curr = month_map.get(target_key) if target_key else None
            prev = month_map.get(prev_key) if prev_key else None
            delta, rate = delta_rate(curr, prev)
            pair_flag = selected_by_pair.get((cat, sub))
            tags = "、".join(pair_flag.get("ReasonTags") or []) if pair_flag else None
            score = pair_flag.get("AnomalyScore") if pair_flag else None
            key_text = pair_flag.get("AnomalyKey") if pair_flag else None
            if curr is None and delta is None and score is None:
                continue
            matrix_rows.append([
                cat,
                sub,
                curr,
                prev,
                delta,
                rate,
                score,
                tags,
                detail_count_map.get((target_key, cat, sub)) if target_key else None,
                key_text,
            ])

        matrix_rows = sorted(
            matrix_rows,
            key=lambda x: (
                x[6] or 0,
                abs(x[4] or 0),
                abs(x[2] or 0),
            ),
            reverse=True,
        )[:matrix_top_n]
        if not matrix_rows:
            matrix_rows = [[None] * len(matrix_headers)]
        matrix_start_row = row_cursor
        self._write_table(ws, matrix_start_row, 1, matrix_headers, matrix_rows)
        matrix_key_col = 1 + matrix_headers.index("明细键")
        row_cursor += len(matrix_rows) + 2

        # C) 目标月异常Top（综合评分）
        ws.cell(row=row_cursor, column=1).value = "C. 异常Top（目标月综合评分）"
        ws.cell(row=row_cursor, column=1).font = Font(name="微软雅黑", bold=True, color="1F4E79")
        row_cursor += 1

        target_flags = [f for f in flags if not target_key or f.get("MonthStr") == target_key]
        top_flags = target_flags[:anomaly_top_n]
        top_headers = [
            "月份", "费用类别", "子科目", "本期金额", "上期金额",
            "环比增量", "环比增速", "同比增量", "同比增速",
            "异常评分", "异常标签", "近12月出现", "明细键"
        ]
        top_rows = []
        for f in top_flags:
            top_rows.append([
                self._month_key_to_label(f.get("MonthStr")),
                f.get("Category"),
                f.get("Subcategory"),
                f.get("Amount"),
                f.get("PrevAmount"),
                f.get("Delta"),
                f.get("Rate"),
                f.get("YoYDelta"),
                f.get("YoYRate"),
                f.get("AnomalyScore"),
                "、".join(f.get("ReasonTags") or []),
                f.get("ActiveMonths12"),
                f.get("AnomalyKey"),
            ])
        if not top_rows:
            top_rows = [[None] * len(top_headers)]
        top_start_row = row_cursor
        self._write_table(ws, top_start_row, 1, top_headers, top_rows)
        top_key_col = 1 + top_headers.index("明细键")
        row_cursor += len(top_rows) + 2

        # D) 异常明细（按异常键）
        ws.cell(row=row_cursor, column=1).value = "D. 异常项目明细（按异常键关联）"
        ws.cell(row=row_cursor, column=1).font = Font(name="微软雅黑", bold=True, color="1F4E79")
        row_cursor += 1

        grouped_details = {}
        for key, group in df.groupby(['MonthStr', 'Category', 'Subcategory']):
            grouped_details[key] = group.sort_values(by='AmountAbs', ascending=False)

        detail_headers = ["明细键", "月份", "部门", "费用类别", "子科目", "摘要", "金额", "异常评分", "异常标签"]
        detail_rows = []
        key_to_detail_row = {}
        detail_header_row = row_cursor
        detail_data_start = detail_header_row + 1
        for f in top_flags:
            detail_key = f.get("AnomalyKey")
            key = (f.get("MonthStr"), f.get("Category"), f.get("Subcategory"))
            g = grouped_details.get(key)
            if g is None or g.empty:
                continue
            take = g.head(max(1, int(detail_lines_per_key)))
            for _, rec in take.iterrows():
                if detail_key and detail_key not in key_to_detail_row:
                    key_to_detail_row[detail_key] = detail_data_start + len(detail_rows)
                detail_rows.append([
                    detail_key,
                    rec.get("MonthLabel"),
                    rec.get("Department"),
                    rec.get("Category"),
                    rec.get("Subcategory"),
                    rec.get("Summary"),
                    rec.get("Amount"),
                    f.get("AnomalyScore"),
                    "、".join(f.get("ReasonTags") or []),
                ])
        if not detail_rows:
            detail_rows = [["-", None, None, None, None, None, None, None, None]]
        self._write_table(ws, detail_header_row, 1, detail_headers, detail_rows)
        row_cursor += len(detail_rows) + 2

        # E) 目标月部门费用贡献
        ws.cell(row=row_cursor, column=1).value = "E. 目标月部门费用贡献"
        ws.cell(row=row_cursor, column=1).font = Font(name="微软雅黑", bold=True, color="1F4E79")
        row_cursor += 1
        dept_headers = ["部门", "销售费用", "管理费用", "财务费用", "其他费用", "费用合计", "占本月费用比", "费用率", "最大费用类别"]
        dept_rows = []
        if target_df is not None and not target_df.empty:
            dept_df = target_df.copy()
            dept_df["Department"] = dept_df["Department"].fillna("未填部门").astype(str).replace({"": "未填部门", "nan": "未填部门"})
            revenue_target = metrics_by_month.get(target_key, {}).get("revenue") if target_key and metrics_by_month.get(target_key) else None
            month_total = dept_df["Amount"].sum()
            for dept, group in dept_df.groupby("Department"):
                cat_map = group.groupby("Category")["Amount"].sum().to_dict()
                sales_amt = cat_map.get("销售费用") or 0
                admin_amt = cat_map.get("管理费用") or 0
                financial_amt = cat_map.get("财务费用") or 0
                total_amt = group["Amount"].sum()
                other_amt = total_amt - sales_amt - admin_amt - financial_amt
                dominant = None
                if cat_map:
                    dominant = max(cat_map.items(), key=lambda kv: abs(kv[1] or 0))[0]
                dept_rows.append([
                    dept,
                    sales_amt,
                    admin_amt,
                    financial_amt,
                    other_amt,
                    total_amt,
                    total_amt / month_total if month_total else None,
                    total_amt / revenue_target if revenue_target else None,
                    dominant,
                ])
        dept_rows = sorted(dept_rows, key=lambda x: abs(x[5] or 0), reverse=True)
        if not dept_rows:
            dept_rows = [["-", None, None, None, None, None, None, None, None]]
        self._write_table(ws, row_cursor, 1, dept_headers, dept_rows)
        row_cursor += len(dept_rows) + 2

        # F) 财务费用拆解
        ws.cell(row=row_cursor, column=1).value = "F. 财务费用拆解"
        ws.cell(row=row_cursor, column=1).font = Font(name="微软雅黑", bold=True, color="1F4E79")
        row_cursor += 1
        finance_headers = ["子科目", "借方发生", "贷方发生", "净额", "占财务费用比", "明细笔数", "主要摘要"]
        finance_rows = []
        if target_df is not None and not target_df.empty:
            finance_df = target_df[target_df["Category"] == "财务费用"].copy()
            finance_total = finance_df["Amount"].sum() if not finance_df.empty else 0
            if not finance_df.empty:
                finance_df["Subcategory"] = finance_df["Subcategory"].fillna("未填子科目").astype(str).replace({"": "未填子科目", "nan": "未填子科目"})
                for sub, group in finance_df.groupby("Subcategory"):
                    debit_sum = group.get("DebitAmount", pd.Series(dtype="float64")).apply(self._to_float).fillna(0).sum()
                    credit_sum = group.get("CreditAmount", pd.Series(dtype="float64")).apply(self._to_float).fillna(0).sum()
                    net_amount = group["Amount"].sum()
                    top_summary = None
                    top_group = group.sort_values(by="AmountAbs", ascending=False)
                    if not top_group.empty:
                        raw_summary = top_group.iloc[0].get("Summary")
                        if raw_summary is not None and not pd.isna(raw_summary):
                            top_summary = str(raw_summary)
                    finance_rows.append([
                        sub,
                        debit_sum,
                        credit_sum,
                        net_amount,
                        net_amount / finance_total if finance_total else None,
                        len(group),
                        top_summary,
                    ])
        finance_rows = sorted(finance_rows, key=lambda x: abs(x[3] or 0), reverse=True)
        if not finance_rows:
            finance_rows = [["-", None, None, None, None, None, None]]
        self._write_table(ws, row_cursor, 1, finance_headers, finance_rows)
        row_cursor += len(finance_rows) + 2

        # G) 年度异常回顾Top
        ws.cell(row=row_cursor, column=1).value = "G. 年度异常回顾Top（非目标月）"
        ws.cell(row=row_cursor, column=1).font = Font(name="微软雅黑", bold=True, color="1F4E79")
        row_cursor += 1
        history_headers = [
            "月份", "费用类别", "子科目", "本期金额", "上期金额",
            "环比增量", "环比增速", "异常评分", "异常标签", "明细键"
        ]
        history_rows = []
        history_flags = [f for f in flags if not target_key or f.get("MonthStr") != target_key][:min(anomaly_top_n, 30)]
        for f in history_flags:
            history_rows.append([
                self._month_key_to_label(f.get("MonthStr")),
                f.get("Category"),
                f.get("Subcategory"),
                f.get("Amount"),
                f.get("PrevAmount"),
                f.get("Delta"),
                f.get("Rate"),
                f.get("AnomalyScore"),
                "、".join(f.get("ReasonTags") or []),
                f.get("AnomalyKey"),
            ])
        if not history_rows:
            history_rows = [["-", None, None, None, None, None, None, None, None, None]]
        self._write_table(ws, row_cursor, 1, history_headers, history_rows)
        row_cursor += len(history_rows) + 2

        budget_targets, _, (ratio_threshold, pp_threshold, days_threshold) = self._read_budget_targets(wb)
        month_keys_warn = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)

        # H) 费用预算偏差
        ws.cell(row=row_cursor, column=1).value = "H. 费用预算偏差"
        ws.cell(row=row_cursor, column=1).font = Font(name="微软雅黑", bold=True, color="1F4E79")
        row_cursor += 1
        budget_headers = ["月份", "指标", "目标值", "实际值", "偏差", "偏差率", "建议"]
        budget_rows = []
        for m_key in month_keys_warn:
            data = metrics_by_month.get(m_key, {})
            revenue = data.get("revenue")
            sales_expense = data.get("sales_expense")
            admin_expense = data.get("admin_expense")
            financial_expense = data.get("financial_expense")
            metric_items = [
                ("销售费用率", "sales_rate", sales_expense / revenue if (sales_expense is not None and revenue not in (None, 0)) else None),
                ("管理费用率", "admin_rate", admin_expense / revenue if (admin_expense is not None and revenue not in (None, 0)) else None),
                ("财务费用率", "financial_rate", financial_expense / revenue if (financial_expense is not None and revenue not in (None, 0)) else None),
            ]
            for label, key, actual in metric_items:
                target = budget_targets.get(m_key, {}).get(key)
                if target is None and actual is None:
                    continue
                variance = actual - target if (actual is not None and target is not None) else None
                variance_rate = variance / target if (variance is not None and target not in (None, 0)) else None
                suggestion = "按费用率偏差追踪责任部门和主要科目" if variance is not None and variance > 0 else ""
                budget_rows.append([
                    self._month_key_to_label(m_key),
                    label,
                    target,
                    actual,
                    variance,
                    variance_rate,
                    suggestion,
                ])
        if not budget_rows:
            budget_rows = [["-", "无预算目标", None, None, None, None, "请在目标_预算中维护费用率目标"]]
        self._write_table(ws, row_cursor, 1, budget_headers, budget_rows)
        row_cursor += len(budget_rows) + 2

        # I) 经营指标预警（预算阈值）
        ws.cell(row=row_cursor, column=1).value = "I. 经营指标预警（预算阈值）"
        ws.cell(row=row_cursor, column=1).font = Font(name="微软雅黑", bold=True, color="1F4E79")
        row_cursor += 1
        metric_warn_headers = ["月份", "指标", "本期值", "上期值", "变化值", "变化率/差值", "阈值", "级别", "建议"]
        metric_warn_rows = []
        prev_data = None
        for m_key in month_keys_warn:
            data = metrics_by_month.get(m_key, {})
            revenue = data.get("revenue")
            profit = data.get("operating_profit")
            cost_rate = data.get("cost_rate")
            sales_expense = data.get("sales_expense")
            admin_expense = data.get("admin_expense")
            financial_expense = data.get("financial_expense")
            sales_rate = sales_expense / revenue if (sales_expense is not None and revenue not in (None, 0)) else None
            admin_rate = admin_expense / revenue if (admin_expense is not None and revenue not in (None, 0)) else None
            financial_rate = financial_expense / revenue if (financial_expense is not None and revenue not in (None, 0)) else None
            inv_start = data.get("inventory_start")
            inv_end = data.get("inventory_end")
            avg_inv = (inv_start + inv_end) / 2 if (inv_start is not None and inv_end is not None) else inv_end
            inv_days = avg_inv / data.get("cost") * 365 if (avg_inv and data.get("cost")) else None
            if prev_data:
                def check_ratio(label, current, previous):
                    if current is None or previous is None:
                        return
                    delta = current - previous
                    rate = delta / previous if previous else None
                    if rate is not None and abs(rate) >= ratio_threshold:
                        metric_warn_rows.append([
                            self._month_key_to_label(m_key), label, current, previous, delta,
                            rate, ratio_threshold, "高", "关注业务波动原因",
                        ])

                def check_pp(label, current, previous):
                    if current is None or previous is None:
                        return
                    delta = current - previous
                    if abs(delta) >= pp_threshold:
                        metric_warn_rows.append([
                            self._month_key_to_label(m_key), label, current, previous, delta,
                            delta, pp_threshold, "中", "检查费用率或成本率变化",
                        ])

                def check_days(label, current, previous):
                    if current is None or previous is None:
                        return
                    delta = current - previous
                    if abs(delta) >= days_threshold:
                        metric_warn_rows.append([
                            self._month_key_to_label(m_key), label, current, previous, delta,
                            delta, days_threshold, "中", "检查周转效率",
                        ])

                check_ratio("主营业务收入", revenue, prev_data.get("revenue"))
                check_ratio("营业利润", profit, prev_data.get("operating_profit"))
                check_ratio("应收账款余额", data.get("ar_balance"), prev_data.get("ar_balance"))
                check_ratio("存货期末余额", inv_end, prev_data.get("inventory_end"))
                check_pp("成本率", cost_rate, prev_data.get("cost_rate"))
                check_pp("销售费用率", sales_rate, prev_data.get("sales_rate"))
                check_pp("管理费用率", admin_rate, prev_data.get("admin_rate"))
                check_pp("财务费用率", financial_rate, prev_data.get("financial_rate"))
                check_days("存货周转天数", inv_days, prev_data.get("inv_days"))

            prev_data = {
                "revenue": revenue,
                "operating_profit": profit,
                "cost_rate": cost_rate,
                "sales_rate": sales_rate,
                "admin_rate": admin_rate,
                "financial_rate": financial_rate,
                "ar_balance": data.get("ar_balance"),
                "inventory_end": inv_end,
                "inv_days": inv_days,
            }

        if not metric_warn_rows:
            metric_warn_rows = [["-", "无异常", None, None, None, None, None, "低", ""]]
        self._write_table(ws, row_cursor, 1, metric_warn_headers, metric_warn_rows)
        self._add_expense_diagnostic_charts(
            ws,
            rows_a,
            top_rows,
            dept_rows,
            finance_rows,
            matrix_rows=matrix_rows,
            subject_compare_rows=subject_compare_rows,
        )

        # 将“明细键”在 B/C 区域链接到 D 区域对应明细行。
        if key_to_detail_row:
            for r in range(matrix_start_row + 1, matrix_start_row + 1 + len(matrix_rows)):
                key_cell = ws.cell(row=r, column=matrix_key_col)
                key_val = key_cell.value
                target_row = key_to_detail_row.get(key_val)
                if key_val and target_row:
                    self._apply_hyperlink(key_cell, f"#'{self.expense_analysis_sheet_name}'!A{target_row}")
                    key_cell.font = Font(color="0563C1", underline="single")
            for r in range(top_start_row + 1, top_start_row + 1 + len(top_rows)):
                key_cell = ws.cell(row=r, column=top_key_col)
                key_val = key_cell.value
                target_row = key_to_detail_row.get(key_val)
                if key_val and target_row:
                    self._apply_hyperlink(key_cell, f"#'{self.expense_analysis_sheet_name}'!A{target_row}")
                    key_cell.font = Font(color="0563C1", underline="single")

    def _update_forecast_sheet(self, wb, metrics_by_month, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "预测与滚动预算", insert_after="异常预警")
        headers = ["月份", "指标", "类型", "数值"]
        month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
        metrics = {
            "主营业务收入": "revenue",
            "营业利润": "operating_profit",
            "主营业务成本": "cost",
            "销售费用": "sales_expense",
            "管理费用": "admin_expense",
            "财务费用": "financial_expense",
            "应收账款余额": "ar_balance",
            "存货期末余额": "inventory_end",
        }

        def linear_forecast(values, steps):
            clean = [(i, v) for i, v in enumerate(values) if v is not None]
            if len(clean) < 2:
                avg = sum(v for _, v in clean) / len(clean) if clean else None
                return [avg for _ in range(steps)]
            n = len(clean)
            xs = [i for i, _ in clean]
            ys = [v for _, v in clean]
            sum_x = sum(xs)
            sum_y = sum(ys)
            sum_xy = sum(x * y for x, y in clean)
            sum_x2 = sum(x * x for x in xs)
            denom = n * sum_x2 - sum_x * sum_x
            if denom == 0:
                avg = sum_y / n
                return [avg for _ in range(steps)]
            slope = (n * sum_xy - sum_x * sum_y) / denom
            intercept = (sum_y - slope * sum_x) / n
            start = len(values)
            return [slope * (start + i) + intercept for i in range(steps)]

        rows = []
        for label, key in metrics.items():
            series = []
            for m_key in month_keys:
                series.append(metrics_by_month.get(m_key, {}).get(key))
                rows.append([self._month_key_to_label(m_key), label, "实际", metrics_by_month.get(m_key, {}).get(key)])
            forecasts = linear_forecast(series[-3:], 3)
            for idx, value in enumerate(forecasts, start=1):
                year, month = self._add_months(int(target_year), int(target_month), idx)
                rows.append([f"{year}/{month:02d}", label, "预测", value])

        if not rows:
            rows = [[None] * len(headers)]
        self._write_table(ws, 1, 1, headers, rows)

        # Chart for revenue and profit forecast
        chart_rows = []
        for idx, m_key in enumerate(month_keys):
            chart_rows.append([
                self._month_key_to_label(m_key),
                metrics_by_month.get(m_key, {}).get("revenue"),
                metrics_by_month.get(m_key, {}).get("operating_profit"),
            ])
        for i in range(1, 4):
            year, month = self._add_months(int(target_year), int(target_month), i)
            label = f"{year}/{month:02d}"
            series = [metrics_by_month.get(m_key, {}).get("revenue") for m_key in month_keys[-3:]]
            forecast_revenue = linear_forecast(series[-3:], 3)[i - 1] if series else None
            series_p = [metrics_by_month.get(m_key, {}).get("operating_profit") for m_key in month_keys[-3:]]
            forecast_profit = linear_forecast(series_p[-3:], 3)[i - 1] if series_p else None
            chart_rows.append([label, forecast_revenue, forecast_profit])

        if chart_rows:
            start_col = 6
            self._write_table(ws, 1, start_col, ["月份", "收入", "利润"], chart_rows)
            anchor = f"{get_column_letter(start_col + 4)}2"
            self._add_line_chart_by_columns(ws, start_col, [start_col + 1, start_col + 2], 1, 2, 1 + len(chart_rows), "收入/利润预测", anchor)
            self._write_chart_note(ws, start_col + 4, 1, "图表说明：收入与利润的滚动预测趋势。")

        header_map = self._get_header_map(ws, 1)
        indicator_col = header_map.get("指标")
        self._reorder_month_rows_desc(ws, group_by_cols=[indicator_col] if indicator_col else None)

    def _discover_entity_dirs(self):
        if not os.path.exists(self.base_data_dir):
            return []
        entries = []
        for name in os.listdir(self.base_data_dir):
            path = os.path.join(self.base_data_dir, name)
            if not os.path.isdir(path):
                continue
            if name.startswith("."):
                continue
            if any(f.endswith(".xlsx") or f.endswith(".xls") for f in os.listdir(path)):
                entries.append(path)
        return entries

    def _update_entity_summary_sheet(self, wb, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "多主体汇总", insert_after="预测与滚动预算")
        entity_dirs = self._discover_entity_dirs()
        if not entity_dirs:
            self._write_table(ws, 1, 1, ["提示"], [["未检测到多主体数据目录"]])
            return
        headers = ["主体", "主营业务收入", "主营业务成本", "营业利润", "销售费用", "管理费用", "财务费用", "应收账款余额", "存货期末余额"]
        rows = []
        for path in entity_dirs:
            name = os.path.basename(path)
            sub = ReportGenerator(path)
            sub.load_all_data()
            metrics = sub._build_monthly_metrics(target_year, target_month, year_scope)
            key = f"{target_year}-{int(target_month):02d}"
            data = metrics.get(key, {})
            rows.append([
                name,
                data.get("revenue"),
                data.get("cost"),
                data.get("operating_profit"),
                data.get("sales_expense"),
                data.get("admin_expense"),
                data.get("financial_expense"),
                data.get("ar_balance"),
                data.get("inventory_end"),
            ])
        rows = sorted(rows, key=lambda x: x[1] if x[1] is not None else 0, reverse=True)
        self._write_table(ws, 1, 1, headers, rows)
        if rows:
            start_col = len(headers) + 2
            chart_rows = [[r[0], r[1]] for r in rows]
            self._write_table(ws, 1, start_col, ["主体", "收入"], chart_rows)
            anchor = f"{get_column_letter(start_col + 3)}2"
            self._add_bar_chart_from_table(ws, 1, start_col, 1 + len(chart_rows), "主体收入对比", anchor)
            self._write_chart_note(ws, start_col + 3, 1, "图表说明：多主体收入对比。")

    def _update_currency_summary_sheet(self, wb, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "币种汇总", insert_after="多主体汇总")
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无销售数据"]])
            return
        df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, year_scope)
        currency_col = None
        for col in df.columns:
            if "币种" in str(col) or "Currency" in str(col) or "币别" in str(col):
                currency_col = col
                break
        if not currency_col:
            self._write_table(ws, 1, 1, ["提示"], [["未检测到币种字段"]])
            return
        df['Revenue'] = self._extract_sales_revenue(df)
        df['RevenueFC'] = pd.to_numeric(df.get('销售出库外币金额合计'), errors='coerce') if '销售出库外币金额合计' in df.columns else None
        df = self._attach_sales_cost(df, target_year, target_month, year_scope)
        rows = []
        for cur, group in df.groupby(currency_col):
            metrics = self._calc_sales_metrics_from_group(group)
            revenue = metrics.get('revenue')
            cost = metrics.get('cost')
            profit = metrics.get('profit')
            revenue_fc = self._sum_numeric_or_none(group.get('RevenueFC')) if 'RevenueFC' in group else None
            implied_rate = (revenue / revenue_fc) if (revenue is not None and revenue_fc not in (None, 0)) else None
            rows.append([cur, revenue, cost, profit, metrics.get('margin'), revenue_fc, implied_rate])
        headers = ["币种", "销售收入", "销售成本", "毛利润", "毛利率", "外币收入", "隐含汇率"]
        rows = sorted(rows, key=lambda x: x[1] if x[1] is not None else 0, reverse=True)
        self._write_table(ws, 1, 1, headers, rows)

    def _add_drilldown_links(self, wb):
        link_map = {
            "经营指标": ["利润表", self.expense_analysis_sheet_name],
            "费用对比": [self.expense_analysis_sheet_name],
            "费用明细": [self.expense_detail_sheet_name],
            "预算执行与偏差": ["目标_预算"],
            "产品贡献毛利": ["按产品汇总_含合计"],
            "品类贡献毛利": ["按品类汇总(按月)"],
            "客户贡献与回款": ["应收账款账龄分析"],
            "渠道贡献": ["客户贡献与回款"],
            "存货健康度": ["明细_销售与库存"],
            "货币资金分析": ["现金流量表(估算)", "资产负债表"],
            "应付账款分析": ["资金链预警", "资产负债表"],
        }
        for sheet, targets in link_map.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            col = ws.max_column + 1
            row = 1
            for target in targets:
                if target not in wb.sheetnames:
                    continue
                cell = ws.cell(row=row, column=col)
                cell.value = f"跳转: {target}"
                cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{target}'!A1")
                cell.style = "Hyperlink"
                row += 1

    def _update_directory_sheet(self, wb):
        if "目录" not in wb.sheetnames:
            return
        ws = wb["目录"]

        # 先修复目录历史错链：A列标题应始终指向同名Sheet，不存在则清空链接。
        for r in range(1, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            link_cell = ws.cell(row=r, column=1)
            if name is None or str(name).strip() == "":
                link_cell.hyperlink = None
                continue
            sheet_name = str(name).strip()
            if sheet_name in wb.sheetnames:
                link_cell.hyperlink = Hyperlink(ref=link_cell.coordinate, location=f"'{sheet_name}'!A1")
                link_cell.style = "Hyperlink"
            else:
                link_cell.hyperlink = None

        desired = [
            (self.expense_analysis_sheet_name, "费用趋势/环比/异常/明细一页整合", "费用分析主入口"),
            (self.expense_detail_sheet_name, "费用异常明细明细清单（按异常键）", "费用明细关联入口"),
            ("货币资金分析", "银行/现金明细与资产负债表货币资金覆盖分析", "资金分析入口"),
            ("应付账款分析", "应付明细、供应商余额与报表应付覆盖分析", "负债分析入口"),
            ("连续年度分析", "按连续年份汇总财务指标、产品贡献与品类结构", "跨年度分析入口"),
            ("年度汇总总览", "年度利润/指标/资产负债/费用异常一页汇总", "年度汇报入口"),
        ]

        existing_rows = {}
        last_used_row = 1
        for r in range(1, ws.max_row + 1):
            row_has_value = False
            for c in range(1, 4):
                val = ws.cell(row=r, column=c).value
                if val not in (None, ""):
                    row_has_value = True
            if row_has_value:
                last_used_row = max(last_used_row, r)

            name = ws.cell(row=r, column=1).value
            if name not in (None, ""):
                existing_rows[str(name).strip()] = r

        next_row = last_used_row + 1
        for sheet_name, desc, tip in desired:
            if sheet_name not in wb.sheetnames:
                continue

            row_idx = existing_rows.get(sheet_name)
            if row_idx is None:
                row_idx = next_row
                next_row += 1
                if row_idx > 1:
                    try:
                        self._copy_row_style(ws, row_idx - 1, row_idx, max_col=3)
                    except Exception:
                        pass
                ws.cell(row=row_idx, column=1).value = sheet_name
                ws.cell(row=row_idx, column=2).value = desc
                ws.cell(row=row_idx, column=3).value = tip
                existing_rows[sheet_name] = row_idx
            else:
                if ws.cell(row=row_idx, column=2).value in (None, ""):
                    ws.cell(row=row_idx, column=2).value = desc
                if ws.cell(row=row_idx, column=3).value in (None, ""):
                    ws.cell(row=row_idx, column=3).value = tip

            link_cell = ws.cell(row=row_idx, column=1)
            link_cell.hyperlink = Hyperlink(ref=link_cell.coordinate, location=f"'{sheet_name}'!A1")
            link_cell.style = "Hyperlink"

    def _generate_extended_reports(self, wb, metrics_by_month_scoped, target_year, target_month, metrics_by_month_all=None):
        self._log_audit("开始生成扩展报表")
        self._update_dashboard(wb, metrics_by_month_scoped, target_year, target_month)
        self._update_cashflow_sheet(wb, target_year, target_month, self.year_scope)
        self._update_cash_balance_analysis_sheet(wb, metrics_by_month_scoped, target_year, target_month, self.year_scope)
        self._update_budget_variance_sheet(wb, metrics_by_month_scoped, target_year, target_month, self.year_scope)
        self._update_product_contribution_sheet(wb, target_year, target_month, self.year_scope)
        self._update_category_contribution_sheet(wb, target_year, target_month, self.year_scope)
        self._update_customer_profit_sheet(wb, target_year, target_month, self.year_scope)
        self._update_channel_profit_sheet(wb, target_year, target_month, self.year_scope)
        self._update_inventory_health_sheet(wb, metrics_by_month_scoped, target_year, target_month, self.year_scope)
        self._update_slow_moving_inventory_sheet(wb, target_year, target_month)
        repl_auto_by_product = not self._bool_param(
            (self.report_params or {}).get("replenishment", {}).get("manual", True),
            default=True,
        )
        repl_params, cash_params, warning_notes = self._resolve_warning_params(
            metrics_by_month_scoped,
            target_year,
            target_month,
            self.year_scope,
        )
        self._update_replenishment_alert_sheet(
            wb,
            target_year,
            target_month,
            self.year_scope,
            param_source_note=warning_notes.get("replenishment"),
            use_product_params=repl_auto_by_product,
            **repl_params,
        )
        self._update_expense_structure_sheet(wb, metrics_by_month_scoped, target_year, target_month, self.year_scope)
        self._update_anomaly_sheet(wb, metrics_by_month_scoped, target_year, target_month, self.year_scope)
        self._update_expense_diagnostic_center(wb, metrics_by_month_scoped, target_year, target_month, self.year_scope)
        self._update_forecast_sheet(wb, metrics_by_month_scoped, target_year, target_month, self.year_scope)
        self._update_entity_summary_sheet(wb, target_year, target_month, self.year_scope)
        self._update_currency_summary_sheet(wb, target_year, target_month, self.year_scope)

        self._update_cashflow_alert_sheet(
            wb,
            metrics_by_month_scoped,
            target_year,
            target_month,
            self.year_scope,
            param_source_note=warning_notes.get("cashflow"),
            **cash_params,
        )
        self._update_ap_analysis_sheet(wb, metrics_by_month_scoped, target_year, target_month, self.year_scope)

        self._update_staff_efficiency_sheet(wb, target_year, target_month, self.year_scope)
        self._update_dupont_sheet(wb, metrics_by_month_scoped, target_year, target_month, self.year_scope)
        self._update_customer_churn_sheet(wb, target_year, target_month, self.year_scope)

        self._generate_annual_reports(wb, metrics_by_month_scoped, target_year, target_month, self.year_scope)
        self._update_continuous_year_analysis_sheet(
            wb,
            metrics_by_month_all or self._build_monthly_metrics(target_year, target_month, "all"),
            target_year,
            target_month,
        )
        self._write_data_quality_sheet(wb)
        self._write_audit_log_sheet(wb)
        # 已合并且失效的表直接删除；费用明细保留作为明细底表。
        removed_sheet_names = [
            "同比经营分析",
            "环比经营分析",
            "品类贡献毛利",
            "费用明细环比分析",
            "费用结构与弹性",
            "异常预警",
            "年度费用异常Top",
        ]
        self._delete_sheets_if_exist(wb, removed_sheet_names)
        self._remove_directory_entries(wb, removed_sheet_names)
        self._update_directory_sheet(wb)
        self._add_drilldown_links(wb)
        self._log_audit("扩展报表生成完成")

    def _annual_month_keys(self, target_year):
        return [f"{target_year}-{m:02d}" for m in range(1, 13)]

    def _annual_data_gaps(self, target_year):
        expected = self._annual_month_keys(target_year)
        profit_missing = [m for m in expected if m not in self.data.get("profit", {})]
        asset_missing = [m for m in expected if m not in self.data.get("asset", {})]
        return {
            "profit": profit_missing,
            "asset": asset_missing,
        }

    def _format_month_keys(self, month_keys):
        if not month_keys:
            return "无"
        return "、".join(self._month_key_to_label(m) for m in month_keys)

    def _write_annual_data_gap_note(self, ws, row, gaps):
        if not gaps:
            return
        profit_missing = gaps.get("profit") or []
        asset_missing = gaps.get("asset") or []
        parts = [
            f"利润表缺月: {self._format_month_keys(profit_missing)}",
            f"资产负债表缺月: {self._format_month_keys(asset_missing)}",
        ]
        ws.cell(row=row, column=1).value = "数据完整性：" + "；".join(parts)

    def _continuous_years_for_report(self, target_year=None, target_month=None):
        month_keys = set()
        for category in ("profit", "asset", "sales", "cost"):
            for key in self.data.get(category, {}).keys():
                key = str(key or "").strip()
                if self._is_valid_month_key(key):
                    month_keys.add(key)

        sales_df = self._get_sales_df()
        if sales_df is not None and not sales_df.empty:
            work = sales_df.copy()
            if "MonthStr" not in work.columns and "ParsedDate" in work.columns:
                work["MonthStr"] = pd.to_datetime(work["ParsedDate"], errors="coerce").dt.strftime("%Y-%m")
            if "MonthStr" in work.columns:
                for key in work["MonthStr"].dropna().astype(str).unique():
                    key = str(key).strip()
                    if self._is_valid_month_key(key):
                        month_keys.add(key)

        if target_year and target_month:
            limit_key = f"{int(target_year):04d}-{int(target_month):02d}"
            month_keys = {key for key in month_keys if key <= limit_key}

        if not month_keys:
            return [int(target_year)] if target_year else []

        min_year = min(int(key[:4]) for key in month_keys)
        max_year = int(target_year) if target_year else max(int(key[:4]) for key in month_keys)
        return list(range(min_year, max_year + 1))

    def _annual_expected_months(self, year, target_year=None, target_month=None):
        end_month = 12
        if target_year and target_month and int(year) == int(target_year):
            end_month = int(target_month)
        return [f"{int(year):04d}-{month:02d}" for month in range(1, end_month + 1)]

    def _average_or_none(self, values):
        vals = [v for v in values if v is not None]
        return sum(vals) / len(vals) if vals else None

    def _safe_change_rate(self, current, previous):
        if current is None or previous in (None, 0):
            return None
        return (current - previous) / previous

    def _first_non_empty_value(self, series):
        if series is None:
            return None
        for value in series:
            if value is None or pd.isna(value):
                continue
            text = str(value).strip()
            if text and text.lower() != "nan":
                return value
        return None

    def _last_non_none_value(self, values):
        for value in reversed(list(values or [])):
            if value is not None:
                return value
        return None

    def _format_continuous_year_sheet(self, ws):
        percent_tokens = ("毛利率", "费用率", "利润率", "净利率", "占比", "同比", "变动")
        amount_tokens = ("收入", "成本", "利润", "费用", "应收", "存货", "金额", "资产", "资金")
        for row in range(1, ws.max_row + 1):
            headers = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
            if not any(headers):
                continue
            for col, header in enumerate(headers, start=1):
                text = str(header or "")
                if not text:
                    continue
                for data_row in range(row + 1, ws.max_row + 1):
                    first_val = ws.cell(row=data_row, column=1).value
                    if isinstance(first_val, str) and first_val.startswith("说明"):
                        break
                    cell = ws.cell(row=data_row, column=col)
                    if cell.value is None:
                        continue
                    if "周转天数" in text or text in {"DSO", "DIO", "DPO", "CCC"}:
                        cell.number_format = "0.0"
                    elif "周转率" in text:
                        cell.number_format = "0.00"
                    elif any(token in text for token in percent_tokens):
                        cell.number_format = "0.00%"
                    elif any(token in text for token in amount_tokens):
                        cell.number_format = '#,##0.00'

        for col in range(1, min(ws.max_column, 18) + 1):
            letter = get_column_letter(col)
            if col in (2, 3):
                ws.column_dimensions[letter].width = 14
            elif col in (4, 5, 6):
                ws.column_dimensions[letter].width = 16
            else:
                ws.column_dimensions[letter].width = 12
        ws.freeze_panes = "A4"

    def _prepare_continuous_year_sales_df(self, target_year, target_month):
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            return pd.DataFrame()

        df = sales_df.copy()
        if "MonthStr" not in df.columns:
            if "ParsedDate" in df.columns:
                df["MonthStr"] = pd.to_datetime(df["ParsedDate"], errors="coerce").dt.strftime("%Y-%m")
            else:
                date_col = next((c for c in df.columns if "日期" in str(c) or "Date" in str(c)), None)
                if date_col:
                    df["ParsedDate"] = pd.to_datetime(df[date_col], errors="coerce")
                    df["MonthStr"] = df["ParsedDate"].dt.strftime("%Y-%m")
        if "MonthStr" not in df.columns:
            return pd.DataFrame()

        df["MonthStr"] = df["MonthStr"].astype(str).str.strip()
        df = df[df["MonthStr"].apply(self._is_valid_month_key)].copy()
        df = self._filter_df_by_scope(df, target_year, target_month, "all")
        if df is None or df.empty:
            return pd.DataFrame()

        df["Revenue"] = self._extract_sales_revenue(df)
        df["Qty"] = pd.to_numeric(df.get("数量"), errors="coerce")
        df = self._attach_sales_cost(df, target_year, target_month, "all")
        df["Year"] = df["MonthStr"].astype(str).str[:4].astype(int)

        code_col = "品目编码" if "品目编码" in df.columns else None
        name_col = next((c for c in ["品目名", "产品名称", "品目名规格"] if c in df.columns), None)
        category_col = "品目组合1名" if "品目组合1名" in df.columns else None
        df["ProductCode"] = df[code_col].astype(str).str.strip() if code_col else ""
        df["ProductName"] = df[name_col].astype(str).str.strip() if name_col else ""
        df["Category"] = df[category_col].astype(str).str.strip() if category_col else "未分类"
        df.loc[df["ProductCode"].str.lower().isin(["nan", "none"]), "ProductCode"] = ""
        df.loc[df["ProductName"].str.lower().isin(["nan", "none"]), "ProductName"] = ""
        df.loc[df["Category"].str.lower().isin(["", "nan", "none"]), "Category"] = "未分类"
        df["ProductKey"] = df["ProductCode"].where(df["ProductCode"] != "", df["ProductName"])
        df = df[df["ProductKey"].astype(str).str.strip() != ""].copy()
        return df

    def _build_continuous_year_financial_rows(self, metrics_by_month, years, target_year, target_month):
        rows = []
        annual_by_year = {}
        for year in years:
            expected_months = self._annual_expected_months(year, target_year, target_month)
            available_months = [m for m in expected_months if m in metrics_by_month]

            def vals(key):
                return [metrics_by_month.get(m, {}).get(key) for m in expected_months]

            def avg_metric(key):
                return self._average_or_none(vals(key))

            def sum_components(month_key, component_keys):
                data = metrics_by_month.get(month_key, {})
                components = [data.get(key) for key in component_keys]
                if not any(value is not None for value in components):
                    return None
                return sum(value or 0 for value in components)

            revenue = self._safe_sum(vals("revenue"))
            cost = self._safe_sum(vals("cost"))
            gross_profit = self._safe_sum(vals("gross_profit"))
            if gross_profit is None and revenue is not None and cost is not None:
                gross_profit = revenue - cost
            sales_expense = self._safe_sum(vals("sales_expense"))
            admin_expense = self._safe_sum(vals("admin_expense"))
            financial_expense = self._safe_sum(vals("financial_expense"))
            total_expense = self._safe_sum([sales_expense, admin_expense, financial_expense])
            operating_profit = self._safe_sum(vals("operating_profit"))
            net_profit = self._safe_sum(vals("net_profit"))
            ar_balance = self._last_non_none_value(vals("ar_balance"))
            inventory_end = self._last_non_none_value(vals("inventory_end"))
            avg_ar = avg_metric("ar_balance")
            avg_ap = avg_metric("ap_balance")
            avg_inventory = avg_metric("inventory_end")
            avg_total_assets = avg_metric("total_assets")

            avg_current_assets = avg_metric("current_assets")
            if avg_current_assets is None:
                current_asset_values = [
                    sum_components(m, ["cash", "ar_balance", "inventory_end"])
                    for m in expected_months
                ]
                avg_current_assets = self._average_or_none(current_asset_values)

            operating_working_capital_values = [
                sum_components(m, ["ar_balance", "inventory_end", "cash"])
                for m in expected_months
            ]
            ap_values = vals("ap_balance")
            working_capital_values = []
            for assets_value, ap_value in zip(operating_working_capital_values, ap_values):
                if assets_value is None and ap_value is None:
                    working_capital_values.append(None)
                else:
                    working_capital_values.append((assets_value or 0) - (ap_value or 0))
            avg_working_capital = self._average_or_none(working_capital_values)

            period_days = sum(self._month_days(m) for m in available_months) if available_months else 365
            inventory_turnover = cost / avg_inventory if cost not in (None, 0) and avg_inventory not in (None, 0) else None
            inventory_turnover_days = (avg_inventory / cost * period_days) if avg_inventory is not None and cost not in (None, 0) else None
            ar_turnover = revenue / avg_ar if revenue not in (None, 0) and avg_ar not in (None, 0) else None
            dso = (avg_ar / revenue * period_days) if avg_ar is not None and revenue not in (None, 0) else None
            ap_turnover = cost / avg_ap if cost not in (None, 0) and avg_ap not in (None, 0) else None
            dpo = (avg_ap / cost * period_days) if avg_ap is not None and cost not in (None, 0) else None
            ccc = (dso + inventory_turnover_days - dpo) if dso is not None and inventory_turnover_days is not None and dpo is not None else None
            total_asset_turnover = revenue / avg_total_assets if revenue not in (None, 0) and avg_total_assets not in (None, 0) else None
            current_fund_turnover = revenue / avg_current_assets if revenue not in (None, 0) and avg_current_assets not in (None, 0) else None
            working_capital_turnover = revenue / avg_working_capital if revenue not in (None, 0) and avg_working_capital not in (None, 0) else None

            gross_margin = gross_profit / revenue if revenue not in (None, 0) and gross_profit is not None else None
            sales_expense_rate = sales_expense / revenue if revenue not in (None, 0) and sales_expense is not None else None
            admin_expense_rate = admin_expense / revenue if revenue not in (None, 0) and admin_expense is not None else None
            financial_expense_rate = financial_expense / revenue if revenue not in (None, 0) and financial_expense is not None else None
            total_expense_rate = total_expense / revenue if revenue not in (None, 0) and total_expense is not None else None
            operating_margin = operating_profit / revenue if revenue not in (None, 0) and operating_profit is not None else None
            net_margin = net_profit / revenue if revenue not in (None, 0) and net_profit is not None else None
            month_range = (
                f"{self._month_key_to_label(available_months[0])}-{self._month_key_to_label(available_months[-1])}"
                if available_months else "无财务月份"
            )

            annual_by_year[year] = {
                "revenue": revenue,
                "net_profit": net_profit,
                "gross_margin": gross_margin,
            }
            prev = annual_by_year.get(year - 1, {})
            rows.append([
                year,
                f"{len(available_months)}/{len(expected_months)}",
                month_range,
                revenue,
                cost,
                gross_profit,
                gross_margin,
                sales_expense,
                sales_expense_rate,
                admin_expense,
                admin_expense_rate,
                financial_expense,
                financial_expense_rate,
                total_expense,
                total_expense_rate,
                operating_profit,
                operating_margin,
                net_profit,
                net_margin,
                ar_balance,
                inventory_end,
                avg_total_assets,
                avg_current_assets,
                avg_working_capital,
                total_asset_turnover,
                current_fund_turnover,
                working_capital_turnover,
                ar_turnover,
                dso,
                inventory_turnover,
                inventory_turnover_days,
                ap_turnover,
                dpo,
                ccc,
                self._safe_change_rate(revenue, prev.get("revenue")),
                self._safe_change_rate(net_profit, prev.get("net_profit")),
                (gross_margin - prev.get("gross_margin")) if gross_margin is not None and prev.get("gross_margin") is not None else None,
            ])
        return rows

    def _build_continuous_year_product_tables(self, years, target_year, target_month, top_n=5):
        df = self._prepare_continuous_year_sales_df(target_year, target_month)
        if df.empty:
            return [], [], []

        product_group = df.groupby(["Year", "ProductKey"], dropna=False).agg(
            product_code=("ProductCode", self._first_non_empty_value),
            product_name=("ProductName", self._first_non_empty_value),
            category=("Category", self._first_non_empty_value),
            qty=("Qty", self._sum_numeric_or_none),
            revenue=("Revenue", self._sum_numeric_or_none),
            cost=("Cost", self._sum_numeric_or_none),
        ).reset_index()
        product_group["profit"] = product_group.apply(
            lambda r: (r["revenue"] - r["cost"])
            if pd.notna(r.get("revenue")) and pd.notna(r.get("cost"))
            else None,
            axis=1,
        )
        product_group["margin"] = product_group.apply(
            lambda r: r["profit"] / r["revenue"]
            if pd.notna(r.get("profit")) and pd.notna(r.get("revenue")) and r["revenue"] != 0
            else None,
            axis=1,
        )

        category_group = df.groupby(["Year", "Category"], dropna=False).agg(
            qty=("Qty", self._sum_numeric_or_none),
            revenue=("Revenue", self._sum_numeric_or_none),
            cost=("Cost", self._sum_numeric_or_none),
        ).reset_index()
        category_group["profit"] = category_group.apply(
            lambda r: (r["revenue"] - r["cost"])
            if pd.notna(r.get("revenue")) and pd.notna(r.get("cost"))
            else None,
            axis=1,
        )
        category_group["margin"] = category_group.apply(
            lambda r: r["profit"] / r["revenue"]
            if pd.notna(r.get("profit")) and pd.notna(r.get("revenue")) and r["revenue"] != 0
            else None,
            axis=1,
        )

        overview_rows = []
        product_top_rows = []
        category_top_rows = []
        for year in years:
            year_products = product_group[product_group["Year"] == year].copy()
            year_categories = category_group[category_group["Year"] == year].copy()
            total_qty = self._sum_numeric_or_none(year_products.get("qty"))
            total_revenue = self._sum_numeric_or_none(year_products.get("revenue"))
            total_cost = self._sum_numeric_or_none(year_products.get("cost"))
            total_profit = (total_revenue - total_cost) if total_revenue is not None and total_cost is not None else None
            total_margin = total_profit / total_revenue if total_profit is not None and total_revenue not in (None, 0) else None

            year_products["_sort_revenue"] = pd.to_numeric(year_products.get("revenue"), errors="coerce").fillna(0)
            year_categories["_sort_revenue"] = pd.to_numeric(year_categories.get("revenue"), errors="coerce").fillna(0)
            top_product = year_products.sort_values("_sort_revenue", ascending=False).head(1)
            top_category = year_categories.sort_values("_sort_revenue", ascending=False).head(1)

            top_product_name = None
            top_product_revenue = None
            top_product_share = None
            if not top_product.empty:
                row = top_product.iloc[0]
                top_product_name = row.get("product_name") or row.get("product_code") or row.get("ProductKey")
                top_product_revenue = row.get("revenue")
                top_product_share = top_product_revenue / total_revenue if total_revenue not in (None, 0) and top_product_revenue is not None else None

            top_category_name = None
            top_category_revenue = None
            top_category_share = None
            if not top_category.empty:
                row = top_category.iloc[0]
                top_category_name = row.get("Category")
                top_category_revenue = row.get("revenue")
                top_category_share = top_category_revenue / total_revenue if total_revenue not in (None, 0) and top_category_revenue is not None else None

            overview_rows.append([
                year,
                int(year_products["ProductKey"].nunique()) if not year_products.empty else 0,
                int(year_categories["Category"].nunique()) if not year_categories.empty else 0,
                total_qty,
                total_revenue,
                total_cost,
                total_profit,
                total_margin,
                top_product_name,
                top_product_revenue,
                top_product_share,
                top_category_name,
                top_category_revenue,
                top_category_share,
            ])

            if not year_products.empty:
                top_products = year_products.sort_values(["_sort_revenue", "ProductKey"], ascending=[False, True]).head(top_n)
                for rank, (_, row) in enumerate(top_products.iterrows(), start=1):
                    revenue = row.get("revenue")
                    product_top_rows.append([
                        year,
                        rank,
                        row.get("product_code"),
                        row.get("product_name") or row.get("product_code") or row.get("ProductKey"),
                        row.get("category"),
                        row.get("qty"),
                        revenue,
                        row.get("cost"),
                        row.get("profit"),
                        row.get("margin"),
                        revenue / total_revenue if total_revenue not in (None, 0) and revenue is not None else None,
                    ])

            if not year_categories.empty:
                top_categories = year_categories.sort_values(["_sort_revenue", "Category"], ascending=[False, True]).head(top_n)
                for rank, (_, row) in enumerate(top_categories.iterrows(), start=1):
                    revenue = row.get("revenue")
                    category_top_rows.append([
                        year,
                        rank,
                        row.get("Category"),
                        row.get("qty"),
                        revenue,
                        row.get("cost"),
                        row.get("profit"),
                        row.get("margin"),
                        revenue / total_revenue if total_revenue not in (None, 0) and revenue is not None else None,
                    ])

        return overview_rows, product_top_rows, category_top_rows

    def _update_continuous_year_analysis_sheet(self, wb, metrics_by_month, target_year, target_month):
        if not target_year or not target_month:
            return

        insert_after = "年度汇总总览" if "年度汇总总览" in wb.sheetnames else "经营指标"
        ws = self._prepare_sheet(wb, "连续年度分析", insert_after=insert_after)
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
        ws._charts = []

        years = self._continuous_years_for_report(target_year, target_month)
        if not years:
            self._write_table(ws, 1, 1, ["提示"], [["无可用于连续年度分析的数据"]])
            return

        ws.cell(row=1, column=1).value = "连续年度分析"
        ws.cell(row=1, column=1).font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
        ws.cell(row=2, column=1).value = f"分析期间：{years[0]}年 至 {years[-1]}年，截至 {int(target_year)}年{int(target_month):02d}月"

        financial_headers = [
            "年份", "覆盖月份", "月份范围", "主营业务收入", "主营业务成本", "毛利润", "毛利率",
            "销售费用", "销售费用率", "管理费用", "管理费用率", "财务费用", "财务费用率",
            "三项费用合计", "三项费用率", "营业利润", "营业利润率", "净利润", "净利率",
            "年末应收账款", "年末存货", "平均总资产", "平均流动资金", "平均营运资金",
            "总资产周转率", "流动资金周转率", "营运资金周转率",
            "应收账款周转率", "DSO", "存货周转率", "存货周转天数",
            "应付账款周转率", "DPO", "CCC", "收入同比", "净利润同比", "毛利率变动",
        ]
        financial_rows = self._build_continuous_year_financial_rows(
            metrics_by_month or {},
            years,
            target_year,
            target_month,
        )
        row_cursor = 4
        self._write_table(ws, row_cursor, 1, financial_headers, financial_rows)
        if financial_rows:
            chart_anchor = f"{get_column_letter(len(financial_headers) + 2)}4"
            self._add_line_chart_by_columns(
                ws,
                1,
                [
                    financial_headers.index("主营业务收入") + 1,
                    financial_headers.index("营业利润") + 1,
                    financial_headers.index("净利润") + 1,
                ],
                row_cursor,
                row_cursor + 1,
                row_cursor + len(financial_rows),
                "连续年度收入与利润趋势",
                chart_anchor,
            )
        row_cursor += len(financial_rows) + 3

        overview_rows, product_top_rows, category_top_rows = self._build_continuous_year_product_tables(
            years,
            target_year,
            target_month,
        )
        product_overview_headers = [
            "年份", "活跃产品数", "活跃品类数", "销量", "销售收入", "销售成本", "毛利润", "毛利率",
            "收入最高产品", "最高产品收入", "最高产品占比", "收入最高品类", "最高品类收入", "最高品类占比",
        ]
        self._write_table(ws, row_cursor, 1, product_overview_headers, overview_rows or [[year, 0, 0, None, None, None, None, None, None, None, None, None, None, None] for year in years])
        if overview_rows:
            self._add_line_chart_by_columns(
                ws,
                1,
                [5, 7],
                row_cursor,
                row_cursor + 1,
                row_cursor + len(overview_rows),
                "连续年度产品收入与毛利",
                "V22",
            )
        row_cursor += max(len(overview_rows), len(years)) + 3

        product_headers = [
            "年份", "排名", "品目编码", "产品名称", "品类", "销量", "销售收入", "销售成本", "毛利润", "毛利率", "收入占比",
        ]
        self._write_table(ws, row_cursor, 1, product_headers, product_top_rows or [["无产品数据", None, None, None, None, None, None, None, None, None, None]])
        row_cursor += max(len(product_top_rows), 1) + 3

        category_headers = [
            "年份", "排名", "品类", "销量", "销售收入", "销售成本", "毛利润", "毛利率", "收入占比",
        ]
        self._write_table(ws, row_cursor, 1, category_headers, category_top_rows or [["无品类数据", None, None, None, None, None, None, None, None]])
        row_cursor += max(len(category_top_rows), 1) + 2

        self._write_chart_note(
            ws,
            1,
            row_cursor,
            "说明：本表按连续年份汇总；目标年度仅统计至报告月份，历史年度统计已加载的全年数据。产品成本优先取成本表单价，缺失时使用销售出库供应价。",
        )
        self._format_continuous_year_sheet(ws)
        self._log_audit("连续年度分析报表生成完成")

    def _generate_annual_reports(self, wb, metrics_by_month, target_year, target_month, year_scope=None):
        if not target_year or not target_month:
            return
        try:
            if int(target_month) != 12:
                return
        except Exception:
            return
        self._log_audit("开始生成年度报表")
        self._update_annual_profit_sheet(wb, metrics_by_month, target_year)
        self._update_annual_metrics_sheet(wb, metrics_by_month, target_year)
        self._update_annual_balance_sheet(wb, target_year)
        self._update_annual_expense_anomaly_sheet(wb, target_year, target_month, year_scope)
        self._update_annual_summary_sheet(wb, metrics_by_month, target_year, target_month, year_scope)
        self._log_audit("年度报表生成完成")

    def _update_annual_profit_sheet(self, wb, metrics_by_month, target_year):
        ws = self._prepare_sheet(wb, "年度利润表", insert_after="利润表")
        months = [f"{target_year}-{m:02d}" for m in range(1, 13)]
        headers = ["项目"] + [f"{target_year}/{m:02d}" for m in range(1, 13)] + ["全年合计"]
        rows = []
        items = [
            ("主营业务收入", "revenue", True),
            ("主营业务成本", "cost", True),
            ("毛利润", "gross_profit", True),
            ("税金及附加", "tax_surcharges", False),
            ("销售费用", "sales_expense", True),
            ("管理费用", "admin_expense", True),
            ("研发费用", "rd_expense", False),
            ("财务费用", "financial_expense", False),
            ("其他收益", "other_income", False),
            ("投资收益", "investment_income", False),
            ("信用减值损失", "credit_impairment_loss", False),
            ("资产减值损失", "asset_impairment_loss", False),
            ("资产处置收益", "asset_disposal_gain", False),
            ("营业利润", "operating_profit", True),
            ("营业外收入", "non_operating_income", False),
            ("营业外支出", "non_operating_expense", False),
            ("利润总额", "total_profit", False),
            ("所得税费用", "income_tax", False),
            ("净利润", "net_profit", True),
        ]
        for label, key, required in items:
            row_vals = [label]
            month_vals = []
            for m_key in months:
                data = metrics_by_month.get(m_key, {})
                if key == "gross_profit":
                    val = data.get("gross_profit")
                    if val is None:
                        revenue = data.get("revenue")
                        cost = data.get("cost")
                        val = (revenue - cost) if revenue is not None and cost is not None else None
                else:
                    val = data.get(key)
                row_vals.append(val)
                month_vals.append(val)
            if required or any(v is not None for v in month_vals):
                row_vals.append(self._safe_sum(month_vals))
                rows.append(row_vals)
        self._write_table(ws, 1, 1, headers, rows)
        note_row = len(rows) + 3
        self._write_chart_note(ws, 1, note_row, "说明：年度利润表按月汇总，仅在12月生成。")
        self._write_annual_data_gap_note(ws, note_row + 1, self._annual_data_gaps(target_year))

        self._reorder_month_columns_simple_desc(ws, header_row=1)

    def _update_annual_metrics_sheet(self, wb, metrics_by_month, target_year):
        ws = self._prepare_sheet(wb, "年度经营指标", insert_after="经营指标")
        value_cols = [
            ("主营业务收入", "revenue"),
            ("主营业务成本", "cost"),
            ("毛利润", "gross_profit"),
            ("销售费用", "sales_expense"),
            ("管理费用", "admin_expense"),
            ("财务费用", "financial_expense"),
            ("营业利润", "operating_profit"),
            ("净利润", "net_profit"),
            ("应收账款余额", "ar_balance"),
            ("存货期末余额", "inventory_end"),
        ]
        ratio_cols = [
            ("成本率", "cost_rate"),
            ("销售费用率", "sales_expense_rate"),
            ("管理费用率", "admin_expense_rate"),
            ("财务费用率", "financial_expense_rate"),
            ("营业利润率", "operating_profit_rate"),
            ("净利润率", "net_profit_rate"),
            ("存货周转天数", "inventory_turnover_days"),
        ]
        headers = ["月份"] + [c[0] for c in value_cols] + [c[0] for c in ratio_cols]
        rows = []
        months = [f"{target_year}-{m:02d}" for m in range(1, 13)]
        sum_totals = {
            "revenue": [],
            "cost": [],
            "gross_profit": [],
            "sales_expense": [],
            "admin_expense": [],
            "financial_expense": [],
            "operating_profit": [],
            "net_profit": [],
        }
        balance_totals = {
            "ar_balance": [],
            "inventory_end": [],
        }
        ratio_totals = {
            "inventory_turnover_days": [],
        }
        for m_key in months:
            data = metrics_by_month.get(m_key, {})
            revenue = data.get("revenue")
            cost = data.get("cost")
            gross_profit = data.get("gross_profit")
            if gross_profit is None and revenue is not None and cost is not None:
                gross_profit = revenue - cost
            row = [self._month_key_to_label(m_key)]
            row_values = {
                "revenue": revenue,
                "cost": cost,
                "gross_profit": gross_profit,
                "sales_expense": data.get("sales_expense"),
                "admin_expense": data.get("admin_expense"),
                "financial_expense": data.get("financial_expense"),
                "operating_profit": data.get("operating_profit"),
                "net_profit": data.get("net_profit"),
                "ar_balance": data.get("ar_balance"),
                "inventory_end": data.get("inventory_end"),
            }
            for _, key in value_cols:
                row.append(row_values.get(key))
            for _, key in ratio_cols:
                row.append(data.get(key))
            rows.append(row)
            for key in sum_totals:
                sum_totals[key].append(row_values.get(key))
            for key in balance_totals:
                balance_totals[key].append(row_values.get(key))
            ratio_totals["inventory_turnover_days"].append(data.get("inventory_turnover_days"))

        def last_non_none(values):
            for v in reversed(values):
                if v is not None:
                    return v
            return None

        def avg(values):
            vals = [v for v in values if v is not None]
            return sum(vals) / len(vals) if vals else None

        total_revenue = self._safe_sum(sum_totals["revenue"])
        total_cost = self._safe_sum(sum_totals["cost"])
        total_sales_expense = self._safe_sum(sum_totals["sales_expense"])
        total_admin_expense = self._safe_sum(sum_totals["admin_expense"])
        total_financial_expense = self._safe_sum(sum_totals["financial_expense"])
        total_operating_profit = self._safe_sum(sum_totals["operating_profit"])
        total_net_profit = self._safe_sum(sum_totals["net_profit"])
        total_gross_profit = None
        if total_revenue is not None and total_cost is not None:
            total_gross_profit = total_revenue - total_cost
        else:
            total_gross_profit = self._safe_sum(sum_totals["gross_profit"])

        cost_rate_total = (total_cost / total_revenue) if (total_revenue not in (None, 0) and total_cost is not None) else None
        sales_rate_total = (total_sales_expense / total_revenue) if (total_revenue not in (None, 0) and total_sales_expense is not None) else None
        admin_rate_total = (total_admin_expense / total_revenue) if (total_revenue not in (None, 0) and total_admin_expense is not None) else None
        financial_rate_total = (total_financial_expense / total_revenue) if (total_revenue not in (None, 0) and total_financial_expense is not None) else None
        operating_rate_total = (total_operating_profit / total_revenue) if (total_revenue not in (None, 0) and total_operating_profit is not None) else None
        net_rate_total = (total_net_profit / total_revenue) if (total_revenue not in (None, 0) and total_net_profit is not None) else None

        rows.append([
            "全年合计",
            total_revenue,
            total_cost,
            total_gross_profit,
            total_sales_expense,
            total_admin_expense,
            total_financial_expense,
            total_operating_profit,
            total_net_profit,
            last_non_none(balance_totals["ar_balance"]),
            last_non_none(balance_totals["inventory_end"]),
            cost_rate_total,
            sales_rate_total,
            admin_rate_total,
            financial_rate_total,
            operating_rate_total,
            net_rate_total,
            avg(ratio_totals["inventory_turnover_days"]),
        ])
        self._write_table(ws, 1, 1, headers, rows)
        note_row = len(rows) + 3
        self._write_chart_note(ws, 1, note_row, "说明：年度经营指标按月汇总，仅在12月生成。")
        self._write_annual_data_gap_note(ws, note_row + 1, self._annual_data_gaps(target_year))

        self._reorder_month_rows_desc(ws)

    def _update_annual_balance_sheet(self, wb, target_year):
        ws = self._prepare_sheet(wb, "年度资产负债表", insert_after="资产负债表")
        asset_keys = [k for k in self.data.get("asset", {}).keys() if str(k).startswith(f"{target_year}-")]
        if not asset_keys:
            self._write_table(ws, 1, 1, ["提示"], [["无年度资产数据"]])
            return
        start_key = sorted(asset_keys)[0]
        end_key = sorted(asset_keys)[-1]
        start_metrics = self._extract_asset_metrics(self.data["asset"].get(start_key), start_key)
        end_metrics = self._extract_asset_metrics(self.data["asset"].get(end_key), end_key)

        headers = ["项目", f"期初({self._month_key_to_label(start_key)})", f"期末({self._month_key_to_label(end_key)})", "变动"]
        rows = []
        items = [
            ("货币资金", "cash", True),
            ("应收账款余额", "ar_balance", True),
            ("存货期末余额", "inventory_end", True),
            ("固定资产", "fixed_assets", True),
            ("流动资产合计", "current_assets", False),
            ("非流动资产合计", "non_current_assets", False),
            ("资产总计", "total_assets", True),
            ("应付账款", "ap_balance", True),
            ("短期借款", "short_debt", True),
            ("长期借款", "long_debt", True),
            ("应付票据", "notes_payable", False),
            ("应交税费", "taxes_payable", False),
            ("其他应付款", "other_payables", False),
            ("预收账款", "advances_from_customers", False),
            ("流动负债合计", "current_liabilities", False),
            ("非流动负债合计", "non_current_liabilities", False),
            ("负债合计", "total_liabilities", True),
            ("实收资本", "paid_in_capital", False),
            ("本年利润", "current_year_profit", False),
            ("利润分配", "profit_distribution", False),
            ("未分配利润", "retained_earnings", False),
            ("所有者权益合计", "total_equity", True),
        ]
        for label, key, required in items:
            start_val = start_metrics.get(key)
            end_val = end_metrics.get(key)
            if not required and start_val is None and end_val is None:
                continue
            delta = end_val - start_val if start_val is not None and end_val is not None else None
            rows.append([label, start_val, end_val, delta])
        self._write_table(ws, 1, 1, headers, rows)
        note_row = len(rows) + 3
        self._write_chart_note(ws, 1, note_row, "说明：年度资产负债表取年初与年末对比，仅在12月生成。")
        self._write_annual_data_gap_note(ws, note_row + 1, self._annual_data_gaps(target_year))

    def _update_annual_summary_sheet(self, wb, metrics_by_month, target_year, target_month, year_scope=None, anomaly_top_n=10):
        ws = self._prepare_sheet(wb, "年度汇总总览", insert_after="年度费用异常Top")

        title = f"{target_year}年度汇总总览"
        ws.cell(row=1, column=1).value = title
        ws.cell(row=1, column=1).font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
        ws.cell(row=2, column=1).value = f"生成期间：截至 {target_year}/{int(target_month):02d}"

        months = [f"{target_year}-{m:02d}" for m in range(1, 13)]

        def last_non_none(values):
            for v in reversed(values):
                if v is not None:
                    return v
            return None

        revenue_vals = [metrics_by_month.get(m, {}).get("revenue") for m in months]
        cost_vals = [metrics_by_month.get(m, {}).get("cost") for m in months]
        sales_expense_vals = [metrics_by_month.get(m, {}).get("sales_expense") for m in months]
        admin_expense_vals = [metrics_by_month.get(m, {}).get("admin_expense") for m in months]
        financial_expense_vals = [metrics_by_month.get(m, {}).get("financial_expense") for m in months]
        operating_profit_vals = [metrics_by_month.get(m, {}).get("operating_profit") for m in months]
        net_profit_vals = [metrics_by_month.get(m, {}).get("net_profit") for m in months]
        ar_vals = [metrics_by_month.get(m, {}).get("ar_balance") for m in months]
        inventory_vals = [metrics_by_month.get(m, {}).get("inventory_end") for m in months]

        total_revenue = self._safe_sum(revenue_vals)
        total_cost = self._safe_sum(cost_vals)
        total_sales_expense = self._safe_sum(sales_expense_vals)
        total_admin_expense = self._safe_sum(admin_expense_vals)
        total_financial_expense = self._safe_sum(financial_expense_vals)
        total_operating_profit = self._safe_sum(operating_profit_vals)
        total_net_profit = self._safe_sum(net_profit_vals)
        total_gross_profit = (total_revenue - total_cost) if (total_revenue is not None and total_cost is not None) else None

        kpi_rows = [
            ["主营业务收入(全年)", total_revenue, "来自年度经营指标按月汇总"],
            ["主营业务成本(全年)", total_cost, "来自年度经营指标按月汇总"],
            ["毛利润(全年)", total_gross_profit, "收入-成本"],
            ["毛利率(全年)", (total_gross_profit / total_revenue) if total_revenue else None, "毛利润/收入"],
            ["销售费用(全年)", total_sales_expense, "来自年度经营指标按月汇总"],
            ["管理费用(全年)", total_admin_expense, "来自年度经营指标按月汇总"],
            ["财务费用(全年)", total_financial_expense, "来自年度经营指标按月汇总"],
            ["营业利润(全年)", total_operating_profit, "来自年度经营指标按月汇总"],
            ["净利润(全年)", total_net_profit, "来自年度经营指标按月汇总"],
            ["净利率(全年)", (total_net_profit / total_revenue) if total_revenue else None, "净利润/收入"],
            ["年末应收账款余额", last_non_none(ar_vals), "取年度最后有值月份"],
            ["年末存货余额", last_non_none(inventory_vals), "取年度最后有值月份"],
        ]

        start_row = 4
        self._write_table(ws, start_row, 1, ["核心指标", "年度值", "说明"], kpi_rows)
        row_cursor = start_row + len(kpi_rows) + 2

        trend_rows = []
        for m in months:
            data = metrics_by_month.get(m, {})
            revenue = data.get("revenue")
            cost = data.get("cost")
            gross_profit = data.get("gross_profit")
            if gross_profit is None and revenue is not None and cost is not None:
                gross_profit = revenue - cost
            operating_profit = data.get("operating_profit")
            net_profit = data.get("net_profit")
            if any(v is not None for v in [revenue, cost, gross_profit, operating_profit, net_profit]):
                trend_rows.append([
                    self._month_key_to_label(m),
                    revenue,
                    cost,
                    gross_profit,
                    operating_profit,
                    net_profit,
                ])
        if not trend_rows:
            trend_rows = [["无年度数据", None, None, None, None, None]]

        trend_header_row = row_cursor
        self._write_table(ws, trend_header_row, 1, ["月度趋势", "收入", "成本", "毛利润", "营业利润", "净利润"], trend_rows)
        trend_has_data = any(any(v is not None for v in row[1:]) for row in trend_rows)
        if trend_has_data:
            trend_data_start = trend_header_row + 1
            trend_data_end = trend_header_row + len(trend_rows)
            self._add_line_chart_by_columns(
                ws,
                1,
                [2, 5, 6],
                trend_header_row,
                trend_data_start,
                trend_data_end,
                "年度收入与利润趋势",
                "I4",
            )
            self._write_chart_note(ws, 9, 2, "图表说明：年度收入、营业利润、净利润月度趋势。")
        row_cursor += len(trend_rows) + 2

        asset_rows = []
        asset_keys = [k for k in self.data.get("asset", {}).keys() if str(k).startswith(f"{target_year}-")]
        if asset_keys:
            start_key = sorted(asset_keys)[0]
            end_key = sorted(asset_keys)[-1]
            start_metrics = self._extract_asset_metrics(self.data["asset"].get(start_key), start_key)
            end_metrics = self._extract_asset_metrics(self.data["asset"].get(end_key), end_key)
            for label, key in [
                ("货币资金", "cash"),
                ("应收账款余额", "ar_balance"),
                ("存货期末余额", "inventory_end"),
                ("资产总计", "total_assets"),
                ("负债合计", "total_liabilities"),
                ("所有者权益合计", "total_equity"),
            ]:
                start_val = start_metrics.get(key)
                end_val = end_metrics.get(key)
                delta = (end_val - start_val) if (start_val is not None and end_val is not None) else None
                asset_rows.append([label, start_val, end_val, delta])
            asset_headers = ["资产负债关键项", f"期初({self._month_key_to_label(start_key)})", f"期末({self._month_key_to_label(end_key)})", "变动"]
        else:
            asset_headers = ["资产负债关键项", "期初", "期末", "变动"]
            asset_rows = [["无年度资产数据", None, None, None]]

        self._write_table(ws, row_cursor, 1, asset_headers, asset_rows)
        row_cursor += len(asset_rows) + 2

        anomaly_rows = []
        expense_df = self._get_expense_df()
        if expense_df is not None and not expense_df.empty:
            flags = self._collect_expense_mom_flags(expense_df, target_year, target_month, "current")
            flags = sorted(
                flags,
                key=lambda x: (
                    x.get("AnomalyScore") or 0,
                    abs(x.get("Delta") or 0),
                    abs(x.get("Amount") or 0),
                ),
                reverse=True,
            )[:anomaly_top_n]
            for f in flags:
                anomaly_rows.append([
                    self._month_key_to_label(f.get("MonthStr")),
                    f.get("Category"),
                    f.get("Subcategory"),
                    f.get("Amount"),
                    f.get("Delta"),
                    f.get("Rate"),
                    f.get("AnomalyScore"),
                    "、".join(f.get("ReasonTags") or []),
                ])
        if not anomaly_rows:
            anomaly_rows = [["无异常项", None, None, None, None, None, None, None]]

        self._write_table(
            ws,
            row_cursor,
            1,
            ["费用异常Top", "费用类别", "子科目", "本期金额", "环比增量", "环比增速", "异常评分", "异常标签"],
            anomaly_rows,
        )
        row_cursor += len(anomaly_rows) + 2

        self._write_chart_note(ws, 1, row_cursor, "说明：本页整合年度经营指标、利润、资产负债与费用异常，便于一次性汇报。")
        self._write_annual_data_gap_note(ws, row_cursor + 1, self._annual_data_gaps(target_year))

    def _get_expense_df(self):
        if not self.data['expense']:
            return None
        frames = []
        for df in self.data['expense'].values():
            frames.append(df)
        if not frames:
            return None
        all_df = pd.concat(frames, ignore_index=True)
        all_df.columns = [str(c).strip().rstrip('\t') for c in all_df.columns]
        return all_df

    def _classify_expense_behavior(self, row):
        category = "" if pd.isna(row.get('Category')) else str(row.get('Category') or "")
        subcategory = "" if pd.isna(row.get('Subcategory')) else str(row.get('Subcategory') or "")
        fields = [category, subcategory, row.get('Summary'), row.get('科目名'), row.get('摘要')]
        text = "".join("" if pd.isna(v) else str(v) for v in fields)
        text_lower = text.lower()

        if any(k in text for k in ["存货盘亏", "存货盘盈", "盘亏", "盘盈", "盘点", "库存调整"]):
            return "inventory_adjustment"
        if any(k in text for k in ["汇兑损益", "期末调汇", "汇率", "调汇"]):
            return "financial_adjustment"
        bank_terms = ["U盾", "u盾", "网银", "银行费用", "银行手续费", "账户管理", "转账手续费", "电汇手续费"]
        if (
            any(k in text for k in bank_terms)
            or ("手续费" in text and ("财务费用" in category or "银行" in text))
            or ("bank" in text_lower and "fee" in text_lower)
        ):
            return "financial_adjustment"

        entertainment_terms = ["交际应酬", "招待", "宴请", "客户吃饭", "厂家吃饭", "红包", "舞狮"]
        if any(k in text for k in entertainment_terms):
            return "variable" if "销售" in category else "fixed"

        personnel_document_terms = ["签证", "健康证", "照片", "居留", "证照", "认证", "无犯罪记录", "海牙"]
        if any(k in text for k in personnel_document_terms):
            return "fixed"

        professional_terms = ["律师", "法律", "咨询", "顾问", "审计", "年账", "做年账"]
        if any(k in text for k in professional_terms):
            return "fixed"

        maintenance_terms = ["维修", "修理", "电线", "断路器", "插板", "插座", "耗材", "工具", "清洁用品"]
        if any(k in text for k in maintenance_terms):
            return "fixed"

        if any(k in text for k in ["佣金", "运费", "运输", "物流", "车费", "快递", "提成", "市场", "广告", "推广", "手续费", "文件费", "装卸", "报关", "代理", "验货", "包装"]):
            return "variable"
        if any(k in text for k in ["工资", "房租", "物业", "水电", "折旧", "办公", "福利", "保险", "生活", "车辆", "差旅", "利息", "自由港", "通讯"]):
            return "fixed"
        return "unclassified"

    def _expense_behavior_history_key(self, row):
        def clean(value):
            if value is None or pd.isna(value):
                return ""
            return str(value).strip().replace("\u3000", " ")

        category = clean(row.get('Category'))
        subcategory = clean(row.get('Subcategory'))
        if not category or not subcategory:
            return None

        generic_subcategories = {"其他", "其它", "未填子科目", "nan", "None"}
        if subcategory in generic_subcategories:
            summary = clean(row.get('Summary') or row.get('摘要'))
            if not summary:
                return None
            normalized_summary = re.sub(r"\s+", "", summary)
            normalized_summary = re.sub(r"\d+(?:\.\d+)?", "#", normalized_summary)
            return (category, subcategory, normalized_summary[:24])

        return (category, subcategory)

    def _build_expense_behavior_history_profiles(self, df):
        profiles = {}
        if df is None or df.empty:
            return profiles

        for _, row in df.iterrows():
            key = self._expense_behavior_history_key(row)
            if not key:
                continue
            amount = row.get('Amount')
            month_key = row.get('MonthStr')
            if not month_key or amount is None or pd.isna(amount) or amount == 0:
                continue

            bucket = row.get('_RuleBehavior')
            if not bucket:
                bucket = self._classify_expense_behavior(row)

            profile = profiles.setdefault(key, {
                "bucket_abs": {},
                "bucket_count": {},
                "monthly_amount": {},
                "category": key[0],
            })
            profile["monthly_amount"][month_key] = profile["monthly_amount"].get(month_key, 0) + amount
            if bucket and bucket != "unclassified":
                profile["bucket_abs"][bucket] = profile["bucket_abs"].get(bucket, 0) + abs(amount)
                profile["bucket_count"][bucket] = profile["bucket_count"].get(bucket, 0) + 1
        return profiles

    def _infer_expense_behavior_from_history(self, row, profiles):
        key = self._expense_behavior_history_key(row)
        if not key:
            return None
        profile = profiles.get(key)
        if not profile:
            return None

        bucket_abs = profile.get("bucket_abs") or {}
        if bucket_abs:
            ranked = sorted(bucket_abs.items(), key=lambda x: x[1], reverse=True)
            top_bucket, top_abs = ranked[0]
            total_abs = sum(bucket_abs.values())
            top_count = (profile.get("bucket_count") or {}).get(top_bucket, 0)
            if total_abs and top_abs / total_abs >= 0.7 and top_count >= 1:
                return top_bucket

        monthly_values = [
            abs(v)
            for v in (profile.get("monthly_amount") or {}).values()
            if v is not None and abs(v) > 0
        ]
        if len(monthly_values) < 3:
            return None

        mean_value = sum(monthly_values) / len(monthly_values)
        if not mean_value:
            return None
        variance = sum((v - mean_value) ** 2 for v in monthly_values) / len(monthly_values)
        coefficient_of_variation = (variance ** 0.5) / mean_value
        category = str(profile.get("category") or "")

        if coefficient_of_variation <= 0.75:
            return "fixed"
        if "财务费用" in category:
            return "financial_adjustment"
        if "销售费用" in category and coefficient_of_variation >= 1.0:
            return "variable"
        return None

    def _calculate_expense_behavior_totals(self, target_year=None, target_month=None, year_scope=None):
        raw_df = self._get_expense_df()
        df = self._prepare_expense_analysis_df(raw_df, target_year, target_month, year_scope)
        if df is None or df.empty:
            return {
                "variable": {},
                "fixed": {},
                "financial_adjustment": {},
                "inventory_adjustment": {},
                "unclassified": {},
                "total": {},
            }

        df = df.copy()
        df["_RuleBehavior"] = df.apply(self._classify_expense_behavior, axis=1)
        history_profiles = self._build_expense_behavior_history_profiles(df)

        totals = {
            "variable": {},
            "fixed": {},
            "financial_adjustment": {},
            "inventory_adjustment": {},
            "unclassified": {},
            "total": {},
        }
        for _, row in df.iterrows():
            month_key = row.get('MonthStr')
            amount = row.get('Amount')
            if not month_key or amount is None or pd.isna(amount) or amount == 0:
                continue
            bucket = row.get("_RuleBehavior") or self._classify_expense_behavior(row)
            if bucket == "unclassified":
                bucket = self._infer_expense_behavior_from_history(row, history_profiles) or "unclassified"
            if bucket not in totals:
                bucket = "unclassified"
            totals[bucket][month_key] = totals[bucket].get(month_key, 0) + amount
            totals["total"][month_key] = totals["total"].get(month_key, 0) + amount
        return totals

    def _calculate_expense_keyword_totals(self, target_year=None, target_month=None, year_scope=None):
        totals = self._calculate_expense_behavior_totals(target_year, target_month, year_scope)
        return totals.get("variable", {}), totals.get("fixed", {})

    def _update_expense_mom_sheet(self, ws, target_year, target_month, year_scope=None):
        raw_df = self._get_expense_df()
        df = self._prepare_expense_analysis_df(raw_df, target_year, target_month, year_scope)
        if df is None or df.empty:
            return

        df = df[df['Category'].notna() & df['Subcategory'].notna()]
        if df.empty:
            return

        summary = (
            df.groupby(['Category', 'Subcategory', 'MonthStr'])['Amount']
            .sum()
            .reset_index()
        )

        value_cols, rate_cols, delta_cols = self._get_expense_mom_column_maps(ws, header_row=1)

        score_col = self._ensure_header_column(ws, "异常评分")
        tag_col = self._ensure_header_column(ws, "异常标签")
        reason_col = self._ensure_header_column(ws, "异常原因")
        month_col = self._ensure_header_column(ws, "异常月份")
        freq_col = self._ensure_header_column(ws, "近12月出现")
        detail_count_col = self._ensure_header_column(ws, "明细笔数")
        detail_key_col = self._ensure_header_column(ws, "明细键")
        self._apply_header_style(ws, 1, max_col=ws.max_column)

        flags = self._collect_expense_mom_flags(df, target_year, target_month, year_scope)
        selected_by_pair = self._select_expense_display_flags(flags, target_year, target_month)

        detail_count_map = (
            df.groupby(['MonthStr', 'Category', 'Subcategory'])['Amount']
            .size()
            .to_dict()
        )

        rows = []
        for (cat, sub), group in summary.groupby(['Category', 'Subcategory']):
            row = {'部门': '合计', '费用类别': cat, '子科目': sub}
            for _, rec in group.iterrows():
                row[rec['MonthStr']] = rec['Amount']
            rows.append(row)

        rows.sort(key=lambda x: (x['费用类别'], x['子科目']))

        # Clear old rows
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                self._safe_set_cell_value(ws, r, c, None)

        row_idx = 2
        for row in rows:
            ws.cell(row=row_idx, column=1).value = row['部门']
            ws.cell(row=row_idx, column=2).value = row['费用类别']
            ws.cell(row=row_idx, column=3).value = row['子科目']

            # Fill month values
            for m_key, col in value_cols.items():
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    self._safe_set_cell_value(ws, row_idx, col, None)
                    continue
                self._safe_set_cell_value(ws, row_idx, col, row.get(m_key))

            # Fill MoM rate and delta
            for m_key, col in rate_cols.items():
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    self._safe_set_cell_value(ws, row_idx, col, None)
                    continue
                prev_year, prev_month = map(int, m_key.split('-'))
                prev_key = None
                if prev_month > 1:
                    prev_key = f"{prev_year}-{prev_month-1:02d}"
                curr = row.get(m_key)
                prev = row.get(prev_key) if prev_key else None
                rate = None
                if curr is not None and prev not in (None, 0):
                    rate = (curr - prev) / prev
                self._safe_set_cell_value(ws, row_idx, col, rate)

            for m_key, col in delta_cols.items():
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    self._safe_set_cell_value(ws, row_idx, col, None)
                    continue
                prev_year, prev_month = map(int, m_key.split('-'))
                prev_key = None
                if prev_month > 1:
                    prev_key = f"{prev_year}-{prev_month-1:02d}"
                curr = row.get(m_key)
                prev = row.get(prev_key) if prev_key else None
                delta = None
                if curr is not None and prev is not None:
                    delta = curr - prev
                self._safe_set_cell_value(ws, row_idx, col, delta)

            pair = (row['费用类别'], row['子科目'])
            selected_flag = selected_by_pair.get(pair)

            if selected_flag:
                tags = selected_flag.get("ReasonTags") or []
                self._safe_set_cell_value(ws, row_idx, score_col, selected_flag.get("AnomalyScore"))
                self._safe_set_cell_value(ws, row_idx, tag_col, "、".join(tags))
                self._safe_set_cell_value(ws, row_idx, reason_col, selected_flag.get("ReasonText"))
                self._safe_set_cell_value(ws, row_idx, month_col, self._month_key_to_label(selected_flag.get("MonthStr")))
                self._safe_set_cell_value(ws, row_idx, freq_col, selected_flag.get("ActiveMonths12"))
                self._safe_set_cell_value(
                    ws,
                    row_idx,
                    detail_count_col,
                    detail_count_map.get((selected_flag.get("MonthStr"), selected_flag.get("Category"), selected_flag.get("Subcategory"))),
                )
                key_text = selected_flag.get("AnomalyKey")
                self._safe_set_cell_value(ws, row_idx, detail_key_col, key_text)
                key_cell = ws.cell(row=row_idx, column=detail_key_col)
                if key_text:
                    target_sheet = self.expense_detail_sheet_name
                    target_row = (self.expense_detail_key_row_map or {}).get(key_text)
                    if target_sheet in ws.parent.sheetnames:
                        anchor_row = target_row if target_row else 2
                        self._apply_hyperlink(key_cell, f"#'{target_sheet}'!A{anchor_row}")
                        key_cell.font = Font(color="0563C1", underline="single")
                    else:
                        key_cell.hyperlink = None
                else:
                    key_cell.hyperlink = None
            else:
                self._safe_set_cell_value(ws, row_idx, score_col, None)
                self._safe_set_cell_value(ws, row_idx, tag_col, None)
                self._safe_set_cell_value(ws, row_idx, reason_col, None)
                self._safe_set_cell_value(ws, row_idx, month_col, None)
                self._safe_set_cell_value(ws, row_idx, freq_col, None)
                self._safe_set_cell_value(ws, row_idx, detail_count_col, None)
                self._safe_set_cell_value(ws, row_idx, detail_key_col, None)
                ws.cell(row=row_idx, column=detail_key_col).hyperlink = None

            row_idx += 1

        self._reorder_month_columns_grouped_by_suffix_desc(ws, header_row=1)
        self._refresh_expense_mom_conditional_formatting(
            ws,
            data_start_row=2,
            data_end_row=max(2, row_idx - 1),
            rate_threshold=0.5,
            delta_threshold=10000,
        )

    def _apply_cvp_sheet_format(self, ws):
        widths = {
            1: 10,
            2: 8,
            3: 14,
            4: 14,
            5: 14,
            6: 12,
            7: 14,
            8: 14,
            9: 16,
            10: 14,
            11: 12,
        }
        for col_idx, width in widths.items():
            letter = get_column_letter(col_idx)
            current = ws.column_dimensions[letter].width or 0
            ws.column_dimensions[letter].width = max(current, width)

        amount_cols = (3, 4, 5, 7, 8, 9, 10)
        percent_cols = (6, 11)
        for r in range(2, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label is None or str(label).strip() == "":
                continue
            for c in amount_cols:
                ws.cell(row=r, column=c).number_format = '#,##0.00'
            for c in percent_cols:
                ws.cell(row=r, column=c).number_format = '0.0%'

    def _update_cvp_sheet(self, ws, metrics_by_month, target_year, target_month, year_scope=None):
        var_by_month, fixed_by_month = self._calculate_expense_keyword_totals(target_year, target_month, year_scope)
        min_margin_ratio = 0.01

        def _valid_margin_ratio(value):
            return value is not None and not pd.isna(value) and value >= min_margin_ratio

        # Write Headers
        headers = ["月份", "合计", "销售收入", "变动成本", "贡献毛利", "贡献毛利率", "固定成本", "总成本", "盈亏平衡点", "安全边际", "安全边际率"]
        for idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=idx).value = h
        self._apply_header_style(ws, 1)

        total_revenue = 0
        total_variable = 0
        total_fixed = 0

        total_row = None
        for r in range(2, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if not label:
                continue
            if str(label).strip() == '合计':
                total_row = r
                continue
            month_key = self._label_to_month_key(label)
            if not self._month_key_in_scope(month_key, target_year, target_month, year_scope):
                for c in range(2, 12):
                    self._safe_set_cell_value(ws, r, c, None)
                continue

            data = metrics_by_month.get(month_key)
            if not data:
                continue
            revenue = data.get('revenue')
            base_cost = data.get('cost')
            variable_cost = None
            if revenue is not None:
                base = base_cost or 0
                if isinstance(base, float) and pd.isna(base):
                    base = 0
                var_extra = var_by_month.get(month_key, 0)
                if isinstance(var_extra, float) and pd.isna(var_extra):
                    var_extra = 0
                variable_cost = base + (var_extra or 0)

            contribution = revenue - variable_cost if revenue is not None and variable_cost is not None else None
            ratio = contribution / revenue if (contribution is not None and revenue not in (None, 0)) else None
            fixed_cost = fixed_by_month.get(month_key, 0) if revenue is not None else None
            if fixed_cost is None or (isinstance(fixed_cost, float) and pd.isna(fixed_cost)):
                fixed_cost = 0
            
            total_cost = (variable_cost or 0) + (fixed_cost or 0) if revenue is not None else None

            if _valid_margin_ratio(ratio):
                break_even = fixed_cost / ratio
                margin = revenue - break_even
                margin_ratio = margin / revenue if (margin is not None and revenue not in (None, 0)) else None
            else:
                break_even = None
                margin = None
                margin_ratio = None

            ws.cell(row=r, column=2).value = '合计'
            ws.cell(row=r, column=3).value = revenue
            ws.cell(row=r, column=4).value = variable_cost
            ws.cell(row=r, column=5).value = contribution
            ws.cell(row=r, column=6).value = ratio
            ws.cell(row=r, column=7).value = fixed_cost
            ws.cell(row=r, column=8).value = total_cost
            ws.cell(row=r, column=9).value = break_even
            ws.cell(row=r, column=10).value = margin
            ws.cell(row=r, column=11).value = margin_ratio

            if revenue:
                total_revenue += revenue
            if variable_cost:
                total_variable += variable_cost
            if fixed_cost:
                total_fixed += fixed_cost

        if total_row:
            total_contribution = total_revenue - total_variable
            total_ratio = total_contribution / total_revenue if total_revenue else None
            total_cost_sum = total_variable + total_fixed
            if _valid_margin_ratio(total_ratio):
                total_break_even = total_fixed / total_ratio
                total_margin = total_revenue - total_break_even
                total_margin_ratio = total_margin / total_revenue if total_revenue else None
            else:
                total_break_even = None
                total_margin = None
                total_margin_ratio = None

            ws.cell(row=total_row, column=3).value = total_revenue
            ws.cell(row=total_row, column=4).value = total_variable
            ws.cell(row=total_row, column=5).value = total_contribution
            ws.cell(row=total_row, column=6).value = total_ratio
            ws.cell(row=total_row, column=7).value = total_fixed
            ws.cell(row=total_row, column=8).value = total_cost_sum
            ws.cell(row=total_row, column=9).value = total_break_even
            ws.cell(row=total_row, column=10).value = total_margin
            ws.cell(row=total_row, column=11).value = total_margin_ratio

        self._reorder_month_rows_desc(ws)
        self._apply_cvp_sheet_format(ws)

    def _update_sales_inventory_detail_sheet(self, ws, target_year, target_month):
        if not target_year or not target_month:
            return
        month_key = f"{target_year}-{int(target_month):02d}"
        cost_df = self.data['cost'].get(month_key)
        if cost_df is None or cost_df.empty:
            return

        def find_col_contains(keyword):
            for c in cost_df.columns:
                if keyword in str(c):
                    return c
            return None

        def find_col_suffix(suffix):
            for c in cost_df.columns:
                if str(c).endswith(suffix):
                    return c
            return None

        code_col = find_col_contains('品目编码')
        name_col = find_col_contains('品目名')
        type_col = find_col_contains('品目类型')
        process_col = find_col_contains('生产流程名')
        q_start = find_col_suffix('期初')
        p_start = find_col_suffix('期初.1')
        a_start = find_col_suffix('期初.2')
        q_add = find_col_suffix('增加')
        p_add = find_col_suffix('增加.1')
        a_add = find_col_suffix('增加.2')
        q_reduce = find_col_suffix('减少')
        p_reduce = find_col_suffix('减少.1')
        a_reduce = find_col_suffix('减少.2')
        q_end = find_col_suffix('期末')
        p_end = find_col_suffix('期末.1')
        a_end = find_col_suffix('期末.2')

        if not code_col:
            return

        sales_df = self.data['sales'].get(month_key)
        sales_map = {}
        if sales_df is not None and not sales_df.empty:
            sales_df = sales_df.copy()
            sales_df['Revenue'] = self._extract_sales_revenue(sales_df)
            sales_df['Qty'] = pd.to_numeric(sales_df.get('数量'), errors='coerce')
            sales_df = self._attach_sales_cost(sales_df, target_year, target_month, year_scope=None)
            for code, group in sales_df.groupby('品目编码'):
                if code is None or pd.isna(code):
                    continue
                metrics = self._calc_sales_metrics_from_group(group)
                sales_map[str(code).strip()] = {
                    'revenue': metrics.get('revenue'),
                    'qty': metrics.get('qty'),
                    'cost': metrics.get('cost'),
                    'name': group.get('品目名').dropna().iloc[0] if '品目名' in group and not group.get('品目名').dropna().empty else None,
                    'category': group.get('品目组合1名').dropna().iloc[0] if '品目组合1名' in group and not group.get('品目组合1名').dropna().empty else None,
                }

        # Clear old data
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                self._safe_set_cell_value(ws, r, c, None)

        row_idx = 2
        month_label = f"{target_year}/{int(target_month):02d}"
        for _, row in cost_df.iterrows():
            code = row.get(code_col)
            if not code or pd.isna(code):
                continue
            code = str(code).strip()
            name_spec = row.get(name_col)
            self._safe_set_cell_value(ws, row_idx, 1, code)
            self._safe_set_cell_value(ws, row_idx, 2, name_spec)
            self._safe_set_cell_value(ws, row_idx, 3, row.get(type_col))
            self._safe_set_cell_value(ws, row_idx, 4, row.get(process_col))
            self._safe_set_cell_value(ws, row_idx, 5, self._to_float(row.get(q_start)))
            self._safe_set_cell_value(ws, row_idx, 6, self._to_float(row.get(p_start)))
            self._safe_set_cell_value(ws, row_idx, 7, self._to_float(row.get(a_start)))
            self._safe_set_cell_value(ws, row_idx, 8, self._to_float(row.get(q_add)))
            self._safe_set_cell_value(ws, row_idx, 9, self._to_float(row.get(p_add)))
            self._safe_set_cell_value(ws, row_idx, 10, self._to_float(row.get(a_add)))
            self._safe_set_cell_value(ws, row_idx, 11, self._to_float(row.get(q_reduce)))
            self._safe_set_cell_value(ws, row_idx, 12, self._to_float(row.get(p_reduce)))
            self._safe_set_cell_value(ws, row_idx, 13, self._to_float(row.get(a_reduce)))
            self._safe_set_cell_value(ws, row_idx, 14, self._to_float(row.get(q_end)))
            self._safe_set_cell_value(ws, row_idx, 15, self._to_float(row.get(p_end)))
            self._safe_set_cell_value(ws, row_idx, 16, self._to_float(row.get(a_end)))
            self._safe_set_cell_value(ws, row_idx, 17, month_label)

            sales = sales_map.get(code, {})
            revenue = sales.get('revenue')
            qty = sales.get('qty')
            cost = sales.get('cost')
            profit = revenue - cost if revenue is not None and cost is not None else None
            margin = profit / revenue if (profit is not None and revenue not in (None, 0)) else None
            resolved_category = self._resolve_uncategorized_product(sales.get('category'), name_spec)
            self._safe_set_cell_value(ws, row_idx, 18, revenue)
            self._safe_set_cell_value(ws, row_idx, 19, qty)
            self._safe_set_cell_value(ws, row_idx, 20, sales.get('name') or name_spec)
            self._safe_set_cell_value(ws, row_idx, 21, cost)
            self._safe_set_cell_value(ws, row_idx, 22, 0)
            self._safe_set_cell_value(ws, row_idx, 23, 0)
            self._safe_set_cell_value(ws, row_idx, 24, profit)
            self._safe_set_cell_value(ws, row_idx, 25, margin)
            self._safe_set_cell_value(ws, row_idx, 26, resolved_category)

            row_idx += 1

    def _update_balance_sheet(self, ws, target_year, target_month, year_scope=None):
        month_cols = {}
        for col in range(2, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if isinstance(header, str) and '/' in header:
                month_key = self._label_to_month_key(header)
                if month_key:
                    month_cols[month_key] = col

        for month_key, col in month_cols.items():
            if not self._month_key_in_scope(month_key, target_year, target_month, year_scope):
                for r in range(2, ws.max_row + 1):
                    self._safe_set_cell_value(ws, r, col, None)
                continue
            df = self.data['asset'].get(month_key)
            if df is None or df.empty:
                continue
            balance_map = self._extract_balance_value_map(df, target_year, target_month)
            if not balance_map:
                continue

            alias_map = {
                '存货期末余额': '存货',
                '存货余额': '存货',
                '存货期初余额': '存货',
                '应收账款余额': '应收账款',
                '流动资产合计': '流动资产',
                '非流动资产合计': '非流动资产',
                '流动负债合计': '流动负债',
                '非流动负债合计': '非流动负债',
            }

            for r in range(2, ws.max_row + 1):
                label = ws.cell(row=r, column=1).value
                if not label:
                    continue
                lookup = self._normalize_balance_label(label)
                if not lookup:
                    continue

                candidates = [lookup, f"{lookup}："]
                alias = alias_map.get(lookup)
                if alias:
                    candidates.extend([alias, f"{alias}："])
                if lookup.endswith('合计'):
                    base = lookup[:-2]
                    candidates.extend([base, f"{base}："])

                val = None
                for key in candidates:
                    if key in balance_map:
                        val = balance_map[key]
                        break
                if val is not None:
                    ws.cell(row=r, column=col).value = val

    def _update_budget_sheet(self, ws, metrics_by_month, target_year, target_month, year_scope=None):
        header_row = None
        for r in range(1, min(ws.max_row, 50) + 1):
            if str(ws.cell(row=r, column=1).value).strip() == '月份':
                header_row = r
                break
        if header_row is None:
            return

        header_map = self._get_header_map(ws, header_row)
        if "财务费用率目标" not in header_map:
            src_col = header_map.get("管理费用率目标") or header_map.get("存货周转天数目标") or ws.max_column
            financial_col = self._ensure_header_column(ws, "财务费用率目标", header_row=header_row)
            if src_col and src_col != financial_col:
                self._copy_column_style(ws, src_col, financial_col, max_row=ws.max_row)
            ws.cell(row=header_row, column=financial_col).value = "财务费用率目标"
            self._apply_header_style(ws, header_row, start_col=financial_col, max_col=financial_col)
            header_map = self._get_header_map(ws, header_row)

        target_columns = {
            "revenue": header_map.get("主营业务收入目标"),
            "profit": header_map.get("营业利润目标"),
            "profit_rate": header_map.get("营业利润率目标"),
            "cost_rate": header_map.get("成本率目标"),
            "sales_rate": header_map.get("销售费用率目标"),
            "admin_rate": header_map.get("管理费用率目标"),
            "financial_rate": header_map.get("财务费用率目标"),
            "ar_balance": header_map.get("应收账款余额目标"),
            "inventory_end": header_map.get("存货期末余额目标"),
            "inventory_days": header_map.get("存货周转天数目标"),
        }

        for r in range(header_row + 1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if not label:
                continue
            month_key = self._label_to_month_key(label)
            if not self._month_key_in_scope(month_key, target_year, target_month, year_scope):
                for c in target_columns.values():
                    if c:
                        self._safe_set_cell_value(ws, r, c, None)
                continue
            data = metrics_by_month.get(month_key)
            if not data:
                continue
            revenue = data.get('revenue')
            profit = data.get('operating_profit')
            cost = data.get('cost')
            sales_exp = data.get('sales_expense')
            admin_exp = data.get('admin_expense')
            financial_exp = data.get('financial_expense')
            ar_balance = data.get('ar_balance')
            inv_end = data.get('inventory_end')
            inv_start = data.get('inventory_start')

            profit_rate = profit / revenue if (profit is not None and revenue not in (None, 0)) else None
            cost_rate = cost / revenue if (cost is not None and revenue not in (None, 0)) else None
            sales_rate = sales_exp / revenue if (sales_exp is not None and revenue not in (None, 0)) else None
            admin_rate = admin_exp / revenue if (admin_exp is not None and revenue not in (None, 0)) else None
            financial_rate = financial_exp / revenue if (financial_exp is not None and revenue not in (None, 0)) else None

            avg_inv = None
            if inv_start is not None and inv_end is not None:
                avg_inv = (inv_start + inv_end) / 2
            elif inv_end is not None:
                avg_inv = inv_end
            inv_days = (avg_inv / cost * 365) if avg_inv is not None and cost else None

            values = {
                "revenue": revenue,
                "profit": profit,
                "profit_rate": profit_rate,
                "cost_rate": cost_rate,
                "sales_rate": sales_rate,
                "admin_rate": admin_rate,
                "financial_rate": financial_rate,
                "ar_balance": ar_balance,
                "inventory_end": inv_end,
                "inventory_days": inv_days,
            }
            for key, value in values.items():
                col = target_columns.get(key)
                if col:
                    ws.cell(row=r, column=col).value = value

        self._reorder_month_rows_desc(ws)

    def _get_sales_df(self):
        if self.sales_df is not None:
            return self.sales_df
        if self.data['sales']:
            frames = []
            for df in self.data['sales'].values():
                frames.append(df)
            if frames:
                self.sales_df = pd.concat(frames, ignore_index=True)
                return self.sales_df
        return None

    def _find_cost_col_contains(self, df, keyword):
        for c in df.columns:
            if keyword in str(c):
                return c
        return None

    def _find_cost_col_suffix(self, df, suffix):
        for c in df.columns:
            if str(c).endswith(suffix):
                return c
        return None

    def _is_summary_or_footer_label(self, value):
        if value is None:
            return False
        text = re.sub(r'\s+', '', str(value).strip())
        if not text:
            return False
        upper = text.upper()
        if upper in {
            "合计",
            "总计",
            "總計",
            "小计",
            "小計",
            "累计",
            "累計",
            "商品合计",
            "商品总计",
            "商品總計",
            "TOTAL",
            "SUBTOTAL",
        }:
            return True
        return bool(re.match(r'^20\d{2}[/.-]\d{1,2}[/.-]\d{1,2}', text))

    def _get_unit_cost_map_for_month(self, month_key):
        cost_df = self.data['cost'].get(month_key)
        if cost_df is None or cost_df.empty:
            return {}

        code_col = self._find_cost_col_contains(cost_df, '品目编码')
        unit_col = self._find_cost_col_suffix(cost_df, '减少.1')
        if not unit_col and not cost_df.empty:
            for c in cost_df.columns:
                first_val = cost_df[c].iloc[0] if len(cost_df) > 0 else None
                if str(first_val).strip() == '单价':
                    unit_col = c
                    break

        if not code_col or not unit_col:
            return {}

        df = cost_df[[code_col, unit_col]].copy()
        df = df[df[code_col].notna()]
        df = df[df[code_col].astype(str).str.strip() != '品目编码']
        df = df[~df[code_col].apply(self._is_summary_or_footer_label)]
        df = df[~df[unit_col].astype(str).str.strip().isin(['单价', '数量', '金额'])]
        df['UnitCost'] = pd.to_numeric(df[unit_col], errors='coerce')
        df = df[df['UnitCost'].notna()]
        df['Code'] = df[code_col].astype(str).str.strip()
        return df.groupby('Code')['UnitCost'].last().to_dict()

    def _build_cost_inventory_map(self, df_cost, month_key=None):
        """将成本表整理为按品目编码索引的库存/成本快照。"""
        month_map = {}
        if df_cost is None or df_cost.empty:
            return month_map

        code_col = next((c for c in df_cost.columns if '品目编码' in str(c)), None)
        name_col = next((c for c in df_cost.columns if '品目名' in str(c)), None)
        qty_start_col = next((c for c in df_cost.columns if str(c).endswith('期初')), None)
        amt_start_col = next((c for c in df_cost.columns if str(c).endswith('期初.2')), None)
        qty_end_col = next((c for c in df_cost.columns if str(c).endswith('期末')), None)
        amt_end_col = next((c for c in df_cost.columns if str(c).endswith('期末.2')), None)
        if not code_col:
            return month_map

        unit_cost_map = self._get_unit_cost_map_for_month(month_key) if month_key else {}
        for _, row in df_cost.iterrows():
            raw_code = row.get(code_col)
            if raw_code is None or pd.isna(raw_code):
                continue
            code = str(raw_code).strip()
            if not code or code == "品目编码" or self._is_summary_or_footer_label(code):
                continue

            qty_end = self._to_float(row.get(qty_end_col))
            amt_end = self._to_float(row.get(amt_end_col))
            latest_cost = unit_cost_map.get(code)
            if latest_cost is None and qty_end not in (None, 0) and amt_end is not None:
                latest_cost = amt_end / qty_end

            month_map[code] = {
                'name_spec': row.get(name_col),
                'qty_start': self._to_float(row.get(qty_start_col)),
                'amt_start': self._to_float(row.get(amt_start_col)),
                'qty_end': qty_end,
                'amt_end': amt_end,
                'unit_cost': latest_cost,
            }
        return month_map

    def get_latest_product_inventory_snapshot(self):
        """返回最新成本月份的品目库存/成本快照，供基础数据同步使用。"""
        valid_months = sorted(
            month_key for month_key in self.data['cost'].keys()
            if re.match(r'^20\d{2}-\d{2}$', str(month_key or '').strip())
        )
        if not valid_months:
            return {"month_key": None, "records": []}

        latest_month = valid_months[-1]
        month_map = self._build_cost_inventory_map(self.data['cost'].get(latest_month), latest_month)
        sales_meta = {}
        sales_df = self._get_sales_df()
        if sales_df is not None and not sales_df.empty and '品目编码' in sales_df.columns:
            sales_df = sales_df.copy()
            if 'MonthStr' not in sales_df.columns:
                if 'ParsedDate' in sales_df.columns:
                    sales_df['MonthStr'] = pd.to_datetime(sales_df['ParsedDate'], errors='coerce').dt.strftime('%Y-%m')
                else:
                    date_col = next((c for c in sales_df.columns if '日期' in str(c) or 'Date' in str(c)), None)
                    if date_col:
                        sales_df['ParsedDate'] = pd.to_datetime(sales_df[date_col], errors='coerce')
                        sales_df['MonthStr'] = sales_df['ParsedDate'].dt.strftime('%Y-%m')
            if 'MonthStr' in sales_df.columns:
                sales_df = sales_df[sales_df['MonthStr'].notna()].copy()
                sales_df = sales_df[sales_df['MonthStr'].astype(str) <= latest_month]
                if not sales_df.empty:
                    sales_df['Code'] = sales_df['品目编码'].astype(str).str.strip()
                    sales_df = sales_df[sales_df['Code'] != ""]
                    name_col = next((c for c in ['品目名', '产品名称', '品名', '商品名称'] if c in sales_df.columns), None)
                    if name_col:
                        sales_df['_resolved_name'] = sales_df[name_col]
                    else:
                        sales_df['_resolved_name'] = None
                    sales_df = sales_df.sort_values(['MonthStr']).drop_duplicates('Code', keep='last')
                    for _, row in sales_df.iterrows():
                        code = row.get('Code')
                        if not code:
                            continue
                        sales_meta[code] = {
                            'product_name': row.get('_resolved_name'),
                            'product_type': self._resolve_uncategorized_product(
                                row.get('品目组合1名'),
                                row.get('_resolved_name'),
                            ),
                        }

        records = []
        for code, snapshot in month_map.items():
            qty = snapshot.get("qty_end")
            cost = snapshot.get("unit_cost")
            if qty is None and cost is None:
                continue
            meta = sales_meta.get(code, {})
            name_spec = snapshot.get("name_spec")
            product_name = meta.get("product_name") or name_spec
            records.append({
                "code": code,
                "name": product_name,
                "product_name": product_name,
                "product_type": self._resolve_uncategorized_product(
                    meta.get("product_type"),
                    name_spec or product_name,
                ),
                "spec_info": name_spec,
                "specification": name_spec,
                "latest_inventory_qty": qty,
                "latest_inventory_cost": cost,
                "latest_inventory_date": latest_month,
            })
        return {"month_key": latest_month, "records": records}

    def _build_business_partner_frame(self, df, fallback_month=None):
        if df is None or df.empty:
            return pd.DataFrame(columns=["MonthStr", "PartnerCode", "PartnerName", "PartnerCategory", "AccountSubject"])

        frame = df.copy()
        if "MonthStr" not in frame.columns:
            if "ParsedDate" in frame.columns:
                frame["MonthStr"] = pd.to_datetime(frame["ParsedDate"], errors="coerce").dt.strftime("%Y-%m")
            else:
                date_col = next((c for c in frame.columns if '日期' in str(c) or 'Date' in str(c)), None)
                if date_col:
                    frame["ParsedDate"] = pd.to_datetime(frame[date_col], errors="coerce")
                    frame["MonthStr"] = frame["ParsedDate"].dt.strftime("%Y-%m")
        if "MonthStr" not in frame.columns and fallback_month:
            frame["MonthStr"] = fallback_month
        if "MonthStr" not in frame.columns:
            return pd.DataFrame(columns=["MonthStr", "PartnerCode", "PartnerName", "PartnerCategory", "AccountSubject"])

        code_col = self._pick_first_column(
            frame,
            ["往来单位编码", "客户编码", "供应商编码", "当地编码", "客户代码", "供应商代码"],
        )
        if not code_col:
            return pd.DataFrame(columns=["MonthStr", "PartnerCode", "PartnerName", "PartnerCategory", "AccountSubject"])

        name_col = self._pick_first_column(
            frame,
            ["往来单位名", "客户名称", "客户名", "客户", "收货公司", "供应商名称", "供应商"],
            ["编码"],
        )
        subject_code_col = self._pick_first_column(
            frame,
            ["科目编码", "会计科目编码"],
            ["往来", "品目"],
        )
        subject_name_col = self._pick_first_column(
            frame,
            ["科目名", "会计科目", "科目"],
            ["编码", "往来", "品目"],
        )
        category_col = self._pick_first_column(
            frame,
            ["往来单位分级组合名", "客户分类", "供应商分类", "分类"],
        )

        result = pd.DataFrame()
        result["MonthStr"] = frame["MonthStr"].astype(str).str.strip()
        result["PartnerCode"] = frame[code_col].astype(str).str.strip()
        result["PartnerName"] = (
            frame[name_col].astype(str).str.strip() if name_col else None
        )
        result["PartnerCategory"] = (
            frame[category_col].astype(str).str.strip() if category_col else None
        )
        if subject_code_col or subject_name_col:
            subject_code_series = (
                frame[subject_code_col].astype(str).str.strip() if subject_code_col else pd.Series([None] * len(frame), index=frame.index)
            )
            subject_name_series = (
                frame[subject_name_col].astype(str).str.strip() if subject_name_col else pd.Series([None] * len(frame), index=frame.index)
            )
            result["AccountSubject"] = [
                f"[{code}] {name}".strip()
                if code and code.lower() != "nan" and name and name.lower() != "nan"
                else (code if code and code.lower() != "nan" else (name if name and name.lower() != "nan" else None))
                for code, name in zip(subject_code_series, subject_name_series)
            ]
        else:
            result["AccountSubject"] = None
        result = result[
            (result["PartnerCode"] != "")
            & (result["PartnerCode"].str.lower() != "nan")
            & result["MonthStr"].str.match(r"^20\d{2}-\d{2}$", na=False)
        ].copy()
        if "PartnerName" in result.columns:
            result.loc[result["PartnerName"].str.lower() == "nan", "PartnerName"] = None
        if "PartnerCategory" in result.columns:
            result.loc[result["PartnerCategory"].str.lower() == "nan", "PartnerCategory"] = None
        if "AccountSubject" in result.columns:
            result.loc[result["AccountSubject"].astype(str).str.lower() == "nan", "AccountSubject"] = None
        return result

    def get_latest_business_partner_snapshot(self):
        """返回经营报告数据中可识别的往来单位快照。"""
        frames = []
        sales_df = self._get_sales_df()
        if sales_df is not None and not sales_df.empty:
            frames.append(self._build_business_partner_frame(sales_df))
        if self.ar_detail_df is not None and not self.ar_detail_df.empty:
            frames.append(self._build_business_partner_frame(self.ar_detail_df))

        frames = [frame for frame in frames if frame is not None and not frame.empty]
        if not frames:
            return {"month_key": None, "records": []}

        merged = pd.concat(frames, ignore_index=True)
        merged = merged[merged["MonthStr"].str.match(r"^20\d{2}-\d{2}$", na=False)].copy()
        if merged.empty:
            return {"month_key": None, "records": []}

        latest_month = sorted(merged["MonthStr"].dropna().astype(str).unique())[-1]
        merged = merged.sort_values(["MonthStr"]).reset_index(drop=True)

        latest_by_code = {}
        for _, row in merged.iterrows():
            code = str(row.get("PartnerCode") or "").strip()
            if not code:
                continue
            current = latest_by_code.setdefault(
                code,
                {
                    "code": code,
                    "name": None,
                    "local_code": code,
                    "category": None,
                    "account_subject": None,
                    "latest_seen_date": None,
                },
            )
            month_key = str(row.get("MonthStr") or "").strip()
            if month_key:
                current["latest_seen_date"] = month_key
            partner_name = row.get("PartnerName")
            if partner_name is not None and str(partner_name).strip():
                current["name"] = str(partner_name).strip()
            partner_category = row.get("PartnerCategory")
            if partner_category is not None and str(partner_category).strip():
                current["category"] = str(partner_category).strip()
            account_subject = row.get("AccountSubject")
            if account_subject is not None and str(account_subject).strip():
                current["account_subject"] = str(account_subject).strip()

        return {
            "month_key": latest_month,
            "records": list(latest_by_code.values()),
        }

    def _attach_sales_cost(self, df, target_year=None, target_month=None, year_scope=None):
        if df is None or df.empty:
            return df
        if 'Revenue' not in df.columns:
            df['Revenue'] = self._extract_sales_revenue(df)
        else:
            df['Revenue'] = pd.to_numeric(df.get('Revenue'), errors='coerce')
            if not df['Revenue'].notna().any():
                df['Revenue'] = self._extract_sales_revenue(df)
        if 'Qty' not in df.columns:
            df['Qty'] = pd.to_numeric(df.get('数量'), errors='coerce')
        else:
            df['Qty'] = pd.to_numeric(df.get('Qty'), errors='coerce')
        if 'MonthStr' not in df.columns:
            date_col = next((c for c in df.columns if '日期' in str(c) or 'Date' in str(c)), None)
            if date_col:
                df['ParsedDate'] = pd.to_datetime(df[date_col], errors='coerce')
                df['MonthStr'] = df['ParsedDate'].dt.strftime('%Y-%m')
        if 'MonthStr' not in df.columns and target_year and target_month:
            df['MonthStr'] = f"{target_year}-{int(target_month):02d}"

        df['_UnitCost'] = None
        df['CostSource'] = None
        if 'MonthStr' in df.columns and '品目编码' in df.columns:
            for m_key, idx in df.groupby('MonthStr').groups.items():
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    continue
                unit_map = self._get_unit_cost_map_for_month(m_key)
                if not unit_map:
                    continue
                codes = df.loc[idx, '品目编码'].astype(str).str.strip()
                df.loc[idx, '_UnitCost'] = codes.map(unit_map)

        df['Cost'] = df['Qty'] * pd.to_numeric(df['_UnitCost'], errors='coerce')
        df.loc[df['Cost'].notna(), 'CostSource'] = 'unit_cost'
        if '销售出库供应价合计' in df.columns:
            raw_cost = pd.to_numeric(df.get('销售出库供应价合计'), errors='coerce')
            if self._can_use_raw_cost_fallback(df, raw_cost, df.get('Revenue')):
                mask = df['Cost'].isna() & raw_cost.notna()
                df.loc[mask, 'Cost'] = raw_cost.loc[mask]
                df.loc[mask, 'CostSource'] = 'raw_fallback'
        df.drop(columns=['_UnitCost'], errors='ignore', inplace=True)
        return df

    def _update_category_month_sheet(self, ws, target_year, target_month, year_scope=None):
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            return
        if 'ParsedDate' not in sales_df.columns:
            return
        df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, year_scope)
        annual_df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, "current")
        if (df is None or df.empty) and (annual_df is None or annual_df.empty):
            return

        def prepare_sales_scope(source_df, scope):
            if source_df is None or source_df.empty:
                return pd.DataFrame()
            work = source_df.copy()
            work['Category'] = work.get('品目组合1名')
            work['Revenue'] = self._extract_sales_revenue(work)
            work['Qty'] = pd.to_numeric(work.get('数量'), errors='coerce')
            return self._attach_sales_cost(work, target_year, target_month, scope)

        df = prepare_sales_scope(df, year_scope)
        annual_df = prepare_sales_scope(annual_df, "current")
        if df.empty and annual_df.empty:
            return

        header_map = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                header_map[str(header).strip()] = col

        month_metric_cols = {}
        for header, col in header_map.items():
            if '_' not in header:
                continue
            prefix, suffix = header.split('_', 1)
            m_key = self._label_to_month_key(prefix)
            if not m_key:
                continue
            metric = None
            if '毛利润' in suffix or '销售利润' in suffix:
                metric = 'profit'
            elif '销售成本' in suffix:
                metric = 'cost'
            elif '销售收入' in suffix:
                metric = 'revenue'
            elif '销售数量' in suffix:
                metric = 'qty'
            if metric:
                month_metric_cols.setdefault(prefix, {})[metric] = col

        annual_metric_cols = {
            'qty': header_map.get('年销售数量'),
            'revenue': header_map.get('年销售收入'),
            'cost': header_map.get('年销售成本'),
            'profit': header_map.get('年毛利润'),
        }
        category_col = header_map.get('产品大类', 1)

        metric_cols_to_clear = [c for c in annual_metric_cols.values() if c]
        for cols in month_metric_cols.values():
            metric_cols_to_clear.extend(cols.values())
        metric_cols_to_clear = sorted(set(metric_cols_to_clear))

        def _to_number(value):
            return self._to_float(value)

        def _share(value, total):
            if value is None or total in (None, 0):
                return None
            return value / total

        def _is_total_label(value):
            if value is None:
                return False
            text = str(value).strip()
            if not text:
                return False
            if self._normalize_category(text) == '合计':
                return True
            return text.upper() == 'TOTAL'

        def _apply_metric_format(row_idx, metric, col_idx):
            if not col_idx:
                return
            if metric == 'qty':
                ws.cell(row=row_idx, column=col_idx).number_format = '#,##0'
            else:
                ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

        def _apply_share_format(row_idx, col_idx):
            if not col_idx:
                return
            ws.cell(row=row_idx, column=col_idx).number_format = '0.00%'

        cat_stats = {}
        for cat, group in df.groupby(['Category']):
            if isinstance(cat, tuple):
                cat = cat[0] if cat else None
            if cat is None:
                continue
            if isinstance(cat, float) and pd.isna(cat):
                continue
            display = self._strip_category_share_suffix(cat)
            cat_key = self._normalize_category(display)
            if not cat_key or cat_key == '合计':
                continue

            metrics = self._calc_sales_metrics_from_group(group)
            revenue_val = metrics.get('revenue')
            cost_val = metrics.get('cost')
            qty_val = metrics.get('qty')
            profit_val = metrics.get('profit')

            month_values = {}
            for m_key, m_group in group.groupby('MonthStr'):
                m_label = self._month_key_to_label(m_key)
                m_metrics = self._calc_sales_metrics_from_group(m_group)
                m_revenue = m_metrics.get('revenue')
                m_cost = m_metrics.get('cost')
                m_qty = m_metrics.get('qty')
                m_profit = m_metrics.get('profit')
                month_values[m_label] = {
                    'profit': m_profit,
                    'cost': m_cost,
                    'revenue': m_revenue,
                    'qty': m_qty,
                }

            cat_stats[cat_key] = {
                'display': str(display).strip(),
                'qty': qty_val,
                'revenue': revenue_val,
                'cost': cost_val,
                'profit': profit_val,
                'sort_revenue': revenue_val,
                'month': month_values,
            }

        annual_cat_stats = {}
        if annual_df is not None and not annual_df.empty:
            for cat, group in annual_df.groupby(['Category']):
                if isinstance(cat, tuple):
                    cat = cat[0] if cat else None
                if cat is None:
                    continue
                if isinstance(cat, float) and pd.isna(cat):
                    continue
                display = self._strip_category_share_suffix(cat)
                cat_key = self._normalize_category(display)
                if not cat_key or cat_key == '合计':
                    continue
                metrics = self._calc_sales_metrics_from_group(group)
                annual_cat_stats[cat_key] = {
                    'display': str(display).strip(),
                    'qty': metrics.get('qty'),
                    'revenue': metrics.get('revenue'),
                    'cost': metrics.get('cost'),
                    'profit': metrics.get('profit'),
                }

        for cat_key, stat in cat_stats.items():
            annual_stat = annual_cat_stats.get(cat_key)
            for metric in ('qty', 'revenue', 'cost', 'profit'):
                stat[metric] = annual_stat.get(metric) if annual_stat else None

        for cat_key, annual_stat in annual_cat_stats.items():
            if cat_key in cat_stats:
                continue
            cat_stats[cat_key] = {
                'display': annual_stat.get('display') or cat_key,
                'qty': annual_stat.get('qty'),
                'revenue': annual_stat.get('revenue'),
                'cost': annual_stat.get('cost'),
                'profit': annual_stat.get('profit'),
                'sort_revenue': annual_stat.get('revenue'),
                'month': {},
            }

        if not cat_stats:
            return

        month_totals = {}
        for m_key, m_group in df.groupby('MonthStr'):
            m_label = self._month_key_to_label(m_key)
            m_metrics = self._calc_sales_metrics_from_group(m_group)
            m_revenue = m_metrics.get('revenue')
            m_cost = m_metrics.get('cost')
            m_qty = m_metrics.get('qty')
            month_totals[m_label] = {
                'profit': m_metrics.get('profit'),
                'cost': m_cost,
                'revenue': m_revenue,
                'qty': m_qty,
            }

        total_stats = self._calc_sales_metrics_from_group(annual_df) if annual_df is not None and not annual_df.empty else {}

        existing_order = []
        existing_display_map = {}
        first_data_row = 2
        total_row = None
        for r in range(2, ws.max_row + 1):
            raw_val = ws.cell(row=r, column=category_col).value
            if raw_val is None:
                continue
            text = str(raw_val).strip()
            if not text:
                continue
            if text.endswith('占比'):
                text = text[:-2]
            if _is_total_label(text):
                if total_row is None:
                    total_row = r
                continue
            cat_key = self._normalize_category(text)
            if not cat_key or cat_key == '合计':
                continue
            if cat_key not in existing_order:
                existing_order.append(cat_key)
            if cat_key not in existing_display_map:
                existing_display_map[cat_key] = text
            first_data_row = min(first_data_row, r)

        ordered_keys = [k for k in existing_order if k in cat_stats]
        for k in sorted(cat_stats.keys(), key=lambda x: abs(cat_stats[x].get('sort_revenue') or cat_stats[x].get('revenue') or 0), reverse=True):
            if k not in ordered_keys:
                ordered_keys.append(k)
        if not ordered_keys:
            return

        required_rows = len(ordered_keys) * 2
        required_last_row = first_data_row + required_rows - 1

        if total_row is None:
            total_row = required_last_row + 1
            self._insert_rows_preserve_merges(ws, total_row, 1)
            style_row = first_data_row if first_data_row <= ws.max_row else 1
            if style_row < total_row:
                self._copy_row_style(ws, style_row, total_row, max_col=max(category_col, ws.max_column))
        elif total_row <= required_last_row:
            insert_count = required_last_row - total_row + 1
            self._insert_rows_preserve_merges(ws, total_row, insert_count)
            for offset in range(insert_count):
                dst_row = total_row + offset
                src_row = total_row - 1 if total_row > 2 else total_row + insert_count
                if 1 <= src_row <= ws.max_row:
                    self._copy_row_style(ws, src_row, dst_row, max_col=max(category_col, ws.max_column))
            total_row += insert_count

        clear_end_row = max(required_last_row, total_row)
        for r in range(first_data_row, clear_end_row + 1):
            self._safe_set_cell_value(ws, r, category_col, None)
            for col in metric_cols_to_clear:
                self._safe_set_cell_value(ws, r, col, None)

        category_value_rows = {}
        for idx, cat_key in enumerate(ordered_keys):
            value_row = first_data_row + idx * 2
            share_row = value_row + 1
            stat = cat_stats[cat_key]
            cat_name = existing_display_map.get(cat_key) or stat['display'] or cat_key
            category_value_rows[cat_key] = value_row
            self._safe_set_cell_value(ws, value_row, category_col, cat_name)
            self._safe_set_cell_value(ws, share_row, category_col, f"{cat_name}占比")

            for metric, col_idx in annual_metric_cols.items():
                if not col_idx:
                    continue
                value = stat.get(metric)
                ws.cell(row=value_row, column=col_idx).value = value
                _apply_metric_format(value_row, metric, col_idx)
                share_val = _share(value, total_stats.get(metric))
                ws.cell(row=share_row, column=col_idx).value = share_val
                _apply_share_format(share_row, col_idx)

            for m_label, metric_cols in month_metric_cols.items():
                month_data = stat['month'].get(m_label, {})
                month_total = month_totals.get(m_label, {})
                for metric, col_idx in metric_cols.items():
                    value = month_data.get(metric)
                    ws.cell(row=value_row, column=col_idx).value = value
                    _apply_metric_format(value_row, metric, col_idx)
                    share_val = _share(value, month_total.get(metric))
                    ws.cell(row=share_row, column=col_idx).value = share_val
                    _apply_share_format(share_row, col_idx)

        self._safe_set_cell_value(ws, total_row, category_col, '合计')
        for metric, col_idx in annual_metric_cols.items():
            if not col_idx:
                continue
            ws.cell(row=total_row, column=col_idx).value = total_stats.get(metric)
            _apply_metric_format(total_row, metric, col_idx)

        for m_label, metric_cols in month_metric_cols.items():
            m_total = month_totals.get(m_label, {})
            for metric, col_idx in metric_cols.items():
                ws.cell(row=total_row, column=col_idx).value = m_total.get(metric)
                _apply_metric_format(total_row, metric, col_idx)

        if target_year and target_month:
            for m_label, metric_cols in month_metric_cols.items():
                m_key = self._label_to_month_key(m_label)
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    for r in range(first_data_row, total_row + 1):
                        for col in metric_cols.values():
                            self._safe_set_cell_value(ws, r, col, None)

        self._reorder_month_columns_grouped_by_suffix_desc(ws, header_row=1)

    def _fill_product_summary_total(self, ws, target_year, target_month, year_scope=None):
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            sales_df = pd.DataFrame(columns=['MonthStr', '品目编码', '品目名', '品目组合1名', '数量'])
        else:
            sales_df = sales_df.copy()
        if 'MonthStr' not in sales_df.columns:
            if 'ParsedDate' in sales_df.columns:
                sales_df['MonthStr'] = pd.to_datetime(sales_df['ParsedDate'], errors='coerce').dt.strftime('%Y-%m')
            else:
                date_col = next((c for c in sales_df.columns if '日期' in str(c) or 'Date' in str(c)), None)
                if date_col:
                    parsed = pd.to_datetime(sales_df[date_col], errors='coerce')
                    sales_df['MonthStr'] = parsed.dt.strftime('%Y-%m')
        display_sales_df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, year_scope)
        annual_sales_df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, "current")

        cost_df = None
        if target_year and target_month:
            cost_df = self.data['cost'].get(f"{target_year}-{int(target_month):02d}")
        if cost_df is None and self.data['cost']:
            cost_df = self.data['cost'][sorted(self.data['cost'].keys())[-1]]
        has_cost_data = cost_df is not None and not cost_df.empty

        if (display_sales_df is None or display_sales_df.empty) and (annual_sales_df is None or annual_sales_df.empty) and not has_cost_data:
            return

        header_map = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                header_map[str(header).strip()] = col

        code_col = header_map.get('品目编码')
        if not code_col:
            return

        def _refresh_header_map():
            refreshed = {}
            for c in range(1, ws.max_column + 1):
                value = ws.cell(row=1, column=c).value
                if value:
                    refreshed[str(value).strip()] = c
            return refreshed

        def _ensure_summary_column(header, reference_headers=None):
            nonlocal header_map
            if header in header_map:
                return header_map[header]
            reference_headers = reference_headers or []
            src_col = None
            for ref in reference_headers:
                if ref in header_map:
                    src_col = header_map[ref]
                    break
            if src_col is None:
                src_col = ws.max_column if ws.max_column else 1
            new_col = ws.max_column + 1
            if src_col:
                self._copy_column_style(ws, src_col, new_col)
            ws.cell(row=1, column=new_col).value = header
            header_map[header] = new_col
            return new_col

        def _move_named_columns_after(anchor_header, headers_to_move):
            nonlocal header_map, code_col
            header_map = _refresh_header_map()
            anchor_col = header_map.get(anchor_header)
            ordered_cols = [
                header_map[h]
                for h in headers_to_move
                if h in header_map
            ]
            if not anchor_col or not ordered_cols:
                return False
            expected_cols = list(range(anchor_col + 1, anchor_col + 1 + len(ordered_cols)))
            if ordered_cols == expected_cols:
                return False

            snapshots = [
                self._snapshot_column(ws, col, ws.max_row)
                for col in ordered_cols
            ]
            for col in sorted(ordered_cols, reverse=True):
                ws.delete_cols(col, 1)
            header_map = _refresh_header_map()
            anchor_col = header_map.get(anchor_header)
            if not anchor_col:
                return False
            insert_at = anchor_col + 1
            ws.insert_cols(insert_at, len(snapshots))
            for offset, snapshot in enumerate(snapshots):
                self._restore_column(ws, insert_at + offset, snapshot, ws.max_row)
            header_map = _refresh_header_map()
            code_col = header_map.get('品目编码')
            return True

        monthly_average_headers = {
            'qty': '月销售数量平均',
            'revenue': '月销售收入平均',
            'cost': '月销售成本平均',
            'profit': '月销售利润平均',
            'margin': '月毛利率平均',
        }
        monthly_average_cols = {
            metric: _ensure_summary_column(monthly_header, [annual_header])
            for metric, monthly_header, annual_header in [
                ('qty', monthly_average_headers['qty'], '年销售数量平均'),
                ('revenue', monthly_average_headers['revenue'], '年销售收入平均'),
                ('cost', monthly_average_headers['cost'], '年销售成本平均'),
                ('profit', monthly_average_headers['profit'], '年销售利润平均'),
                ('margin', monthly_average_headers['margin'], '年毛利率平均'),
            ]
        }
        _move_named_columns_after('年毛利率平均', list(monthly_average_headers.values()))
        monthly_average_cols = {
            metric: header_map.get(header)
            for metric, header in monthly_average_headers.items()
        }

        def _is_total_label(v):
            return self._is_summary_or_footer_label(v)

        def _normalize_output(v):
            if pd.isna(v):
                return None
            if hasattr(v, "item"):
                try:
                    return v.item()
                except Exception:
                    pass
            return v

        def prepare_sales_scope(source_df, scope):
            if source_df is None or source_df.empty:
                return pd.DataFrame(columns=['MonthStr', '品目编码', '品目名', '品目组合1名', 'Qty', 'Revenue', 'Cost'])
            work = source_df.copy()
            work['Revenue'] = self._extract_sales_revenue(work)
            work['Qty'] = pd.to_numeric(work.get('数量'), errors='coerce')
            return self._attach_sales_cost(work, target_year, target_month, scope)

        sales_df = prepare_sales_scope(display_sales_df, year_scope)
        annual_sales_df = prepare_sales_scope(annual_sales_df, "current")
        if sales_df.empty and annual_sales_df.empty and not has_cost_data:
            return

        month_cols = {}
        for header, col in header_map.items():
            m = re.match(r'(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?_(.+)', str(header))
            if not m:
                continue
            label = f"{m.group(1)}/{int(m.group(2)):02d}"
            suffix = m.group(3)
            metric = None
            if '期初库存数量' in suffix:
                metric = 'start_qty'
            elif '期初库存金额' in suffix:
                metric = 'start_amt'
            elif '期末库存数量' in suffix:
                metric = 'end_qty'
            elif '期末库存金额' in suffix:
                metric = 'end_amt'
            elif '库存数量变动' in suffix:
                metric = 'qty_change'
            elif '库存金额变动' in suffix:
                metric = 'amt_change'
            elif '毛利率' in suffix or '利润率' in suffix or suffix.endswith('率'):
                metric = 'margin'
            elif '销售利润' in suffix or '毛利润' in suffix:
                metric = 'profit'
            elif '销售成本' in suffix:
                metric = 'cost'
            elif '销售数量' in suffix:
                metric = 'qty'
            elif '销售收入' in suffix or '销售金额' in suffix:
                metric = 'revenue'
            if metric:
                month_cols.setdefault(label, {})[metric] = col

        if not month_cols and target_year and target_month:
            target_label = f"{target_year}/{int(target_month):02d}"
            default_suffixes = ['销售数量', '销售收入', '销售成本', '销售利润', '毛利率']
            last_col = ws.max_column
            start_col = last_col + 1
            for offset, suffix in enumerate(default_suffixes):
                col = start_col + offset
                ws.cell(row=1, column=col).value = f"{target_label}_{suffix}"
                self._copy_column_style(ws, last_col, col)
                header_map[f"{target_label}_{suffix}"] = col
                metric = None
                if suffix == '毛利率':
                    metric = 'margin'
                elif suffix == '销售利润':
                    metric = 'profit'
                elif suffix == '销售成本':
                    metric = 'cost'
                elif suffix == '销售数量':
                    metric = 'qty'
                elif suffix == '销售收入':
                    metric = 'revenue'
                if metric:
                    month_cols.setdefault(target_label, {})[metric] = col

        def build_sales_by_code(source_df):
            if source_df is None or source_df.empty or '品目编码' not in source_df.columns:
                empty = pd.DataFrame(columns=[
                    'year_qty', 'year_revenue', 'year_cost', 'year_cost_non_na',
                    'category', 'product_name', 'year_profit'
                ])
                empty.index.name = '品目编码'
                return empty
            work = source_df.copy()
            work['品目编码'] = work['品目编码'].astype(str).str.strip()
            grouped = work.groupby('品目编码').agg(
                year_qty=('Qty', 'sum'),
                year_revenue=('Revenue', 'sum'),
                year_cost=('Cost', lambda s: s.sum(min_count=1)),
                year_cost_non_na=('Cost', lambda s: int(pd.to_numeric(s, errors='coerce').notna().sum())),
                category=('品目组合1名', 'first'),
                product_name=('品目名', 'first')
            ).reset_index()
            grouped.loc[grouped['year_cost_non_na'] == 0, 'year_cost'] = None
            grouped['year_profit'] = grouped.apply(
                lambda r: (r['year_revenue'] - r['year_cost'])
                if pd.notna(r.get('year_revenue')) and pd.notna(r.get('year_cost'))
                else None,
                axis=1
            )
            grouped.set_index('品目编码', inplace=True)
            return grouped

        def build_product_meta_by_code(source_df):
            if source_df is None or source_df.empty or '品目编码' not in source_df.columns:
                empty = pd.DataFrame(columns=['category', 'product_name'])
                empty.index.name = '品目编码'
                return empty
            work = source_df.copy()
            work['品目编码'] = work['品目编码'].astype(str).str.strip()
            meta = work.groupby('品目编码').agg(
                category=('品目组合1名', 'first'),
                product_name=('品目名', 'first'),
            )
            return meta

        sales_by_code = build_sales_by_code(annual_sales_df)
        display_by_code = build_product_meta_by_code(sales_df)

        monthly_summary = sales_df.groupby(['MonthStr', '品目编码']).agg(
            qty=('Qty', 'sum'),
            revenue=('Revenue', 'sum'),
            cost=('Cost', lambda s: s.sum(min_count=1)),
            cost_non_na=('Cost', lambda s: int(pd.to_numeric(s, errors='coerce').notna().sum())),
        ).reset_index()
        monthly_summary.loc[monthly_summary['cost_non_na'] == 0, 'cost'] = None
        monthly_summary['profit'] = monthly_summary.apply(
            lambda r: (r['revenue'] - r['cost'])
            if pd.notna(r.get('revenue')) and pd.notna(r.get('cost'))
            else None,
            axis=1
        )
        monthly_summary['margin'] = monthly_summary.apply(
            lambda r: r['profit'] / r['revenue']
            if pd.notna(r.get('profit')) and pd.notna(r.get('revenue')) and r['revenue'] != 0
            else None,
            axis=1
        )
        monthly_map = {}
        for _, row in monthly_summary.iterrows():
            code = row.get('品目编码')
            if code is None or pd.isna(code):
                continue
            monthly_map[(str(row.get('MonthStr')), str(code).strip())] = row
        months_count_by_code = {}
        total_months_count = 0
        if annual_sales_df is not None and not annual_sales_df.empty and '品目编码' in annual_sales_df.columns:
            annual_count_df = annual_sales_df.copy()
            annual_count_df['品目编码'] = annual_count_df['品目编码'].astype(str).str.strip()
            months_count_by_code = (
                annual_count_df.groupby('品目编码')['MonthStr']
                .nunique()
                .to_dict()
            )
            total_months_count = int(annual_count_df['MonthStr'].nunique())

        def build_cost_month_map(df_cost):
            month_map = {}
            if df_cost is None or df_cost.empty:
                return month_map
            code_col_name = next((c for c in df_cost.columns if '品目编码' in str(c)), None)
            name_col = next((c for c in df_cost.columns if '品目名' in str(c)), None)
            qty_start_col = next((c for c in df_cost.columns if str(c).endswith('期初')), None)
            amt_start_col = next((c for c in df_cost.columns if str(c).endswith('期初.2')), None)
            qty_end_col = next((c for c in df_cost.columns if str(c).endswith('期末')), None)
            amt_end_col = next((c for c in df_cost.columns if str(c).endswith('期末.2')), None)
            if not code_col_name:
                return month_map
            for _, row in df_cost.iterrows():
                code = row.get(code_col_name)
                if code is None or pd.isna(code):
                    continue
                code = str(code).strip()
                if not code or _is_total_label(code):
                    continue
                month_map[code] = {
                    'name_spec': row.get(name_col),
                    'qty_start': self._to_float(row.get(qty_start_col)),
                    'amt_start': self._to_float(row.get(amt_start_col)),
                    'qty_end': self._to_float(row.get(qty_end_col)),
                    'amt_end': self._to_float(row.get(amt_end_col)),
                }
            return month_map

        cost_by_month = {}
        for m_key, m_df in self.data['cost'].items():
            if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                continue
            month_map = build_cost_month_map(m_df)
            if month_map:
                cost_by_month[m_key] = month_map

        cost_map = {}
        target_cost_key = f"{target_year}-{int(target_month):02d}" if target_year and target_month else None
        if target_cost_key and target_cost_key in cost_by_month:
            cost_map = cost_by_month.get(target_cost_key, {})
        elif cost_by_month:
            cost_map = cost_by_month[sorted(cost_by_month.keys())[-1]]
        elif cost_df is not None and not cost_df.empty:
            cost_map = build_cost_month_map(cost_df)

        def _row_value(row, key):
            if row is None:
                return None
            try:
                val = row.get(key)
            except Exception:
                return None
            if val is None:
                return None
            try:
                if pd.isna(val):
                    return None
            except Exception:
                pass
            return val

        def _first_value(*values):
            for val in values:
                if val is None:
                    continue
                text = str(val).strip()
                if not text or text.lower() == "nan":
                    continue
                return val
            return None

        # 补齐模板中不存在但当期有销售的品目编码，避免12月出现“有数据无行可写”
        existing_codes = {}
        total_row = None
        total_marker_cols = [1, code_col]
        for optional_col in (
            header_map.get('产品名称'),
            header_map.get('品目名规格'),
            header_map.get('产品大类'),
        ):
            if optional_col and optional_col not in total_marker_cols:
                total_marker_cols.append(optional_col)
        for r in range(2, ws.max_row + 1):
            if total_row is None and any(
                _is_total_label(ws.cell(row=r, column=c).value)
                for c in total_marker_cols
            ):
                total_row = r
                continue

            raw_code = ws.cell(row=r, column=code_col).value
            if raw_code is None:
                continue
            code = str(raw_code).strip()
            if not code or _is_total_label(code):
                continue
            if code not in existing_codes:
                existing_codes[code] = r

        missing_codes = []
        source_codes = []
        for idx in list(display_by_code.index) + list(sales_by_code.index) + list(cost_map.keys()):
            code = str(idx).strip()
            if code and code not in source_codes:
                source_codes.append(code)
        for code in source_codes:
            code = str(code).strip()
            if code and code not in existing_codes:
                missing_codes.append(code)
        if missing_codes:
            insert_row = total_row if total_row else ws.max_row + 1
            style_row = insert_row - 1 if insert_row > 2 else 2
            for code in missing_codes:
                self._insert_rows_preserve_merges(ws, insert_row)
                self._copy_row_style(ws, style_row, insert_row)
                for c in range(1, ws.max_column + 1):
                    self._safe_set_cell_value(ws, insert_row, c, None)
                ws.cell(row=insert_row, column=code_col).value = code
                sales_row = sales_by_code.loc[code] if code in sales_by_code.index else None
                meta_row = display_by_code.loc[code] if code in display_by_code.index else sales_row
                cost_row = cost_map.get(code, {})
                if '产品名称' in header_map and meta_row is not None:
                    ws.cell(row=insert_row, column=header_map['产品名称']).value = _first_value(
                        _row_value(meta_row, 'product_name'),
                        cost_row.get('name_spec'),
                        code,
                    )
                elif '产品名称' in header_map:
                    ws.cell(row=insert_row, column=header_map['产品名称']).value = _first_value(
                        cost_row.get('name_spec'),
                        code,
                    )
                resolved_category = self._resolve_uncategorized_product(
                    meta_row.get('category') if meta_row is not None else None,
                    cost_row.get('name_spec')
                )
                if '产品大类' in header_map and resolved_category is not None:
                    ws.cell(row=insert_row, column=header_map['产品大类']).value = resolved_category
                style_row = insert_row
                # 插入发生在合计行之前时，合计行会整体下移，需同步行号。
                if total_row is not None and insert_row <= total_row:
                    total_row += 1
                insert_row += 1

        if total_row is None:
            total_row = ws.max_row + 1
            if ws.max_row >= 2:
                self._copy_row_style(ws, ws.max_row, total_row)
            self._safe_set_cell_value(ws, total_row, 1, '合计')

        total_vals = {
            'current_qty': 0,
            'current_amt': 0,
            'year_qty': 0,
            'year_revenue': 0,
            'year_cost': 0,
            'year_profit': 0,
            'year_start_amt': 0,
            'year_end_amt': 0,
        }
        for r in range(2, ws.max_row + 1):
            if total_row and r == total_row:
                continue
            code = ws.cell(row=r, column=code_col).value
            if not code:
                continue
            code = str(code).strip()
            if _is_total_label(code):
                continue
            sales_row = sales_by_code.loc[code] if code in sales_by_code.index else None
            meta_row = display_by_code.loc[code] if code in display_by_code.index else sales_row
            cost_row = cost_map.get(code, {})
            resolved_category = self._resolve_uncategorized_product(
                _row_value(meta_row, 'category'),
                cost_row.get('name_spec')
            )
            if '产品名称' in header_map:
                product_name = _first_value(
                    _row_value(meta_row, 'product_name'),
                    cost_row.get('name_spec'),
                    code,
                )
                if product_name is not None:
                    ws.cell(row=r, column=header_map['产品名称']).value = product_name
            if '品目名规格' in header_map and cost_row.get('name_spec') is not None:
                ws.cell(row=r, column=header_map['品目名规格']).value = cost_row.get('name_spec')
            if '产品大类' in header_map and resolved_category is not None:
                ws.cell(row=r, column=header_map['产品大类']).value = resolved_category
            if '当前库存数量' in header_map:
                val = cost_row.get('qty_end')
                ws.cell(row=r, column=header_map['当前库存数量']).value = val
            if '当前库存金额' in header_map:
                val = cost_row.get('amt_end')
                ws.cell(row=r, column=header_map['当前库存金额']).value = val
            if '年销售数量合计' in header_map:
                val = _normalize_output(sales_row['year_qty']) if sales_row is not None else None
                ws.cell(row=r, column=header_map['年销售数量合计']).value = val
            if '年销售收入合计' in header_map:
                val = _normalize_output(sales_row['year_revenue']) if sales_row is not None else None
                ws.cell(row=r, column=header_map['年销售收入合计']).value = val
            if '年销售成本合计' in header_map:
                val = _normalize_output(sales_row['year_cost']) if sales_row is not None else None
                ws.cell(row=r, column=header_map['年销售成本合计']).value = val
            if '年销售利润合计' in header_map:
                val = _normalize_output(sales_row['year_profit']) if sales_row is not None else None
                ws.cell(row=r, column=header_map['年销售利润合计']).value = val
            if '年初存货金额' in header_map:
                val = cost_row.get('amt_start')
                ws.cell(row=r, column=header_map['年初存货金额']).value = val
            if '年末存货金额' in header_map:
                val = cost_row.get('amt_end')
                ws.cell(row=r, column=header_map['年末存货金额']).value = val

            months_count = months_count_by_code.get(code, 0)
            year_qty = sales_row['year_qty'] if sales_row is not None else None
            year_revenue = sales_row['year_revenue'] if sales_row is not None else None
            year_cost = sales_row['year_cost'] if sales_row is not None else None
            year_profit = sales_row['year_profit'] if sales_row is not None else None
            year_qty = self._to_float(year_qty)
            year_revenue = self._to_float(year_revenue)
            year_cost = self._to_float(year_cost)
            year_profit = self._to_float(year_profit)
            if year_qty is not None:
                total_vals['year_qty'] += year_qty
            if year_revenue is not None:
                total_vals['year_revenue'] += year_revenue
            if year_cost is not None:
                total_vals['year_cost'] += year_cost
            if year_profit is not None:
                total_vals['year_profit'] += year_profit
            start_amt = cost_row.get('amt_start')
            end_amt = cost_row.get('amt_end')
            end_qty = cost_row.get('qty_end')
            if start_amt is not None:
                total_vals['year_start_amt'] += start_amt
            if end_amt is not None:
                total_vals['year_end_amt'] += end_amt
                total_vals['current_amt'] += end_amt
            if end_qty is not None:
                total_vals['current_qty'] += end_qty
            avg_inventory = None
            if start_amt is not None and end_amt is not None:
                avg_inventory = (start_amt + end_amt) / 2
            elif end_amt is not None:
                avg_inventory = end_amt
            elif start_amt is not None:
                avg_inventory = start_amt
            turnover = (year_cost / avg_inventory) if (year_cost is not None and avg_inventory) else None
            turnover_days = (365 / turnover) if turnover else None
            avg_qty = (year_qty / months_count) if (months_count and year_qty is not None) else None
            avg_revenue = (year_revenue / months_count) if (months_count and year_revenue is not None) else None
            avg_cost = (year_cost / months_count) if (months_count and year_cost is not None) else None
            avg_profit = (year_profit / months_count) if (months_count and year_profit is not None) else None
            avg_margin = (avg_profit / avg_revenue) if (avg_profit is not None and avg_revenue) else None

            if '年平均存货' in header_map:
                ws.cell(row=r, column=header_map['年平均存货']).value = avg_inventory
            if '存货周转率' in header_map:
                ws.cell(row=r, column=header_map['存货周转率']).value = turnover
            if '存货周转天数' in header_map:
                ws.cell(row=r, column=header_map['存货周转天数']).value = turnover_days
            if '年销售数量平均' in header_map:
                ws.cell(row=r, column=header_map['年销售数量平均']).value = avg_qty
            if '年销售收入平均' in header_map:
                ws.cell(row=r, column=header_map['年销售收入平均']).value = avg_revenue
            if '年销售成本平均' in header_map:
                ws.cell(row=r, column=header_map['年销售成本平均']).value = avg_cost
            if '年销售利润平均' in header_map:
                ws.cell(row=r, column=header_map['年销售利润平均']).value = avg_profit
            if '年毛利率平均' in header_map:
                ws.cell(row=r, column=header_map['年毛利率平均']).value = avg_margin
            if monthly_average_cols.get('qty'):
                ws.cell(row=r, column=monthly_average_cols['qty']).value = avg_qty
            if monthly_average_cols.get('revenue'):
                ws.cell(row=r, column=monthly_average_cols['revenue']).value = avg_revenue
            if monthly_average_cols.get('cost'):
                ws.cell(row=r, column=monthly_average_cols['cost']).value = avg_cost
            if monthly_average_cols.get('profit'):
                ws.cell(row=r, column=monthly_average_cols['profit']).value = avg_profit
            if monthly_average_cols.get('margin'):
                ws.cell(row=r, column=monthly_average_cols['margin']).value = avg_margin

            if month_cols:
                for m_label, metric_cols in month_cols.items():
                    m_key = self._label_to_month_key(m_label)
                    if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                        continue
                    m_row = monthly_map.get((m_key, code))
                    c_map = cost_by_month.get(m_key, {}).get(code, {})
                    qty_change = None
                    amt_change = None
                    if c_map.get('qty_start') is not None and c_map.get('qty_end') is not None:
                        qty_change = c_map.get('qty_end') - c_map.get('qty_start')
                    if c_map.get('amt_start') is not None and c_map.get('amt_end') is not None:
                        amt_change = c_map.get('amt_end') - c_map.get('amt_start')
                    if 'start_qty' in metric_cols:
                        ws.cell(row=r, column=metric_cols['start_qty']).value = c_map.get('qty_start')
                    if 'start_amt' in metric_cols:
                        ws.cell(row=r, column=metric_cols['start_amt']).value = c_map.get('amt_start')
                    if 'end_qty' in metric_cols:
                        ws.cell(row=r, column=metric_cols['end_qty']).value = c_map.get('qty_end')
                    if 'end_amt' in metric_cols:
                        ws.cell(row=r, column=metric_cols['end_amt']).value = c_map.get('amt_end')
                    if 'qty_change' in metric_cols:
                        ws.cell(row=r, column=metric_cols['qty_change']).value = qty_change
                    if 'amt_change' in metric_cols:
                        ws.cell(row=r, column=metric_cols['amt_change']).value = amt_change
                    if m_row is None:
                        continue
                    if 'qty' in metric_cols:
                        ws.cell(row=r, column=metric_cols['qty']).value = _normalize_output(m_row.get('qty'))
                    if 'revenue' in metric_cols:
                        ws.cell(row=r, column=metric_cols['revenue']).value = _normalize_output(m_row.get('revenue'))
                    if 'cost' in metric_cols:
                        ws.cell(row=r, column=metric_cols['cost']).value = _normalize_output(m_row.get('cost'))
                    if 'profit' in metric_cols:
                        ws.cell(row=r, column=metric_cols['profit']).value = _normalize_output(m_row.get('profit'))
                    if 'margin' in metric_cols:
                        ws.cell(row=r, column=metric_cols['margin']).value = _normalize_output(m_row.get('margin'))

        current_qty_total = self._safe_sum([item.get('qty_end') for item in cost_map.values()])
        current_amt_total = self._safe_sum([item.get('amt_end') for item in cost_map.values()])
        if current_qty_total is not None:
            total_vals['current_qty'] = current_qty_total
        if current_amt_total is not None:
            total_vals['current_amt'] = current_amt_total
            total_vals['year_end_amt'] = current_amt_total

        if total_row:
            if '当前库存数量' in header_map:
                ws.cell(row=total_row, column=header_map['当前库存数量']).value = total_vals['current_qty']
            if '当前库存金额' in header_map:
                ws.cell(row=total_row, column=header_map['当前库存金额']).value = total_vals['current_amt']
            if '年销售数量合计' in header_map:
                ws.cell(row=total_row, column=header_map['年销售数量合计']).value = total_vals['year_qty']
            if '年销售收入合计' in header_map:
                ws.cell(row=total_row, column=header_map['年销售收入合计']).value = total_vals['year_revenue']
            if '年销售成本合计' in header_map:
                ws.cell(row=total_row, column=header_map['年销售成本合计']).value = total_vals['year_cost']
            if '年销售利润合计' in header_map:
                ws.cell(row=total_row, column=header_map['年销售利润合计']).value = total_vals['year_profit']
            if '年初存货金额' in header_map:
                ws.cell(row=total_row, column=header_map['年初存货金额']).value = total_vals['year_start_amt']
            if '年末存货金额' in header_map:
                ws.cell(row=total_row, column=header_map['年末存货金额']).value = total_vals['year_end_amt']

            # 合计行采用分母加权口径，避免简单算术平均造成偏差。
            total_avg_inventory = None
            if total_vals['year_start_amt'] is not None and total_vals['year_end_amt'] is not None:
                total_avg_inventory = (total_vals['year_start_amt'] + total_vals['year_end_amt']) / 2
            elif total_vals['year_end_amt'] is not None:
                total_avg_inventory = total_vals['year_end_amt']
            elif total_vals['year_start_amt'] is not None:
                total_avg_inventory = total_vals['year_start_amt']

            total_turnover = (
                total_vals['year_cost'] / total_avg_inventory
                if (total_vals['year_cost'] is not None and total_avg_inventory)
                else None
            )
            total_turnover_days = (365 / total_turnover) if total_turnover else None

            # 年均口径按期内月数归一，体现整体经营月均表现。
            avg_months = total_months_count if total_months_count else None
            total_avg_qty = (total_vals['year_qty'] / avg_months) if avg_months else None
            total_avg_revenue = (total_vals['year_revenue'] / avg_months) if avg_months else None
            total_avg_cost = (total_vals['year_cost'] / avg_months) if avg_months else None
            total_avg_profit = (total_vals['year_profit'] / avg_months) if avg_months else None
            total_avg_margin = (
                total_avg_profit / total_avg_revenue
                if (total_avg_profit is not None and total_avg_revenue)
                else None
            )

            if '年平均存货' in header_map:
                ws.cell(row=total_row, column=header_map['年平均存货']).value = total_avg_inventory
            if '存货周转率' in header_map:
                ws.cell(row=total_row, column=header_map['存货周转率']).value = total_turnover
            if '存货周转天数' in header_map:
                ws.cell(row=total_row, column=header_map['存货周转天数']).value = total_turnover_days
            if '年销售数量平均' in header_map:
                ws.cell(row=total_row, column=header_map['年销售数量平均']).value = total_avg_qty
            if '年销售收入平均' in header_map:
                ws.cell(row=total_row, column=header_map['年销售收入平均']).value = total_avg_revenue
            if '年销售成本平均' in header_map:
                ws.cell(row=total_row, column=header_map['年销售成本平均']).value = total_avg_cost
            if '年销售利润平均' in header_map:
                ws.cell(row=total_row, column=header_map['年销售利润平均']).value = total_avg_profit
            if '年毛利率平均' in header_map:
                ws.cell(row=total_row, column=header_map['年毛利率平均']).value = total_avg_margin
            if monthly_average_cols.get('qty'):
                ws.cell(row=total_row, column=monthly_average_cols['qty']).value = total_avg_qty
            if monthly_average_cols.get('revenue'):
                ws.cell(row=total_row, column=monthly_average_cols['revenue']).value = total_avg_revenue
            if monthly_average_cols.get('cost'):
                ws.cell(row=total_row, column=monthly_average_cols['cost']).value = total_avg_cost
            if monthly_average_cols.get('profit'):
                ws.cell(row=total_row, column=monthly_average_cols['profit']).value = total_avg_profit
            if monthly_average_cols.get('margin'):
                ws.cell(row=total_row, column=monthly_average_cols['margin']).value = total_avg_margin

            if month_cols:
                month_totals = monthly_summary.groupby('MonthStr').agg(
                    qty=('qty', 'sum'),
                    revenue=('revenue', 'sum'),
                    cost=('cost', lambda s: s.sum(min_count=1)),
                ).reset_index()
                month_totals['profit'] = month_totals.apply(
                    lambda r: (r['revenue'] - r['cost'])
                    if pd.notna(r.get('revenue')) and pd.notna(r.get('cost'))
                    else None,
                    axis=1
                )
                month_totals['margin'] = month_totals.apply(
                    lambda r: r['profit'] / r['revenue']
                    if pd.notna(r.get('profit')) and pd.notna(r.get('revenue')) and r['revenue'] != 0
                    else None,
                    axis=1
                )
                totals_map = {str(r['MonthStr']): r for _, r in month_totals.iterrows()}
                cost_totals_map = {}
                for m_key, month_map in cost_by_month.items():
                    start_qty = self._safe_sum([v.get('qty_start') for v in month_map.values()])
                    start_amt = self._safe_sum([v.get('amt_start') for v in month_map.values()])
                    end_qty = self._safe_sum([v.get('qty_end') for v in month_map.values()])
                    end_amt = self._safe_sum([v.get('amt_end') for v in month_map.values()])
                    qty_change = (end_qty - start_qty) if start_qty is not None and end_qty is not None else None
                    amt_change = (end_amt - start_amt) if start_amt is not None and end_amt is not None else None
                    cost_totals_map[str(m_key)] = {
                        'start_qty': start_qty,
                        'start_amt': start_amt,
                        'end_qty': end_qty,
                        'end_amt': end_amt,
                        'qty_change': qty_change,
                        'amt_change': amt_change,
                    }
                for m_label, metric_cols in month_cols.items():
                    m_key = self._label_to_month_key(m_label)
                    if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                        continue
                    m_row = totals_map.get(m_key)
                    c_row = cost_totals_map.get(m_key, {})
                    if 'start_qty' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['start_qty']).value = c_row.get('start_qty')
                    if 'start_amt' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['start_amt']).value = c_row.get('start_amt')
                    if 'end_qty' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['end_qty']).value = c_row.get('end_qty')
                    if 'end_amt' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['end_amt']).value = c_row.get('end_amt')
                    if 'qty_change' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['qty_change']).value = c_row.get('qty_change')
                    if 'amt_change' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['amt_change']).value = c_row.get('amt_change')
                    if m_row is None:
                        continue
                    if 'qty' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['qty']).value = _normalize_output(m_row.get('qty'))
                    if 'revenue' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['revenue']).value = _normalize_output(m_row.get('revenue'))
                    if 'cost' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['cost']).value = _normalize_output(m_row.get('cost'))
                    if 'profit' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['profit']).value = _normalize_output(m_row.get('profit'))
                    if 'margin' in metric_cols:
                        ws.cell(row=total_row, column=metric_cols['margin']).value = _normalize_output(m_row.get('margin'))

        if target_year and target_month and month_cols:
            for m_label, metric_cols in month_cols.items():
                m_key = self._label_to_month_key(m_label)
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    for r in range(2, ws.max_row + 1):
                        for col in metric_cols.values():
                            self._safe_set_cell_value(ws, r, col, None)

        self._reorder_month_columns_desc(ws, header_row=1)
        header_map = _refresh_header_map()

        for header, col in header_map.items():
            text = str(header or "").strip()
            letter = get_column_letter(col)
            is_annual_average = ("年" in text) and ("平均" in text or "均" in text)
            keep_visible = ("毛利率" in text) or ("利润率" in text)
            if is_annual_average:
                ws.column_dimensions[letter].hidden = not keep_visible
            elif text in monthly_average_headers.values():
                ws.column_dimensions[letter].hidden = False

        if total_row and total_row != 2:
            if self._move_row_preserve(ws, total_row, 2):
                total_row = 2

        # --- Pareto Chart Logic ---
        # 1. Extract Data
        pareto_data = []
        name_col_idx = header_map.get('产品名称') or header_map.get('品目名规格')
        rev_col_idx = header_map.get('年销售收入合计')
        
        if name_col_idx and rev_col_idx:
            for r in range(2, ws.max_row + 1):
                # Skip total row
                if r == total_row:
                    continue
                name = ws.cell(row=r, column=name_col_idx).value
                rev = self._to_float(ws.cell(row=r, column=rev_col_idx).value)
                if name and rev is not None and rev > 0:
                    pareto_data.append((name, rev))
        
        if pareto_data:
            # 2. Sort
            pareto_data.sort(key=lambda x: x[1], reverse=True)
            
            # 3. Calculate Cumulative %
            total_rev = sum(x[1] for x in pareto_data)
            cum_rev = 0
            pareto_rows = []
            for name, rev in pareto_data:
                cum_rev += rev
                cum_pct = cum_rev / total_rev if total_rev else 0
                pareto_rows.append([name, rev, cum_pct])
            
            # Limit to Top 20 or 80% for clarity if list is huge, but let's show all or Top 30
            pareto_rows = pareto_rows[:30] 

            # 4. Write to new area
            start_col = ws.max_column + 2
            headers = ["产品(按收入排序)", "销售收入", "累计占比"]
            self._write_table(ws, 1, start_col, headers, pareto_rows)
            
            # 5. Add Chart
            anchor = f"{get_column_letter(start_col + 4)}2"
            self._add_pareto_chart(
                ws, 
                start_col,     # Cats: Name
                start_col + 1, # Data: Revenue
                start_col + 2, # Line: Cum %
                1, 
                2, 
                1 + len(pareto_rows), 
                "产品销售帕累托分析 (ABC分析)", 
                anchor
            )
            self._write_chart_note(ws, start_col + 4, 1, "图表说明：前20%的产品通常贡献80%的收入(二八定律)。")

    def _product_compare_display_label(self, code, name=None):
        code_text = str(code or "").strip()
        name_text = str(name or "").strip()
        if not name_text or name_text.lower() == "nan" or name_text == code_text:
            return code_text
        if code_text and name_text.startswith(code_text):
            return name_text
        return f"{code_text} {name_text}".strip()

    def _product_compare_month_keys(self, target_year, target_month, source_month_keys=None):
        target_key = f"{target_year}-{int(target_month):02d}"
        trend_keys = [
            f"{target_year}-{month:02d}"
            for month in range(1, int(target_month) + 1)
        ]
        prev_year, prev_month = self._add_months(int(target_year), int(target_month), -12)
        mom_year, mom_month = self._add_months(int(target_year), int(target_month), -1)
        compare_keys = [
            f"{prev_year}-{prev_month:02d}",
            f"{mom_year}-{mom_month:02d}",
            target_key,
        ]
        # ``Series.unique()`` returns a NumPy array.  Using ``or`` on an
        # array with multiple elements attempts an ambiguous truth-value
        # check and raises ``ValueError``.
        valid_keys = set(source_month_keys) if source_month_keys is not None else set()
        ordered = []
        for key in trend_keys + compare_keys:
            if valid_keys and key not in valid_keys:
                continue
            if key not in ordered:
                ordered.append(key)
        return ordered

    def _update_hidden_chart_data(self, ws, target_year, target_month, year_scope=None):
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            return
        if 'ParsedDate' not in sales_df.columns and 'MonthStr' not in sales_df.columns:
            return

        source_df = sales_df.copy()
        if 'MonthStr' not in source_df.columns:
            source_df['MonthStr'] = pd.to_datetime(source_df['ParsedDate'], errors='coerce').dt.strftime('%Y-%m')
        source_df = source_df[source_df['MonthStr'].notna()].copy()
        source_df['MonthStr'] = source_df['MonthStr'].astype(str)
        source_df = self._filter_df_by_scope(source_df, target_year, target_month, "all")
        if source_df.empty:
            return

        required_month_keys = self._product_compare_month_keys(
            target_year,
            target_month,
            source_df['MonthStr'].dropna().astype(str).unique(),
        )
        if required_month_keys:
            source_df = source_df[source_df['MonthStr'].isin(required_month_keys)].copy()
        if source_df.empty:
            return

        source_df['Revenue'] = self._extract_sales_revenue(source_df)
        source_df['Qty'] = pd.to_numeric(source_df.get('数量'), errors='coerce')
        source_df = self._attach_sales_cost(source_df, target_year, target_month, "all")
        if '品目编码' not in source_df.columns:
            return

        source_df['Code'] = source_df['品目编码'].astype(str).str.strip()
        source_df = source_df[(source_df['Code'] != '') & (source_df['Code'].str.lower() != 'nan')].copy()
        if source_df.empty:
            return
        name_col = next((c for c in ['品目名', '产品名称', '品名', '商品名称'] if c in source_df.columns), None)
        source_df['ProductName'] = source_df[name_col].astype(str).str.strip() if name_col else source_df['Code']

        cost_by_month = {}
        for m_key, cost_df in self.data['cost'].items():
            if m_key not in required_month_keys:
                continue
            cost_by_month[m_key] = self._build_cost_inventory_map(cost_df, m_key)

        if target_year and target_month:
            start_label = f"{target_year}/01/01"
            end_label = f"{target_year}/{int(target_month):02d}/01"
            period_prefix = f"{start_label}-{end_label}"
        else:
            period_prefix = "期间"

        base_headers = [
            '期初库存数量', '期初库存金额', '期末库存数量', '期末库存金额',
            '销售收入', '销售数量', '销售成本', '销售利润', '毛利率',
            '库存数量变动', '库存金额变动'
        ]

        headers = ['产品名称']
        headers += [f"{period_prefix}_{h}" for h in base_headers]
        for m_key in required_month_keys:
            m_prefix = self._month_key_to_period_label(m_key)
            headers += [f"{m_prefix}_{h}" for h in base_headers]

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max(ws.max_column, len(headers))):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = header
        self._apply_header_style(ws, 1, max_col=len(headers))

        current_keys = [
            f"{target_year}-{month:02d}"
            for month in range(1, int(target_month) + 1)
        ]
        current_df = source_df[source_df['MonthStr'].isin(current_keys)].copy()
        sort_df = current_df if not current_df.empty else source_df
        code_stats = (
            sort_df.groupby('Code', dropna=False)
            .agg(
                revenue=('Revenue', lambda s: s.sum(min_count=1)),
                name=('ProductName', self._first_non_empty_value),
            )
            .reset_index()
        )
        for code in sorted(set().union(*[set(v.keys()) for v in cost_by_month.values()]) if cost_by_month else set()):
            if code not in set(code_stats['Code'].astype(str)):
                code_stats = pd.concat(
                    [
                        code_stats,
                        pd.DataFrame([{'Code': code, 'revenue': None, 'name': None}]),
                    ],
                    ignore_index=True,
                )
        code_stats['revenue'] = pd.to_numeric(code_stats['revenue'], errors='coerce')
        code_stats = code_stats.sort_values(['revenue', 'Code'], ascending=[False, True], na_position='last')

        first_month = current_keys[0] if current_keys else (required_month_keys[0] if required_month_keys else None)
        last_month = current_keys[-1] if current_keys else (required_month_keys[-1] if required_month_keys else None)
        row_idx = 2
        for _, row in code_stats.iterrows():
            code = str(row.get('Code') or '').strip()
            if not code:
                continue
            name = row.get('name')
            if not name or str(name).strip().lower() == "nan":
                for month_map in cost_by_month.values():
                    if code in month_map and month_map[code].get('name_spec'):
                        name = month_map[code].get('name_spec')
                        break
            ws.cell(row=row_idx, column=1).value = self._product_compare_display_label(code, name)

            period_group = current_df[current_df['Code'] == code]
            period_metrics = self._calc_sales_metrics_from_group(period_group)
            start_cost = cost_by_month.get(first_month, {}).get(code, {}) if first_month else {}
            end_cost = cost_by_month.get(last_month, {}).get(code, {}) if last_month else {}
            values = [
                start_cost.get('qty_start'),
                start_cost.get('amt_start'),
                end_cost.get('qty_end'),
                end_cost.get('amt_end'),
                period_metrics.get('revenue'),
                period_metrics.get('qty'),
                period_metrics.get('cost'),
                period_metrics.get('profit'),
                period_metrics.get('margin'),
            ]
            qty_change = (
                end_cost.get('qty_end') - start_cost.get('qty_start')
                if start_cost.get('qty_start') is not None and end_cost.get('qty_end') is not None
                else None
            )
            amt_change = (
                end_cost.get('amt_end') - start_cost.get('amt_start')
                if start_cost.get('amt_start') is not None and end_cost.get('amt_end') is not None
                else None
            )
            values += [qty_change, amt_change]

            col = 2
            for value in values:
                ws.cell(row=row_idx, column=col).value = value
                col += 1

            for m_key in required_month_keys:
                m_group = source_df[(source_df['MonthStr'] == m_key) & (source_df['Code'] == code)]
                m_metrics = self._calc_sales_metrics_from_group(m_group)
                c_map = cost_by_month.get(m_key, {}).get(code, {})
                qty_change = (
                    c_map.get('qty_end') - c_map.get('qty_start')
                    if c_map.get('qty_start') is not None and c_map.get('qty_end') is not None
                    else None
                )
                amt_change = (
                    c_map.get('amt_end') - c_map.get('amt_start')
                    if c_map.get('amt_start') is not None and c_map.get('amt_end') is not None
                    else None
                )
                for value in [
                    c_map.get('qty_start'),
                    c_map.get('amt_start'),
                    c_map.get('qty_end'),
                    c_map.get('amt_end'),
                    m_metrics.get('revenue'),
                    m_metrics.get('qty'),
                    m_metrics.get('cost'),
                    m_metrics.get('profit'),
                    m_metrics.get('margin'),
                    qty_change,
                    amt_change,
                ]:
                    ws.cell(row=row_idx, column=col).value = value
                    col += 1
            row_idx += 1

        for col in range(2, len(headers) + 1):
            header = str(ws.cell(row=1, column=col).value or "")
            if "毛利率" in header:
                fmt = "0.00%"
            elif "数量" in header:
                fmt = "#,##0"
            else:
                fmt = "#,##0.00"
            for row in range(2, row_idx):
                ws.cell(row=row, column=col).number_format = fmt

        ws.sheet_state = "hidden"

    def _update_product_compare_sheet(self, ws, target_year, target_month):
        if not target_year or not target_month:
            return
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            return

        wb = ws.parent
        if "图表数据源_隐藏" in wb.sheetnames:
            hidden_ws = wb["图表数据源_隐藏"]
        else:
            hidden_ws = wb.create_sheet("图表数据源_隐藏")
        self._update_hidden_chart_data(hidden_ws, target_year, target_month, self.year_scope)

        product_labels = []
        for r in range(2, hidden_ws.max_row + 1):
            label = hidden_ws.cell(row=r, column=1).value
            if label is None or str(label).strip() == "":
                continue
            product_labels.append(str(label).strip())
        if not product_labels:
            return

        selected_metric = str(ws.cell(row=1, column=2).value or "").strip()
        metric_options = [
            "销售收入",
            "销售数量",
            "销售成本",
            "销售利润",
            "毛利率",
            "库存数量变动",
            "库存金额变动",
            "期初库存数量",
            "期初库存金额",
            "期末库存数量",
            "期末库存金额",
        ]
        if selected_metric not in metric_options:
            selected_metric = "销售收入"

        product_by_code = {}
        for label in product_labels:
            code = str(label).split()[0] if str(label).split() else label
            product_by_code.setdefault(code, label)

        def resolve_selector_to_label(selector_text):
            text = str(selector_text or '').strip()
            if not text:
                return None
            if text in product_labels:
                return text
            first_token = re.split(r'\s+', text)[0]
            if first_token in product_by_code:
                return product_by_code[first_token]
            for label in product_labels:
                if text and text in label:
                    return label
            return None

        selector_count = 5
        selected_labels = []
        for col in range(2, 2 + selector_count):
            v = ws.cell(row=2, column=col).value
            if v is None:
                continue
            label = resolve_selector_to_label(v)
            if label and label not in selected_labels:
                selected_labels.append(label)
        for label in product_labels:
            if label not in selected_labels:
                selected_labels.append(label)
            if len(selected_labels) >= selector_count:
                break
        selected_labels = selected_labels[:selector_count]

        ws._charts = []
        ws.data_validations = DataValidationList()
        clear_rows = max(ws.max_row, 45)
        clear_cols = max(ws.max_column, 12)
        for row in ws.iter_rows(min_row=1, max_row=clear_rows, min_col=1, max_col=clear_cols):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.hyperlink = None

        ws.cell(row=1, column=1).value = "1. 选择对比指标:"
        ws.cell(row=1, column=2).value = selected_metric
        ws.cell(row=2, column=1).value = "2. 选择对比产品 (可多选):"
        for idx, label in enumerate(selected_labels, start=2):
            ws.cell(row=2, column=idx).value = label
        for col in range(2, 2 + selector_count):
            ws.cell(row=4, column=col).value = f"={get_column_letter(col)}$2"

        product_last_row = max(2, len(product_labels) + 1)
        metric_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(metric_options)}"',
            allow_blank=False,
        )
        product_dv = DataValidation(
            type="list",
            formula1=f"='图表数据源_隐藏'!$A$2:$A${product_last_row}",
            allow_blank=True,
        )
        ws.add_data_validation(metric_dv)
        ws.add_data_validation(product_dv)
        metric_dv.add("B1")
        product_dv.add(f"B2:{get_column_letter(1 + selector_count)}2")

        month_start_row = 5
        row_end = month_start_row + int(target_month) - 1
        ws.cell(row=4, column=1).value = "月份"

        hidden_name = "'图表数据源_隐藏'"

        def value_formula(product_ref, period_expr):
            return (
                f'=IF({product_ref}="","",IFERROR(INDEX({hidden_name}!$A:$ZZ,'
                f'MATCH({product_ref},{hidden_name}!$A:$A,0),'
                f'MATCH({period_expr}&"_"&$B$1,{hidden_name}!$1:$1,0)),""))'
            )

        for month in range(1, int(target_month) + 1):
            row = month_start_row + month - 1
            period_label = self._month_key_to_period_label(f"{target_year}-{month:02d}")
            ws.cell(row=row, column=1).value = period_label
            for col in range(2, 2 + selector_count):
                ws.cell(row=row, column=col).value = value_formula(f"{get_column_letter(col)}$2", f"$A{row}")

        self._apply_header_style(ws, 4, max_col=1 + selector_count)
        trend_number_format = "0.00%" if selected_metric == "毛利率" else "#,##0.00"
        for r in range(month_start_row, row_end + 1):
            for c in range(2, 2 + selector_count):
                ws.cell(row=r, column=c).number_format = trend_number_format

        self._add_line_chart_by_columns(
            ws,
            1,
            list(range(2, 2 + selector_count)),
            4,
            month_start_row,
            row_end,
            "多产品动态趋势对比",
            "H2",
        )

        target_key = f"{target_year}-{int(target_month):02d}"
        yoy_year, yoy_month = self._add_months(int(target_year), int(target_month), -12)
        mom_year, mom_month = self._add_months(int(target_year), int(target_month), -1)
        target_label = self._month_key_to_period_label(target_key)
        yoy_label = self._month_key_to_period_label(f"{yoy_year}-{yoy_month:02d}")
        mom_label = self._month_key_to_period_label(f"{mom_year}-{mom_month:02d}")

        def write_change_table(start_row, title, prior_header, prior_label, delta_header, rate_header):
            ws.cell(row=start_row - 1, column=1).value = title
            headers = ["产品", "本期值", prior_header, delta_header, rate_header]
            for offset, header in enumerate(headers, start=1):
                ws.cell(row=start_row, column=offset).value = header
            self._apply_header_style(ws, start_row, max_col=len(headers))

            for idx in range(selector_count):
                row = start_row + idx + 1
                selector_col = 2 + idx
                selector_letter = get_column_letter(selector_col)
                product_ref = f"${selector_letter}$2"
                ws.cell(row=row, column=1).value = f"={product_ref}"
                ws.cell(row=row, column=2).value = value_formula(product_ref, f'"{target_label}"')
                ws.cell(row=row, column=3).value = value_formula(product_ref, f'"{prior_label}"')
                ws.cell(row=row, column=4).value = f'=IF(OR(B{row}="",C{row}=""),"",B{row}-C{row})'
                ws.cell(row=row, column=5).value = f'=IF(OR(C{row}="",C{row}=0,D{row}=""),"",D{row}/C{row})'
                ws.cell(row=row, column=2).number_format = trend_number_format
                ws.cell(row=row, column=3).number_format = trend_number_format
                ws.cell(row=row, column=4).number_format = trend_number_format
                ws.cell(row=row, column=5).number_format = "0.00%"
            return start_row + selector_count

        yoy_start = max(18, row_end + 3)
        yoy_end = write_change_table(
            yoy_start,
            "选中产品同比分析",
            "上年同期",
            yoy_label,
            "同比增量",
            "同比增速",
        )
        self._add_bar_chart_by_columns(
            ws,
            1,
            [5],
            yoy_start,
            yoy_start + 1,
            yoy_end,
            "选中产品同比增速",
            f"H{yoy_start}",
        )

        mom_start = yoy_end + 4
        mom_end = write_change_table(
            mom_start,
            "选中产品环比分析",
            "上月值",
            mom_label,
            "环比增量",
            "环比增速",
        )
        self._add_bar_chart_by_columns(
            ws,
            1,
            [5],
            mom_start,
            mom_start + 1,
            mom_end,
            "选中产品环比增速",
            f"H{mom_start}",
        )

        self._write_chart_note(
            ws,
            8,
            mom_end + 3,
            "图表说明：B1切换指标、B2:F2切换产品后，趋势、同比、环比表格和图表会随公式联动更新。",
        )

        ws.column_dimensions["A"].width = 26
        for col in range(2, 2 + selector_count):
            ws.column_dimensions[get_column_letter(col)].width = 15
        for col in range(8, 14):
            ws.column_dimensions[get_column_letter(col)].width = 14

        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
        except Exception:
            pass

    def _update_chart_titles(self, wb, target_year, target_month):
        if not target_year or not target_month:
            return
        range_text = f"{target_year}/01-{target_year}/{int(target_month):02d}"
        for ws in wb.worksheets:
            for ch in ws._charts:
                title = ch.title
                if title and title.tx and title.tx.rich and title.tx.rich.p:
                    for p in title.tx.rich.p:
                        for run in p.r:
                            if run.t and '202' in run.t:
                                run.t = re.sub(r'20\d{2}/\d{2}\s*-\s*20\d{2}/\d{2}', range_text, run.t)

    def _normalize_party_name_key(self, value):
        if value is None or pd.isna(value):
            return ""
        text = str(value).strip().replace("\t", "")
        if not text or text.lower() in {"nan", "none", "nat"}:
            return ""
        return re.sub(r"\s+", " ", text)

    def _month_end_datetime(self, year, month):
        start = datetime(int(year), int(month), 1)
        end_dt = (start.replace(day=28) + pd.Timedelta(days=4)).replace(day=1) - pd.Timedelta(days=1)
        return end_dt if isinstance(end_dt, datetime) else end_dt.to_pydatetime()

    def _ar_turnover_period_bounds(self, target_year, target_month, year_scope=None):
        if year_scope is None:
            year_scope = self.year_scope
        year_scope = self._normalize_year_scope(year_scope)
        as_of = self._month_end_datetime(target_year, target_month)
        if year_scope == "current":
            start = datetime(int(target_year), 1, 1)
        else:
            start = datetime(int(target_year), int(target_month), 1)
            sales_df = self._get_sales_df()
            if sales_df is not None and not sales_df.empty and "MonthStr" in sales_df.columns:
                scoped = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, year_scope)
                month_keys = sorted(
                    str(v)
                    for v in scoped.get("MonthStr", pd.Series(dtype=object)).dropna().unique()
                    if re.match(r"^20\d{2}-\d{2}$", str(v))
                )
                if month_keys:
                    year, month = month_keys[0].split("-", 1)
                    start = datetime(int(year), int(month), 1)
        period_days = max((as_of.date() - start.date()).days + 1, 1)
        return start, as_of, period_days

    def _customer_sales_revenue_by_scope(self, target_year, target_month, year_scope=None):
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            return {}
        df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, year_scope)
        if df is None or df.empty:
            return {}
        cust_col = self._pick_ledger_column(
            df,
            ["往来单位名", "客户名称", "客户名", "客户", "收货公司", "对应往来单位名"],
        )
        if not cust_col:
            return {}
        df["Revenue"] = self._extract_sales_revenue(df)
        df["CustomerKey"] = df[cust_col].apply(self._normalize_party_name_key)
        df = df[(df["CustomerKey"] != "") & pd.to_numeric(df["Revenue"], errors="coerce").notna()].copy()
        if df.empty:
            return {}
        return pd.to_numeric(df["Revenue"], errors="coerce").groupby(df["CustomerKey"]).sum(min_count=1).to_dict()

    def _get_statement_ar_balance(self, target_year, target_month):
        month_key = f"{target_year}-{int(target_month):02d}" if target_year and target_month else None
        if not month_key:
            return None
        asset_df = self.data.get("asset", {}).get(month_key)
        if asset_df is None or asset_df.empty:
            return None
        balance_map = self._extract_balance_value_map(asset_df, target_year, target_month)
        for key in ("应收账款", "应收账款余额", "应收账款净额", "应收账款净值"):
            if key in balance_map:
                return balance_map[key]
            key_with_colon = f"{key}："
            if key_with_colon in balance_map:
                return balance_map[key_with_colon]
        return None

    def _update_ar_aging_sheet(self, ws, target_year, target_month):
        if self.ar_detail_df is None or self.ar_detail_df.empty:
            return
        if not target_year or not target_month:
            return

        df = self._drop_older_source_duplicate_rows(self.ar_detail_df.copy())
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].apply(lambda v: v.strip().replace('\t', '') if isinstance(v, str) else v)

        date_col = "ParsedDate" if "ParsedDate" in df.columns else None
        if not date_col:
            for c in df.columns:
                if '日期' in c:
                    date_col = c
                    break
        if not date_col:
            return

        cust_col = self._pick_ledger_column(
            df,
            ["往来单位名", "账页往来单位名", "客户名", "客户", "对应往来单位名"],
        )
        if not cust_col:
            return

        debit_col = None
        credit_col = None
        for c in df.columns:
            if '借方金额' in c:
                debit_col = c
            if '贷方金额' in c:
                credit_col = c
        if not debit_col or not credit_col:
            return

        def parse_date(val):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return None
            s = str(val)
            m = re.search(r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})', s)
            if m:
                s = m.group(1)
            try:
                return pd.to_datetime(s).to_pydatetime()
            except Exception:
                return None

        df['TxnDate'] = df[date_col].apply(parse_date)
        df = df.dropna(subset=['TxnDate', cust_col])
        df['CustomerKey'] = df[cust_col].apply(self._normalize_party_name_key)
        df = df[df['CustomerKey'] != ""].copy()
        df['Debit'] = df[debit_col].apply(self._to_float)
        df['Credit'] = df[credit_col].apply(self._to_float)
        df['Amount'] = df['Debit'].fillna(0) - df['Credit'].fillna(0)
        df = df[df['Amount'] != 0]

        period_start, as_of, period_days = self._ar_turnover_period_bounds(
            target_year,
            target_month,
            self.year_scope,
        )
        beginning_as_of = period_start - pd.Timedelta(days=1)
        df = df[df['TxnDate'] <= as_of].copy()
        if df.empty:
            return
        revenue_by_customer = self._customer_sales_revenue_by_scope(
            target_year,
            target_month,
            self.year_scope,
        )

        bucket_defs = [
            (30, '1-30天', 15, 0.1),
            (60, '31-60天', 45, 0.2),
            (90, '61-90天', 75, 0.4),
            (120, '91-120天', 105, 0.6),
            (10**9, '120天以上', 150, 1.0),
        ]

        results = []
        total_receivable = 0.0

        def build_open_items(group, cutoff_dt):
            open_items = []
            for _, row in group.sort_values('TxnDate').iterrows():
                if row['TxnDate'] > cutoff_dt:
                    continue
                amt = row['Amount']
                if amt > 0:
                    open_items.append([row['TxnDate'], amt])
                else:
                    credit = -amt
                    i = 0
                    while credit > 0 and i < len(open_items):
                        if open_items[i][1] <= credit:
                            credit -= open_items[i][1]
                            open_items[i][1] = 0
                        else:
                            open_items[i][1] -= credit
                            credit = 0
                        i += 1
                    open_items = [item for item in open_items if item[1] > 0]
            return open_items

        for cust, group in df.groupby(cust_col):
            group = group.sort_values('TxnDate')
            customer_key = self._normalize_party_name_key(cust)
            open_items = build_open_items(group, as_of)
            has_beginning_history = bool((group['TxnDate'] <= beginning_as_of).any())
            beginning_items = build_open_items(group, beginning_as_of) if has_beginning_history else []
            beginning_total = sum(item[1] for item in beginning_items) if has_beginning_history else None

            buckets = {name: 0.0 for _, name, _, _ in bucket_defs}
            for d, amt in open_items:
                age_days = (as_of - d).days
                if age_days < 0:
                    age_days = 0
                for limit, name, _, _ in bucket_defs:
                    if age_days <= limit:
                        buckets[name] += amt
                        break

            total = sum(buckets.values())
            if total <= 0:
                continue
            total_receivable += total

            weighted_age = 0.0
            risk_coeff = 0.0
            for limit, name, mid, weight in bucket_defs:
                weighted_age += buckets[name] * mid
                risk_coeff += buckets[name] * weight

            avg_age = weighted_age / total if total else 0
            over120 = buckets['120天以上']
            over120_ratio = over120 / total if total else 0

            results.append({
                'customer': cust,
                'customer_key': customer_key,
                'total': total,
                'beginning_total': beginning_total,
                'risk_coeff': risk_coeff,
                'over120_ratio': over120_ratio,
                'avg_age': avg_age,
                'buckets': buckets,
            })

        results.sort(key=lambda x: x['total'], reverse=True)

        if total_receivable == 0:
            return

        raw_total_receivable = total_receivable
        statement_ar_balance = self._get_statement_ar_balance(target_year, target_month)
        reconciled_to_statement = False
        if statement_ar_balance is not None and statement_ar_balance > 0 and abs(statement_ar_balance - total_receivable) > 0.01:
            factor = statement_ar_balance / total_receivable
            for item in results:
                item['raw_total'] = item['total']
                item['total'] *= factor
                if item.get('beginning_total') is not None:
                    item['beginning_total'] *= factor
                item['risk_coeff'] *= factor
                for bucket_name in item['buckets']:
                    item['buckets'][bucket_name] *= factor
            total_receivable = statement_ar_balance
            reconciled_to_statement = True

        for item in results:
            revenue = revenue_by_customer.get(item.get('customer_key') or "")
            beginning_total = item.get('beginning_total')
            ending_total = item.get('total')
            avg_ar = ((beginning_total + ending_total) / 2) if beginning_total is not None else ending_total
            if avg_ar is not None and avg_ar <= 0:
                avg_ar = None
            turnover = (revenue / avg_ar) if (revenue not in (None, 0) and avg_ar not in (None, 0)) else None
            turnover_days = (period_days / turnover) if turnover not in (None, 0) else None
            item['sales_revenue'] = revenue
            item['avg_ar'] = avg_ar
            item['turnover'] = turnover
            item['turnover_days'] = turnover_days

        headers = [
            "往来单位名",
            "总欠款",
            "占比",
            "风险系数",
            "120天以上占比",
            "平均账龄(天)",
            "1-30天",
            "31-60天",
            "61-90天",
            "91-120天",
            "120天以上",
            "销售收入",
            "期初应收",
            "平均应收",
            "应收账款周转率",
            "周转天数",
        ]

        ws._charts = []
        clear_cols = max(ws.max_column, len(headers) + 18)
        for r in range(1, ws.max_row + 1):
            for c in range(1, clear_cols + 1):
                self._safe_set_cell_value(ws, r, c, None)
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col).value = header
        self._apply_header_style(ws, 1, 1, len(headers))

        row = 2
        for item in results:
            ws.cell(row=row, column=1).value = item['customer']
            ws.cell(row=row, column=2).value = item['total']
            ws.cell(row=row, column=3).value = item['total'] / total_receivable
            ws.cell(row=row, column=4).value = item['risk_coeff']
            ws.cell(row=row, column=5).value = item['over120_ratio']
            ws.cell(row=row, column=6).value = item['avg_age']
            ws.cell(row=row, column=7).value = item['buckets']['1-30天']
            ws.cell(row=row, column=8).value = item['buckets']['31-60天']
            ws.cell(row=row, column=9).value = item['buckets']['61-90天']
            ws.cell(row=row, column=10).value = item['buckets']['91-120天']
            ws.cell(row=row, column=11).value = item['buckets']['120天以上']
            ws.cell(row=row, column=12).value = item.get('sales_revenue')
            ws.cell(row=row, column=13).value = item.get('beginning_total')
            ws.cell(row=row, column=14).value = item.get('avg_ar')
            ws.cell(row=row, column=15).value = item.get('turnover')
            ws.cell(row=row, column=16).value = item.get('turnover_days')
            row += 1

        summary_row = row + 1
        total_sales_revenue = self._safe_sum(item.get('sales_revenue') for item in results)
        beginning_values = [item.get('beginning_total') for item in results if item.get('beginning_total') is not None]
        total_beginning = sum(beginning_values) if beginning_values else None
        total_avg_ar = ((total_beginning + total_receivable) / 2) if total_beginning is not None else total_receivable
        if total_avg_ar is not None and total_avg_ar <= 0:
            total_avg_ar = None
        total_turnover = (
            total_sales_revenue / total_avg_ar
            if total_sales_revenue not in (None, 0) and total_avg_ar not in (None, 0)
            else None
        )
        total_turnover_days = (period_days / total_turnover) if total_turnover not in (None, 0) else None

        ws.cell(row=summary_row, column=1).value = "合计"
        ws.cell(row=summary_row, column=2).value = total_receivable
        ws.cell(row=summary_row, column=3).value = 1
        ws.cell(row=summary_row, column=5).value = (
            sum(item['buckets']['120天以上'] for item in results) / total_receivable
            if total_receivable else None
        )
        ws.cell(row=summary_row, column=7).value = sum(item['buckets']['1-30天'] for item in results)
        ws.cell(row=summary_row, column=8).value = sum(item['buckets']['31-60天'] for item in results)
        ws.cell(row=summary_row, column=9).value = sum(item['buckets']['61-90天'] for item in results)
        ws.cell(row=summary_row, column=10).value = sum(item['buckets']['91-120天'] for item in results)
        ws.cell(row=summary_row, column=11).value = sum(item['buckets']['120天以上'] for item in results)
        ws.cell(row=summary_row, column=12).value = total_sales_revenue
        ws.cell(row=summary_row, column=13).value = total_beginning
        ws.cell(row=summary_row, column=14).value = total_avg_ar
        ws.cell(row=summary_row, column=15).value = total_turnover
        ws.cell(row=summary_row, column=16).value = total_turnover_days
        self._apply_header_style(ws, summary_row, 1, len(headers))

        for r in range(2, summary_row + 1):
            for c in [2, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
                ws.cell(row=r, column=c).number_format = '#,##0.00'
            for c in [3, 5]:
                ws.cell(row=r, column=c).number_format = '0.00%'
            ws.cell(row=r, column=15).number_format = '0.00'
            ws.cell(row=r, column=16).number_format = '0.0'

        duplicate_summary = getattr(self, "_last_ledger_duplicate_summary", {}) or {}
        duplicate_rows = duplicate_summary.get("rows") or 0
        if reconciled_to_statement or duplicate_rows:
            detail_row = summary_row + 1
            if reconciled_to_statement:
                ws.cell(row=detail_row, column=1).value = "明细账龄原始合计"
                ws.cell(row=detail_row, column=2).value = raw_total_receivable
                detail_row += 1

            if duplicate_rows:
                ws.cell(row=detail_row, column=1).value = "已剔除重复录入行数"
                ws.cell(row=detail_row, column=2).value = duplicate_rows
                detail_row += 1
                ws.cell(row=detail_row, column=1).value = "已剔除重复录入借方金额"
                ws.cell(row=detail_row, column=2).value = duplicate_summary.get("debit")
                detail_row += 1
                ws.cell(row=detail_row, column=1).value = "已剔除重复录入贷方金额"
                ws.cell(row=detail_row, column=2).value = duplicate_summary.get("credit")
                detail_row += 1
                ws.cell(row=detail_row, column=1).value = "已剔除重复录入净额"
                ws.cell(row=detail_row, column=2).value = duplicate_summary.get("net")
                detail_row += 1

            if reconciled_to_statement:
                ws.cell(row=detail_row, column=1).value = "资产负债表应收账款"
                ws.cell(row=detail_row, column=2).value = statement_ar_balance
                detail_row += 1
                ws.cell(row=detail_row, column=1).value = "校准前差异"
                ws.cell(row=detail_row, column=2).value = statement_ar_balance - raw_total_receivable
                detail_row += 1
                ws.cell(row=detail_row, column=1).value = "校准后差异"
                ws.cell(row=detail_row, column=2).value = total_receivable - statement_ar_balance
                detail_row += 1

            ws.cell(row=detail_row, column=1).value = "说明"
            ws.cell(row=detail_row, column=2).value = (
                "客户账龄按明细账龄结构分配，并按资产负债表应收账款余额校准；"
                "已剔除重复录入金额按旧来源重复行的借方、贷方和净额列示；"
                "校准前差异通常还可能来自结转/期初行或报表与明细口径不一致；"
                f"周转率=销售收入/平均应收，周转天数={period_days}天/周转率，"
                "缺期初应收时以期末应收作为平均应收兜底。"
            )
        else:
            detail_row = summary_row + 1
            ws.cell(row=detail_row, column=1).value = "说明"
            ws.cell(row=detail_row, column=2).value = (
                f"周转率=销售收入/平均应收，周转天数={period_days}天/周转率；"
                "销售收入按当前报表范围汇总，缺期初应收时以期末应收作为平均应收兜底。"
            )

        chart_start_col = len(headers) + 2
        chart_anchor_col = chart_start_col + 5
        turnover_chart_rows = [
            [
                self._short_chart_label(item['customer']),
                item.get('total'),
                item.get('turnover'),
                item.get('turnover_days'),
            ]
            for item in results[:10]
        ]
        if turnover_chart_rows:
            self._write_table(
                ws,
                1,
                chart_start_col,
                ["客户", "总欠款", "应收账款周转率", "周转天数"],
                turnover_chart_rows,
            )
            self._add_combo_chart(
                ws,
                chart_start_col,
                [chart_start_col + 1],
                [chart_start_col + 2],
                1,
                2,
                1 + len(turnover_chart_rows),
                "应收余额与周转率Top客户",
                f"{get_column_letter(chart_anchor_col)}2",
            )
            self._write_chart_note(
                ws,
                chart_anchor_col,
                1,
                "图表说明：柱形展示应收余额，折线展示同一客户应收账款周转率。",
            )
            self._add_detailed_analysis_box(ws, chart_anchor_col, 14, "应收周转率", [
                "周转率越高，表示销售形成现金回收的效率越高。",
                "余额高且周转率低的客户应优先跟进账期和回款计划。",
                "有余额但无销售收入的客户，通常需要检查历史欠款或长期挂账。"
            ])

        # 客户数量过多时 100% 堆积图会严重压缩，改为仅展示 TopN 以保证可读性。
        chart_top_n = min(len(results), 20)
        if chart_top_n > 0:
            anchor_stacked = f"{get_column_letter(chart_anchor_col)}27"
            self._add_stacked_bar_chart(
                ws,
                1,  # Categories: Customer
                [7, 8, 9, 10, 11],  # Data: Buckets
                1,
                2,
                1 + chart_top_n,
                f"应收账款账龄结构 (Top{chart_top_n}客户 100%堆积)",
                anchor_stacked,
                percent=True,
            )
            self._write_chart_note(
                ws,
                chart_anchor_col,
                26,
                f"图表说明：Top{chart_top_n}客户应收账款的账龄结构分布(100%堆积)。",
            )
    def check_data_completeness(self, target_year, target_month):
        target_key = f"{target_year}-{int(target_month):02d}"
        missing_cats = []
        for cat in ['profit', 'cost', 'expense', 'asset', 'sales']:
            if target_key not in self.data[cat]:
                missing_cats.append(cat)
        has_ar_detail = self.ar_detail_df is not None and not self.ar_detail_df.empty
        has_monthly_ar = target_key in self.data.get('ar', {})
        if not has_ar_detail and not has_monthly_ar:
            missing_cats.append('ar')
        return missing_cats

    def _find_month_column(self, ws, target_year, target_month, header_row=1):
        target_label = f"{target_year}/{int(target_month):02d}"
        for col in range(1, ws.max_column + 1):
            label = self._normalize_month_label(ws.cell(row=header_row, column=col).value)
            if label == target_label:
                return col
        return None

    def _find_row_by_label(self, ws, label, col=1, start_row=1):
        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None and str(v).strip() == label:
                return r
        return None

    def _validate_generated_report(self, report_path, target_year, target_month, year_scope=None):
        """校验已生成报表关键值；返回问题列表。"""
        issues = []
        if not target_year or not target_month:
            return issues

        if year_scope is None:
            year_scope = self.year_scope

        dq_summary = self._get_data_quality_summary_for_scope(
            target_year=target_year,
            target_month=target_month,
            year_scope=year_scope,
        )
        if dq_summary.get("ERROR", 0) > 0:
            issues.append({
                "severity": "WARN",
                "sheet": "数据质量检查",
                "message": (
                    f"源数据质量存在 ERROR={dq_summary.get('ERROR', 0)}、"
                    f"WARN={dq_summary.get('WARN', 0)}，建议先处理再使用报告结论"
                ),
            })
        elif dq_summary.get("WARN", 0) > 0:
            issues.append({
                "severity": "WARN",
                "sheet": "数据质量检查",
                "message": f"源数据质量存在 WARN={dq_summary.get('WARN', 0)}，请结合数据质量检查页复核",
            })

        month_key = f"{target_year}-{int(target_month):02d}"
        profit_df = self.data['profit'].get(month_key)
        asset_df = self.data['asset'].get(month_key)
        if profit_df is None or profit_df.empty:
            issues.append({"severity": "ERROR", "sheet": "利润表", "message": f"缺少源数据 {month_key}"})
        if asset_df is None or asset_df.empty:
            issues.append({"severity": "ERROR", "sheet": "资产负债表", "message": f"缺少源数据 {month_key}"})
        if issues:
            return issues

        profit_source = {}
        name_col = profit_df.columns[0]
        for _, row in profit_df.iterrows():
            key = self._normalize_profit_label(row.get(name_col))
            if not key:
                continue
            val = self._row_value_for_month(row, profit_df.columns[1:], month_key)
            if val is None:
                continue
            profit_source[key] = val

        asset_source = self._extract_balance_value_map(asset_df, target_year, target_month)
        metrics_by_month = self._build_monthly_metrics(target_year, target_month, year_scope)
        target_metrics = metrics_by_month.get(month_key, {})

        try:
            wb_val = openpyxl.load_workbook(report_path, data_only=True)
            wb_formula = openpyxl.load_workbook(report_path, data_only=False)
        except Exception as e:
            issues.append({"severity": "ERROR", "sheet": "文件", "message": f"无法打开生成报表: {e}"})
            return issues

        # 利润表关键值校验
        if '利润表' in wb_val.sheetnames:
            ws = wb_val['利润表']
            month_col = self._find_month_column(ws, target_year, target_month, header_row=1)
            if not month_col:
                issues.append({"severity": "ERROR", "sheet": "利润表", "message": f"缺少月份列 {target_year}/{int(target_month):02d}"})
            else:
                checks = [
                    ("主营业务收入", "主营业务收入"),
                    ("主营业务成本", "主营业务成本"),
                    ("销售费用", "销售费用"),
                    ("管理费用", "管理费用"),
                    ("二、营业利润", "营业利润"),
                    ("营业利润", "营业利润"),
                    ("三、利润总额", "利润总额"),
                    ("四、净利润", "净利润"),
                ]
                for report_label, source_key in checks:
                    expected = profit_source.get(source_key)
                    if expected is None:
                        continue
                    row_idx = self._find_row_by_label(ws, report_label, col=1, start_row=2)
                    if not row_idx:
                        issues.append({"severity": "WARN", "sheet": "利润表", "message": f"缺少指标行: {report_label}"})
                        continue
                    actual = self._to_float(ws.cell(row=row_idx, column=month_col).value)
                    if actual is None or abs(actual - expected) > 1e-6:
                        issues.append({
                            "severity": "ERROR",
                            "sheet": "利润表",
                            "message": f"{report_label}不一致: 报表={actual}, 源={expected}"
                        })

            # 利润表“全年汇总/全年合计”应与月度列求和一致，避免模板残留值。
            annual_cols = []
            scoped_month_cols = []
            annual_tokens = ("全年汇总", "全年合计", "本年累计", "年累计", "累计")
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header is None:
                    continue
                h_text = str(header).strip()
                m_key = self._label_to_month_key(h_text) or self._extract_month_key_from_text(h_text)
                if m_key and self._month_key_in_scope(m_key, target_year, target_month, "current"):
                    scoped_month_cols.append(col)
                if any(token in h_text for token in annual_tokens):
                    annual_cols.append(col)
            if annual_cols and scoped_month_cols:
                for r in range(2, ws.max_row + 1):
                    label = ws.cell(row=r, column=1).value
                    if label is None or str(label).strip() == "":
                        continue
                    vals = []
                    for col in scoped_month_cols:
                        v = self._to_float(ws.cell(row=r, column=col).value)
                        if v is not None:
                            vals.append(v)
                    if not vals:
                        continue
                    expected_total = sum(vals)
                    for col in annual_cols:
                        actual_total = self._to_float(ws.cell(row=r, column=col).value)
                        if actual_total is None:
                            continue
                        if abs(actual_total - expected_total) > 1e-6:
                            issues.append({
                                "severity": "ERROR",
                                "sheet": "利润表",
                                "message": f"全年汇总不一致[{label}]: 报表={actual_total}, 月和={expected_total}"
                            })
                            break
        else:
            issues.append({"severity": "ERROR", "sheet": "利润表", "message": "缺少Sheet: 利润表"})

        # 年度汇总总览关键年度值校验（若存在）。
        if '年度汇总总览' in wb_val.sheetnames:
            ws = wb_val['年度汇总总览']
            month_keys = [
                m for m in self._annual_month_keys(int(target_year))
                if self._month_key_in_scope(m, target_year, target_month, year_scope)
            ]

            def _sum_metric(metric_key):
                vals = [metrics_by_month.get(m, {}).get(metric_key) for m in month_keys]
                return self._safe_sum(vals)

            expected_map = {
                "主营业务收入(全年)": _sum_metric("revenue"),
                "主营业务成本(全年)": _sum_metric("cost"),
                "财务费用(全年)": _sum_metric("financial_expense"),
                "营业利润(全年)": _sum_metric("operating_profit"),
                "净利润(全年)": _sum_metric("net_profit"),
            }
            row_map = {}
            for r in range(1, ws.max_row + 1):
                label = ws.cell(row=r, column=1).value
                if label is None:
                    continue
                row_map[str(label).strip()] = r
            for label, expected in expected_map.items():
                if expected is None:
                    continue
                row_idx = row_map.get(label)
                if not row_idx:
                    continue
                actual = self._to_float(ws.cell(row=row_idx, column=2).value)
                if actual is None or abs(actual - expected) > 1e-6:
                    issues.append({
                        "severity": "ERROR",
                        "sheet": "年度汇总总览",
                        "message": f"{label}不一致: 报表={actual}, 源={expected}"
                    })

        # 资产负债表关键值校验
        if '资产负债表' in wb_val.sheetnames:
            ws = wb_val['资产负债表']
            month_col = self._find_month_column(ws, target_year, target_month, header_row=1)
            if not month_col:
                issues.append({"severity": "ERROR", "sheet": "资产负债表", "message": f"缺少月份列 {target_year}/{int(target_month):02d}"})
            else:
                checks = [
                    ("流动资产", "流动资产"),
                    ("流动资产合计", "流动资产"),
                    ("货币资金", "货币资金"),
                    ("应收账款", "应收账款"),
                    ("存货", "存货"),
                    ("负债合计", "负债合计"),
                    ("所有者权益合计", "所有者权益合计"),
                    ("负债和所有者权益总计", "负债和所有者权益总计"),
                ]
                for report_label, source_key in checks:
                    expected = asset_source.get(source_key)
                    if expected is None:
                        continue
                    row_idx = self._find_row_by_label(ws, report_label, col=1, start_row=2)
                    if not row_idx:
                        issues.append({"severity": "WARN", "sheet": "资产负债表", "message": f"缺少指标行: {report_label}"})
                        continue
                    actual = self._to_float(ws.cell(row=row_idx, column=month_col).value)
                    if actual is None or abs(actual - expected) > 1e-6:
                        issues.append({
                            "severity": "ERROR",
                            "sheet": "资产负债表",
                            "message": f"{report_label}不一致: 报表={actual}, 源={expected}"
                        })
        else:
            issues.append({"severity": "ERROR", "sheet": "资产负债表", "message": "缺少Sheet: 资产负债表"})

        # 应收账龄合计应与资产负债表应收账款余额一致；账龄结构可来自明细分配。
        if '应收账款账龄分析' in wb_val.sheetnames:
            ws = wb_val['应收账款账龄分析']
            expected_ar = asset_source.get("应收账款") or asset_source.get("应收账款：")
            total_row = self._find_row_by_label(ws, "合计", col=1, start_row=2)
            if expected_ar is not None and total_row:
                actual_ar = self._to_float(ws.cell(row=total_row, column=2).value)
                if actual_ar is None or abs(actual_ar - expected_ar) > 0.01:
                    issues.append({
                        "severity": "ERROR",
                        "sheet": "应收账款账龄分析",
                        "message": f"账龄合计与资产负债表应收账款不一致: 账龄={actual_ar}, 资产负债表={expected_ar}"
                    })
            elif expected_ar is not None:
                issues.append({
                    "severity": "WARN",
                    "sheet": "应收账款账龄分析",
                    "message": "缺少合计行，无法核对应收账款账龄合计"
                })

        # 经营指标关键值校验
        if '经营指标' in wb_val.sheetnames:
            ws = wb_val['经营指标']
            header_map = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    header_map[str(header).strip()] = col

            month_row_map = {}
            scoped_rows = []
            for r in range(2, ws.max_row + 1):
                mk = self._label_to_month_key(ws.cell(row=r, column=1).value)
                if not mk:
                    continue
                month_row_map[mk] = r
                if self._month_key_in_scope(mk, target_year, target_month, year_scope):
                    scoped_rows.append(r)

            target_row = month_row_map.get(month_key)
            if not target_row:
                issues.append({
                    "severity": "ERROR",
                    "sheet": "经营指标",
                    "message": f"缺少目标月份行: {target_year}/{int(target_month):02d}"
                })
            else:
                metric_checks = [
                    ("主营业务收入", "revenue"),
                    ("主营业务成本", "cost"),
                    ("销售费用", "sales_expense"),
                    ("管理费用", "admin_expense"),
                    ("财务费用", "financial_expense"),
                    ("营业利润", "operating_profit"),
                    ("应收账款余额", "ar_balance"),
                    ("存货期末余额", "inventory_end"),
                    ("存货期初余额", "inventory_start"),
                    ("主营业务成本成本率", "cost_rate"),
                    ("销售费用率", "sales_expense_rate"),
                    ("管理费用率", "admin_expense_rate"),
                    ("财务费用率", "financial_expense_rate"),
                    ("营业利润率", "operating_profit_rate"),
                    ("存货周转天数", "inventory_turnover_days"),
                ]
                for header, key in metric_checks:
                    col = header_map.get(header)
                    if not col:
                        continue
                    expected = target_metrics.get(key)
                    if expected is None:
                        continue
                    actual = self._to_float(ws.cell(row=target_row, column=col).value)
                    if actual is None or abs(actual - expected) > 1e-6:
                        issues.append({
                            "severity": "ERROR",
                            "sheet": "经营指标",
                            "message": f"{header}不一致: 报表={actual}, 源={expected}"
                        })

            if scoped_rows:
                sorted_rows = sorted(scoped_rows)
                expected_rows = list(range(sorted_rows[0], sorted_rows[-1] + 1))
                if sorted_rows != expected_rows:
                    issues.append({
                        "severity": "WARN",
                        "sheet": "经营指标",
                        "message": f"月份行不连续: 行{sorted_rows[0]}到行{sorted_rows[-1]}之间存在空洞"
                    })

            # 经营指标“合计”行应与当前口径月度数据一致，避免模板残留值。
            summary_labels = {"合计", "全年合计", "全年汇总", "本年累计", "年累计"}
            summary_row = None
            for r in range(2, ws.max_row + 1):
                label = ws.cell(row=r, column=1).value
                if isinstance(label, str) and label.strip() in summary_labels:
                    summary_row = r
                    break
            if summary_row:
                scoped_month_keys = [
                    m for m in sorted(month_row_map.keys())
                    if self._month_key_in_scope(m, target_year, target_month, year_scope)
                ]
                checks = [
                    ("主营业务收入", "revenue"),
                    ("主营业务成本", "cost"),
                    ("销售费用", "sales_expense"),
                    ("管理费用", "admin_expense"),
                    ("财务费用", "financial_expense"),
                    ("营业利润", "operating_profit"),
                ]
                for header, key in checks:
                    col = header_map.get(header)
                    if not col:
                        continue
                    expected = self._safe_sum([metrics_by_month.get(m, {}).get(key) for m in scoped_month_keys])
                    if expected is None:
                        continue
                    actual = self._to_float(ws.cell(row=summary_row, column=col).value)
                    if actual is None or abs(actual - expected) > 1e-6:
                        issues.append({
                            "severity": "ERROR",
                            "sheet": "经营指标",
                            "message": f"合计行{header}不一致: 报表={actual}, 月和={expected}"
                        })
        else:
            issues.append({"severity": "ERROR", "sheet": "经营指标", "message": "缺少Sheet: 经营指标"})

        # 费用对比“合计”行与月度汇总一致性校验。
        if '费用对比' in wb_val.sheetnames:
            ws = wb_val['费用对比']
            header_map = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    header_map[str(header).strip()] = col
            summary_labels = {"合计", "全年合计", "全年汇总", "本年累计", "年累计"}
            summary_row = None
            for r in range(2, ws.max_row + 1):
                label = ws.cell(row=r, column=1).value
                if isinstance(label, str) and label.strip() in summary_labels:
                    summary_row = r
                    break
            if summary_row:
                scoped_month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
                checks = [
                    ("主营业务收入", "revenue"),
                    ("主营业务成本", "cost"),
                    ("销售费用", "sales_expense"),
                    ("管理费用", "admin_expense"),
                    ("财务费用", "financial_expense"),
                    ("营业利润", "operating_profit"),
                ]
                for header, key in checks:
                    col = header_map.get(header)
                    if not col:
                        continue
                    expected = self._safe_sum([metrics_by_month.get(m, {}).get(key) for m in scoped_month_keys])
                    if expected is None:
                        continue
                    actual = self._to_float(ws.cell(row=summary_row, column=col).value)
                    if actual is None or abs(actual - expected) > 1e-6:
                        issues.append({
                            "severity": "ERROR",
                            "sheet": "费用对比",
                            "message": f"合计行{header}不一致: 报表={actual}, 月和={expected}"
                        })

        # 仪表盘公式/范围校验
        if '仪表盘' in wb_formula.sheetnames and '经营指标' in wb_formula.sheetnames:
            dash_ws = wb_formula['仪表盘']
            metric_ws = wb_formula['经营指标']
            month_rows = []
            month_row_map = {}
            for r in range(2, metric_ws.max_row + 1):
                m_key = self._label_to_month_key(metric_ws.cell(row=r, column=1).value)
                if not m_key:
                    continue
                month_row_map[m_key] = r
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    continue
                month_rows.append(r)
            if month_rows:
                start_row = min(month_rows)
                end_row = max(month_rows)
                target_row = month_row_map.get(month_key)
                expected_dv = f"='经营指标'!$A${start_row}:$A${end_row}"
                def normalize_dv_formula(formula):
                    if formula is None:
                        return None
                    text = str(formula).strip()
                    if text.startswith("="):
                        text = text[1:]
                    text = text.replace("'", "")
                    text = re.sub(r"\s+", "", text)
                    return text
                expected_dv_norm = normalize_dv_formula(expected_dv)
                dv_ok = False
                for dv in dash_ws.data_validations.dataValidation:
                    if 'B3' in str(dv.sqref) and normalize_dv_formula(dv.formula1) == expected_dv_norm:
                        dv_ok = True
                        break
                if not dv_ok:
                    issues.append({"severity": "WARN", "sheet": "仪表盘", "message": f"B3下拉范围未匹配: {expected_dv}"})

                range_pattern = re.compile(r"'经营指标'!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)")
                bad_formula_cells = []
                bad_syntax_cells = []
                for row in dash_ws.iter_rows(min_row=1, max_row=dash_ws.max_row, min_col=1, max_col=dash_ws.max_column):
                    for cell in row:
                        val = cell.value
                        if not (isinstance(val, str) and val.startswith('=') and '经营指标' in val):
                            continue
                        if '=IF(=IFERROR(' in val or '/=IFERROR(' in val or '-=IFERROR(' in val:
                            bad_syntax_cells.append(cell.coordinate)
                        for m in range_pattern.finditer(val):
                            c1, r1, c2, r2 = m.groups()
                            if c1 != c2:
                                continue
                            if int(r1) != start_row or int(r2) != end_row:
                                bad_formula_cells.append(cell.coordinate)
                                break
                if bad_formula_cells:
                    issues.append({
                        "severity": "ERROR",
                        "sheet": "仪表盘",
                        "message": f"经营指标引用范围异常: {', '.join(sorted(set(bad_formula_cells))[:10])}"
                    })
                if bad_syntax_cells:
                    issues.append({
                        "severity": "WARN",
                        "sheet": "仪表盘",
                        "message": f"可能存在无效公式: {', '.join(sorted(set(bad_syntax_cells))[:10])}"
                    })

                # 目标预算区环比公式应使用“上月日期匹配”，不能依赖 MATCH(...)-1。
                old_mom_cells = []
                target_table_header = self._find_header_row_by_keyword(dash_ws, "指标", max_row=80)
                if target_table_header:
                    mom_rows = range(target_table_header + 1, min(dash_ws.max_row, target_table_header + 20) + 1)
                else:
                    mom_rows = range(35, min(dash_ws.max_row, 44) + 1)
                for r in mom_rows:
                    addr = f"F{r}"
                    val = dash_ws[addr].value
                    if not (isinstance(val, str) and val.startswith("=")):
                        continue
                    if "MATCH($B$3" in val and "-1" in val:
                        old_mom_cells.append(addr)
                if old_mom_cells:
                    issues.append({
                        "severity": "ERROR",
                        "sheet": "仪表盘",
                        "message": f"目标预算区环比公式仍为旧逻辑: {', '.join(old_mom_cells)}"
                    })

                # 目标月份图表点有效性校验：避免图表范围正确但最后一个月值为空。
                if target_row and (start_row <= target_row <= end_row):
                    series_pattern = re.compile(r"'经营指标'!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)")
                    missing_points = []
                    for chart_idx, chart in enumerate(dash_ws._charts, start=1):
                        for series_idx, series in enumerate(chart.series, start=1):
                            ref = None
                            if series.val and series.val.numRef and series.val.numRef.f:
                                ref = str(series.val.numRef.f)
                            if not ref or "经营指标" not in ref:
                                continue
                            m = series_pattern.search(ref)
                            if not m:
                                continue
                            c1, r1, c2, r2 = m.groups()
                            if c1 != c2:
                                continue
                            if not (int(r1) <= target_row <= int(r2)):
                                continue
                            col_idx = column_index_from_string(c1)
                            point_val = self._to_float(metric_ws.cell(row=target_row, column=col_idx).value)
                            if point_val is None:
                                missing_points.append(f"图{chart_idx}-序列{series_idx}({c1}{target_row})")
                    if missing_points:
                        issues.append({
                            "severity": "ERROR",
                            "sheet": "仪表盘",
                            "message": f"目标月份图表点为空: {', '.join(missing_points[:8])}"
                        })

        return issues

    def validate_report_file(self, report_path, target_year, target_month, year_scope="current"):
        """公开校验接口，供外部脚本调用。"""
        return self._validate_generated_report(report_path, target_year, target_month, year_scope)

    def _apply_output_sheet_filter(self, wb, output_sheets=None):
        """按输出表选择隐藏未勾选的 Sheet，返回被隐藏的 Sheet 名称。"""
        if output_sheets is None:
            return []

        selected = []
        seen = set()
        for name in output_sheets:
            text = str(name or "").strip()
            if not text or text in seen:
                continue
            selected.append(text)
            seen.add(text)
        if not selected:
            raise ValueError("至少需要选择一个输出表。")

        selected_set = set(selected)
        existing_selected = [name for name in wb.sheetnames if name in selected_set]
        if not existing_selected:
            raise ValueError("所选输出表在最终报告中不存在，请重新选择。")

        hidden = []
        for name in wb.sheetnames:
            ws = wb[name]
            if name in selected_set:
                ws.sheet_state = "visible"
            else:
                ws.sheet_state = "hidden"
                hidden.append(name)
        if wb.active is None or wb.active.title not in selected_set:
            wb.active = wb[existing_selected[0]]

        missing = [name for name in selected if name not in existing_selected]
        if missing:
            print("以下勾选的输出表未在最终报告中生成，已忽略: " + ", ".join(missing))
        if hidden:
            print(f"已按输出表设置显示 {len(existing_selected)} 个 Sheet，隐藏 {len(hidden)} 个 Sheet。")
        return hidden

    def generate_report(
        self,
        template_path,
        output_path,
        target_year=None,
        target_month=None,
        year_scope="current",
        replenishment_params=None,
        cashflow_params=None,
        include_ai_placeholders=False,
        fail_on_validation_error=True,
        fail_on_data_quality_error=False,
        allow_generated_report_template=False,
        output_sheets=None,
    ):
        print(f"\n正在根据模板 {template_path} 生成报告...")
        try:
            self.year_scope = self._normalize_year_scope(year_scope or "current")
            if replenishment_params:
                self.report_params.setdefault("replenishment", {}).update(replenishment_params)
            if cashflow_params:
                self.report_params.setdefault("cashflow", {}).update(cashflow_params)
            self.report_params.setdefault("ai", {})["include_placeholders"] = bool(include_ai_placeholders)
            self._log_audit(f"开始生成报告: {output_path}")
            wb = openpyxl.load_workbook(template_path)

            has_risk, risk_reasons = self._detect_template_risk(wb, template_path)
            if has_risk and not allow_generated_report_template:
                reason_text = "；".join(risk_reasons)
                raise ValueError(f"模板疑似为已生成报告文件，不建议直接复用。{reason_text}")

            if fail_on_data_quality_error:
                dq = self.get_data_quality_summary()
                if dq.get("ERROR", 0) > 0:
                    raise ValueError(
                        f"数据质量检查存在 {dq.get('ERROR', 0)} 条 ERROR，已阻止生成（可关闭 fail_on_data_quality_error）。"
                    )

            changed_sheets = set()
            
            # --- 1. Global Metadata Update (Update all Titles) ---
            if target_year and target_month:
                self._update_all_sheet_titles(wb, target_year, target_month)
                if '仪表盘' in wb.sheetnames:
                    self._update_dashboard_date(wb['仪表盘'], target_year, target_month)
                changed_sheets = self._ensure_report_month_structure(wb, target_year, target_month)
                for name in changed_sheets:
                    if name in wb.sheetnames:
                        wb[name]._charts = []

            # --- 2. Fill Statements ---
            if '利润表' in wb.sheetnames:
                self._fill_profit_sheet(wb['利润表'], target_year, target_month, self.year_scope)
            if '按产品汇总(含合计数)' in wb.sheetnames:
                self._fill_product_summary(
                    wb['按产品汇总(含合计数)'], target_year, target_month, self.year_scope
                )
            if '按产品汇总_含合计' in wb.sheetnames:
                self._fill_product_summary_total(
                    wb['按产品汇总_含合计'], target_year, target_month, self.year_scope
                )
            if '费用明细' in wb.sheetnames and target_year and target_month:
                self._fill_expense_details(
                    wb['费用明细'],
                    target_year,
                    target_month,
                    self.year_scope,
                    wb=wb,
                )
            if '利润表' in wb.sheetnames and '费用明细' in wb.sheetnames and target_year and target_month:
                self._highlight_profit_expense_anomalies(
                    wb,
                    target_year,
                    target_month,
                    self.year_scope,
                )

            # --- 3. Sync derived sheets & charts ---
            metrics_by_month_all = self._build_monthly_metrics(target_year, target_month, "all")
            metrics_by_month_scoped = self._build_monthly_metrics(target_year, target_month, self.year_scope)
            if '经营指标' in wb.sheetnames:
                self._update_management_metrics_sheet(
                    wb['经营指标'],
                    metrics_by_month_scoped,
                    target_year,
                    target_month,
                    self.year_scope,
                    metrics_by_month_all=metrics_by_month_all,
                )
            if '费用对比' in wb.sheetnames:
                self._update_expense_compare_sheet(
                    wb['费用对比'], metrics_by_month_scoped, target_year, target_month, self.year_scope
                )
            if '同比经营分析' in wb.sheetnames:
                self._update_compare_sheet(
                    wb['同比经营分析'], metrics_by_month_all, 'yoy', target_year, target_month, self.year_scope
                )
            if '环比经营分析' in wb.sheetnames:
                self._update_compare_sheet(
                    wb['环比经营分析'], metrics_by_month_scoped, 'mom', target_year, target_month, self.year_scope
                )
            if '按品类汇总(按月)' in wb.sheetnames:
                self._update_category_month_sheet(
                    wb['按品类汇总(按月)'], target_year, target_month, self.year_scope
                )
            if '明细_销售与库存' in wb.sheetnames:
                self._update_sales_inventory_detail_sheet(wb['明细_销售与库存'], target_year, target_month)
            if '费用明细环比分析' in wb.sheetnames:
                self._update_expense_mom_sheet(
                    wb['费用明细环比分析'], target_year, target_month, self.year_scope
                )
            if '本量利分析' in wb.sheetnames:
                self._update_cvp_sheet(
                    wb['本量利分析'], metrics_by_month_scoped, target_year, target_month, self.year_scope
                )
            if '资产负债表' in wb.sheetnames:
                self._update_balance_sheet(wb['资产负债表'], target_year, target_month, self.year_scope)
            if '目标_预算' in wb.sheetnames:
                self._update_budget_sheet(
                    wb['目标_预算'], metrics_by_month_scoped, target_year, target_month, self.year_scope
                )
            if '图表数据源_隐藏' in wb.sheetnames:
                self._update_hidden_chart_data(
                    wb['图表数据源_隐藏'], target_year, target_month, self.year_scope
                )
            if '产品对比(动态图表)' in wb.sheetnames:
                self._update_product_compare_sheet(wb['产品对比(动态图表)'], target_year, target_month)
            if '应收账款账龄分析' in wb.sheetnames:
                self._update_ar_aging_sheet(wb['应收账款账龄分析'], target_year, target_month)
            self._update_yoy_multi_year_sheet(wb, metrics_by_month_all, target_year, target_month, self.year_scope)
            self._generate_extended_reports(
                wb,
                metrics_by_month_scoped,
                target_year,
                target_month,
                metrics_by_month_all=metrics_by_month_all,
            )
            self._ensure_template_charts(wb, template_path)
            self._ensure_report_charts(wb)
            self._repair_missing_chart_categories(wb)
            self._audit_chart_counts(wb)
            self._update_chart_titles(wb, target_year, target_month)
            self._update_dashboard_controls(wb, target_year, target_month, self.year_scope)
            self._rewrite_chart_month_categories_for_history(wb, self.year_scope)
            self._trim_chart_data_ranges(wb)
            self._ensure_chart_month_axis_order(wb)
            self._append_chart_notes_below(wb, target_year, target_month)
            self._update_manager_guide_sheet(wb, target_year, target_month)
            self._normalize_internal_hyperlinks(wb)
            if not self._should_include_ai_placeholders():
                self._remove_ai_placeholder_texts(wb)
            self._apply_current_scope_visibility(wb, target_year, target_month, self.year_scope)

            wb.save(output_path)
            validation_issues = self._validate_generated_report(
                output_path,
                target_year,
                target_month,
                self.year_scope,
            )
            if validation_issues:
                print(f"自动校验发现 {len(validation_issues)} 条问题：")
                for item in validation_issues[:20]:
                    print(f"[{item.get('severity', 'WARN')}] {item.get('sheet', '未知')}: {item.get('message', '')}")
                error_count = sum(
                    1 for item in validation_issues
                    if str(item.get("severity") or "WARN").upper() == "ERROR"
                )
                warn_count = len(validation_issues) - error_count
                self._log_audit(
                    f"自动校验发现 {len(validation_issues)} 条问题 (ERROR={error_count}, WARN={warn_count})"
                )
                if error_count > 0 and fail_on_validation_error:
                    print("检测到自动校验 ERROR，已标记本次报告生成失败。")
                    self._log_audit("报告生成失败：自动校验存在 ERROR")
                    return False
            else:
                print("自动校验通过：关键指标与公式范围正常。")
                self._log_audit("自动校验通过")

            hidden_sheets = self._apply_output_sheet_filter(wb, output_sheets)
            if hidden_sheets:
                wb.save(output_path)
                self._log_audit(f"按输出表设置隐藏 {len(hidden_sheets)} 个 Sheet")

            print(f"报告生成成功: {output_path}")
            self._log_audit(f"报告生成成功: {output_path}")
            return True
        except Exception as e:
            print(f"生成报告失败: {e}")
            import traceback; traceback.print_exc()
            self._log_audit(f"报告生成失败: {e}")
            return False

    def generate_batch_reports(
        self,
        template_path,
        output_dir,
        target_year,
        months=None,
        year_scope="current",
        replenishment_params=None,
        cashflow_params=None,
        include_ai_placeholders=False,
        fail_on_validation_error=True,
        fail_on_data_quality_error=False,
        allow_generated_report_template=False,
        output_sheets=None,
    ):
        if months is None:
            months = [int(m[-2:]) for m in self.list_available_months() if m.startswith(f"{target_year}-")]
        os.makedirs(output_dir, exist_ok=True)
        summary = []
        for month in sorted(set(months)):
            output_path = os.path.join(output_dir, f"{target_year}年{int(month):02d}月_经营分析报告.xlsx")
            success = self.generate_report(
                template_path,
                output_path,
                str(target_year),
                str(month),
                year_scope=year_scope,
                replenishment_params=replenishment_params,
                cashflow_params=cashflow_params,
                include_ai_placeholders=include_ai_placeholders,
                fail_on_validation_error=fail_on_validation_error,
                fail_on_data_quality_error=fail_on_data_quality_error,
                allow_generated_report_template=allow_generated_report_template,
                output_sheets=output_sheets,
            )
            summary.append((month, output_path, "成功" if success else "失败"))
        summary_path = os.path.join(output_dir, f"批量生成摘要_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        with open(summary_path, "w", encoding="utf-8") as f:
            for month, path, status in summary:
                f.write(f"{target_year}-{int(month):02d}\t{status}\t{path}\n")
        print(f"批量生成完成，摘要文件: {summary_path}")
        return summary

    def generate_continuous_batch_reports(
        self,
        template_path,
        output_dir,
        start_year,
        start_month,
        end_year,
        end_month,
        year_scope="current",
        replenishment_params=None,
        cashflow_params=None,
        include_ai_placeholders=False,
        fail_on_validation_error=True,
        fail_on_data_quality_error=False,
        allow_generated_report_template=False,
        output_sheets=None,
    ):
        """按连续月份区间批量生成经营分析报告。"""
        month_keys = self.build_month_range(start_year, start_month, end_year, end_month)
        os.makedirs(output_dir, exist_ok=True)
        summary = []
        for month_key in month_keys:
            year, month = month_key.split("-")
            output_path = os.path.join(output_dir, f"{year}年{int(month):02d}月_经营分析报告.xlsx")
            success = self.generate_report(
                template_path,
                output_path,
                year,
                month,
                year_scope=year_scope,
                replenishment_params=replenishment_params,
                cashflow_params=cashflow_params,
                include_ai_placeholders=include_ai_placeholders,
                fail_on_validation_error=fail_on_validation_error,
                fail_on_data_quality_error=fail_on_data_quality_error,
                allow_generated_report_template=allow_generated_report_template,
                output_sheets=output_sheets,
            )
            summary.append((month_key, output_path, "成功" if success else "失败"))

        summary_path = os.path.join(output_dir, f"连续批量生成摘要_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        with open(summary_path, "w", encoding="utf-8") as f:
            for month_key, path, status in summary:
                f.write(f"{month_key}\t{status}\t{path}\n")
        print(f"连续批量生成完成，摘要文件: {summary_path}")
        return summary

    def _update_all_sheet_titles(self, wb, year, month):
        """搜索并替换所有 Sheet 中的年份/月份文本"""
        month_num = int(month)
        new_full_str = f"{year}年{month_num}月"

        def replace_month_text(text):
            if not isinstance(text, str):
                return text
            text = re.sub(r"(20\d{2})\s*年\s*\d{1,2}\s*月", new_full_str, text)
            text = re.sub(
                r"(?<!\d)(\d{1,2})\s*(月度|月份|月)(?!\d)",
                lambda m: f"{month_num}{m.group(2)}",
                text,
            )
            return text
        
        for ws in wb.worksheets:
            # Only scan first 5 rows for titles
            for r in range(1, 6):
                for c in range(1, 15): # Scan first 15 columns
                    cell = ws.cell(row=r, column=c)
                    new_val = replace_month_text(cell.value)
                    if new_val != cell.value:
                        cell.value = new_val

    def _update_dashboard_date(self, ws, year, month):
        date_str = f"{year}/{int(month):02d}"
        cell = ws['B2']
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = date_str
                print(f"仪表盘日期已更新为: {date_str} (in merged cell)")
                return
        cell.value = date_str
        print(f"仪表盘日期已更新为: {date_str}")

    def _update_dashboard_controls(self, wb, target_year, target_month, year_scope=None):
        if '仪表盘' not in wb.sheetnames or '经营指标' not in wb.sheetnames:
            return
        if not target_year or not target_month:
            return

        dash_ws = wb['仪表盘']
        metric_ws = wb['经营指标']
        month_rows = []
        for r in range(2, metric_ws.max_row + 1):
            m_key = self._label_to_month_key(metric_ws.cell(row=r, column=1).value)
            if not m_key:
                continue
            if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                continue
            month_rows.append(r)
        if not month_rows:
            return
        start_row = min(month_rows)
        end_row = max(month_rows)
        month_formula = f"='经营指标'!$A${start_row}:$A${end_row}"

        selector_updated = False
        for dv in dash_ws.data_validations.dataValidation:
            # Only normalize selector bound to B3.
            if dv.type == "list" and "B3" in str(dv.sqref):
                dv.formula1 = month_formula
                selector_updated = True
        if not selector_updated:
            dv = DataValidation(type="list", formula1=month_formula, allow_blank=True)
            dash_ws.add_data_validation(dv)
            dv.add("B3")

        dash_ws['B3'].value = f"{target_year}/{int(target_month):02d}"

        range_pattern = re.compile(r"((?:'经营指标'|经营指标)!\$([A-Z]+)\$)(\d+)(:\$([A-Z]+)\$)(\d+)")

        def sanitize_formula(formula):
            if not isinstance(formula, str) or not formula.startswith('='):
                return formula
            fixed = formula
            fixed = fixed.replace("=IF(=IFERROR(", "=IF(IFERROR(")
            fixed = fixed.replace("/=IFERROR(", "/IFERROR(")
            fixed = fixed.replace("-=IFERROR(", "-IFERROR(")
            return fixed

        def update_metric_range(formula):
            if not formula or "经营指标" not in formula:
                return formula

            def repl(match):
                col1 = match.group(2)
                col2 = match.group(5)
                if col1 != col2:
                    return match.group(0)
                return f"经营指标!${col1}${start_row}{match.group(4)}{end_row}"

            return range_pattern.sub(repl, str(formula))

        for chart in dash_ws._charts:
            for series in chart.series:
                if series.val and series.val.numRef and series.val.numRef.f:
                    series.val.numRef.f = update_metric_range(series.val.numRef.f)
                if series.cat and series.cat.strRef and series.cat.strRef.f:
                    series.cat.strRef.f = update_metric_range(series.cat.strRef.f)
                if series.cat and series.cat.numRef and series.cat.numRef.f:
                    series.cat.numRef.f = update_metric_range(series.cat.numRef.f)

        # 仪表盘中的INDEX/MATCH公式也需要扩展到最新月份行
        for row in dash_ws.iter_rows(min_row=1, max_row=dash_ws.max_row, min_col=1, max_col=dash_ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('=') and '经营指标' in cell.value:
                    cell.value = update_metric_range(sanitize_formula(cell.value))

        self._ensure_dashboard_financial_expense_rate(wb)
        self._repair_dashboard_target_table_formulas(
            dash_ws,
            budget_header_row=self._dashboard_budget_header_row(wb),
        )

    def _fill_profit_sheet(self, ws, target_year, target_month, year_scope=None):
        month_col_map = {} 
        header_row = None
        for r in range(1, 6):
            row_vals = [c.value for c in ws[r]]
            if any(isinstance(v, datetime) or (isinstance(v, str) and '202' in v) for v in row_vals):
                header_row = r
                for idx, cell in enumerate(ws[r], 1):
                    val = cell.value
                    month_key = self._extract_month_key_from_text(val)
                    if month_key:
                        month_col_map[month_key] = idx
                break
        
        if not month_col_map: return

        # IMPORTANT: Clear data outside the selected scope
        if target_year and target_month:
            for m_str, col_idx in month_col_map.items():
                if not self._month_key_in_scope(m_str, target_year, target_month, year_scope):
                    for r in range(header_row + 1, ws.max_row + 1):
                        cell = ws.cell(row=r, column=col_idx)
                        if isinstance(cell, MergedCell):
                            continue
                        cell.value = None

        month_keys = self._filter_month_keys(self.data['profit'].keys(), target_year, target_month, year_scope)
        # Clear in-scope month columns before refilling to avoid stale values.
        for m_str in month_keys:
            col_idx = month_col_map.get(m_str)
            if not col_idx:
                continue
            for r in range((header_row or 1) + 1, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        def build_indicator_maps():
            indicator_row_map = {}
            indicator_row_map_norm = {}
            indicator_row_map_norm_all = {}
            for row_idx in range((header_row or 1) + 1, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=1).value
                if cell_val is None:
                    continue
                raw = str(cell_val).strip().replace(" ", "").replace("\u3000", "")
                if not raw:
                    continue
                indicator_row_map[raw] = row_idx
                norm = self._normalize_profit_label(raw)
                if norm:
                    indicator_row_map_norm_all.setdefault(norm, []).append(row_idx)
                    if norm not in indicator_row_map_norm:
                        indicator_row_map_norm[norm] = row_idx
            return indicator_row_map, indicator_row_map_norm, indicator_row_map_norm_all

        indicator_row_map, indicator_row_map_norm, indicator_row_map_norm_all = build_indicator_maps()

        # 补齐模板中缺失的明细行（例如 管理费用-装卸费），避免“有源数据但无行可写”。
        missing_detail_rows = []
        seen_missing = set()
        for month_key in month_keys:
            df = self.data['profit'].get(month_key)
            if df is None or df.empty:
                continue
            for _, src_row in df.iterrows():
                raw_label = src_row.iloc[0] if len(src_row) > 0 else None
                if pd.isna(raw_label):
                    continue
                raw_name = str(raw_label).strip().replace(" ", "").replace("\u3000", "")
                if not raw_name or raw_name.lower() == "nan":
                    continue
                if '-' not in raw_name:
                    continue
                if re.match(r'^20\d{2}[/-]\d{1,2}[/-]\d{1,2}', raw_name):
                    continue
                norm_name = self._normalize_profit_label(raw_name)
                if raw_name in indicator_row_map:
                    continue
                if norm_name and norm_name in indicator_row_map_norm_all:
                    continue
                if raw_name in seen_missing:
                    continue
                seen_missing.add(raw_name)
                missing_detail_rows.append(raw_name)

        if missing_detail_rows:
            for raw_name in missing_detail_rows:
                prefix = raw_name.split('-', 1)[0] + '-'
                insert_row = None
                style_row = None

                # 优先插入到同类费用段末尾，保持表结构可读。
                for r in range((header_row or 1) + 1, ws.max_row + 1):
                    val = ws.cell(row=r, column=1).value
                    label = str(val).strip().replace(" ", "").replace("\u3000", "") if val is not None else ""
                    if label.startswith(prefix):
                        style_row = r
                        insert_row = r + 1

                if insert_row is None:
                    for r in range((header_row or 1) + 1, ws.max_row + 1):
                        val = ws.cell(row=r, column=1).value
                        if isinstance(val, str) and val.strip() == "AI分析":
                            insert_row = r
                            break
                if insert_row is None:
                    for r in range((header_row or 1) + 1, ws.max_row + 1):
                        val = ws.cell(row=r, column=1).value
                        if val is None or str(val).strip() == "":
                            insert_row = r
                            break
                if insert_row is None:
                    insert_row = ws.max_row + 1
                if style_row is None:
                    style_row = insert_row - 1 if insert_row > (header_row or 1) + 1 else (header_row or 1) + 1

                self._insert_rows_preserve_merges(ws, insert_row)
                self._copy_row_style(ws, style_row, insert_row)
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=insert_row, column=c)
                    if isinstance(cell, MergedCell):
                        continue
                    cell.value = None
                ws.cell(row=insert_row, column=1).value = raw_name

            indicator_row_map, indicator_row_map_norm, indicator_row_map_norm_all = build_indicator_maps()

        alias_groups = {
            "营业收入": ["主营业务收入", "营业收入"],
            "主营业务收入": ["主营业务收入", "营业收入"],
            "营业成本": ["主营业务成本", "营业成本"],
            "主营业务成本": ["主营业务成本", "营业成本"],
            "营业利润": ["营业利润"],
            "净利润": ["净利润"],
            "利润总额": ["利润总额"],
        }

        def can_fuzzy_match(source_name, target_name):
            # 仅允许同层级名称做兜底匹配，避免“销售费用-xx”覆盖“销售费用”
            if not source_name or not target_name:
                return False
            if '-' in source_name or '-' in target_name:
                return False
            if len(source_name) < 3 or len(target_name) < 3:
                return False
            return (source_name in target_name) or (target_name in source_name)

        for month_key in month_keys:
            df = self.data['profit'].get(month_key)
            if df is None:
                continue
            target_col = month_col_map.get(month_key)
            if not target_col: continue
            exact_written_rows = set()
            for _, row in df.iterrows():
                raw_label = row.iloc[0] if len(row) > 0 else None
                if pd.isna(raw_label):
                    continue
                raw_name = str(raw_label).strip().replace(" ", "").replace("\u3000", "")
                if not raw_name or raw_name.lower() == "nan":
                    continue
                name_clean = re.sub(r'^[一二三四五六七八九十0-9]+[、\.]', '', raw_name)
                norm_name = self._normalize_profit_label(raw_name)
                lookup_names = alias_groups.get(norm_name, [norm_name] if norm_name else [])

                target_rows = []
                match_mode = None
                for lookup_name in lookup_names:
                    map_row = indicator_row_map_norm.get(lookup_name)
                    if map_row:
                        target_rows = [map_row]
                        match_mode = "exact"
                        break
                if not target_rows:
                    target_row = indicator_row_map.get(raw_name)
                    if target_row:
                        target_rows = [target_row]
                        match_mode = "exact"
                if not target_rows and name_clean:
                    for map_name, map_row in indicator_row_map_norm.items():
                        if can_fuzzy_match(name_clean, map_name):
                            target_rows = [map_row]
                            match_mode = "fuzzy"
                            break
                if target_rows:
                    value = self._row_value_for_month(row, df.columns[1:], month_key)
                    if value is not None:
                        for target_row in target_rows:
                            if match_mode == "fuzzy" and target_row in exact_written_rows:
                                continue
                            target_cell = ws.cell(row=target_row, column=target_col)
                            if isinstance(target_cell, MergedCell):
                                continue
                            self._safe_set_cell_value(ws, target_row, target_col, value)
                            if match_mode == "exact":
                                exact_written_rows.add(target_row)
            self._fill_profit_category_rows(ws, target_col, month_key, header_row)
            self._sync_profit_duplicate_rows(ws, target_col)

        self._reposition_profit_rows(ws, header_row)
        self._refresh_profit_annual_totals(ws, header_row, month_col_map, target_year, target_month, year_scope)

    def _highlight_profit_expense_anomalies(self, wb, target_year, target_month, year_scope=None):
        if "利润表" not in wb.sheetnames or "费用明细" not in wb.sheetnames:
            return

        profit_ws = wb["利润表"]
        expense_ws = wb["费用明细"]
        target_month_key = f"{int(target_year):04d}-{int(target_month):02d}" if target_year and target_month else None

        month_col_map = {}
        header_row = None
        for r in range(1, 6):
            row_has_month = False
            for idx, cell in enumerate(profit_ws[r], 1):
                month_key = self._extract_month_key_from_text(cell.value)
                if month_key:
                    month_col_map[month_key] = idx
                    row_has_month = True
            if row_has_month:
                header_row = r
                break
        if not header_row or not month_col_map:
            return

        raw_df = self._get_expense_df()
        scoped_expense_df = self._prepare_expense_analysis_df(raw_df, target_year, target_month, year_scope)
        if scoped_expense_df is None or scoped_expense_df.empty:
            return

        flags = self._collect_expense_mom_flags(scoped_expense_df, target_year, target_month, year_scope)
        if not flags:
            return

        def _norm(v):
            if v is None:
                return ""
            return str(v).strip().replace(" ", "").replace("\u3000", "")

        def _parse_profit_expense_label(label):
            if label is None:
                return None, None
            text = _norm(label)
            if not text:
                return None, None
            normalized = self._normalize_profit_label(text)
            if not normalized:
                return None, None
            m = re.match(r"^(销售费用|管理费用|财务费用|研发费用)\s*[-—_/\\\\]\s*(.+)$", normalized)
            if m:
                return _norm(m.group(1)), _norm(m.group(2))
            if normalized in ("销售费用", "管理费用", "财务费用", "研发费用"):
                return _norm(normalized), None
            return None, None

        flag_exact = {}
        for f in flags:
            m_key = f.get("MonthStr")
            if not m_key or m_key not in month_col_map:
                continue
            if target_month_key and m_key != target_month_key:
                continue
            tags = set(f.get("ReasonTags") or [])
            # 利润表仅标记“环比异常”的当月费用项，避免低频/占比类告警带来噪声。
            if "环比异常" not in tags:
                continue
            cat = _norm(f.get("Category"))
            sub = _norm(f.get("Subcategory"))
            score = f.get("AnomalyScore") or 0
            if not cat or not sub:
                continue
            key_exact = (m_key, cat, sub)
            if key_exact not in flag_exact or score > (flag_exact[key_exact].get("AnomalyScore") or 0):
                flag_exact[key_exact] = f

        if not flag_exact:
            return

        header_map = self._get_header_map(expense_ws, 1)
        month_key_col = header_map.get("月份键") or 8
        month_label_col = header_map.get("月份") or 1
        category_col = header_map.get("费用类别") or 3
        subcategory_col = header_map.get("子科目") or 4

        detail_exact = {}
        detail_by_category = {}
        detail_by_month = {}
        for r in range(2, expense_ws.max_row + 1):
            m_key = expense_ws.cell(row=r, column=month_key_col).value
            if not m_key:
                m_key = self._extract_month_key_from_text(expense_ws.cell(row=r, column=month_label_col).value)
            if not m_key:
                continue
            m_key = str(m_key).strip().replace("/", "-")

            cat = _norm(expense_ws.cell(row=r, column=category_col).value)
            sub = _norm(expense_ws.cell(row=r, column=subcategory_col).value)
            if m_key not in detail_by_month:
                detail_by_month[m_key] = r
            if cat and (m_key, cat) not in detail_by_category:
                detail_by_category[(m_key, cat)] = r
            if cat and sub and (m_key, cat, sub) not in detail_exact:
                detail_exact[(m_key, cat, sub)] = r

        if not detail_by_month:
            return

        for r in range(header_row + 1, profit_ws.max_row + 1):
            cat, sub = _parse_profit_expense_label(profit_ws.cell(row=r, column=1).value)
            if not cat or not sub:
                continue

            for m_key, c in month_col_map.items():
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    continue
                if target_month_key and m_key != target_month_key:
                    continue

                flag = flag_exact.get((m_key, cat, sub))
                if flag is None:
                    continue

                target_row = None
                target_row = detail_exact.get((m_key, cat, sub))
                if target_row is None:
                    target_row = detail_by_category.get((m_key, cat))
                if target_row is None:
                    target_row = detail_by_month.get(m_key)
                if target_row is None:
                    continue

                cell = profit_ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
                    continue

                self._apply_hyperlink(cell, f"#'费用明细'!A{target_row}")
                font = copy.copy(cell.font) if cell.font else Font(name="微软雅黑", size=10)
                font.color = "00FF0000"
                font.underline = "single"
                cell.font = font

    def _refresh_profit_annual_totals(self, ws, header_row, month_col_map, target_year, target_month, year_scope=None):
        if not header_row or not month_col_map:
            return

        annual_cols = []
        annual_tokens = ("全年汇总", "全年合计", "本年累计", "年累计", "累计")
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if not isinstance(val, str):
                continue
            text = val.strip()
            if any(token in text for token in annual_tokens):
                annual_cols.append(col)
        if not annual_cols:
            return

        scoped_month_cols = []
        for m_key, col_idx in month_col_map.items():
            if self._month_key_in_scope(m_key, target_year, target_month, "current"):
                scoped_month_cols.append(col_idx)
        scoped_month_cols = sorted(set(scoped_month_cols))
        if not scoped_month_cols:
            return

        for r in range((header_row or 1) + 1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label is None or str(label).strip() == "":
                continue

            vals = []
            for col_idx in scoped_month_cols:
                v = self._to_float(ws.cell(row=r, column=col_idx).value)
                if v is not None:
                    vals.append(v)
            total_val = sum(vals) if vals else None

            for annual_col in annual_cols:
                self._safe_set_cell_value(ws, r, annual_col, total_val)

    def _reposition_profit_rows(self, ws, header_row):
        if not header_row:
            return False

        def clean_label(val):
            if val is None:
                return ""
            return str(val).strip().replace(" ", "").replace("\u3000", "")

        def find_row_exact(options):
            for r in range(header_row + 1, ws.max_row + 1):
                val = clean_label(ws.cell(row=r, column=1).value)
                if val in options:
                    return r
            return None

        def find_first_row(predicate):
            for r in range(header_row + 1, ws.max_row + 1):
                val = clean_label(ws.cell(row=r, column=1).value)
                if predicate(val):
                    return r
            return None

        def find_last_row(predicate):
            last = None
            for r in range(header_row + 1, ws.max_row + 1):
                val = clean_label(ws.cell(row=r, column=1).value)
                if predicate(val):
                    last = r
            return last

        moved = False

        # 1) "一、营业收入" -> before "主营业务收入"
        revenue_header_row = find_row_exact({"一、营业收入", "1、营业收入", "一.营业收入", "一．营业收入"})
        revenue_base_row = find_first_row(
            lambda v: (v in {"主营业务收入", "营业收入"}) and ("-" not in v)
        )
        if revenue_header_row and revenue_base_row:
            moved |= self._move_row_preserve(ws, revenue_header_row, revenue_base_row)

        # 2) "减：营业成本" -> before "主营业务成本"
        cost_header_row = find_row_exact({"减：营业成本", "减:营业成本", "减：主营业务成本", "减:主营业务成本"})
        cost_base_row = find_first_row(
            lambda v: (v in {"主营业务成本", "营业成本"}) and ("-" not in v)
        )
        if cost_header_row and cost_base_row:
            moved |= self._move_row_preserve(ws, cost_header_row, cost_base_row)

        # 3) "税金及附加"/"营业税金及附加" -> after成本段
        def is_cost_block(label):
            if not label:
                return False
            base = label
            if base.startswith("减：") or base.startswith("减:"):
                base = base[2:]
            return base.startswith("主营业务成本") or base.startswith("营业成本")

        cost_block_last = find_last_row(is_cost_block)
        if cost_block_last:
            insert_at = cost_block_last + 1
            for name in ["税金及附加", "营业税金及附加"]:
                row = find_row_exact({name})
                if row:
                    moved |= self._move_row_preserve(ws, row, insert_at)
                    insert_at += 1

        return moved

    def _fill_product_summary(self, ws, target_year, target_month, year_scope=None):
        col_map, header_row = {}, 1
        for idx, cell in enumerate(ws[header_row], 1):
            if cell.value:
                col_map[str(cell.value).strip().replace('/', '-')] = idx

        # Clear future months.
        if target_year and target_month:
            for full_key, col_idx in col_map.items():
                m_match = re.match(r'(20\d{2})\s*[./-]\s*(\d{1,2})\s*月?', full_key)
                if not m_match:
                    continue
                m_key = f"{m_match.group(1)}-{int(m_match.group(2)):02d}"
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    for r in range(header_row + 1, ws.max_row + 1):
                        self._safe_set_cell_value(ws, r, col_idx, None)

        prod_row_map = {}
        total_row = None
        for r in range(header_row + 1, ws.max_row + 1):
            p_name = ws.cell(row=r, column=1).value
            if not p_name:
                continue
            label = str(p_name).strip()
            if label == '合计' and total_row is None:
                total_row = r
                continue
            if label and label.lower() != 'nan' and label not in prod_row_map:
                prod_row_map[label] = r

        month_keys = self._filter_month_keys(self.data['sales'].keys(), target_year, target_month, year_scope)
        if not month_keys:
            return

        monthly_product_map = {}
        monthly_total_map = {}
        for month_key in month_keys:
            df = self.data['sales'].get(month_key)
            if df is None or df.empty:
                continue
            src = df.copy()
            if 'MonthStr' not in src.columns:
                src['MonthStr'] = month_key

            src['Revenue'] = self._extract_sales_revenue(src)
            src['Qty'] = pd.to_numeric(src.get('数量'), errors='coerce')
            src = self._attach_sales_cost(src, target_year, target_month, year_scope)

            name_col = None
            for candidate in ['品目名', '产品名称', '品名', '商品名称']:
                if candidate in src.columns:
                    name_col = candidate
                    break
            if not name_col:
                continue

            src['ProductName'] = src[name_col].astype(str).str.strip()
            src = src[(src['ProductName'] != '') & (src['ProductName'].str.lower() != 'nan')]
            if src.empty:
                continue

            grouped = (
                src.groupby(['MonthStr', 'ProductName'], dropna=False)
                .agg(
                    qty=('Qty', 'sum'),
                    revenue=('Revenue', 'sum'),
                    cost=('Cost', lambda s: s.sum(min_count=1)),
                    cost_non_na=('Cost', lambda s: int(pd.to_numeric(s, errors='coerce').notna().sum())),
                )
                .reset_index()
            )
            grouped.loc[grouped['cost_non_na'] == 0, 'cost'] = None
            grouped['profit'] = grouped.apply(
                lambda r: (r['revenue'] - r['cost'])
                if pd.notna(r.get('revenue')) and pd.notna(r.get('cost'))
                else None,
                axis=1
            )
            grouped['margin'] = grouped.apply(
                lambda r: (r['profit'] / r['revenue'])
                if pd.notna(r.get('profit')) and pd.notna(r.get('revenue')) and r['revenue'] != 0
                else None,
                axis=1
            )

            month_group = grouped[grouped['MonthStr'] == month_key]
            if month_group.empty:
                continue

            month_total = {
                'qty': month_group['qty'].sum(min_count=1),
                'revenue': month_group['revenue'].sum(min_count=1),
                'cost': month_group['cost'].sum(min_count=1),
            }
            if pd.notna(month_total['revenue']) and pd.notna(month_total['cost']):
                month_total['profit'] = month_total['revenue'] - month_total['cost']
            else:
                month_total['profit'] = None
            month_total['margin'] = (
                month_total['profit'] / month_total['revenue']
                if month_total['profit'] is not None and month_total['revenue'] not in (None, 0)
                else None
            )
            monthly_total_map[month_key] = month_total

            for _, rec in month_group.iterrows():
                monthly_product_map[(month_key, str(rec['ProductName']).strip())] = {
                    'qty': rec.get('qty'),
                    'revenue': rec.get('revenue'),
                    'cost': rec.get('cost'),
                    'profit': rec.get('profit'),
                    'margin': rec.get('margin'),
                }

        metric_map = {
            '销售收入': 'revenue',
            '销售数量': 'qty',
            '销售成本': 'cost',
            '销售利润': 'profit',
            '毛利率': 'margin',
        }

        def _normalize_output(v):
            if pd.isna(v):
                return None
            if hasattr(v, "item"):
                try:
                    return v.item()
                except Exception:
                    pass
            return v

        for month_key in month_keys:
            metric_cols = {}
            for metric_name, metric_key in metric_map.items():
                col_idx = col_map.get(f"{month_key}_{metric_name}")
                if col_idx:
                    metric_cols[metric_key] = col_idx
            if not metric_cols:
                continue

            for row_idx in prod_row_map.values():
                for col_idx in metric_cols.values():
                    self._safe_set_cell_value(ws, row_idx, col_idx, None)

            for product_name, row_idx in prod_row_map.items():
                metrics = monthly_product_map.get((month_key, product_name))
                if not metrics:
                    continue
                for metric_key, col_idx in metric_cols.items():
                    self._safe_set_cell_value(ws, row_idx, col_idx, _normalize_output(metrics.get(metric_key)))

            if total_row is not None:
                total_metrics = monthly_total_map.get(month_key, {})
                for metric_key, col_idx in metric_cols.items():
                    self._safe_set_cell_value(ws, total_row, col_idx, _normalize_output(total_metrics.get(metric_key)))

    def _collect_expense_mom_flags(
        self,
        df,
        target_year,
        target_month,
        year_scope=None,
        rate_threshold=0.5,
        delta_threshold=10000,
        yoy_rate_threshold=0.35,
        zscore_threshold=3.0,
        share_jump_threshold=0.08,
        rare_month_threshold=2,
        rare_amount_threshold=5000,
    ):
        scoped = self._prepare_expense_analysis_df(df, target_year, target_month, year_scope)
        if scoped is None or scoped.empty:
            return []

        scoped = scoped[scoped['Category'].notna() & scoped['Subcategory'].notna()]
        if scoped.empty:
            return []

        summary = (
            scoped.groupby(['Category', 'Subcategory', 'MonthStr'])['Amount']
            .sum()
            .reset_index()
        )
        if summary.empty:
            return []

        month_abs_totals = (
            summary.groupby('MonthStr')['Amount']
            .apply(lambda s: s.abs().sum())
            .to_dict()
        )
        summary['MonthAbsTotal'] = summary['MonthStr'].map(month_abs_totals)
        summary['Share'] = summary.apply(
            lambda r: abs(r['Amount']) / r['MonthAbsTotal'] if r['MonthAbsTotal'] else None,
            axis=1,
        )

        def month_shift(month_key, offset):
            try:
                p = pd.Period(month_key, freq='M') + offset
                return f"{p.year}-{p.month:02d}"
            except Exception:
                return None

        flags = []
        for (cat, sub), group in summary.groupby(['Category', 'Subcategory']):
            group = group.sort_values(by='MonthStr')
            month_map = {row['MonthStr']: row['Amount'] for _, row in group.iterrows()}
            share_map = {row['MonthStr']: row['Share'] for _, row in group.iterrows()}
            for m_key in sorted(month_map.keys()):
                if not self._month_key_in_scope(m_key, target_year, target_month, year_scope):
                    continue

                prev_key = month_shift(m_key, -1)
                yoy_key = month_shift(m_key, -12)
                curr = month_map.get(m_key)
                prev = month_map.get(prev_key) if prev_key else None
                yoy = month_map.get(yoy_key) if yoy_key else None
                if curr is None:
                    continue

                delta = (curr - prev) if prev is not None else None
                rate = (delta / prev) if prev not in (None, 0) else None

                yoy_delta = (curr - yoy) if yoy is not None else None
                yoy_rate = (yoy_delta / yoy) if yoy not in (None, 0) else None

                hist_keys = [k for k in sorted(month_map.keys()) if k < m_key][-12:]
                hist_values = [month_map.get(k) for k in hist_keys if month_map.get(k) is not None]
                robust_z = None
                if len(hist_values) >= 4:
                    hist_series = pd.Series(hist_values, dtype='float64')
                    median = hist_series.median()
                    mad = (hist_series - median).abs().median()
                    if mad and mad > 1e-9:
                        robust_z = 0.6745 * (curr - median) / mad

                share_curr = share_map.get(m_key)
                share_hist_start = month_shift(m_key, -6)
                share_hist = [
                    share_map.get(k)
                    for k in sorted(share_map.keys())
                    if k < m_key and (share_hist_start is None or k >= share_hist_start) and share_map.get(k) is not None
                ]
                share_median = pd.Series(share_hist, dtype='float64').median() if len(share_hist) >= 3 else None
                share_delta = (share_curr - share_median) if (share_curr is not None and share_median is not None) else None

                active_window_start = month_shift(m_key, -11)
                active_months_12 = 0
                for k in month_map.keys():
                    if k > m_key:
                        continue
                    if active_window_start and k < active_window_start:
                        continue
                    if abs(month_map.get(k) or 0) >= 100:
                        active_months_12 += 1

                mom_alert = rate is not None and abs(rate) > rate_threshold and abs(delta or 0) > delta_threshold
                yoy_alert = yoy_rate is not None and abs(yoy_rate) > yoy_rate_threshold and abs(yoy_delta or 0) > delta_threshold
                z_alert = robust_z is not None and abs(robust_z) >= zscore_threshold and abs(curr) >= rare_amount_threshold
                share_alert = share_delta is not None and share_delta >= share_jump_threshold and abs(curr) >= rare_amount_threshold
                rare_alert = active_months_12 <= rare_month_threshold and abs(curr) >= rare_amount_threshold
                new_item_alert = prev in (None, 0) and active_months_12 <= 1 and abs(curr) >= max(delta_threshold, rare_amount_threshold * 2)

                reason_tags = []
                reason_texts = []
                if mom_alert:
                    reason_tags.append("环比异常")
                    reason_texts.append(f"环比 {rate:+.1%}，增量 {delta:+,.0f}")
                if yoy_alert:
                    reason_tags.append("同比异常")
                    reason_texts.append(f"同比 {yoy_rate:+.1%}，增量 {yoy_delta:+,.0f}")
                if z_alert:
                    reason_tags.append("历史偏离")
                    reason_texts.append(f"稳健Z值 {robust_z:+.2f}")
                if share_alert:
                    reason_tags.append("占比跃升")
                    reason_texts.append(f"占比较历史中位 +{share_delta:.1%}")
                if rare_alert:
                    reason_tags.append("低频项目")
                    reason_texts.append(f"近12月出现 {active_months_12} 次")
                if new_item_alert:
                    reason_tags.append("新增大额")
                    reason_texts.append("上月无发生，本月首次大额出现")

                if not reason_tags:
                    continue

                score = 0.0
                if mom_alert:
                    score += 35
                if yoy_alert:
                    score += 20
                if z_alert:
                    score += 18
                if share_alert:
                    score += 15
                if rare_alert:
                    score += 12
                if new_item_alert:
                    score += 20
                if rate is not None:
                    score += min(8, abs(rate) * 5)
                if yoy_rate is not None:
                    score += min(6, abs(yoy_rate) * 3)
                if robust_z is not None:
                    score += min(8, abs(robust_z))
                if share_delta is not None and share_delta > 0:
                    score += min(6, share_delta * 40)
                score = round(min(100, score), 1)

                if score < 30:
                    continue

                flags.append({
                    "MonthStr": m_key,
                    "Category": cat,
                    "Subcategory": sub,
                    "Delta": delta,
                    "Rate": rate,
                    "Amount": curr,
                    "PrevAmount": prev,
                    "YoYAmount": yoy,
                    "YoYDelta": yoy_delta,
                    "YoYRate": yoy_rate,
                    "RobustZ": robust_z,
                    "Share": share_curr,
                    "ShareMedian": share_median,
                    "ShareDelta": share_delta,
                    "ActiveMonths12": active_months_12,
                    "AnomalyScore": score,
                    "ReasonTags": reason_tags,
                    "ReasonText": "；".join(reason_texts),
                    "AnomalyKey": f"{m_key}|{cat}|{sub}",
                })
        return flags

    def _fill_expense_details(self, ws, target_year, target_month, year_scope=None, wb=None):
        self.expense_detail_key_row_map = {}
        wb_obj = wb if wb is not None else ws.parent

        main_headers = ["月份", "部门", "费用类别", "子科目", "摘要", "金额", "异常标签", "月份键"]
        detail_headers = ["月份", "部门", "费用类别", "子科目", "摘要", "金额", "环比增量", "环比增速", "异常评分", "异常标签", "明细键"]

        raw_df = self._get_expense_df()
        df = self._prepare_expense_analysis_df(raw_df, target_year, target_month, year_scope)

        # 先清理主表历史残留，确保只保留一段主表头。
        start_row = 2
        clear_end = max(ws.max_row, start_row + (len(df) if df is not None else 0) + 300)
        for r in range(start_row, clear_end + 1):
            for c in range(1, 9):
                self._safe_set_cell_value(ws, r, c, None)
        max_clear_col = max(ws.max_column, 9)
        for r in range(1, min(clear_end, ws.max_row) + 1):
            for c in range(9, max_clear_col + 1):
                self._safe_set_cell_value(ws, r, c, None)

        for idx, h in enumerate(main_headers, start=1):
            self._safe_set_cell_value(ws, 1, idx, h)

        if df is None or df.empty:
            ws.auto_filter.ref = "A1:H1"
            if wb_obj is not None:
                detail_ws = self._prepare_sheet(wb_obj, self.expense_detail_sheet_name, insert_after="费用明细")
                detail_ws.cell(row=1, column=1).value = "异常项目明细（按明细键）"
                detail_ws.cell(row=1, column=1).font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
                self._write_table(detail_ws, 2, 1, detail_headers, [])
                detail_ws.cell(row=3, column=1).value = "无可关联的异常项目明细"
            return

        df = df.sort_values(by=['MonthStr', 'AmountAbs'], ascending=[False, False])

        flags = self._collect_expense_mom_flags(df, target_year, target_month, year_scope)
        flags = sorted(
            flags,
            key=lambda x: (
                x.get("AnomalyScore") or 0,
                abs(x.get("Delta") or 0),
                abs(x.get("Amount") or 0),
            ),
            reverse=True,
        )
        selected_by_pair = self._select_expense_display_flags(flags, target_year, target_month)
        selected_flags = sorted(
            list(selected_by_pair.values()),
            key=lambda x: (
                x.get("AnomalyScore") or 0,
                abs(x.get("Delta") or 0),
                abs(x.get("Amount") or 0),
            ),
            reverse=True,
        )
        flag_map = {
            (f['MonthStr'], f['Category'], f['Subcategory']): f
            for f in selected_flags
        }

        current_row = start_row
        for _, row in df.iterrows():
            key = (row.get('MonthStr'), row.get('Category'), row.get('Subcategory'))
            flag = flag_map.get(key)
            tags = ""
            if flag:
                tags = "、".join(flag.get("ReasonTags") or [])
            summary_val = row.get('Summary')
            if pd.isna(summary_val):
                summary_val = None
            self._safe_set_cell_value(ws, current_row, 1, row.get('MonthLabel'))
            self._safe_set_cell_value(ws, current_row, 2, row.get('Department'))
            self._safe_set_cell_value(ws, current_row, 3, row.get('Category'))
            self._safe_set_cell_value(ws, current_row, 4, row.get('Subcategory'))
            self._safe_set_cell_value(ws, current_row, 5, summary_val)
            self._safe_set_cell_value(ws, current_row, 6, row.get('Amount'))
            self._safe_set_cell_value(ws, current_row, 7, tags if tags else "normal")
            self._safe_set_cell_value(ws, current_row, 8, row.get('MonthStr'))
            current_row += 1

        last_row = max(start_row, current_row - 1)
        ws.auto_filter.ref = f"A1:H{last_row}"

        detail_rows = []
        top_lines_per_flag = 5
        max_detail_rows = 500
        grouped_details = {}
        for key, group in df.groupby(['MonthStr', 'Category', 'Subcategory']):
            grouped_details[key] = group.sort_values(by='AmountAbs', ascending=False)

        for f in selected_flags:
            key = (f.get('MonthStr'), f.get('Category'), f.get('Subcategory'))
            detail_df = grouped_details.get(key)
            if detail_df is None or detail_df.empty:
                continue
            picks = detail_df.head(top_lines_per_flag)
            for _, row in picks.iterrows():
                detail_rows.append([
                    row.get('MonthLabel'),
                    row.get('Department'),
                    row.get('Category'),
                    row.get('Subcategory'),
                    row.get('Summary'),
                    row.get('Amount'),
                    f.get('Delta'),
                    f.get('Rate'),
                    f.get('AnomalyScore'),
                    "、".join(f.get('ReasonTags') or []),
                    f.get('AnomalyKey'),
                ])
                if len(detail_rows) >= max_detail_rows:
                    break
            if len(detail_rows) >= max_detail_rows:
                break

        if wb_obj is None:
            return

        detail_ws = self._prepare_sheet(wb_obj, self.expense_detail_sheet_name, insert_after="费用明细")
        detail_ws.cell(row=1, column=1).value = f"{target_year}/{int(target_month):02d} 异常项目明细（按明细键）"
        detail_ws.cell(row=1, column=1).font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

        self._write_table(detail_ws, 2, 1, detail_headers, detail_rows if detail_rows else [])
        if detail_rows:
            end_row = 2 + len(detail_rows)
            detail_ws.auto_filter.ref = f"A2:K{end_row}"
            for r_idx, row in enumerate(detail_rows, start=3):
                key_text = row[-1]
                if key_text and key_text not in self.expense_detail_key_row_map:
                    self.expense_detail_key_row_map[key_text] = r_idx
                key_cell = detail_ws.cell(row=r_idx, column=11)
                if key_text:
                    key_cell.font = Font(color="0563C1")
        else:
            detail_ws.cell(row=3, column=1).value = "无可关联的异常项目明细"

    def _update_staff_efficiency_sheet(self, wb, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "销售人效分析", insert_after="渠道贡献")
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无销售数据"]])
            return
        df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, year_scope)
        
        staff_col = None
        for col in ['职员', '职员(负责)名', '销售员']:
            if col in df.columns:
                staff_col = col
                break
        if not staff_col:
            self._write_table(ws, 1, 1, ["提示"], [["缺少职员字段"]])
            return

        df['Revenue'] = self._extract_sales_revenue(df)
        df['Qty'] = pd.to_numeric(df.get('数量'), errors='coerce')
        df = self._attach_sales_cost(df, target_year, target_month, year_scope)
        
        rows = []
        for staff, group in df.groupby(staff_col):
            if not staff or pd.isna(staff): continue
            metrics = self._calc_sales_metrics_from_group(group)
            revenue = metrics.get('revenue')
            profit = metrics.get('profit')
            margin = metrics.get('margin')
            months = group['MonthStr'].nunique()
            avg_rev = (revenue / months) if (revenue is not None and months) else revenue
            rows.append([staff, revenue, profit, margin, avg_rev])
            
        headers = ["职员", "总销售额", "总毛利", "毛利率", "月均销售额"]
        rows = sorted(rows, key=lambda x: x[1] if x[1] is not None else 0, reverse=True)
        self._write_table(ws, 1, 1, headers, rows)
        
        top_rows = rows[:10]
        if top_rows:
            start_col = len(headers) + 2
            chart_rows = [[r[0], r[1]] for r in top_rows]
            self._write_table(ws, 1, start_col, ["职员", "销售额"], chart_rows)
            anchor = f"{get_column_letter(start_col + 3)}2"
            self._add_bar_chart_from_table(ws, 1, start_col, 1 + len(chart_rows), "职员销售额Top (人效)", anchor)
            self._add_detailed_analysis_box(ws, start_col + 3, 14, "人效分析", [
                "展示销售额最高的Top10员工。",
                "月均销售额反映了员工的持续产出能力。",
                "关注高销售额但低毛利的员工，可能存在为了冲量而牺牲利润的情况。"
            ])

    def _update_dupont_sheet(self, wb, metrics_by_month, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "杜邦分析", insert_after="经营指标")
        headers = ["月份", "净资产收益率(ROE)", "销售净利率", "资产周转率", "权益乘数"]
        rows = []
        month_keys = self._filter_month_keys(metrics_by_month.keys(), target_year, target_month, year_scope)
        
        for m_key in month_keys:
            data = metrics_by_month.get(m_key, {})
            revenue = data.get("revenue")
            net_profit = data.get("net_profit")
            assets = self._safe_sum([
                data.get("cash"), data.get("ar_balance"), data.get("inventory_end"), data.get("fixed_assets")
            ])
            liabilities = self._safe_sum([
                data.get("ap_balance"), data.get("short_debt"), data.get("long_debt")
            ])
            equity = (assets - liabilities) if (assets is not None and liabilities is not None) else None
            
            net_margin = net_profit / revenue if (net_profit is not None and revenue) else None
            asset_turnover = revenue / assets if (revenue and assets) else None
            equity_multiplier = assets / equity if (assets and equity) else None
            
            roe = None
            if net_margin and asset_turnover and equity_multiplier:
                roe = net_margin * asset_turnover * equity_multiplier
            elif net_profit is not None and equity:
                roe = net_profit / equity
                
            rows.append([self._month_key_to_label(m_key), roe, net_margin, asset_turnover, equity_multiplier])
            
        if not rows:
            rows = [[None]*len(headers)]
        self._write_table(ws, 1, 1, headers, rows)
        
        anchor = f"{get_column_letter(len(headers)+2)}2"
        self._add_line_chart_by_columns(ws, 1, [2, 3, 4, 5], 1, 2, 1 + len(rows), "杜邦分析趋势", anchor)
        self._add_detailed_analysis_box(ws, len(headers)+2, 14, "杜邦分析", [
            "ROE (净资产收益率) = 销售净利率 × 资产周转率 × 权益乘数。",
            "销售净利率：衡量产品盈利能力 (越高越好)。",
            "资产周转率：衡量资产利用效率 (越高越好)。",
            "权益乘数：衡量财务杠杆 (过高增加风险，过低影响回报)。",
            "观察ROE变化是由哪个因子驱动的。"
        ])

        self._reorder_month_rows_desc(ws)

    def _update_annual_expense_anomaly_sheet(self, wb, target_year, target_month, year_scope=None, top_n=30):
        ws = self._prepare_sheet(wb, "年度费用异常Top", insert_after="费用明细环比分析")
        if not target_year or not target_month:
            self._write_table(ws, 1, 1, ["提示"], [["缺少期间参数"]])
            return

        df = self._get_expense_df()
        if df is None or df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无费用数据"]])
            return

        df = df.copy()
        if 'MonthStr' not in df.columns:
            date_col = next((c for c in df.columns if '日期' in str(c) or 'Date' in str(c)), None)
            if date_col:
                df['ParsedDate'] = pd.to_datetime(df[date_col], errors='coerce')
                df['MonthStr'] = df['ParsedDate'].dt.strftime('%Y-%m')

        if 'MonthStr' not in df.columns:
            self._write_table(ws, 1, 1, ["提示"], [["费用数据缺少月份字段"]])
            return

        # Restrict to target year and <= target_month
        limit_key = f"{target_year}-{int(target_month):02d}"
        df = df[df['MonthStr'].astype(str).str.startswith(f"{target_year}-")]
        df = df[df['MonthStr'] <= limit_key]
        if df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["目标年度无费用数据"]])
            return

        flags = self._collect_expense_mom_flags(
            df,
            target_year,
            target_month,
            year_scope="current",
        )

        if not flags:
            self._write_table(ws, 1, 1, ["提示"], [["未发现符合条件的费用异常"]])
            return

        rows = []
        for f in flags:
            rows.append([
                self._month_key_to_label(f.get("MonthStr")),
                f.get("Category"),
                f.get("Subcategory"),
                f.get("Amount"),
                f.get("PrevAmount"),
                f.get("Delta"),
                f.get("Rate"),
                f.get("AnomalyScore"),
                "、".join(f.get("ReasonTags") or []),
                f.get("ActiveMonths12"),
            ])

        rows = sorted(
            rows,
            key=lambda x: (
                x[7] if x and x[7] is not None else 0,
                abs(x[5]) if x and x[5] is not None else 0,
            ),
            reverse=True,
        )[:top_n]
        headers = ["月份", "费用类别", "子科目", "本期金额", "上期金额", "环比增量", "环比增速", "异常评分", "异常标签", "近12月出现"]
        self._write_table(ws, 1, 1, headers, rows)
        self._write_chart_note(ws, 1, len(rows) + 3, "说明：综合环比/同比/历史偏离/占比跃升/低频项目识别后的费用异常Top。")

    def _update_customer_churn_sheet(self, wb, target_year, target_month, year_scope=None):
        ws = self._prepare_sheet(wb, "客户流失与留存", insert_after="客户贡献与回款")
        sales_df = self._get_sales_df()
        if sales_df is None or sales_df.empty:
            self._write_table(ws, 1, 1, ["提示"], [["无销售数据"]])
            return
            
        df = self._filter_df_by_scope(sales_df.copy(), target_year, target_month, "all") 
        if '往来单位名' not in df.columns:
            return
            
        monthly_custs = {}
        month_keys = sorted(df['MonthStr'].unique())
        for m in month_keys:
            monthly_custs[m] = set(df[df['MonthStr'] == m]['往来单位名'].dropna().unique())
            
        headers = ["月份", "活跃客户数", "新增客户", "流失客户", "回流客户", "净增"]
        rows = []
        
        target_keys = self._filter_month_keys(month_keys, target_year, target_month, year_scope)
        
        for i, m in enumerate(month_keys):
            if m not in target_keys: continue
            
            current = monthly_custs[m]
            if i == 0:
                rows.append([self._month_key_to_label(m), len(current), len(current), 0, 0, len(current)])
                continue
                
            prev_m = month_keys[i-1]
            prev = monthly_custs[prev_m]
            
            active = len(current)
            churned = len(prev - current)
            gained = len(current - prev)
            net = gained - churned
            
            rows.append([self._month_key_to_label(m), active, gained, churned, 0, net])
            
        self._write_table(ws, 1, 1, headers, rows)
        
        anchor = f"{get_column_letter(len(headers)+2)}2"
        self._add_combo_chart(
            ws, 1,
            [6], 
            [2], 
            1, 2, 1 + len(rows),
            "客户变动趋势 (净增vs活跃总量)",
            anchor
        )
        self._add_detailed_analysis_box(ws, len(headers)+2, 14, "客户流失与留存", [
            "活跃客户数：本月有交易记录的客户总量。",
            "新增客户：本月新出现(或回流)的客户。",
            "流失客户：上月活跃但本月无交易的客户。",
            "净增 = 新增 - 流失。净增持续为负预示着市场萎缩风险。"
        ])

        self._reorder_month_rows_desc(ws)

    def _add_detailed_analysis_box(self, ws, col, start_row, title, elements_desc, ai_mode=None):
        """
        在图表下方添加详细说明；可按需附加 AI 分析占位符。
        """
        if ai_mode is None:
            ai_mode = self._should_include_ai_placeholders()
        # 1. Title
        title_cell = ws.cell(row=start_row, column=col)
        title_cell.value = f"▼ {title} 解读指南"
        title_cell.font = Font(bold=True, size=11, color="1F4E79")
        
        # 2. Elements Description
        r = start_row + 1
        for desc in elements_desc:
            cell = ws.cell(row=r, column=col)
            cell.value = f"• {desc}"
            cell.font = Font(size=10, color="444444")
            r += 1
            
        # 3. AI Placeholder
        if ai_mode:
            r += 1
            ai_header = ws.cell(row=r, column=col)
            ai_header.value = "🤖 AI 智能分析"
            ai_header.font = Font(bold=True, color="E26B0A")
            
            r += 1
            ai_box = ws.cell(row=r, column=col)
            ai_box.value = "(此处预留用于 AI 模型根据图表数据生成的深度洞察与建议...)"
            ai_box.font = Font(italic=True, color="AAAAAA", size=9)
            ai_box.alignment = Alignment(wrap_text=True, vertical="top")
            ai_box.border = Border(
                left=Side(style='dotted', color="CCCCCC"),
                right=Side(style='dotted', color="CCCCCC"),
                top=Side(style='dotted', color="CCCCCC"),
                bottom=Side(style='dotted', color="CCCCCC")
            )
            # Merge rows for the box, be careful with columns
            try:
                ws.merge_cells(start_row=r, start_column=col, end_row=r+4, end_column=col+6)
            except Exception:
                # Ignore merge errors if the target range overlaps with existing merges.
                pass

if __name__ == "__main__":
    base_dir = r"C:\Users\123\Downloads\亿看智能识别系统\基础资料"
    template = r"C:\Users\123\Downloads\亿看智能识别系统\11月汇总结果_整理优化美化_含仪表盘_目标预算达成异常 (1).xlsx"
    output = r"C:\Users\123\Downloads\亿看智能识别系统\Generated_Report.xlsx"
    
    gen = ReportGenerator(base_dir)
    gen.load_all_data()
    success = gen.generate_report(template, output, "2025", "11", year_scope="current")
    
    if success:
        print("\n是否启动本地 LM Studio AI 分析? (确保 LM Studio 服务已开启端口 1234)")
        # In a real interactive CLI we might ask input(), but for automation we can default to trying or just print instructions.
        # Here we will try to import and run it if available.
        try:
            from local_llm_analyzer import LocalLLMAnalyzer
            print("正在尝试连接本地 LM Studio 进行分析...")
            backend = None
            try:
                import json
                from base_data_manager import BaseDataManager
                mgr = BaseDataManager()
                configs = mgr.get_all_configs()
                task_map = json.loads(configs.get("ai_task_map", "{}"))
                backends = json.loads(configs.get("ai_backends", "[]"))
                backend_name = task_map.get("report_analysis")
                if backend_name and backend_name != "(未配置)":
                    backend = next((b for b in backends if b.get("name") == backend_name), None)
                if not backend:
                    backend = {
                        "provider": configs.get("ai_provider", "lm_studio"),
                        "api_key": configs.get("ai_api_key", configs.get("api_key", "")),
                        "base_url": configs.get("ai_base_url", "http://localhost:1234/v1"),
                        "model": configs.get("ai_model_name", "local-model"),
                    }
            except Exception:
                backend = None

            if backend:
                provider = backend.get("provider") or "lm_studio"
                api_key = backend.get("api_key") or ""
                base_url = backend.get("base_url") or "http://localhost:1234/v1"
                if provider == "lm_studio":
                    model = backend.get("model") or "local-model"
                else:
                    model = backend.get("model") or "glm-4-flash"
                analyzer = LocalLLMAnalyzer(
                    api_base=base_url,
                    model=model,
                    api_key=api_key,
                    provider=provider,
                    enable_chart_recognition=True,
                )
            else:
                analyzer = LocalLLMAnalyzer(enable_chart_recognition=True)
            analysis_output = output.replace(".xlsx", "_AI_Analysis.md")
            analyzer.analyze_report(output, analysis_output, embed_to_excel=True)
        except ImportError:
            print("未找到 local_llm_analyzer 模块，跳过 AI 分析。")
        except Exception as e:
            print(f"AI 分析启动失败: {e}")
