import urllib.request
import urllib.error
import json
import math
import re
import pandas as pd
import openpyxl
import os
import time
from datetime import date, datetime
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, column_index_from_string

class LocalLLMAnalyzer:
    def __init__(
        self,
        api_base="http://localhost:1234/v1",
        model="local-model",
        timeout=120,
        provider="lm_studio",
        api_key="",
        base_data_dir="",
        enable_chart_recognition=False,
    ):
        self.api_base = (api_base or "http://localhost:1234/v1").rstrip("/")
        self.provider = provider or "lm_studio"
        default_model = "glm-4-flash" if self.provider == "zhipu" else "local-model"
        self.model = model or default_model
        self.timeout = timeout
        self.api_key = api_key or "lm-studio"
        self.ai_client = None
        self.init_error = ""
        self.max_sheet_rows = 80
        self.max_sheet_cols = 24
        self.max_chunk_count = 4
        self.max_chunk_rows = 60
        self.max_summary_chars = 18000
        self.max_related_context_chars = 2800
        self.max_chart_context_chars = 2600
        self.base_data_dir = (base_data_dir or "").strip()
        self.enable_chart_recognition = bool(enable_chart_recognition)
        self._related_cache = {}

        if self.provider == "zhipu":
            try:
                from zhipuai import ZhipuAI
                if not api_key:
                    raise ValueError("Zhipu API Key is missing")
                self.ai_client = ZhipuAI(api_key=api_key)
            except Exception as e:
                self.init_error = str(e)

    def _chat_completion(self, messages):
        if self.provider == "zhipu":
            if not self.ai_client:
                return f"Error initializing ZhipuAI client: {self.init_error}"
            try:
                response = self.ai_client.chat.completions.create(
                    model=self.model,
                    messages=messages,
                    temperature=0.7,
                )
                return response.choices[0].message.content
            except Exception as e:
                return f"Error during analysis: {e}"

        url = f"{self.api_base}/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}"
        }
        data = {
            "model": self.model,
            "messages": messages,
            "temperature": 0.7,
            "stream": False
        }
        
        try:
            print(f"DEBUG: calling {url} with model {self.model}...")
            req = urllib.request.Request(url, data=json.dumps(data).encode('utf-8'), headers=headers)
            with urllib.request.urlopen(req, timeout=self.timeout) as response:
                result = json.loads(response.read().decode('utf-8'))
                return result['choices'][0]['message']['content']
        except urllib.error.URLError as e:
            print(f"DEBUG: Failed to connect to {url}. Error: {e}")
            return f"Error connecting to LM Studio: {e}. Please ensure LM Studio is running and the server is started at {self.api_base}."
        except Exception as e:
            print(f"DEBUG: Unexpected error calling {url}. Error: {e}")
            return f"Error during analysis: {e}"

    def _truncate_analysis_text(self, text: str, max_len: int = 32000) -> str:
        if not text:
            return ""
        if len(text) <= max_len:
            return text
        return text[: max_len - 40] + "\n...(内容过长已截断)"

    def _is_token_limit_error(self, text: str) -> bool:
        if not text:
            return False
        s = str(text).lower()
        keywords = [
            "context length",
            "maximum context",
            "max context",
            "too many tokens",
            "token limit",
            "prompt is too long",
            "超出",
            "上下文长度",
            "token",
        ]
        if "error" not in s and "失败" not in s:
            return False
        return any(k in s for k in keywords)

    def _select_key_columns(self, df: pd.DataFrame, max_cols: int = 24):
        if df is None or df.empty:
            return []

        cols = [str(c) for c in df.columns]
        must_keywords = ["月", "month", "date", "日期", "时间", "品目", "产品", "客户", "部门", "指标", "项目"]
        first_pass = []
        for c in cols:
            cl = c.lower()
            if any(k in c for k in must_keywords) or any(k in cl for k in must_keywords):
                first_pass.append(c)

        numeric_cols = []
        for c in cols:
            try:
                s = pd.to_numeric(df[c], errors="coerce")
                if s.notna().any():
                    numeric_cols.append((c, int(s.notna().sum())))
            except Exception:
                continue
        numeric_cols.sort(key=lambda x: x[1], reverse=True)
        numeric_pick = [c for c, _ in numeric_cols[: max(6, max_cols // 2)]]

        selected = []
        for c in first_pass + numeric_pick + cols:
            if c not in selected:
                selected.append(c)
            if len(selected) >= max_cols:
                break
        return selected

    def _dataframe_to_prompt_table(self, df: pd.DataFrame, max_rows=None, max_cols=None) -> str:
        if df is None or df.empty:
            return "(empty)"
        max_rows = max_rows or self.max_sheet_rows
        max_cols = max_cols or self.max_sheet_cols

        key_cols = self._select_key_columns(df, max_cols=max_cols)
        safe_df = df[key_cols].copy() if key_cols else df.copy()
        if len(safe_df.index) > max_rows:
            head_rows = max_rows // 2
            tail_rows = max_rows - head_rows
            safe_df = pd.concat([safe_df.head(head_rows), safe_df.tail(tail_rows)], ignore_index=True)
            truncated_note = f"\n...(Truncated, total rows: {df.shape[0]}, sampled head+tail)"
        else:
            truncated_note = ""

        try:
            table_text = safe_df.to_markdown(index=False)
        except Exception:
            table_text = safe_df.to_string(index=False)
        return table_text + truncated_note

    def _build_df_profile(self, df: pd.DataFrame) -> str:
        if df is None or df.empty:
            return "DataFrame is empty."
        lines = [f"rows={df.shape[0]}, cols={df.shape[1]}"]
        cols = [str(c) for c in df.columns]
        lines.append("columns: " + ", ".join(cols[:30]) + (" ..." if len(cols) > 30 else ""))

        numeric_summaries = []
        for c in cols:
            s = pd.to_numeric(df[c], errors="coerce")
            if s.notna().sum() == 0:
                continue
            numeric_summaries.append(
                f"{c}: count={int(s.notna().sum())}, mean={float(s.mean()):.2f}, min={float(s.min()):.2f}, max={float(s.max()):.2f}"
            )
            if len(numeric_summaries) >= 8:
                break
        if numeric_summaries:
            lines.append("numeric summary:")
            lines.extend(numeric_summaries)
        return "\n".join(lines)

    def _normalize_month_key(self, value):
        if value is None:
            return None
        m = re.search(r"(20\d{2})\s*[年/-]\s*(\d{1,2})", str(value))
        if not m:
            return None
        return f"{m.group(1)}-{int(m.group(2)):02d}"

    def _parse_target_period(self, excel_path):
        name = os.path.basename(excel_path)
        m = re.search(r"(20\d{2})\s*年\s*(\d{1,2})\s*月", name)
        if m:
            return m.group(1), f"{int(m.group(2)):02d}"
        m = re.search(r"(20\d{2})[/-](\d{1,2})", name)
        if m:
            return m.group(1), f"{int(m.group(2)):02d}"
        return None, None

    def _detect_base_data_dir(self, excel_path):
        if self.base_data_dir and os.path.isdir(self.base_data_dir):
            return self.base_data_dir
        report_dir = os.path.dirname(excel_path)
        candidates = [
            os.path.join(report_dir, "基础资料"),
            os.path.join(report_dir, "基础数据"),
        ]
        for path in candidates:
            if os.path.isdir(path):
                return path
        return ""

    def _ensure_month_column(self, df, default_month=None):
        if df is None or df.empty:
            return df
        out = df.copy()
        if "MonthStr" in out.columns:
            out["MonthStr"] = out["MonthStr"].astype(str).str.strip()
            return out
        date_col = next((c for c in out.columns if "日期" in str(c) or "Date" in str(c)), None)
        if date_col:
            parsed = pd.to_datetime(out[date_col], errors="coerce")
            out["MonthStr"] = parsed.dt.strftime("%Y-%m")
            return out
        if default_month:
            out["MonthStr"] = default_month
        return out

    def _select_text_column(self, df, candidates, contains_keyword=None):
        for col in candidates:
            if col in df.columns:
                return col
        if contains_keyword:
            for col in df.columns:
                if contains_keyword in str(col):
                    return col
        return None

    def _extract_expense_amount_series(self, df):
        debit_col = self._select_text_column(df, ["借方金额"], contains_keyword="借方金额")
        credit_col = self._select_text_column(df, ["贷方金额"], contains_keyword="贷方金额")
        if debit_col and credit_col:
            debit = pd.to_numeric(df[debit_col], errors="coerce").fillna(0)
            credit = pd.to_numeric(df[credit_col], errors="coerce").fillna(0)
            return debit - credit
        for col in ["金额", "本期值", "外币借方金额"]:
            if col in df.columns:
                s = pd.to_numeric(df[col], errors="coerce")
                if s.notna().any():
                    return s
        for col in df.columns:
            if "金额" in str(col):
                s = pd.to_numeric(df[col], errors="coerce")
                if s.notna().any():
                    return s
        return pd.Series([None] * len(df), index=df.index, dtype="float64")

    def _format_number(self, value):
        if value is None or pd.isna(value):
            return "NA"
        try:
            return f"{float(value):,.2f}"
        except Exception:
            return str(value)

    def _format_prompt_value(self, value):
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return "NA"
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, (int, float)):
            return self._format_number(value)
        text = str(value).strip().replace("\n", " ")
        if len(text) > 36:
            text = text[:33] + "..."
        return text or "NA"

    def _extract_chart_title_text(self, chart):
        title = getattr(chart, "title", None)
        if title is None:
            return ""
        if isinstance(title, str):
            return title.strip()
        try:
            tx = getattr(title, "tx", None)
            rich = getattr(tx, "rich", None) if tx else None
            if rich and getattr(rich, "p", None):
                parts = []
                for para in rich.p:
                    runs = getattr(para, "r", None) or []
                    for run in runs:
                        t = getattr(run, "t", None)
                        if t:
                            parts.append(str(t))
                    fields = getattr(para, "fld", None) or []
                    for fld in fields:
                        t = getattr(fld, "t", None)
                        if t:
                            parts.append(str(t))
                if parts:
                    return "".join(parts).strip()
        except Exception:
            pass
        return str(title).strip()

    def _parse_excel_range_formula(self, formula):
        if not formula:
            return None
        expr = str(formula).strip()
        if expr.startswith("="):
            expr = expr[1:].strip()
        m = re.match(
            r"^(?:'((?:[^']|'')+)'|([^'!]+))!\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$",
            expr,
            re.IGNORECASE,
        )
        if not m:
            return None

        sheet_name = (m.group(1) or m.group(2) or "").replace("''", "'")
        c1 = column_index_from_string(m.group(3).upper())
        r1 = int(m.group(4))
        c2 = column_index_from_string((m.group(5) or m.group(3)).upper())
        r2 = int(m.group(6) or m.group(4))
        min_col, max_col = sorted((c1, c2))
        min_row, max_row = sorted((r1, r2))
        return {
            "sheet": sheet_name,
            "min_col": min_col,
            "max_col": max_col,
            "min_row": min_row,
            "max_row": max_row,
        }

    def _collect_formula_values(self, wb, formula, max_cells=2000):
        parsed = self._parse_excel_range_formula(formula)
        if not parsed:
            return []
        sheet_name = parsed["sheet"]
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        values = []
        for row in range(parsed["min_row"], parsed["max_row"] + 1):
            for col in range(parsed["min_col"], parsed["max_col"] + 1):
                values.append(ws.cell(row=row, column=col).value)
                if len(values) >= max_cells:
                    return values
        return values

    def _sample_value_text(self, values, limit=6):
        if not values:
            return "[]"
        filtered = []
        for value in values:
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            filtered.append(value)
        if not filtered:
            return "[]"
        if len(filtered) <= limit:
            picks = filtered
        else:
            head = max(1, limit // 2)
            tail = max(1, limit - head)
            picks = filtered[:head] + filtered[-tail:]
        rendered = [self._format_prompt_value(v) for v in picks]
        return "[" + ", ".join(rendered) + "]"

    def _extract_series_name(self, wb, series, idx):
        title = getattr(series, "title", None)
        if title is None:
            return f"系列{idx}"
        value = getattr(title, "v", None)
        if value:
            return str(value)
        str_ref = getattr(title, "strRef", None)
        if str_ref and getattr(str_ref, "f", None):
            name_vals = self._collect_formula_values(wb, str_ref.f, max_cells=1)
            if name_vals:
                return self._format_prompt_value(name_vals[0])
            return str(str_ref.f)
        text = str(title).strip()
        if not text:
            return f"系列{idx}"
        return text if len(text) <= 48 else (text[:45] + "...")

    def _build_chart_context_for_sheet(self, wb, sheet_name):
        if sheet_name not in wb.sheetnames:
            return ""
        ws = wb[sheet_name]
        charts = list(getattr(ws, "_charts", []) or [])
        if not charts:
            return ""

        max_charts = min(4, len(charts))
        lines = [f"【图表识别】本sheet共识别到 {len(charts)} 个图表。"]
        for chart_idx, chart in enumerate(charts[:max_charts], start=1):
            chart_type = chart.__class__.__name__
            chart_title = self._extract_chart_title_text(chart) or "(无标题)"
            series_list = list(getattr(chart, "series", []) or [])
            lines.append(
                f"- 图表{chart_idx}: 类型={chart_type}, 标题={chart_title}, 系列数={len(series_list)}"
            )
            for series_idx, series in enumerate(series_list[:5], start=1):
                series_name = self._extract_series_name(wb, series, series_idx)
                cat_formula = ""
                val_formula = ""
                try:
                    cat_formula = series.cat.numRef.f
                except Exception:
                    cat_formula = ""
                try:
                    val_formula = series.val.numRef.f
                except Exception:
                    val_formula = ""

                cat_values = self._collect_formula_values(wb, cat_formula, max_cells=240) if cat_formula else []
                val_values = self._collect_formula_values(wb, val_formula, max_cells=240) if val_formula else []
                cat_sample = self._sample_value_text(cat_values, limit=6)
                val_sample = self._sample_value_text(val_values, limit=6)

                num = pd.to_numeric(pd.Series(val_values), errors="coerce").dropna() if val_values else pd.Series(dtype="float64")
                stats = ""
                if not num.empty:
                    stats = f"，最小={self._format_number(num.min())}，最大={self._format_number(num.max())}"

                lines.append(
                    f"  - 系列{series_idx}({series_name}): 分类样本={cat_sample}；数值样本={val_sample}{stats}"
                )
                if val_formula:
                    lines.append(f"    数值范围={val_formula}")
            if len(series_list) > 5:
                lines.append(f"  - 其余系列省略 {len(series_list) - 5} 个")

        if len(charts) > max_charts:
            lines.append(f"- 其余图表省略 {len(charts) - max_charts} 个")
        return self._truncate_analysis_text("\n".join(lines), max_len=self.max_chart_context_chars)

    def _collect_sheet_chart_context_map(self, excel_path, target_sheets=None):
        if not self.enable_chart_recognition:
            return {}
        if not os.path.exists(excel_path):
            return {}

        contexts = {}
        wb = None
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheets = list(target_sheets or wb.sheetnames)
            for name in sheets:
                if name not in wb.sheetnames:
                    continue
                block = self._build_chart_context_for_sheet(wb, name)
                if block:
                    contexts[name] = block
        except Exception as e:
            print(f"图表识别初始化失败: {e}")
        finally:
            if wb is not None:
                wb.close()
        return contexts

    def analyze_chart_context(self, sheet_name, context=""):
        prompt = f"""
你是资深经营分析师。请基于以下图表识别结果，解读 sheet "{sheet_name}" 的业务表现。
{context}

要求:
1. 识别主要趋势、结构变化与异常点。
2. 结合图表类型和样本数据判断潜在风险。
3. 给出可执行建议。

请严格输出:
【分析】
- 3~5条关键发现
【结论】
- 1~3条可执行建议
"""
        messages = [
            {"role": "system", "content": "你是严谨、简洁的财务分析专家。"},
            {"role": "user", "content": prompt},
        ]
        result = self._chat_completion(messages)
        if self._is_token_limit_error(result):
            retry_prompt = f"""
请仅基于以下图表识别结果输出精简结论：
{self._truncate_analysis_text(context, max_len=1600)}

格式:
【分析】
- 3条
【结论】
- 2条
"""
            result = self._chat_completion(
                [
                    {"role": "system", "content": "你是严谨、简洁的财务分析专家。"},
                    {"role": "user", "content": retry_prompt},
                ]
            )
        return result

    def _build_expense_deep_dive(self, gen, target_year=None):
        frames = []
        for month_key, df in gen.data.get("expense", {}).items():
            if df is None or df.empty:
                continue
            if target_year and not str(month_key).startswith(f"{target_year}-"):
                continue
            scoped = self._ensure_month_column(df, default_month=month_key)
            frames.append(scoped)
        if not frames:
            return ""

        data = pd.concat(frames, ignore_index=True)
        if "MonthStr" not in data.columns:
            return ""
        data = data[data["MonthStr"].notna()].copy()
        if data.empty:
            return ""

        data["_Amount"] = self._extract_expense_amount_series(data)
        data = data[data["_Amount"].notna()]
        if data.empty:
            return ""

        monthly_total = data.groupby("MonthStr")["_Amount"].sum().sort_index()
        if monthly_total.empty:
            return ""
        monthly_delta = monthly_total.diff()
        rank = monthly_delta.abs().sort_values(ascending=False)
        months = [m for m in rank.index if isinstance(m, str)][:3]
        if not months:
            months = list(monthly_total.tail(3).index)

        subject_col = self._select_text_column(data, ["科目名", "子科目"], contains_keyword="科目")
        dept_col = self._select_text_column(data, ["部门名", "部门"], contains_keyword="部门")
        lines = ["【跨表补充-费用明细（基础资料）】"]
        for month_key in months:
            m_df = data[data["MonthStr"] == month_key].copy()
            if m_df.empty:
                continue
            total = monthly_total.get(month_key)
            delta = monthly_delta.get(month_key)
            line = f"- {month_key}: 总费用={self._format_number(total)}, 较上月变化={self._format_number(delta)}"
            if subject_col:
                top_subject = (
                    m_df.groupby(subject_col)["_Amount"].sum()
                    .sort_values(key=lambda s: s.abs(), ascending=False)
                    .head(4)
                )
                if not top_subject.empty:
                    detail = "；".join([f"{idx}:{self._format_number(v)}" for idx, v in top_subject.items()])
                    line += f"，Top科目={detail}"
            if dept_col:
                top_dept = (
                    m_df.groupby(dept_col)["_Amount"].sum()
                    .sort_values(key=lambda s: s.abs(), ascending=False)
                    .head(2)
                )
                if not top_dept.empty:
                    detail = "；".join([f"{idx}:{self._format_number(v)}" for idx, v in top_dept.items()])
                    line += f"，Top部门={detail}"
            lines.append(line)
        return "\n".join(lines)

    def _build_sales_deep_dive(self, gen, target_year=None):
        sales_df = gen._get_sales_df()
        if sales_df is None or sales_df.empty:
            return ""
        data = sales_df.copy()
        if target_year:
            if "MonthStr" not in data.columns:
                data = self._ensure_month_column(data)
            if "MonthStr" in data.columns:
                data = data[data["MonthStr"].astype(str).str.startswith(f"{target_year}-")]
        if data.empty:
            return ""

        data["Revenue"] = gen._extract_sales_revenue(data)
        data = gen._attach_sales_cost(data, target_year=target_year, target_month=None, year_scope="all")
        if "MonthStr" not in data.columns:
            return ""
        data = data[data["MonthStr"].notna()].copy()
        if data.empty:
            return ""

        monthly = (
            data.groupby("MonthStr")[["Revenue", "Cost"]]
            .sum(min_count=1)
            .sort_index()
        )
        monthly["Profit"] = monthly["Revenue"] - monthly["Cost"]
        monthly["Delta"] = monthly["Revenue"].diff()
        rank = monthly["Delta"].abs().sort_values(ascending=False)
        months = [m for m in rank.index if isinstance(m, str)][:2]
        if not months:
            months = list(monthly.tail(2).index)

        product_col = self._select_text_column(data, ["品目名", "产品名称", "品名"], contains_keyword="品目")
        lines = ["【跨表补充-销售明细（基础资料）】"]
        for month_key in months:
            row = monthly.loc[month_key]
            line = (
                f"- {month_key}: 收入={self._format_number(row.get('Revenue'))}, "
                f"成本={self._format_number(row.get('Cost'))}, "
                f"利润={self._format_number(row.get('Profit'))}, "
                f"收入环比变化={self._format_number(row.get('Delta'))}"
            )
            if product_col:
                m_df = data[data["MonthStr"] == month_key].copy()
                if not m_df.empty:
                    top_products = (
                        m_df.groupby(product_col)["Revenue"].sum()
                        .sort_values(ascending=False)
                        .head(3)
                    )
                    if not top_products.empty:
                        detail = "；".join([f"{idx}:{self._format_number(v)}" for idx, v in top_products.items()])
                        line += f"，Top产品={detail}"
            lines.append(line)
        return "\n".join(lines)

    def _build_kpi_snapshot(self, excel_path, target_year=None, target_month=None):
        if not os.path.exists(excel_path):
            return ""
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
            if "经营指标" not in wb.sheetnames:
                wb.close()
                return ""
            ws = wb["经营指标"]
            headers = {str(ws.cell(row=1, column=c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
            target_key = f"{target_year}-{int(target_month):02d}" if target_year and target_month else None
            target_row = None
            for r in range(2, ws.max_row + 1):
                key = self._normalize_month_key(ws.cell(row=r, column=1).value)
                if key and key == target_key:
                    target_row = r
                    break
            if target_row is None:
                wb.close()
                return ""
            key_metrics = ["主营业务收入", "主营业务成本", "销售费用", "管理费用", "营业利润", "应收账款余额", "存货期末余额"]
            parts = []
            for metric in key_metrics:
                col = headers.get(metric)
                if not col:
                    continue
                parts.append(f"{metric}={self._format_number(ws.cell(row=target_row, column=col).value)}")
            wb.close()
            if not parts:
                return ""
            return "【跨表补充-当期KPI】\n- " + "；".join(parts)
        except Exception:
            return ""

    def _prepare_related_cache(self, excel_path):
        cache = self._related_cache.get(excel_path)
        if cache:
            return cache

        target_year, target_month = self._parse_target_period(excel_path)
        base_dir = self._detect_base_data_dir(excel_path)
        expense_pack = ""
        sales_pack = ""
        if base_dir:
            try:
                from report_generator import ReportGenerator
                gen = ReportGenerator(base_dir)
                gen.load_all_data()
                expense_pack = self._build_expense_deep_dive(gen, target_year=target_year)
                sales_pack = self._build_sales_deep_dive(gen, target_year=target_year)
            except Exception as e:
                expense_pack = f"【跨表补充-费用明细（基础资料）】\n- 读取失败: {e}"

        kpi_pack = self._build_kpi_snapshot(excel_path, target_year=target_year, target_month=target_month)
        cache = {
            "target_year": target_year,
            "target_month": target_month,
            "expense_pack": expense_pack,
            "sales_pack": sales_pack,
            "kpi_pack": kpi_pack,
        }
        self._related_cache[excel_path] = cache
        return cache

    def _build_related_context_for_sheet(self, excel_path, sheet_name, df):
        cache = self._prepare_related_cache(excel_path)
        blocks = []

        kpi_pack = cache.get("kpi_pack", "")
        expense_pack = cache.get("expense_pack", "")
        sales_pack = cache.get("sales_pack", "")

        if kpi_pack and sheet_name not in {"经营指标", "年度经营指标"}:
            blocks.append(kpi_pack)

        expense_keywords = ["费用", "预算", "同比经营分析", "环比经营分析", "经营指标", "利润表"]
        if expense_pack and any(k in sheet_name for k in expense_keywords):
            blocks.append(expense_pack)

        sales_keywords = ["销售", "产品", "客户", "渠道", "库存", "贡献", "本量利"]
        if sales_pack and any(k in sheet_name for k in sales_keywords):
            blocks.append(sales_pack)

        if not blocks:
            return ""
        return self._truncate_analysis_text("\n\n".join(blocks), max_len=self.max_related_context_chars)

    def _analyze_large_dataframe(self, df, sheet_name, context=""):
        key_cols = self._select_key_columns(df, max_cols=self.max_sheet_cols)
        work_df = df[key_cols].copy() if key_cols else df.copy()
        total_rows = len(work_df.index)
        chunk_size = self.max_chunk_rows
        chunk_count = min(self.max_chunk_count, max(1, int(math.ceil(total_rows / chunk_size))))
        chunk_analyses = []

        for i in range(chunk_count):
            start = i * chunk_size
            end = min((i + 1) * chunk_size, total_rows)
            chunk_df = work_df.iloc[start:end].copy()
            if chunk_df.empty:
                continue
            chunk_table = self._dataframe_to_prompt_table(
                chunk_df,
                max_rows=self.max_chunk_rows,
                max_cols=self.max_sheet_cols,
            )
            prompt = f"""
你是资深经营分析师。请分析 sheet "{sheet_name}" 的第 {i+1}/{chunk_count} 个数据分片。
{context}

分片范围: 行 {start+1}-{end} / 总行数 {total_rows}
数据:
{chunk_table}

请严格输出:
【分析】
- 3条以内关键发现（趋势/异常/风险）
【结论】
- 1~2条可执行建议
"""
            messages = [
                {"role": "system", "content": "你是严谨、简洁的财务分析专家。"},
                {"role": "user", "content": prompt},
            ]
            chunk_result = self._chat_completion(messages)
            chunk_result = self._truncate_analysis_text(chunk_result, max_len=2500)
            chunk_analyses.append(f"分片{i+1}:\n{chunk_result}")
            time.sleep(0.5)

        profile_text = self._build_df_profile(work_df)
        synthesis_prompt = f"""
你是资深经营分析师。下面是同一张大表（{sheet_name}）的分片分析结果，请合并成最终结论。
{context}

数据概况:
{profile_text}

分片分析:
{chr(10).join(chunk_analyses)}

请严格输出:
【分析】
- 综合趋势、关键异常、潜在原因（4~6条）
【结论】
- 2~3条优先级最高的建议（可执行）
"""
        messages = [
            {"role": "system", "content": "你是严谨、简洁的财务分析专家。"},
            {"role": "user", "content": synthesis_prompt},
        ]
        result = self._chat_completion(messages)
        if self._is_token_limit_error(result):
            # 极端情况下再次降维，只用概况+分片标题。
            lightweight_chunks = "\n".join([f"分片{i+1}: {txt[:400]}" for i, txt in enumerate(chunk_analyses)])
            retry_prompt = f"""
请基于以下概要输出最终结论，不展开细节。
sheet={sheet_name}
{context}
{profile_text}
{lightweight_chunks}

格式:
【分析】
- 3~5条
【结论】
- 2条
"""
            result = self._chat_completion(
                [
                    {"role": "system", "content": "你是严谨、简洁的财务分析专家。"},
                    {"role": "user", "content": retry_prompt},
                ]
            )
        return result

    def _split_analysis_and_conclusion(self, text: str):
        if not text:
            return "", ""
        raw = text.strip()
        analysis = raw
        conclusion = ""
        if "【结论】" in raw:
            parts = raw.split("【结论】", 1)
            analysis = parts[0].replace("【分析】", "").strip()
            conclusion = parts[1].strip()
        elif "结论:" in raw:
            parts = raw.split("结论:", 1)
            analysis = parts[0].replace("【分析】", "").strip()
            conclusion = parts[1].strip()
        elif "结论：" in raw:
            parts = raw.split("结论：", 1)
            analysis = parts[0].replace("【分析】", "").strip()
            conclusion = parts[1].strip()
        return analysis, conclusion

    def _find_last_used_row(self, ws):
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        for row in range(max_row, 0, -1):
            for col in range(1, max_col + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None and str(val).strip() != "":
                    return row
        return 1

    def _find_analysis_anchor(self, ws):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() == "AI分析":
                    return cell.row
        return None

    def _safe_merge(self, ws, start_row, start_col, end_row, end_col):
        if end_col <= start_col:
            return
        cell_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        if any(str(rng) == cell_range for rng in ws.merged_cells.ranges):
            return
        ws.merge_cells(cell_range)

    def _write_analysis_to_sheet(self, ws, analysis_text: str, conclusion_text: str = ""):
        if not analysis_text and not conclusion_text:
            return
        analysis_text = self._truncate_analysis_text((analysis_text or "").strip())
        conclusion_text = self._truncate_analysis_text((conclusion_text or "").strip())
        anchor_row = self._find_analysis_anchor(ws)
        if anchor_row:
            label_row = anchor_row
        else:
            label_row = self._find_last_used_row(ws) + 2
        value_row = label_row + 1
        max_col = max(1, min(ws.max_column or 1, 12))

        label_cell = ws.cell(row=label_row, column=1)
        label_cell.value = "AI分析"
        label_cell.font = Font(bold=True)
        self._safe_merge(ws, label_row, 1, label_row, max_col)

        if analysis_text:
            value_cell = ws.cell(row=value_row, column=1)
            value_cell.value = analysis_text
            value_cell.alignment = Alignment(wrap_text=True, vertical="top")
            self._safe_merge(ws, value_row, 1, value_row, max_col)
            ws.row_dimensions[value_row].height = 120

        if conclusion_text:
            conclusion_label_row = value_row + 2
            conclusion_value_row = conclusion_label_row + 1
            conclusion_label_cell = ws.cell(row=conclusion_label_row, column=1)
            conclusion_label_cell.value = "AI结论"
            conclusion_label_cell.font = Font(bold=True)
            self._safe_merge(ws, conclusion_label_row, 1, conclusion_label_row, max_col)

            conclusion_cell = ws.cell(row=conclusion_value_row, column=1)
            conclusion_cell.value = conclusion_text
            conclusion_cell.alignment = Alignment(wrap_text=True, vertical="top")
            self._safe_merge(ws, conclusion_value_row, 1, conclusion_value_row, max_col)
            ws.row_dimensions[conclusion_value_row].height = 80

    def _write_analyses_to_workbook(self, excel_path: str, analyses_by_sheet: dict, output_path: str = None):
        if not analyses_by_sheet:
            return
        wb = openpyxl.load_workbook(excel_path)
        for sheet_name, analysis in analyses_by_sheet.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            if isinstance(analysis, dict):
                self._write_analysis_to_sheet(ws, analysis.get("analysis", ""), analysis.get("conclusion", ""))
            else:
                parsed_analysis, parsed_conclusion = self._split_analysis_and_conclusion(analysis)
                self._write_analysis_to_sheet(ws, parsed_analysis, parsed_conclusion)
        wb.save(output_path or excel_path)

    def analyze_dataframe(self, df, sheet_name, context=""):
        """
        Analyze a specific dataframe using the local LLM.
        """
        if df is None or df.empty:
            return "【分析】\n- 表格为空，暂无可分析数据。\n【结论】\n- 无需处理。"

        if df.shape[0] > self.max_sheet_rows or df.shape[1] > self.max_sheet_cols:
            return self._analyze_large_dataframe(df, sheet_name, context)

        csv_text = self._dataframe_to_prompt_table(df, max_rows=self.max_sheet_rows, max_cols=self.max_sheet_cols)

        prompt = f"""
You are a senior financial analyst. Analyze the following financial data from the sheet "{sheet_name}".
{context}

Data:
{csv_text}

Please return the result strictly in this format (do not add extra titles):
【分析】
- Key trends and anomalies.
- Significant changes compared to previous periods (if applicable).
- Potential areas of concern or opportunities.
【结论】
- One to three concise conclusions that can be placed back into the same sheet.
"""
        messages = [
            {"role": "system", "content": "You are a helpful expert financial analyst."},
            {"role": "user", "content": prompt}
        ]
        
        print(f"Sending data from '{sheet_name}' to Local LLM...")
        analysis = self._chat_completion(messages)
        if self._is_token_limit_error(analysis):
            slim_table = self._dataframe_to_prompt_table(df, max_rows=40, max_cols=16)
            retry_prompt = f"""
You are a senior financial analyst. The previous request exceeded token limits.
Please provide a concise analysis based on this reduced dataset from "{sheet_name}".
{context}

Data:
{slim_table}

Format:
【分析】
- 3-5 key findings
【结论】
- 1-2 actions
"""
            analysis = self._chat_completion(
                [
                    {"role": "system", "content": "You are a helpful expert financial analyst."},
                    {"role": "user", "content": retry_prompt},
                ]
            )
        return analysis

    def _collect_sheet_analyses(self, excel_path, sheets_to_analyze, include_related_context=False):
        analyses = {}
        all_texts = []
        chart_context_map = self._collect_sheet_chart_context_map(
            excel_path,
            target_sheets=list(sheets_to_analyze.keys()),
        )
        with pd.ExcelFile(excel_path) as xls:
            for sheet, context in sheets_to_analyze.items():
                if sheet not in xls.sheet_names:
                    continue
                print(f"Analyzing sheet: {sheet}...")
                try:
                    df = pd.read_excel(xls, sheet_name=sheet)
                    df.dropna(how='all', inplace=True)
                    df.dropna(axis=1, how='all', inplace=True)
                    chart_context = chart_context_map.get(sheet, "")
                    if df.empty:
                        if chart_context:
                            analysis = self.analyze_chart_context(sheet, f"{context}\n\n{chart_context}")
                            parsed_analysis, parsed_conclusion = self._split_analysis_and_conclusion(analysis)
                            analyses[sheet] = {
                                "analysis": parsed_analysis or analysis,
                                "conclusion": parsed_conclusion,
                            }
                            combined = parsed_analysis or analysis
                            if parsed_conclusion:
                                combined += f"\n结论: {parsed_conclusion}"
                            all_texts.append(self._truncate_analysis_text(combined, max_len=2200))
                        else:
                            analyses[sheet] = {"analysis": "表格为空，未生成AI分析。", "conclusion": ""}
                        continue
                    full_context = context
                    if include_related_context:
                        related = self._build_related_context_for_sheet(excel_path, sheet, df)
                        if related:
                            full_context = f"{context}\n\n{related}"
                    if chart_context:
                        full_context = f"{full_context}\n\n{chart_context}"
                    analysis = self.analyze_dataframe(df, sheet, full_context)
                    parsed_analysis, parsed_conclusion = self._split_analysis_and_conclusion(analysis)
                    analyses[sheet] = {
                        "analysis": parsed_analysis or analysis,
                        "conclusion": parsed_conclusion,
                    }
                    combined = parsed_analysis or analysis
                    if parsed_conclusion:
                        combined += f"\n结论: {parsed_conclusion}"
                    all_texts.append(self._truncate_analysis_text(combined, max_len=2200))
                    time.sleep(1)
                except Exception as e:
                    analyses[sheet] = {"analysis": f"表格分析失败: {e}", "conclusion": ""}
        return analyses, all_texts

    def analyze_reconciliation_report(self, excel_path, output_md_path=None, embed_to_excel=False):
        """
        Specific analysis for Reconciliation Reports.
        """
        if not os.path.exists(excel_path):
            return f"Error: File not found {excel_path}"

        print(f"Starting AI Analysis for Reconciliation Report: {excel_path}")
        
        sheets_to_analyze = {
            "汇总": "Summary of reconciliation results. Focus on the count of unmatched items and the overall match rate.",
            "当地未匹配(建议新增)": "Items present in Local system but missing in Yikan. Analyze common dates, large amounts, or missing patterns.",
            "亿看未匹配(建议清理)": "Items present in Yikan but missing in Local. Analyze if these are old outstanding items, duplicates, or specific types.",
            "AI建议匹配": "AI suggested matches. Review the confidence and types of matches found."
        }
        
        full_report = f"# AI Reconciliation Analysis Report\n\n**Source File**: {os.path.basename(excel_path)}\n**Date**: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}\n\n"
        
        analyses = []
        try:
            analyses_by_sheet, analyses = self._collect_sheet_analyses(
                excel_path,
                sheets_to_analyze,
                include_related_context=False,
            )
            for sheet, analysis in analyses_by_sheet.items():
                if isinstance(analysis, dict):
                    analysis_text = analysis.get("analysis", "")
                    conclusion_text = analysis.get("conclusion", "")
                else:
                    analysis_text, conclusion_text = self._split_analysis_and_conclusion(analysis)

                if analysis_text.startswith("表格为空"):
                    full_report += f"## Analysis: {sheet}\n\n*Sheet was empty (Good news for unmatched items).*\n\n"
                elif analysis_text.startswith("表格分析失败"):
                    full_report += f"## Analysis: {sheet}\n\n*Error processing sheet: {analysis_text}*\n\n"
                else:
                    block = analysis_text
                    if conclusion_text:
                        block += f"\n\n结论：{conclusion_text}"
                    full_report += f"## Analysis: {sheet}\n\n{block}\n\n---\n\n"

            # Generate Final Executive Summary
            if analyses:
                print("Generating Final Executive Summary...")
                final_summary = self.generate_summary_of_analyses(analyses)
                full_report = f"# Executive Summary\n\n{final_summary}\n\n---\n\n" + full_report
            if embed_to_excel and analyses_by_sheet:
                self._write_analyses_to_workbook(excel_path, analyses_by_sheet)
        except Exception as e:
            return f"Failed to read Excel file: {e}"

        if output_md_path:
            with open(output_md_path, 'w', encoding='utf-8') as f:
                f.write(full_report)
            print(f"Analysis saved to: {output_md_path}")
            return full_report
        else:
            return full_report

    def generate_summary_of_analyses(self, analyses_texts):
        """
        Synthesize multiple analysis sections into a cohesive executive summary.
        """
        if not analyses_texts:
            return "未生成可汇总的分表分析。"

        normalized = [self._truncate_analysis_text(t, max_len=2200) for t in analyses_texts if t]
        combined_text = "\n\n".join(normalized)

        # Large report: hierarchical summarize to avoid token overflow.
        if len(combined_text) > self.max_summary_chars:
            batch_size = 6
            stage1 = []
            for i in range(0, len(normalized), batch_size):
                chunk = normalized[i:i + batch_size]
                chunk_text = "\n\n".join(chunk)
                chunk_prompt = f"""
You are a lead financial auditor. Summarize the following analyses batch into 4-6 bullets.
Focus on major risks, anomalies, and actions.

Analyses:
{chunk_text}
"""
                stage1_summary = self._chat_completion(
                    [
                        {"role": "system", "content": "You are an expert financial auditor."},
                        {"role": "user", "content": chunk_prompt},
                    ]
                )
                stage1.append(self._truncate_analysis_text(stage1_summary, max_len=1600))

            combined_text = "\n\n".join(stage1)

        prompt = f"""
You are a lead financial auditor. Review the following individual analyses of different parts of a financial report/reconciliation.
Synthesize them into a high-level Executive Summary.

Key requirements:
1. Highlight the most critical issues (e.g. large discrepancies, missing data).
2. Provide actionable recommendations.
3. Be professional and concise.

Analyses to summarize:
{combined_text}
"""
        messages = [
            {"role": "system", "content": "You are an expert financial auditor."},
            {"role": "user", "content": prompt}
        ]
        summary = self._chat_completion(messages)
        if self._is_token_limit_error(summary):
            fallback_prompt = f"""
Please output only the top 5 findings and top 3 actions from these analyses:
{self._truncate_analysis_text(combined_text, max_len=8000)}
"""
            summary = self._chat_completion(
                [
                    {"role": "system", "content": "You are an expert financial auditor."},
                    {"role": "user", "content": fallback_prompt},
                ]
            )
        return summary

    def analyze_report(self, excel_path, output_md_path=None, embed_to_excel=False):
        """
        Reads key sheets from the Excel report and generates an analysis for each.
        Dispatcher that detects report type.
        """
        if "对账" in os.path.basename(excel_path) or "Reconciliation" in os.path.basename(excel_path):
            return self.analyze_reconciliation_report(excel_path, output_md_path, embed_to_excel=embed_to_excel)
            
        # Default to Business Operation Report analysis
        if not os.path.exists(excel_path):
            return f"Error: File not found {excel_path}"

        print(f"Starting AI Analysis for report: {excel_path}")
        
        # Define default contexts; fall back to generic prompt for other sheets
        base_contexts = {
            "利润表": "This is the Profit and Loss statement. Focus on Revenue, Cost, Gross Profit, and Net Profit trends.",
            "同比经营分析": "Year-over-Year (YoY) analysis. Compare current month/period with the same period last year.",
            "环比经营分析": "Month-over-Month (MoM) analysis. Compare current month with the previous month.",
            "按产品汇总_含合计": "Sales performance by product. Identify top sellers and low performers.",
            "费用明细": "Detailed expense breakdown. Look for unusual spikes in specific expense categories.",
            "经营指标": "Key Business Indicators (KPIs). Analyze the overall health and performance metrics.",
            "资产负债表": "Balance sheet. Focus on asset/liability structure and key balance changes.",
            "现金流量表(估算)": "Cash flow estimation. Review operating/financing/investing signals.",
            "年度利润表": "Annual P&L summary. Emphasize full-year performance and key ratios.",
            "年度经营指标": "Annual KPI summary. Highlight the strongest and weakest months.",
            "年度资产负债表": "Annual balance comparison between beginning and end of year.",
        }

        exclude_exact = {
            "目录",
            "图表数据源_隐藏",
            "产品对比(动态图表)",
        }
        exclude_contains = ["图表数据源", "隐藏", "动态图表"]

        sheets_to_analyze = {}
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        for sheet_name in sheet_names:
            if sheet_name in exclude_exact:
                continue
            if any(token in sheet_name for token in exclude_contains):
                continue
            context = base_contexts.get(
                sheet_name,
                f"This is a financial sheet named \"{sheet_name}\". Provide concise analysis and conclusions based on the data layout."
            )
            sheets_to_analyze[sheet_name] = context
        
        full_report = f"# AI Business Analysis Report\n\n**Source File**: {os.path.basename(excel_path)}\n**Date**: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}\n\n"
        analyses = []
        
        try:
            analyses_by_sheet, analyses = self._collect_sheet_analyses(
                excel_path,
                sheets_to_analyze,
                include_related_context=True,
            )
            for sheet, analysis in analyses_by_sheet.items():
                if isinstance(analysis, dict):
                    analysis_text = analysis.get("analysis", "")
                    conclusion_text = analysis.get("conclusion", "")
                else:
                    analysis_text, conclusion_text = self._split_analysis_and_conclusion(analysis)

                if analysis_text.startswith("表格为空"):
                    full_report += f"## Analysis: {sheet}\n\n*Sheet was empty.*\n\n"
                elif analysis_text.startswith("表格分析失败"):
                    full_report += f"## Analysis: {sheet}\n\n*Error processing sheet: {analysis_text}*\n\n"
                else:
                    block = analysis_text
                    if conclusion_text:
                        block += f"\n\n结论：{conclusion_text}"
                    full_report += f"## Analysis: {sheet}\n\n{block}\n\n---\n\n"
            for sheet in sheets_to_analyze:
                if sheet not in analyses_by_sheet:
                    print(f"Skipping {sheet} (not found in workbook)")

            # Generate Final Executive Summary
            if analyses:
                print("Generating Final Executive Summary...")
                final_summary = self.generate_summary_of_analyses(analyses)
                full_report = f"# Executive Summary\n\n{final_summary}\n\n---\n\n" + full_report
            if embed_to_excel and analyses_by_sheet:
                self._write_analyses_to_workbook(excel_path, analyses_by_sheet)

        except Exception as e:
            return f"Failed to read Excel file: {e}"

        if output_md_path:
            with open(output_md_path, 'w', encoding='utf-8') as f:
                f.write(full_report)
            print(f"Analysis saved to: {output_md_path}")
            return full_report
        else:
            return full_report

if __name__ == "__main__":
    # Example usage
    # Ensure LM Studio is running with a model loaded and server started at port 1234
    analyzer = LocalLLMAnalyzer()
    
    # Check if we have a report to analyze
    report_file = r"C:\Users\123\Downloads\亿看智能识别系统\Generated_Report.xlsx" 
    if os.path.exists(report_file):
        analyzer.analyze_report(report_file, "AI_Analysis_Result.md")
    else:
        print(f"Report file not found at {report_file}. Please run report_generator.py first.")
