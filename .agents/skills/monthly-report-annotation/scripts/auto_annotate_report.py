# -*- coding: utf-8 -*-
"""
月度经营分析报告 Excel 批注自动化工具脚本
Auto Annotate Monthly Management Report Script
"""

import sys
import io
import openpyxl
from openpyxl.comments import Comment

def create_adaptive_comment(text, author="经营分析系统"):
    """创建自适应尺寸的 Excel 批注对象"""
    comment = Comment(text, author)
    lines = text.splitlines()
    max_len = max(len(l) for l in lines) if lines else 20
    comment.width = max(380, min(550, max_len * 15))
    comment.height = max(180, len(lines) * 19 + 30)
    return comment

def apply_comments_to_workbook(file_path, sheet_comments_dict):
    """
    批量将批注写入 Excel 工作簿
    sheet_comments_dict: {
        'SheetName': {
            'CellCoord': 'Comment Text',
            ...
        }
    }
    """
    print(f"正在加载 Excel 文件: {file_path} ...")
    wb = openpyxl.load_workbook(file_path)
    total_added = 0

    for sheet_name, cell_dict in sheet_comments_dict.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = 0
            for coord, txt in cell_dict.items():
                ws[coord].comment = create_adaptive_comment(txt)
                count += 1
            print(f"  - [{sheet_name}] 成功插入 {count} 处批注")
            total_added += count
        else:
            print(f"  - ⚠️ 警告：工作表 [{sheet_name}] 不存在，已跳过")

    print(f"正在保存工作簿 (共写入 {total_added} 处批注) ...")
    wb.save(file_path)
    print("✅ 保存完成！")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        target_file = sys.argv[1]
        print(f"目标分析文件: {target_file}")
    else:
        print("请传入目标 Excel 文件路径")
