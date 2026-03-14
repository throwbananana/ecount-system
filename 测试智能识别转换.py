# -*- coding: utf-8 -*-
"""
测试智能识别转换功能
演示如何正确使用摘要智能识别自动填充科目编码
"""

import pandas as pd
from openpyxl import load_workbook
from base_data_manager import BaseDataManager
from summary_intelligence import SummaryIntelligence

def test_conversion():
    print("=" * 70)
    print("测试智能识别转换")
    print("=" * 70)

    # 1. 加载源数据
    print("\n1. 加载源数据...")
    source_file = "其他应收-巴拿马01.xlsx"
    df = pd.read_excel(source_file, header=0)
    print(f"   源数据行数: {len(df)}")
    print(f"   源数据列: {list(df.columns)}")

    # 2. 初始化智能识别器
    print("\n2. 初始化智能识别器...")
    mgr = BaseDataManager()
    recognizer = SummaryIntelligence(mgr)
    print("   ✓ 智能识别器初始化成功")

    # 3. 加载模板
    print("\n3. 加载模板...")
    template_wb = load_workbook("Template.xlsx")
    template_ws = template_wb.active

    # 获取模板列名
    template_columns = []
    for i in range(1, template_ws.max_column + 1):
        col_name = template_ws.cell(1, i).value
        if col_name:
            template_columns.append(col_name)
    print(f"   模板列数: {len(template_columns)}")
    print(f"   模板包含: 日期、序号、科目编码、对方科目、默认账户、摘要等...")

    # 4. 转换数据
    print("\n4. 开始转换...")
    print(f"   {'源摘要':<50} {'科目编码':<10} {'部门':<10} {'金额':<15}")
    print("-" * 90)

    output_rows = []
    for idx, row in df.head(10).iterrows():  # 只转换前10行做演示
        # 获取摘要
        summary = str(row.get('摘要', '')).strip()
        if not summary or summary == 'nan':
            continue

        # 智能识别
        recognized = recognizer.recognize(summary, use_ai=False)

        # 构建输出行
        out_row = {
            '日期': row.get('发生日期', ''),
            '序号': idx + 1,
            '科目编码': recognized.get('科目编码', ''),
            '对方科目': '',  # 需要扩展识别规则
            '默认账户': '',  # 需要扩展识别规则
            '摘要': summary,
            '部门': recognized.get('部门', ''),
            '金额': recognized.get('金额', row.get('支出明细金额', '')),
            '往来单位编码': recognized.get('往来单位编码', ''),
            '往来单位名': recognized.get('往来单位名', ''),
        }

        output_rows.append(out_row)

        # 打印识别结果
        print(f"   {summary[:48]:<50} {out_row['科目编码']:<10} {out_row['部门']:<10} {str(out_row['金额']):<15}")

    # 5. 保存结果
    print(f"\n5. 保存转换结果...")
    output_df = pd.DataFrame(output_rows)
    output_file = "测试转换结果_with_智能识别.xlsx"

    # 使用模板作为基础
    output_wb = load_workbook("Template.xlsx")
    output_ws = output_wb.active

    # 清除模板中的数据行（保留表头）
    if output_ws.max_row > 1:
        output_ws.delete_rows(2, output_ws.max_row)

    # 写入数据
    for row_idx, row_data in enumerate(output_rows, start=2):
        for col_idx, col_name in enumerate(template_columns, start=1):
            value = row_data.get(col_name, '')
            output_ws.cell(row_idx, col_idx).value = value

    output_wb.save(output_file)
    print(f"   ✓ 已保存到: {output_file}")

    # 6. 统计
    print("\n6. 转换统计:")
    filled_count = sum(1 for r in output_rows if r.get('科目编码'))
    print(f"   总行数: {len(output_rows)}")
    print(f"   已填充科目编码: {filled_count} 行")
    print(f"   填充率: {filled_count/len(output_rows)*100:.1f}%")

    print("\n" + "=" * 70)
    print("✅ 测试完成！")
    print("=" * 70)
    print("\n提示：")
    print("  • 在GUI中使用时，请务必勾选'启用摘要智能识别'选项")
    print("  • 如需自动填充'对方科目'和'默认账户'，请告诉我具体的填充规则")
    print("  • 可以使用'使用AI深度识别'选项获得更准确的结果（需要联网）")

if __name__ == "__main__":
    test_conversion()
