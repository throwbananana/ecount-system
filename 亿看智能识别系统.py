# -*- coding: utf-8 -*-
"""
一般凭证 Excel 模板转换工具（带 GUI + 自动识别列匹配）

功能：
1. 读取同目录下的 Template.xlsx（一般凭证模板），自动获取表头和批注。
2. 选择任意原始 Excel 文件 + 工作表。
3. 【自动识别匹配列】：
   - 表头规范化（去空格/特殊符号/大小写）
   - 包含/被包含判断
   - difflib 相似度打分
   - 同义词映射（如“科目代码”≈“科目编码”）
4. 用户可以在 GUI 中微调映射。
5. 根据每个模板表头的规则（日期、金额、小数位、长度等）进行转换。
6. 将转换结果填入模板，保留原有表头与批注。

依赖：
    pip install pandas openpyxl
"""

import os
import sys
import calendar
import math
import difflib
import re
import json
import hashlib
import threading
from types import SimpleNamespace
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from typing import Optional, List, Dict, Any
import ast
from functools import lru_cache

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog, scrolledtext

import pandas as pd
from openpyxl import load_workbook
from shipping_module import ShippingModule
from openpyxl.styles import PatternFill
from treeview_tools import attach_treeview_tools, install_smart_restore_header

from base_data_manager import BaseDataManager
from summary_intelligence import SummaryIntelligence, DEFAULT_API_KEY
from reconciliation_module import StandardReconciler
from export_format_manager import (
    apply_export_format,
    get_active_export_format_name,
    get_export_format_names,
    load_export_formats,
    open_export_format_editor,
    open_export_format_center,
    save_export_formats,
    set_active_export_format,
)
try:
    from bank_parser import BankParser
except ImportError:
    BankParser = None
try:
    from excel_merger import ExcelMergerGUI
except ImportError:
    ExcelMergerGUI = None
try:
    from report_generator import ReportGenerator
except ImportError:
    ReportGenerator = None
try:
    from folder_processor import FolderProcessorGUI
except ImportError:
    FolderProcessorGUI = None

# 图片智能识别模块（可选导入）
try:
    from image_recognition_gui import open_image_recognition_window
    HAS_IMAGE_RECOGNITION = True
except ImportError:
    HAS_IMAGE_RECOGNITION = False

# 文档识别模块 (DANFE)（可选导入）
try:
    from danfe_recognition_gui import open_danfe_recognition_window
    HAS_DANFE_RECOGNITION = True
except ImportError:
    HAS_DANFE_RECOGNITION = False

TEMPLATE_FILE = "Template.xlsx"
APP_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(APP_DIR, "config.json")

# ======== 默认字段规则（兜底用） ========
DEFAULT_FIELD_RULES = {
    "凭证日期": {"type": "date"},
    "日期": {"type": "date"},
    "序号": {"type": "text", "max_len": 4},
    "会计凭证No.": {"type": "text", "max_len": 30},
    "摘要编码": {"type": "text", "max_len": 2},
    "摘要": {"type": "text", "max_len": 200},
    "摘要名": {"type": "text", "max_len": 200},
    "类型": {"type": "text", "max_len": 1},
    "科目编码": {"type": "text", "max_len": 8},
    "对方科目": {"type": "text", "max_len": 8},
    "默认账户": {"type": "text", "max_len": 30},
    "往来单位编码": {"type": "text", "max_len": 30},
    "往来单位名": {"type": "text", "max_len": 100},
    "金额": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
    "外币金额": {"type": "number", "max_int_len": 15, "max_decimal_len": 4},
    "汇率": {"type": "number", "max_int_len": 14, "max_decimal_len": 6},
    "部门": {"type": "text", "max_len": 50},
    # 销售出库模板新增字段
    "职员": {"type": "text", "max_len": 50},
    "发货仓库": {"type": "text", "max_len": 50},
    "交易类型": {"type": "text", "max_len": 2},
    "货币": {"type": "text", "max_len": 10},
    "收货公司": {"type": "text", "max_len": 200},
    "PEDIDO No": {"type": "text", "max_len": 200},
    "No.FACTURA": {"type": "text", "max_len": 200},
    "CONSIGNADO": {"type": "text", "max_len": 200},
    "CONDI. PAGO": {"type": "text", "max_len": 200},
    "CON DESTINO A": {"type": "text", "max_len": 200},
    "EMBARCADO VIA": {"type": "text", "max_len": 200},
    "MARCA": {"type": "text", "max_len": 200},
    "CARGOS": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
    "品目编码": {"type": "text", "max_len": 20},
    "品目名": {"type": "text", "max_len": 100},
    "规格": {"type": "text", "max_len": 100},
    "Mo.TAMA": {"type": "text", "max_len": 200},
    "COLOR": {"type": "text", "max_len": 200},
    "装数": {"type": "number", "max_int_len": 15, "max_decimal_len": 0},
    "CAJA": {"type": "number", "max_int_len": 12, "max_decimal_len": 0},
    "数量": {"type": "number", "max_int_len": 12, "max_decimal_len": 0},
    "单价": {"type": "number", "max_int_len": 12, "max_decimal_len": 6},
    "供应价": {"type": "number", "max_int_len": 12, "max_decimal_len": 2},
    "附带费用": {"type": "number", "max_int_len": 12, "max_decimal_len": 2},
    "增值税": {"type": "number", "max_int_len": 12, "max_decimal_len": 2},
    "生成生产入库": {"type": "text", "max_len": 1},
    "体积": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
    "总体积": {"type": "number", "max_int_len": 15, "max_decimal_len": 1},
    "重量": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
    "总重量": {"type": "number", "max_int_len": 15, "max_decimal_len": 2},
}

# ======== 默认同义词（兜底用） ========
DEFAULT_FIELD_SYNONYMS = {
    "凭证日期": ["日期", "记账日期", "制单日期"],
    "日期": ["日期", "fecha", "fecha日期", "日期fecha", "日期 fecha", "fecha 日期"],
    "序号": ["行号", "行次", "序号", "订单号", "单据号", "no.pedi", "pedido", "No.PEDI."],
    "会计凭证No.": ["凭证号", "凭证编号", "会计凭证号"],
    "摘要编码": ["摘要代码", "摘要编码"],
    "摘要": ["摘要", "摘要名", "摘要说明", "说明", "备注", "descrip", "内容", "descripción", "DESCRIPCION", "描述", "description"],
    "摘要名": ["摘要名", "摘要", "摘要说明", "说明", "备注", "descrip", "内容", "描述"],
    "类型": ["借贷标志", "借贷", "方向", "类型"],
    "科目编码": ["科目代码", "科目编号", "会计科目", "科目", "code", "codig", "codigo"],
    "对方科目": ["对方科目", "对应科目", "反方科目", "贷方科目", "借方科目"],
    "默认账户": ["默认账户", "账户", "账户编码", "银行账户", "账号"],
    "往来单位编码": ["往来单位代码", "客户编码", "供应商编码", "往来编码", "客戶 cliente", "cliente", "客户"],
    "往来单位名": ["往来单位名称", "客户名称", "供应商名称", "单位名称", "cliente", "客戶"],
    "金额": ["金额", "本币金额", "发生额", "金额1", "合计", "total", "本位币金额", "本地金额", "计价金额"],
    "外币金额": ["外币金额", "原币金额", "合计", "total", "合计 total"],
    "汇率": ["汇率", "折算汇率"],
    "部门": ["部门", "部门名称", "成本中心"],
    # 销售出库模板补充
    "发货仓库": ["发货仓库", "出货仓库", "仓库"],
    "品目编码": ["品目编码", "物料编码", "商品编码", "代号", "代號", "codig", "代號 codigo", "代號 codig"],
    "品目名": ["品目名", "品目名称", "商品名称", "内容", "descrip", "descripcion", "descripción"],
    "数量": ["数量", "数量 canti", "canti", "qty"],
    "单价": ["单价", "价格", "precio", "价格 precio"],
    "供应价": ["供应价", "合计", "total", "合计 total", "importe"],
    "交易类型": ["交易类型", "交易编码", "单据类型"],
    "货币": ["货币", "币别", "currency", "货币代码"],
    "收货公司": ["收货公司", "收货方", "cliente", "客户"],
    "PEDIDO No": ["pedido", "No.PEDI.", "订单号", "pedido no"],
    "No.FACTURA": ["factura", "发票号"],
    "CONSIGNADO": ["consignado"],
    "CONDI. PAGO": ["支付条件", "付款条件", "pago"],
    "CON DESTINO A": ["destino"],
    "EMBARCADO VIA": ["embarcado", "via"],
    "MARCA": ["marca", "品牌"],
    "装数": ["装数", "装箱数"],
    "CAJA": ["caja", "箱数"],
    "附带费用": ["附加费用", "费用", "运费"],
    "增值税": ["增值税", "税额", "营业税"],
}

def load_config():
    """尝试从 config.json 加载配置，失败则返回默认值。"""
    rules = DEFAULT_FIELD_RULES
    synonyms = DEFAULT_FIELD_SYNONYMS
    header_schemes = {}
    
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                if "FIELD_RULES" in config:
                    rules = config["FIELD_RULES"]
                if "FIELD_SYNONYMS" in config:
                    synonyms = config["FIELD_SYNONYMS"]
                if "HEADER_SCHEMES" in config:
                    header_schemes = config["HEADER_SCHEMES"]
                print(f"已加载配置文件：{CONFIG_FILE}")
        except Exception as e:
            print(f"加载配置文件失败，使用默认配置。错误：{e}")
    else:
        print("配置文件不存在，使用默认配置。")
        
    return rules, synonyms, header_schemes

# 初始化全局配置
FIELD_RULES, FIELD_SYNONYMS, HEADER_SCHEMES = load_config()

def save_header_schemes(schemes: dict):
    config = {}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
        except Exception:
            config = {}
    config["HEADER_SCHEMES"] = schemes
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=4)

# 转换模式
MODE_GENERAL_VOUCHER = "通用凭证模式"
MODE_SALES_OUTBOUND = "销售出库模式"
MODE_CUSTOM = "自定义模式"
MODE_ORIGINAL = "原格式模式(不使用模板)"

# 通用凭证模式的字段列表 (用于 UI 过滤)
GENERAL_VOUCHER_FIELDS = {
    "凭证日期", "日期", "序号", "会计凭证No.", "摘要编码", "摘要",
    "类型", "科目编码", "往来单位编码", "往来单位名", "金额",
    "外币金额", "汇率", "部门", "对方科目", "默认账户"
}

# 针对销售出库格式的快捷映射候选（用于缺省映射兜底）
SALES_OUTBOUND_MAPPING_CANDIDATES = {
    "日期": ["日期 FECHA", "fecha", "日期"],
    "序号": ["No.PEDI.", "pedido", "订单号", "序号"],
    "PEDIDO No": ["No.PEDI.", "pedido", "订单号"],
    "往来单位编码": ["客戶 CLIENTE", "cliente", "客户编码", "客户"],
    "往来单位名": ["客戶 CLIENTE", "cliente", "客户名称"],
    "收货公司": ["客戶 CLIENTE", "cliente"],
    "品目编码": ["代號 CODIGO", "codig", "代号", "品目编码"],
    "品目名": ["内容 DESCRIPCION", "descrip", "内容", "商品名称"],
    "数量": ["数量 CANTI", "canti", "数量"],
    "单价": ["价格 PRECIO", "precio", "价格", "单价"],
    "供应价": ["合计 TOTAL", "total", "合计"],
    "外币金额": ["合计 TOTAL", "total", "合计"],
    "金额": ["合计 TOTAL", "total", "合计"],
    "摘要": ["内容 DESCRIPCION", "descrip", "摘要"],
}


class TemplateHeader:
    def __init__(self, name, col_idx, comment=""):
        self.name = name
        self.col_idx = col_idx
        self.comment = comment or ""


def load_template_headers(template_path=TEMPLATE_FILE):
    """从模板文件中读取表头和批注。"""
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"未找到模板文件：{template_path}")

    wb = load_workbook(template_path, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(
                TemplateHeader(
                    str(cell.value).strip(),
                    cell.column,
                    cell.comment.text if cell.comment else "",
                )
            )
    return headers, wb, ws


def safe_str(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def format_date(value):
    """转成 YYYYMMDD 字符串格式，若失败则返回空字符串。"""
    if value is None or (isinstance(value, float) and math.isnan(value)) or safe_str(value) == "":
        return ""

    if isinstance(value, (datetime, date)):
        return value.strftime("%Y%m%d")

    s = safe_str(value)

    # 1. 尝试多种字符串格式解析
    formats = (
        "%Y%m%d", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", 
        "%m/%d/%Y", "%m/%d/%y", 
        "%d/%m/%Y", "%d/%m/%y", 
        "%m-%d-%Y", "%m-%d-%y",
        "%d-%m-%Y", "%d-%m-%y",
        "%d-%b-%Y", "%d-%b-%y"
    )
    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            # 处理 2 位年份 (例如 25 -> 2025)
            if dt.year < 100:
                dt = dt.replace(year=2000 + dt.year)
            return dt.strftime("%Y%m%d")
        except ValueError:
            continue

    # 2. 尝试解析 Excel 序列号
    try:
        if s.isdigit() and len(s) == 8:
            dt = datetime.strptime(s, "%Y%m%d")
            return dt.strftime("%Y%m%d")
            
        serial = float(s)
        if 1 <= serial <= 100000:
            base = datetime(1899, 12, 30)
            dt = base + timedelta(days=serial)
            return dt.strftime("%Y%m%d")
    except Exception:
        pass

    # 3. 兜底解析：纯数字尝试
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 8:
        try:
            dt = datetime.strptime(digits, "%Y%m%d")
            return dt.strftime("%Y%m%d")
        except ValueError:
            pass

    return ""


def format_number(value, max_int_len=15, max_decimal_len=2):
    """根据整数位/小数位限制格式化数字。"""
    if value is None or (isinstance(value, float) and math.isnan(value)) or safe_str(value) == "":
        return ""

    s = safe_str(value).replace(",", "")

    try:
        d = Decimal(s)
    except InvalidOperation:
        return s

    if max_decimal_len > 0:
        quantize_str = "1." + ("0" * max_decimal_len)
        d = d.quantize(Decimal(quantize_str))
    else:
        d = d.quantize(Decimal("1"))

    s_full = format(d, "f")
    if "." in s_full:
        int_part, dec_part = s_full.split(".")
    else:
        int_part, dec_part = s_full, ""

    negative = int_part.startswith("-")
    int_digits = int_part.lstrip("-")

    if len(int_digits) > max_int_len:
        int_digits = int_digits[-max_int_len:]

    int_part = ("-" if negative else "") + int_digits

    if max_decimal_len == 0:
        return int_part

    if len(dec_part) > max_decimal_len:
        dec_part = dec_part[:max_decimal_len]

    return int_part + "." + dec_part if dec_part else int_part


def apply_text_length(value, max_len=None):
    s = safe_str(value)
    if max_len and len(s) > max_len:
        return s[:max_len]
    return s


def convert_value(header_name, src_value, custom_format=None):
    """根据字段规则把原始值转换成模板要求格式。"""
    rule = FIELD_RULES.get(header_name, {"type": "text"})
    t = rule.get("type", "text")

    # 应用自定义格式覆盖
    if custom_format:
        if "decimal" in custom_format and (t == "number" or header_name in ["金额", "外币金额", "汇率"]):
            return format_number(
                src_value,
                max_int_len=rule.get("max_int_len", 15),
                max_decimal_len=custom_format["decimal"],
            )
        if "max_len" in custom_format:
            return apply_text_length(src_value, custom_format["max_len"])

    if t == "date":
        return format_date(src_value)
    if t == "number":
        return format_number(
            src_value,
            max_int_len=rule.get("max_int_len", 15),
            max_decimal_len=rule.get("max_decimal_len", 2),
        )
    return apply_text_length(src_value, rule.get("max_len"))


# ========== 自动匹配相关工具函数 ==========

EMPTY_OPTION = "<留空>"
CUSTOM_INPUT_OPTION = "<自定义输入...>"
DEFAULT_VALUE_ALIASES = {
    "默认部门": "部门",
    "默认科目": "科目编码",
    "默认科目编码": "科目编码",
    "账户": "默认账户",
    "默认账户编码": "默认账户",
    "默认货币": "货币",
    "默认仓库": "默认仓库",
}
COMPOSITE_PREFIX = "[综合]"
BUILTIN_COMPOSITE_FIELDS = {
    "借贷列辅助": {
        "type": "debit_credit",
        "description": "借方/贷方合并成带符号金额，借为正数，贷为负数",
    }
}


def normalize_header(s: str) -> str:
    """表头规范化：去空格、符号、大小写、全角半角等。"""
    if s is None:
        return ""
    s = str(s)
    # 全角转半角
    def _dbc_to_sbc(ch):
        code = ord(ch)
        if code == 0x3000:
            return " "
        if 0xFF01 <= code <= 0xFF5E:
            return chr(code - 0xFEE0)
        return ch

    s = "".join(_dbc_to_sbc(ch) for ch in s)
    # 去掉常见的分隔符号
    s = re.sub(r"[\s\-\_/（）\(\)【】\[\]：:，,\.。]", "", s)
    return s.lower()


def score_similarity(template_name: str, src_name: str, template_key: str) -> float:
    """
    给某个模板表头 & 源表头 的匹配打分：
    1. 规范名是否完全相等
    2. 包含关系
    3. 同义词命中
    4. difflib 模糊匹配
    综合出一个 0~1 分数。
    """
    raw_t = template_name
    raw_s = src_name
    n_t = normalize_header(raw_t)
    n_s = normalize_header(raw_s)

    # 完全相等
    if n_t and n_t == n_s:
        return 1.0

    score = 0.0

    # 包含关系
    if n_t and n_s:
        if n_t in n_s or n_s in n_t:
            score = max(score, 0.85)

    # 同义词匹配
    syns = FIELD_SYNONYMS.get(template_key, [])
    for syn in syns:
        n_syn = normalize_header(syn)
        if n_syn and (n_syn == n_s or n_syn in n_s or n_s in n_syn):
            score = max(score, 0.9)

    # difflib 相似度
    ratio = difflib.SequenceMatcher(None, n_t, n_s).ratio()
    score = max(score, ratio * 0.9)  # 稍微降一点权重

    return score


# --- 安全表达式求值（用于综合字段表达式） ---
ALLOWED_AST_NODES = (
    ast.Expression, ast.BinOp, ast.UnaryOp, ast.Num, ast.Str, ast.Constant,
    ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Mod, ast.Pow, ast.Load,
    ast.Call, ast.Name, ast.Attribute, ast.Tuple, ast.List, ast.Dict,
    ast.Compare, ast.Eq, ast.NotEq, ast.Gt, ast.GtE, ast.Lt, ast.LtE,
    ast.BoolOp, ast.And, ast.Or, ast.IfExp, ast.Subscript, ast.Index, ast.Slice,
)
ALLOWED_CALL_NAMES = {"col", "field", "mapped", "composite", "lookup", "abs", "round", "min", "max", "Decimal", "float", "int", "str"}
MAX_COMPOSITE_EXPR_LEN = 1000
MAX_COMPOSITE_AST_NODES = 200
MAX_COMPOSITE_DEPTH = 20


@lru_cache(maxsize=256)
def _compile_safe_expr(expr: str):
    """编译并缓存安全表达式，限制长度与AST复杂度。"""
    if not isinstance(expr, str):
        raise ValueError("表达式必须是字符串")
    if not expr.strip():
        raise ValueError("表达式不能为空")
    if len(expr) > MAX_COMPOSITE_EXPR_LEN:
        raise ValueError(f"表达式过长(>{MAX_COMPOSITE_EXPR_LEN})")

    try:
        tree = ast.parse(expr, mode="eval")
    except SyntaxError as e:
        raise ValueError(f"表达式语法错误: {e}")

    node_count = 0
    for node in ast.walk(tree):
        node_count += 1
        if node_count > MAX_COMPOSITE_AST_NODES:
            raise ValueError("表达式过于复杂，请简化后再试")
        if not isinstance(node, ALLOWED_AST_NODES):
            raise ValueError(f"不支持的表达式元素: {type(node).__name__}")
        if isinstance(node, ast.Call):
            if isinstance(node.func, ast.Name):
                fn = node.func.id
            elif isinstance(node.func, ast.Attribute):
                # 禁止对象方法调用，避免逃逸
                raise ValueError("不支持对象方法调用")
            else:
                raise ValueError("不支持的调用形式")
            if fn not in ALLOWED_CALL_NAMES:
                raise ValueError(f"不允许调用函数: {fn}")
        if isinstance(node, ast.Attribute):
            # 禁止属性访问，避免 __getattr__ 等
            raise ValueError("不支持属性访问")

    return compile(tree, "<composite_expr>", "eval")


def safe_eval_expr(expr: str, env: Dict[str, Any]):
    """安全地执行用户表达式，只允许白名单节点/函数。"""
    code = _compile_safe_expr(expr)
    return eval(code, {"__builtins__": {}}, env)


class ExcelConverterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("一般凭证 Excel 格式转换工具（含自动识别列匹配）")

        self.template_headers = []
        self.template_workbook = None
        self.template_sheet = None

        self.input_path = ""
        self.input_df = None
        self.input_columns = []
        self.mapping_vars = {}
        self.custom_composite_fields = {}
        self._skip_auto_mapping_once = False

        self.sheet_var = tk.StringVar()
        self.input_path_var = tk.StringVar()
        
        default_template = TEMPLATE_FILE
        if os.path.exists("Template_通用凭证.xlsx"):
             default_template = "Template_通用凭证.xlsx"
        self.template_path_var = tk.StringVar(value=default_template)
        
        self.convert_mode_var = tk.StringVar(value=MODE_GENERAL_VOUCHER)
        self.header_scheme_var = tk.StringVar(value="")
        self.export_scheme_override_var = tk.BooleanVar(value=True)
        self.export_format_var = tk.StringVar(value=get_active_export_format_name("main_export"))
        self.debug_var = tk.BooleanVar(value=False)
        self.search_field_var = tk.StringVar(value="")
        self.current_columns = []
        self.sort_states = {}
        self._prefer_prompt_default_account = False

        self.manual_debit_col_var = tk.StringVar(value=EMPTY_OPTION)
        self.manual_credit_col_var = tk.StringVar(value=EMPTY_OPTION)
        self.manual_dc_col_var = tk.StringVar(value=EMPTY_OPTION) # 新增：借贷标志列 (辅助)
        self.field_formats = {}  # 新增：字段格式自定义 (如：{"金额": {"decimal": 2}})

        # 高级选项变量（提前初始化供菜单使用）
        self.enable_smart_recognition = tk.BooleanVar(value=True)
        self.use_ai_var = tk.BooleanVar(value=False)
        self.use_foreign_currency_var = tk.BooleanVar(value=False)
        self.auto_balance_var = tk.BooleanVar(value=False)
        self.split_amount_var = tk.BooleanVar(value=False)
        
        # 对账高级选项
        self.recon_fuzzy_var = tk.BooleanVar(value=True)
        self.recon_use_ai_analysis = tk.BooleanVar(value=False)
        self.invert_bank_var = tk.BooleanVar(value=False)
        self.recon_export_voucher_var = tk.BooleanVar(value=True) # Default True for convenience

        # 摘要匹配模块
        self.summary_match_target_path_var = tk.StringVar()
        self.summary_match_target_sheet_var = tk.StringVar()
        self.summary_match_source_path_var = tk.StringVar()
        self.summary_match_source_sheet_var = tk.StringVar()
        self.summary_match_target_df = None
        self.summary_match_source_df = None
        self.summary_match_target_columns = []
        self.summary_match_source_columns = []
        self.summary_match_target_vars = {
            "summary": tk.StringVar(),
            "date": tk.StringVar(),
            "amount": tk.StringVar(),
            "debit": tk.StringVar(),
            "credit": tk.StringVar(),
        }
        self.summary_match_source_vars = {
            "summary": tk.StringVar(),
            "date": tk.StringVar(),
            "amount": tk.StringVar(),
            "debit": tk.StringVar(),
            "credit": tk.StringVar(),
        }
        self.summary_match_export_format_var = tk.StringVar(
            value=get_active_export_format_name("summary_match_export")
        )
        self.summary_match_date_tol_var = tk.StringVar(value="3")
        self.summary_match_amount_abs_tol_var = tk.StringVar(value="0.01")
        self.summary_match_amount_pct_tol_var = tk.StringVar(value="1")
        self.summary_match_keep_original_var = tk.BooleanVar(value=True)
        self.summary_match_unique_var = tk.BooleanVar(value=True)
        self.summary_match_sync_direction_var = tk.BooleanVar(value=False) # 新增：同步借贷方向/正负号
        self.summary_match_fill_missing_var = tk.BooleanVar(value=False)   # 新增：填充缺失记录
        self.summary_match_status_var = tk.StringVar(value="等待加载数据...")
        self.summary_match_result_df = None

        # 基础数据管理器 (先初始化，以便读取配置)
        self.base_data_mgr = None
        self._init_base_data()
        install_smart_restore_header(self.root, base_data_mgr=self.base_data_mgr)

        # 读取持久化配置
        db_configs = {}
        if self.base_data_mgr:
            db_configs = self.base_data_mgr.get_all_configs()
            
            # 从 db_configs 加载高级选项的持久化状态
            self.enable_smart_recognition.set(db_configs.get("setting_enable_smart_recognition", "True") == "True")
            self.use_ai_var.set(db_configs.get("setting_use_ai_var", "False") == "True")
            self.use_foreign_currency_var.set(db_configs.get("setting_use_foreign_currency_var", "False") == "True")
            self.auto_balance_var.set(db_configs.get("setting_auto_balance_var", "False") == "True")
            self.split_amount_var.set(db_configs.get("setting_split_amount_var", "False") == "True")
            
            # 加载对账高级选项的持久化状态
            self.recon_fuzzy_var.set(db_configs.get("setting_recon_fuzzy_var", "True") == "True")
            self.recon_use_ai_analysis.set(db_configs.get("setting_recon_use_ai_analysis", "False") == "True")
            self.invert_bank_var.set(db_configs.get("setting_invert_bank_var", "False") == "True")

            # 加载全局综合字段
            raw_comp = db_configs.get("app_global_composite_fields", "{}")
            try:
                saved_comp = json.loads(raw_comp)
                if isinstance(saved_comp, dict):
                    self.custom_composite_fields.update(saved_comp)
            except Exception as e:
                print(f"加载全局综合字段失败: {e}")

            raw_field_formats = db_configs.get("main_field_formats", "")
            if raw_field_formats:
                try:
                    parsed_formats = json.loads(raw_field_formats)
                    if isinstance(parsed_formats, dict):
                        self.field_formats = parsed_formats
                except Exception as e:
                    print(f"加载字段格式配置失败: {e}")
        self._global_field_formats = dict(self.field_formats)

        # 默认值组
        self.default_value_groups = {}
        self.active_default_group = None
        raw_groups = db_configs.get("default_value_groups", "")
        if raw_groups:
            try:
                parsed_groups = json.loads(raw_groups)
                if isinstance(parsed_groups, dict):
                    self.default_value_groups = parsed_groups
            except Exception as e:
                print(f"加载默认值组失败: {e}")
        raw_active_group = db_configs.get("default_value_group_active")
        if raw_active_group in self.default_value_groups:
            self.active_default_group = raw_active_group

        # AI 设置 (优先使用数据库中的 Key)
        self.api_key = db_configs.get("api_key", DEFAULT_API_KEY)

        # 默认值 (从数据库加载)
        # 过滤掉 api_key 和已持久化的设置，剩下的视为 default_values
        legacy_defaults = {k: v for k, v in db_configs.items() if k not in [
            "api_key",
            "setting_enable_smart_recognition",
            "setting_use_ai_var",
            "setting_use_foreign_currency_var",
            "setting_auto_balance_var",
            "setting_split_amount_var",
            "setting_recon_fuzzy_var",
            "setting_recon_use_ai_analysis",
            "setting_invert_bank_var",
            "main_field_formats",
            "default_value_groups",
            "default_value_group_active"
        ]}
        if self.default_value_groups:
            if not self.active_default_group:
                self.active_default_group = next(iter(self.default_value_groups.keys()))
            self.default_values = dict(self.default_value_groups.get(self.active_default_group, {}))
        else:
            self.default_values = legacy_defaults
            if legacy_defaults:
                self.default_value_groups = {"默认": dict(legacy_defaults)}
                self.active_default_group = "默认"
                if self.base_data_mgr:
                    self._persist_default_value_groups()

        # 摘要智能识别器
        self.summary_recognizer = None
        self._init_summary_recognizer()

        self._build_ui()
        self._load_template()
        self._ensure_default_main_export_format()
        self._on_mode_changed()

        # 在窗口关闭时调用 _on_closing
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _save_settings(self):
        """保存高级选项的当前状态到数据库"""
        if self.base_data_mgr:
            try:
                self.base_data_mgr.set_config("setting_enable_smart_recognition", str(self.enable_smart_recognition.get()))
                self.base_data_mgr.set_config("setting_use_ai_var", str(self.use_ai_var.get()))
                self.base_data_mgr.set_config("setting_use_foreign_currency_var", str(self.use_foreign_currency_var.get()))
                self.base_data_mgr.set_config("setting_auto_balance_var", str(self.auto_balance_var.get()))
                self.base_data_mgr.set_config("setting_split_amount_var", str(self.split_amount_var.get()))
                
                # 对账选项
                self.base_data_mgr.set_config("setting_recon_fuzzy_var", str(self.recon_fuzzy_var.get()))
                self.base_data_mgr.set_config("setting_recon_use_ai_analysis", str(self.recon_use_ai_analysis.get()))
                self.base_data_mgr.set_config("setting_invert_bank_var", str(self.invert_bank_var.get()))
                
                self.log_message("高级选项设置已保存。")
            except Exception as e:
                self.log_message(f"保存高级选项设置失败: {e}")

    def _persist_field_formats(self):
        if not self.base_data_mgr:
            return
        try:
            payload = json.dumps(self.field_formats, ensure_ascii=False)
            self.base_data_mgr.set_config("main_field_formats", payload)
            self._global_field_formats = dict(self.field_formats)
        except Exception as e:
            self.log_message(f"保存字段格式配置失败: {e}")

    def _on_closing(self):
        """处理窗口关闭事件，保存设置并退出"""
        self._save_settings()
        self.root.destroy()

    def _show_ai_settings(self):
        """显示 AI 设置对话框 (新版: 多模型配置 + 任务分发)"""
        dialog = tk.Toplevel(self.root)
        dialog.title("AI 智能配置中心")
        dialog.geometry("700x550")
        
        # 读取配置
        db_configs = self.base_data_mgr.get_all_configs() if self.base_data_mgr else {}
        
        # 1. 解析后端列表 (JSON存储)
        # 格式: {"name": "Zhipu-Main", "provider": "zhipu", "api_key": "...", "model": "..."}
        backends_json = db_configs.get("ai_backends", "[]")
        try:
            backends = json.loads(backends_json)
            if not isinstance(backends, list): backends = []
        except: backends = []

        # 如果为空，尝试迁移旧配置
        if not backends:
            old_provider = db_configs.get("ai_provider", "zhipu")
            old_key = db_configs.get("ai_api_key", self.api_key)
            old_url = db_configs.get("ai_base_url", "http://localhost:1234/v1")
            old_model = db_configs.get("ai_model_name", "local-model")
            
            if old_provider == "zhipu" and old_key:
                backends.append({"name": "默认智谱", "provider": "zhipu", "api_key": old_key, "model": "glm-4-flash"})
            elif old_provider == "lm_studio":
                backends.append({"name": "默认本地", "provider": "lm_studio", "base_url": old_url, "model": old_model})

        # 2. 解析任务分配 (JSON存储)
        # 格式: {"summary_rec": "默认智谱", "formula_gen": "默认本地", ...}
        task_map_json = db_configs.get("ai_task_map", "{}")
        try:
            task_map = json.loads(task_map_json)
        except: task_map = {}

        # 界面布局
        notebook = ttk.Notebook(dialog)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        # === Tab 1: 模型源配置 ===
        tab_models = ttk.Frame(notebook)
        notebook.add(tab_models, text="模型源配置")

        # 左侧列表
        m_paned = ttk.PanedWindow(tab_models, orient="horizontal")
        m_paned.pack(fill="both", expand=True, padx=5, pady=5)
        
        m_left = ttk.Frame(m_paned, width=150)
        m_paned.add(m_left, weight=1)
        
        m_list = tk.Listbox(m_left, exportselection=False)
        m_list.pack(fill="both", expand=True)
        
        m_btns = ttk.Frame(m_left)
        m_btns.pack(fill="x")
        
        # 右侧表单
        m_right = ttk.LabelFrame(m_paned, text="模型详情", padding=10)
        m_paned.add(m_right, weight=3)

        # 表单变量
        var_name = tk.StringVar()
        var_provider = tk.StringVar()
        var_key = tk.StringVar()
        var_url = tk.StringVar()
        var_model = tk.StringVar()

        current_idx = None

        def clear_form():
            nonlocal current_idx
            var_name.set("")
            var_provider.set("zhipu")
            var_key.set("")
            var_url.set("http://localhost:1234/v1")
            var_model.set("")
            current_idx = None
            try:
                m_list.selection_clear(0, "end")
            except Exception:
                pass

        def load_form(evt):
            nonlocal current_idx
            sel = m_list.curselection()
            if not sel: return
            idx = sel[0]
            current_idx = idx
            data = backends[idx]
            var_name.set(data.get("name", ""))
            var_provider.set(data.get("provider", "zhipu"))
            var_key.set(data.get("api_key", ""))
            var_url.set(data.get("base_url", ""))
            var_model.set(data.get("model", ""))
            _toggle_fields()

        m_list.bind("<<ListboxSelect>>", load_form)

        ttk.Label(m_right, text="名称:").grid(row=0, column=0, sticky="e", pady=2)
        ttk.Entry(m_right, textvariable=var_name).grid(row=0, column=1, sticky="ew", pady=2)

        ttk.Label(m_right, text="类型:").grid(row=1, column=0, sticky="e", pady=2)
        p_frame = ttk.Frame(m_right)
        p_frame.grid(row=1, column=1, sticky="w")
        
        def _toggle_fields():
            if var_provider.get() == "zhipu":
                url_entry.config(state="disabled")
            else:
                url_entry.config(state="normal")

        ttk.Radiobutton(p_frame, text="智谱AI", variable=var_provider, value="zhipu", command=_toggle_fields).pack(side="left")
        ttk.Radiobutton(p_frame, text="LM Studio/OpenAI", variable=var_provider, value="lm_studio", command=_toggle_fields).pack(side="left")

        ttk.Label(m_right, text="API Key:").grid(row=2, column=0, sticky="e", pady=2)
        ttk.Entry(m_right, textvariable=var_key).grid(row=2, column=1, sticky="ew", pady=2)

        ttk.Label(m_right, text="Base URL:").grid(row=3, column=0, sticky="e", pady=2)
        url_entry = ttk.Entry(m_right, textvariable=var_url)
        url_entry.grid(row=3, column=1, sticky="ew", pady=2)

        ttk.Label(m_right, text="Model:").grid(row=4, column=0, sticky="e", pady=2)
        ttk.Entry(m_right, textvariable=var_model).grid(row=4, column=1, sticky="ew", pady=2)

        test_btn = ttk.Button(m_right, text="测试连接")
        test_btn.grid(row=5, column=1, sticky="e", pady=(6, 0))
        
        m_right.columnconfigure(1, weight=1)
        
        def refresh_list():
            m_list.delete(0, "end")
            for b in backends:
                name = b.get("name", "未命名")
                provider = b.get("provider", "")
                model = b.get("model", "")
                suffix = f"{provider}/{model}" if provider or model else ""
                label = f"{name} ({suffix})" if suffix else name
                m_list.insert("end", label)

        def persist_backends():
            if self.base_data_mgr:
                self.base_data_mgr.set_config("ai_backends", json.dumps(backends, ensure_ascii=False))

        def persist_task_map():
            if self.base_data_mgr:
                self.base_data_mgr.set_config("ai_task_map", json.dumps(task_map, ensure_ascii=False))

        def refresh_task_values():
            names = get_backend_names()
            for cb in combos:
                cb['values'] = names
            for tid, tv in task_vars.items():
                if tv.get() not in names:
                    tv.set("(未配置)")

        def _unique_backend_name(name, exclude_idx=None):
            existing = {b.get("name") for i, b in enumerate(backends) if i != exclude_idx}
            if name not in existing:
                return name
            base = name or "模型"
            i = 2
            candidate = f"{base}-{i}"
            while candidate in existing:
                i += 1
                candidate = f"{base}-{i}"
            return candidate

        def _get_selected_index():
            sel = m_list.curselection()
            if sel:
                return sel[0]
            return current_idx

        def add_backend():
            name = var_name.get().strip() or f"模型-{len(backends)+1}"
            name = _unique_backend_name(name)
            new_b = {
                "name": name,
                "provider": var_provider.get(),
                "api_key": var_key.get().strip(),
                "base_url": var_url.get().strip(),
                "model": var_model.get().strip()
            }
            backends.append(new_b)
            refresh_list()
            clear_form()
            persist_backends()
            refresh_task_values()

        def _run_test_on_thread(cfg: Dict[str, str]):
            def _post_to_ui(fn):
                """跨线程安全地投递 UI 回调；窗口已销毁时静默忽略。"""
                for widget in (dialog, self.root):
                    try:
                        if widget and widget.winfo_exists():
                            widget.after(0, fn)
                            return
                    except Exception:
                        continue

            def _safe_restore_test_btn():
                try:
                    if test_btn and test_btn.winfo_exists():
                        test_btn.config(state="normal")
                except Exception:
                    pass

            def _finish_ok(msg: str):
                _safe_restore_test_btn()
                try:
                    if dialog and dialog.winfo_exists():
                        messagebox.showinfo("测试成功", msg, parent=dialog)
                except Exception:
                    pass

            def _finish_err(msg: str):
                _safe_restore_test_btn()
                try:
                    if dialog and dialog.winfo_exists():
                        messagebox.showerror("测试失败", msg, parent=dialog)
                except Exception:
                    pass

            provider = (cfg.get("provider") or "zhipu").strip()
            api_key = cfg.get("api_key") or ""
            base_url = cfg.get("base_url") or "http://localhost:1234/v1"
            model = cfg.get("model") or ("glm-4-flash" if provider == "zhipu" else "local-model")

            if provider == "zhipu" and not api_key:
                api_key = os.environ.get("YIKAN_AI_API_KEY") or os.environ.get("ZHIPU_API_KEY") or ""
            if provider == "zhipu" and not api_key:
                _post_to_ui(lambda: _finish_err("未配置智谱 API Key"))
                return

            try:
                if provider == "zhipu":
                    from zhipuai import ZhipuAI
                    client = ZhipuAI(api_key=api_key)
                    resp = client.chat.completions.create(
                        model=model,
                        messages=[{"role": "user", "content": "ping"}],
                        temperature=0.1
                    )
                    content = resp.choices[0].message.content
                    _post_to_ui(lambda: _finish_ok(f"返回内容: {str(content)[:120]}"))
                else:
                    from openai import OpenAI
                    client = OpenAI(base_url=base_url, api_key=api_key or "lm-studio")
                    resp = client.chat.completions.create(
                        model=model,
                        messages=[{"role": "user", "content": "ping"}],
                        temperature=0.1,
                        max_tokens=20
                    )
                    content = resp.choices[0].message.content
                    _post_to_ui(lambda: _finish_ok(f"返回内容: {str(content)[:120]}"))
            except Exception as e:
                _post_to_ui(lambda err=str(e): _finish_err(err))

        def test_backend():
            cfg = {
                "provider": var_provider.get(),
                "api_key": var_key.get().strip(),
                "base_url": var_url.get().strip(),
                "model": var_model.get().strip(),
            }
            test_btn.config(state="disabled")
            threading.Thread(target=_run_test_on_thread, args=(cfg,), daemon=True).start()

        test_btn.config(command=test_backend)

        def update_backend():
            idx = _get_selected_index()
            if idx is None:
                return
            old_name = backends[idx].get("name", "")
            new_name = var_name.get().strip() or old_name or f"模型-{idx+1}"
            new_name = _unique_backend_name(new_name, exclude_idx=idx)
            backends[idx] = {
                "name": new_name,
                "provider": var_provider.get(),
                "api_key": var_key.get().strip(),
                "base_url": var_url.get().strip(),
                "model": var_model.get().strip()
            }
            refresh_list()
            try:
                m_list.selection_set(idx)
                m_list.activate(idx)
            except Exception:
                pass
            # 同步任务映射里使用的名称
            if new_name != old_name:
                for tid, tv in task_vars.items():
                    if tv.get() == old_name:
                        tv.set(new_name)
                if task_map.get("smart_summary") == old_name:
                    task_map["smart_summary"] = new_name
                for tid, val in list(task_map.items()):
                    if val == old_name:
                        task_map[tid] = new_name
                persist_task_map()
            persist_backends()
            refresh_task_values()
            
        def del_backend():
            idx = _get_selected_index()
            if idx is None:
                return
            removed_name = backends[idx].get("name", "")
            del backends[idx]
            refresh_list()
            clear_form()
            if removed_name:
                for tid, tv in task_vars.items():
                    if tv.get() == removed_name:
                        tv.set("(未配置)")
                for tid, val in list(task_map.items()):
                    if val == removed_name:
                        task_map[tid] = "(未配置)"
                persist_task_map()
            persist_backends()
            refresh_task_values()

        ttk.Button(m_btns, text="新增", command=add_backend).pack(side="left", fill="x", expand=True)
        ttk.Button(m_btns, text="保存修改", command=update_backend).pack(side="left", fill="x", expand=True)
        ttk.Button(m_btns, text="删除", command=del_backend).pack(side="left", fill="x", expand=True)
        
        refresh_list()

        # === Tab 2: 任务分发 ===
        tab_tasks = ttk.Frame(notebook)
        notebook.add(tab_tasks, text="功能模块分配")

        tasks = [
            ("smart_summary", "摘要智能识别", "用于自动识别科目编码、摘要内容"),
            ("formula_gen", "AI 公式生成", "用于综合字段中根据描述生成表达式"),
            ("image_rec", "图片智能识别", "用于识别单据图片内容"),
            ("reconciliation", "对账智能分析", "用于分析对账差异原因"),
            ("report_analysis", "报告智能解读", "用于对经营报告/对账报告进行AI解读"),
        ]
        
        task_vars = {}
        
        row = 0
        ttk.Label(tab_tasks, text="功能模块", font=("", 10, "bold")).grid(row=row, column=0, padx=10, pady=5, sticky="w")
        ttk.Label(tab_tasks, text="选择使用的模型", font=("", 10, "bold")).grid(row=row, column=1, padx=10, pady=5, sticky="w")
        row += 1
        
        def get_backend_names():
            return ["(未配置)"] + [b["name"] for b in backends]

        combos = []

        for tid, tname, tdesc in tasks:
            ttk.Label(tab_tasks, text=tname).grid(row=row, column=0, padx=10, pady=2, sticky="nw")
            ttk.Label(tab_tasks, text=tdesc, foreground="gray", font=("", 8)).grid(row=row+1, column=0, padx=10, pady=(0, 5), sticky="w")
            
            tv = tk.StringVar(value=task_map.get(tid, "(未配置)"))
            cb = ttk.Combobox(tab_tasks, textvariable=tv, values=get_backend_names(), state="readonly", width=30)
            cb.grid(row=row, column=1, rowspan=2, padx=10, pady=5)
            
            # 下拉框获得焦点时刷新列表，防止新建模型后这里没更新
            def _refresh_cb_values(event, c=cb):
                c['values'] = get_backend_names()
            cb.bind("<Button-1>", _refresh_cb_values)
            
            task_vars[tid] = tv
            combos.append(cb)
            row += 2

        # 保存所有配置
        def save_all():
            # 先把当前表单的修改写回 backends（避免只点“保存所有配置”导致改动丢失）
            update_backend()
            # 1. 保存后端
            if self.base_data_mgr:
                self.base_data_mgr.set_config("ai_backends", json.dumps(backends))
            
            # 2. 保存任务映射
            new_map = {tid: tv.get() for tid, tv in task_vars.items()}
            if self.base_data_mgr:
                self.base_data_mgr.set_config("ai_task_map", json.dumps(new_map))
                
                # 兼容旧逻辑：如果摘要识别选了某个智谱模型，更新旧的 key
                summary_backend_name = new_map.get("smart_summary")
                found = next((b for b in backends if b["name"] == summary_backend_name), None)
                if found:
                    self.base_data_mgr.set_config("ai_provider", found["provider"])
                    self.base_data_mgr.set_config("ai_api_key", found.get("api_key", ""))
                    self.base_data_mgr.set_config("api_key", found.get("api_key", ""))
                    self.base_data_mgr.set_config("ai_base_url", found.get("base_url", ""))
                    self.base_data_mgr.set_config("ai_model_name", found.get("model", ""))
                    
                    # 更新运行时识别器
                    if self.summary_recognizer:
                        self.summary_recognizer.update_config(
                            provider=found["provider"],
                            api_key=found.get("api_key", ""),
                            base_url=found.get("base_url", ""),
                            model_name=found.get("model", "")
                        )

            messagebox.showinfo("成功", "AI 配置已保存")
            dialog.destroy()

        ttk.Button(dialog, text="保存所有配置", command=save_all).pack(pady=10)

    def _get_ai_backend_for_task(self, task_type: str, allow_legacy: bool = True) -> Optional[Dict[str, Any]]:
        if not task_type:
            return None
        configs = self.base_data_mgr.get_all_configs() if self.base_data_mgr else {}
        try:
            task_map = json.loads(configs.get("ai_task_map", "{}"))
        except Exception:
            task_map = {}
        try:
            backends = json.loads(configs.get("ai_backends", "[]"))
        except Exception:
            backends = []

        backend_name = task_map.get(task_type)
        backend = None
        if backend_name and backend_name != "(未配置)":
            backend = next((b for b in backends if b.get("name") == backend_name), None)

        if backend:
            return backend
        if not allow_legacy:
            return None
        return {
            "name": "(legacy)",
            "provider": self.default_values.get("ai_provider", "zhipu"),
            "api_key": self.default_values.get("ai_api_key", self.api_key),
            "base_url": self.default_values.get("ai_base_url", "http://localhost:1234/v1"),
            "model": self.default_values.get("ai_model_name", "local-model"),
        }

    def _normalize_ai_backend(self, backend: Optional[Dict[str, Any]]) -> Optional[Dict[str, str]]:
        if not backend:
            return None
        provider = (backend.get("provider") or "zhipu").strip()
        api_key = backend.get("api_key") or ""
        if not api_key and provider == "zhipu":
            api_key = os.environ.get("YIKAN_AI_API_KEY") or os.environ.get("ZHIPU_API_KEY") or ""
        if not api_key and provider == "zhipu":
            api_key = os.environ.get("YIKAN_AI_API_KEY") or os.environ.get("ZHIPU_API_KEY") or ""
        base_url = backend.get("base_url") or "http://localhost:1234/v1"
        if provider == "lm_studio":
            model = backend.get("model") or "local-model"
        else:
            model = backend.get("model") or "glm-4-flash"
        return {
            "name": backend.get("name") or "",
            "provider": provider,
            "api_key": api_key,
            "base_url": base_url,
            "model": model,
        }

    def _build_ai_context(self, task_type: str) -> Optional[Any]:
        backend = self._normalize_ai_backend(self._get_ai_backend_for_task(task_type, allow_legacy=True))
        if not backend:
            return None
        provider = backend["provider"]
        api_key = backend["api_key"]
        base_url = backend["base_url"]
        model = backend["model"]
        client = None

        if provider == "lm_studio":
            try:
                from openai import OpenAI
                client = OpenAI(base_url=base_url, api_key=api_key or "lm-studio")
            except Exception as exc:
                self.log_message(f"LM Studio 客户端初始化失败: {exc}")
                return None
        elif provider == "zhipu":
            try:
                from zhipuai import ZhipuAI
                if not api_key:
                    self.log_message("未配置智谱 API Key，AI 分析不可用。")
                    return None
                client = ZhipuAI(api_key=api_key)
            except Exception as exc:
                self.log_message(f"智谱AI 客户端初始化失败: {exc}")
                return None
        else:
            self.log_message(f"未知 AI Provider: {provider}")
            return None

        return SimpleNamespace(
            ai_client=client,
            ai_model_name=model,
            ai_base_url=base_url,
            ai_provider=provider,
        )

    def call_ai_service(self, task_type: str, prompt: str) -> Optional[str]:
        """统一 AI 调用接口"""
        if not self.base_data_mgr:
            return None

        configs = self.base_data_mgr.get_all_configs()
        try:
            task_map = json.loads(configs.get("ai_task_map", "{}"))
            if not isinstance(task_map, dict):
                task_map = {}
        except Exception:
            task_map = {}
        try:
            backends = json.loads(configs.get("ai_backends", "[]"))
            if not isinstance(backends, list):
                backends = []
        except Exception:
            backends = []

        backend_name = task_map.get(task_type)
        if not backend_name:
            print(f"任务 {task_type} 未分配模型")
            return None

        backend = next((b for b in backends if b["name"] == backend_name), None)
        if not backend:
            print(f"未找到模型配置: {backend_name}")
            return None

        provider = (backend.get("provider") or "zhipu").strip()
        api_key = backend.get("api_key") or ""
        base_url = backend.get("base_url") or "http://localhost:1234/v1"
        model = backend.get("model") or ("glm-4-flash" if provider == "zhipu" else "local-model")

        def _post_json(url: str, payload: Dict[str, Any], headers: Dict[str, str]):
            data = json.dumps(payload).encode("utf-8")
            try:
                import requests
                resp = requests.post(url, json=payload, headers=headers, timeout=30)
                resp.raise_for_status()
                return resp.json()
            except ImportError:
                import urllib.request
                import urllib.error
                req = urllib.request.Request(url, data=data, headers=headers)
                with urllib.request.urlopen(req, timeout=30) as response:
                    return json.loads(response.read().decode("utf-8"))

        try:
            if provider == "zhipu":
                if not api_key:
                    print(f"AI 调用失败 ({backend_name}): 未配置 API Key")
                    return None
                from zhipuai import ZhipuAI
                client = ZhipuAI(api_key=api_key)
                response = client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": prompt}]
                )
                return response.choices[0].message.content
            else:
                headers = {"Content-Type": "application/json"}
                if api_key:
                    headers["Authorization"] = f"Bearer {api_key}"

                payload = {
                    "model": model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1
                }

                url = f"{base_url.rstrip('/')}/chat/completions"
                result = _post_json(url, payload, headers)
                return result["choices"][0]["message"]["content"]
        except Exception as e:
            print(f"AI 调用失败 ({backend_name}): {e}")
            return None

    def _open_image_recognition(self):
        """打开图片智能识别窗口"""
        global HAS_IMAGE_RECOGNITION
        if not HAS_IMAGE_RECOGNITION:
            # 尝试动态导入
            try:
                from image_recognition_gui import open_image_recognition_window
                HAS_IMAGE_RECOGNITION = True
            except ImportError:
                messagebox.showerror(
                    "功能不可用",
                    "图片智能识别模块未找到。\n\n"
                    "请确保以下文件存在：\n"
                    "• image_intelligence.py\n"
                    "• image_recognition_gui.py\n\n"
                    "依赖安装可在图片识别窗口中选择。"
                )
                return

        # 获取图片识别任务的AI设置
        backend = self._normalize_ai_backend(self._get_ai_backend_for_task("image_rec", allow_legacy=True))
        if backend:
            api_key = backend["api_key"]
            ai_provider = backend["provider"]
            base_url = backend["base_url"]
            model_name = backend["model"]
        else:
            api_key = self.default_values.get("ai_api_key", self.api_key)
            ai_provider = self.default_values.get("ai_provider", "zhipu")
            base_url = self.default_values.get("ai_base_url", "http://localhost:1234/v1")
            model_name = self.default_values.get("ai_model_name", "local-model")
        template_path = self.template_path_var.get() or TEMPLATE_FILE

        # 打开图片识别窗口
        try:
            from image_recognition_gui import open_image_recognition_window
            open_image_recognition_window(
                parent=self.root,
                api_key=api_key,
                ai_provider=ai_provider,
                base_url=base_url,
                model_name=model_name,
                template_path=template_path
            )
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开图片识别窗口：{e}")

    def _open_danfe_recognition(self):
        """打开巴西 DANFE 识别窗口"""
        global HAS_DANFE_RECOGNITION
        if not HAS_DANFE_RECOGNITION:
            try:
                from danfe_recognition_gui import open_danfe_recognition_window
                HAS_DANFE_RECOGNITION = True
            except ImportError:
                messagebox.showerror(
                    "功能不可用",
                    "巴西 DANFE 识别模块未找到。\n\n"
                    "请确保以下文件存在：\n"
                    "• danfe_recognition_module.py\n"
                    "• danfe_recognition_gui.py"
                )
                return

        try:
            from danfe_recognition_gui import open_danfe_recognition_window
            open_danfe_recognition_window(self.root)
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开 DANFE 识别窗口：{e}")

    def _build_document_recognition_tab(self):
        """构建文档识别标签页"""
        doc_frame = ttk.Frame(self.notebook)
        self.notebook.add(doc_frame, text="文档识别 (Docs)")

        main_panel = ttk.Frame(doc_frame, padding=20)
        main_panel.pack(fill="both", expand=True)

        ttk.Label(main_panel, text="文档智能识别中心", font=("", 14, "bold")).pack(pady=(0, 20))
        
        info_text = (
            "本模块支持对特定格式的文档进行批量识别、数据提取并转换为记账凭证。\n"
            "当前支持类型：\n"
            "• 巴西 NF-e (DANFE) 电子发票: 自动提取 Access Key、金额、日期、发行人等。"
        )
        ttk.Label(main_panel, text=info_text, justify="left").pack(pady=10)

        btn_frame = ttk.Frame(main_panel)
        btn_frame.pack(pady=20)

        danfe_btn = ttk.Button(
            btn_frame, 
            text="启动 巴西 DANFE 识别工具", 
            command=self._open_danfe_recognition,
            width=30
        )
        danfe_btn.pack(pady=10)
        
        ttk.Label(main_panel, text="更多文档类型识别功能正在开发中...", foreground="gray").pack(pady=20)

    def _open_export_format_center(self):
        open_export_format_center(self.root, base_data_mgr=self.base_data_mgr)

    def _build_folder_processor_tab(self):
        """构建文件夹汇总平铺标签页"""
        if FolderProcessorGUI:
            proc_frame = ttk.Frame(self.notebook)
            self.notebook.add(proc_frame, text="文件夹平铺汇总")
            # 修复：必须保存实例引用，防止垃圾回收导致 Windows 拖拽回调崩溃 (Python 3.13)
            self.folder_processor = FolderProcessorGUI(proc_frame)
        else:
            proc_frame = ttk.Frame(self.notebook)
            self.notebook.add(proc_frame, text="文件夹平铺汇总")
            ttk.Label(proc_frame, text="无法加载 FolderProcessorGUI 模块，请检查 folder_processor.py 是否存在。", foreground="red").pack(padx=20, pady=20)

    def _init_base_data(self):
        """初始化基础数据"""
        try:
            self.base_data_mgr = BaseDataManager()
            stats = self.base_data_mgr.get_statistics()
            total_records = sum(stats.values())
            if total_records == 0:
                # 数据库为空，尝试自动导入
                result = self.base_data_mgr.import_all_data()
                if result["success"]:
                    print(f"基础数据自动导入成功: {result['message']}")
                else:
                    print(f"基础数据自动导入提示: {result['message']}")
        except Exception as e:
            print(f"初始化基础数据失败: {e}")
            self.base_data_mgr = None

    def _init_summary_recognizer(self):
        """初始化摘要智能识别器"""
        try:
            if self.base_data_mgr:
                self.summary_recognizer = SummaryIntelligence(self.base_data_mgr, self.default_values)
                print("摘要智能识别器初始化成功")
            else:
                print("摘要智能识别器初始化失败：基础数据管理器未就绪")
        except Exception as e:
            print(f"初始化摘要识别器失败: {e}")
            self.summary_recognizer = None

    def _update_recognizer_defaults(self):
        """更新识别器的默认值"""
        if self.summary_recognizer:
            self.summary_recognizer.default_values = self.default_values

    def _load_schemes_to_combo(self):
        """加载保存的映射方案到下拉框"""
        base_modes = [MODE_GENERAL_VOUCHER, MODE_SALES_OUTBOUND, MODE_CUSTOM, MODE_ORIGINAL]
        schemes = []
        if self.base_data_mgr:
            try:
                schemes = self.base_data_mgr.get_mapping_schemes()
            except Exception as e:
                print(f"加载方案失败: {e}")
        
        # 组合模式列表：标准模式 + 分隔符 + 自定义方案
        values = base_modes
        if schemes:
            values = base_modes + ["--- 自定义方案 ---"] + [f"方案: {s['name']}" for s in schemes]
        
        self.mode_combo["values"] = values
        
        # 缓存方案数据以便查找
        self.cached_schemes = {s['name']: s for s in schemes}

    def _save_current_scheme(self):
        """保存当前映射方案"""
        if not self.input_df is not None:
             messagebox.showwarning("提示", "请先加载Excel并配置好映射关系")
             return

        # 弹窗输入名称
        name = tk.simpledialog.askstring("保存方案", "请输入方案名称:")
        if not name:
            return
            
        # 收集当前状态
        current_mode = self.convert_mode_var.get()
        # 如果当前已经是方案模式，需要找到它的原始Base Mode
        if current_mode.startswith("方案: "):
            scheme_name = current_mode.replace("方案: ", "")
            if scheme_name in self.cached_schemes:
                base_mode = self.cached_schemes[scheme_name]["base_mode"]
            else:
                base_mode = MODE_GENERAL_VOUCHER # 默认回退
        elif current_mode == "--- 自定义方案 ---":
             return
        else:
            base_mode = current_mode
            
        template_path = self.template_path_var.get()
        
        # 收集映射关系
        mapping = {}
        for header, var in self.mapping_vars.items():
            val = var.get()
            if val and val != EMPTY_OPTION:
                mapping[header] = val
        
        # 辅助列持久化
        if self.manual_debit_col_var.get() != EMPTY_OPTION:
            mapping["__aux_debit_col__"] = self.manual_debit_col_var.get()
        if self.manual_credit_col_var.get() != EMPTY_OPTION:
            mapping["__aux_credit_col__"] = self.manual_credit_col_var.get()
        if self.manual_dc_col_var.get() != EMPTY_OPTION:
            mapping["__aux_dc_col__"] = self.manual_dc_col_var.get()
        
        # 字段格式自定义持久化
        if self.field_formats:
            mapping["__field_formats__"] = self.field_formats
                
        if self.base_data_mgr:
            res = self.base_data_mgr.save_mapping_scheme(name, base_mode, template_path, mapping, self.custom_composite_fields)
            if res["success"]:
                messagebox.showinfo("成功", "方案已保存")
                self._load_schemes_to_combo() # 刷新列表
                # 自动选中新方案
                self.mode_combo.set(f"方案: {name}")
            else:
                messagebox.showerror("错误", res["message"])

    def _delete_current_scheme(self):
        """删除当前选中的方案"""
        current_mode = self.convert_mode_var.get()
        if not current_mode.startswith("方案: "):
             messagebox.showinfo("提示", "当前未选择自定义方案，无法删除")
             return
             
        name = current_mode.replace("方案: ", "")
        if messagebox.askyesno("确认", f"确定要删除方案 '{name}' 吗？"):
            if self.base_data_mgr:
                res = self.base_data_mgr.delete_mapping_scheme(name)
                if res["success"]:
                    messagebox.showinfo("成功", "方案已删除")
                    self._load_schemes_to_combo()
                    self.mode_combo.set(MODE_GENERAL_VOUCHER)
                    self._on_mode_changed()
                else:
                    messagebox.showerror("错误", res["message"])

    # ---------- 综合字段管理 ----------
    def _get_all_composite_fields(self) -> Dict[str, Dict]:
        merged = dict(BUILTIN_COMPOSITE_FIELDS)
        merged.update(self.custom_composite_fields or {})
        return merged

    def _composite_option_label(self, name: str) -> str:
        return f"{COMPOSITE_PREFIX} {name}"

    def _composite_option_labels(self) -> List[str]:
        return [self._composite_option_label(n) for n in self._get_all_composite_fields().keys()]

    def _is_composite_option(self, value: Optional[str]) -> bool:
        return isinstance(value, str) and value.startswith(f"{COMPOSITE_PREFIX} ")

    def _extract_composite_name(self, value: str) -> Optional[str]:
        if not self._is_composite_option(value):
            return None
        return value[len(COMPOSITE_PREFIX) + 1:]

    def _get_mapping_source_options(self) -> List[str]:
        return [EMPTY_OPTION] + self.input_columns + self._composite_option_labels() + [CUSTOM_INPUT_OPTION]

    def _prompt_custom_mapping(self, header_name: str, preset_name: Optional[str] = None, force_create_const: bool = False):
        if self.input_df is None:
            return
        if preset_name is None:
            raw_name = simpledialog.askstring("自定义字段", "请输入源列名（若不存在可创建常量列）:", parent=self.root)
            if raw_name is None:
                self.mapping_vars.get(header_name, tk.StringVar(value=EMPTY_OPTION)).set(EMPTY_OPTION)
                return
            col_name = str(raw_name).strip()
        else:
            col_name = str(preset_name).strip()
        if not col_name:
            self.mapping_vars.get(header_name, tk.StringVar(value=EMPTY_OPTION)).set(EMPTY_OPTION)
            return
        if col_name.startswith("="):
            label = self._ensure_inline_composite(header_name, col_name)
            if label and header_name in self.mapping_vars:
                self._skip_auto_mapping_once = True
                self._refresh_mapping_sources_after_composite_change()
                self.mapping_vars[header_name].set(label)
            return
        if col_name not in self.input_df.columns:
            if force_create_const:
                self.input_df[col_name] = col_name
            else:
                create = messagebox.askyesno("列不存在", f"源表不存在列“{col_name}”。是否创建常量列？")
                if not create:
                    self.mapping_vars.get(header_name, tk.StringVar(value=EMPTY_OPTION)).set(EMPTY_OPTION)
                    return
                const_val = simpledialog.askstring("常量列", f"请输入“{col_name}”的常量值:", parent=self.root)
                if const_val is None:
                    self.mapping_vars.get(header_name, tk.StringVar(value=EMPTY_OPTION)).set(EMPTY_OPTION)
                    return
                self.input_df[col_name] = const_val
            self.input_columns = [str(c) for c in self.input_df.columns]
            self._skip_auto_mapping_once = True
            self._refresh_mapping_sources_after_composite_change()
        if header_name in self.mapping_vars:
            self.mapping_vars[header_name].set(col_name)

    def _on_mapping_selected(self, header_name: str, var: tk.StringVar):
        selected = var.get()
        if selected == CUSTOM_INPUT_OPTION:
            self._prompt_custom_mapping(header_name)
            return
        self.log_message(f"手动修改映射: {header_name} <== {selected}")

    def _on_mapping_value_committed(self, header_name: str, var: tk.StringVar):
        selected = str(var.get()).strip()
        if not selected:
            var.set(EMPTY_OPTION)
            return
        if selected == CUSTOM_INPUT_OPTION:
            self._prompt_custom_mapping(header_name)
            return
        if selected.startswith("="):
            label = self._ensure_inline_composite(header_name, selected)
            if label:
                self._skip_auto_mapping_once = True
                self._refresh_mapping_sources_after_composite_change()
                if header_name in self.mapping_vars:
                    self.mapping_vars[header_name].set(label)
            return
        if self._is_composite_option(selected) or selected == EMPTY_OPTION:
            self.log_message(f"手动修改映射: {header_name} <== {selected}")
            return
        options = set(self._get_mapping_source_options())
        if selected in options:
            self.log_message(f"手动修改映射: {header_name} <== {selected}")
            return
        self._prompt_custom_mapping(header_name, preset_name=selected, force_create_const=True)

    def _add_mapping_source_column(self):
        if self.input_df is None:
            messagebox.showwarning("提示", "请先选择原始 Excel 文件。")
            return
        col_name = simpledialog.askstring("添加字段", "请输入新的源列名:", parent=self.root)
        if col_name is None:
            return
        col_name = str(col_name).strip()
        if not col_name:
            return
        if col_name in self.input_df.columns:
            messagebox.showinfo("提示", f"列“{col_name}”已存在。")
            return
        const_val = simpledialog.askstring("常量列", f"请输入“{col_name}”的常量值:", parent=self.root)
        if const_val is None:
            return
        self.input_df[col_name] = const_val
        self.input_columns = [str(c) for c in self.input_df.columns]
        self._skip_auto_mapping_once = True
        self._refresh_mapping_sources_after_composite_change()

    def _remove_mapping_source_column(self):
        if self.input_df is None:
            messagebox.showwarning("提示", "请先选择原始 Excel 文件。")
            return
        col_name = simpledialog.askstring(
            "删除字段",
            f"请输入要删除的源列名:\n{', '.join(self.input_columns)}",
            parent=self.root
        )
        if col_name is None:
            return
        col_name = str(col_name).strip()
        if not col_name:
            return
        if col_name not in self.input_df.columns:
            messagebox.showwarning("提示", f"源列“{col_name}”不存在。")
            return
        if not messagebox.askyesno("确认删除", f"确定删除源列“{col_name}”？\n已映射到该列的字段将被清空。"):
            return
        try:
            self.input_df.drop(columns=[col_name], inplace=True)
        except Exception as exc:
            messagebox.showerror("错误", f"删除源列失败：{exc}")
            return
        self.input_columns = [str(c) for c in self.input_df.columns]
        self._skip_auto_mapping_once = True
        self._refresh_mapping_sources_after_composite_change()

    def _refresh_mapping_sources_after_composite_change(self):
        """更新映射下拉选项，尽量保留已选值"""
        if self.input_df is None:
            return
        previous = {h: var.get() for h, var in self.mapping_vars.items()}
        self._create_mapping_widgets()
        options = set(self._get_mapping_source_options())
        for header, val in previous.items():
            if header in self.mapping_vars and val in options:
                self.mapping_vars[header].set(val)

    def _get_effective_mode_for_mapping(self):
        raw_mode = self.convert_mode_var.get()
        if raw_mode.startswith("方案: "):
            scheme_name = raw_mode.replace("方案: ", "")
            if hasattr(self, "cached_schemes") and scheme_name in self.cached_schemes:
                return self.cached_schemes[scheme_name]["base_mode"], scheme_name
            return MODE_GENERAL_VOUCHER, scheme_name
        return raw_mode, None

    def _resolve_default_value(self, name: str):
        if name in self.default_values:
            return self.default_values.get(name)
        alias = DEFAULT_VALUE_ALIASES.get(name)
        if alias and alias in self.default_values:
            return self.default_values.get(alias)
        if name.startswith("默认"):
            stripped = name.replace("默认", "", 1)
            if stripped in self.default_values:
                return self.default_values.get(stripped)
        return None

    def _get_prompt_counter_subject_account_pair(self):
        """读取“生成前参数确认”中的对方科目-账户编码绑定。"""
        subject_code = self._normalize_subject_code(self.default_values.get("对方科目") or "")
        account_code = self._normalize_code_value(self._resolve_default_value("默认账户"))
        if not subject_code or not account_code:
            return "", ""
        return subject_code, account_code

    def _normalize_code_value(self, value: Any) -> str:
        if value is None:
            return ""
        s = str(value).strip()
        if not s or s.lower() == "nan":
            return ""
        if s.endswith(".0"):
            s = s[:-2]
        return s

    def _get_cash_business_override(self, summary: Any) -> Dict[str, str]:
        """现金存/取业务强制覆盖：主科目1001，往来单位编码100102。"""
        if summary is None:
            return {}
        s = str(summary).lower().replace(" ", "")
        if not s:
            return {}

        direct_hits = (
            "现金存bac", "现金取bac", "存现金bac", "取现金bac",
            "现金存st", "现金取st", "存现金st", "取现金st",
        )
        if any(k in s for k in direct_hits):
            return {"科目编码": "1001", "往来单位编码": "100102"}

        has_cash = ("现金" in s) or ("cash" in s)
        has_action = any(k in s for k in ("存", "取", "deposit", "withdraw"))
        has_bank = ("bac" in s) or ("st" in s)
        if has_cash and has_action and has_bank:
            return {"科目编码": "1001", "往来单位编码": "100102"}

        return {}

    def _get_bank_account_rows(self) -> List[Dict[str, Any]]:
        cache = getattr(self, "_bank_account_rows_cache", None)
        if cache is not None:
            return cache
        rows: List[Dict[str, Any]] = []
        if self.base_data_mgr:
            try:
                rows = self.base_data_mgr.query("bank_account") or []
            except Exception:
                rows = []
        self._bank_account_rows_cache = rows
        return rows

    def _invalidate_bank_account_cache(self):
        self._bank_account_rows_cache = None

    def _resolve_linked_default_account(
        self,
        subject_code: str,
        current_value: Any = None,
        partner_code: Any = None,
        allow_fallback: bool = True
    ) -> str:
        """默认账户优先按对方科目匹配，未命中时回退到输入值/默认值/往来编码。"""
        current_code = self._normalize_code_value(current_value)
        preferred = self._normalize_code_value(self._resolve_default_value("默认账户"))
        partner_fallback = self._normalize_code_value(partner_code)
        subj = self._normalize_subject_code(subject_code)
        if not subj:
            if not allow_fallback:
                return ""
            prefer_default_first = bool(getattr(self, "_prefer_prompt_default_account", False))
            fallback_order = [preferred, current_code, partner_fallback] if prefer_default_first else [current_code, preferred, partner_fallback]
            for candidate in fallback_order:
                if candidate:
                    return candidate
            return ""

        def _match_subject(account_code: str, account_subject: str) -> bool:
            code = self._normalize_code_value(account_code)
            subject_text = str(account_subject or "")
            if not code:
                return False
            if f"({subj})" in subject_text or f"[{subj}]" in subject_text:
                return True
            # 兜底：账户编码与科目前缀一致也视为匹配
            if code.startswith(subj):
                return True
            return False

        def _pick_if_match(target_code: str) -> str:
            if not target_code:
                return ""
            for row in self._get_bank_account_rows():
                code = self._normalize_code_value(row.get("code"))
                if code != target_code:
                    continue
                if _match_subject(code, row.get("account_subject", "")):
                    return code
                return ""
            return ""

        prefer_default_first = bool(getattr(self, "_prefer_prompt_default_account", False))
        if prefer_default_first:
            matched_preferred = _pick_if_match(preferred)
            if matched_preferred:
                return matched_preferred
            matched_current = _pick_if_match(current_code)
            if matched_current:
                return matched_current
        else:
            matched_current = _pick_if_match(current_code)
            if matched_current:
                return matched_current
            matched_preferred = _pick_if_match(preferred)
            if matched_preferred:
                return matched_preferred

        candidates: List[str] = []
        for row in self._get_bank_account_rows():
            code = self._normalize_code_value(row.get("code"))
            if not code:
                continue
            account_subject = str(row.get("account_subject", "") or "")
            if _match_subject(code, account_subject):
                candidates.append(code)
        if candidates:
            return candidates[0]

        # 未匹配到账户科目时，不置空：优先保留输入值，其次默认值，再回退往来单位编码
        if not allow_fallback:
            return ""
        fallback_order = [preferred, current_code, partner_fallback] if prefer_default_first else [current_code, preferred, partner_fallback]
        for candidate in fallback_order:
            if candidate:
                return candidate
        return ""

    def _ensure_inline_composite(self, header_name: str, expr: str) -> Optional[str]:
        expr = str(expr).strip()
        if expr.startswith("="):
            expr = expr[1:].strip()
        if not expr:
            return None
        key_seed = f"{header_name}|{expr}"
        key = f"expr_{hashlib.md5(key_seed.encode('utf-8')).hexdigest()[:10]}"
        if key not in self.custom_composite_fields:
            self.custom_composite_fields[key] = {
                "type": "expression",
                "expression": expr,
                "desc": f"inline:{header_name}",
            }
        return self._composite_option_label(key)

    def _apply_auto_mapping(self):
        if getattr(self, "_skip_auto_mapping_once", False):
            self._skip_auto_mapping_once = False
            return
        if not self.base_data_mgr or self.input_df is None or not self.template_headers:
            return
        effective_mode, scheme_name = self._get_effective_mode_for_mapping()
        if effective_mode == MODE_ORIGINAL:
            self._debug_log("跳过自动映射: 原格式模式")
            return
        if scheme_name:
            self._debug_log("跳过自动映射: 方案模式")
            return
        template_path = self.template_path_var.get() or TEMPLATE_FILE
        saved = self.base_data_mgr.get_auto_mapping(
            template_path,
            effective_mode,
            list(self.input_df.columns)
        )
        if not saved:
            return
        mapping = saved.get("mapping") or {}
        composite = saved.get("composite") or {}
        if composite:
            self.custom_composite_fields = composite
        options = set(self._get_mapping_source_options())
        for header, col in mapping.items():
            if header in self.mapping_vars and col in options:
                self.mapping_vars[header].set(col)
        if mapping:
            self.log_message("已自动应用历史字段映射。")
            self._debug_log(f"自动映射字段数: {len(mapping)}")

    def _save_auto_mapping(self, mapping: Dict[str, Optional[str]]):
        if not self.base_data_mgr or self.input_df is None or not mapping:
            return
        effective_mode, scheme_name = self._get_effective_mode_for_mapping()
        if effective_mode == MODE_ORIGINAL:
            self._debug_log("跳过保存自动映射: 原格式模式")
            return
        if scheme_name:
            self._debug_log("跳过保存自动映射: 方案模式")
            return
        template_path = self.template_path_var.get() or TEMPLATE_FILE
        cleaned = {k: v for k, v in mapping.items() if v}
        if not cleaned:
            return
        self.base_data_mgr.save_auto_mapping(
            template_path,
            effective_mode,
            list(self.input_df.columns),
            cleaned,
            self.custom_composite_fields
        )
        self._debug_log(f"已保存自动映射: {len(cleaned)} 字段")

    def _open_composite_field_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("综合字段设置")
        dialog.geometry("800x600")

        container = ttk.Frame(dialog, padding=12)
        container.pack(fill="both", expand=True)

        help_text = (
            "表达式使用受限 Python 风格，可用 helper：\n"
            "- col('源列名'): 读取原始Excel列\n"
            "- lookup('表名', '查找列', 值, '目标列'): 查基础数据 (可用表: currency, department...)\n"
            "- field('模板字段名') / composite('综合名') / abs/round/min/max/Decimal\n"
            "AI 助手: 在下方输入自然语言描述 (如: 如果币种是USD则取金额，否则取金额除以汇率)，点击生成。"
        )
        ttk.Label(container, text=help_text, justify="left").pack(anchor="w", pady=(0, 8))

        tree = ttk.Treeview(container, columns=("name", "type", "expr", "desc", "scope"), show="headings", height=6)
        tree.heading("name", text="名称")
        tree.heading("type", text="类型")
        tree.heading("expr", text="表达式/规则")
        tree.heading("desc", text="说明")
        tree.heading("scope", text="来源")
        tree.column("expr", width=260)
        tree.column("desc", width=160)
        tree.pack(fill="x", pady=(0, 8))
        attach_treeview_tools(tree)

        form = ttk.Frame(container)
        form.pack(fill="x", pady=(0, 8))

        ttk.Label(form, text="名称:").grid(row=0, column=0, sticky="e", padx=5, pady=3)
        name_var = tk.StringVar()
        ttk.Entry(form, textvariable=name_var, width=24).grid(row=0, column=1, sticky="w")

        # --- AI 生成区域 ---
        ttk.Label(form, text="AI 描述:").grid(row=1, column=0, sticky="e", padx=5, pady=3)
        ai_desc_var = tk.StringVar()
        ai_entry = ttk.Entry(form, textvariable=ai_desc_var, width=50)
        ai_entry.grid(row=1, column=1, sticky="w")
        
        # 结果反馈标签
        result_label = ttk.Label(form, text="", foreground="gray")
        result_label.grid(row=4, column=1, columnspan=2, sticky="w")

        def run_test_formula(code=None):
            """测试运行表达式"""
            expr = code if code is not None else expr_var.get().strip()
            if not expr:
                result_label.config(text="⚠️ 表达式为空", foreground="orange")
                return
            
            sample_row = {}
            warning_msg = ""
            
            if self.input_df is None or self.input_df.empty:
                warning_msg = " (无源数据，col()将返回空)"
            else:
                # 取第一行非空数据作为样本
                try:
                    sample_row = self.input_df.iloc[0].to_dict()
                    for i in range(min(5, len(self.input_df))):
                         r = self.input_df.iloc[i].to_dict()
                         if any(str(v).strip() for v in r.values() if pd.notna(v)):
                             sample_row = r
                             break
                except Exception:
                    pass

            # 构建简易测试环境
            def mock_col(name):
                return sample_row.get(name)
            
            def mock_lookup(table, key, val, target):
                if not self.base_data_mgr: return "DB_Not_Init"
                return self.base_data_mgr.lookup_value(table, key, val, target)
            
            def mock_field(tmpl_name):
                # 1. 尝试映射
                var = self.mapping_vars.get(tmpl_name)
                src_col = var.get() if var else None
                if src_col and src_col != EMPTY_OPTION:
                    if self._is_composite_option(src_col):
                         return f"[Composite:{src_col}]"
                    return mock_col(src_col)
                # 2. 尝试默认值
                return self._resolve_default_value(tmpl_name)

            # 简单的 mock 其他函数
            env = {
                "col": mock_col,
                "field": mock_field,
                "mapped": mock_field,
                "composite": lambda x: f"[Comp:{x}]",
                "lookup": mock_lookup,
                "abs": abs, "round": round, "min": min, "max": max, "Decimal": Decimal,
                "float": float, "int": int, "str": str,
                "derived": {}, "smart": {}
            }

            try:
                res = safe_eval_expr(expr, env)
                result_label.config(text=f"✅ 测试通过，预览结果: {res}{warning_msg}", foreground="green")
            except Exception as e:
                result_label.config(text=f"❌ 执行错误: {e}", foreground="red")

        def generate_formula():
            user_desc = ai_desc_var.get().strip()
            if not user_desc:
                messagebox.showwarning("提示", "请输入描述")
                return
            
            # 构建 Prompt 上下文
            columns_str = ", ".join(self.input_columns)
            
            # 1. 字段规则摘要
            rules_summary = []
            for f, r in list(FIELD_RULES.items())[:15]: # 仅列出常用字段避免Token溢出
                rtype = r.get('type', 'text')
                limit = ""
                if rtype == 'number':
                    limit = f"(小数位:{r.get('max_decimal_len', 2)})"
                elif rtype == 'text':
                    limit = f"(最大:{r.get('max_len', '?')}字)"
                rules_summary.append(f"{f}: {rtype}{limit}")
            rules_str = ", ".join(rules_summary) + "..."

            # 2. 基础数据表结构 (提示给AI)
            tables_desc = (
                "- currency (code, name, exchange_rate)\n"
                "- department (code, name)\n"
                "- warehouse (code, name)\n"
                "- account_subject (code_name, is_subject, debit_credit_type)\n"
                "- product (code, name, spec_info, unit, in_price, out_price)\n"
                "- business_partner (code, name, contact_person, tax_number)"
            )

            prompt = f"""
            你是一个 Python 表达式生成器，用于 Excel 数据转换工具。请根据用户描述生成一个合法的 Python 表达式。
            
            【上下文信息】
            1. 原始 Excel 列名: {columns_str}
            2. 目标模板字段规则: {rules_str}
            3. 可用的基础数据表(用于lookup):
            {tables_desc}
            
            【可用函数与变量】
            - col('列名'): 获取原始 Excel 列的值 (注意：通常返回字符串，数值计算需转 float)
            - lookup('表名', '查找键列', 查找值, '目标值列'): 在基础数据库中查找
              例: lookup('currency', 'name', col('币种'), 'exchange_rate')
            - field('模板字段名'): 引用已映射的字段值 (如果未映射，会自动尝试获取系统设置的默认值)
              例: field('默认账户') 可以获取设置好的默认科目
            - composite('综合字段名'): 引用其他已定义的综合字段
            - Python 内置: abs(), round(), min(), max(), Decimal(), str(), float(), int()
            - 逻辑运算: if/else (e.g. A if cond else B)
            
            【用户需求】
            {user_desc}
            
            【生成要求】
            - 只返回 Python 表达式代码，不要包含 ```python 标记或任何解释文字。
            - 确保引用的列名存在于“原始 Excel 列名”中。
            - 涉及金额计算时，请务必使用 float() 转换，防止字符串拼接。
            - 优先使用 field('字段') 来获取已知的模板字段，特别是涉及默认值的场景。
            """
            
            btn_gen.config(state="disabled", text="生成中...")
            result_label.config(text="🤖 AI 思考中...", foreground="blue")
            self.root.update()

            def _run_ai():
                try:
                    result = self.call_ai_service("formula_gen", prompt)
                    def _on_success():
                        if result:
                            code = result.strip().replace("```python", "").replace("```", "").strip()
                            expr_var.set(code)
                            run_test_formula(code)
                        else:
                            messagebox.showerror("失败", "AI 未返回结果，请检查配置或网络。")
                            result_label.config(text="❌ AI 调用失败", foreground="red")
                    self.root.after(0, _on_success)
                except Exception as e:
                    def _on_error():
                        messagebox.showerror("错误", f"生成失败: {e}")
                        result_label.config(text="❌ 系统异常", foreground="red")
                    self.root.after(0, _on_error)
                finally:
                    self.root.after(0, lambda: btn_gen.config(state="normal", text="AI 生成"))

            threading.Thread(target=_run_ai, daemon=True).start()

        btn_gen = ttk.Button(form, text="AI 生成", command=generate_formula)
        btn_gen.grid(row=1, column=2, padx=5, sticky="w")

        ttk.Label(form, text="表达式:").grid(row=2, column=0, sticky="e", padx=5, pady=3)
        expr_var = tk.StringVar()
        expr_entry = ttk.Entry(form, textvariable=expr_var, width=58)
        expr_entry.grid(row=2, column=1, sticky="w")
        
        # 手动测试按钮
        btn_test = ttk.Button(form, text="▶ 试运行", command=lambda: run_test_formula(), width=8)
        btn_test.grid(row=2, column=2, padx=5, sticky="w")

        ttk.Label(form, text="说明:").grid(row=5, column=0, sticky="e", padx=5, pady=3)
        desc_var = tk.StringVar()
        ttk.Entry(form, textvariable=desc_var, width=58).grid(row=5, column=1, sticky="w")

        def refresh_tree():
            for item in tree.get_children():
                tree.delete(item)
            all_fields = self._get_all_composite_fields()
            for name, cfg in all_fields.items():
                scope = "内置" if name in BUILTIN_COMPOSITE_FIELDS else "自定义"
                c_type = cfg.get("type", "expression")
                expr = "借贷列合并(借为正/贷为负)" if c_type == "debit_credit" else cfg.get("expression", "")
                desc = cfg.get("description", "")
                tree.insert("", "end", values=(name, c_type, expr, desc, scope))

        def on_tree_select(event=None):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            name_var.set(vals[0])
            expr_var.set("" if vals[1] == "debit_credit" else vals[2])
            desc_var.set(vals[3])
            if vals[0] in BUILTIN_COMPOSITE_FIELDS:
                expr_entry_state = "disabled"
            else:
                expr_entry_state = "normal"
            expr_entry.configure(state=expr_entry_state)

        tree.bind("<<TreeviewSelect>>", on_tree_select)

        def save_or_update():
            name = name_var.get().strip()
            expr = expr_var.get().strip()
            desc = desc_var.get().strip()
            if not name:
                messagebox.showwarning("提示", "请填写名称")
                return
            if name in self.input_columns or self._composite_option_label(name) in self.input_columns:
                messagebox.showwarning("提示", "名称与源表列冲突，请更换名称")
                return
            if name in BUILTIN_COMPOSITE_FIELDS:
                messagebox.showinfo("提示", "内置综合字段不可修改，请使用其他名称。")
                return
            if not expr:
                messagebox.showwarning("提示", "请填写表达式")
                return
            # 简单语法校验，提前反馈
            try:
                safe_eval_expr(expr, {"col": lambda x: None, "field": lambda x: None, "mapped": lambda x: None, "composite": lambda x: None, "lookup": lambda a,b,c,d: None, "abs": abs, "round": round, "min": min, "max": max, "Decimal": Decimal, "float": float, "int": int, "str": str})
            except Exception as e:
                messagebox.showerror("错误", f"表达式无效：{e}")
                return
            self.custom_composite_fields[name] = {
                "type": "expression",
                "expression": expr,
                "description": desc,
            }
            # 持久化保存
            if self.base_data_mgr:
                try:
                    self.base_data_mgr.set_config("app_global_composite_fields", json.dumps(self.custom_composite_fields))
                except Exception as e:
                    print(f"保存综合字段失败: {e}")

            refresh_tree()
            self._refresh_mapping_sources_after_composite_change()
            messagebox.showinfo("成功", f"已保存综合字段: {name}")

        def delete_selected():
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            name = vals[0]
            if name in BUILTIN_COMPOSITE_FIELDS:
                messagebox.showinfo("提示", "内置综合字段不可删除。")
                return
            if name in self.custom_composite_fields:
                self.custom_composite_fields.pop(name, None)
                # 持久化保存
                if self.base_data_mgr:
                    try:
                        self.base_data_mgr.set_config("app_global_composite_fields", json.dumps(self.custom_composite_fields))
                    except Exception as e:
                        print(f"保存综合字段失败: {e}")

                refresh_tree()
                self._refresh_mapping_sources_after_composite_change()

        def reset_form():
            name_var.set("")
            expr_var.set("")
            desc_var.set("")
            ai_desc_var.set("")
            expr_entry.configure(state="normal")
            for item in tree.selection():
                tree.selection_remove(item)

        btns = ttk.Frame(container)
        btns.pack(fill="x")
        ttk.Button(btns, text="新建", command=reset_form, width=8).pack(side="left", padx=3)
        ttk.Button(btns, text="保存/更新", command=save_or_update, width=12).pack(side="left", padx=3)
        ttk.Button(btns, text="删除", command=delete_selected, width=8).pack(side="left", padx=3)
        ttk.Button(btns, text="关闭", command=dialog.destroy, width=8).pack(side="right", padx=3)

        refresh_tree()
        dialog.transient(self.root)
        dialog.grab_set()

    def _build_ui(self):
        # 菜单栏
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # 文件菜单
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="文件", menu=file_menu)
        file_menu.add_command(label="导入基础数据", command=self.import_base_data)
        file_menu.add_command(label="批量导入基础数据Excel", command=self.import_base_data_batch)
        file_menu.add_command(label="导出基础数据模板", command=self.export_base_data_templates)
        file_menu.add_command(label="导入AI训练数据", command=self.import_ai_training_data)
        file_menu.add_command(label="查看统计信息", command=self.show_base_data_stats)
        file_menu.add_command(label="刷新识别缓存", command=self.refresh_recognition_cache)
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.root.quit)

        # 设置菜单（高级选项集中管理）
        settings_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="设置", menu=settings_menu)
        settings_menu.add_checkbutton(label="启用摘要智能识别（自动填充字段）", variable=self.enable_smart_recognition)
        settings_menu.add_checkbutton(label="启用 AI 深度识别 (科目判断)", variable=self.use_ai_var)
        settings_menu.add_checkbutton(label="启用外币模式 (自动计算本币)", variable=self.use_foreign_currency_var)
        settings_menu.add_checkbutton(label="自动生成对方分录（借贷平衡）", variable=self.auto_balance_var)
        settings_menu.add_checkbutton(label="强制拆分金额为借/贷两列 (输出)", variable=self.split_amount_var)
        settings_menu.add_separator()
        settings_menu.add_command(label="导出格式中心...", command=self._open_export_format_center)
        settings_menu.add_command(label="AI 设置...", command=self._show_ai_settings)
        settings_menu.add_command(label="设置默认值...", command=self._show_default_values_dialog)
        settings_menu.add_command(label="图片智能识别...", command=self._open_image_recognition)
        settings_menu.add_command(label="巴西 DANFE 识别...", command=self._open_danfe_recognition)

        # 创建标签页控件
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)

        # 标签页1: Excel转换
        self._build_excel_converter_tab()

        # 标签页2: 摘要匹配
        self._build_summary_match_tab()

        # 标签页3: 基础数据管理
        self._build_base_data_tab()

        # 标签页4: 智能对账
        self._build_reconciliation_tab()

        # 标签页5: 控制台
        self._build_console_tab()

        # 标签页6: 批量合并
        self._build_batch_merge_tab()

        # 标签页7: 报关清单汇总
        self._build_shipping_tab()

        # 标签页8: 经营报告
        self._build_report_tab()

        # 标签页9: 文档识别
        self._build_document_recognition_tab()
        self._build_folder_processor_tab()

    def _build_reconciliation_tab(self):
        """构建智能对账标签页"""
        recon_frame = ttk.Frame(self.notebook)
        self.notebook.add(recon_frame, text="智能对账系统")

        main_panel = ttk.Frame(recon_frame, padding=10)
        main_panel.pack(fill="both", expand=True)

        # --- 0. 对账类型选择 ---
        type_frame = ttk.Frame(main_panel)
        type_frame.pack(fill="x", pady=(0, 5))
        ttk.Label(type_frame, text="对账类型 (Type):").pack(side="left")
        self.recon_type = tk.StringVar(value="应收账款 (AR)")
        type_cb = ttk.Combobox(
            type_frame,
            textvariable=self.recon_type,
            values=["应收账款 (AR)", "现金帐 (Cash)", "银行对账 (Bank)", "自定义 (Custom)"],
            state="readonly",
            width=25,
        )
        type_cb.pack(side="left", padx=5)

        # --- 1. 文件选择区域 ---
        file_frame = ttk.LabelFrame(main_panel, text="数据源选择 (支持 Excel/PDF)", padding=10)
        file_frame.pack(fill="x", pady=5)

        # 当地系统文件
        ttk.Label(file_frame, text="当地系统文件 (Local):").grid(row=0, column=0, sticky="e", pady=5)
        self.recon_local_path = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.recon_local_path, width=50).grid(row=0, column=1, sticky="ew", padx=5)
        local_btn_frame = ttk.Frame(file_frame)
        local_btn_frame.grid(row=0, column=2, padx=5, sticky="w")
        ttk.Button(
            local_btn_frame,
            text="浏览...",
            command=lambda: self._select_file(
                self.recon_local_path,
                [("Excel/PDF Files", "*.xlsx;*.xls;*.pdf"), ("Excel Files", "*.xlsx;*.xls"), ("PDF Files", "*.pdf")],
            ),
        ).pack(side="left")
        ttk.Button(local_btn_frame, text="预浏览", command=self._preview_recon_local_data).pack(side="left", padx=5)

        # 亿看系统文件
        ttk.Label(file_frame, text="亿看系统文件 (Yikan):").grid(row=1, column=0, sticky="e", pady=5)
        self.recon_yikan_path = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.recon_yikan_path, width=50).grid(row=1, column=1, sticky="ew", padx=5)
        yikan_btn_frame = ttk.Frame(file_frame)
        yikan_btn_frame.grid(row=1, column=2, padx=5, sticky="w")
        ttk.Button(
            yikan_btn_frame,
            text="浏览...",
            command=lambda: self._select_file(self.recon_yikan_path, [("Excel Files", "*.xlsx;*.xls")]),
        ).pack(side="left")
        ttk.Button(yikan_btn_frame, text="预浏览", command=self._preview_recon_yikan_data).pack(side="left", padx=5)

        file_frame.columnconfigure(1, weight=1)

        # --- 2. 选项区域 ---
        opt_frame = ttk.LabelFrame(main_panel, text="对账选项", padding=10)
        opt_frame.pack(fill="x", pady=5)

        ttk.Label(opt_frame, text="日期范围 (YYYY-MM-DD 或 YYYY-MM):").pack(side="left")
        self.recon_start_date = tk.StringVar()
        self.recon_end_date = tk.StringVar()
        
        ttk.Entry(opt_frame, textvariable=self.recon_start_date, width=12).pack(side="left", padx=5)
        ttk.Label(opt_frame, text="至").pack(side="left")
        ttk.Entry(opt_frame, textvariable=self.recon_end_date, width=12).pack(side="left", padx=5)
        
        ttk.Label(opt_frame, text="(留空则匹配所有日期)", foreground="gray").pack(side="left", padx=5)

        ttk.Checkbutton(opt_frame, text="启用编码模糊匹配", variable=self.recon_fuzzy_var).pack(side="left", padx=5)

        ttk.Checkbutton(opt_frame, text="启用AI智能错误检索", variable=self.recon_use_ai_analysis).pack(side="left", padx=5)

        # 新增：自主选择是否强制人工匹配表头
        self.recon_force_manual_mapping_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_frame, text="强制人工匹配表头", variable=self.recon_force_manual_mapping_var).pack(side="left", padx=5)

        ttk.Checkbutton(opt_frame, text="反转银行借贷方向 (Bank Only)", variable=self.invert_bank_var).pack(side="left", padx=5)

        ttk.Checkbutton(opt_frame, text="同时导出未匹配项为凭证导入文件", variable=self.recon_export_voucher_var).pack(side="left", padx=5)

        # --- 3. 操作按钮 ---
        btn_frame = ttk.Frame(main_panel, padding=10)
        btn_frame.pack(fill="x")
        
        ttk.Button(btn_frame, text="开始对账", command=self._start_reconciliation, width=20).pack(side="left", padx=5)
        
        # --- 4. 日志区域 ---
        log_frame = ttk.LabelFrame(main_panel, text="执行日志", padding=10)
        log_frame.pack(fill="both", expand=True, pady=5)
        
        self.recon_log_text = tk.Text(log_frame, height=15, width=80)
        self.recon_log_text.pack(fill="both", expand=True)

    def _select_file(self, var, filetypes=None):
        ftypes = filetypes or [("Excel Files", "*.xlsx;*.xls")]
        f = filedialog.askopenfilename(filetypes=ftypes)
        if f:
            var.set(f)

    def _log_recon(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.recon_log_text.insert("end", f"[{ts}] {msg}\n")
        self.recon_log_text.see("end")
        self.root.update_idletasks()

    def _init_recon_header_mapping_store(self):
        if hasattr(self, "_recon_header_mapping_store"):
            return
        self._recon_header_mapping_store = {"version": 1, "local": {}, "yikan": {}}
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self._recon_header_mapping_path = os.path.join(
            base_dir,
            "reconciliation_header_mapping.json",
        )
        if not os.path.exists(self._recon_header_mapping_path):
            return
        try:
            with open(self._recon_header_mapping_path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            if isinstance(data, dict):
                self._recon_header_mapping_store.update(data)
        except Exception as exc:
            self._log_recon(f"警告: 加载对账表头映射失败: {exc}")

    def _get_recon_header_signature(self, columns):
        norm_cols = [normalize_header(c) for c in columns]
        norm_cols = [c for c in norm_cols if c]
        norm_cols = sorted(set(norm_cols))
        signature_src = "|".join(norm_cols)
        return hashlib.sha1(signature_src.encode("utf-8")).hexdigest()

    def _get_saved_recon_mapping(self, context_key, columns):
        self._init_recon_header_mapping_store()
        signature = self._get_recon_header_signature(columns)
        context = self._recon_header_mapping_store.get(context_key, {})
        entry = context.get(signature)
        if not entry:
            return None
        mapping = entry.get("mapping", {})
        if not isinstance(mapping, dict):
            return None
        return {k: v for k, v in mapping.items() if v in columns}

    def _save_recon_mapping(self, context_key, columns, mapping):
        if not mapping:
            return
        self._init_recon_header_mapping_store()
        signature = self._get_recon_header_signature(columns)
        context = self._recon_header_mapping_store.setdefault(context_key, {})
        context[signature] = {
            "columns": [str(c) for c in columns],
            "mapping": {str(k): str(v) for k, v in mapping.items()},
        }
        try:
            with open(self._recon_header_mapping_path, "w", encoding="utf-8") as handle:
                json.dump(self._recon_header_mapping_store, handle, ensure_ascii=True, indent=2)
        except Exception as exc:
            self._log_recon(f"警告: 保存对账表头映射失败: {exc}")

    def _apply_recon_mapping(self, df, mapping, overwrite=False):
        if not mapping:
            return
        for target, source in mapping.items():
            if source in df.columns and (overwrite or target not in df.columns):
                df[target] = df[source]

    def _guess_recon_mapping(self, columns, targets):
        norm_cols = {col: normalize_header(col) for col in columns}
        mapping = {}
        for target in targets:
            best_col = None
            best_score = 0.0
            t_norm = normalize_header(target)
            for col, col_norm in norm_cols.items():
                if not t_norm or not col_norm:
                    continue
                if t_norm == col_norm:
                    score = 1.0
                elif t_norm in col_norm or col_norm in t_norm:
                    score = 0.85
                else:
                    score = difflib.SequenceMatcher(None, t_norm, col_norm).ratio()
                if score > best_score:
                    best_score = score
                    best_col = col
            if best_col and best_score >= 0.7:
                mapping[target] = best_col
        return mapping

    def _prompt_recon_header_mapping(self, title, columns, specs, preferred_mapping=None):
        if not columns:
            return None

        win = tk.Toplevel(self.root)
        win.title(title)
        win.transient(self.root)
        win.grab_set()
        win.geometry("640x520")

        info_text = "请选择每个标准字段对应的源表头，带 * 为必填。"
        ttk.Label(win, text=info_text, wraplength=600).pack(padx=10, pady=(10, 5))

        body = ttk.Frame(win)
        body.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        none_label = "(不使用)"
        col_labels = [str(c) for c in columns]
        display_to_col = {str(c): c for c in columns}
        guess_mapping = self._guess_recon_mapping(columns, [spec["key"] for spec in specs])

        vars_map = {}
        for row, spec in enumerate(specs):
            label = spec["label"]
            if spec.get("required"):
                label = f"{label} *"
            ttk.Label(body, text=label).grid(row=row, column=0, sticky="w", pady=2)
            var = tk.StringVar()
            combo = ttk.Combobox(
                body,
                textvariable=var,
                values=[none_label] + col_labels,
                state="readonly",
                width=40,
            )
            preferred = None
            if preferred_mapping:
                preferred = preferred_mapping.get(spec["key"])
                if preferred not in columns:
                    preferred = None
            if preferred:
                var.set(str(preferred))
            elif spec["key"] in columns:
                var.set(str(spec["key"]))
            elif spec["key"] in guess_mapping:
                var.set(str(guess_mapping[spec["key"]]))
            else:
                var.set(none_label)
            combo.grid(row=row, column=1, sticky="w", pady=2)
            vars_map[spec["key"]] = var

        body.columnconfigure(1, weight=1)

        btn_frame = ttk.Frame(win)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        def on_apply():
            missing = []
            for spec in specs:
                if not spec.get("required"):
                    continue
                if spec["key"] in columns:
                    continue
                sel = vars_map[spec["key"]].get()
                if not sel or sel == none_label:
                    missing.append(spec["label"])

            amount_sel = vars_map.get("金额", tk.StringVar()).get()
            debit_sel = vars_map.get("借方", tk.StringVar()).get()
            credit_sel = vars_map.get("贷方", tk.StringVar()).get()
            has_amount = ("金额" in columns) or (amount_sel and amount_sel != none_label)
            has_debit_credit = (
                ("借方" in columns) or (debit_sel and debit_sel != none_label)
            ) and (
                ("贷方" in columns) or (credit_sel and credit_sel != none_label)
            )
            if not has_amount and not has_debit_credit:
                missing.append("金额 (或 借方+贷方)")

            if missing:
                messagebox.showwarning(
                    "Missing Columns",
                    "请选择必填字段: " + ", ".join(missing),
                )
                return

            result = {}
            for spec in specs:
                sel = vars_map[spec["key"]].get()
                if sel and sel != none_label:
                    result[spec["key"]] = display_to_col.get(sel, sel)
            win.result = result
            win.destroy()

        def on_cancel():
            win.result = None
            win.destroy()

        ttk.Button(btn_frame, text="应用 (Apply)", command=on_apply).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="取消 (Cancel)", command=on_cancel).pack(side=tk.RIGHT)

        win.protocol("WM_DELETE_WINDOW", on_cancel)
        win.wait_window()
        return getattr(win, "result", None)

    def _is_recon_numeric_value(self, value):
        if pd.isna(value):
            return True
        if isinstance(value, (int, float, Decimal)):
            return True
        text = str(value).strip()
        if not text:
            return True
        if text.startswith("(") and text.endswith(")"):
            text = "-" + text[1:-1].strip()
        text = text.replace(",", "").replace(" ", "")
        if not text:
            return True
        if re.match(r"^[+-]?\d+(\.\d+)?$", text):
            return True
        try:
            float(text)
            return True
        except Exception:
            return False

    def _validate_recon_numeric_columns(self, df, columns, source_label):
        invalid_by_col = {}
        for col in columns:
            if col not in df.columns:
                continue
            samples = []
            for val in df[col].tolist():
                if not self._is_recon_numeric_value(val):
                    samples.append(val)
                    if len(samples) >= 5:
                        break
            if samples:
                invalid_by_col[col] = samples
        if not invalid_by_col:
            return True
        lines = [f"{col}: {samples}" for col, samples in invalid_by_col.items()]
        msg = (
            f"{source_label} 数值列包含非数字内容，请检查映射或源数据：\n"
            + "\n".join(lines)
        )
        self._log_recon(msg)
        messagebox.showerror("错误", msg)
        return False

    def _recon_column_has_data(self, df, col):
        if col not in df.columns:
            return False
        series = df[col].dropna()
        if series.empty:
            return False
        if series.dtype == object:
            return any(str(v).strip() for v in series.tolist())
        return True

    def _log_recon_df_stats(self, df, label):
        if df is None:
            return
        cols = [str(c) for c in df.columns]
        self._log_recon(f"{label}行数: {len(df)}; 列数: {len(cols)}")
        self._log_recon(f"{label}列: {cols}")
        if "凭证日期" in df.columns:
            dates = df["凭证日期"].dropna()
            if not dates.empty:
                sample = dates.head(3).astype(str).tolist()
                self._log_recon(f"{label}凭证日期样例: {sample}")
        if "金额" in df.columns:
            amounts = df["金额"].dropna()
            if not amounts.empty:
                sample = amounts.head(3).astype(str).tolist()
                self._log_recon(f"{label}金额样例: {sample}")

    def _ensure_recon_amount_type(self, df):
        if "金额" in df.columns:
            return
        if "借方" not in df.columns or "贷方" not in df.columns:
            return
        d_vals = pd.to_numeric(df["借方"], errors="coerce").fillna(0)
        c_vals = pd.to_numeric(df["贷方"], errors="coerce").fillna(0)
        df["金额"] = d_vals - c_vals
        if "类型" not in df.columns:
            types = []
            for d_val, c_val in zip(d_vals, c_vals):
                if abs(d_val) > 0.001:
                    types.append("3")
                elif abs(c_val) > 0.001:
                    types.append("4")
                else:
                    types.append("")
            df["类型"] = types

    def _recon_required_ready(self, df, require_code, require_date):
        has_amount = False
        if "金额" in df.columns:
            has_amount = self._recon_column_has_data(df, "金额")
        elif "借方" in df.columns and "贷方" in df.columns:
            has_amount = (
                self._recon_column_has_data(df, "借方")
                or self._recon_column_has_data(df, "贷方")
            )
        if not has_amount:
            return False
        if require_code and not self._recon_column_has_data(df, "往来单位编码"):
            return False
        if require_date and not self._recon_column_has_data(df, "凭证日期"):
            return False
        return True

    def _prepare_recon_dataframe(self, df, source_label, context_key, reconciler, recon_type, require_date):
        is_bank = "Bank" in recon_type or "银行" in recon_type
        is_custom = "Custom" in recon_type or "自定义" in recon_type
        if is_custom:
            require_date = True
        require_code = not is_bank and not is_custom

        original_columns = list(df.columns)
        saved_mapping = self._get_saved_recon_mapping(context_key, original_columns)
        if saved_mapping:
            self._apply_recon_mapping(df, saved_mapping, overwrite=False)
            self._log_recon(f"{source_label}已应用历史表头映射。")
            self._log_recon(f"{source_label}历史映射: {saved_mapping}")

        
        # Check force flag
        force_manual = False
        if hasattr(self, "recon_force_manual_mapping_var"):
            force_manual = self.recon_force_manual_mapping_var.get()

        if force_manual or not self._recon_required_ready(df, require_code, require_date):
            df_auto = reconciler.map_columns_smart(df)
            
            # Only use auto result if NOT forced and it is ready
            if not force_manual and self._recon_required_ready(df_auto, require_code, require_date):
                self._log_recon(f"{source_label}已使用自动表头识别。")
                df = df_auto
            else:
                seq_label = "序号/债券" if is_custom else "序号"
                specs = [
                    {"key": "凭证日期", "label": "凭证日期", "required": require_date},
                    {"key": "序号", "label": seq_label, "required": False},
                    {"key": "摘要", "label": "摘要", "required": False},
                    {"key": "往来单位编码", "label": "往来单位编码", "required": require_code},
                    {"key": "金额", "label": "金额", "required": False, "numeric": True},
                    {"key": "类型", "label": "类型", "required": False},
                    {"key": "借方", "label": "借方 (可选)", "required": False, "numeric": True},
                    {"key": "贷方", "label": "贷方 (可选)", "required": False, "numeric": True},
                ]
                reason = "强制人工干预" if force_manual else "表头不标准"
                self._log_recon(f"{source_label}{reason}，进入人工匹配...")
                mapping = self._prompt_recon_header_mapping(
                    f"{source_label}表头匹配",
                    original_columns,
                    specs,
                    preferred_mapping=saved_mapping,
                )
                if mapping is None:
                    self._log_recon("已取消人工表头匹配。")
                    return None
                df = df.copy()
                self._apply_recon_mapping(df, mapping, overwrite=True)
                self._log_recon(f"{source_label}人工映射: {mapping}")
                amount_mapped = ("金额" in mapping) or ("金额" in df.columns)
                if amount_mapped:
                    numeric_cols = ["金额"]
                else:
                    numeric_cols = [c for c in ["借方", "贷方"] if c in df.columns]
                if not self._validate_recon_numeric_columns(df, numeric_cols, source_label):
                    return None
                self._ensure_recon_amount_type(df)
                self._save_recon_mapping(context_key, original_columns, mapping)

        self._ensure_recon_amount_type(df)
        if is_bank and self.invert_bank_var.get():
            if not df.attrs.get("bank_inverted") and "借方" in df.columns and "贷方" in df.columns:
                if source_label.startswith("当地"):
                    self._log_recon(f"{source_label}应用银行借贷反转(Excel)...")
                    df["Temp"] = df["借方"]
                    df["借方"] = df["贷方"]
                    df["贷方"] = df["Temp"]
                    del df["Temp"]
                    df.attrs["bank_inverted"] = True
                    self._ensure_recon_amount_type(df)
        if is_bank and "往来单位编码" not in df.columns:
            df["往来单位编码"] = "BANK"

        counts = self._restore_codes_in_df(
            df,
            title=f"{source_label}智能还原编码",
            show_progress=False,
            show_alerts=False
        )
        if counts["account"] or counts["partner"]:
            self._log_recon(
                f"{source_label}智能还原编码: 科目={counts['account']} 条, 往来={counts['partner']} 条"
            )

        return df

    def _parse_recon_date_input(self, value, is_end=False):
        value = (value or "").strip()
        if not value:
            return None
        for fmt in ["%Y-%m-%d", "%Y%m%d", "%Y/%m/%d"]:
            try:
                return datetime.strptime(value, fmt)
            except Exception:
                pass
        match = re.match(r"^(\d{4})[-/](\d{1,2})$", value)
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            try:
                last_day = calendar.monthrange(year, month)[1]
            except Exception:
                return None
            day = last_day if is_end else 1
            return datetime(year, month, day)
        return None

    def _load_recon_local_df(self, local_path, recon_type):
        if local_path.lower().endswith(".pdf"):
            if not BankParser:
                self._log_recon("错误: 未找到 BankParser 模块，无法解析 PDF。请确保 bank_parser.py 存在。")
                messagebox.showerror("错误", "未找到 BankParser 模块，无法解析 PDF。")
                return None
            self._log_recon("检测到PDF文件，正在解析...")
            df_local = BankParser.parse_pdf(local_path, use_ocr=True)
            if df_local is None or df_local.empty:
                self._log_recon("错误: PDF解析失败或无数据（含OCR）。请确认PDF为文字版或检查OCR识别结果。")
                messagebox.showerror("错误", "PDF解析失败或无数据（含OCR），请确认PDF为文字版或检查OCR识别结果。")
                return None
            rename_map = {
                "Date": "凭证日期",
                "Doc": "序号",
                "Desc": "摘要",
                "Debit": "借方",
                "Credit": "贷方",
            }
            df_local.rename(columns=rename_map, inplace=True)
            df_local["往来单位编码"] = "BANK"
            if self.invert_bank_var.get():
                self._log_recon("应用银行借贷反转...")
                df_local["Temp"] = df_local["借方"]
                df_local["借方"] = df_local["贷方"]
                df_local["贷方"] = df_local["Temp"]
                del df_local["Temp"]
                df_local.attrs["bank_inverted"] = True
            try:
                excel_out = local_path + ".converted.xlsx"
                df_local.to_excel(excel_out, index=False)
                self._log_recon(f"PDF已转换为Excel: {os.path.basename(excel_out)}")
            except Exception:
                pass
            return df_local

        try:
            df_local = pd.read_excel(local_path)
        except Exception as exc:
            self._log_recon(f"错误: 读取当地系统Excel失败: {exc}")
            messagebox.showerror("错误", f"读取当地系统Excel失败: {exc}")
            return None
        return df_local

    def _load_recon_yikan_df(self, yikan_path):
        try:
            df_yikan = pd.read_excel(yikan_path)
        except Exception as exc:
            self._log_recon(f"错误: 读取亿看系统Excel失败: {exc}")
            messagebox.showerror("错误", f"读取亿看系统Excel失败: {exc}")
            return None
        return df_yikan

    def _show_recon_data_preview(self, df, title):
        win = tk.Toplevel(self.root)
        win.title(title)
        win.geometry("900x600")

        info_frame = ttk.Frame(win, padding=5)
        info_frame.pack(fill=tk.X)
        info = f"Rows: {len(df)} | Columns: {list(df.columns)}"
        ttk.Label(info_frame, text=info, wraplength=880).pack(side=tk.LEFT)

        frame = ttk.Frame(win)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = list(df.columns)
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(column=0, row=0, sticky="nsew")
        vsb.grid(column=1, row=0, sticky="ns")
        hsb.grid(column=0, row=1, sticky="ew")
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)

        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=100, minwidth=50)

        for _, row in df.head(200).iterrows():
            vals = [str(v) for v in row]
            tree.insert("", tk.END, values=vals)

        attach_treeview_tools(tree)

        win.transient(self.root)
        win.grab_set()

    def _preview_recon_local_data(self):
        local_path = self.recon_local_path.get().strip()
        if not local_path or not os.path.exists(local_path):
            messagebox.showwarning("提示", "请选择有效的当地系统文件")
            return
        recon_type = self.recon_type.get()
        df_local = self._load_recon_local_df(local_path, recon_type)
        if df_local is None:
            return
        recon_ai_context = self._build_ai_context("reconciliation") or self.summary_recognizer
        reconciler = StandardReconciler(self.base_data_mgr, recon_ai_context)
        reconciler.set_logger(self._log_recon)
        require_date = bool(self.recon_start_date.get().strip() or self.recon_end_date.get().strip())
        df_local = self._prepare_recon_dataframe(
            df_local,
            "当地系统",
            "local",
            reconciler,
            recon_type,
            require_date,
        )
        if df_local is None:
            return
        self._show_recon_data_preview(df_local, "当地系统预处理预览")

    def _preview_recon_yikan_data(self):
        yikan_path = self.recon_yikan_path.get().strip()
        if not yikan_path or not os.path.exists(yikan_path):
            messagebox.showwarning("提示", "请选择有效的亿看系统文件")
            return
        recon_type = self.recon_type.get()
        df_yikan = self._load_recon_yikan_df(yikan_path)
        if df_yikan is None:
            return
        recon_ai_context = self._build_ai_context("reconciliation") or self.summary_recognizer
        reconciler = StandardReconciler(self.base_data_mgr, recon_ai_context)
        reconciler.set_logger(self._log_recon)
        require_date = bool(self.recon_start_date.get().strip() or self.recon_end_date.get().strip())
        df_yikan = self._prepare_recon_dataframe(
            df_yikan,
            "亿看系统",
            "yikan",
            reconciler,
            recon_type,
            require_date,
        )
        if df_yikan is None:
            return
        self._show_recon_data_preview(df_yikan, "亿看系统预处理预览")

    def _show_reconciliation_result_preview(self, results, reconciler=None):
        """显示对账结果预览窗口"""
        preview_window = tk.Toplevel(self.root)
        preview_window.title("智能对账结果预览")
        preview_window.geometry("1000x700")
        
        # Summary
        matched = results['matched']
        unmatched_l = results['unmatched_local']
        unmatched_y = results['unmatched_yikan']
        direction_mismatch = results.get('direction_mismatch', pd.DataFrame())
        ai_sugg = results.get('ai_suggestions', pd.DataFrame())
        
        info_frame = ttk.Frame(preview_window, padding=10)
        info_frame.pack(fill="x")
        
        summary_text = (
            f"已匹配: {len(matched)} 行  |  "
            f"当地未匹配: {len(unmatched_l)} 行  |  "
            f"亿看未匹配: {len(unmatched_y)} 行  |  "
            f"方向不一致: {len(direction_mismatch)} 行  |  "
            f"AI建议: {len(ai_sugg)} 条"
        )
        ttk.Label(info_frame, text=summary_text, font=("Arial", 11, "bold")).pack(anchor="w")
        ttk.Label(info_frame, text="请确认以下结果，无误后点击“确认并导出”", foreground="gray").pack(anchor="w")

        # Notebook
        nb = ttk.Notebook(preview_window)
        nb.pack(fill="both", expand=True, padx=10, pady=5)
        
        def add_tab(name, df):
            frame = ttk.Frame(nb)
            nb.add(frame, text=f"{name} ({len(df)})")
            
            if df.empty:
                ttk.Label(frame, text="无数据").pack(padx=10, pady=10)
                return

            cols = list(df.columns)
            tree = ttk.Treeview(frame, columns=cols, show="headings")
            vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            tree.grid(column=0, row=0, sticky="nsew")
            vsb.grid(column=1, row=0, sticky="ns")
            hsb.grid(column=0, row=1, sticky="ew")
            frame.grid_columnconfigure(0, weight=1)
            frame.grid_rowconfigure(0, weight=1)
            
            for c in cols:
                tree.heading(c, text=c)
                tree.column(c, width=100, minwidth=50)
                
            for _, row in df.head(100).iterrows(): # Preview first 100
                vals = [str(v) for v in row]
                tree.insert("", "end", values=vals)
                
            if len(df) > 100:
                ttk.Label(frame, text=f"仅显示前 100 行预览，共 {len(df)} 行").grid(column=0, row=2, sticky="w")
            attach_treeview_tools(tree)

        add_tab("已匹配明细", matched)
        add_tab("当地未匹配", unmatched_l)
        add_tab("亿看未匹配", unmatched_y)
        if not direction_mismatch.empty:
            add_tab("方向不一致", direction_mismatch)
        if not ai_sugg.empty:
            add_tab("AI建议匹配", ai_sugg)

        # Buttons
        btn_frame = ttk.Frame(preview_window, padding=10)
        btn_frame.pack(fill="x")
        
        # --- AI Smart Analysis Button ---
        def do_ai_analysis():
            if not reconciler:
                messagebox.showerror("错误", "对账器未初始化，无法分析")
                return
            
            if not reconciler.summary_intelligence or not reconciler.summary_intelligence.ai_client:
                messagebox.showwarning("提示", "AI 模型未配置或不可用，请先在控制台或设置中配置 AI 模型。")
                return

            if unmatched_l.empty or unmatched_y.empty:
                messagebox.showinfo("提示", "未匹配项为空，无需分析")
                return
                
            btn_ai.config(state="disabled", text="AI 分析中...")
            self._log_recon("正在进行AI未匹配项分析(调用本地/远程模型)...")
            preview_window.update()
            
            try:
                # Call analysis
                suggestions = reconciler.analyze_mismatches_with_ai(unmatched_l, unmatched_y)
                
                btn_ai.config(state="normal", text="AI 智能分析未匹配项")
                self._log_recon(f"AI 分析完成，找到 {len(suggestions)} 条建议。")
                
                if not suggestions:
                    messagebox.showinfo("AI 分析结果", "AI 未发现明显的潜在匹配。")
                    return
                    
                # Show suggestions in new window
                ai_win = tk.Toplevel(preview_window)
                ai_win.title(f"AI 分析建议 ({len(suggestions)} 条)")
                ai_win.geometry("900x500")
                
                cols = list(suggestions[0].keys()) if suggestions else []
                tree = ttk.Treeview(ai_win, columns=cols, show="headings")
                vsb = ttk.Scrollbar(ai_win, orient="vertical", command=tree.yview)
                hsb = ttk.Scrollbar(ai_win, orient="horizontal", command=tree.xview)
                tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
                
                tree.pack(side="left", fill="both", expand=True)
                vsb.pack(side="right", fill="y")
                hsb.pack(side="bottom", fill="x")
                
                for c in cols:
                    tree.heading(c, text=c)
                    tree.column(c, width=150)
                    
                for s in suggestions:
                    vals = [str(s.get(c, "")) for c in cols]
                    tree.insert("", "end", values=vals)
                attach_treeview_tools(tree)
                
            except Exception as e:
                btn_ai.config(state="normal", text="AI 智能分析未匹配项")
                self._log_recon(f"AI 分析失败: {e}")
                messagebox.showerror("错误", f"AI 分析失败: {e}")

        btn_ai = ttk.Button(btn_frame, text="AI 智能分析未匹配项", command=do_ai_analysis, width=25)
        btn_ai.pack(side="left", padx=10)
        
        result = {"confirmed": False, "open_in_converter": False}
        
        def confirm(open_in_converter=False):
            result["confirmed"] = True
            result["open_in_converter"] = open_in_converter
            preview_window.destroy()
            
        def cancel():
            preview_window.destroy()
            
        ttk.Button(btn_frame, text="确认并导出", command=lambda: confirm(False)).pack(side="right", padx=10)
        ttk.Button(btn_frame, text="导出并转入凭证转换", command=lambda: confirm(True)).pack(side="right", padx=10)
        ttk.Button(btn_frame, text="取消", command=cancel).pack(side="right")
        
        preview_window.transient(self.root)
        preview_window.grab_set()
        self.root.wait_window(preview_window)
        
        return result

    def _start_reconciliation(self):
        local_path = self.recon_local_path.get()
        yikan_path = self.recon_yikan_path.get()
        recon_type = self.recon_type.get()
        
        if not local_path or not os.path.exists(local_path):
            messagebox.showwarning("提示", "请选择有效的当地系统文件")
            return
        if not yikan_path or not os.path.exists(yikan_path):
            messagebox.showwarning("提示", "请选择有效的亿看系统文件")
            return
            
        try:
            # 1. 解析日期
            s_date = None
            e_date = None
            s_str = self.recon_start_date.get().strip()
            e_str = self.recon_end_date.get().strip()
            
            if s_str:
                s_date = self._parse_recon_date_input(s_str, is_end=False)
            if e_str:
                e_date = self._parse_recon_date_input(e_str, is_end=True)
            
            if (s_str and not s_date) or (e_str and not e_date):
                self._log_recon("错误: 日期格式无效，请使用 YYYY-MM-DD 或 YYYY-MM")
                return

            # 2. 读取/解析文件
            self._log_recon("正在读取文件...")
            df_local = None
            
            # --- Local File Handling ---
            df_local = self._load_recon_local_df(local_path, recon_type)
            if df_local is None:
                return

            # --- Yikan File Handling ---
            df_yikan = self._load_recon_yikan_df(yikan_path)
            if df_yikan is None:
                return
            self._log_recon_df_stats(df_local, "当地系统原始")
            self._log_recon_df_stats(df_yikan, "亿看系统原始")
            
            # 3. 初始化对账器
            recon_ai_context = self._build_ai_context("reconciliation") or self.summary_recognizer
            reconciler = StandardReconciler(self.base_data_mgr, recon_ai_context)
            reconciler.set_logger(self._log_recon)

            # 3.5 智能表头识别 + 人工映射 + 持久化
            self._log_recon("检查表头格式...")
            require_date = bool(s_date or e_date)

            df_local = self._prepare_recon_dataframe(
                df_local,
                "当地系统",
                "local",
                reconciler,
                recon_type,
                require_date,
            )
            if df_local is None:
                return
            self._log_recon_df_stats(df_local, "当地系统预处理后")

            df_yikan = self._prepare_recon_dataframe(
                df_yikan,
                "亿看系统",
                "yikan",
                reconciler,
                recon_type,
                require_date,
            )
            if df_yikan is None:
                return
            self._log_recon_df_stats(df_yikan, "亿看系统预处理后")

            config = {
                'start_date': s_date,
                'end_date': e_date,
                'fuzzy_code': self.recon_fuzzy_var.get(),
                'use_ai_analysis': self.recon_use_ai_analysis.get(),
                'recon_type': recon_type,
                'require_same_direction': ("Bank" in recon_type or "银行" in recon_type),
            }
            
            # 4. 执行对账
            self._log_recon(f"开始执行对账逻辑 (类型: {recon_type})...")
            results = reconciler.reconcile(df_local, df_yikan, config)
            
            # 4.5 显示结果预览 (新增: 传入 reconciler)
            preview_result = self._show_reconciliation_result_preview(results, reconciler=reconciler)
            if not preview_result.get("confirmed"):
                self._log_recon("用户取消了导出。")
                return

            # 5. 导出结果
            self._log_recon("正在导出结果...")
            default_name = f"智能对账报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            out_path = filedialog.asksaveasfilename(initialfile=default_name, defaultextension=".xlsx")
            
            if out_path:
                voucher_path = None
                with pd.ExcelWriter(out_path) as writer:
                    # 汇总页
                    matched = results['matched']
                    unmatched_l = results['unmatched_local']
                    unmatched_y = results['unmatched_yikan']
                    direction_mismatch = results.get('direction_mismatch', pd.DataFrame())
                    ai_suggestions = results.get('ai_suggestions', pd.DataFrame())
                    
                    summary = pd.DataFrame({
                        '项目': ['已匹配', '当地未匹配 (需新增)', '亿看未匹配 (需减少)', '方向不一致 (需核对)', 'AI 建议匹配'],
                        '数量': [len(matched), len(unmatched_l), len(unmatched_y), len(direction_mismatch), len(ai_suggestions)]
                    })
                    summary.to_excel(writer, sheet_name='汇总', index=False)
                    
                    if not matched.empty:
                        matched.to_excel(writer, sheet_name='已匹配明细', index=False)
                    
                    if not unmatched_l.empty:
                        unmatched_l.to_excel(writer, sheet_name='当地未匹配(建议新增)', index=False)
                        
                    if not unmatched_y.empty:
                        unmatched_y.to_excel(writer, sheet_name='亿看未匹配(建议清理)', index=False)

                    if not direction_mismatch.empty:
                        direction_mismatch.to_excel(writer, sheet_name='方向不一致(需核对)', index=False)
                        
                    if not ai_suggestions.empty:
                        ai_suggestions.to_excel(writer, sheet_name='AI建议匹配', index=False)
                        
                self._log_recon(f"报告已保存至: {out_path}")
                
                # Export Voucher if requested
                if self.recon_export_voucher_var.get() and not unmatched_l.empty:
                    try:
                        import add_voucher_export
                        # Generate filename: same as report but with _voucher suffix
                        base, ext = os.path.splitext(out_path)
                        voucher_path = f"{base}_补录凭证{ext}"
                        
                        # We need to map the unmatched_l columns to what export_to_voucher expects
                        # export_to_voucher expects: '当地日期', '当地单号', '当地映射编码', '当地借方', '当地贷方', '当地摘要'
                        # But unmatched_l is in Standard Format: '凭证日期', '序号', '会计凭证No.', '摘要', '往来单位编码', '金额', '类型'
                        # We need to adapt it.
                        
                        # Adapter logic:
                        # Create a temp DF with expected columns
                        df_for_export = unmatched_l.copy()
                        df_for_export.rename(columns={
                            '凭证日期': '当地日期',
                            '序号': '当地单号',
                            '摘要': '当地摘要',
                            '往来单位编码': '当地映射编码'
                        }, inplace=True)
                        
                        # Handle Amount/Debit/Credit from Standard Format
                        # Standard has '金额' and '类型' (3=Debit, 4=Credit)
                        debits = []
                        credits = []
                        for _, row in df_for_export.iterrows():
                            amt = row.get('金额', 0)
                            try: amt = float(amt)
                            except: amt = 0
                            
                            t = str(row.get('类型', ''))
                            if t == '3':
                                debits.append(amt)
                                credits.append(0)
                            elif t == '4':
                                debits.append(0)
                                credits.append(amt)
                            else:
                                # Default to debit if positive?
                                debits.append(amt)
                                credits.append(0)
                        
                        df_for_export['当地借方'] = debits
                        df_for_export['当地贷方'] = credits
                        
                        add_voucher_export.export_to_voucher(df_for_export, voucher_path)
                        self._log_recon(f"凭证导入文件已生成: {voucher_path}")
                        
                    except Exception as e:
                        self._log_recon(f"凭证导出失败: {e}")
                        messagebox.showwarning("部分成功", f"对账报告已生成，但凭证导出失败: {e}")

                if preview_result.get("open_in_converter"):
                    self._open_in_converter(out_path, mode=MODE_ORIGINAL)

                # --- AI Analysis Integration ---
                if messagebox.askyesno("AI 智能分析", "对账报告已生成。是否立即启动本地 AI 对报告进行深度解读？\n(需确保 LM Studio 服务已开启)"):
                    def _run_ai_analysis():
                        try:
                            self.root.after(0, lambda: self._log_recon("启动本地 AI 分析..."))
                            from local_llm_analyzer import LocalLLMAnalyzer
                            backend = self._normalize_ai_backend(
                                self._get_ai_backend_for_task("report_analysis", allow_legacy=True)
                            )
                            analyzer = LocalLLMAnalyzer(
                                api_base=backend["base_url"],
                                model=backend["model"],
                                api_key=backend["api_key"],
                                provider=backend["provider"],
                            )
                            analysis_output = out_path.replace(".xlsx", "_AI_Analysis.md")
                            analyzer.analyze_report(out_path, analysis_output)
                            self.root.after(0, lambda: self._log_recon(f"AI 分析报告已生成: {analysis_output}"))
                            self.root.after(0, lambda: messagebox.showinfo("AI 分析完成", f"AI 解读报告已生成：\n{analysis_output}"))
                        except ImportError:
                            self.root.after(0, lambda: self._log_recon("错误: 未找到 local_llm_analyzer 模块。"))
                            self.root.after(0, lambda: messagebox.showerror("错误", "未找到 AI 分析模块 (local_llm_analyzer)。"))
                        except Exception as e:
                            self.root.after(0, lambda: self._log_recon(f"AI 分析失败: {e}"))
                            self.root.after(0, lambda: messagebox.showerror("AI 分析失败", str(e)))

                    threading.Thread(target=_run_ai_analysis, daemon=True).start()
                    messagebox.showinfo("AI 分析已启动", "AI 分析已在后台启动，完成后会提示。")
                else:
                    messagebox.showinfo("成功", "对账完成，报告已生成。")

            else:
                self._log_recon("导出已取消")

        except Exception as e:
            self._log_recon(f"发生错误: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("错误", f"对账过程中发生异常:\n{e}")

    def _build_console_tab(self):
        """构建控制台标签页"""
        console_frame = ttk.Frame(self.notebook)
        self.notebook.add(console_frame, text="控制台 (Console)")

        # 工具栏
        toolbar = ttk.Frame(console_frame)
        toolbar.pack(fill="x", padx=5, pady=5)

        ttk.Button(toolbar, text="清空日志", command=self.clear_console).pack(side="left", padx=5)
        ttk.Button(toolbar, text="导出日志", command=self.export_console_log).pack(side="left", padx=5)
        ttk.Button(toolbar, text="测试 AI 连接", command=self.test_ai_connection).pack(side="left", padx=5)
        ttk.Checkbutton(toolbar, text="调试日志", variable=self.debug_var).pack(side="left", padx=5)

        # 日志文本框
        self.console_text = scrolledtext.ScrolledText(console_frame, state="disabled", font=("Consolas", 9))
        self.console_text.pack(fill="both", expand=True, padx=5, pady=5)

        # 初始欢迎信息
        self.log_message("欢迎使用亿看智能识别系统！")
        try:
             self.log_message(f"当前模式: {self.convert_mode_var.get()}")
        except:
             pass

    def test_ai_connection(self):
        """测试 AI 连接"""
        if not self.summary_recognizer:
            messagebox.showerror("错误", "智能识别器未初始化")
            return
        
        self.log_message("正在测试 AI 连接...")
        
        try:
            # Check config
            provider = self.summary_recognizer.ai_provider
            base_url = self.summary_recognizer.ai_base_url
            model = self.summary_recognizer.ai_model_name
            
            self.log_message(f"配置信息: Provider={provider}, URL={base_url}, Model={model}")
            
            if not self.summary_recognizer.ai_client:
                self.log_message("错误: AI Client 为空，请检查配置（API Key/Provider）。")
                messagebox.showerror("连接失败", "AI Client 未初始化，请检查配置。")
                return

            # Send test request
            prompt = "Hello, are you working? Please reply with 'Yes'."
            
            # Compatible call
            if provider == "lm_studio" or True: # Use standard interface for all if wrapper supports it
                 response = self.summary_recognizer.ai_client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.1,
                    timeout=10 # Short timeout for test
                )
                 reply = response.choices[0].message.content.strip()
            
            self.log_message(f"AI 回复: {reply}")
            messagebox.showinfo("连接成功", f"AI 连接成功！\n\n回复: {reply}")
            
        except Exception as e:
            self.log_message(f"连接测试失败: {e}")
            messagebox.showerror("连接失败", f"测试失败:\n{e}\n\n请检查本地模型是否启动，或URL/端口是否正确。")

    def log_message(self, msg: str):
        """向控制台输出日志"""
        if not hasattr(self, "console_text"):
            print(msg) # Fallback
            return 
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        full_msg = f"[{timestamp}] {msg}\n"
        
        try:
            self.console_text.config(state="normal")
            self.console_text.insert("end", full_msg)
            self.console_text.see("end")
            self.console_text.config(state="disabled")
        except Exception as e:
            print(f"Error logging to console: {e}")
        
        # 同时打印到标准输出
        print(msg)

    def _set_status_text(self, text: str):
        label = getattr(self, "status_label", None)
        if label:
            try:
                label.config(text=text)
            except Exception:
                pass

    def _debug_log(self, msg: str):
        if getattr(self, "debug_var", None) is not None and self.debug_var.get():
            self.log_message(f"[DEBUG] {msg}")

    def _is_tk_crash_prone_runtime(self) -> bool:
        """检测已知的 Tk 崩溃高风险运行时（Windows 商店版 Python 3.13 + Tk 8.6.15）"""
        try:
            exe = (sys.executable or "").lower()
            is_store_python = ("windowsapps" in exe) and ("pythonsoftwarefoundation.python" in exe)
            tk_patch = str(self.root.tk.call("info", "patchlevel"))
            return is_store_python and sys.version_info >= (3, 13) and tk_patch.startswith("8.6.15")
        except Exception:
            return False

    def _should_skip_preprocessing_preview(self) -> bool:
        """是否应跳过预处理预览"""
        return False

    def clear_console(self):
        self.console_text.config(state="normal")
        self.console_text.delete("1.0", "end")
        self.console_text.config(state="disabled")

    def export_console_log(self):
        log_content = self.console_text.get("1.0", "end")
        if not log_content.strip():
            messagebox.showinfo("提示", "日志为空，无需导出。")
            return
            
        f = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")],
            initialfile=f"ConsoleLog_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        if f:
            try:
                with open(f, "w", encoding="utf-8") as file:
                    file.write(log_content)
                messagebox.showinfo("成功", f"日志已导出至: {f}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {e}")

    def _build_excel_converter_tab(self):
        """构建Excel转换标签页"""
        excel_frame = ttk.Frame(self.notebook)
        self.excel_tab_frame = excel_frame
        self.notebook.add(excel_frame, text="Excel凭证转换")

        # 顶部：文件选择
        top = ttk.Frame(excel_frame, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="原始 Excel：").grid(row=0, column=0, sticky="e")
        entry = ttk.Entry(top, textvariable=self.input_path_var, width=50)
        entry.grid(row=0, column=1, sticky="we", padx=5)
        ttk.Button(top, text="浏览...", command=self.select_input_file).grid(
            row=0, column=2, padx=5
        )

        ttk.Label(top, text="工作表：").grid(row=1, column=0, sticky="e", pady=(5, 0))
        self.sheet_combo = ttk.Combobox(top, textvariable=self.sheet_var, state="disabled", width=20)
        self.sheet_combo.grid(row=1, column=1, sticky="w", padx=5, pady=(5, 0))
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_changed)

        # 转换模式
        ttk.Label(top, text="转换模式：").grid(row=2, column=0, sticky="e", pady=(5, 0))
        self.mode_combo = ttk.Combobox(
            top,
            textvariable=self.convert_mode_var,
            values=[MODE_GENERAL_VOUCHER, MODE_SALES_OUTBOUND, MODE_CUSTOM, MODE_ORIGINAL],
            state="readonly",
            width=20
        )
        self.mode_combo.grid(row=2, column=1, sticky="w", padx=5, pady=(5, 0))
        self.mode_combo.bind("<<ComboboxSelected>>", self._on_mode_changed)
        
        ttk.Label(top, text="表头方案：").grid(row=2, column=2, sticky="e", pady=(5, 0))
        self.header_scheme_combo = ttk.Combobox(
            top,
            textvariable=self.header_scheme_var,
            values=sorted(HEADER_SCHEMES.keys()),
            state="readonly",
            width=20
        )
        self.header_scheme_combo.grid(row=2, column=3, sticky="w", padx=5, pady=(5, 0))
        self.header_scheme_combo.bind("<<ComboboxSelected>>", lambda e: self._apply_header_scheme_to_mapping())
        
        self.export_scheme_chk = ttk.Checkbutton(
            top,
            text="按表头方案导出(忽略模板样式)",
            variable=self.export_scheme_override_var
        )
        self.export_scheme_chk.grid(row=3, column=1, columnspan=2, sticky="w", padx=5, pady=(4, 0))

        # 目标模板 (Custom Mode)
        ttk.Label(top, text="目标模板：").grid(row=3, column=0, sticky="e", pady=(5, 0))
        self.template_entry = ttk.Entry(top, textvariable=self.template_path_var, width=50, state="readonly")
        self.template_entry.grid(row=3, column=1, sticky="we", padx=5, pady=(5, 0))
        self.template_btn = ttk.Button(top, text="浏览...", command=self.select_template_file, state="disabled")
        self.template_btn.grid(row=3, column=2, padx=5, pady=(5, 0))
        self.preserve_mapping_var = tk.BooleanVar(value=False)
        self.preserve_mapping_chk = ttk.Checkbutton(
            top,
            text="保留当前映射",
            variable=self.preserve_mapping_var,
            state="disabled"
        )
        self.preserve_mapping_chk.grid(row=3, column=3, padx=5, pady=(5, 0))

        # 智能/高级选项已移至“设置”菜单

        # 借贷列辅助选择（模板仅有单列金额时使用）
        dc_frame = ttk.LabelFrame(top, text="借贷列辅助（模板只有“金额”时请设置）")
        dc_frame.grid(row=4, column=1, columnspan=2, sticky="we", padx=5, pady=(6, 0))
        ttk.Label(dc_frame, text="借方列:").grid(row=0, column=0, sticky="e", padx=5, pady=3)
        self.debit_col_combo = ttk.Combobox(
            dc_frame,
            textvariable=self.manual_debit_col_var,
            state="readonly",
            width=28,
            values=[EMPTY_OPTION] + self.input_columns
        )
        self.debit_col_combo.grid(row=0, column=1, sticky="w", padx=3, pady=3)

        ttk.Label(dc_frame, text="贷方列:").grid(row=0, column=2, sticky="e", padx=5, pady=3)
        self.credit_col_combo = ttk.Combobox(
            dc_frame,
            textvariable=self.manual_credit_col_var,
            state="readonly",
            width=28,
            values=[EMPTY_OPTION] + self.input_columns
        )
        self.credit_col_combo.grid(row=0, column=3, sticky="w", padx=3, pady=3)

        ttk.Label(top, text="导出格式：").grid(row=5, column=0, sticky="e", pady=(6, 0))
        self.export_format_combo = ttk.Combobox(
            top,
            textvariable=self.export_format_var,
            values=get_export_format_names("main_export"),
            state="readonly",
            width=20
        )
        self.export_format_combo.grid(row=5, column=1, sticky="w", padx=5, pady=(6, 0))
        self.export_format_combo.bind("<<ComboboxSelected>>", self._on_export_format_changed)
        ttk.Button(top, text="设置...", command=self._open_export_format_editor).grid(
            row=5, column=2, padx=5, pady=(6, 0)
        )

        # 提示信息
        ttk.Label(
            top,
            text="注：开始转换前会先询问“对方科目/默认账户/是否外币/汇率”，也可在基础数据中预设默认值",
            foreground="gray"
        ).grid(row=6, column=1, columnspan=2, sticky="w", padx=5)

        # 按钮区 (默认值 & 方案管理)
        btn_frame = ttk.Frame(top)
        btn_frame.grid(row=4, column=2, rowspan=2, padx=5, sticky="n")

        ttk.Button(btn_frame, text="保存当前映射...", command=self._save_current_scheme, width=15).pack(pady=2)
        ttk.Button(btn_frame, text="综合字段...", command=self._open_composite_field_dialog, width=15).pack(pady=2)
        ttk.Button(btn_frame, text="删除当前方案", command=self._delete_current_scheme, width=15).pack(pady=2)

        top.grid_columnconfigure(1, weight=1)
        
        # 初始化方案列表
        self.cached_schemes = {}
        self._load_schemes_to_combo()

        # 中部：字段映射
        mid = ttk.LabelFrame(excel_frame, text="字段映射（模板列 -> 原始列）", padding=10)
        mid.pack(fill="both", expand=True, padx=10, pady=5)
        # 可滚动区域，避免表头过多时看不到后续行
        self.mapping_canvas = tk.Canvas(mid, highlightthickness=0)
        mapping_scrollbar = ttk.Scrollbar(mid, orient="vertical", command=self.mapping_canvas.yview)
        mapping_scrollbar_x = ttk.Scrollbar(mid, orient="horizontal", command=self.mapping_canvas.xview)
        self.mapping_container = ttk.Frame(self.mapping_canvas)
        self.mapping_container.bind(
            "<Configure>",
            lambda e: self.mapping_canvas.configure(scrollregion=self.mapping_canvas.bbox("all"))
        )
        self.mapping_canvas.create_window((0, 0), window=self.mapping_container, anchor="nw")
        self.mapping_canvas.configure(yscrollcommand=mapping_scrollbar.set, xscrollcommand=mapping_scrollbar_x.set)
        self.mapping_canvas.pack(side="left", fill="both", expand=True)
        mapping_scrollbar.pack(side="right", fill="y")
        mapping_scrollbar_x.pack(side="bottom", fill="x")
        self.mapping_frame = self.mapping_container

        # 底部：操作按钮
        bottom = ttk.Frame(excel_frame, padding=10)
        bottom.pack(fill="x")

        ttk.Button(bottom, text="自动识别匹配", command=self.auto_match).pack(side="left")
        ttk.Button(bottom, text="预览映射", command=self._preview_mapping).pack(side="left", padx=10)
        ttk.Button(bottom, text="智能还原编码", command=self._restore_base_data_codes).pack(side="left", padx=10)
        ttk.Button(bottom, text="开始转换并导出", command=self.do_convert).pack(side="right")

        self._refresh_export_format_options()

    def _build_summary_match_tab(self):
        """构建摘要匹配标签页"""
        match_frame = ttk.Frame(self.notebook)
        self.notebook.add(match_frame, text="摘要匹配")

        main_frame = ttk.Frame(match_frame, padding=10)
        main_frame.pack(fill="both", expand=True)

        file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding=10)
        file_frame.pack(fill="x", pady=(0, 8))

        target_frame = ttk.LabelFrame(file_frame, text="替换摘要文件", padding=8)
        target_frame.pack(fill="x", pady=(0, 6))

        ttk.Label(target_frame, text="文件:").grid(row=0, column=0, sticky="e")
        ttk.Entry(target_frame, textvariable=self.summary_match_target_path_var, width=60).grid(
            row=0, column=1, sticky="we", padx=5
        )
        ttk.Button(target_frame, text="浏览...", command=lambda: self._select_summary_match_file("target")).grid(
            row=0, column=2, padx=5
        )

        ttk.Label(target_frame, text="工作表:").grid(row=1, column=0, sticky="e", pady=(5, 0))
        self.summary_match_target_sheet_combo = ttk.Combobox(
            target_frame, textvariable=self.summary_match_target_sheet_var, state="disabled", width=20
        )
        self.summary_match_target_sheet_combo.grid(row=1, column=1, sticky="w", padx=5, pady=(5, 0))
        self.summary_match_target_sheet_combo.bind(
            "<<ComboboxSelected>>", lambda e: self._load_summary_match_sheet("target")
        )
        target_frame.columnconfigure(1, weight=1)

        source_frame = ttk.LabelFrame(file_frame, text="摘要源文件明细", padding=8)
        source_frame.pack(fill="x")

        ttk.Label(source_frame, text="文件:").grid(row=0, column=0, sticky="e")
        ttk.Entry(source_frame, textvariable=self.summary_match_source_path_var, width=60).grid(
            row=0, column=1, sticky="we", padx=5
        )
        ttk.Button(source_frame, text="浏览...", command=lambda: self._select_summary_match_file("source")).grid(
            row=0, column=2, padx=5
        )

        ttk.Label(source_frame, text="工作表:").grid(row=1, column=0, sticky="e", pady=(5, 0))
        self.summary_match_source_sheet_combo = ttk.Combobox(
            source_frame, textvariable=self.summary_match_source_sheet_var, state="disabled", width=20
        )
        self.summary_match_source_sheet_combo.grid(row=1, column=1, sticky="w", padx=5, pady=(5, 0))
        self.summary_match_source_sheet_combo.bind(
            "<<ComboboxSelected>>", lambda e: self._load_summary_match_sheet("source")
        )
        source_frame.columnconfigure(1, weight=1)

        column_frame = ttk.LabelFrame(main_frame, text="列选择", padding=10)
        column_frame.pack(fill="x", pady=(0, 8))

        col_target_frame = ttk.LabelFrame(column_frame, text="替换摘要文件列", padding=8)
        col_target_frame.pack(side="left", fill="both", expand=True, padx=(0, 6))
        self._build_summary_match_column_selectors(col_target_frame, self.summary_match_target_vars)

        col_source_frame = ttk.LabelFrame(column_frame, text="摘要源文件列", padding=8)
        col_source_frame.pack(side="left", fill="both", expand=True)
        self._build_summary_match_column_selectors(col_source_frame, self.summary_match_source_vars)

        settings_frame = ttk.LabelFrame(main_frame, text="匹配规则", padding=10)
        settings_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(settings_frame, text="日期允许差值(天):").grid(row=0, column=0, sticky="e")
        ttk.Entry(settings_frame, textvariable=self.summary_match_date_tol_var, width=8).grid(
            row=0, column=1, sticky="w", padx=5
        )

        ttk.Label(settings_frame, text="金额绝对差值:").grid(row=0, column=2, sticky="e")
        ttk.Entry(settings_frame, textvariable=self.summary_match_amount_abs_tol_var, width=10).grid(
            row=0, column=3, sticky="w", padx=5
        )

        ttk.Label(settings_frame, text="金额百分比差值(%):").grid(row=0, column=4, sticky="e")
        ttk.Entry(settings_frame, textvariable=self.summary_match_amount_pct_tol_var, width=8).grid(
            row=0, column=5, sticky="w", padx=5
        )

        # 新增：日期解析选项
        self.summary_match_target_dayfirst_var = tk.BooleanVar(value=False)
        self.summary_match_source_dayfirst_var = tk.BooleanVar(value=True) # 默认源文件（通常是银行流水）为日领先
        self.summary_match_use_local_ai_var = tk.BooleanVar(value=False) # 新增本地AI开关
        
        ttk.Checkbutton(settings_frame, text="目标文件日领先(DD-MM)", variable=self.summary_match_target_dayfirst_var).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=5, pady=2
        )
        ttk.Checkbutton(settings_frame, text="源文件日领先(DD-MM)", variable=self.summary_match_source_dayfirst_var).grid(
            row=1, column=2, columnspan=2, sticky="w", padx=5, pady=2
        )
        ttk.Checkbutton(settings_frame, text="启用本地AI辅助匹配", variable=self.summary_match_use_local_ai_var).grid(
            row=1, column=4, columnspan=2, sticky="w", padx=5, pady=2
        )

        ttk.Checkbutton(settings_frame, text="不重复匹配", variable=self.summary_match_unique_var).grid(
            row=0, column=6, padx=8
        )
        ttk.Checkbutton(settings_frame, text="保留原摘要列", variable=self.summary_match_keep_original_var).grid(
            row=0, column=7, padx=8
        )
        ttk.Checkbutton(settings_frame, text="同步借贷方向/正负号", variable=self.summary_match_sync_direction_var).grid(
            row=1, column=6, padx=8
        )
        # ttk.Checkbutton(settings_frame, text="填充缺失记录", variable=self.summary_match_fill_missing_var).grid(
        #    row=1, column=7, padx=8
        # )

        settings_frame.columnconfigure(7, weight=1)

        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill="x", pady=(0, 8))
        ttk.Button(action_frame, text="开始匹配并替换", command=self._run_summary_match, width=18).pack(
            side="left"
        )
        self.summary_match_export_btn = ttk.Button(
            action_frame, text="导出结果", command=self._export_summary_match_result, width=12, state="disabled"
        )
        self.summary_match_export_btn.pack(side="left", padx=6)
        ttk.Button(
            action_frame,
            text="智能还原编码",
            command=self._restore_summary_match_codes,
            width=12
        ).pack(side="left", padx=6)

        ttk.Label(action_frame, text="导出格式:").pack(side="left")
        self.summary_match_export_format_combo = ttk.Combobox(
            action_frame,
            textvariable=self.summary_match_export_format_var,
            values=get_export_format_names("summary_match_export"),
            state="readonly",
            width=14
        )
        self.summary_match_export_format_combo.pack(side="left", padx=4)
        self.summary_match_export_format_combo.bind("<<ComboboxSelected>>", self._on_summary_match_export_format_changed)
        ttk.Button(
            action_frame,
            text="设置",
            command=self._open_summary_match_export_format_editor,
            width=6
        ).pack(side="left", padx=4)

        ttk.Label(action_frame, textvariable=self.summary_match_status_var, foreground="gray").pack(
            side="left", padx=10
        )

        preview_frame = ttk.LabelFrame(main_frame, text="预览 (前30行)", padding=10)
        preview_frame.pack(fill="both", expand=True)

        self.summary_match_preview_container = ttk.Frame(preview_frame)
        self.summary_match_preview_container.pack(fill="both", expand=True)

        self._refresh_summary_match_export_format_options()

    def _build_summary_match_column_selectors(self, parent, vars_map):
        none_label = "(不使用)"
        labels = [
            ("summary", "摘要列"),
            ("date", "日期列"),
            ("amount", "金额列"),
            ("debit", "借方列"),
            ("credit", "贷方列"),
        ]
        for row, (key, label) in enumerate(labels):
            ttk.Label(parent, text=f"{label}:").grid(row=row, column=0, sticky="e", pady=3)
            combo = ttk.Combobox(parent, textvariable=vars_map[key], values=[none_label], state="readonly", width=26)
            combo.grid(row=row, column=1, sticky="w", padx=5, pady=3)
            vars_map[f"{key}_combo"] = combo
            vars_map[key].set(none_label)
        parent.columnconfigure(1, weight=1)

    def _select_summary_match_file(self, kind):
        path = filedialog.askopenfilename(
            title="选择数据文件 (Excel 或 PDF)",
            filetypes=[("Excel 文件", "*.xlsx;*.xls"), ("PDF 文件", "*.pdf"), ("所有文件", "*.*")],
        )
        if not path:
            return
        if kind == "target":
            self.summary_match_target_path_var.set(path)
        else:
            self.summary_match_source_path_var.set(path)
        
        # 处理 PDF
        if path.lower().endswith(".pdf"):
            if kind == "target":
                self.summary_match_target_sheet_var.set("[PDF 数据]")
                self.summary_match_target_sheet_combo["values"] = ["[PDF 数据]"]
            else:
                self.summary_match_source_sheet_var.set("[PDF 数据]")
                self.summary_match_source_sheet_combo["values"] = ["[PDF 数据]"]
            self._load_summary_match_sheet(kind)
            return

        try:
            sheet_names = self._get_excel_sheet_names(path)
        except Exception as e:
            messagebox.showerror("错误", f"读取工作表列表失败：\n{e}")
            return
        if kind == "target":
            combo = self.summary_match_target_sheet_combo
            var = self.summary_match_target_sheet_var
        else:
            combo = self.summary_match_source_sheet_combo
            var = self.summary_match_source_sheet_var
        combo["values"] = sheet_names
        combo.configure(state="readonly")
        if sheet_names:
            var.set(sheet_names[0])
            self._load_summary_match_sheet(kind)

    def _load_summary_match_sheet(self, kind):
        path = self.summary_match_target_path_var.get() if kind == "target" else self.summary_match_source_path_var.get()
        sheet = self.summary_match_target_sheet_var.get() if kind == "target" else self.summary_match_source_sheet_var.get()
        if not path or not sheet:
            return
        try:
            if path.lower().endswith(".pdf"):
                self.summary_match_status_var.set("正在解析 PDF，请稍候...")
                self.root.update_idletasks()
                df = BankParser.parse_pdf(path)
                if df is None or df.empty:
                    messagebox.showwarning("解析失败", "PDF 中未识别到有效的银行流水表格。")
                    return
            else:
                df = self._read_excel_sheet(path, sheet)
        except Exception as e:
            messagebox.showerror("错误", f"读取数据失败：\n{e}")
            return
        if kind == "target":
            self.summary_match_target_df = df
            self.summary_match_target_columns = [str(c) for c in df.columns]
            self._update_summary_match_column_options("target")
        else:
            self.summary_match_source_df = df
            self.summary_match_source_columns = [str(c) for c in df.columns]
            self._update_summary_match_column_options("source")
        self.summary_match_status_var.set("数据加载成功，等待匹配...")

    def _update_summary_match_column_options(self, kind):
        if kind == "target":
            columns = self.summary_match_target_columns
            vars_map = self.summary_match_target_vars
        else:
            columns = self.summary_match_source_columns
            vars_map = self.summary_match_source_vars

        none_label = "(不使用)"
        for key in ["summary", "date", "amount", "debit", "credit"]:
            combo = vars_map.get(f"{key}_combo")
            if not combo:
                continue
            combo["values"] = [none_label] + columns

        guesses = {
            "summary": self._guess_summary_match_column(columns, ["摘要", "摘要名", "说明", "备注", "description", "descrip", "desc"]),
            "date": self._guess_summary_match_column(columns, ["凭证日期", "日期", "date"]),
            "amount": self._guess_summary_match_column(columns, ["金额", "amount", "本币", "金额本币", "原币", "balance"]),
            "debit": self._guess_summary_match_column(columns, ["借方", "借", "debit"]),
            "credit": self._guess_summary_match_column(columns, ["贷方", "贷", "credit"]),
        }

        for key, guess in guesses.items():
            current = vars_map[key].get()
            if current not in columns:
                if guess:
                    vars_map[key].set(guess)
                else:
                    vars_map[key].set(none_label)

    def _guess_summary_match_column(self, columns, keywords):
        if not columns:
            return ""
        norm_keywords = [normalize_header(k) for k in keywords]
        for col in columns:
            norm_col = normalize_header(col)
            for nk in norm_keywords:
                if nk and nk in norm_col:
                    return col
        return ""

    def _parse_summary_match_date(self, value, dayfirst=False):
        if value is None or pd.isna(value):
            return None
        if isinstance(value, (pd.Timestamp, datetime)):
            return value.date()
        if isinstance(value, date):
            return value
            
        s_val = str(value).strip()
        if not s_val:
            return None

        # 1. 尝试使用 pd.to_datetime 解析
        try:
            # 优先处理带有斜杠的格式
            if "/" in s_val:
                # 尝试多种组合解析
                formats = ["%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%d/%m/%y"]
                if dayfirst:
                    formats = ["%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y", "%m/%d/%y"]
                
                for fmt in formats:
                    try:
                        dt = datetime.strptime(s_val, fmt)
                        # 处理 2 位年份
                        if dt.year < 100:
                            year = 2000 + dt.year
                            dt = dt.replace(year=year)
                        return dt.date()
                    except ValueError:
                        continue

            dt = pd.to_datetime(s_val, dayfirst=dayfirst, errors="coerce")
            if pd.notna(dt):
                # 特殊处理 2 位年份：pd.to_datetime 默认行为可能不一致
                y = dt.year
                if y < 100:
                    y += 2000
                elif 70 <= y <= 99:
                    y += 1900
                elif 0 <= y <= 69:
                    y += 2000
                return date(y, dt.month, dt.day)
        except Exception:
            pass
            
        # 2. 兜底尝试数字解析 (Excel 序列号)
        if s_val.replace(".", "").isdigit():
            try:
                dt = pd.to_datetime(float(s_val), unit="d", origin="1899-12-30", errors="coerce")
                return dt.date() if pd.notna(dt) else None
            except Exception:
                pass
                
        return None

    def _parse_summary_match_number(self, value):
        if value is None or pd.isna(value):
            return None
        if isinstance(value, (int, float, Decimal)):
            return float(value)
        text = str(value).strip()
        if not text:
            return None
        if text.startswith("(") and text.endswith(")"):
            text = "-" + text[1:-1].strip()
        text = text.replace(",", "").replace(" ", "")
        try:
            return float(text)
        except Exception:
            return None

    def _extract_summary_match_amount(self, row, mapping):
        amount_col = mapping.get("amount")
        debit_col = mapping.get("debit")
        credit_col = mapping.get("credit")
        direction = None
        amount = None
        
        # 1. 优先尝试从金额列提取
        if amount_col:
            amount = self._parse_summary_match_number(row.get(amount_col))
            
        # 2. 如果金额为空（或为0且有借贷列），尝试从借方/贷方提取
        if amount is None or (isinstance(amount, float) and abs(amount) < 0.0001):
            debit_val = self._parse_summary_match_number(row.get(debit_col)) if debit_col else None
            credit_val = self._parse_summary_match_number(row.get(credit_col)) if credit_col else None
            
            if debit_val is not None and abs(debit_val) > 0.0001:
                amount = abs(debit_val)
                direction = "debit"
            elif credit_val is not None and abs(credit_val) > 0.0001:
                amount = abs(credit_val)
                direction = "credit"
                
        if amount is None:
            return None, None
        return abs(amount), direction

    def _summary_match_get_mapping(self, kind):
        vars_map = self.summary_match_target_vars if kind == "target" else self.summary_match_source_vars
        none_label = "(不使用)"
        mapping = {}
        for key in ["summary", "date", "amount", "debit", "credit"]:
            val = vars_map[key].get()
            if val and val != none_label:
                mapping[key] = val
        return mapping

    def _ensure_unique_column_name(self, columns, base_name):
        if base_name not in columns:
            return base_name
        idx = 1
        while True:
            name = f"{base_name}_{idx}"
            if name not in columns:
                return name
            idx += 1



    def _run_summary_match(self):
        if self.summary_match_target_df is None or self.summary_match_source_df is None:
            messagebox.showwarning("提示", "请先加载替换摘要文件和摘要源文件ảng。")
            return

        target_mapping = self._summary_match_get_mapping("target")
        source_mapping = self._summary_match_get_mapping("source")

        for label, mapping in [("替换摘要文件", target_mapping), ("摘要源文件", source_mapping)]:
            if "summary" not in mapping or "date" not in mapping:
                messagebox.showwarning("提示", f"{label}必须选择摘要列和日期列ảng بمن。")
                return
            has_amount = "amount" in mapping
            has_debit_credit = "debit" in mapping or "credit" in mapping
            if not has_amount and not has_debit_credit:
                messagebox.showwarning("提示", f"{label}请至少选择金额列，或借方/贷方列ảng بمن。")
                return

        try:
            date_tol_days = int(float(self.summary_match_date_tol_var.get().strip() or "0"))
            amount_abs_tol = float(self.summary_match_amount_abs_tol_var.get().strip() or "0")
            amount_pct_tol_raw = float(self.summary_match_amount_pct_tol_var.get().strip() or "0")
            if amount_pct_tol_raw >= 100:
                self.log_message("⚠️ 警告: 金额百分比差值设为 100% 或更高，容易产生误匹配ảng بمن بمن。")
            amount_pct_tol = amount_pct_tol_raw / 100.0 if amount_pct_tol_raw > 1 else amount_pct_tol_raw
        except Exception:
            messagebox.showerror("错误", "配置参数必须是数字ảng بمن。")
            return

        t_dayfirst = self.summary_match_target_dayfirst_var.get()
        s_dayfirst = self.summary_match_source_dayfirst_var.get()
        use_ai_assistant = self.summary_match_use_local_ai_var.get()

        self.log_message(f"开始摘要匹配 (日期容差: {date_tol_days}天, 金额容差: {amount_abs_tol}, 百分比: {amount_pct_tol*100:.1f}%)ảng بمن")
        self.log_message(f"配置: 日期规则[T:{'DD-MM' if t_dayfirst else 'MM-DD'}, S:{'DD-MM' if s_dayfirst else 'MM-DD'}], AI辅助:{'开启' if use_ai_assistant else '关闭'}ảng بمن")

        source_entries = []
        source_by_date = {}
        for idx, row in self.summary_match_source_df.iterrows():
            s_date = self._parse_summary_match_date(row.get(source_mapping["date"]), dayfirst=s_dayfirst)
            s_amount, s_dir = self._extract_summary_match_amount(row, source_mapping)
            if s_date is None or s_amount is None: continue
            entry = {
                "idx": idx,
                "date": s_date,
                "amount": s_amount,
                "direction": s_dir,
                "summary": str(row.get(source_mapping["summary"], "")),
            }
            source_entries.append(entry)
            source_by_date.setdefault(s_date, []).append(entry)

        if not source_entries:
            messagebox.showwarning("提示", "摘要源文件没有可用的记录ảng بمن。")
            return

        result_df = self.summary_match_target_df.copy()
        target_summary_col = target_mapping["summary"]
        if self.summary_match_keep_original_var.get():
            new_col = self._ensure_unique_column_name(list(result_df.columns), "原摘要")
            result_df[new_col] = result_df[target_summary_col]

        used_sources = set()
        matched_count = 0
        preview_rows = []

        self.log_message("--- 匹配明细 --ảng بمن")
        for idx, row in result_df.iterrows():
            t_date = self._parse_summary_match_date(row.get(target_mapping["date"]), dayfirst=t_dayfirst)
            t_amount, t_dir = self._extract_summary_match_amount(row, target_mapping)
            original_summary = str(row.get(target_summary_col, ""))
            
            # 准备预览用的初始借贷/金额
            p_debit, p_credit, p_signed_amt = "", "", ""
            if "debit" in target_mapping or "credit" in target_mapping:
                p_debit = self._parse_summary_match_number(row.get(target_mapping.get("debit"))) if "debit" in target_mapping else 0
                p_credit = self._parse_summary_match_number(row.get(target_mapping.get("credit"))) if "credit" in target_mapping else 0
                p_debit = p_debit if p_debit else ""
                p_credit = p_credit if p_credit else ""
            else:
                p_signed_amt = t_amount if t_dir != "credit" else -t_amount

            if t_date is None or t_amount is None:
                if p_signed_amt != "":
                    preview_rows.append(["未匹配", t_date, p_signed_amt, original_summary, ""])
                else:
                    preview_rows.append(["未匹配", t_date, p_debit, p_credit, original_summary, ""])
                continue

            best = None
            for delta in range(-date_tol_days, date_tol_days + 1):
                cand_date = t_date + timedelta(days=delta)
                for entry in source_by_date.get(cand_date, []):
                    if self.summary_match_unique_var.get() and entry["idx"] in used_sources: continue
                    
                    # --- 方向校验容错逻辑 ---
                    # 如果启用了“同步方向”选项，则不强制校验方向一致性（因为目标文件可能本身就是错的，需要靠匹配来纠正）
                    # 否则，只有当双方都能明确识别出借/贷方向时，才进行严格校验
                    if not self.summary_match_sync_direction_var.get():
                        if t_dir and entry["direction"]:
                            if t_dir != entry["direction"]:
                                continue
                    
                    diff = abs(entry["amount"] - t_amount)
                    tol = max(amount_abs_tol, abs(t_amount) * amount_pct_tol)
                    if diff > tol + 0.0001: continue

                    t_text = original_summary.strip()
                    s_text = entry["summary"].strip()
                    text_score = score_similarity(t_text, s_text, t_text) if t_text and s_text else 0.0
                    if use_ai_assistant and self.summary_recognizer and (0.1 < text_score < 0.8):
                        ai_score = self.summary_recognizer.calculate_ai_similarity(t_text, s_text)
                        text_score = max(text_score, ai_score)

                    # 计算方向不匹配惩罚 (新逻辑)
                    dir_penalty = 0.0
                    if t_dir and entry["direction"] and t_dir != entry["direction"]:
                        dir_penalty = 1.0

                    norm_diff = diff / (abs(t_amount) + 1)
                    # 将 dir_penalty 加入评分元组，优先级仅次于金额差异
                    score = (norm_diff * 10, dir_penalty, abs(delta), 1.0 - text_score)
                    if best is None or score < best["score"]:
                        best = {"entry": entry, "score": score, "delta": delta, "diff": diff, "text_score": text_score}

            if best:
                entry = best["entry"]
                new_summary = entry["summary"]
                is_weak_match = (best["diff"] > abs(t_amount) * 0.2) and (best["text_score"] < 0.2)
                if is_weak_match:
                    if p_signed_amt != "":
                        preview_rows.append(["未匹配", t_date, p_signed_amt, original_summary, ""])
                    else:
                        preview_rows.append(["未匹配", t_date, p_debit, p_credit, original_summary, ""])
                    continue
                
                result_df.at[idx, target_summary_col] = new_summary
                
                # --- 同步方向和金额 ---
                s_dir = entry["direction"]
                s_amt = entry["amount"]
                
                if self.summary_match_sync_direction_var.get():
                    t_amt_col = target_mapping.get("amount")
                    t_debit_col = target_mapping.get("debit")
                    t_credit_col = target_mapping.get("credit")
                    
                    if t_amt_col:
                        signed_val = s_amt if s_dir == "debit" else -s_amt
                        result_df.at[idx, t_amt_col] = signed_val
                        p_signed_amt = signed_val
                    if t_debit_col:
                        val = s_amt if s_dir == "debit" else 0
                        result_df.at[idx, t_debit_col] = val
                        p_debit = val if val else ""
                    if t_credit_col:
                        val = s_amt if s_dir == "credit" else 0
                        result_df.at[idx, t_credit_col] = val
                        p_credit = val if val else ""
                
                matched_count += 1
                if self.summary_match_unique_var.get(): used_sources.add(entry["idx"])
                log_msg = f"  [成功] 行{idx}: {t_date} {t_amount:.2f} <==> 源:[金额:{entry['amount']:.2f}, 方向:{entry['direction']}] (相似度:{best['text_score']:.2f})ảng بمن"
                self.log_message(log_msg)
                
                if p_signed_amt != "":
                    preview_rows.append(["已匹配", t_date, p_signed_amt, original_summary, new_summary])
                else:
                    preview_rows.append(["已匹配", t_date, p_debit, p_credit, original_summary, new_summary])
            else:
                if p_signed_amt != "":
                    preview_rows.append(["未匹配", t_date, p_signed_amt, original_summary, ""])
                else:
                    preview_rows.append(["未匹配", t_date, p_debit, p_credit, original_summary, ""])

        self.log_message("----------------ảng بمن")
        self.summary_match_result_df = result_df
        self.summary_match_status_var.set(f"匹配完成：共 {len(result_df)} 行，匹配 {matched_count} 行")
        self.log_message(f"摘要匹配完成，共处理 {len(result_df)} 行，成功关联 {matched_count} 行ảng بمن。")
        self._refresh_summary_match_preview(preview_rows[:30], target_mapping)
        self.summary_match_export_btn.config(state="normal")

    def _refresh_summary_match_preview(self, rows, mapping):
        for w in self.summary_match_preview_container.winfo_children():
            w.destroy()

        # 根据用户映射动态决定预览列
        if "debit" in mapping or "credit" in mapping:
            columns = ["匹配状态", "日期", "借方", "贷方", "原摘要", "新摘要"]
        else:
            columns = ["匹配状态", "日期", "金额(带号)", "原摘要", "新摘要"]

        tree_frame = ttk.Frame(self.summary_match_preview_container)
        tree_frame.pack(fill="both", expand=True)

        tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")

        tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
        )
        tree_scroll_y.config(command=tree.yview)
        tree_scroll_x.config(command=tree.xview)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=160, minwidth=80)

        for row in rows:
            tree.insert("", "end", values=row)

        tree.pack(side="left", fill="both", expand=True)
        tree_scroll_y.pack(side="right", fill="y")
        tree_scroll_x.pack(side="bottom", fill="x")
        attach_treeview_tools(tree)

    def _on_summary_match_export_format_changed(self, event=None):
        name = self.summary_match_export_format_var.get().strip()
        set_active_export_format("summary_match_export", name)

    def _open_summary_match_export_format_editor(self):
        headers = list(self.summary_match_result_df.columns) if self.summary_match_result_df is not None else []
        open_export_format_editor(
            self.root,
            "summary_match_export",
            headers,
            title="导出格式设置 - 摘要匹配",
            base_data_mgr=self.base_data_mgr
        )
        self._refresh_summary_match_export_format_options()

    def _refresh_summary_match_export_format_options(self):
        if hasattr(self, "summary_match_export_format_combo"):
            names = get_export_format_names("summary_match_export")
            self.summary_match_export_format_combo["values"] = names
            active = get_active_export_format_name("summary_match_export")
            if active:
                self.summary_match_export_format_var.set(active)

    def _restore_summary_match_codes(self):
        if self.summary_match_result_df is None:
            messagebox.showwarning("提示", "暂无可处理的数据，请先执行匹配。")
            return
        counts = self._restore_codes_in_df(self.summary_match_result_df, title="摘要匹配-智能还原编码")
        messagebox.showinfo(
            "结果",
            f"智能还原完成！\n\n- 修复科目编码: {counts['account']} 条\n- 修复往来单位编码: {counts['partner']} 条"
        )

    def _export_summary_match_result(self):
        if self.summary_match_result_df is None:
            messagebox.showwarning("提示", "暂无可导出的匹配结果，请先执行匹配。")
            return
        path = filedialog.asksaveasfilename(
            title="保存摘要替换结果",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not path:
            return
        try:
            if hasattr(self, "summary_match_tree") and hasattr(self.summary_match_tree, "_treeview_tools"):
                headers, rows = self.summary_match_tree._treeview_tools.get_visual_data()
            else:
                headers = list(self.summary_match_result_df.columns)
                rows = self.summary_match_result_df.values.tolist()
                
            mapped_headers, mapped_rows, mapped = apply_export_format(
                "summary_match_export",
                headers,
                rows,
                base_data_mgr=self.base_data_mgr
            )
            if mapped:
                df = pd.DataFrame(mapped_rows, columns=mapped_headers)
            else:
                df = self.summary_match_result_df
            df.to_excel(path, index=False)
            messagebox.showinfo("完成", f"已导出替换结果：\n{path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{e}")

    def _build_base_data_tab(self):
        """构建基础数据管理标签页"""
        base_frame = ttk.Frame(self.notebook)
        self.notebook.add(base_frame, text="基础数据管理")

        # 左侧：数据类型列表
        left_frame = ttk.Frame(base_frame, width=200)
        left_frame.pack(side="left", fill="y", padx=5, pady=5)

        ttk.Label(left_frame, text="数据类型", font=("", 10, "bold")).pack(pady=5)

        self.data_types_base = {
            "币种": "currency",
            "部门": "department",
            "仓库": "warehouse",
            "科目编码": "account_subject",
            "品目信息": "product",
            "往来单位": "business_partner",
            "账户": "bank_account",
            "智能识别缓存": "smart_recognition_cache",
            "智能识别规则": "recognition_rules"
        }

        self.current_table = tk.StringVar(value="currency")

        self.base_data_type_frame = ttk.Frame(left_frame)
        self.base_data_type_frame.pack(fill="both", expand=True)
        self._refresh_base_data_type_buttons()

        ttk.Separator(left_frame, orient="horizontal").pack(fill="x", padx=5, pady=6)
        ttk.Button(left_frame, text="新增分类", command=self._add_custom_category).pack(fill="x", padx=5, pady=2)
        ttk.Button(left_frame, text="删除分类", command=self._delete_custom_category).pack(fill="x", padx=5, pady=2)

        # 右侧：数据显示和编辑区域
        right_frame = ttk.Frame(base_frame)
        right_frame.pack(side="right", fill="both", expand=True, padx=5, pady=5)

        # 顶部：搜索和操作按钮
        top_bar = ttk.Frame(right_frame)
        top_bar.pack(fill="x", pady=(0, 5))

        ttk.Label(top_bar, text="搜索:").pack(side="left", padx=5)
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(top_bar, textvariable=self.search_var, width=30)
        self.search_entry.pack(side="left", padx=5)
        ttk.Label(top_bar, text="字段:").pack(side="left")
        self.search_field_combo = ttk.Combobox(top_bar, textvariable=self.search_field_var, width=12, state="readonly")
        self.search_field_combo.pack(side="left", padx=(0, 5))
        ttk.Button(top_bar, text="搜索", command=self._search_base_data).pack(side="left", padx=2)
        ttk.Button(top_bar, text="显示全部", command=self._load_base_data_table).pack(side="left", padx=2)
        self.base_data_show_favorites_only_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            top_bar,
            text="仅显示收藏",
            variable=self.base_data_show_favorites_only_var,
            command=self._refresh_base_data_view
        ).pack(side="left", padx=(4, 2))
        ttk.Button(top_bar, text="导出模板", command=self.export_base_data_templates).pack(side="left", padx=6)
        ttk.Button(top_bar, text="批量导入", command=self.import_base_data_batch).pack(side="left", padx=2)

        # 编辑按钮
        ttk.Button(top_bar, text="新增", command=self._add_base_data_record).pack(side="right", padx=2)
        ttk.Button(top_bar, text="编辑", command=self._edit_base_data_record).pack(side="right", padx=2)
        ttk.Button(top_bar, text="删除", command=self._delete_base_data_record).pack(side="right", padx=2)
        ttk.Button(top_bar, text="加入收藏", command=lambda: self._update_base_data_favorites(True)).pack(side="right", padx=2)
        ttk.Button(top_bar, text="取消收藏", command=lambda: self._update_base_data_favorites(False)).pack(side="right", padx=2)

        # 中部：数据表格
        table_frame = ttk.Frame(right_frame)
        table_frame.pack(fill="both", expand=True)

        # 滚动条
        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical")
        scrollbar_y.pack(side="right", fill="y")

        scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal")
        scrollbar_x.pack(side="bottom", fill="x")

        # 树形视图
        self.base_data_tree = ttk.Treeview(
            table_frame,
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set,
            selectmode="extended"  # 启用多选
        )
        self.base_data_tree.pack(fill="both", expand=True)
        self.base_data_tree_tools = attach_treeview_tools(self.base_data_tree)
        self.base_data_tree.bind("<<TreeviewPaste>>", self._on_base_data_paste)

        scrollbar_y.config(command=self.base_data_tree.yview)
        scrollbar_x.config(command=self.base_data_tree.xview)

        # 双击编辑 (支持单元格直接编辑)
        self.base_data_tree.bind("<Double-1>", self._on_tree_double_click)

    def _on_base_data_paste(self, event=None):
        """处理基础数据表格的批量粘贴"""
        table_name = self.current_table.get()
        if not table_name:
            return
            
        # 统计更新行
        # 注意：TreeviewTools 已经把值更新到 Treeview 界面上了
        # 我们现在需要把界面上的值同步到数据库
        
        if not messagebox.askyesno("保存确认", f"检测到批量粘贴操作，是否将更改保存到数据库表 [{table_name}]？\n\n注意：这将根据第一列的 ID 更新所有对应行。"):
            # 如果不保存，建议刷新以还原
            self._load_base_data_table()
            return
            
        success_count = 0
        error_count = 0
        
        columns = self.base_data_tree["columns"]
        items = self.base_data_tree.get_children("")
        
        for iid in items:
            values = self.base_data_tree.item(iid, "values")
            if not values:
                continue
                
            try:
                record_id = int(values[0])
                # 构造更新数据，跳过 ID 列
                data = {}
                for i in range(1, len(columns)):
                    if i < len(values):
                        data[columns[i]] = values[i]
                
                result = self.base_data_mgr.update_record(table_name, record_id, data)
                if result and result.get("success"):
                    success_count += 1
                else:
                    error_count += 1
            except (ValueError, IndexError, TypeError):
                # 可能这一行没有 ID，或者是空行
                error_count += 1
                
        self.log_message(f"批量粘贴更新完成：成功 {success_count} 行，失败 {error_count} 行。")
        self._set_status_text(f"保存完成：成功 {success_count} 条")
        
        # 刷新视图确保一致性
        self._load_base_data_table()

        # 底部：状态栏
        self.status_label = ttk.Label(right_frame, text="就绪", relief="sunken")
        self.status_label.pack(fill="x", pady=(5, 0))

        # 初始化加载第一个表
        self._load_base_data_table()

    def _refresh_base_data_type_buttons(self):
        for w in self.base_data_type_frame.winfo_children():
            w.destroy()

        self.data_types = dict(self.data_types_base)
        custom_categories = []
        if self.base_data_mgr:
            try:
                custom_categories = self.base_data_mgr.list_custom_categories()
            except Exception:
                custom_categories = []
        if custom_categories:
            self.data_types["自定义类"] = None
            for cat in custom_categories:
                label = cat.get("display_name") or cat.get("name")
                name_key = cat.get("name")
                if label and name_key:
                    self.data_types[label] = f"custom:{name_key}"

        values = [v for v in self.data_types.values() if v]
        if self.current_table.get() not in values and values:
            self.current_table.set(values[0])

        for label, table_name in self.data_types.items():
            if table_name is None:
                ttk.Label(self.base_data_type_frame, text=label, foreground="gray").pack(
                    fill="x", padx=5, pady=(6, 2)
                )
                continue
            btn = ttk.Radiobutton(
                self.base_data_type_frame,
                text=label,
                variable=self.current_table,
                value=table_name,
                command=self._load_base_data_table
            )
            btn.pack(fill="x", padx=5, pady=2)

    def _add_custom_category(self):
        if not self.base_data_mgr:
            messagebox.showwarning("提示", "基础数据管理器未初始化。")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("新增自定义分类")
        dialog.geometry("520x420")
        dialog.transient(self.root)
        dialog.grab_set()

        name_var = tk.StringVar()
        display_var = tk.StringVar()

        ttk.Label(dialog, text="分类名称(显示):").grid(row=0, column=0, padx=10, pady=8, sticky="e")
        ttk.Entry(dialog, textvariable=display_var, width=30).grid(row=0, column=1, padx=6, pady=8, sticky="w")
        ttk.Label(dialog, text="分类标识(可选):").grid(row=1, column=0, padx=10, pady=8, sticky="e")
        ttk.Entry(dialog, textvariable=name_var, width=30).grid(row=1, column=1, padx=6, pady=8, sticky="w")

        fields_frame = ttk.LabelFrame(dialog, text="字段(除 code/name 之外)")
        fields_frame.grid(row=2, column=0, columnspan=2, padx=10, pady=8, sticky="nsew")
        dialog.grid_rowconfigure(2, weight=1)
        dialog.grid_columnconfigure(1, weight=1)

        field_tree = ttk.Treeview(fields_frame, columns=("name", "type"), show="headings", height=8)
        field_tree.heading("name", text="字段名")
        field_tree.heading("type", text="类型")
        field_tree.column("name", width=200)
        field_tree.column("type", width=100)
        field_tree.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        fsb = ttk.Scrollbar(fields_frame, orient="vertical", command=field_tree.yview)
        fsb.pack(side="right", fill="y")
        field_tree.configure(yscrollcommand=fsb.set)

        def add_field():
            sub = tk.Toplevel(dialog)
            sub.title("新增字段")
            sub.geometry("300x160")
            sub.transient(dialog)
            sub.grab_set()
            fname = tk.StringVar()
            ftype = tk.StringVar(value="text")
            ttk.Label(sub, text="字段名:").pack(pady=8)
            ttk.Entry(sub, textvariable=fname, width=24).pack()
            ttk.Label(sub, text="类型:").pack(pady=6)
            ttk.Combobox(sub, textvariable=ftype, values=["text", "number", "date", "bool"], state="readonly").pack()

            def confirm():
                n = fname.get().strip()
                if not n or n in ("id", "code", "name"):
                    messagebox.showwarning("提示", "字段名不能为空，且不能为 id/code/name。")
                    return
                for item in field_tree.get_children():
                    if field_tree.item(item, "values")[0] == n:
                        messagebox.showwarning("提示", "字段名已存在。")
                        return
                field_tree.insert("", "end", values=(n, ftype.get()))
                sub.destroy()

            ttk.Button(sub, text="确定", command=confirm).pack(pady=10)

        def remove_field():
            for item in field_tree.selection():
                field_tree.delete(item)

        btns = ttk.Frame(dialog)
        btns.grid(row=3, column=0, columnspan=2, pady=8)
        ttk.Button(btns, text="新增字段", command=add_field).pack(side="left", padx=6)
        ttk.Button(btns, text="删除字段", command=remove_field).pack(side="left", padx=6)

        def save_category():
            display_name = display_var.get().strip()
            name_key = name_var.get().strip() or display_name
            if not display_name:
                messagebox.showwarning("提示", "分类名称不能为空。")
                return
            fields = []
            for item in field_tree.get_children():
                fname, ftype = field_tree.item(item, "values")
                fields.append({"name": fname, "type": ftype})
            res = self.base_data_mgr.add_custom_category(name_key, display_name, fields)
            if not res.get("success"):
                messagebox.showerror("错误", f"创建失败：{res.get('message')}")
                return
            dialog.destroy()
            self._refresh_base_data_type_buttons()

        action = ttk.Frame(dialog)
        action.grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(action, text="保存", command=save_category).pack(side="right", padx=6)
        ttk.Button(action, text="取消", command=dialog.destroy).pack(side="right")

    def _delete_custom_category(self):
        if not self.base_data_mgr:
            messagebox.showwarning("提示", "基础数据管理器未初始化。")
            return
        table_name = self.current_table.get()
        if not table_name.startswith("custom:"):
            messagebox.showwarning("提示", "请先选择一个自定义分类。")
            return
        name_key = table_name.split("custom:", 1)[1]
        if not messagebox.askyesno("确认", f"删除自定义分类：{name_key}？\n此操作会删除该分类下的所有记录。"):
            return
        res = self.base_data_mgr.delete_custom_category(name_key)
        if not res.get("success"):
            messagebox.showerror("错误", f"删除失败：{res.get('message')}")
            return
        self._refresh_base_data_type_buttons()
        self._load_base_data_table()

    def _on_tree_double_click(self, event):
        """处理双击事件：尝试单元格编辑，否则弹出对话框"""
        region = self.base_data_tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = self.base_data_tree.identify_column(event.x)
        item_id = self.base_data_tree.identify_row(event.y)
        
        if not item_id:
            return

        table_name = self.current_table.get()
        
        # 获取列索引 (例如 #1, #2...)
        col_index = int(column.replace("#", "")) - 1
        columns = self.base_data_tree["columns"]
        
        # 防止索引越界（虽然理论上不应该）
        if col_index >= len(columns):
            return
            
        col_name = columns[col_index]

        # 1. 禁止编辑 ID 列 (通常是第0列)
        if col_index == 0 or col_name == "id":
            return

        # 2. 智能识别缓存表：禁止编辑摘要和创建时间，只允许编辑科目编码
        if table_name == "smart_recognition_cache":
            if col_name == "account_code":
                self._show_entry_editor(item_id, col_index, col_name)
            else:
                # 其他列(摘要)太长或只读，弹窗显示更合适，或者直接忽略
                self._edit_base_data_record()
            return

        # 3. 其他所有基础数据表：允许直接编辑所有非ID列
        self._show_entry_editor(item_id, col_index, col_name)

    def _show_entry_editor(self, item_id, col_index, col_name):
        """在Treeview单元格上显示输入框"""
        # 获取单元格坐标
        bbox = self.base_data_tree.bbox(item_id, column=col_name)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # 获取当前值
        current_values = self.base_data_tree.item(item_id, "values")
        current_val = current_values[col_index]

        # 创建输入框
        entry = ttk.Entry(self.base_data_tree, width=width)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, current_val)
        entry.select_range(0, tk.END)
        entry.focus()

        def save_edit(event=None):
            new_val = entry.get().strip()
            # 销毁输入框
            entry.destroy()
            
            # 如果值没变，不进行数据库操作
            if new_val == str(current_val):
                return

            table_name = self.current_table.get()
            record_id = int(current_values[0]) # 假设第一列总是ID
            
            result = None
            
            # 分发更新逻辑
            if table_name == "smart_recognition_cache":
                if col_name == "account_code":
                    result = self.base_data_mgr.update_cached_recognition(record_id, new_val)
            else:
                # 通用表更新
                # 注意：对于数字类型的列，可能需要转换类型，这里暂且都当字符串存，sqlite是弱类型的，通常没问题
                # 但为了严谨，如果base_data_manager有类型转换逻辑更好。
                # 这里为了通用性，直接传递。
                data = {col_name: new_val}
                result = self.base_data_mgr.update_record(table_name, record_id, data)
            
            if result and result["success"]:
                # 更新界面显示
                new_values = list(current_values)
                new_values[col_index] = new_val
                self.base_data_tree.item(item_id, values=new_values)
                self._set_status_text(f"更新成功: ID {record_id} [{col_name}] -> {new_val}")
            else:
                msg = result["message"] if result else "未知错误"
                messagebox.showerror("错误", msg)

        def cancel_edit(event=None):
            entry.destroy()

        # 绑定事件
        entry.bind("<Return>", save_edit)
        # entry.bind("<FocusOut>", lambda e: entry.destroy()) # 失去焦点暂不保存
        # 为了体验更像Excel，失去焦点通常意味着保存。但为了防止误触，我们这里设定：
        # 如果点击了表格其他地方，会触发Treeview的点击，导致FocusOut。
        # 暂且设为：失去焦点 = 取消 (您之前似乎倾向于安全)
        # 如果您希望失去焦点自动保存，请告诉我。
        entry.bind("<FocusOut>", lambda e: entry.destroy()) 
        entry.bind("<Escape>", cancel_edit)

    def _load_template(self, template_path=None):
        path = template_path if template_path else self.template_path_var.get()
        if not path:
             path = TEMPLATE_FILE
             
        try:
            headers, wb, ws = load_template_headers(path)
        except FileNotFoundError as e:
            messagebox.showerror("错误", f"读取模板失败：\n{e}\n\n请确认 {path} 是否存在。")
            return
        except Exception as e:
            messagebox.showerror("错误", f"读取模板失败：\n{e}")
            return

        self.template_headers = headers
        self.template_workbook = wb
        self.template_sheet = ws
        self._debug_log(f"已加载模板: {path}, 表头数={len(headers)}")
    
    def _load_original_headers(self):
        """原格式模式：使用源文件表头作为目标表头"""
        if self.input_df is None:
            self.template_headers = []
            self.template_workbook = None
            self.template_sheet = None
            return
        headers = []
        for idx, col in enumerate(self.input_df.columns, start=1):
            headers.append(TemplateHeader(str(col), idx, ""))
        self.template_headers = headers
        self.template_workbook = None
        self.template_sheet = None
        self._debug_log(f"已加载原格式表头: {len(headers)} 列")

    def select_template_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            preserve = self.preserve_mapping_var.get()
            previous = {h: var.get() for h, var in self.mapping_vars.items()} if preserve else {}
            self.template_path_var.set(filename)
            self.log_message(f"已选择模板文件: {filename}")
            self._load_template()
            if preserve:
                self._skip_auto_mapping_once = True
            self._create_mapping_widgets()
            if preserve and previous:
                options = set(self._get_mapping_source_options())
                for header, val in previous.items():
                    if header in self.mapping_vars and val in options:
                        self.mapping_vars[header].set(val)

    def _on_mode_changed(self, event=None):
        """切换转换模式"""
        raw_mode = self.convert_mode_var.get()
        
        # 处理分隔符选择
        if raw_mode.startswith("---"):
            self.mode_combo.set(MODE_GENERAL_VOUCHER)
            return

        base_title = "一般凭证 Excel 格式转换工具（含自动识别列匹配）"
        
        # 判断是否是自定义方案
        scheme_data = None
        mode = raw_mode
        if raw_mode.startswith("方案: "):
            scheme_name = raw_mode.replace("方案: ", "")
            scheme_data = self.cached_schemes.get(scheme_name)
            if scheme_data:
                mode = scheme_data["base_mode"]
                try:
                    comp_json = scheme_data.get("composite_json")
                    self.custom_composite_fields = json.loads(comp_json) if comp_json else {}
                except Exception:
                    self.custom_composite_fields = {}
                # 方案附带的模板路径
                if scheme_data["template_path"] and os.path.exists(scheme_data["template_path"]):
                    self.template_path_var.set(scheme_data["template_path"])
                base_title += f" - {scheme_name}"
            else:
                # 方案丢失或错误，回退
                mode = MODE_GENERAL_VOUCHER

        if mode == MODE_CUSTOM:
            self.template_entry.config(state="normal")
            self.template_btn.config(state="normal")
            self.preserve_mapping_chk.config(state="normal")
            if not scheme_data: # 纯自定义模式，非方案
                base_title += " - 自定义模式"
        elif mode == MODE_ORIGINAL:
            self.template_path_var.set(MODE_ORIGINAL)
            self.template_entry.config(state="readonly")
            self.template_btn.config(state="disabled")
            self.preserve_mapping_chk.config(state="disabled")
            base_title += " - 原格式模式"
        else:
            if not scheme_data: # 如果不是方案，重置为对应的默认模板
                if mode == MODE_GENERAL_VOUCHER and os.path.exists("Template_通用凭证.xlsx"):
                     self.template_path_var.set("Template_通用凭证.xlsx")
                else:
                     self.template_path_var.set(TEMPLATE_FILE)

            self.template_entry.config(state="readonly")
            self.template_btn.config(state="disabled")
            self.preserve_mapping_chk.config(state="disabled")
            if mode == MODE_SALES_OUTBOUND:
                base_title += " - 销售出库模式"

        self.root.title(base_title)
        self._debug_log(f"切换模式: raw={raw_mode}, effective={mode}, template={self.template_path_var.get()}")

        # 重新加载模板
        # 优化：如果是方案模式，且模板路径没变，可能不需要重载，但为了保险起见还是重载
        if event is not None or scheme_data:
            if mode == MODE_ORIGINAL:
                self._load_original_headers()
            else:
                self._load_template()
            
        # 切换模式后重建映射区域
        if self.input_df is not None:
            if mode == MODE_ORIGINAL:
                self._load_original_headers()
            self._create_mapping_widgets()
            
            # 如果是方案模式，应用保存的映射
            if scheme_data and scheme_data.get("mapping_json"):
                try:
                    mapping = json.loads(scheme_data["mapping_json"])
                    valid_sources = set(self._get_mapping_source_options())
                    # 方案未包含字段格式时，回退到全局已保存格式
                    self.field_formats = dict(getattr(self, "_global_field_formats", {}))
                    for header, col in mapping.items():
                        if header == "__aux_debit_col__":
                            if col in self.input_columns: self.manual_debit_col_var.set(col)
                        elif header == "__aux_credit_col__":
                            if col in self.input_columns: self.manual_credit_col_var.set(col)
                        elif header == "__aux_dc_col__":
                            if col in self.input_columns: self.manual_dc_col_var.set(col)
                        elif header == "__field_formats__":
                            if isinstance(col, dict):
                                self.field_formats = col
                        elif header in self.mapping_vars and col in valid_sources:
                            self.mapping_vars[header].set(col)
                except Exception as e:
                    print(f"应用方案映射失败: {e}")

    @staticmethod
    def _is_legacy_xls(path: str) -> bool:
        return str(path or "").lower().endswith(".xls")

    def _get_excel_sheet_names(self, path: str) -> List[str]:
        """获取 Excel 工作表名，兼容 .xlsx 与 .xls。"""
        if self._is_legacy_xls(path):
            try:
                xls = pd.ExcelFile(path, engine="xlrd")
                return list(xls.sheet_names)
            except ImportError:
                raise RuntimeError("当前环境缺少 xlrd，无法读取 .xls 文件。请先安装：pip install xlrd")
            except Exception as e:
                raise RuntimeError(f".xls 工作表读取失败: {e}")

        wb = load_workbook(path, read_only=True, data_only=True)
        return list(wb.sheetnames)

    def _read_excel_sheet(self, path: str, sheet_name: str):
        """读取指定工作表，兼容 .xlsx 与 .xls。"""
        if self._is_legacy_xls(path):
            try:
                return pd.read_excel(path, sheet_name=sheet_name, engine="xlrd")
            except ImportError:
                raise RuntimeError("当前环境缺少 xlrd，无法读取 .xls 文件。请先安装：pip install xlrd")
        return pd.read_excel(path, sheet_name=sheet_name)

    def select_input_file(self):
        path = filedialog.askopenfilename(
            title="选择原始 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx;*.xls"), ("所有文件", "*.*")],
        )
        if not path:
            return

        self.input_path = path
        self.input_path_var.set(path)
        self.log_message(f"已选择源文件: {path}")

        try:
            sheet_names = self._get_excel_sheet_names(path)
        except Exception as e:
            messagebox.showerror("错误", f"读取工作表列表失败：\n{e}")
            return

        self.sheet_combo["values"] = sheet_names
        self.sheet_combo.configure(state="readonly")
        if sheet_names:
            self.sheet_var.set(sheet_names[0])
            self._load_input_sheet()

    def _open_in_converter(self, path: str, mode: Optional[str] = MODE_ORIGINAL):
        if not path or not os.path.exists(path):
            messagebox.showwarning("提示", "导出文件不存在，无法转入凭证转换。")
            return
        if mode and self.convert_mode_var.get() != mode:
            self.convert_mode_var.set(mode)
            self._on_mode_changed()
        if hasattr(self, "excel_tab_frame"):
            self.notebook.select(self.excel_tab_frame)
        self.input_path = path
        self.input_path_var.set(path)
        self.log_message(f"已转入凭证转换: {path}")
        try:
            sheet_names = self._get_excel_sheet_names(path)
        except Exception as e:
            messagebox.showerror("错误", f"读取工作表列表失败：\n{e}")
            return
        self.sheet_combo["values"] = sheet_names
        self.sheet_combo.configure(state="readonly")
        if sheet_names:
            self.sheet_var.set(sheet_names[0])
            self._load_input_sheet()

    def _on_sheet_changed(self, event=None):
        self._load_input_sheet()

    def _load_input_sheet(self):
        if not self.input_path or not self.sheet_var.get():
            return
        try:
            sheet_name = self.sheet_var.get()
            self.log_message(f"正在加载工作表: [{sheet_name}] ...")
            self.input_df = self._read_excel_sheet(self.input_path, sheet_name)
            
            rows, cols = self.input_df.shape
            self.log_message(f"工作表加载成功: {rows} 行, {cols} 列")
            self.log_message(f"包含列: {', '.join(str(c) for c in self.input_df.columns)}")
        except Exception as e:
            messagebox.showerror("错误", f"读取 Excel 数据失败：\n{e}")
            self.log_message(f"加载工作表失败: {e}")
            return

        self.input_columns = [str(c) for c in self.input_df.columns]
        self._debug_log(f"已加载工作表: {sheet_name}, 行={rows}, 列={cols}")
        self._debug_log(f"列示例: {self.input_columns[:12]}")
        effective_mode, _ = self._get_effective_mode_for_mapping()
        if effective_mode == MODE_ORIGINAL:
            self._load_original_headers()
        self._create_mapping_widgets()
        self._refresh_debit_credit_options()

    def _get_header_display_list(self):
        """根据表头方案返回映射界面的显示顺序"""
        headers = self.template_headers or []
        scheme_name = self.header_scheme_var.get().strip()
        if not scheme_name:
            return headers
        scheme = HEADER_SCHEMES.get(scheme_name)
        if not scheme:
            return headers
        name_map = {h.name: h for h in headers}
        ordered = []
        used = set()
        for item in scheme:
            name = item.get("name")
            if name in name_map:
                ordered.append(name_map[name])
                used.add(name)
        for h in headers:
            if h.name not in used:
                ordered.append(h)
        return ordered

    def _apply_header_scheme_to_mapping(self):
        if not self.template_headers:
            return
        self._debug_log(f"应用表头方案到映射: {self.header_scheme_var.get()}")
        self._create_mapping_widgets()

    def _on_export_format_changed(self, event=None):
        name = self.export_format_var.get().strip()
        set_active_export_format("main_export", name)

    def _open_export_format_editor(self):
        source_headers = [h.name for h in self.template_headers] if self.template_headers else list(self.input_columns)
        open_export_format_editor(
            self.root,
            "main_export",
            source_headers,
            title="导出格式设置 - 凭证转换",
            base_data_mgr=self.base_data_mgr
        )
        self._refresh_export_format_options()

    def _ensure_default_main_export_format(self):
        try:
            names = get_export_format_names("main_export")
            if names:
                return
            if not self.template_headers:
                return
            mapping = [{"output": h.name, "source": h.name, "default": ""} for h in self.template_headers]
            formats = load_export_formats()
            module = formats.get("main_export")
            if not isinstance(module, dict):
                module = {"active": "", "formats": {}}
            module_formats = module.get("formats")
            if not isinstance(module_formats, dict):
                module_formats = {}
            default_name = "系统默认-模板同名"
            module_formats[default_name] = mapping
            module["formats"] = module_formats
            module["active"] = default_name
            formats["main_export"] = module
            save_export_formats(formats)
            if hasattr(self, "export_format_var"):
                self.export_format_var.set(default_name)
            self._refresh_export_format_options()
        except Exception as e:
            self._debug_log(f"初始化默认导出格式失败: {e}")

    def _refresh_export_format_options(self):
        if not hasattr(self, "export_format_combo"):
            return
        names = get_export_format_names("main_export")
        self.export_format_combo["values"] = names
        active = get_active_export_format_name("main_export")
        if active:
            self.export_format_var.set(active)

    def _create_mapping_widgets(self):
        for w in self.mapping_frame.winfo_children():
            w.destroy()

        ttk.Label(self.mapping_frame, text="模板列名", width=20).grid(
            row=0, column=0, padx=3, pady=3, sticky="w"
        )
        ttk.Label(self.mapping_frame, text="来源列（原始 Excel）", width=30).grid(
            row=0, column=1, padx=3, pady=3, sticky="w"
        )
        ttk.Label(self.mapping_frame, text="批注", width=6).grid(
            row=0, column=3, padx=3, pady=3, sticky="w"
        )
        ttk.Label(self.mapping_frame, text="格式", width=6).grid(
            row=0, column=4, padx=3, pady=3, sticky="w"
        )
        ttk.Button(self.mapping_frame, text="添加字段", command=self._add_mapping_source_column, width=10).grid(
            row=0, column=5, padx=3, pady=3, sticky="w"
        )
        ttk.Button(self.mapping_frame, text="删除字段", command=self._remove_mapping_source_column, width=10).grid(
            row=0, column=6, padx=3, pady=3, sticky="w"
        )

        self.mapping_vars = {}

        # 解析当前有效模式
        raw_mode = self.convert_mode_var.get()
        effective_mode = raw_mode
        if raw_mode.startswith("方案: "):
            scheme_name = raw_mode.replace("方案: ", "")
            if hasattr(self, "cached_schemes") and scheme_name in self.cached_schemes:
                effective_mode = self.cached_schemes[scheme_name]["base_mode"]
            else:
                effective_mode = MODE_GENERAL_VOUCHER

        row_idx = 1
        display_headers = self._get_header_display_list()
        self._debug_log(f"构建映射界面: 显示表头数={len(display_headers)}")
        for header in display_headers:
            # 通用凭证模式下，仅显示核心字段
            # if effective_mode == MODE_GENERAL_VOUCHER:
            #    if header.name not in GENERAL_VOUCHER_FIELDS:
            #        continue

            ttk.Label(self.mapping_frame, text=header.name).grid(
                row=row_idx, column=0, padx=3, pady=3, sticky="w"
            )

            var = tk.StringVar(value=EMPTY_OPTION)
            combo_values = self._get_mapping_source_options()
            combo = ttk.Combobox(
                self.mapping_frame,
                textvariable=var,
                values=combo_values,
                state="normal",
                width=30,
            )
            combo.grid(row=row_idx, column=1, padx=3, pady=3, sticky="w")
            
            # 绑定选择事件，支持自定义输入
            combo.bind("<<ComboboxSelected>>", lambda e, h=header.name, v=var: self._on_mapping_selected(h, v))
            combo.bind("<FocusOut>", lambda e, h=header.name, v=var: self._on_mapping_value_committed(h, v))
            combo.bind("<Return>", lambda e, h=header.name, v=var: self._on_mapping_value_committed(h, v))

            self.mapping_vars[header.name] = var
            if effective_mode == MODE_ORIGINAL and header.name in self.input_columns:
                var.set(header.name)

            btn_view = ttk.Button(
                self.mapping_frame,
                text="查看",
                command=lambda h=header: self.show_comment(h),
                width=6,
            )
            btn_view.grid(row=row_idx, column=3, padx=3, pady=3, sticky="w")

            btn_fmt = ttk.Button(
                self.mapping_frame,
                text="格式",
                command=lambda h=header.name: self._open_field_format_dialog(h),
                width=6,
            )
            btn_fmt.grid(row=row_idx, column=4, padx=3, pady=3, sticky="w")

            row_idx += 1

        self._apply_auto_mapping()

        # 更新借/贷辅助下拉的可选值
        self._refresh_debit_credit_options()

    def _open_field_format_dialog(self, header_name):
        """打开字段格式自定义对话框"""
        top = tk.Toplevel(self.root)
        top.title(f"字段格式自定义: {header_name}")
        top.geometry("400x300")
        top.transient(self.root)
        top.grab_set()

        current_fmt = self.field_formats.get(header_name, {})
        
        main_frame = ttk.Frame(top, padding=20)
        main_frame.pack(fill="both", expand=True)

        ttk.Label(main_frame, text=f"正在设置字段 【{header_name}】 的输出格式", font=("", 10, "bold")).pack(pady=(0, 15))

        # 数字格式设置
        num_frame = ttk.LabelFrame(main_frame, text="数字/金额设置", padding=10)
        num_frame.pack(fill="x", pady=5)

        ttk.Label(num_frame, text="保留小数位数:").grid(row=0, column=0, sticky="e", padx=5)
        decimal_var = tk.StringVar(value=str(current_fmt.get("decimal", "")))
        decimal_entry = ttk.Entry(num_frame, textvariable=decimal_var, width=10)
        decimal_entry.grid(row=0, column=1, sticky="w")
        ttk.Label(num_frame, text="(空则使用默认规则)", foreground="gray").grid(row=0, column=2, padx=5)

        # 文本长度设置
        text_frame = ttk.LabelFrame(main_frame, text="文本/常规设置", padding=10)
        text_frame.pack(fill="x", pady=5)

        ttk.Label(text_frame, text="最大字符长度:").grid(row=0, column=0, sticky="e", padx=5)
        length_var = tk.StringVar(value=str(current_fmt.get("max_len", "")))
        length_entry = ttk.Entry(text_frame, textvariable=length_var, width=10)
        length_entry.grid(row=0, column=1, sticky="w")

        def save_and_close():
            fmt = {}
            try:
                d = decimal_var.get().strip()
                if d: fmt["decimal"] = int(d)
                
                l = length_var.get().strip()
                if l: fmt["max_len"] = int(l)
                
                if fmt:
                    self.field_formats[header_name] = fmt
                    self.log_message(f"已更新字段 [{header_name}] 的格式自定义: {fmt}")
                elif header_name in self.field_formats:
                    del self.field_formats[header_name]
                    self.log_message(f"已清除字段 [{header_name}] 的格式自定义")
                self._persist_field_formats()
                
                top.destroy()
            except ValueError:
                messagebox.showerror("错误", "请输入有效的数字")

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=20)
        ttk.Button(btn_frame, text="确定", command=save_and_close).pack(side="right", padx=5)
        ttk.Button(btn_frame, text="取消", command=top.destroy).pack(side="right", padx=5)

    def _refresh_debit_credit_options(self):
        """根据当前输入列刷新借/贷下拉选项"""
        if not hasattr(self, "debit_col_combo") or not hasattr(self, "credit_col_combo"):
            return
        options = [EMPTY_OPTION] + self.input_columns
        # 记住当前选项，若不存在则重置为空
        cur_debit = self.manual_debit_col_var.get()
        cur_credit = self.manual_credit_col_var.get()
        self.debit_col_combo["values"] = options
        self.credit_col_combo["values"] = options
        if cur_debit not in self.input_columns:
            self.manual_debit_col_var.set(EMPTY_OPTION)
        if cur_credit not in self.input_columns:
            self.manual_credit_col_var.set(EMPTY_OPTION)
        # 预计算借/贷关键列，提高后续行扫描性能
        self._auto_debit_cols = []
        self._auto_credit_cols = []
        debit_syn = {"借", "借方", "借方金额", "借方本币", "借方金额(本币)", "debit", "dr", "debito", "cargo", "deudor"}
        credit_syn = {"贷", "贷方", "贷方金额", "贷方本币", "贷方金额(本币)", "credit", "cr", "credito", "abono", "haber"}
        for c in self.input_columns:
            n = normalize_header(c)
            if n in debit_syn:
                self._auto_debit_cols.append(c)
            if n in credit_syn:
                self._auto_credit_cols.append(c)

    def _derive_debit_credit_context(self, src_row) -> Dict[str, Any]:
        """识别当前行的借/贷金额、类型等上下文信息"""
        manual_debit_col = self.manual_debit_col_var.get()
        manual_credit_col = self.manual_credit_col_var.get()
        manual_dc_col = self.manual_dc_col_var.get() # 新增

        debit_cols = []
        credit_cols = []
        manual_debit_set = bool(manual_debit_col and manual_debit_col in self.input_columns and manual_debit_col != EMPTY_OPTION)
        manual_credit_set = bool(manual_credit_col and manual_credit_col in self.input_columns and manual_credit_col != EMPTY_OPTION)
        if manual_debit_set:
            debit_cols.append(manual_debit_col)
        if manual_credit_set:
            credit_cols.append(manual_credit_col)

        # 若用户显式选择了借/贷辅助列，则以用户选择为准，不再自动猜测对应方向
        if not manual_debit_set:
            for c in getattr(self, "_auto_debit_cols", []):
                if c not in debit_cols:
                    debit_cols.append(c)
        if not manual_credit_set:
            for c in getattr(self, "_auto_credit_cols", []):
                if c not in credit_cols:
                    credit_cols.append(c)

        def _is_zero(val):
            try:
                return float(val) == 0.0
            except Exception:
                return False

        def _first_non_empty(cols):
            for c in cols:
                v = src_row.get(c, None)
                if v is not None and not (isinstance(v, float) and pd.isna(v)) and str(v).strip() != "":
                    if _is_zero(v):
                        continue
                    return v
            return None

        debit_val = _first_non_empty(debit_cols)
        credit_val = _first_non_empty(credit_cols)
        derived_amount = None
        derived_type = None

        # 优先使用具体的借/贷金额列
        if debit_val is not None:
            derived_amount = debit_val
            derived_type = "3"
        elif credit_val is not None:
            derived_amount = credit_val
            derived_type = "4"
        
        # 如果还没确定方向，尝试使用借贷标志列
        if derived_type is None and manual_dc_col and manual_dc_col != EMPTY_OPTION:
            dc_val = str(src_row.get(manual_dc_col, "")).strip().lower()
            if dc_val:
                # 借方关键词
                debit_syn = {"借", "借方", "debit", "dr", "3", "1", "debito", "cargo", "deudor"}
                # 贷方关键词
                credit_syn = {"贷", "贷方", "credit", "cr", "4", "2", "credito", "abono", "haber"}
                
                if any(s in dc_val for s in debit_syn if not s.isdigit()) or dc_val in dc_val in debit_syn:
                    derived_type = "3"
                elif any(s in dc_val for s in credit_syn if not s.isdigit()) or dc_val in dc_val in credit_syn:
                    derived_type = "4"
                
                # 如果通过标志列确定了方向，但 derived_amount 还是空的，则尝试从已映射的金额列获取
                if derived_type and derived_amount is None:
                    # 尝试从“金额”或“外币金额”映射中获取原始数值
                    target_fields = ["金额", "外币金额"]
                    for tf in target_fields:
                        if tf in self.mapping_vars:
                            src_c = self.mapping_vars[tf].get()
                            if src_c and src_c != EMPTY_OPTION and not src_c.startswith(COMPOSITE_PREFIX):
                                val = src_row.get(src_c)
                                if val is not None and not pd.isna(val) and str(val).strip() != "":
                                    derived_amount = val
                                    break

        alt_amount_signed = None
        income_cols = ["收入明细金额", "收入金额", "收入"]
        expense_cols = ["支出明细金额", "支出金额", "支出", "费用金额"]
        income_val = _first_non_empty(income_cols)
        expense_val = _first_non_empty(expense_cols)
        if income_val is not None:
            try:
                alt_amount_signed = float(income_val)
            except Exception:
                alt_amount_signed = income_val
        elif expense_val is not None:
            try:
                alt_amount_signed = -float(expense_val)
            except Exception:
                alt_amount_signed = -abs(expense_val) if isinstance(expense_val, (int, float)) else expense_val

        if derived_amount is None and alt_amount_signed is not None:
            try:
                f_abs = abs(float(alt_amount_signed))
                derived_amount = f_abs
                derived_type = "3" if float(alt_amount_signed) >= 0 else "4"
            except Exception:
                derived_amount = alt_amount_signed

        signed_amount_from_derived = None
        if derived_amount is not None and derived_type:
            try:
                f_amt = float(derived_amount)
                signed_amount_from_derived = f_amt if derived_type == "3" else -abs(f_amt)
            except (ValueError, TypeError):
                signed_amount_from_derived = derived_amount

        return {
            "debit_val": debit_val,
            "credit_val": credit_val,
            "derived_amount": derived_amount,
            "derived_type": derived_type,
            "signed_amount": signed_amount_from_derived,
            "alt_amount_signed": alt_amount_signed,
        }

    def _compute_composite_value(
        self,
        name: str,
        src_row,
        mapping: Dict[str, Optional[str]],
        derived_ctx: Dict[str, Any],
        smart_data: Optional[Dict[str, Any]] = None,
        cache: Optional[Dict[str, Any]] = None,
        visited: Optional[set] = None,
        depth: int = 0,
    ):
        comps = self._get_all_composite_fields()
        if name not in comps:
            return None
        if visited is None:
            visited = set()
        if name in visited:
            print(f"综合字段[{name}]存在循环引用")
            return None
        if depth > MAX_COMPOSITE_DEPTH:
            print(f"综合字段[{name}]嵌套过深，已停止计算")
            return None
        visited.add(name)

        cfg = comps[name]
        c_type = cfg.get("type", "expression")

        if c_type == "debit_credit":
            val = derived_ctx.get("signed_amount")
            if val is None:
                val = derived_ctx.get("alt_amount_signed")
            if cache is not None:
                cache[name] = val
            visited.discard(name)
            return val

        expr = cfg.get("expression") or cfg.get("expr")
        if not expr:
            visited.discard(name)
            return None

        def col(col_name: str):
            try:
                return src_row.get(col_name, None)
            except Exception:
                try:
                    return src_row[col_name]
                except Exception:
                    return None

        def field(tmpl_name: str):
            src_col = mapping.get(tmpl_name)
            if not src_col or src_col == EMPTY_OPTION:
                # 尝试获取默认值（含别名）
                return self._resolve_default_value(tmpl_name)
            if self._is_composite_option(src_col):
                inner_name = self._extract_composite_name(src_col)
                if cache is not None and inner_name in cache:
                    return cache[inner_name]
                return self._compute_composite_value(inner_name, src_row, mapping, derived_ctx, smart_data, cache, visited, depth + 1)
            return col(src_col)

        def composite_fn(c_name: str):
            if cache is not None and c_name in cache:
                return cache[c_name]
            return self._compute_composite_value(c_name, src_row, mapping, derived_ctx, smart_data, cache, visited, depth + 1)

        def lookup_fn(table: str, key_col: str, val: Any, target_col: str):
            if not self.base_data_mgr:
                return None
            return self.base_data_mgr.lookup_value(table, key_col, val, target_col)

        env = {
            "col": col,
            "field": field,
            "mapped": field,
            "composite": composite_fn,
            "lookup": lookup_fn,
            "derived": derived_ctx or {},
            "smart": smart_data or {},
            "abs": abs,
            "round": round,
            "min": min,
            "max": max,
            "Decimal": Decimal,
            "float": float,
            "int": int,
            "str": str,
        }

        try:
            val = safe_eval_expr(expr, env)
            if cache is not None:
                cache[name] = val
            visited.discard(name)
            return val
        except Exception as e:
            print(f"综合字段[{name}]计算失败: {e}")
            visited.discard(name)
            return None

    # ---------- 预览映射 ----------
    def _preview_mapping(self):
        """预览当前映射配置及示例数据转换效果"""
        if self.input_df is None:
            messagebox.showwarning("提示", "请先选择原始 Excel 文件。")
            return
        if not self.template_headers:
            messagebox.showwarning("提示", "未加载模板表头。")
            return

        # 创建预览窗口
        preview_win = tk.Toplevel(self.root)
        preview_win.title("映射预览 - 模板列与源数据对照")
        preview_win.geometry("1100x600")

        main_frame = ttk.Frame(preview_win, padding=10)
        main_frame.pack(fill="both", expand=True)

        # 顶部说明
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill="x", pady=(0, 10))

        # 解析当前有效模式
        raw_mode = self.convert_mode_var.get()
        effective_mode = raw_mode
        if raw_mode.startswith("方案: "):
            scheme_name = raw_mode.replace("方案: ", "")
            if hasattr(self, "cached_schemes") and scheme_name in self.cached_schemes:
                effective_mode = self.cached_schemes[scheme_name]["base_mode"]
            else:
                effective_mode = MODE_GENERAL_VOUCHER

        ttk.Label(info_frame, text=f"当前模式: {raw_mode}", font=("", 10, "bold")).pack(side="left")
        template_label = self.template_path_var.get()
        if effective_mode == MODE_ORIGINAL:
            template_label = MODE_ORIGINAL
        ttk.Label(info_frame, text=f"  |  模板文件: {template_label}", foreground="gray").pack(side="left", padx=10)
        ttk.Label(info_frame, text=f"  |  源文件: {os.path.basename(self.input_path)}", foreground="gray").pack(side="left")
        template_has_type = any(h.name == "类型" for h in self.template_headers)
        manual_debit_col = self.manual_debit_col_var.get()
        manual_credit_col = self.manual_credit_col_var.get()
        use_raw_output = effective_mode == MODE_ORIGINAL

        # 创建 Treeview 显示映射预览
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill="both", expand=True)

        columns = ("template_col", "source_col", "sample_source", "sample_converted", "field_rule")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=20)

        tree.heading("template_col", text="模板列名")
        tree.heading("source_col", text="映射源列")
        tree.heading("sample_source", text="源数据示例")
        tree.heading("sample_converted", text="转换后示例")
        tree.heading("field_rule", text="字段规则")

        tree.column("template_col", width=120, minwidth=100)
        tree.column("source_col", width=150, minwidth=100)
        tree.column("sample_source", width=250, minwidth=150)
        tree.column("sample_converted", width=250, minwidth=150)
        tree.column("field_rule", width=200, minwidth=100)

        # 滚动条
        scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        attach_treeview_tools(tree)

        # 标签样式
        tree.tag_configure("mapped", background="#e6ffe6")  # 已映射 - 浅绿
        tree.tag_configure("unmapped", background="#fff0f0")  # 未映射 - 浅红
        tree.tag_configure("default", background="#fff9e6")  # 使用默认值 - 浅黄

        # 获取前3行示例数据
        sample_df = self.input_df.head(3) if not self.input_df.empty else pd.DataFrame()

        current_mapping = {}
        for header in self.template_headers:
            var = self.mapping_vars.get(header.name)
            if var:
                sel = var.get()
                current_mapping[header.name] = None if sel == EMPTY_OPTION else sel

        # 填充数据
        for header in self._get_header_display_list():
            # 通用凭证模式下，仅显示核心字段
            # if effective_mode == MODE_GENERAL_VOUCHER:
            #    if header.name not in GENERAL_VOUCHER_FIELDS:
            #        continue

            template_col = header.name
            var = self.mapping_vars.get(header.name)
            source_col = var.get() if var else EMPTY_OPTION

            # 获取示例值
            sample_source = ""
            sample_converted = ""

            if source_col and source_col != EMPTY_OPTION and source_col in sample_df.columns:
                # 获取源数据示例（最多3个值）
                sample_vals = sample_df[source_col].dropna().astype(str).tolist()[:3]
                sample_source = " | ".join(sample_vals) if sample_vals else "(空)"

                # 获取转换后示例
                if use_raw_output:
                    sample_converted = sample_source or "(空)"
                else:
                    converted_vals = []
                    for val in sample_vals[:3]:
                        try:
                            converted = convert_value(header.name, val, self.field_formats.get(header.name))
                            converted_vals.append(str(converted) if converted else "(空)")
                        except:
                            converted_vals.append("(转换错误)")
                    sample_converted = " | ".join(converted_vals) if converted_vals else "(空)"
            elif self._is_composite_option(source_col):
                comp_name = self._extract_composite_name(source_col)
                sample_source = f"{comp_name or '综合字段'}"
                preview_vals = []
                for _, row in sample_df.iterrows():
                    ctx = self._derive_debit_credit_context(row)
                    comp_cache = {}
                    comp_val = self._compute_composite_value(comp_name, row, current_mapping, ctx, smart_data={}, cache=comp_cache)
                    if comp_val is not None:
                        preview_vals.append(comp_val)
                    if len(preview_vals) >= 3:
                        break

                if preview_vals:
                    if use_raw_output:
                        sample_converted = " | ".join(str(v) for v in preview_vals)
                    else:
                        converted_vals = []
                        for val in preview_vals:
                            try:
                                converted_vals.append(str(convert_value(header.name, val, self.field_formats.get(header.name))))
                            except:
                                converted_vals.append("(转换错误)")
                        sample_converted = " | ".join(converted_vals)
                else:
                    sample_converted = "(空)"
            elif (not use_raw_output) and header.name == "金额" and (not template_has_type) and (
                (manual_debit_col and manual_debit_col in sample_df.columns) or
                (manual_credit_col and manual_credit_col in sample_df.columns)
            ):
                # 使用手动选择的借/贷列展示预览（带符号）
                sample_source_parts = []
                if manual_debit_col and manual_debit_col in sample_df.columns:
                    debit_vals = sample_df[manual_debit_col].dropna().astype(str).tolist()[:3]
                    if debit_vals:
                        sample_source_parts.append(f"借:{' | '.join(debit_vals)}")
                if manual_credit_col and manual_credit_col in sample_df.columns:
                    credit_vals = sample_df[manual_credit_col].dropna().astype(str).tolist()[:3]
                    if credit_vals:
                        sample_source_parts.append(f"贷:{' | '.join(credit_vals)}")
                sample_source = " ; ".join(sample_source_parts) if sample_source_parts else "(未映射)"

                preview_vals = []
                for _, row in sample_df.iterrows():
                    v = None
                    if manual_debit_col and manual_debit_col in sample_df.columns:
                        raw_d = row[manual_debit_col]
                        if raw_d is not None and not (isinstance(raw_d, float) and pd.isna(raw_d)) and str(raw_d).strip() != "":
                            try:
                                v = float(raw_d)
                            except:
                                v = raw_d
                    if v is None and manual_credit_col and manual_credit_col in sample_df.columns:
                        raw_c = row[manual_credit_col]
                        if raw_c is not None and not (isinstance(raw_c, float) and pd.isna(raw_c)) and str(raw_c).strip() != "":
                            try:
                                v = -float(raw_c)
                            except:
                                v = raw_c
                    if v is not None:
                        preview_vals.append(v)
                    if len(preview_vals) >= 3:
                        break

                if preview_vals:
                    converted_vals = []
                    for val in preview_vals:
                        try:
                            converted_vals.append(str(convert_value(header.name, val, self.field_formats.get(header.name))))
                        except:
                            converted_vals.append("(转换错误)")
                    sample_converted = " | ".join(converted_vals)
                else:
                    sample_converted = "(空)"
            elif source_col == EMPTY_OPTION:
                # 检查是否有默认值
                default_val = self.default_values.get(header.name)
                if default_val and not use_raw_output:
                    sample_source = f"[默认值]"
                    sample_converted = str(default_val)
                else:
                    sample_source = "(未映射)"
                    sample_converted = "(空)"

            # 获取字段规则描述
            rule = FIELD_RULES.get(header.name, {"type": "text"})
            rule_desc = f"类型: {rule.get('type', 'text')}"
            if rule.get("max_len"):
                rule_desc += f", 最大长度: {rule['max_len']}"
            if rule.get("max_int_len"):
                rule_desc += f", 整数位: {rule['max_int_len']}"
            if rule.get("max_decimal_len") is not None:
                rule_desc += f", 小数位: {rule['max_decimal_len']}"

            # 确定标签
            tags = ()
            manual_mapped = (
                header.name == "金额"
                and (not template_has_type)
                and (
                    (manual_debit_col and manual_debit_col in self.input_columns) or
                    (manual_credit_col and manual_credit_col in self.input_columns)
                )
            )
            if (source_col and source_col != EMPTY_OPTION) or manual_mapped:
                tags = ("mapped",)
            elif self.default_values.get(header.name):
                tags = ("default",)
            else:
                tags = ("unmapped",)

            tree.insert("", "end", values=(template_col, source_col, sample_source, sample_converted, rule_desc), tags=tags)

        # 底部统计和按钮
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill="x", pady=(10, 0))

        # 统计信息
        total_fields = len([h for h in self.template_headers if effective_mode != MODE_GENERAL_VOUCHER or h.name in GENERAL_VOUCHER_FIELDS])
        mapped_count = sum(1 for h in self.template_headers
                          if (effective_mode != MODE_GENERAL_VOUCHER or h.name in GENERAL_VOUCHER_FIELDS)
                          and self.mapping_vars.get(h.name)
                          and self.mapping_vars[h.name].get() != EMPTY_OPTION)
        # 手动借/贷列合并时，将金额视为已映射
        if (not template_has_type) and any([
            manual_debit_col and manual_debit_col in self.input_columns,
            manual_credit_col and manual_credit_col in self.input_columns
        ]):
            amt_var = self.mapping_vars.get("金额")
            already_counted = amt_var and amt_var.get() != EMPTY_OPTION
            if not already_counted:
                mapped_count += 1
        if effective_mode == MODE_ORIGINAL:
            default_count = 0
        else:
            default_count = sum(1 for h in self.template_headers
                               if (effective_mode != MODE_GENERAL_VOUCHER or h.name in GENERAL_VOUCHER_FIELDS)
                               and (not self.mapping_vars.get(h.name) or self.mapping_vars[h.name].get() == EMPTY_OPTION)
                               and self.default_values.get(h.name))
        unmapped_count = total_fields - mapped_count - default_count

        stats_text = f"总字段: {total_fields}  |  已映射: {mapped_count}  |  使用默认值: {default_count}  |  未映射: {unmapped_count}"
        ttk.Label(bottom_frame, text=stats_text, font=("", 9)).pack(side="left")

        # 图例
        legend_frame = ttk.Frame(bottom_frame)
        legend_frame.pack(side="left", padx=20)
        tk.Label(legend_frame, text="  ", bg="#e6ffe6", width=3).pack(side="left")
        ttk.Label(legend_frame, text="已映射").pack(side="left", padx=(2, 10))
        tk.Label(legend_frame, text="  ", bg="#fff9e6", width=3).pack(side="left")
        ttk.Label(legend_frame, text="默认值").pack(side="left", padx=(2, 10))
        tk.Label(legend_frame, text="  ", bg="#fff0f0", width=3).pack(side="left")
        ttk.Label(legend_frame, text="未映射").pack(side="left", padx=(2, 0))

        ttk.Button(bottom_frame, text="关闭", command=preview_win.destroy).pack(side="right")

        preview_win.transient(self.root)
        preview_win.grab_set()

    def _calculate_content_score(self, template_header: str, src_col: str, samples: list) -> float:
        """计算内容匹配度分数 (0.0 ~ 1.0)"""
        if not samples:
            return 0.0
            
        str_samples = [str(s).strip() for s in samples if str(s).strip()]
        if not str_samples:
            return 0.0

        # 1. 摘要字段：检查智能识别缓存命中率 & 文本特征
        if template_header in ["摘要", "摘要名", "说明", "用途"]:
            hits = 0
            text_len_score = 0
            negative_score = 0
            
            if self.base_data_mgr:
                for s in str_samples:
                    # 负面特征：纯数字、日期格式、过短
                    if s.replace('.','').replace('-','').isdigit() or len(s) < 2:
                        negative_score += 1
                        continue
                    if re.search(r'\d{4}[-/]\d{1,2}', s):
                        negative_score += 1
                        continue
                        
                    # 正面特征1: 缓存命中 (模糊匹配)
                    if self.base_data_mgr.get_cached_recognition_fuzzy(s):
                        hits += 1
                    # 正面特征2: 长度适中且包含中文或空格
                    if len(s) > 4:
                        text_len_score += 1
            
            total = len(str_samples)
            if negative_score / total > 0.5:
                return 0.0
                
            cache_score = (hits / total)
            feature_score = (text_len_score / total)
            
            # 如果缓存命中率高，给予极高分
            if cache_score > 0.3: return 0.95
            return max(cache_score * 1.0, feature_score * 0.8)

        # 2. 日期字段
        if template_header in ["凭证日期", "日期", "业务日期"]:
            dates = 0
            for s in str_samples:
                # 简单日期检测 (YYYY-MM-DD, YYYYMMDD, YYYY/MM/DD, 2025.01.01)
                if re.search(r'\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2}', s) or \
                   (re.search(r'\d{8}', s) and s.startswith('20')):
                    dates += 1
            return (dates / len(str_samples)) * 1.0

        # 3. 金额字段
        if template_header in ["金额", "外币金额", "借方金额", "贷方金额", "金额1", "供应价"]:
            # 负面特征：列名包含 ID, Code, Date, 号, 码 (防止误判 ID 为金额)
            norm_col = src_col.lower()
            if any(k in norm_col for k in ["id", "code", "date", "no.", "号", "码", "日期"]):
                return 0.0
                
            nums = 0
            for s in str_samples:
                try:
                    # 去除千分位
                    val = float(s.replace(',', ''))
                    # 排除 0 或 极小的数 (可能是状态码) - 视情况而定，这里暂不排除
                    nums += 1
                except:
                    pass
            return (nums / len(str_samples)) * 1.0
            
        return 0.0

    # ---------- 自动识别匹配 ----------
    def auto_match(self):
        if not self.input_columns:
            messagebox.showwarning("提示", "请先选择原始 Excel 文件并加载工作表。")
            return

        self.log_message("开始自动识别匹配...")
        self._debug_log(f"源列数量: {len(self.input_columns)}")

        # 预取示例值 (扩大到10行以提高内容分析准确性)
        sample_values = {}
        if self.input_df is not None and not self.input_df.empty:
            head = self.input_df.head(10)
            for col in self.input_columns:
                vals = head[col].dropna().astype(str).tolist()
                sample_values[col] = vals

        used_columns = set() # 记录已使用的源列，防止重复映射

        # 1. 收集所有可能的匹配及其分数
        all_matches = []
        for header in self.template_headers:
            if header.name not in self.mapping_vars:
                continue
            
            # 计算该表头对每一列的评分
            col_scores = []
            # 扩展源列范围，纳入综合字段
            candidate_cols = self.input_columns + self._composite_option_labels()
            for src_col in candidate_cols:
                if self._is_composite_option(src_col):
                    # 综合字段评分逻辑：如果名字高度相似，给高分
                    c_name = self._extract_composite_name(src_col)
                    name_score = score_similarity(header.name, c_name, header.name)
                    # 综合字段通常没有样本值用于内容评分，赋予名字更高权重
                    total_score = name_score
                    if total_score > 0.8: # 综合字段要求更高的名字匹配度
                         col_scores.append((src_col, total_score))
                else:
                    name_score = score_similarity(header.name, src_col, header.name)
                    content_score = self._calculate_content_score(header.name, src_col, sample_values.get(src_col, []))
                    
                    if header.name in ["摘要", "摘要名", "凭证日期", "日期", "金额", "外币金额"]:
                        if content_score == 0 and name_score < 0.9: 
                            total_score = name_score * 0.5
                        else:
                            total_score = name_score * 0.4 + content_score * 0.6
                    else:
                        total_score = name_score
                    
                    if total_score > 0.5: # 初步筛选
                        col_scores.append((src_col, total_score))
            
            # 排序取最佳
            col_scores.sort(key=lambda x: x[1], reverse=True)
            if col_scores:
                best_col, best_score = col_scores[0]
                all_matches.append({
                    "header": header,
                    "col": best_col,
                    "score": best_score,
                    "is_amount": header.name in ["金额", "外币金额", "借方金额", "贷方金额", "金额1", "供应价"]
                })
            else:
                # 即使没有高分匹配，也要保留条目以便后续处理（如AI匹配或置空）
                all_matches.append({
                    "header": header,
                    "col": None,
                    "score": 0,
                    "is_amount": header.name in ["金额", "外币金额", "借方金额", "贷方金额", "金额1", "供应价"]
                })

        # 2. 按分数降序处理，优先满足高置信度的匹配
        all_matches.sort(key=lambda x: x["score"], reverse=True)

        for match in all_matches:
            header = match["header"]
            col = match["col"]
            score = match["score"]
            is_amount = match["is_amount"]
            
            final_col = None

            # 策略：如果分数够高且未被占用（或允许复用），则采纳
            if col and score >= 0.6:
                if col not in used_columns or is_amount:
                    final_col = col
                    used_columns.add(col)
            
            # 如果自动算法未匹配到，且启用了AI，尝试AI
            if not final_col:
                # 尝试 AI 辅助匹配
                ai_col = None
                ai_client = getattr(self.summary_recognizer, "ai_client", None) if self.summary_recognizer else None
                if self.use_ai_var.get() and self.summary_recognizer and ai_client:
                    try:
                        ai_samples = {k: v[:5] for k, v in sample_values.items()}
                        ai_col = self._ai_match_column(header.name, self.input_columns, ai_samples)
                    except Exception as e:
                        print(f"AI 列匹配失败: {e}")
                        ai_col = None
                
                if ai_col:
                    # AI 推荐的列也要检查唯一性
                    if ai_col not in used_columns or is_amount:
                        final_col = ai_col
                        used_columns.add(ai_col)

            # 设置结果
            if final_col:
                self.mapping_vars[header.name].set(final_col)
                self.log_message(f"  [自动匹配] {header.name} <== {final_col} (分数: {score:.2f})")
            else:
                self.mapping_vars[header.name].set(EMPTY_OPTION)

        self.log_message("自动识别匹配完成。")

        # --- 自动匹配借贷辅助列 ---
        debit_keywords = ["借方", "借方金额", "debit", "dr", "借方本币"]
        credit_keywords = ["贷方", "贷方金额", "credit", "cr", "贷方本币"]
        dc_keywords = ["借贷", "借贷标志", "借贷类型", "dc", "direction", "tipo", "sentido"]
        
        # 仅当未手动设置时才尝试自动匹配
        if self.manual_dc_col_var.get() == EMPTY_OPTION:
            for col in self.input_columns:
                norm_col = normalize_header(col)
                if any(k in norm_col for k in dc_keywords):
                    self.manual_dc_col_var.set(col)
                    break

        if self.manual_debit_col_var.get() == EMPTY_OPTION:
            for col in self.input_columns:
                norm_col = normalize_header(col)
                if any(k in norm_col for k in debit_keywords):
                    self.manual_debit_col_var.set(col)
                    break
        
        if self.manual_credit_col_var.get() == EMPTY_OPTION:
            for col in self.input_columns:
                norm_col = normalize_header(col)
                if any(k in norm_col for k in credit_keywords):
                    self.manual_credit_col_var.set(col)
                    break

        messagebox.showinfo("自动识别完成", "已根据表头名称自动匹配列，请检查映射结果，如有需要可手动调整。")

    def _ai_match_column(self, template_header: str, src_columns: list, samples: dict) -> str:
        """调用 AI 辅助选择最合适的源列"""
        if not src_columns:
            return ""

        # 获取模板字段的批注/规则信息
        header_obj = next((h for h in self.template_headers if h.name == template_header), None)
        comment = header_obj.comment if header_obj else ""
        rule = FIELD_RULES.get(template_header, {"type": "text"})
        rule_desc = f"字段类型: {rule.get('type', 'text')}"
        if rule.get("max_len"):
            rule_desc += f"，最大长度: {rule['max_len']}"
        if rule.get("max_int_len"):
            rule_desc += f"，整数位: {rule['max_int_len']}"
        if rule.get("max_decimal_len") is not None:
            rule_desc += f"，小数位: {rule['max_decimal_len']}"

        # 构建提示
        column_desc = []
        for col in src_columns:
            vals = [v for v in samples.get(col, []) if str(v).strip() != ""]
            # 去重并截取前5个有代表性的样本
            seen = []
            filtered = []
            for v in vals:
                if v in seen:
                    continue
                seen.append(v)
                filtered.append(v)
                if len(filtered) >= 5:
                    break
            # 简单数值检测，给出范围
            numeric_vals = []
            for v in filtered:
                try:
                    numeric_vals.append(float(v))
                except Exception:
                    pass
            stats = ""
            if numeric_vals:
                stats = f" (数值范围: {min(numeric_vals)}~{max(numeric_vals)})"
            sample_str = " | ".join(str(v) for v in filtered) if filtered else ""
            column_desc.append(f"{col}: {sample_str}{stats}")

        all_template_headers = ", ".join(h.name for h in self.template_headers)

        prompt = f"""你是表头映射助手。请把源表列映射到模板字段。
模板字段: {template_header}
字段规则: {rule_desc}
字段批注: {comment or '无'}
模板全部字段: {all_template_headers}
源表列及样例值:
{chr(10).join(column_desc)}

要求：
1) 选择最匹配模板字段的源表列名。
2) 只能返回源表中存在的列名（原样返回）。
3) 如果不确定或没有合适的列，返回 None。
"""

        ai_client = getattr(self.summary_recognizer, "ai_client", None)
        if not ai_client:
            return ""

        resp = ai_client.chat.completions.create(
            model=getattr(self.summary_recognizer, "ai_model_name", "glm-4-flash"),
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
        )
        result = resp.choices[0].message.content.strip()
        # 直接返回匹配到的列名
        for col in src_columns:
            if col == result:
                return col
        # 简单包含匹配
        for col in src_columns:
            if result in col or col in result:
                return col
        return ""

    def show_comment(self, header: TemplateHeader):
        text = header.comment.strip() if header.comment else "该列暂无批注说明。"
        messagebox.showinfo(f"{header.name} 批注", text)

    def _get_base_data_favorites_key(self, table_name: str) -> str:
        return f"base_data_favorites_{table_name}"

    def _load_base_data_favorites(self, table_name: str) -> List[str]:
        if not self.base_data_mgr or not table_name:
            return []
        try:
            raw_favs = self.base_data_mgr.get_config(self._get_base_data_favorites_key(table_name), "[]")
            parsed = json.loads(raw_favs) if raw_favs else []
            if isinstance(parsed, list):
                return [str(v).strip() for v in parsed if str(v).strip()]
        except Exception:
            pass
        return []

    def _save_base_data_favorites(self, table_name: str, favorites: List[str]):
        if not self.base_data_mgr or not table_name:
            return
        try:
            cleaned = [str(v).strip() for v in favorites if str(v).strip()]
            self.base_data_mgr.set_config(
                self._get_base_data_favorites_key(table_name),
                json.dumps(cleaned, ensure_ascii=True)
            )
        except Exception:
            pass

    def _get_base_data_favorite_key(
        self,
        table_name: str,
        row_data: Dict[str, Any],
        display_cols: Optional[List[str]] = None
    ) -> str:
        code = str(row_data.get("code", "") or "").strip()
        name = str(row_data.get("name", "") or "").strip()
        code_name = str(row_data.get("code_name", "") or "").strip()

        if table_name == "account_subject":
            if code:
                return code
            if code_name:
                m = re.search(r"\[(\d+)\]", code_name)
                if m:
                    return m.group(1)
                return code_name

        if code:
            return code
        if name:
            return name

        cols = display_cols or list(row_data.keys())
        for col in cols:
            if col == "id":
                continue
            value = str(row_data.get(col, "") or "").strip()
            if value:
                return value
        return ""

    def _open_base_data_selector(
        self,
        parent,
        table_name,
        target_var,
        title="选择基础数据",
        favorites_only_default=False,
        lock_favorites_only=False
    ):
        """打开基础数据选择对话框"""
        if not self.base_data_mgr:
            messagebox.showwarning("提示", "基础数据管理器未初始化，无法选择。")
            return

        # 检查表名是否合法
        try:
            self.base_data_mgr._assert_valid_table(table_name)
        except ValueError:
            messagebox.showwarning("提示", f"不支持的基础数据表: {table_name}")
            return

        # 使用传入的 parent 作为父窗口
        target_parent = parent if parent else self.root
        selector = tk.Toplevel(target_parent)
        selector.title(title)
        selector.geometry("700x500")
        
        # 关键：设置为模态并置顶于父窗口
        selector.transient(target_parent)
        selector.grab_set()

        # 搜索框
        top_frame = ttk.Frame(selector, padding=5)
        top_frame.pack(fill="x")
        ttk.Label(top_frame, text="搜索:").pack(side="left")
        search_var = tk.StringVar()
        entry_search = ttk.Entry(top_frame, textvariable=search_var)
        entry_search.pack(side="left", fill="x", expand=True, padx=5)
        show_favorites_only_var = tk.BooleanVar(value=bool(favorites_only_default))
        fav_chk = ttk.Checkbutton(top_frame, text="仅显示收藏", variable=show_favorites_only_var)
        fav_chk.pack(side="right")
        if lock_favorites_only:
            fav_chk.state(["disabled"])
        
        # 数据列表
        tree_frame = ttk.Frame(selector, padding=5)
        tree_frame.pack(fill="both", expand=True)
        
        columns = self.base_data_mgr.get_table_columns(table_name)
        # 优化列显示
        display_cols = [c for c in columns if c not in ["id", "match_items", "search_keyword"]]
        
        tree = ttk.Treeview(tree_frame, columns=display_cols, show="headings")
        for col in display_cols:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        
        scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")

        # 加载收藏
        favorites = self._load_base_data_favorites(table_name)

        def _favorite_key(row):
            return self._get_base_data_favorite_key(table_name, row, display_cols)

        # 加载数据
        all_data = self.base_data_mgr.query(table_name)
        
        def populate_tree(data):
            for item in tree.get_children():
                tree.delete(item)
            fav_set = set(favorites)
            fav_rows = []
            normal_rows = []
            for row in data:
                key = _favorite_key(row)
                if key and key in fav_set:
                    fav_rows.append(row)
                else:
                    normal_rows.append(row)
            if show_favorites_only_var.get():
                ordered = fav_rows
            else:
                ordered = fav_rows + normal_rows
            for row in ordered:
                vals = [row.get(c, "") for c in display_cols]
                key = _favorite_key(row)
                tags = ["row"]
                if key and key in fav_set:
                    tags.append("favorite")
                tree.insert("", "end", values=vals, tags=tuple(tags))

        populate_tree(all_data)
        tree.tag_configure("favorite", background="#fff7d6")

        # 搜索功能
        def do_search(*args):
            kw = search_var.get().lower().strip()
            if not kw:
                populate_tree(all_data)
                return
            filtered = []
            for row in all_data:
                # 全字段模糊搜索
                if any(kw in str(v).lower() for v in row.values()):
                    filtered.append(row)
            populate_tree(filtered)

        search_var.trace("w", do_search)
        show_favorites_only_var.trace("w", do_search)

        # 确认选择
        def on_select(event=None):
            selection = tree.selection()
            if not selection:
                return
            item = tree.item(selection[0])
            vals = item["values"]
            if not vals:
                return
            
            row_data = dict(zip(display_cols, vals))
            selected_val = ""
            
            # 特殊处理 account_subject
            if table_name == "account_subject":
                code_name = row_data.get("code_name", "") or row_data.get("code", "")
                m = re.search(r'\[(\d+)\]', str(code_name))
                if m:
                    selected_val = m.group(1)
                else:
                    selected_val = str(code_name).strip()
            else:
                # 优先取 code，其次 name
                if "code" in row_data:
                    selected_val = row_data["code"]
                elif "name" in row_data:
                    selected_val = row_data["name"]
                # 部门表特殊情况
                elif table_name == "department" and "部门编码" in row_data: # 实际上已统一为code
                     selected_val = row_data.get("code")
                else:
                    # 兜底：取第一列
                    selected_val = vals[0]
            
            if selected_val:
                target_var.set(selected_val)
                selector.destroy()

        tree.bind("<Double-1>", on_select)
        
        btn_frame = ttk.Frame(selector, padding=5)
        btn_frame.pack(fill="x")

        def _update_favorites(add: bool):
            selection = tree.selection()
            if not selection:
                return
            current = list(favorites)
            changed = False
            for item in selection:
                values = tree.item(item, "values")
                row_data = dict(zip(display_cols, values))
                key = _favorite_key(row_data)
                if not key:
                    continue
                if add and key not in current:
                    current.append(key)
                    changed = True
                if not add and key in current:
                    current.remove(key)
                    changed = True
            if changed:
                favorites[:] = current
                self._save_base_data_favorites(table_name, favorites)
                populate_tree(all_data)

        ttk.Button(btn_frame, text="加入收藏", command=lambda: _update_favorites(True)).pack(side="left")
        ttk.Button(btn_frame, text="取消收藏", command=lambda: _update_favorites(False)).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="选择", command=on_select).pack(side="right")
        ttk.Button(btn_frame, text="取消", command=selector.destroy).pack(side="right", padx=5)
        return selector

    def _persist_default_value_groups(self):
        """持久化默认值组配置"""
        if not self.base_data_mgr:
            return
        try:
            payload = json.dumps(self.default_value_groups, ensure_ascii=True)
            self.base_data_mgr.set_config("default_value_groups", payload)
            if self.active_default_group:
                self.base_data_mgr.set_config("default_value_group_active", self.active_default_group)
        except Exception as e:
            print(f"保存默认值组失败: {e}")

    def _sync_current_default_group(self):
        """同步当前默认值到活跃默认组"""
        group_name = self.active_default_group or "默认"
        if not self.default_value_groups:
            self.default_value_groups = {}
        self.default_value_groups[group_name] = dict(self.default_values)
        self.active_default_group = group_name
        if self.base_data_mgr:
            self._persist_default_value_groups()

    def _normalize_subject_code(self, value: Any) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        match = re.search(r"\[(\d+)\]", raw)
        if match:
            return match.group(1)
        if raw.endswith(".0"):
            raw = raw[:-2]
        return raw

    def _ask_pre_convert_subject_and_account(self, effective_mode: str) -> bool:
        """转换前询问并设置对方科目、部门、费用上级科目限制、默认账户、是否使用外币及汇率"""
        if effective_mode not in [MODE_GENERAL_VOUCHER, MODE_CUSTOM]:
            return True

        dialog = tk.Toplevel(self.root)
        dialog.title("生成前参数确认")
        dialog.geometry("620x400")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        subject_var = tk.StringVar(value=self._normalize_subject_code(
            self.default_values.get("对方科目") or self.default_values.get("科目编码", "")
        ))
        department_var = tk.StringVar(value=str(self.default_values.get("部门", "") or "").strip())
        expense_parent_options = {
            "1.6601(销售费用)": "6601",
            "2.6602（管理费用）": "6602",
            "3.6603（财务费用）": "6603",
            "4.不受限制": "UNLIMITED",
        }
        expense_parent_reverse = {v: k for k, v in expense_parent_options.items()}
        expense_parent_limit = str(self.default_values.get("费用科目上级限制", "") or "").strip()
        expense_parent_var = tk.StringVar(
            value=expense_parent_reverse.get(expense_parent_limit, "4.不受限制")
        )
        account_var = tk.StringVar(value=str(self._resolve_default_value("默认账户") or "").strip())
        use_foreign_var = tk.BooleanVar(value=self.use_foreign_currency_var.get())
        rate_var = tk.StringVar(value=str(self.default_values.get("汇率", "") or "1"))
        save_group_var = tk.BooleanVar(value=True)
        result = {"ok": False}

        body = ttk.Frame(dialog, padding=12)
        body.pack(fill="both", expand=True)

        ttk.Label(
            body,
            text="生成前请确认“对方科目 / 部门 / 费用上级科目限制 / 默认账户 / 外币模式 / 汇率”：",
            font=("", 10, "bold")
        ).pack(anchor="w", pady=(0, 8))

        row_subject = ttk.Frame(body)
        row_subject.pack(fill="x", pady=4)
        ttk.Label(row_subject, text="对方科目:", width=10).pack(side="left")
        ttk.Entry(row_subject, textvariable=subject_var, width=34).pack(side="left", fill="x", expand=True, padx=4)
        ttk.Button(
            row_subject,
            text="选择...",
            width=8,
            command=lambda: self._open_base_data_selector(dialog, "account_subject", subject_var, "选择对方科目")
        ).pack(side="left", padx=(4, 0))

        row_department = ttk.Frame(body)
        row_department.pack(fill="x", pady=4)
        ttk.Label(row_department, text="部门:", width=10).pack(side="left")
        ttk.Entry(row_department, textvariable=department_var, width=34).pack(side="left", fill="x", expand=True, padx=4)
        ttk.Button(
            row_department,
            text="选择...",
            width=8,
            command=lambda: self._open_base_data_selector(dialog, "department", department_var, "选择部门")
        ).pack(side="left", padx=(4, 0))

        row_expense_parent = ttk.Frame(body)
        row_expense_parent.pack(fill="x", pady=4)
        ttk.Label(row_expense_parent, text="费用上级:", width=10).pack(side="left")
        expense_parent_combo = ttk.Combobox(
            row_expense_parent,
            textvariable=expense_parent_var,
            values=list(expense_parent_options.keys()),
            state="readonly",
            width=34,
        )
        expense_parent_combo.pack(side="left", fill="x", expand=True, padx=4)

        row_account = ttk.Frame(body)
        row_account.pack(fill="x", pady=4)
        ttk.Label(row_account, text="默认账户:", width=10).pack(side="left")
        ttk.Entry(row_account, textvariable=account_var, width=34).pack(side="left", fill="x", expand=True, padx=4)
        ttk.Button(
            row_account,
            text="选择...",
            width=8,
            command=lambda: self._open_base_data_selector(dialog, "bank_account", account_var, "选择默认账户")
        ).pack(side="left", padx=(4, 0))

        row_foreign = ttk.Frame(body)
        row_foreign.pack(fill="x", pady=4)
        ttk.Label(row_foreign, text="外币模式:", width=10).pack(side="left")
        ttk.Checkbutton(row_foreign, text="使用外币模式", variable=use_foreign_var).pack(side="left", padx=4)

        row_rate = ttk.Frame(body)
        row_rate.pack(fill="x", pady=4)
        ttk.Label(row_rate, text="汇率:", width=10).pack(side="left")
        rate_entry = ttk.Entry(row_rate, textvariable=rate_var, width=20)
        rate_entry.pack(side="left", padx=4)
        ttk.Label(row_rate, text="例如: 7.01031", foreground="gray").pack(side="left", padx=4)

        def _toggle_rate_entry(*_args):
            rate_entry.config(state="normal" if use_foreign_var.get() else "disabled")

        use_foreign_var.trace_add("write", _toggle_rate_entry)
        _toggle_rate_entry()

        ttk.Checkbutton(body, text="保存到当前默认值组", variable=save_group_var).pack(anchor="w", pady=(8, 0))
        ttk.Label(
            body,
            text="提示：默认账户优先按“对方科目”匹配；无匹配时保留输入或回退往来编码。",
            foreground="gray"
        ).pack(anchor="w", pady=(4, 0))

        btn_frame = ttk.Frame(body)
        btn_frame.pack(fill="x", pady=(12, 0))

        def _confirm():
            subject = self._normalize_subject_code(subject_var.get())
            department = str(department_var.get() or "").strip()
            expense_parent_label = str(expense_parent_var.get() or "").strip()
            expense_parent_limit = expense_parent_options.get(expense_parent_label, "UNLIMITED")
            account = str(account_var.get() or "").strip()
            use_foreign = bool(use_foreign_var.get())
            rate_text = str(rate_var.get() or "").strip()

            if not subject and self.auto_balance_var.get():
                proceed = messagebox.askyesno(
                    "提示",
                    "当前未设置“对方科目”，自动生成对方分录时将回退到默认科目或 1002。是否继续？",
                    parent=dialog
                )
                if not proceed:
                    return

            if not department:
                messagebox.showwarning("提示", "请选择部门。", parent=dialog)
                return

            if use_foreign:
                if not rate_text:
                    messagebox.showwarning("提示", "已启用外币模式，请输入汇率。", parent=dialog)
                    return
                try:
                    rate_val = float(rate_text)
                    if rate_val <= 0:
                        messagebox.showwarning("提示", "汇率必须大于 0。", parent=dialog)
                        return
                except Exception:
                    messagebox.showwarning("提示", "汇率格式无效，请输入数字。", parent=dialog)
                    return

            if subject:
                self.default_values["对方科目"] = subject
            else:
                self.default_values.pop("对方科目", None)

            if account:
                self.default_values["默认账户"] = account
                self.default_values["账户"] = account
                self._prefer_prompt_default_account = True
            else:
                self.default_values.pop("默认账户", None)
                self.default_values.pop("账户", None)
                self._prefer_prompt_default_account = False

            self.default_values["部门"] = department
            self.default_values["默认部门"] = department

            self.default_values["费用科目上级限制"] = expense_parent_limit

            if rate_text:
                self.default_values["汇率"] = rate_text

            self.use_foreign_currency_var.set(use_foreign)
            if self.base_data_mgr:
                self.base_data_mgr.set_config("setting_use_foreign_currency_var", str(use_foreign))

            if save_group_var.get():
                self._sync_current_default_group()

            self._update_recognizer_defaults()
            self.log_message(
                f"生成前设置已确认: 对方科目={subject or '(空)'}，部门={department}，"
                f"费用上级限制={expense_parent_limit}，默认账户={account or '(空)'}，"
                f"外币模式={'开启' if use_foreign else '关闭'}，汇率={rate_text or '(空)'}"
            )
            result["ok"] = True
            dialog.destroy()

        def _cancel():
            dialog.destroy()

        ttk.Button(btn_frame, text="取消本次生成", command=_cancel, width=12).pack(side="right")
        ttk.Button(btn_frame, text="确认并继续", command=_confirm, width=12).pack(side="right", padx=6)

        dialog.protocol("WM_DELETE_WINDOW", _cancel)
        self.root.wait_window(dialog)
        return result["ok"]

    def _show_default_values_dialog(self):
        """显示默认值设置对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("设置默认值")
        dialog.geometry("650x500")

        # 创建主框架
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill="both", expand=True)

        ttk.Label(
            main_frame,
            text="为未映射或未识别的字段设置默认值：",
            font=("", 10, "bold")
        ).pack(anchor="w", pady=(0, 5))
        
        ttk.Label(
             main_frame, 
             text="提示：点击蓝色标题或双击输入框可选择基础数据", 
             foreground="blue", 
             font=("", 9)
        ).pack(anchor="w", pady=(0, 10))

        # 默认值组
        local_groups = dict(self.default_value_groups)
        local_active_group = self.active_default_group or (next(iter(local_groups.keys())) if local_groups else "默认")

        def _group_values_to_entries(group_name):
            values = local_groups.get(group_name, {})
            for field_name, var in entry_vars.items():
                var.set(values.get(field_name, ""))

        def _entries_to_group_values():
            result = {}
            for field_name, var in entry_vars.items():
                value = var.get().strip()
                if value:
                    result[field_name] = value
            return result

        group_frame = ttk.Frame(main_frame)
        group_frame.pack(fill="x", pady=(0, 6))
        ttk.Label(group_frame, text="默认值组:").pack(side="left")
        group_var = tk.StringVar(value=local_active_group)
        group_combo = ttk.Combobox(group_frame, textvariable=group_var, state="readonly", width=18)
        group_combo.pack(side="left", padx=5)

        def refresh_group_options():
            names = list(local_groups.keys())
            if not names:
                names = ["默认"]
            group_combo["values"] = names
            if group_var.get() not in names:
                group_var.set(names[0])

        def on_group_selected(_event=None):
            _group_values_to_entries(group_var.get())

        def add_group():
            name = simpledialog.askstring("新增默认值组", "请输入默认值组名称：", parent=dialog)
            if not name:
                return
            name = name.strip()
            if not name:
                return
            if name in local_groups:
                messagebox.showwarning("提示", "该默认值组已存在。")
                return
            local_groups[name] = _entries_to_group_values()
            refresh_group_options()
            group_var.set(name)
            _group_values_to_entries(name)

        ttk.Button(group_frame, text="新增组", command=add_group, width=8).pack(side="left", padx=5)
        group_combo.bind("<<ComboboxSelected>>", on_group_selected)
        refresh_group_options()

        # 常用字段的默认值设置
        default_fields = [
            ("部门", "部门编码或名称，例如: 10001"),
            ("科目编码", "默认科目编码，例如: 1002"),
            ("对方科目", "默认对方科目编码，例如: 1001 (当智能识别无法确定时使用)"),
            ("默认账户", "默认账户编码，例如: BAC001"),
            ("外币金额", "外币金额，例如: 0"),
            ("汇率", "汇率，例如: 1 或 6.5"),
            ("序号", "序号，例如: 1"),
            ("会计凭证No.", "会计凭证号"),
            ("发货仓库", "出库仓库编码或名称（销售出库）"),
            ("默认仓库", "备用默认仓库编码/名称（为空时回退到发货仓库）"),
            ("交易类型", "销售出库交易类型编码，例如: 11 或 12"),
            ("货币", "默认货币编码，例如: USD 或 CNY"),
            ("职员", "经办人编码或名称"),
            ("收货公司", "默认收货公司/客户名称"),
            ("生成生产入库", "Y/N，留空默认为 N"),
        ]

        # 字段对应表映射
        FIELD_TABLE_MAP = {
            "部门": "department",
            "科目编码": "account_subject",
            "对方科目": "account_subject",
            "默认账户": "bank_account",
            "发货仓库": "warehouse",
            "默认仓库": "warehouse",
            "货币": "currency",
            "收货公司": "business_partner",
        }

        entry_vars = {
            field_name: tk.StringVar(value=self.default_values.get(field_name, ""))
            for field_name, _ in default_fields
        }

        def _build_default_fields_panel(parent_frame, favorites_only=False, lock_favorites_only=False, only_mapped=False):
            container = ttk.Frame(parent_frame)
            container.pack(fill="both", expand=True)

            canvas = tk.Canvas(container)
            scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            for field_name, hint in default_fields:
                table_name = FIELD_TABLE_MAP.get(field_name)
                if only_mapped and not table_name:
                    continue

                frame = ttk.Frame(scrollable_frame)
                frame.pack(fill="x", pady=5)

                lbl = ttk.Label(frame, text=f"{field_name}:", width=15)
                lbl.pack(side="left")

                var = entry_vars[field_name]
                entry = ttk.Entry(frame, textvariable=var, width=30)
                entry.pack(side="left", padx=5)

                if table_name:
                    lbl.configure(foreground="blue", cursor="hand2")

                    def _open_selector(t=table_name, v=var, f=field_name, fo=favorites_only, lf=lock_favorites_only):
                        self._open_base_data_selector(
                            dialog,
                            t,
                            v,
                            f"选择{f}",
                            favorites_only_default=fo,
                            lock_favorites_only=lf
                        )

                    lbl.bind("<Button-1>", lambda e, op=_open_selector: op())
                    entry.bind("<Double-1>", lambda e, op=_open_selector: op())
                    ttk.Button(frame, text="...", width=3, command=_open_selector).pack(side="left", padx=(0, 5))

                ttk.Label(frame, text=hint, foreground="gray", font=("", 8)).pack(side="left")

        fields_notebook = ttk.Notebook(main_frame)
        fields_notebook.pack(fill="both", expand=True)
        tab_all = ttk.Frame(fields_notebook)
        tab_favorites = ttk.Frame(fields_notebook)
        fields_notebook.add(tab_all, text="全部字段")
        fields_notebook.add(tab_favorites, text="仅显示收藏")

        _build_default_fields_panel(tab_all, favorites_only=False, lock_favorites_only=False, only_mapped=False)
        _build_default_fields_panel(tab_favorites, favorites_only=True, lock_favorites_only=True, only_mapped=True)

        # 底部按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))

        def save_defaults(close_dialog=True):
            group_name = group_var.get().strip() or "默认"
            local_groups[group_name] = _entries_to_group_values()

            # 同步到全局并持久化
            self.default_value_groups = local_groups
            self.active_default_group = group_name
            self.default_values = dict(local_groups.get(group_name, {}))
            if self.base_data_mgr:
                self._persist_default_value_groups()

            # 更新识别器
            self._update_recognizer_defaults()

            messagebox.showinfo("成功", f"已保存默认值设置\n当前组: {group_name}\n共设置 {len(self.default_values)} 个默认值")
            if close_dialog:
                dialog.destroy()

        def clear_defaults():
            # 清空所有默认值
            for var in entry_vars.values():
                var.set("")
            messagebox.showinfo("成功", "已清空当前默认值输入")

        ttk.Button(button_frame, text="保存", command=lambda: save_defaults(True), width=12).pack(side="right", padx=5)
        ttk.Button(button_frame, text="应用", command=lambda: save_defaults(False), width=12).pack(side="right", padx=5)
        ttk.Button(button_frame, text="清空", command=clear_defaults, width=12).pack(side="right")
        ttk.Button(button_frame, text="取消", command=dialog.destroy, width=12).pack(side="right")

        # 使窗口模态
        dialog.transient(self.root)
        dialog.grab_set()

    # ---------- 转换与导出 ----------
    def _restore_base_data_codes(self):
        """尝试还原基础数据编码 (往来单位 & 品目信息)"""
        if self.input_df is None:
            messagebox.showwarning("提示", "请先加载 Excel 文件。")
            return

        if not self.base_data_mgr:
            messagebox.showwarning("提示", "基础数据管理器未初始化。")
            return

        # 1. 确定映射关系
        def get_mapped_col(field_name):
            var = self.mapping_vars.get(field_name)
            if var:
                val = var.get()
                if val and val != EMPTY_OPTION and not self._is_composite_option(val):
                    return val
            return None

        col_account_code = get_mapped_col("科目编码")
        col_account_name = get_mapped_col("科目名称") or get_mapped_col("科目名")
        col_partner_code = get_mapped_col("往来单位编码")
        col_partner_name = get_mapped_col("往来单位名")
        col_item_code = get_mapped_col("品目编码")
        col_item_name = get_mapped_col("品目名")

        # 如果未找到映射，弹出对话框让用户手动选择
        if not any([col_account_code, col_account_name, col_partner_code, col_partner_name, col_item_code, col_item_name]):
            # 准备列选项
            columns = self.input_columns

            # 创建弹窗
            dialog = tk.Toplevel(self.root)
            dialog.title("选择还原字段")
            dialog.geometry("400x300")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="未检测到已映射的字段，请手动选择对应的列：").pack(pady=10, padx=10, anchor="w")

            frame = ttk.Frame(dialog)
            frame.pack(fill="x", padx=10)

            vars_manual = {}

            def add_row(label, key):
                r = ttk.Frame(frame)
                r.pack(fill="x", pady=2)
                ttk.Label(r, text=label, width=15).pack(side="left")
                v = tk.StringVar(value=EMPTY_OPTION)
                cb = ttk.Combobox(r, textvariable=v, values=[EMPTY_OPTION] + columns, state="readonly")
                cb.pack(side="left", fill="x", expand=True)
                vars_manual[key] = v

            add_row("科目编码:", "account_code")
            add_row("科目名称:", "account_name")
            add_row("往来单位编码:", "partner_code")
            add_row("往来单位名称:", "partner_name")
            add_row("品目编码:", "item_code")
            add_row("品目名称:", "item_name")

            result_container = {}

            def on_confirm():
                result_container["account_code"] = vars_manual["account_code"].get()
                result_container["account_name"] = vars_manual["account_name"].get()
                result_container["partner_code"] = vars_manual["partner_code"].get()
                result_container["partner_name"] = vars_manual["partner_name"].get()
                result_container["item_code"] = vars_manual["item_code"].get()
                result_container["item_name"] = vars_manual["item_name"].get()

                # 简单的空值处理
                for k, v in result_container.items():
                    if v == EMPTY_OPTION:
                        result_container[k] = None

                dialog.destroy()

            ttk.Button(dialog, text="开始还原", command=on_confirm).pack(pady=20)

            self.root.wait_window(dialog)

            if not result_container:  # 用户关闭窗口或取消
                return

            col_partner_code = result_container.get("partner_code")
            col_partner_name = result_container.get("partner_name")
            col_item_code = result_container.get("item_code")
            col_item_name = result_container.get("item_name")

            if not any([col_account_code, col_account_name, col_partner_code, col_partner_name, col_item_code, col_item_name]):
                return

        # 2. 遍历处理
        count_account = 0
        count_partner = 0
        count_item = 0
        
        # 进度窗口
        progress_win = tk.Toplevel(self.root)
        progress_win.title("正在还原编码...")
        progress_win.geometry("350x120")
        
        ttk.Label(progress_win, text="正在智能分析与还原，请稍候...").pack(pady=20)
        progress_bar = ttk.Progressbar(progress_win, mode='indeterminate')
        progress_bar.pack(fill='x', padx=20)
        progress_bar.start()
        
        # 居中显示
        progress_win.transient(self.root)
        progress_win.grab_set()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 175
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 60
        progress_win.geometry(f"+{x}+{y}")
        
        self.root.update()

        try:
            # --- 处理科目编码 ---
            if col_account_code:
                target_col = col_account_code
                ref_col = col_account_name
                if target_col in self.input_df.columns:
                    for idx in self.input_df.index:
                        current_val = safe_str(self.input_df.at[idx, target_col])
                        match = self.base_data_mgr.find_best_match("account_subject", current_val)
                        if not match and ref_col and ref_col in self.input_df.columns:
                            ref_val = safe_str(self.input_df.at[idx, ref_col])
                            match = self.base_data_mgr.find_best_match("account_subject", ref_val)
                        if match and match != current_val:
                            self.input_df.at[idx, target_col] = match
                            count_account += 1

            # --- 处理往来单位 ---
            if col_partner_code:
                target_col = col_partner_code
                ref_col = col_partner_name
                
                # 检查 target_col 是否存在
                if target_col in self.input_df.columns:
                    for idx in self.input_df.index:
                        current_val = safe_str(self.input_df.at[idx, target_col])
                        
                        # 尝试匹配
                        match = self.base_data_mgr.find_best_match("business_partner", current_val)
                        
                        if not match and ref_col and ref_col in self.input_df.columns:
                            ref_val = safe_str(self.input_df.at[idx, ref_col])
                            match = self.base_data_mgr.find_best_match("business_partner", ref_val)
                        
                        if match and match != current_val:
                            self.input_df.at[idx, target_col] = match
                            count_partner += 1

            # --- 处理品目信息 ---
            if col_item_code:
                target_col = col_item_code
                ref_col = col_item_name
                
                if target_col in self.input_df.columns:
                    for idx in self.input_df.index:
                        current_val = safe_str(self.input_df.at[idx, target_col])
                        match = self.base_data_mgr.find_best_match("product", current_val)
                        
                        if not match and ref_col and ref_col in self.input_df.columns:
                            ref_val = safe_str(self.input_df.at[idx, ref_col])
                            match = self.base_data_mgr.find_best_match("product", ref_val)
                            
                        if match and match != current_val:
                            self.input_df.at[idx, target_col] = match
                            count_item += 1
            
            # 清除缓存
            self.base_data_mgr.clear_lookup_cache()

        except Exception as e:
            messagebox.showerror("错误", f"处理过程中出错: {e}")
        finally:
            progress_win.destroy()
        
        msg = (
            "智能还原完成！\n\n"
            f"- 修复科目编码: {count_account} 条\n"
            f"- 修复往来单位编码: {count_partner} 条\n"
            f"- 修复品目编码: {count_item} 条\n\n"
            "请点击“预览映射”查看结果。"
        )
        messagebox.showinfo("结果", msg)

    def _restore_codes_in_df(self, df: pd.DataFrame, title: str = "智能还原编码",
                             show_progress: bool = True, show_alerts: bool = True):
        if df is None:
            if show_alerts:
                messagebox.showwarning("提示", "暂无数据可处理。")
            return {"account": 0, "partner": 0}
        if not self.base_data_mgr:
            if show_alerts:
                messagebox.showwarning("提示", "基础数据管理器未初始化。")
            return {"account": 0, "partner": 0}

        def pick_col(candidates):
            for name in candidates:
                if name in df.columns:
                    return name
            return None

        col_account_code = pick_col(["科目编码", "科目代码", "会计科目", "科目"])
        col_account_name = pick_col(["科目名称", "科目名", "科目"])
        col_partner_code = pick_col(["往来单位编码", "客户编码", "供应商编码", "往来编码"])
        col_partner_name = pick_col(["往来单位名", "往来单位名称", "客户名称", "供应商名称", "单位名称"])

        if not any([col_account_code, col_account_name, col_partner_code, col_partner_name]):
            if show_alerts:
                messagebox.showwarning("提示", "未找到科目编码或往来单位编码相关列。")
            return {"account": 0, "partner": 0}

        count_account = 0
        count_partner = 0

        progress_win = None
        if show_progress:
            progress_win = tk.Toplevel(self.root)
            progress_win.title(title)
            progress_win.geometry("350x120")
            ttk.Label(progress_win, text="正在智能分析与还原，请稍候...").pack(pady=20)
            progress_bar = ttk.Progressbar(progress_win, mode='indeterminate')
            progress_bar.pack(fill='x', padx=20)
            progress_bar.start()
            progress_win.transient(self.root)
            progress_win.grab_set()
            x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 175
            y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 60
            progress_win.geometry(f"+{x}+{y}")
            self.root.update()

        try:
            if col_account_code:
                for idx in df.index:
                    current_val = safe_str(df.at[idx, col_account_code])
                    match = self.base_data_mgr.find_best_match("account_subject", current_val)
                    if not match and col_account_name:
                        ref_val = safe_str(df.at[idx, col_account_name])
                        match = self.base_data_mgr.find_best_match("account_subject", ref_val)
                    if match and match != current_val:
                        df.at[idx, col_account_code] = match
                        count_account += 1

            if col_partner_code:
                for idx in df.index:
                    current_val = safe_str(df.at[idx, col_partner_code])
                    match = self.base_data_mgr.find_best_match("business_partner", current_val)
                    if not match and col_partner_name:
                        ref_val = safe_str(df.at[idx, col_partner_name])
                        match = self.base_data_mgr.find_best_match("business_partner", ref_val)
                    if match and match != current_val:
                        df.at[idx, col_partner_code] = match
                        count_partner += 1

            self.base_data_mgr.clear_lookup_cache()
        finally:
            if progress_win is not None:
                progress_win.destroy()

        return {"account": count_account, "partner": count_partner}

    def do_convert(self):
        self.log_message("开始执行转换...")
        self._preview_header_override = False
        self._preview_header_names = None
        self._debug_log(f"当前模式: {self.convert_mode_var.get()}")
        # 强制刷新配置（以防万一）
        if self.base_data_mgr:
            db_configs = self.base_data_mgr.get_all_configs()
            # 默认值优先使用“默认值组”；兼容无默认组时回退旧版散列配置。
            refreshed_defaults = {}
            loaded_from_groups = False
            raw_groups = db_configs.get("default_value_groups", "")
            if raw_groups:
                try:
                    parsed_groups = json.loads(raw_groups)
                    if isinstance(parsed_groups, dict) and parsed_groups:
                        self.default_value_groups = parsed_groups
                        raw_active_group = db_configs.get("default_value_group_active")
                        if raw_active_group in parsed_groups:
                            self.active_default_group = raw_active_group
                        elif self.active_default_group not in parsed_groups:
                            self.active_default_group = next(iter(parsed_groups.keys()))
                        refreshed_defaults = dict(parsed_groups.get(self.active_default_group, {}))
                        loaded_from_groups = True
                except Exception as e:
                    print(f"转换前刷新默认值组失败: {e}")

            if not loaded_from_groups:
                legacy_excluded = {
                    "api_key",
                    "setting_enable_smart_recognition",
                    "setting_use_ai_var",
                    "setting_use_foreign_currency_var",
                    "setting_auto_balance_var",
                    "setting_split_amount_var",
                    "setting_recon_fuzzy_var",
                    "setting_recon_use_ai_analysis",
                    "setting_invert_bank_var",
                    "main_field_formats",
                    "default_value_groups",
                    "default_value_group_active",
                }
                refreshed_defaults = {
                    k: v for k, v in db_configs.items()
                    if k not in legacy_excluded
                }

            # AI 与任务分发配置保留在 default_values，供识别上下文读取。
            for cfg_key in (
                "ai_provider",
                "ai_api_key",
                "ai_base_url",
                "ai_model_name",
                "recognition_priority",
                "ai_backends",
                "ai_task_map",
            ):
                if cfg_key in db_configs:
                    refreshed_defaults[cfg_key] = db_configs[cfg_key]

            self.default_values = refreshed_defaults

        print(f"调试: 当前所有默认值 keys: {list(self.default_values.keys())}")
        print(
            f"调试: 科目编码={self.default_values.get('科目编码')}, "
            f"对方科目={self.default_values.get('对方科目')}, "
            f"默认账户={self.default_values.get('默认账户')}"
        )

        if self.input_df is None:
            messagebox.showwarning("提示", "请先选择原始 Excel 文件。")
            return
        if not self.template_headers:
            messagebox.showerror("错误", "未加载模板表头，无法转换。")
            return

        mapping = {}
        for header in self.template_headers:
            var = self.mapping_vars.get(header.name)
            if not var:
                continue
            sel = var.get()
            mapping[header.name] = None if sel == EMPTY_OPTION else sel

        # 解析当前有效模式（处理自定义方案）
        raw_mode = self.convert_mode_var.get()
        effective_mode = raw_mode
        if raw_mode.startswith("方案: "):
            scheme_name = raw_mode.replace("方案: ", "")
            if hasattr(self, "cached_schemes") and scheme_name in self.cached_schemes:
                effective_mode = self.cached_schemes[scheme_name]["base_mode"]
            else:
                effective_mode = MODE_GENERAL_VOUCHER

        # 本次转换默认不强制覆盖账户；若用户在弹窗中输入了“默认账户”会再置为 True
        self._prefer_prompt_default_account = False
        if not self._ask_pre_convert_subject_and_account(effective_mode):
            self.log_message("用户取消了本次转换。")
            return

        # --- 调试日志：输出当前映射关系 ---
        try:
            _debug_mapping = {k: v for k, v in mapping.items() if v and v != EMPTY_OPTION}
            self.log_message(f"当前字段映射: {json.dumps(_debug_mapping, ensure_ascii=False)}")
        except Exception:
            pass

        self._debug_log(f"映射字段数: {len([k for k,v in mapping.items() if v])}")
        self._debug_log(f"借/贷辅助列: 借={self.manual_debit_col_var.get()}, 贷={self.manual_credit_col_var.get()}")
        self._debug_log(f"自动借列: {getattr(self, '_auto_debit_cols', [])}")
        self._debug_log(f"自动贷列: {getattr(self, '_auto_credit_cols', [])}")
        
        # Log Foreign Currency Mode state
        _use_foreign_mode = self.use_foreign_currency_var.get()
        self.log_message(f"外币模式状态: {'开启' if _use_foreign_mode else '关闭'}")

        # 原格式模式：不做智能识别/默认值/格式化，按映射原样输出
        if effective_mode == MODE_ORIGINAL:
            output_rows = []
            display_headers = self._get_header_display_list()
            for _, src_row in self.input_df.iterrows():
                out_row = []
                comp_cache = {}
                for header in display_headers:
                    src_col = mapping.get(header.name)
                    if src_col:
                        if self._is_composite_option(src_col):
                            comp_name = self._extract_composite_name(src_col)
                            src_value = self._compute_composite_value(
                                comp_name, src_row, mapping, {}, smart_data={}, cache=comp_cache
                            )
                        else:
                            src_value = src_row.get(src_col)
                    else:
                        src_value = ""
                    if src_value is None or (isinstance(src_value, float) and pd.isna(src_value)):
                        src_value = ""
                out_row.append(src_value)
                output_rows.append(out_row)
            if not output_rows:
                messagebox.showwarning("提示", "原始 Excel 中没有数据行。")
                return
            
            # 使用显示表头名导出
            self._preview_header_override = True
            self._preview_header_names = [h.name for h in display_headers]
            self._export_to_excel(output_rows, None)
            return

        # 销售出库模式走专用转换逻辑
        if effective_mode == MODE_SALES_OUTBOUND:
            output_rows, recognition_info, unmatched_info = self._convert_sales_outbound(mapping)
            if not output_rows:
                messagebox.showwarning("提示", "原始 Excel 中没有数据行。")
                return
            need_preview = True
            if need_preview:
                try:
                    preview_result = self._show_preprocessing_preview(output_rows, recognition_info, unmatched_info)
                except Exception as e:
                    import traceback
                    tb = traceback.format_exc()
                    self.log_message(f"预处理预览异常（销售出库）: {e}\n{tb}")
                    preview_result = messagebox.askyesno(
                        "预览异常",
                        f"预处理预览窗口异常。\n\n错误: {e}\n\n是否跳过预览并继续导出？"
                    )
                if not preview_result:
                    return
            self._export_to_excel(output_rows, unmatched_info)
            return

        # 准备转换数据
        output_rows = []
        recognition_info = []  # 存储智能识别信息
        unmatched_info = []

        print(f"调试: 智能识别开关={self.enable_smart_recognition.get()}, AI开关={self.use_ai_var.get()}")

        # 序号生成器状态
        serial_map = {}
        next_serial_id = 1

        # 模板是否包含“类型”列
        template_has_type = any(h.name == "类型" for h in self.template_headers)
        manual_debit_col = self.manual_debit_col_var.get()
        manual_credit_col = self.manual_credit_col_var.get()
        foreign_mode_logged = False
        forced_counter_subject, forced_counter_account = self._get_prompt_counter_subject_account_pair()
        if forced_counter_subject and forced_counter_account:
            self._debug_log(
                f"启用对方科目账户强绑定: 科目={forced_counter_subject} -> 账户={forced_counter_account}"
            )

        for idx, src_row in self.input_df.iterrows():
            # 准备行数据字典（用于智能识别）
            row_dict = {}
            for col_name in self.input_df.columns:
                row_dict[str(col_name)] = src_row[col_name]

            derived_ctx = self._derive_debit_credit_context(src_row)
            if idx < 5:
                self._debug_log(
                    f"行{idx+2}借贷推断: 借={derived_ctx.get('debit_val')}, "
                    f"贷={derived_ctx.get('credit_val')}, 类型={derived_ctx.get('derived_type')}, "
                    f"金额={derived_ctx.get('derived_amount')}"
                )
            derived_amount = derived_ctx.get("derived_amount")
            derived_type = derived_ctx.get("derived_type")
            signed_amount_from_derived = derived_ctx.get("signed_amount")
            alt_amount_signed = derived_ctx.get("alt_amount_signed")

            # 如果启用智能识别，尝试从摘要中提取信息
            smart_data = {}
            row_summary_text = ""
            summary_col_for_row = mapping.get("摘要") or mapping.get("摘要名")
            if summary_col_for_row and summary_col_for_row in src_row.index:
                try:
                    _sv = src_row[summary_col_for_row]
                    if _sv is not None and not pd.isna(_sv):
                        row_summary_text = str(_sv)
                except Exception:
                    row_summary_text = ""
            if self.enable_smart_recognition.get() and self.summary_recognizer:
                # 获取摘要内容（支持多种列名：摘要、摘要名）
                summary_col = mapping.get("摘要") or mapping.get("摘要名")
                if summary_col:
                    summary_value = src_row[summary_col] if summary_col in src_row.index else None
                    if summary_value and not pd.isna(summary_value):
                        # 智能识别
                        # 传递 AI 识别开关和外币模式开关
                        use_ai = self.use_ai_var.get()
                        use_foreign = self.use_foreign_currency_var.get()
                        smart_data = self.summary_recognizer.recognize(str(summary_value), row_dict, use_ai=use_ai, use_foreign_currency=use_foreign)
                        # 记录识别信息
                        if smart_data:
                            recognition_info.append({
                                "row_num": idx + 2,  # Excel行号（从2开始）
                                "summary": str(summary_value),
                                "recognized": smart_data
                            })

            # 现金存/取业务强制覆盖（即使未开启智能识别也生效）
            row_cash_override = self._get_cash_business_override(row_summary_text)
            if row_cash_override:
                smart_data.update(row_cash_override)
                if idx < 20:
                    self._debug_log(
                        f"行{idx+2}命中现金业务覆盖: 科目编码={row_cash_override.get('科目编码')}, "
                        f"往来单位编码={row_cash_override.get('往来单位编码')}"
                    )

            composite_cache = {}

            # --- 外币模式：行级汇率与外币金额 ---
            row_rate_str = None
            row_calc_rate = None
            row_f_amt = None
            if self.use_foreign_currency_var.get():
                # 汇率：优先行级映射/识别，其次默认值
                r_col = mapping.get("汇率")
                if r_col:
                    if self._is_composite_option(r_col):
                        row_rate_str = self._compute_composite_value(self._extract_composite_name(r_col), src_row, mapping, derived_ctx, smart_data, cache=composite_cache)
                    else:
                        row_rate_str = src_row.get(r_col)
                if row_rate_str is None or (isinstance(row_rate_str, float) and pd.isna(row_rate_str)) or str(row_rate_str).strip() == "":
                    row_rate_str = smart_data.get("汇率")
                if row_rate_str is None or str(row_rate_str).strip() == "":
                    row_rate_str = self.default_values.get("汇率", "1")

                # 预先同步汇率精度，确保计算与最终写入一致
                try:
                    formatted_rate_s = convert_value("汇率", row_rate_str, self.field_formats.get("汇率"))
                    row_calc_rate = float(str(formatted_rate_s).replace(",", ""))
                except Exception:
                    try:
                        row_calc_rate = float(row_rate_str)
                    except Exception:
                        row_calc_rate = 1.0

                # 外币金额：仅取行级映射/识别，不使用本币回退
                f_col = mapping.get("外币金额")
                if f_col:
                    if self._is_composite_option(f_col):
                        row_f_amt = self._compute_composite_value(self._extract_composite_name(f_col), src_row, mapping, derived_ctx, smart_data, cache=composite_cache)
                    else:
                        row_f_amt = src_row.get(f_col)
                if row_f_amt is None or (isinstance(row_f_amt, float) and pd.isna(row_f_amt)) or str(row_f_amt).strip() == "":
                    row_f_amt = smart_data.get("外币金额")

            # 构建输出行
            out_row = []
            row_type_inferred_from_amount_sign = False
            row_subject_for_account = ""
            row_counter_subject_for_account = ""
            row_partner_for_account = self._normalize_code_value(
                self.default_values.get("往来单位编码") or self.default_values.get("往来单位", "")
            )

            # 预先解析本行“科目编码”，用于按本行科目匹配往来单位编码，避免列顺序影响
            subject_src_col = mapping.get("科目编码")
            if subject_src_col:
                try:
                    if self._is_composite_option(subject_src_col):
                        comp_name = self._extract_composite_name(subject_src_col)
                        pre_subject_val = self._compute_composite_value(
                            comp_name, src_row, mapping, derived_ctx, smart_data, cache=composite_cache
                        )
                    else:
                        pre_subject_val = src_row.get(subject_src_col)
                    pre_subject_code = self._normalize_subject_code(pre_subject_val)
                    if pre_subject_code:
                        row_subject_for_account = pre_subject_code
                except Exception:
                    pass
            if not row_subject_for_account:
                smart_subject_code = self._normalize_subject_code(smart_data.get("科目编码"))
                if smart_subject_code:
                    row_subject_for_account = smart_subject_code

            # 预先解析本行“对方科目”，避免模板列顺序导致“默认账户”先于“对方科目”时错配
            counter_src_col = mapping.get("对方科目")
            if counter_src_col:
                try:
                    if self._is_composite_option(counter_src_col):
                        comp_name = self._extract_composite_name(counter_src_col)
                        pre_counter_val = self._compute_composite_value(
                            comp_name, src_row, mapping, derived_ctx, smart_data, cache=composite_cache
                        )
                    else:
                        pre_counter_val = src_row.get(counter_src_col)
                    pre_counter_subject = self._normalize_subject_code(pre_counter_val)
                    if pre_counter_subject:
                        row_counter_subject_for_account = pre_counter_subject
                except Exception:
                    pass

            if not row_counter_subject_for_account:
                smart_counter_subject = self._normalize_subject_code(smart_data.get("对方科目"))
                if smart_counter_subject:
                    row_counter_subject_for_account = smart_counter_subject

            # 预先解析本行“往来单位编码”，用于默认账户回退，避免列顺序影响
            partner_src_col = mapping.get("往来单位编码")
            if partner_src_col:
                try:
                    if self._is_composite_option(partner_src_col):
                        comp_name = self._extract_composite_name(partner_src_col)
                        pre_partner_val = self._compute_composite_value(
                            comp_name, src_row, mapping, derived_ctx, smart_data, cache=composite_cache
                        )
                    else:
                        pre_partner_val = src_row.get(partner_src_col)
                    pre_partner_code = self._normalize_code_value(pre_partner_val)
                    if pre_partner_code:
                        row_partner_for_account = pre_partner_code
                except Exception:
                    pass
            if not row_partner_for_account:
                smart_partner_code = self._normalize_code_value(smart_data.get("往来单位编码"))
                if smart_partner_code:
                    row_partner_for_account = smart_partner_code

            for header in self.template_headers:
                # 1. 获取智能识别结果（作为备选）
                smart_value = smart_data.get(header.name, None)
                if header.name == "类型":
                    # 类型仅由源数据/借贷列推断，忽略智能识别结果
                    smart_value = None

                # 2. 尝试获取手动映射的值
                src_col = mapping.get(header.name)
                src_value = None
                if src_col:
                    if self._is_composite_option(src_col):
                        comp_name = self._extract_composite_name(src_col)
                        src_value = self._compute_composite_value(comp_name, src_row, mapping, derived_ctx, smart_data, cache=composite_cache)
                    else:
                        src_value = src_row[src_col]

                # 3. 【核心修复】空值回退逻辑
                # 判断 src_value 是否为"有效空"（None, NaN, 或清除空白后的空字符串）
                is_src_empty = (src_value is None) or \
                               (isinstance(src_value, float) and pd.isna(src_value)) or \
                               (str(src_value).strip() == "")
                
                # 如果手动映射为空，且有智能识别结果，则使用智能识别结果
                if is_src_empty and smart_value is not None:
                    src_value = smart_value
                # 如果既没有手动映射，也没有智能识别，但没有手动映射列（src_col is None），则直接用智能识别（哪怕是None）
                elif src_col is None:
                    src_value = smart_value

                # 现金存/取业务字段强制覆盖（最高优先级）
                if row_cash_override:
                    if header.name == "科目编码":
                        src_value = row_cash_override.get("科目编码", src_value)
                    elif header.name == "往来单位编码":
                        src_value = row_cash_override.get("往来单位编码", src_value)

                # --- 借贷辅助列逻辑 (优先处理) ---
                if header.name == "金额" and not self.use_foreign_currency_var.get():
                    # 检查当前是否已获取到值
                    is_empty = (src_value is None) or \
                               (isinstance(src_value, float) and pd.isna(src_value)) or \
                               (str(src_value).strip() == "")
                    
                    if is_empty and derived_amount is not None:
                        if template_has_type:
                            # 有类型列 -> 填绝对值
                            try:
                                src_value = abs(float(derived_amount))
                            except Exception:
                                src_value = derived_amount
                        else:
                            # 无类型列 -> 填带符号值
                            if signed_amount_from_derived is not None:
                                src_value = signed_amount_from_derived

                if header.name == "金额" and template_has_type and self._is_composite_option(src_col):
                    comp_name = self._extract_composite_name(src_col)
                    comp_cfg = self._get_all_composite_fields().get(comp_name, {})
                    if comp_cfg.get("type") == "debit_credit":
                        try:
                            src_value = abs(float(src_value))
                        except Exception:
                            pass

                # 默认账户仅作内部匹配依据，不作为独立导出列写值
                if header.name == "默认账户":
                    if forced_counter_subject and forced_counter_account and row_subject_for_account == forced_counter_subject:
                        row_partner_for_account = forced_counter_account
                    else:
                        account_subject = row_subject_for_account
                        _linked_for_row = self._resolve_linked_default_account(
                            account_subject, src_value, row_partner_for_account, allow_fallback=False
                        )
                        if _linked_for_row:
                            row_partner_for_account = self._normalize_code_value(_linked_for_row)
                    src_value = ""
                elif header.name == "往来单位编码":
                    if forced_counter_subject and forced_counter_account and row_subject_for_account == forced_counter_subject:
                        # 强规则：指定对方科目对应的账户编码，直接覆盖源值
                        src_value = forced_counter_account
                    else:
                        partner_is_empty = (src_value is None) or \
                                           (isinstance(src_value, float) and pd.isna(src_value)) or \
                                           (str(src_value).strip() == "")
                        if partner_is_empty and row_subject_for_account:
                            account_subject = row_subject_for_account
                            linked_code = self._resolve_linked_default_account(
                                account_subject, None, row_partner_for_account, allow_fallback=False
                            )
                            if linked_code:
                                src_value = linked_code

                # 4. 最后尝试使用静态默认值 (当智能识别未启用或未覆盖时)
                is_src_empty_final = (src_value is None) or \
                               (isinstance(src_value, float) and pd.isna(src_value)) or \
                               (str(src_value).strip() == "")
                if is_src_empty_final:
                    if header.name not in ["默认账户", "往来单位编码"]:
                        default_val = self.default_values.get(header.name)
                        if default_val is not None and str(default_val).strip() != "":
                            src_value = default_val

                # --- 特殊处理：序号自动生成 (解决4字限制，保持单据捆绑) ---
                if header.name == "序号":
                    # 如果有源值（通常是原始单号），则通过映射生成短序号
                    s_val = str(src_value).strip() if src_value is not None and not pd.isna(src_value) else ""
                    if s_val:
                        if s_val not in serial_map:
                            serial_map[s_val] = str(next_serial_id)
                            next_serial_id += 1
                        src_value = serial_map[s_val]
                    else:
                        # 如果没有源值（未映射列），则默认自动递增（假设每一行是一张新凭证，除非后续有合并逻辑）
                        # 或者是为了避免空值报错
                        src_value = str(next_serial_id)
                        next_serial_id += 1
                # -----------------------------------------------------

                # --- 特殊处理：类型自动转换 (借/贷 -> 3/4) ---
                # 预先判断并存储推断的类型，只在处理"类型"列时才赋值给src_value
                inferred_type_for_row = None
                inferred_from_amount_sign = False
                s_val_for_type_inference = str(src_value).strip() if src_value is not None and not pd.isna(src_value) else ""

                # 1. 优先尝试从当前列值（关键字）推断
                if "借" in s_val_for_type_inference: inferred_type_for_row = "3"
                elif "贷" in s_val_for_type_inference: inferred_type_for_row = "4"
                elif "出" in s_val_for_type_inference: inferred_type_for_row = "1"
                elif "入" in s_val_for_type_inference: inferred_type_for_row = "2"
                elif derived_type:
                    inferred_type_for_row = derived_type
                
                # 2. 如果关键字推断失败，且当前列为空，则尝试从金额推断
                # (金额正数为借-3，负数为贷-4)
                if inferred_type_for_row is None:
                     # 无论当前列是否为空（只要没匹配到关键字），都尝试从金额反推
                     # 这样即使"类型"列填了奇怪的东西，也能纠正。但为了安全，还是限制在空值或无法识别的情况
                     
                     if s_val_for_type_inference == "" or (header.name == "类型"):
                        amt_col = mapping.get("金额")
                        raw_amt = None

                        # (A) 优先从“金额”映射列获取
                        if amt_col:
                            if self._is_composite_option(amt_col):
                                comp_name = self._extract_composite_name(amt_col)
                                raw_amt = self._compute_composite_value(comp_name, src_row, mapping, derived_ctx, smart_data, cache=composite_cache)
                            else:
                                raw_amt = src_row[amt_col]

                        # (B) 若“金额”未映射或取不到，再尝试“外币金额”（外币模式常见）
                        if raw_amt is None:
                            f_amt_col = mapping.get("外币金额")
                            if f_amt_col:
                                if self._is_composite_option(f_amt_col):
                                    comp_name = self._extract_composite_name(f_amt_col)
                                    raw_amt = self._compute_composite_value(comp_name, src_row, mapping, derived_ctx, smart_data, cache=composite_cache)
                                else:
                                    raw_amt = src_row.get(f_amt_col)

                        # (C) 外币模式下，最后回退到行级已解析外币金额
                        if raw_amt is None and self.use_foreign_currency_var.get() and row_f_amt is not None:
                            raw_amt = row_f_amt

                        # (D) 根据金额正负推断类型
                        if raw_amt is not None:
                            try:
                                f_amt = float(str(raw_amt).replace(",", ""))
                                if f_amt < 0:
                                    inferred_type_for_row = "4" # 贷
                                    inferred_from_amount_sign = True
                                else:
                                    inferred_type_for_row = "3" # 借
                                    inferred_from_amount_sign = True
                            except (ValueError, TypeError):
                                pass
                
                if header.name == "类型":
                    if inferred_type_for_row is not None:
                        src_value = inferred_type_for_row
                        row_type_inferred_from_amount_sign = inferred_from_amount_sign
                # ---------------------------------------------

                # --- 合并借/贷列为单列金额（模板没有“类型”列时保留符号） ---
                if header.name == "金额" and (not template_has_type):
                    is_empty_amount = (src_value is None) or (isinstance(src_value, float) and pd.isna(src_value)) or (str(src_value).strip() == "")
                    if is_empty_amount and signed_amount_from_derived is not None:
                        src_value = signed_amount_from_derived
                    # 收入/支出兜底：如果没有借贷列，尝试使用收入/支出列的带符号金额
                    if is_empty_amount and src_value is None:
                        if alt_amount_signed is not None:
                            src_value = alt_amount_signed

                # --- 特殊处理：强制拆分金额为借/贷两列 (输出) ---
                if self.split_amount_var.get() and (src_value is None or str(src_value).strip() == ""):
                    h_name = header.name
                    # 简单的关键字匹配判断是否为借贷列 (例如: "借方金额", "贷方金额")
                    is_debit_col = "借" in h_name and "金额" in h_name
                    is_credit_col = "贷" in h_name and "金额" in h_name
                    
                    if is_debit_col or is_credit_col:
                        use_amt = None
                        use_type = None

                        # 1. 优先使用已识别的借/贷 (源文件已有借贷列)
                        if derived_amount is not None and derived_type:
                            use_amt = derived_amount
                            use_type = derived_type
                        
                        # 2. 其次使用智能识别结果
                        if use_amt is None:
                            use_amt = smart_data.get("金额")

                        # 3. 最后尝试从手动映射的"金额"和"类型"列获取
                        if use_amt is None:
                            amt_col = mapping.get("金额")
                            if amt_col:
                                if self._is_composite_option(amt_col):
                                    comp_name = self._extract_composite_name(amt_col)
                                    use_amt = self._compute_composite_value(comp_name, src_row, mapping, derived_ctx, smart_data, cache=composite_cache)
                                else:
                                    use_amt = src_row.get(amt_col)

                            # 尝试推断类型
                            if inferred_type_for_row:
                                use_type = inferred_type_for_row
                            else:
                                type_col = mapping.get("类型")
                                if type_col:
                                    if self._is_composite_option(type_col):
                                        t_val = str(self._compute_composite_value(self._extract_composite_name(type_col), src_row, mapping, derived_ctx, smart_data, cache=composite_cache) or "")
                                    else:
                                        t_val = str(src_row.get(type_col, ""))
                                    if "借" in t_val or "3" in t_val: use_type = "3"
                                    elif "贷" in t_val or "4" in t_val: use_type = "4"

                        # 填值
                        if use_amt is not None:
                            try:
                                f_amt = abs(float(use_amt))
                                if is_debit_col and str(use_type) == "3":
                                    src_value = f_amt
                                elif is_credit_col and str(use_type) == "4":
                                    src_value = f_amt
                            except:
                                pass
                # ---------------------------------------------

                # --- 外币模式 ---
                # 规则：外币有值且金额为空 -> 计算金额；金额有值且外币为空 -> 不反算外币
                if self.use_foreign_currency_var.get():
                    if not foreign_mode_logged:
                        self.log_message(f"ℹ️ [外币模式] 已开启，默认汇率: {self.default_values.get('汇率', '1')}")
                        foreign_mode_logged = True

                    if header.name == "金额":
                        amt_has_explicit_mapping = bool(src_col and src_col != EMPTY_OPTION)
                        is_amt_empty = (src_value is None) or (isinstance(src_value, float) and pd.isna(src_value)) or (str(src_value).strip() == "")
                        # 外币存在且金额未显式映射时，优先按外币换算，避免摘要识别干扰
                        if row_f_amt is not None and not amt_has_explicit_mapping:
                            is_amt_empty = True
                        if is_amt_empty and row_f_amt is not None:
                            try:
                                from decimal import Decimal, ROUND_HALF_UP
                                _d_f = Decimal(str(row_f_amt).replace(",", ""))
                                _d_r = Decimal(str(row_calc_rate))
                                src_value = float((_d_f * _d_r).quantize(Decimal("0.00"), rounding=ROUND_HALF_UP))
                            except Exception:
                                pass

                    elif header.name == "外币金额":
                        # 外币为空时不反算
                        pass

                    elif header.name == "汇率":
                        if src_value is None or str(src_value).strip() == "":
                            src_value = row_rate_str
                # ---------------------------------------------

                # --- 特殊处理：金额取绝对值 ---
                if header.name in ["金额", "外币金额"] and (template_has_type or self.split_amount_var.get()):
                     if src_value is not None:
                         try:
                             # 尝试转为浮点数取绝对值，再转回字符串给后续 format_number 处理
                             f_val = float(src_value)
                             src_value = abs(f_val)
                         except (ValueError, TypeError):
                             pass

                # --- 借贷分列兜底：如果金额列为空，但有借/贷分列，优先使用 ---
                # --- [修正] 避免与外币模式冲突 ---
                # 只有在非外币模式下，才允许此处的逻辑覆盖金额，否则会破坏外币模式的精确计算
                if header.name == "金额" and not self.use_foreign_currency_var.get():
                    if derived_amount is not None:
                        try:
                            f_val = float(derived_amount)
                            if template_has_type and derived_type:
                                src_value = abs(f_val)
                            elif derived_type == "3":
                                src_value = abs(f_val)
                            elif derived_type == "4":
                                src_value = -abs(f_val)
                            else:
                                src_value = f_val
                        except (ValueError, TypeError):
                            src_value = derived_amount

                if header.name == "类型" and (src_value is None or str(src_value).strip() == ""):
                    if derived_type:
                        src_value = derived_type
                    elif self.use_foreign_currency_var.get():
                        fallback_signed = row_f_amt
                        if fallback_signed is None:
                            amt_col = mapping.get("金额")
                            if amt_col:
                                if self._is_composite_option(amt_col):
                                    comp_name = self._extract_composite_name(amt_col)
                                    fallback_signed = self._compute_composite_value(
                                        comp_name, src_row, mapping, derived_ctx, smart_data, cache=composite_cache
                                    )
                                else:
                                    fallback_signed = src_row.get(amt_col)
                        try:
                            if fallback_signed is not None and str(fallback_signed).strip() != "":
                                src_value = "4" if float(str(fallback_signed).replace(",", "")) < 0 else "3"
                                # 标记为“由金额正负推断”，让自动平账按“正负作用于对方科目”规则处理
                                row_type_inferred_from_amount_sign = True
                        except Exception:
                            pass
                # ---------------------------------------------

                if header.name == "科目编码":
                    subject_code = self._normalize_subject_code(src_value)
                    if subject_code:
                        row_subject_for_account = subject_code
                elif header.name == "对方科目":
                    counter_subject = self._normalize_subject_code(src_value)
                    if counter_subject:
                        row_counter_subject_for_account = counter_subject
                    # 对方科目用于自动平衡计算，不单独导出列值
                    src_value = ""
                elif header.name == "往来单位编码":
                    partner_code = self._normalize_code_value(src_value)
                    if partner_code:
                        row_partner_for_account = partner_code

                converted = convert_value(header.name, src_value, self.field_formats.get(header.name))
                out_row.append(converted)

            # 已取消“默认往来单位”自动回填，保留原始/识别得到的往来单位值

            # --- 校验：金额 = 外币 * 汇率 ---
            try:
                # 临时构建索引映射
                _h_map = {h.name: i for i, h in enumerate(self.template_headers)}
                if "金额" in _h_map and "外币金额" in _h_map and "汇率" in _h_map:
                    if _h_map["金额"] < len(out_row) and _h_map["外币金额"] < len(out_row) and _h_map["汇率"] < len(out_row):
                        _val_amt = out_row[_h_map["金额"]]
                        _val_f_amt = out_row[_h_map["外币金额"]]
                        _val_rate = out_row[_h_map["汇率"]]

                        def _safe_float(v):
                            try: return float(str(v).replace(",", ""))
                            except: return 0.0

                        _f_amt = _safe_float(_val_amt)
                        _f_f_amt = _safe_float(_val_f_amt)
                        _f_rate = _safe_float(_val_rate)

                        # 只有当有外币且有汇率时才校验
                        if abs(_f_f_amt) > 0.001 and abs(_f_rate) > 0.0001:
                            # 重新计算理论值，使用 Decimal 以保证最高精度
                            from decimal import Decimal, ROUND_HALF_UP
                            _d_f = Decimal(str(_f_f_amt))
                            _d_r = Decimal(str(_f_rate))
                            _expected = float((_d_f * _d_r).quantize(Decimal("0.00"), rounding=ROUND_HALF_UP))

                            # 允许 0.05 的误差 (主要应对复杂的浮点数累积误差或源数据多位汇率导致的分位差)
                            if abs(_f_amt - _expected) > 0.05:
                                self.log_message(f"⚠️ 警告: 行 {idx+2} 金额不一致! 金额({_f_amt}) != 外币({_f_f_amt}) * 汇率({_f_rate}) [理论值: {_expected:.2f}]")
            except Exception as e:
                print(f"校验金额逻辑出错: {e}")
            
            output_rows.append(out_row)

            # --- 自动生成对方分录 (借贷平衡) ---
            if self.auto_balance_var.get():
                # ... (existing auto-balance logic, omitted for brevity as it's not changing) ...
                # 获取当前行的关键信息
                current_type = ""
                current_amount = ""
                current_serial = ""
                current_subject = ""
                current_partner = ""
                
                # 需要找到对应列的索引
                header_map = {h.name: i for i, h in enumerate(self.template_headers)}
                
                if "类型" in header_map: current_type = out_row[header_map["类型"]]
                if "金额" in header_map: current_amount = out_row[header_map["金额"]]
                if "序号" in header_map: current_serial = out_row[header_map["序号"]]
                if "科目编码" in header_map:
                    current_subject = str(out_row[header_map["科目编码"]]).strip()
                    if current_subject.endswith(".0"):
                        current_subject = current_subject[:-2]
                if "往来单位编码" in header_map: current_partner = str(out_row[header_map["往来单位编码"]]).strip()
                if "部门" in header_map: current_department = str(out_row[header_map["部门"]]).strip()

                # 只有当类型明确为3或4时才生成对方分录
                if current_type in ["3", "4"]:
                    balance_row = list(out_row) # 复制原行数据

                    # 1. 确定借贷方向
                    # 规则：
                    # - 当类型由“金额正负”推断时：正负作用于对方科目，主科目取相反方向
                    # - 其他情况：主科目保持原方向，对方分录取相反方向
                    if row_type_inferred_from_amount_sign:
                        counterparty_type = current_type
                        current_type = "4" if counterparty_type == "3" else "3"
                        if "类型" in header_map:
                            out_row[header_map["类型"]] = current_type
                            balance_row[header_map["类型"]] = counterparty_type
                        self._debug_log(
                            f"行{idx+2}应用'正负作用于对方科目'规则: 当前科目类型={current_type}, 对方科目类型={counterparty_type}"
                        )
                    else:
                        new_type = "4" if current_type == "3" else "3"
                        if "类型" in header_map:
                            out_row[header_map["类型"]] = current_type
                            balance_row[header_map["类型"]] = new_type
                    
                    # 2. 设置对方科目
                    # 从默认值中获取：对方科目优先，回退科目编码
                    target_subject = self._normalize_subject_code(
                        self.default_values.get("对方科目") or self.default_values.get("科目编码", "")
                    )
                    
                    # 【安全检查】如果默认科目与当前行的往来单位编码相同，说明用户可能在设置中填错了（把往来填到了科目里）
                    # 此时强制清空默认科目，避免生成的对方科目变成客户编码
                    if target_subject and target_subject == current_partner:
                        print(f"警告: 默认科目({target_subject})与当前往来({current_partner})相同，疑似配置错误，已忽略默认科目。")
                        target_subject = ""

                    # print(f"调试: 默认对方科目={target_subject}") 

                    final_subject = target_subject
                    final_partner = "" # 最终决定的对方往来
                    
                    # --- 智能推断逻辑 ---
                    # 场景：如果原分录是银行类科目且有往来单位，对方科目优先 1122，避免 1002/1002 平账
                    curr_sub_str = str(current_subject).strip().replace(".0", "")
                    tgt_sub_str = str(target_subject).strip().replace(".0", "")

                    is_bank = curr_sub_str.startswith("100")
                    if not row_cash_override:
                        if is_bank and current_partner:
                            final_subject = "1122"
                            final_partner = current_partner
                        elif curr_sub_str == tgt_sub_str:
                            final_subject = "" # 默认先清空，避免借贷同科目
                            if current_type == "3" and current_partner:
                                final_subject = "1122"
                                final_partner = current_partner

                    # 如果仍未确定对方科目，回退到默认科目或银行科目，确保不为空
                    if not final_subject:
                        final_subject = target_subject or "1002"

                    # 填入最终决定的科目
                    if "科目编码" in header_map:
                        balance_row[header_map["科目编码"]] = final_subject
                    if "对方科目" in header_map:
                        # 对方科目用于内部平衡关系，不作为独立导出列
                        out_row[header_map["对方科目"]] = ""
                        balance_row[header_map["对方科目"]] = ""
                         
                    # 3. 处理往来单位
                    out_linked_acc = ""
                    bal_linked_acc = ""
                    out_partner_val = ""
                    bal_partner_val = ""
                    if "往来单位编码" in header_map:
                        p_idx = header_map["往来单位编码"]
                        out_partner_val = out_row[p_idx] if p_idx < len(out_row) else ""
                        bal_partner_val = balance_row[p_idx] if p_idx < len(balance_row) else ""
                    # 注意：不能把 final_partner 直接写给 balance_row。
                    # final_partner 属于 out_row 的对方科目(final_subject)语义，
                    # 若直接写入 balance_row，会造成“对方科目-账户”方向反转。
                    if "往来单位名" in header_map:
                        balance_row[header_map["往来单位名"]] = "" # 清空名称，让系统自动带出或避免混淆
                    # 账户匹配应按“本行科目编码”，不能按对方科目，否则会发生反向错配
                    out_subject_for_account = current_subject
                    bal_subject_for_account = final_subject
                    out_linked_acc = self._resolve_linked_default_account(
                        out_subject_for_account, None, out_partner_val, allow_fallback=False
                    )
                    bal_linked_acc = self._resolve_linked_default_account(
                        bal_subject_for_account, None, bal_partner_val, allow_fallback=False
                    )
                    if forced_counter_subject and forced_counter_account:
                        if out_subject_for_account == forced_counter_subject:
                            out_linked_acc = forced_counter_account
                        if bal_subject_for_account == forced_counter_subject:
                            bal_linked_acc = forced_counter_account
                    if "默认账户" in header_map:
                        acc_idx = header_map["默认账户"]
                        if acc_idx < len(out_row):
                            out_row[acc_idx] = ""
                        if acc_idx < len(balance_row):
                            balance_row[acc_idx] = ""

                    if "往来单位编码" in header_map:
                        p_idx = header_map["往来单位编码"]
                        def _empty_partner(val):
                            try:
                                if val is None:
                                    return True
                                if isinstance(val, float) and pd.isna(val):
                                    return True
                                s = str(val).strip()
                                return s == "" or s.lower() == "nan"
                            except Exception:
                                return True
                        if p_idx < len(out_row):
                            if forced_counter_subject and forced_counter_account and out_subject_for_account == forced_counter_subject:
                                out_row[p_idx] = forced_counter_account
                            elif _empty_partner(out_row[p_idx]) and out_linked_acc:
                                out_row[p_idx] = out_linked_acc
                        if p_idx < len(balance_row):
                            if forced_counter_subject and forced_counter_account and bal_subject_for_account == forced_counter_subject:
                                balance_row[p_idx] = forced_counter_account
                            elif _empty_partner(balance_row[p_idx]) and bal_linked_acc:
                                balance_row[p_idx] = bal_linked_acc
                    
                    # 部门保持一致（不能空着）
                    if "部门" in header_map:
                        balance_row[header_map["部门"]] = current_department

                    # --- [修正] 外币模式平账逻辑 ---
                    if self.use_foreign_currency_var.get() and "金额" in header_map and "外币金额" in header_map and "汇率" in header_map:
                        try:
                            # 对方分录的金额保持一致，但外币金额应重新计算，防止原行存在脏数据或换算不一致
                            _b_amt = float(str(balance_row[header_map["金额"]]).replace(",", ""))
                            _b_rate = float(str(balance_row[header_map["汇率"]]).replace(",", ""))
                            if _b_rate > 0.0001:
                                balance_row[header_map["外币金额"]] = round(_b_amt / _b_rate, 2)
                        except:
                            pass

                    output_rows.append(balance_row)
            # ---------------------------------------------

        if not output_rows:
            messagebox.showwarning("提示", "原始 Excel 中没有数据行。")
            return

        # ======== 自动排序逻辑 ========
        # 目标：按 [凭证日期] 升序 -> [序号] 升序 排序
        # 这能解决原文件乱序导致导出混乱的问题
        try:
            sort_indices = []
            header_map = {h.name: i for i, h in enumerate(self.template_headers)}
            
            # 1. 凭证日期 (YYYYMMDD)
            date_col_idx = header_map.get("凭证日期")
            if date_col_idx is None:
                date_col_idx = header_map.get("日期")
            
            # 2. 序号
            serial_col_idx = header_map.get("序号")

            if date_col_idx is not None:
                self.log_message("正在对结果进行自动排序 (按日期+序号)...")
                
                def sort_key(row):
                    # 获取日期
                    d_val = row[date_col_idx]
                    # 确保日期是字符串且格式统一，空值排最后
                    if d_val is None: d_val = "99999999"
                    d_str = str(d_val).strip()
                    
                    # 获取序号 (尝试转数字排序，否则按字符串)
                    s_val = 0
                    if serial_col_idx is not None:
                        raw_s = row[serial_col_idx]
                        if raw_s:
                            try:
                                s_val = float(raw_s)
                            except:
                                s_val = str(raw_s)
                        else:
                            s_val = 0
                    
                    return (d_str, s_val)

                # Python 的 sort 是稳定的，这很重要 (保持借贷分录的相对顺序)
                output_rows.sort(key=sort_key)
                
        except Exception as e:
            print(f"排序失败 (非致命错误): {e}")
            self.log_message(f"排序跳过: {e}")
        # ============================

        print(f"调试: 识别到 {len(recognition_info)} 行信息")

        # 显示预处理预览窗口
        # 修改：始终显示预览窗口，以便用户在导出前审核数据 (预处理功能)
        need_preview = True
        
        if need_preview and output_rows:
            try:
                preview_result = self._show_preprocessing_preview(output_rows, recognition_info, unmatched_info)
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                self.log_message(f"预处理预览异常: {e}\n{tb}")
                preview_result = messagebox.askyesno(
                    "预览异常",
                    f"预处理预览窗口异常。\n\n错误: {e}\n\n是否跳过预览并继续导出？"
                )
            if not preview_result:
                return  # 用户取消了转换

        # 执行导出
        self.log_message(f"准备导出数据，共 {len(output_rows)} 行。")
        self._export_to_excel(output_rows, unmatched_info)

    def _build_sales_mapping(self, mapping: Dict[str, str]) -> Dict[str, str]:
        """根据候选列名为销售出库模式补全映射"""
        effective = dict(mapping)
        for target, candidates in SALES_OUTBOUND_MAPPING_CANDIDATES.items():
            if effective.get(target):
                continue
            # 直接命中同名列
            for col in self.input_columns:
                if normalize_header(col) == normalize_header(target):
                    effective[target] = col
                    break
            if effective.get(target):
                continue
            # 候选模糊匹配
            for col in self.input_columns:
                norm_col = normalize_header(col)
                for cand in candidates:
                    norm_cand = normalize_header(cand)
                    if norm_cand and (norm_cand == norm_col or norm_cand in norm_col or norm_col in norm_cand):
                        effective[target] = col
                        break
                if effective.get(target):
                    break
        return effective

    def _convert_sales_outbound(self, mapping):
        """销售出库导入模板转换"""
        self.log_message("进入[销售出库模式]转换逻辑...")
        output_rows = []
        recognition_info = []  # 销售出库暂不使用智能识别预览
        unmatched_info = []  # 未能匹配到基础资料的行
        effective_mapping = self._build_sales_mapping(mapping)

        # 预加载基础资料，便于模糊匹配
        product_rows = []
        partner_rows = []
        if self.base_data_mgr:
            try:
                product_rows = self.base_data_mgr.query("product")
            except Exception:
                product_rows = []
            try:
                partner_rows = self.base_data_mgr.query("business_partner")
            except Exception:
                partner_rows = []

        def _norm(val: str) -> str:
            if val is None:
                return ""
            # 去掉空格/横线/下划线，方便匹配 3WF-600-20L、3WF 600 20L 等
            return re.sub(r"[\s\-_]+", "", str(val)).lower()

        def _code_key(val: str) -> str:
            """生成忽略前导0的编码键，形如 L0021 -> l:21，BC520 -> bc:520"""
            if not val:
                return ""
            s = _norm(val)
            # 前缀字母 + 数字 + 可选尾巴
            m = re.match(r"([a-z]+)(0*)(\d+)(.*)", s)
            if m:
                prefix, _, num, tail = m.groups()
                return f"{prefix}:{num.lstrip('0') or '0'}{tail}"
            # 纯数字前缀（少见）
            m2 = re.match(r"(\d+)(.*)", s)
            if m2:
                num, tail = m2.groups()
                return f":{num.lstrip('0') or '0'}{tail}"
            return s

        def _fuzzy_find(rows, value):
            """在基础资料中模糊匹配 code/name"""
            if not value:
                return None
            n_val = _norm(value)
            if not n_val:
                return None
            key_val = _code_key(value)
            def _parse_match_items(raw):
                if not raw:
                    return []
                if isinstance(raw, list):
                    items = raw
                else:
                    try:
                        items = json.loads(raw)
                    except Exception:
                        items = []
                cleaned = []
                for i in items:
                    if i is None:
                        continue
                    s = str(i).strip()
                    if s:
                        cleaned.append(s)
                return cleaned
            # 1) 编码去0匹配 (如 L0021 vs L021)，优先返回编码长度更长的候选
            key_matches = [r for r in rows if _code_key(r.get("code")) == key_val]
            if key_matches:
                key_matches.sort(key=lambda x: len(str(x.get("code", ""))), reverse=True)
                return key_matches[0]
            # 2) 精确 code
            for r in rows:
                if _norm(r.get("code")) == n_val:
                    return r
            # 3) 精确 name
            for r in rows:
                if _norm(r.get("name")) == n_val:
                    return r
            # 4) 包含匹配
            for r in rows:
                # 4.0) 先尝试映射匹配项
                for alias in _parse_match_items(r.get("match_items")):
                    if n_val == _norm(alias) or n_val in _norm(alias) or _norm(alias) in n_val:
                        return r
                if n_val in _norm(r.get("code")) or _norm(r.get("code")) in n_val:
                    return r
                if n_val in _norm(r.get("name")) or _norm(r.get("name")) in n_val:
                    return r
            return None

        # 默认值
        default_rate = self.default_values.get("汇率", "1")
        try:
            default_rate_val = float(default_rate)
        except Exception:
            default_rate_val = 1.0
        default_currency = self.default_values.get("货币", "") or self.default_values.get("货币编码", "")
        default_department = self.default_values.get("部门", "")
        default_wh = self.default_values.get("发货仓库", "") or self.default_values.get("默认仓库", "")
        default_tran_type = self.default_values.get("交易类型", "11")
        default_staff = self.default_values.get("职员", "")
        default_receiver = self.default_values.get("收货公司", "")
        default_inbound_flag = self.default_values.get("生成生产入库", "N") or "N"

        def _is_empty(val):
            return val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == ""

        current_derived_ctx = {}
        current_comp_cache = {}

        def resolve_mapped_value(row, col):
            if not col:
                return None
            if self._is_composite_option(col):
                comp_name = self._extract_composite_name(col)
                return self._compute_composite_value(
                    comp_name,
                    row,
                    effective_mapping,
                    current_derived_ctx,
                    smart_data={},
                    cache=current_comp_cache
                )
            if col in row:
                return row[col]
            return None

        def pick_value(row, target_name, fallbacks=None):
            """按映射取值"""
            fallbacks = fallbacks or []
            mapped_col = effective_mapping.get(target_name)
            candidates = [mapped_col] + [effective_mapping.get(fb) for fb in fallbacks]
            for col in candidates:
                if not col:
                    continue
                val = resolve_mapped_value(row, col)
                if val is not None and not (isinstance(val, float) and pd.isna(val)):
                    return val
            return None

        order_serial_map = {}
        next_order_serial = 1

        for idx, src_row in self.input_df.iterrows():
            current_derived_ctx = self._derive_debit_credit_context(src_row)
            current_comp_cache = {}
            unmatched_msgs = []
            unmatched_fields = []

            order_no = pick_value(src_row, "序号", fallbacks=["PEDIDO No"])
            if order_no is None:
                order_no = pick_value(src_row, "PEDIDO No")
            order_raw = str(order_no).strip() if order_no not in (None, "") else ""
            if order_raw:
                if order_raw not in order_serial_map:
                    order_serial_map[order_raw] = str(next_order_serial)
                    next_order_serial += 1
                order_serial = order_serial_map[order_raw]
            else:
                order_serial = str(next_order_serial)
                next_order_serial += 1

            customer_code = pick_value(src_row, "往来单位编码")
            customer_name = pick_value(src_row, "往来单位名") or customer_code
            receiver = pick_value(src_row, "收货公司") or customer_name or default_receiver
            product_code = pick_value(src_row, "品目编码")
            product_name = pick_value(src_row, "品目名")
            quantity = pick_value(src_row, "数量")
            unit_price = pick_value(src_row, "单价")
            total_amount = pick_value(src_row, "供应价", fallbacks=["外币金额", "金额1"])
            if total_amount is None:
                total_amount = pick_value(src_row, "外币金额")
            if total_amount is None and quantity is not None and unit_price is not None:
                try:
                    total_amount = float(quantity) * float(unit_price)
                except Exception:
                    pass
            raw_partner_value = customer_code or customer_name
            raw_product_value = product_code or product_name

            # 基础资料匹配（往来单位）
            partner_match = _fuzzy_find(partner_rows, customer_code or customer_name)
            if partner_match:
                customer_code = partner_match.get("code") or customer_code
                customer_name = partner_match.get("name") or customer_name
            elif raw_partner_value:
                unmatched_msgs.append(f"往来单位未匹配: {raw_partner_value}")
                unmatched_fields.extend(["往来单位编码", "往来单位名"])

            # 基础资料匹配（品目）
            product_match = _fuzzy_find(product_rows, product_code or product_name)
            if product_match:
                product_code = product_match.get("code") or product_code
                product_name = product_match.get("name") or product_name
                spec_val = product_match.get("specification") or product_match.get("spec_info")
                mo_tama_val = product_match.get("size_range") or product_match.get("spec_info")
                color_val = product_match.get("color")
                pack_qty_val = product_match.get("pack_qty")
                volume_val = product_match.get("volume")
                weight_val = product_match.get("weight")
            else:
                spec_val = mo_tama_val = color_val = pack_qty_val = volume_val = weight_val = None
                if raw_product_value:
                    unmatched_msgs.append(f"品目未匹配: {raw_product_value}")
                    unmatched_fields.extend(["品目编码", "品目名"])

            # 计算外币金额/本币供应价
            foreign_amount = total_amount
            if foreign_amount is None and quantity is not None and unit_price is not None:
                try:
                    foreign_amount = float(quantity) * float(unit_price)
                except Exception:
                    pass
            local_amount = None
            if foreign_amount is not None:
                try:
                    local_amount = float(foreign_amount) * default_rate_val
                except Exception:
                    local_amount = foreign_amount

            # 体积/重量汇总
            total_volume_val = None
            total_weight_val = None
            if volume_val is not None and quantity not in (None, ""):
                try:
                    total_volume_val = float(volume_val) * float(quantity if quantity is not None else 0)
                except Exception:
                    total_volume_val = volume_val
            if weight_val is not None and quantity not in (None, ""):
                try:
                    total_weight_val = float(weight_val) * float(quantity if quantity is not None else 0)
                except Exception:
                    total_weight_val = weight_val

            summary_text = pick_value(src_row, "摘要") or f"{customer_code or ''} {product_name or product_code or ''}".strip()

            out_row = []
            for header in self.template_headers:
                # 体积/重量类字段强制使用系统计算，不从映射取值
                if header.name in ["体积", "总体积", "重量", "总重量"]:
                    src_col = None
                else:
                    src_col = effective_mapping.get(header.name)
                value = resolve_mapped_value(src_row, src_col)

                # 针对关键字段的兜底赋值
                if header.name == "日期":
                    if value is None:
                        value = pick_value(src_row, "日期")
                elif header.name == "序号":
                    value = order_serial
                elif header.name == "PEDIDO No":
                    value = value if value not in (None, "") else order_no
                elif header.name == "往来单位编码":
                    value = customer_code
                elif header.name == "往来单位名":
                    value = customer_name
                elif header.name == "规格":
                    value = spec_val
                elif header.name == "Mo.TAMA":
                    value = mo_tama_val
                elif header.name == "COLOR":
                    value = color_val
                elif header.name == "装数":
                    value = pack_qty_val
                elif header.name == "CAJA":
                    value = pack_qty_val
                elif header.name == "收货公司":
                    value = value if value not in (None, "") else receiver
                elif header.name == "品目编码":
                    value = product_code
                elif header.name == "品目名":
                    value = product_name
                elif header.name == "数量":
                    value = quantity
                elif header.name == "单价":
                    value = unit_price
                elif header.name == "外币金额":
                    value = foreign_amount
                elif header.name in ["供应价", "金额1"]:
                    value = local_amount
                elif header.name == "货币":
                    value = value if not _is_empty(value) else default_currency
                elif header.name == "汇率":
                    value = value if not _is_empty(value) else default_rate_val
                elif header.name == "交易类型":
                    value = value if not _is_empty(value) else default_tran_type
                elif header.name == "部门":
                    value = value if not _is_empty(value) else default_department
                elif header.name == "发货仓库":
                    value = value if not _is_empty(value) else default_wh
                elif header.name == "职员":
                    value = value if not _is_empty(value) else default_staff
                elif header.name == "摘要":
                    value = summary_text
                elif header.name == "生成生产入库":
                    value = value if not _is_empty(value) else default_inbound_flag
                elif header.name == "体积":
                    value = volume_val
                elif header.name == "总体积":
                    value = total_volume_val
                elif header.name == "重量":
                    value = weight_val
                elif header.name == "总重量":
                    value = total_weight_val

                converted = convert_value(header.name, value, self.field_formats.get(header.name))
                out_row.append(converted)

            output_rows.append(out_row)
            current_output_idx = len(output_rows)

            if unmatched_msgs:
                unique_fields = list(dict.fromkeys(unmatched_fields))
                unmatched_info.append({
                    "row_num": idx + 2,  # Excel 行号（包含表头）
                    "output_index": current_output_idx,
                    "fields": unique_fields,
                    "messages": unmatched_msgs
                })

        return output_rows, recognition_info, unmatched_info

    def _show_preprocessing_preview_compat(self, output_rows, recognition_info, unmatched_info=None):
        """预处理预览兼容模式（用于高风险 Tk 运行时）"""
        unmatched_info = unmatched_info or []
        preview_window = tk.Toplevel(self.root)
        preview_window.title("预处理预览 - 兼容模式")
        preview_window.geometry("1120x700")

        main_frame = ttk.Frame(preview_window, padding=10)
        main_frame.pack(fill="both", expand=True)

        ttk.Label(
            main_frame,
            text=f"智能识别统计：共处理 {len(output_rows)} 行数据，识别到 {len(recognition_info)} 行包含可识别信息",
            font=("Arial", 10, "bold")
        ).pack(anchor="w")
        ttk.Label(
            main_frame,
            text="当前运行时已启用兼容预览：保留格式化表格预览，关闭复杂编辑功能以降低卡死风险",
            foreground="#b35b00"
        ).pack(anchor="w", pady=(2, 4))
        if unmatched_info:
            ttk.Label(
                main_frame,
                text=f"未匹配提醒：{len(unmatched_info)} 行存在基础资料未命中",
                foreground="red"
            ).pack(anchor="w", pady=(0, 6))

        if self.template_headers:
            preview_headers = [h.name for h in self.template_headers]
        else:
            max_col_count = max((len(r) for r in output_rows), default=0)
            preview_headers = [f"列{i + 1}" for i in range(max_col_count)]

        source_headers = list(preview_headers)
        source_rows = output_rows
        display_limit = 500
        use_export_preview_var = tk.BooleanVar(value=True)
        preview_status_var = tk.StringVar(value="")
        preview_tree = None
        tree_scroll_y = None
        tree_scroll_x = None

        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill="both", expand=True)

        def _resolve_display_dataset():
            if use_export_preview_var.get():
                try:
                    mapped_headers, mapped_rows, mapped = apply_export_format(
                        "main_export",
                        source_headers,
                        source_rows,
                        base_data_mgr=self.base_data_mgr
                    )
                    if mapped:
                        return mapped_headers, mapped_rows, True
                except Exception as e:
                    self.log_message(f"兼容预览应用导出格式失败，已回退原始预览: {e}")
            return source_headers, source_rows, False

        unmatched_output_indices = {
            item.get("output_index") for item in unmatched_info if item.get("output_index")
        }

        def _refresh_preview_table():
            nonlocal preview_tree, tree_scroll_y, tree_scroll_x
            headers, rows, mapped_applied = _resolve_display_dataset()
            headers = headers or source_headers
            if not headers:
                max_col_count = max((len(r) for r in rows), default=0)
                headers = [f"列{i + 1}" for i in range(max_col_count)]

            if preview_tree is not None:
                try:
                    preview_tree.destroy()
                except Exception:
                    pass
            if tree_scroll_y is not None:
                try:
                    tree_scroll_y.destroy()
                except Exception:
                    pass
            if tree_scroll_x is not None:
                try:
                    tree_scroll_x.destroy()
                except Exception:
                    pass

            tree_scroll_y = ttk.Scrollbar(table_frame, orient="vertical")
            tree_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal")
            preview_tree = ttk.Treeview(
                table_frame,
                columns=headers,
                show="headings",
                yscrollcommand=tree_scroll_y.set,
                xscrollcommand=tree_scroll_x.set,
                selectmode="browse"
            )
            tree_scroll_y.config(command=preview_tree.yview)
            tree_scroll_x.config(command=preview_tree.xview)
            preview_tree.tag_configure("unmatched", background="#fff1f0")

            for col in headers:
                preview_tree.heading(col, text=col)
                preview_tree.column(col, width=120, minwidth=80, stretch=False)

            rows_to_show = rows[:display_limit]
            for row_idx, row in enumerate(rows_to_show, start=1):
                tags = ("unmatched",) if row_idx in unmatched_output_indices else ()
                values = ["" if v is None else str(v) for v in list(row)]
                if len(values) < len(headers):
                    values.extend([""] * (len(headers) - len(values)))
                preview_tree.insert("", "end", iid=str(row_idx - 1), values=values[:len(headers)], tags=tags)

            preview_tree.pack(side="left", fill="both", expand=True)
            tree_scroll_y.pack(side="right", fill="y")
            tree_scroll_x.pack(side="bottom", fill="x")

            mode_text = "已应用导出格式预览" if mapped_applied else "原始结果预览"
            preview_status_var.set(f"{mode_text} | 显示 {min(len(rows), display_limit)}/{len(rows)} 行，{len(headers)} 列")

        def _open_format_editor_and_refresh():
            self._open_export_format_editor()
            self._refresh_export_format_options()
            _refresh_preview_table()

        toolbar = ttk.Frame(main_frame)
        toolbar.pack(fill="x", pady=(6, 4))
        ttk.Checkbutton(
            toolbar,
            text="按当前导出格式预览",
            variable=use_export_preview_var,
            command=_refresh_preview_table
        ).pack(side="left")
        ttk.Button(toolbar, text="导出格式设置...", command=_open_format_editor_and_refresh).pack(side="left", padx=(6, 0))
        ttk.Button(toolbar, text="刷新预览", command=_refresh_preview_table).pack(side="left", padx=(6, 0))
        ttk.Label(toolbar, textvariable=preview_status_var, foreground="gray").pack(side="right")

        _refresh_preview_table()

        hint_text = f"注: 兼容模式仅显示前 {display_limit} 行（总计 {len(output_rows)} 行），仅用于格式核对"
        ttk.Label(main_frame, text=hint_text, foreground="gray").pack(anchor="w", pady=(5, 0))

        detail_frame = ttk.Frame(main_frame)
        detail_frame.pack(fill="x", pady=(6, 0))
        detail_lines = []
        if recognition_info:
            detail_lines.append(f"识别详情: 共 {len(recognition_info)} 行。")
        if unmatched_info:
            sample = []
            for warn in unmatched_info[:3]:
                row_num = warn.get("row_num", "")
                msgs = warn.get("messages", []) or []
                if msgs:
                    sample.append(f"第{row_num}行: {msgs[0]}")
            if sample:
                detail_lines.append("未匹配样例: " + " | ".join(sample))
            else:
                detail_lines.append(f"未匹配提醒: 共 {len(unmatched_info)} 行。")
        if detail_lines:
            ttk.Label(detail_frame, text=" ".join(detail_lines), foreground="#444").pack(anchor="w")

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))

        result = {"confirmed": False}

        def export_preview():
            path = filedialog.asksaveasfilename(
                title="导出预览为 Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            if not path:
                return
            try:
                export_headers, export_rows, _mapped_applied = _resolve_display_dataset()
                df = pd.DataFrame(export_rows, columns=export_headers)
                df.to_excel(path, index=False)
                messagebox.showinfo("完成", f"已导出预览数据：\n{path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败：{e}")

        def confirm():
            result["confirmed"] = True
            preview_window.destroy()

        def cancel():
            result["confirmed"] = False
            preview_window.destroy()

        ttk.Button(button_frame, text="导出预览为Excel", command=export_preview, width=18).pack(side="left")
        ttk.Button(button_frame, text="确认并导出", command=confirm, width=15).pack(side="right", padx=5)
        ttk.Button(button_frame, text="取消", command=cancel, width=15).pack(side="right")

        preview_window.protocol("WM_DELETE_WINDOW", cancel)
        preview_window.transient(self.root)
        try:
            preview_window.lift()
            preview_window.focus_force()
        except Exception:
            pass
        self.root.wait_window(preview_window)
        return result["confirmed"]

    def _show_preprocessing_preview(self, output_rows, recognition_info, unmatched_info=None):
        """显示预处理预览窗口"""
        if self._is_tk_crash_prone_runtime():
            self.log_message("检测到高风险 Tk 运行时，预处理预览自动切换为兼容模式。")
            return self._show_preprocessing_preview_compat(output_rows, recognition_info, unmatched_info)

        unmatched_info = unmatched_info or []
        preview_window = tk.Toplevel(self.root)
        preview_window.title("预处理预览 - 智能识别结果")
        preview_window.geometry("1200x700")
        base_rows = [list(r) for r in output_rows]
        preview_cell_overrides = {}
        header_items = [{"name": h.name, "idx": i, "default": ""} for i, h in enumerate(self.template_headers)]
        original_signature = [(item["idx"], item["name"]) for item in header_items]

        # 创建主框架
        main_frame = ttk.Frame(preview_window, padding="10")
        main_frame.pack(fill="both", expand=True)

        # 顶部信息
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(
            info_frame,
            text=f"智能识别统计：共处理 {len(output_rows)} 行数据，识别到 {len(recognition_info)} 行包含可识别信息",
            font=("Arial", 10, "bold")
        ).pack(anchor="w")
        if unmatched_info:
            ttk.Label(
                info_frame,
                text=f"未匹配提醒：{len(unmatched_info)} 行存在基础资料未命中，已在下方标红，请手工核对",
                foreground="red"
            ).pack(anchor="w", pady=(2, 0))

        # 创建Notebook分页
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill="both", expand=True)

        # 第一页：识别详情
        detail_frame = ttk.Frame(notebook)
        notebook.add(detail_frame, text="识别详情")

        # 创建滚动区域
        detail_canvas = tk.Canvas(detail_frame)
        detail_scrollbar_y = ttk.Scrollbar(detail_frame, orient="vertical", command=detail_canvas.yview)
        detail_scrollbar_x = ttk.Scrollbar(detail_frame, orient="horizontal", command=detail_canvas.xview)
        detail_scrollable = ttk.Frame(detail_canvas)

        detail_scrollable.bind(
            "<Configure>",
            lambda e: detail_canvas.configure(scrollregion=detail_canvas.bbox("all"))
        )

        detail_canvas.create_window((0, 0), window=detail_scrollable, anchor="nw")
        detail_canvas.configure(yscrollcommand=detail_scrollbar_y.set, xscrollcommand=detail_scrollbar_x.set)

        detail_canvas.pack(side="left", fill="both", expand=True)
        detail_scrollbar_y.pack(side="right", fill="y")
        detail_scrollbar_x.pack(side="bottom", fill="x")

        # 显示识别详情
        for info in recognition_info:
            item_frame = ttk.LabelFrame(
                detail_scrollable,
                text=f"第 {info['row_num']} 行",
                padding="10"
            )
            item_frame.pack(fill="x", padx=5, pady=5)

            # 摘要
            summary_frame = ttk.Frame(item_frame)
            summary_frame.pack(fill="x", pady=(0, 5))
            ttk.Label(summary_frame, text="摘要:", font=("Arial", 9, "bold")).pack(side="left")
            ttk.Label(summary_frame, text=info['summary'], wraplength=900).pack(side="left", padx=(5, 0))

            # 识别字段
            recognized_frame = ttk.Frame(item_frame)
            recognized_frame.pack(fill="x")
            ttk.Label(recognized_frame, text="识别结果:", font=("Arial", 9, "bold")).pack(anchor="w")

            rec_text = ttk.Frame(recognized_frame)
            rec_text.pack(fill="x", padx=(20, 0))

            for key, value in info['recognized'].items():
                if value:  # 只显示非空值
                    ttk.Label(
                        rec_text,
                        text=f"• {key}: {value}",
                        foreground="blue"
                    ).pack(anchor="w")

        if unmatched_info:
            ttk.Separator(detail_scrollable, orient="horizontal").pack(fill="x", pady=5)
            ttk.Label(
                detail_scrollable,
                text="未匹配提醒（基础资料未命中/需人工确认）",
                foreground="red",
                font=("Arial", 10, "bold")
            ).pack(anchor="w", pady=(0, 5))

            for warn in unmatched_info:
                warn_frame = ttk.LabelFrame(
                    detail_scrollable,
                    text=f"第 {warn.get('row_num')} 行",
                    padding="10"
                )
                warn_frame.pack(fill="x", padx=5, pady=3)
                for msg in warn.get("messages", []):
                    ttk.Label(
                        warn_frame,
                        text=f"! {msg}",
                        foreground="red"
                    ).pack(anchor="w")

        # 第二页：转换结果预览
        preview_frame = ttk.Frame(notebook)
        notebook.add(preview_frame, text="转换结果预览")

        unmatched_output_indices = {
            item.get("output_index") for item in unmatched_info if item.get("output_index")
        }

        if unmatched_output_indices:
            ttk.Label(
                preview_frame,
                text="标红行包含未匹配基础资料，请核对后再导出",
                foreground="red"
            ).pack(anchor="w", padx=5)

        def _get_preview_rows():
            indices = [item["idx"] for item in header_items]
            rows = []
            for row_idx, row in enumerate(base_rows):
                row_vals = []
                for i, item in zip(indices, header_items):
                    if i is None:
                        override_key = (row_idx, item["name"])
                        if override_key in preview_cell_overrides:
                            row_vals.append(preview_cell_overrides[override_key])
                        else:
                            row_vals.append(item.get("default", ""))
                    else:
                        row_vals.append(row[i] if i < len(row) else "")
                rows.append(row_vals)
            return rows

        preview_tree = None
        tree_scroll_y = None
        tree_scroll_x = None
        show_all_var = tk.BooleanVar(value=False)
        preview_menu = None
        preview_col_widths = {}
        rematch_favorites_only_var = tk.BooleanVar(value=False)
        rematch_skip_cache_var = tk.BooleanVar(value=False)
        rematch_history = {}
        FIELD_TABLE_MAP = {
            "部门": "department",
            "科目编码": "account_subject",
            "对方科目": "account_subject",
            "默认账户": "bank_account",
            "往来单位编码": "business_partner",
            "发货仓库": "warehouse",
            "默认仓库": "warehouse",
            "货币": "currency",
        }
        def _get_header_index(names):
            return next((i for i, item in enumerate(header_items) if item["name"] in names), None)

        def _get_tree_col_index(names):
            if preview_tree is None:
                return None
            cols = list(preview_tree["columns"])
            for n in names:
                if n in cols:
                    return cols.index(n)
            return None

        def _get_cell_value(row_idx: int, col_index: int):
            if row_idx < 0 or row_idx >= len(base_rows):
                return ""
            item = header_items[col_index]
            src_idx = item.get("idx")
            if src_idx is None:
                return preview_cell_overrides.get((row_idx, item["name"]), item.get("default", ""))
            row = base_rows[row_idx]
            return row[src_idx] if src_idx < len(row) else ""

        def _set_cell_value(row_idx: int, col_index: int, value):
            item = header_items[col_index]
            src_idx = item.get("idx")
            if src_idx is None:
                preview_cell_overrides[(row_idx, item["name"])] = value
                return
            while src_idx >= len(base_rows[row_idx]):
                base_rows[row_idx].append("")
            base_rows[row_idx][src_idx] = value

        unmatched_action_cache = {}
        restore_label_map = {
            "account_subject": "科目编码",
            "business_partner": "往来单位",
            "product": "品目信息",
            "currency": "币种",
            "department": "部门",
            "warehouse": "仓库",
            "bank_account": "账户",
        }
        try:
            if self.base_data_mgr:
                for cat in self.base_data_mgr.list_custom_categories():
                    name_key = cat.get("name")
                    display = cat.get("display_name") or name_key
                    if name_key:
                        restore_label_map[f"custom:{name_key}"] = display
        except Exception:
            pass

        def _prompt_unmatched_action(label, code_value, name_value):
            result = {"choice": "skip"}
            dialog = tk.Toplevel(preview_window)
            dialog.title("未匹配提示")
            dialog.transient(preview_window)
            dialog.grab_set()
            msg = f"{label} 未在基础数据中找到：\n\n编码: {code_value}"
            if name_value:
                msg += f"\n名称: {name_value}"
            ttk.Label(dialog, text=msg, justify="left").pack(padx=12, pady=(12, 6))
            ttk.Label(dialog, text="请选择处理方式：", foreground="gray").pack(padx=12, pady=(0, 6), anchor="w")

            btns = ttk.Frame(dialog)
            btns.pack(padx=12, pady=12, fill="x")

            def _choose(val):
                result["choice"] = val
                dialog.destroy()

            ttk.Button(btns, text="不加入", command=lambda: _choose("skip"), width=10).pack(side="right", padx=6)
            ttk.Button(btns, text="继续模糊匹配", command=lambda: _choose("fuzzy"), width=14).pack(side="right", padx=6)
            ttk.Button(btns, text="加入基础数据", command=lambda: _choose("add"), width=12).pack(side="right")

            preview_window.wait_window(dialog)
            return result["choice"]

        def _restore_codes_for_category(category_key: str, code_col=None, name_col=None):
            if not self.base_data_mgr:
                messagebox.showwarning("提示", "基础数据管理器未初始化。")
                return
            table_name = category_key
            label = restore_label_map.get(table_name, str(table_name))
            code_col_index = _get_header_index([code_col]) if code_col else None
            name_col_index = _get_header_index([name_col]) if name_col else None
            if code_col_index is None and name_col_index is None:
                messagebox.showwarning("提示", f"未找到{label}相关列。")
                return
            updated = 0
            for row_idx in range(len(base_rows)):
                current_val = _get_cell_value(row_idx, code_col_index) if code_col_index is not None else ""
                match = self.base_data_mgr.find_best_match(table_name, safe_str(current_val))
                if not match and name_col_index is not None:
                    ref_val = _get_cell_value(row_idx, name_col_index)
                    match = self.base_data_mgr.find_best_match(table_name, safe_str(ref_val))
                if match and str(match) != str(current_val):
                    if code_col_index is not None:
                        _set_cell_value(row_idx, code_col_index, match)
                    updated += 1
                    continue
                if match or not safe_str(current_val):
                    continue

                code_str = safe_str(current_val)
                name_str = safe_str(_get_cell_value(row_idx, name_col_index)) if name_col_index is not None else ""
                cache_key = (table_name, code_str)
                if cache_key in unmatched_action_cache:
                    action = unmatched_action_cache[cache_key]
                else:
                    action = _prompt_unmatched_action(label, code_str, name_str)
                    unmatched_action_cache[cache_key] = action

                if action == "fuzzy":
                    match = self.base_data_mgr.find_best_match(table_name, code_str, min_score=0.6)
                    if not match and name_str:
                        match = self.base_data_mgr.find_best_match(table_name, name_str, min_score=0.6)
                    if match and str(match) != str(current_val):
                        _set_cell_value(row_idx, code_col_index, match)
                        updated += 1
                    else:
                        follow = _prompt_unmatched_action(label, code_str, name_str)
                        if follow == "add":
                            payload = {"code": code_str, "name": name_str or code_str}
                            try:
                                self.base_data_mgr.add_record(table_name, payload)
                                if hasattr(self.base_data_mgr, "clear_lookup_cache"):
                                    self.base_data_mgr.clear_lookup_cache()
                            except Exception:
                                messagebox.showerror("错误", f"加入基础数据失败: {code_str}")
                elif action == "add":
                    payload = {"code": code_str, "name": name_str or code_str}
                    try:
                        self.base_data_mgr.add_record(table_name, payload)
                        if hasattr(self.base_data_mgr, "clear_lookup_cache"):
                            self.base_data_mgr.clear_lookup_cache()
                    except Exception:
                        messagebox.showerror("错误", f"加入基础数据失败: {code_str}")
            if hasattr(self.base_data_mgr, "clear_lookup_cache"):
                self.base_data_mgr.clear_lookup_cache()
            _refresh_preview_tree()
            messagebox.showinfo("完成", f"{label} 智能还原完成，更新 {updated} 行。")

        def _best_alternative_account(summary_text: str, exclude_code: str) -> Optional[str]:
            if not summary_text:
                return None
            summary_lower = str(summary_text).lower()
            candidates = {}
            favorite_codes = set()
            if rematch_favorites_only_var.get() and self.base_data_mgr:
                try:
                    raw_favs = self.base_data_mgr.get_config("base_data_favorites_account_subject", "[]")
                    parsed = json.loads(raw_favs) if raw_favs else []
                    if isinstance(parsed, list):
                        favorite_codes = {str(v).strip() for v in parsed if str(v).strip()}
                except Exception:
                    favorite_codes = set()
            def _collect_candidates(fav_only: bool):
                local_candidates = {}
                account_map = {}
                if self.summary_recognizer and getattr(self.summary_recognizer, "accounts", None):
                    account_map = self.summary_recognizer.accounts
                if account_map:
                    for key, code in account_map.items():
                        code_str = str(code or "").strip()
                        if not code_str or code_str == str(exclude_code):
                            continue
                        if not code_str.isdigit() or len(code_str) < 4:
                            continue
                        if fav_only and favorite_codes and code_str not in favorite_codes:
                            continue
                        kw = str(key or "").strip()
                        if not kw:
                            continue
                        kw_lower = kw.lower()
                        if kw_lower in summary_lower:
                            score = 2.0 + (len(kw_lower) / 100.0)
                        else:
                            score = difflib.SequenceMatcher(None, summary_lower, kw_lower).ratio()
                        if score > local_candidates.get(code_str, 0.0):
                            local_candidates[code_str] = score
                elif self.base_data_mgr:
                    accounts = self.base_data_mgr.query("account_subject")
                    for row in accounts:
                        code = str(row.get("code", "") or "").strip()
                        if not code or code == exclude_code:
                            continue
                        if not code.isdigit() or len(code) < 4:
                            continue
                        if fav_only and favorite_codes and code not in favorite_codes:
                            continue
                        keywords = []
                        code_name = str(row.get("code_name", "") or "").strip()
                        name = str(row.get("name", "") or "").strip()
                        if code_name:
                            keywords.append(code_name)
                        if name and name not in keywords:
                            keywords.append(name)
                        raw_items = str(row.get("match_items", "") or "")
                        if raw_items:
                            for part in re.split(r"[,\n]+", raw_items):
                                part = part.strip()
                                if part and part not in keywords:
                                    keywords.append(part)
                        best_score = 0.0
                        for kw in keywords:
                            kw_lower = kw.lower()
                            if not kw_lower:
                                continue
                            if kw_lower in summary_lower:
                                score = 2.0 + (len(kw_lower) / 100.0)
                            else:
                                score = difflib.SequenceMatcher(None, summary_lower, kw_lower).ratio()
                            if score > best_score:
                                best_score = score
                        local_candidates[code] = best_score
                return local_candidates

            candidates = _collect_candidates(fav_only=bool(favorite_codes))
            if not candidates and favorite_codes:
                candidates = _collect_candidates(fav_only=False)
            if not candidates:
                return exclude_code if exclude_code else None
            best_code, _best_score = max(candidates.items(), key=lambda x: x[1])
            return best_code

        def _open_cell_editor(event):
            if preview_tree is None:
                return
            region = preview_tree.identify_region(event.x, event.y)
            if region != "cell":
                return
            row_id = preview_tree.identify_row(event.y)
            col_id = preview_tree.identify_column(event.x)
            if not row_id or not col_id:
                return
            try:
                row_idx = int(row_id)
            except ValueError:
                return
            col_index = int(col_id.replace("#", "")) - 1
            columns = list(preview_tree["columns"])
            if col_index < 0 or col_index >= len(columns):
                return
            col_name = columns[col_index]

            bbox = preview_tree.bbox(row_id, column=col_name)
            if not bbox:
                return
            x, y, width, height = bbox
            current_values = list(preview_tree.item(row_id, "values"))
            current_val = current_values[col_index] if col_index < len(current_values) else ""

            entry = ttk.Entry(preview_tree)
            entry.place(x=x, y=y, width=width, height=height)
            entry.insert(0, current_val)
            entry.select_range(0, tk.END)
            entry.focus()

            def _save_edit(event=None):
                new_val = entry.get()
                entry.destroy()
                if col_index >= len(current_values):
                    current_values.extend([""] * (col_index + 1 - len(current_values)))
                if new_val == str(current_values[col_index]):
                    return
                current_values[col_index] = new_val
                preview_tree.item(row_id, values=current_values)

                header_item = header_items[col_index]
                src_idx = header_item.get("idx")
                if src_idx is None:
                    preview_cell_overrides[(row_idx, col_name)] = new_val
                else:
                    while src_idx >= len(base_rows[row_idx]):
                        base_rows[row_idx].append("")
                    base_rows[row_idx][src_idx] = new_val

            def _cancel_edit(event=None):
                entry.destroy()

            entry.bind("<Return>", _save_edit)
            entry.bind("<Escape>", _cancel_edit)
            entry.bind("<FocusOut>", _save_edit)

        def _refresh_preview_tree():
            nonlocal preview_tree, tree_scroll_y, tree_scroll_x
            if preview_tree is not None:
                preview_tree.destroy()
            if tree_scroll_y is not None:
                tree_scroll_y.destroy()
            if tree_scroll_x is not None:
                tree_scroll_x.destroy()

            # 创建滚动条
            tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
            tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")

            # 创建Treeview
            preview_tree = ttk.Treeview(
                tree_frame,
                columns=[item["name"] for item in header_items],
                show="headings",
                selectmode="extended",
                yscrollcommand=tree_scroll_y.set,
                xscrollcommand=tree_scroll_x.set
            )
            preview_tree._smart_restore_delegate = _restore_codes_for_category

            tree_scroll_y.config(command=preview_tree.yview)
            tree_scroll_x.config(command=preview_tree.xview)
            preview_tree.tag_configure("unmatched", background="#fff1f0")

            # 设置列标题
            for item in header_items:
                name = item["name"]
                preview_tree.heading(name, text=name)
                width = preview_col_widths.get(name, 120)
                preview_tree.column(name, width=width, minwidth=80, stretch=False)

            # 插入数据（只显示前20行）
            rows = _get_preview_rows()
            if not show_all_var.get():
                rows = rows[:20]
            for display_idx, row in enumerate(rows, 1):
                row_idx = display_idx - 1
                tags = [f"row{display_idx}"]
                if display_idx in unmatched_output_indices:
                    tags.append("unmatched")
                preview_tree.insert("", "end", iid=str(row_idx), values=row, tags=tuple(tags))

            preview_tree.pack(side="left", fill="both", expand=True)
            tree_scroll_y.pack(side="right", fill="y")
            tree_scroll_x.pack(side="bottom", fill="x")
            attach_treeview_tools(preview_tree, allow_reorder=False)
            preview_tree.bind("<Double-1>", _open_cell_editor, add="+")
            if preview_menu is not None:
                preview_tree.bind("<Button-3>", _show_preview_menu, add="+")
            preview_tree.bind(
                "<ButtonRelease-1>",
                lambda _e: preview_col_widths.update({c: preview_tree.column(c, "width") for c in preview_tree["columns"]}),
                add="+"
            )

            if len(output_rows) > 20:
                if show_all_var.get():
                    preview_hint_label.config(text=f"已显示全部 {len(output_rows)} 行")
                else:
                    preview_hint_label.config(text=f"注：仅显示前20行，共 {len(output_rows)} 行")
                preview_hint_label.pack(pady=5)
            else:
                preview_hint_label.pack_forget()

        def _open_header_editor():
            editor = tk.Toplevel(preview_window)
            editor.title("编辑表头与顺序")
            editor.geometry("520x420")
            editor.transient(preview_window)

            working_items = [dict(item) for item in header_items]

            list_frame = ttk.Frame(editor, padding=10)
            list_frame.pack(fill="both", expand=True)

            listbox = tk.Listbox(list_frame, height=12)
            for item in working_items:
                listbox.insert(tk.END, item["name"])
            listbox.pack(side="left", fill="both", expand=True)

            btn_frame = ttk.Frame(list_frame)
            btn_frame.pack(side="right", fill="y", padx=8)

            def _move_selected(delta):
                sel = listbox.curselection()
                if not sel:
                    return
                i = sel[0]
                j = i + delta
                if j < 0 or j >= len(working_items):
                    return
                working_items[i], working_items[j] = working_items[j], working_items[i]
                listbox.delete(0, tk.END)
                for item in working_items:
                    listbox.insert(tk.END, item["name"])
                listbox.selection_set(j)

            ttk.Button(btn_frame, text="上移", command=lambda: _move_selected(-1), width=8).pack(pady=2)
            ttk.Button(btn_frame, text="下移", command=lambda: _move_selected(1), width=8).pack(pady=2)
            ttk.Button(btn_frame, text="删除", command=lambda: _delete_selected(), width=8).pack(pady=2)

            rename_frame = ttk.Frame(editor, padding=(10, 0))
            rename_frame.pack(fill="x")
            ttk.Label(rename_frame, text="新表头名:").pack(side="left")
            rename_var = tk.StringVar()
            rename_entry = ttk.Entry(rename_frame, textvariable=rename_var, width=30)
            rename_entry.pack(side="left", padx=5)

            def _rename_selected():
                sel = listbox.curselection()
                if not sel:
                    return
                new_name = rename_var.get().strip()
                if not new_name:
                    return
                i = sel[0]
                working_items[i]["name"] = new_name
                listbox.delete(i)
                listbox.insert(i, new_name)
                listbox.selection_set(i)

            ttk.Button(rename_frame, text="重命名", command=_rename_selected, width=8).pack(side="left")

            default_frame = ttk.Frame(editor, padding=(10, 0))
            default_frame.pack(fill="x")
            ttk.Label(default_frame, text="默认值:").pack(side="left")
            default_var = tk.StringVar()
            default_entry = ttk.Entry(default_frame, textvariable=default_var, width=30)
            default_entry.pack(side="left", padx=5)

            def _set_default_value():
                sel = listbox.curselection()
                if not sel:
                    return
                i = sel[0]
                working_items[i]["default"] = default_var.get()

            ttk.Button(default_frame, text="设置默认值", command=_set_default_value, width=10).pack(side="left")

            add_frame = ttk.Frame(editor, padding=(10, 0))
            add_frame.pack(fill="x")
            ttk.Label(add_frame, text="新增列名:").pack(side="left")
            add_name_var = tk.StringVar()
            ttk.Entry(add_frame, textvariable=add_name_var, width=25).pack(side="left", padx=5)

            def _add_column():
                name = add_name_var.get().strip()
                if not name:
                    return
                working_items.append({"name": name, "idx": None, "default": ""})
                listbox.insert(tk.END, name)
                listbox.selection_clear(0, tk.END)
                listbox.selection_set(tk.END)

            ttk.Button(add_frame, text="新增", command=_add_column, width=8).pack(side="left")

            def _delete_selected():
                sel = listbox.curselection()
                if not sel:
                    return
                i = sel[0]
                del working_items[i]
                listbox.delete(i)

            action_frame = ttk.Frame(editor, padding=10)
            action_frame.pack(fill="x")

            def _apply_header_changes():
                header_items[:] = working_items
                _refresh_preview_tree()
                editor.destroy()

            ttk.Button(action_frame, text="应用并关闭", command=_apply_header_changes).pack(side="right")
            ttk.Button(action_frame, text="取消", command=editor.destroy).pack(side="right", padx=6)

            drag_state = {"index": None}

            def _drag_start(event):
                drag_state["index"] = listbox.nearest(event.y)

            def _drag_motion(event):
                i = listbox.nearest(event.y)
                j = drag_state.get("index")
                if j is None or i == j or i < 0:
                    return
                working_items.insert(i, working_items.pop(j))
                listbox.delete(0, tk.END)
                for item in working_items:
                    listbox.insert(tk.END, item["name"])
                listbox.selection_set(i)
                drag_state["index"] = i

            listbox.bind("<ButtonPress-1>", _drag_start)
            listbox.bind("<B1-Motion>", _drag_motion)

        # 创建树形视图显示前20行数据
        tree_frame = ttk.Frame(preview_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        toolbar = ttk.Frame(preview_frame)
        toolbar.pack(fill="x", padx=5, pady=(0, 5))
        ttk.Button(toolbar, text="编辑表头/顺序", command=_open_header_editor).pack(side="left")
        ttk.Button(toolbar, text="全选", command=lambda: preview_tree.selection_set(preview_tree.get_children()) if preview_tree else None).pack(side="left", padx=6)
        ttk.Checkbutton(toolbar, text="显示全部", variable=show_all_var, command=_refresh_preview_tree).pack(side="left", padx=6)
        ttk.Label(toolbar, text="双击单元格编辑，回车保存，Esc取消", foreground="gray").pack(side="left", padx=(10, 0))
        ttk.Label(toolbar, text="表头方案:").pack(side="left", padx=(12, 4))
        scheme_var = tk.StringVar()
        scheme_combo = ttk.Combobox(
            toolbar,
            textvariable=scheme_var,
            values=sorted(HEADER_SCHEMES.keys()),
            state="readonly",
            width=20
        )
        scheme_combo.pack(side="left")

        def _apply_header_scheme():
            name = scheme_var.get().strip()
            if not name or name not in HEADER_SCHEMES:
                return
            scheme = HEADER_SCHEMES.get(name) or []
            current_by_name = {item["name"]: item for item in header_items}
            used = set()
            new_items = []
            for s in scheme:
                s_name = s.get("name", "")
                if not s_name:
                    continue
                if s.get("idx") is None and s_name not in current_by_name:
                    new_items.append({"name": s_name, "idx": None, "default": s.get("default", "")})
                    continue
                if s_name in current_by_name:
                    item = current_by_name[s_name]
                    new_items.append({
                        "name": s_name,
                        "idx": item.get("idx"),
                        "default": s.get("default", item.get("default", "")),
                    })
                    used.add(s_name)
            for item in header_items:
                if item["name"] not in used:
                    new_items.append(item)
            header_items[:] = new_items
            _refresh_preview_tree()

        def _save_header_scheme():
            name = simpledialog.askstring("保存表头方案", "请输入方案名称:")
            if not name:
                return
            name = name.strip()
            if not name:
                return
            HEADER_SCHEMES[name] = [
                {"name": item["name"], "idx": item.get("idx"), "default": item.get("default", "")}
                for item in header_items
            ]
            save_header_schemes(HEADER_SCHEMES)
            scheme_combo["values"] = sorted(HEADER_SCHEMES.keys())
            scheme_var.set(name)
            if hasattr(self, "header_scheme_combo"):
                self.header_scheme_combo["values"] = sorted(HEADER_SCHEMES.keys())
            messagebox.showinfo("成功", f"已保存表头方案：{name}")

        ttk.Button(toolbar, text="应用", command=_apply_header_scheme).pack(side="left", padx=4)
        ttk.Button(toolbar, text="保存为方案", command=_save_header_scheme).pack(side="left", padx=4)

        preview_hint_label = ttk.Label(preview_frame, foreground="gray")
        _refresh_preview_tree()

        def _get_summary_text(row_idx: int) -> str:
            summary_col_index = _get_header_index(["摘要", "摘要名"])
            if summary_col_index is None:
                return ""
            item = header_items[summary_col_index]
            src_idx = item.get("idx")
            if src_idx is None:
                return ""
            if row_idx < 0 or row_idx >= len(base_rows):
                return ""
            row = base_rows[row_idx]
            return str(row[src_idx]) if src_idx < len(row) else ""

        def _get_account_code(row_idx: int) -> str:
            account_col_index = _get_header_index(["科目编码"])
            if account_col_index is None:
                return ""
            item = header_items[account_col_index]
            src_idx = item.get("idx")
            if src_idx is None:
                return ""
            row = base_rows[row_idx]
            return str(row[src_idx]) if src_idx < len(row) else ""

        def _set_account_code(row_idx: int, new_code: str):
            account_col_index = _get_header_index(["科目编码"])
            if account_col_index is None:
                return
            item = header_items[account_col_index]
            src_idx = item.get("idx")
            if src_idx is None:
                return
            while src_idx >= len(base_rows[row_idx]):
                base_rows[row_idx].append("")
            base_rows[row_idx][src_idx] = new_code

        def _ai_rematch_selected():
            if not self.summary_recognizer:
                messagebox.showwarning("提示", "智能识别器未初始化，无法 AI 匹配。")
                return
            # 强制使用本地模型进行 AI 匹配
            if not self.use_ai_var.get():
                messagebox.showwarning("提示", "请先勾选“启用 AI 深度识别”，否则无法使用本地模型匹配。")
                return
            ai_ctx = self._build_ai_context("smart_summary") if hasattr(self, "_build_ai_context") else None
            provider = None
            if ai_ctx and getattr(ai_ctx, "default_values", None):
                provider = ai_ctx.default_values.get("ai_provider")
            if provider and str(provider) != "lm_studio":
                messagebox.showwarning("提示", "当前 AI 后端非本地模型，请在 AI 配置中心切换到本地模型。")
                return
            if _get_header_index(["科目编码"]) is None:
                messagebox.showwarning("提示", "未找到“科目编码”列。")
                return
            selection = preview_tree.selection() if preview_tree else ()
            if not selection:
                messagebox.showwarning("提示", "请先选择要匹配的行。")
                return
            updated = 0
            self.log_message(f"AI 重新匹配开始：选中 {len(selection)} 行")
            for row_id in selection:
                try:
                    row_idx = int(row_id)
                except ValueError:
                    continue
                summary_text = _get_summary_text(row_idx)
                current_code = _get_account_code(row_idx)
                self.log_message(f"  行{row_idx+1}: 当前科目={current_code} 摘要={summary_text[:60]}")
                if not summary_text:
                    self.log_message(f"  行{row_idx+1}: 摘要为空，跳过")
                    continue
                normalized_summary = str(summary_text).strip()
                if normalized_summary in rematch_history:
                    last_code = rematch_history.get(normalized_summary)
                    self.log_message(f"  行{row_idx+1}: 摘要已匹配过，跳过（上次={last_code}）")
                    continue
                use_ai = self.use_ai_var.get()
                use_foreign = self.use_foreign_currency_var.get()
                if rematch_skip_cache_var.get():
                    ai_code = self.summary_recognizer.recognize_account_with_ai_core(str(summary_text))
                    ai_result = {"科目编码": ai_code} if ai_code else {}
                else:
                    ai_result = self.summary_recognizer.recognize(str(summary_text), None, use_ai=use_ai, use_foreign_currency=use_foreign)
                self.log_message(f"  行{row_idx+1}: AI 返回={ai_result}")
                new_code = ai_result.get("科目编码") if isinstance(ai_result, dict) else None
                if not new_code or str(new_code) == str(current_code):
                    new_code = _best_alternative_account(summary_text, str(current_code))
                    if new_code:
                        self.log_message(f"  行{row_idx+1}: 候选替代科目={new_code}")
                    else:
                        self.log_message(f"  行{row_idx+1}: 未找到可用替代科目")
                if new_code and str(new_code) != str(current_code):
                    _set_account_code(row_idx, str(new_code))
                    current_values = list(preview_tree.item(row_id, "values"))
                    tree_account_index = _get_tree_col_index(["科目编码"])
                    if tree_account_index is not None and tree_account_index < len(current_values):
                        current_values[tree_account_index] = str(new_code)
                        preview_tree.item(row_id, values=current_values)
                    updated += 1
                    rematch_history[normalized_summary] = str(new_code)
                else:
                    self.log_message(f"  行{row_idx+1}: 未更新（无新科目或与当前相同）")
            messagebox.showinfo("完成", f"已更新 {updated} 行科目编码。")
            self.log_message(f"AI 重新匹配结束：已更新 {updated} 行")

        last_cell_context = {"row_idx": None, "col_index": None, "col_name": None}

        def _choose_base_data_for_cell():
            row_idx = last_cell_context.get("row_idx")
            col_index = last_cell_context.get("col_index")
            col_name = last_cell_context.get("col_name")
            if row_idx is None or col_index is None or not col_name:
                return
            table_name = FIELD_TABLE_MAP.get(col_name)
            if not table_name:
                messagebox.showinfo("提示", "当前列不支持基础数据选择。")
                return
            target_var = tk.StringVar()
            selector = self._open_base_data_selector(preview_window, table_name, target_var, f"选择{col_name}")
            if selector is not None:
                preview_window.wait_window(selector)
            new_val = target_var.get().strip()
            if not new_val:
                return
            header_item = header_items[col_index]
            src_idx = header_item.get("idx")
            if src_idx is None:
                preview_cell_overrides[(row_idx, col_name)] = new_val
            else:
                while src_idx >= len(base_rows[row_idx]):
                    base_rows[row_idx].append("")
                base_rows[row_idx][src_idx] = new_val
            _refresh_preview_tree()

        def _choose_base_data_apply_selected():
            col_index = last_cell_context.get("col_index")
            col_name = last_cell_context.get("col_name")
            if col_index is None or not col_name:
                return
            table_name = FIELD_TABLE_MAP.get(col_name)
            if not table_name:
                messagebox.showinfo("提示", "当前列不支持基础数据选择。")
                return
            selection = preview_tree.selection() if preview_tree else ()
            if not selection:
                messagebox.showwarning("提示", "请先选择要批量应用的行。")
                return
            target_var = tk.StringVar()
            selector = self._open_base_data_selector(preview_window, table_name, target_var, f"选择{col_name}")
            if selector is not None:
                preview_window.wait_window(selector)
            new_val = target_var.get().strip()
            if not new_val:
                return
            header_item = header_items[col_index]
            src_idx = header_item.get("idx")
            for row_id in selection:
                try:
                    row_idx = int(row_id)
                except ValueError:
                    continue
                if src_idx is None:
                    preview_cell_overrides[(row_idx, col_name)] = new_val
                else:
                    while src_idx >= len(base_rows[row_idx]):
                        base_rows[row_idx].append("")
                    base_rows[row_idx][src_idx] = new_val
            _refresh_preview_tree()

        def _append_match_item_to_account():
            selection = preview_tree.selection() if preview_tree else ()
            if not selection:
                messagebox.showwarning("提示", "请先选择要添加匹配项的行。")
                return
            try:
                row_idx = int(selection[0])
            except ValueError:
                return
            summary_text = _get_summary_text(row_idx)
            account_code = _get_account_code(row_idx)
            if not summary_text:
                messagebox.showwarning("提示", "摘要为空，无法添加匹配项。")
                return
            if not account_code:
                messagebox.showwarning("提示", "科目编码为空，无法添加匹配项。")
                return
            alias = simpledialog.askstring("添加匹配项", "请输入要加入的匹配项：", initialvalue=str(summary_text), parent=preview_window)
            if not alias:
                return
            alias = alias.strip()
            if not alias:
                return
            if not self.base_data_mgr:
                messagebox.showwarning("提示", "基础数据管理器未初始化，无法保存匹配项。")
                return

            def _parse_match_items(raw):
                if not raw:
                    return []
                if isinstance(raw, list):
                    return [str(x).strip() for x in raw if str(x).strip()]
                try:
                    parsed = json.loads(raw)
                    if isinstance(parsed, list):
                        return [str(x).strip() for x in parsed if str(x).strip()]
                except Exception:
                    pass
                parts = re.split(r"[\\n,，;；、]+", str(raw))
                return [p.strip() for p in parts if p and p.strip()]

            def _find_subject_record(code):
                rows = self.base_data_mgr.query("account_subject")
                for row in rows:
                    row_code = str(row.get("code", "") or "").strip()
                    if not row_code:
                        code_name = str(row.get("code_name", "") or "").strip()
                        m = re.search(r"\\[(\\d+)\\]", code_name)
                        if m:
                            row_code = m.group(1)
                    if row_code == code:
                        return row
                return None

            record = _find_subject_record(str(account_code).strip())
            if not record or "id" not in record:
                messagebox.showwarning("提示", f"未找到科目编码 {account_code} 的基础数据记录。")
                return
            existing = _parse_match_items(record.get("match_items", ""))
            if alias in existing:
                messagebox.showinfo("提示", "该匹配项已存在，无需重复添加。")
                return
            existing.append(alias)
            data = {"match_items": json.dumps(existing, ensure_ascii=False)}
            result = self.base_data_mgr.update_record("account_subject", int(record["id"]), data)
            if result.get("success"):
                if hasattr(self.base_data_mgr, "clear_lookup_cache"):
                    self.base_data_mgr.clear_lookup_cache()
                if self.summary_recognizer:
                    self.summary_recognizer.refresh_cache()
                messagebox.showinfo("成功", f"已添加匹配项到科目 {account_code}")
            else:
                messagebox.showerror("错误", result.get("message", "保存匹配项失败"))

        def _show_preview_menu(event):
            if preview_tree is None:
                return
            row_id = preview_tree.identify_row(event.y)
            col_id = preview_tree.identify_column(event.x)
            if row_id:
                if row_id not in preview_tree.selection():
                    preview_tree.selection_set(row_id)
                try:
                    row_idx = int(row_id)
                except ValueError:
                    row_idx = None
                try:
                    col_index = int(col_id.replace("#", "")) - 1
                except Exception:
                    col_index = None
                col_name = None
                if col_index is not None:
                    cols = list(preview_tree["columns"])
                    if 0 <= col_index < len(cols):
                        col_name = cols[col_index]
                last_cell_context.update({"row_idx": row_idx, "col_index": col_index, "col_name": col_name})
                if col_name in FIELD_TABLE_MAP:
                    preview_menu.entryconfig("从基础数据选择", state="normal")
                    preview_menu.entryconfig("批量应用到选中行同列", state="normal")
                else:
                    preview_menu.entryconfig("从基础数据选择", state="disabled")
                    preview_menu.entryconfig("批量应用到选中行同列", state="disabled")
                preview_menu.tk_popup(event.x_root, event.y_root)

        preview_menu = tk.Menu(preview_window, tearoff=0)
        preview_menu.add_command(label="从基础数据选择", command=_choose_base_data_for_cell)
        preview_menu.add_command(label="批量应用到选中行同列", command=_choose_base_data_apply_selected)
        preview_menu.add_command(label="将摘要加入科目匹配项", command=_append_match_item_to_account)
        preview_menu.add_checkbutton(label="仅在收藏科目中匹配", variable=rematch_favorites_only_var)
        preview_menu.add_checkbutton(label="强制AI识别（跳过缓存）", variable=rematch_skip_cache_var)
        preview_menu.add_command(label="AI 重新匹配科目（排除当前）", command=_ai_rematch_selected)

        # 底部按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))

        result = {"confirmed": False}

        def export_preview():
            """导出完整预览数据为 Excel（包含所有行）"""
            path = filedialog.asksaveasfilename(
                title="导出预览为 Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            )
            if not path:
                return
            try:
                preview_headers = [item["name"] for item in header_items]
                preview_rows = _get_preview_rows()
                mapped_headers, mapped_rows, mapped = apply_export_format(
                    "main_export",
                    preview_headers,
                    preview_rows,
                    base_data_mgr=self.base_data_mgr
                )
                if mapped:
                    df = pd.DataFrame(mapped_rows, columns=mapped_headers)
                else:
                    df = pd.DataFrame(preview_rows, columns=preview_headers)
                df.to_excel(path, index=False)
                messagebox.showinfo("完成", f"已导出预览数据：\n{path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败：{e}")

        def confirm():
            current_signature = [(item["idx"], item["name"]) for item in header_items]
            if len(base_rows) == len(output_rows):
                for i in range(len(output_rows)):
                    output_rows[i] = base_rows[i]
            if current_signature != original_signature:
                self._preview_header_override = True
                self._preview_header_names = [item["name"] for item in header_items]
                output_rows[:] = _get_preview_rows()
            result["confirmed"] = True
            preview_window.destroy()

        def cancel():
            result["confirmed"] = False
            preview_window.destroy()

        ttk.Button(button_frame, text="导出预览为Excel", command=export_preview, width=18).pack(side="left")
        ttk.Button(button_frame, text="确认并导出", command=confirm, width=15).pack(side="right", padx=5)
        ttk.Button(button_frame, text="取消", command=cancel, width=15).pack(side="right")

        # 使窗口模态
        preview_window.transient(self.root)
        preview_window.grab_set()
        self.root.wait_window(preview_window)

        return result["confirmed"]

    def _export_to_excel(self, output_rows, unmatched_info=None):
        """导出到Excel文件"""
        if not output_rows:
            return

        try:
            source_headers = []
            if getattr(self, "_preview_header_override", False) and self._preview_header_names:
                source_headers = self._preview_header_names
            elif self.template_headers:
                source_headers = [h.name for h in self.template_headers]
            mapped_headers, mapped_rows, mapped = apply_export_format(
                "main_export",
                source_headers,
                output_rows,
                base_data_mgr=self.base_data_mgr
            )
            if mapped:
                save_path = filedialog.asksaveasfilename(
                    title="保存导出文件",
                    defaultextension=".xlsx",
                    filetypes=[("Excel 文件", "*.xlsx")],
                    initialfile="导出结果.xlsx"
                )
                if not save_path:
                    return
                df = pd.DataFrame(mapped_rows, columns=mapped_headers)
                df.to_excel(save_path, index=False)
                self.log_message(f"导出成功(自定义格式): {save_path}")
                messagebox.showinfo("成功", f"成功导出到：\n{save_path}")
                return

            # 解析当前有效模式（处理自定义方案）
            raw_mode = self.convert_mode_var.get()
            effective_mode = raw_mode
            if raw_mode.startswith("方案: "):
                scheme_name = raw_mode.replace("方案: ", "")
                if hasattr(self, "cached_schemes") and scheme_name in self.cached_schemes:
                    effective_mode = self.cached_schemes[scheme_name]["base_mode"]
                else:
                    effective_mode = MODE_GENERAL_VOUCHER

            # 如果选择了表头方案，优先按方案顺序导出（忽略模板列顺序）
            scheme_name = getattr(self, "header_scheme_var", tk.StringVar(value="")).get().strip()
            if scheme_name and scheme_name in HEADER_SCHEMES and self.export_scheme_override_var.get() and not getattr(self, "_preview_header_override", False):
                self._debug_log(f"按表头方案导出: {scheme_name}")
                scheme = HEADER_SCHEMES.get(scheme_name) or []
                name_to_idx = {h.name: i for i, h in enumerate(self.template_headers)}
                ordered_names = []
                ordered_indices = []
                defaults = []
                for item in scheme:
                    name = item.get("name")
                    if not name:
                        continue
                    if name in name_to_idx:
                        ordered_names.append(name)
                        ordered_indices.append(name_to_idx[name])
                        defaults.append(item.get("default", ""))
                    else:
                        ordered_names.append(name)
                        ordered_indices.append(None)
                        defaults.append(item.get("default", ""))
                for h in self.template_headers:
                    if h.name not in ordered_names:
                        ordered_names.append(h.name)
                        ordered_indices.append(name_to_idx[h.name])
                        defaults.append("")

                ordered_rows = []
                for row in output_rows:
                    new_row = []
                    for idx, d in zip(ordered_indices, defaults):
                        if idx is None:
                            new_row.append(d)
                        else:
                            new_row.append(row[idx] if idx < len(row) else "")
                    ordered_rows.append(new_row)

                save_path = filedialog.asksaveasfilename(
                    title="保存导出文件",
                    defaultextension=".xlsx",
                    filetypes=[("Excel 文件", "*.xlsx")],
                    initialfile="导出结果.xlsx"
                )
                if not save_path:
                    return
                df = pd.DataFrame(ordered_rows, columns=ordered_names)
                df.to_excel(save_path, index=False)
                self.log_message(f"导出成功(表头方案): {save_path}")
                messagebox.showinfo("成功", f"成功导出到：\n{save_path}")
                return

            if getattr(self, "_preview_header_override", False) and self._preview_header_names:
                self._debug_log("按预览自定义表头导出")
                save_path = filedialog.asksaveasfilename(
                    title="保存导出文件",
                    defaultextension=".xlsx",
                    filetypes=[("Excel 文件", "*.xlsx")],
                    initialfile="导出结果.xlsx"
                )
                if not save_path:
                    return
                df = pd.DataFrame(output_rows, columns=self._preview_header_names)
                df.to_excel(save_path, index=False)
                self.log_message(f"导出成功: {save_path}")
                messagebox.showinfo("成功", f"成功导出到：\n{save_path}")
                self._preview_header_override = False
                self._preview_header_names = None
                return

            if effective_mode == MODE_ORIGINAL:
                self._debug_log("原格式模式导出")
                save_path = filedialog.asksaveasfilename(
                    title="保存导出文件",
                    defaultextension=".xlsx",
                    filetypes=[("Excel 文件", "*.xlsx")],
                    initialfile="导出结果.xlsx"
                )
                if not save_path:
                    return
                df = pd.DataFrame(output_rows, columns=[h.name for h in self.template_headers])
                df.to_excel(save_path, index=False)
                self.log_message(f"导出成功: {save_path}")
                messagebox.showinfo("成功", f"成功导出到：\n{save_path}")
                return

            # 加载模板
            template_path = self.template_path_var.get() or TEMPLATE_FILE
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"未找到模板文件：{template_path}")

            self.log_message(f"正在加载导出模板: {template_path}")
            wb = load_workbook(template_path)
            ws = wb.active

            # 确定要保留的字段集合
            keep_fields = None
            
            if effective_mode == MODE_GENERAL_VOUCHER:
                # keep_fields = set(GENERAL_VOUCHER_FIELDS)
                keep_fields = None # 不过滤列，完全保留模板结构，解决导出结果与模板不一致的问题
            elif effective_mode == MODE_CUSTOM:
                # 自定义模式下，不进行列过滤，完全保留模板结构
                keep_fields = None

            # 如果设定了过滤规则，执行过滤
            if keep_fields is not None:
                cols_to_delete = []  # indices to delete
                kept_indices = []    # indices to keep
                
                # 遍历 self.template_headers (假设它与当前打开的模板列一一对应)
                for i, header in enumerate(self.template_headers):
                    if header.name in keep_fields:
                        kept_indices.append(i)
                    else:
                        cols_to_delete.append(i)
                
                # 从后往前删除，避免索引偏移
                cols_to_delete.sort(reverse=True)
                for idx in cols_to_delete:
                    ws.delete_cols(idx + 1)
                
                # 同步过滤 output_rows 的数据，只保留剩下的列
                new_output_rows = []
                for row in output_rows:
                    if len(row) == len(self.template_headers):
                        new_row = [row[i] for i in kept_indices]
                        new_output_rows.append(new_row)
                    else:
                        new_output_rows.append(row)
                output_rows = new_output_rows

            # 写入数据
            self.log_message(f"开始写入 {len(output_rows)} 行数据到 Excel...")
            # 从第2行开始（假设第1行是表头）
            start_row = 2
            for i, row_data in enumerate(output_rows):
                if (i + 1) % 500 == 0:
                     self.log_message(f"  正在写入第 {i + 1} 行...")
                for col_idx, value in enumerate(row_data):
                    # 写入单元格（注意：openpyxl行列从1开始）
                    cell = ws.cell(row=start_row, column=col_idx + 1, value=value)
                start_row += 1

            # 保存文件
            save_path = filedialog.asksaveasfilename(
                title="保存导出文件",
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx")],
                initialfile="导出结果.xlsx"
            )
            if not save_path:
                return

            wb.save(save_path)
            self.log_message(f"导出成功: {save_path}")
            
            # 显示成功信息 (含未匹配提示)
            msg = f"成功导出到：\n{save_path}"
            if unmatched_info:
                msg += f"\n\n注意：有 {len(unmatched_info)} 行数据存在基础资料未匹配项，已在日志中记录。"
            
            messagebox.showinfo("成功", msg)

        except Exception as e:
            # 捕获所有异常并提示
            import traceback
            traceback.print_exc()
            self.log_message(f"导出失败: {e}\n{traceback.format_exc()}")
            err_msg = f"保存结果失败:\n{e}"
            messagebox.showerror("错误", err_msg)

    # ========== 基础数据管理标签页功能 ==========

    def _load_base_data_table(self):
        """加载基础数据表"""
        if not self.base_data_mgr:
            self._set_status_text("错误：基础数据管理器未初始化")
            return

        table_name = self.current_table.get()

        try:
            # 获取数据
            if table_name == "smart_recognition_cache":
                results = self.base_data_mgr.get_all_cached_recognitions()
            elif table_name == "recognition_rules":
                results = self.base_data_mgr.get_recognition_rules()
            else:
                results = self.base_data_mgr.query(table_name)

            self._populate_base_data_table(results, table_name=table_name)

        except Exception as e:
            messagebox.showerror("错误", f"加载数据失败：\n{e}")
            self._set_status_text(f"加载失败: {str(e)}")

    def _refresh_base_data_view(self):
        """按当前搜索条件刷新基础数据视图（用于收藏筛选切换）"""
        keyword = self.search_var.get().strip() if hasattr(self, "search_var") else ""
        if keyword:
            self._search_base_data()
        else:
            self._load_base_data_table()

    def _update_base_data_favorites(self, add: bool):
        """将当前选中记录加入/移出收藏"""
        if not self.base_data_mgr:
            return

        table_name = self.current_table.get()
        selection = self.base_data_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择需要收藏的记录。")
            return

        columns = list(self.base_data_tree["columns"])
        if not columns:
            return

        favorites = self._load_base_data_favorites(table_name)
        current = list(favorites)
        changed = False

        for item in selection:
            values = self.base_data_tree.item(item, "values")
            row_data = dict(zip(columns, values))
            key = self._get_base_data_favorite_key(table_name, row_data, columns)
            if not key:
                continue
            if add and key not in current:
                current.append(key)
                changed = True
            if (not add) and key in current:
                current.remove(key)
                changed = True

        if changed:
            self._save_base_data_favorites(table_name, current)
            self._refresh_base_data_view()
            action = "加入" if add else "移除"
            self._set_status_text(f"{table_name} - 已{action}收藏，共 {len(current)} 项")
        else:
            messagebox.showinfo("提示", "选中记录未发生收藏变更。")

    def _search_base_data(self):
        """搜索基础数据"""
        if not self.base_data_mgr:
            return

        keyword = self.search_var.get().strip()
        table_name = self.current_table.get()
        search_field = self.search_field_var.get().strip()

        try:
            if table_name == "smart_recognition_cache":
                # 对缓存表进行搜索 (目前没有直接的 search_by_name，这里简单过滤)
                all_results = self.base_data_mgr.get_all_cached_recognitions()
                results = [r for r in all_results if keyword.lower() in str(r.get("summary", "")).lower() or keyword.lower() in str(r.get("account_code", "")).lower()]
            else:
                results = self.base_data_mgr.query(table_name)
                if keyword and results:
                    # 按指定字段过滤，若字段不存在则对整行做模糊过滤
                    k_lower = keyword.lower()
                    if search_field and search_field in results[0].keys():
                        results = [r for r in results if k_lower in str(r.get(search_field, "")).lower()]
                    else:
                        results = [
                            r for r in results
                            if any(k_lower in str(v).lower() for v in r.values())
                        ]

            self._populate_base_data_table(results, table_name=table_name)

            row_count = len(self.base_data_tree.get_children(""))
            fav_suffix = ""
            if getattr(self, "base_data_show_favorites_only_var", None) and self.base_data_show_favorites_only_var.get():
                fav_suffix = "（仅显示收藏）"
            if keyword:
                self._set_status_text(f"找到 {row_count} 条包含 '{keyword}' 的记录{fav_suffix}")
            else:
                self._set_status_text(f"共 {row_count} 条记录{fav_suffix}")

        except Exception as e:
            messagebox.showerror("错误", f"搜索失败：\n{e}")

    def _populate_base_data_table(self, results, table_name=None):
        """渲染基础数据表格（支持搜索和加载复用）"""
        # 清空现有内容
        for item in self.base_data_tree.get_children():
            self.base_data_tree.delete(item)

        table_name = table_name or self.current_table.get()
        favorites = self._load_base_data_favorites(table_name)
        fav_set = set(favorites)
        show_favorites_only = bool(
            getattr(self, "base_data_show_favorites_only_var", None)
            and self.base_data_show_favorites_only_var.get()
        )
        if show_favorites_only and results:
            filtered = []
            for row in results:
                key = self._get_base_data_favorite_key(table_name, row, list(row.keys()))
                if key and key in fav_set:
                    filtered.append(row)
            results = filtered

        if not results:
            suffix = " (仅显示收藏)" if show_favorites_only else ""
            self._set_status_text(f"{table_name} - 无数据{suffix}")
            # 为空时也要设置列，防止Treeview报错
            if table_name == "smart_recognition_cache":
                columns = ["id", "summary", "match_items", "account_code", "created_at"]
                self.base_data_tree["columns"] = columns
                self.base_data_tree["show"] = "headings"
                for col in columns:
                    heading_text = "映射匹配项" if col == "match_items" else col
                    self.base_data_tree.heading(col, text=heading_text)
                    self.base_data_tree.column(col, width=100)
                    if col == "summary": self.base_data_tree.column(col, width=300)
                    if col == "match_items": self.base_data_tree.column(col, width=220)
                    if col == "created_at": self.base_data_tree.column(col, width=150)
            elif table_name.startswith("custom:"):
                columns = self.base_data_mgr.get_table_columns(table_name)
                self.base_data_tree["columns"] = columns
                self.base_data_tree["show"] = "headings"
                for col in columns:
                    self.base_data_tree.heading(col, text=col)
                    self.base_data_tree.column(col, width=120)
            else:
                # 对于非缓存表，如果无数据，先清空列头，避免显示旧的
                self.base_data_tree["columns"] = []
            return

        # 将匹配项展示为易读的逗号分隔形式
        tables_with_json = {"smart_recognition_cache", "account_subject", "product", "business_partner", "recognition_rules"}
        if table_name in tables_with_json:
            normalized_results = []
            for row in results:
                row_copy = dict(row)
                
                # 处理 match_items
                if "match_items" in row_copy:
                    raw_items = row_copy.get("match_items", "")
                    display_items = ""
                    try:
                        parsed = json.loads(raw_items) if isinstance(raw_items, str) else (raw_items or [])
                        if isinstance(parsed, list):
                            display_items = ", ".join(str(i).strip() for i in parsed if str(i).strip())
                        else:
                            display_items = str(parsed)
                    except Exception:
                        display_items = str(raw_items) if raw_items is not None else ""
                    row_copy["match_items"] = display_items
                
                # 处理 keywords (识别规则表)
                if "keywords" in row_copy:
                    raw_kws = row_copy.get("keywords", "")
                    display_kws = ""
                    try:
                        parsed = json.loads(raw_kws) if isinstance(raw_kws, str) else (raw_kws or [])
                        if isinstance(parsed, list):
                            display_kws = ", ".join(str(i).strip() for i in parsed if str(i).strip())
                        else:
                            display_kws = str(parsed)
                    except Exception:
                        display_kws = str(raw_kws) if raw_kws is not None else ""
                    row_copy["keywords"] = display_kws

                normalized_results.append(row_copy)
            results = normalized_results

        # 获取列名
        columns = list(results[0].keys())
        self.current_columns = columns
        # 更新搜索字段下拉
        self.search_field_combo["values"] = columns
        if self.search_field_var.get() not in columns:
            self.search_field_var.set(columns[0] if columns else "")

        # 配置树形视图列
        self.base_data_tree["columns"] = columns
        self.base_data_tree["show"] = "headings"

        # 设置列标题和宽度，并绑定排序
        for col in columns:
            heading_text = "映射匹配项" if col == "match_items" else col
            self.base_data_tree.heading(col, text=heading_text, command=lambda c=col: self._sort_treeview(c))
            width = 100
            if col == "summary": width = 300
            elif col == "created_at": width = 150
            elif col == "account_code": width = 80
            elif col == "match_items": width = 220
            self.base_data_tree.column(col, width=width, stretch=True)

        # 插入数据
        for row in results:
            values = [row.get(col, "") for col in columns]
            key = self._get_base_data_favorite_key(table_name, row, columns)
            tags = ("favorite",) if key and key in fav_set else ()
            self.base_data_tree.insert("", "end", values=values, tags=tags)
        self.base_data_tree.tag_configure("favorite", background="#fff7d6")

        suffix = " (仅显示收藏)" if show_favorites_only else ""
        self._set_status_text(f"{table_name} - 共 {len(results)} 条记录{suffix}")

    def _sort_treeview(self, column):
        """点击表头排序，数字优先按数值排序"""
        # 读取当前数据
        items = self.base_data_tree.get_children("")
        data = []
        for iid in items:
            values = self.base_data_tree.item(iid, "values")
            data.append((iid, values))

        if not data:
            return

        col_idx = self.base_data_tree["columns"].index(column)

        def _as_number(val):
            try:
                return float(val)
            except (TypeError, ValueError):
                return val

        # 切换排序方向
        ascending = not self.sort_states.get(column, True)
        self.sort_states[column] = ascending

        sorted_data = sorted(
            data,
            key=lambda item: _as_number(item[1][col_idx]),
            reverse=not ascending
        )

        # 重新排列
        for idx, (iid, _) in enumerate(sorted_data):
            self.base_data_tree.move(iid, "", idx)

    def _add_base_data_record(self):
        """新增基础数据记录"""
        table_name = self.current_table.get()
        if table_name == "smart_recognition_cache":
            # 打开缓存专用新增对话框
            self._add_cached_recognition_dialog(self._load_base_data_table)
            return
        elif table_name == "recognition_rules":
            # 打开规则专用新增对话框
            self._recognition_rule_dialog(None, self._load_base_data_table)
            return

        if not self.base_data_mgr:
            return

        columns = self.base_data_mgr.get_table_columns(table_name)

        self._edit_record_dialog(table_name, columns, None, self._load_base_data_table)

    def _add_cached_recognition_dialog(self, refresh_callback):
        """新增智能识别缓存记录（手动）"""
        dialog = tk.Toplevel(self.root)
        dialog.title("新增智能识别缓存")
        dialog.geometry("650x460")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="摘要（原文）:").grid(row=0, column=0, sticky="ne", padx=8, pady=8)
        summary_text = tk.Text(dialog, height=4, width=55, wrap="word")
        summary_text.grid(row=0, column=1, columnspan=3, sticky="we", padx=5, pady=8)

        normalize_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            dialog,
            text="保存前自动标准化摘要/匹配项（去日期/金额噪声，提高命中率）",
            variable=normalize_var
        ).grid(row=1, column=1, columnspan=3, sticky="w", padx=5)

        ttk.Label(dialog, text="科目编码:").grid(row=2, column=0, sticky="e", padx=8, pady=8)
        account_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=account_var, width=30).grid(row=2, column=1, sticky="w", padx=5, pady=8)

        ttk.Label(dialog, text="映射匹配项（多对一，逗号/换行分隔）:").grid(row=3, column=0, sticky="ne", padx=8, pady=4)
        match_items_text = tk.Text(dialog, height=3, width=55, wrap="word")
        match_items_text.grid(row=3, column=1, columnspan=3, sticky="we", padx=5, pady=4)

        preview_var = tk.StringVar()
        ttk.Label(dialog, text="将被缓存的摘要:").grid(row=4, column=0, sticky="ne", padx=8)
        preview_label = ttk.Label(dialog, textvariable=preview_var, wraplength=420, foreground="gray")
        preview_label.grid(row=4, column=1, columnspan=3, sticky="w", padx=5)

        def _normalize(text: str) -> str:
            if not text:
                return ""
            if normalize_var.get() and self.summary_recognizer:
                try:
                    return self.summary_recognizer._normalize_summary(text)
                except Exception:
                    pass
            return text.strip()

        def _parse_match_items() -> List[str]:
            raw = match_items_text.get("1.0", "end").strip()
            if not raw:
                return []
            parts = re.split(r"[\\n,，;；、]+", raw)
            items = []
            for part in parts:
                p = part.strip()
                if not p:
                    continue
                items.append(_normalize(p))
            return items

        def refresh_preview():
            raw = summary_text.get("1.0", "end").strip()
            preview_var.set(_normalize(raw))

        ttk.Button(dialog, text="刷新预览", command=refresh_preview, width=10).grid(row=0, column=4, padx=5, pady=8, sticky="n")

        def save():
            raw_summary = summary_text.get("1.0", "end").strip()
            account_code = account_var.get().strip()

            if not raw_summary:
                messagebox.showwarning("警告", "摘要不能为空！")
                return
            if not account_code:
                messagebox.showwarning("警告", "科目编码不能为空！")
                return

            summary_to_save = _normalize(raw_summary)
            match_items = _parse_match_items()
            try:
                self.base_data_mgr.save_cached_recognition(summary_to_save, account_code, match_items)
                messagebox.showinfo("成功", f"已保存缓存：\n摘要= {summary_to_save}\n科目= {account_code}")
                dialog.destroy()
                refresh_callback()
            except Exception as e:
                messagebox.showerror("错误", f"保存失败：{e}")

        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=4, column=0, columnspan=3, pady=10)
        ttk.Button(button_frame, text="保存", command=save, width=10).pack(side="left", padx=5)
        ttk.Button(button_frame, text="取消", command=dialog.destroy, width=10).pack(side="left", padx=5)

    def _recognition_rule_dialog(self, record_data, refresh_callback):
        """新增/编辑识别规则对话框"""
        is_edit = record_data is not None
        dialog = tk.Toplevel(self.root)
        dialog.title("编辑识别规则" if is_edit else "新增识别规则")
        dialog.geometry("600x550")
        dialog.transient(self.root)
        dialog.grab_set()

        # 初始化数据
        data = record_data or {}
        
        # 布局
        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill="both", expand=True)
        
        row = 0
        # 规则类型
        ttk.Label(frame, text="规则类型:").grid(row=row, column=0, sticky="e", pady=5)
        type_var = tk.StringVar(value=data.get("rule_type", "business"))
        type_combo = ttk.Combobox(frame, textvariable=type_var, values=["business", "account", "department"], state="readonly")
        type_combo.grid(row=row, column=1, sticky="w", pady=5)
        row += 1
        
        # 规则名称
        ttk.Label(frame, text="规则名称:").grid(row=row, column=0, sticky="e", pady=5)
        name_var = tk.StringVar(value=data.get("name", ""))
        ttk.Entry(frame, textvariable=name_var, width=40).grid(row=row, column=1, sticky="w", pady=5)
        row += 1
        
        # 关键词
        ttk.Label(frame, text="关键词 (逗号分隔):").grid(row=row, column=0, sticky="ne", pady=5)
        keywords_text = tk.Text(frame, height=4, width=40, wrap="word")
        keywords_text.grid(row=row, column=1, sticky="w", pady=5)
        
        # 预填充关键词
        raw_kws = data.get("keywords", "")
        if isinstance(raw_kws, str):
            # 尝试解析JSON
            try:
                parsed = json.loads(raw_kws)
                if isinstance(parsed, list):
                    keywords_text.insert("1.0", ", ".join(parsed))
                else:
                    keywords_text.insert("1.0", str(parsed))
            except:
                keywords_text.insert("1.0", raw_kws) # 原样显示
        else:
            keywords_text.insert("1.0", str(raw_kws))
        
        row += 1

        # 科目编码
        ttk.Label(frame, text="科目编码:").grid(row=row, column=0, sticky="e", pady=5)
        acc_var = tk.StringVar(value=data.get("account_code", ""))
        ttk.Entry(frame, textvariable=acc_var, width=20).grid(row=row, column=1, sticky="w", pady=5)
        ttk.Label(frame, text="(业务/科目规则用)").grid(row=row, column=2, sticky="w")
        row += 1

        # 借贷类型
        ttk.Label(frame, text="借贷类型:").grid(row=row, column=0, sticky="e", pady=5)
        trans_type_var = tk.StringVar(value=data.get("transaction_type", ""))
        ttk.Combobox(frame, textvariable=trans_type_var, values=["1", "2"], width=5).grid(row=row, column=1, sticky="w", pady=5)
        ttk.Label(frame, text="(1=出/借, 2=入/贷)").grid(row=row, column=2, sticky="w")
        row += 1

        # 摘要编码
        ttk.Label(frame, text="摘要编码:").grid(row=row, column=0, sticky="e", pady=5)
        summary_code_var = tk.StringVar(value=data.get("summary_code", ""))
        ttk.Entry(frame, textvariable=summary_code_var, width=10).grid(row=row, column=1, sticky="w", pady=5)
        row += 1

        # 部门编码
        ttk.Label(frame, text="部门编码:").grid(row=row, column=0, sticky="e", pady=5)
        dept_code_var = tk.StringVar(value=data.get("dept_code", ""))
        ttk.Entry(frame, textvariable=dept_code_var, width=20).grid(row=row, column=1, sticky="w", pady=5)
        ttk.Label(frame, text="(部门规则用)").grid(row=row, column=2, sticky="w")
        row += 1
        
        # 优先级
        ttk.Label(frame, text="优先级:").grid(row=row, column=0, sticky="e", pady=5)
        priority_var = tk.IntVar(value=data.get("priority", 0))
        ttk.Entry(frame, textvariable=priority_var, width=10).grid(row=row, column=1, sticky="w", pady=5)
        ttk.Label(frame, text="(数字越大越优先)").grid(row=row, column=2, sticky="w")
        row += 1

        def save():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning("提示", "规则名称不能为空")
                return
            
            # 解析关键词
            kw_raw = keywords_text.get("1.0", "end").strip()
            keywords = [k.strip() for k in re.split(r'[,，\n]', kw_raw) if k.strip()]
            
            rule_data = {
                "rule_type": type_var.get(),
                "name": name,
                "keywords": keywords,
                "account_code": acc_var.get().strip() or None,
                "transaction_type": trans_type_var.get().strip() or None,
                "summary_code": summary_code_var.get().strip() or None,
                "dept_code": dept_code_var.get().strip() or None,
                "priority": priority_var.get()
            }
            
            try:
                if is_edit:
                    res = self.base_data_mgr.update_recognition_rule(data["id"], rule_data)
                else:
                    res = self.base_data_mgr.add_recognition_rule(**rule_data)
                
                if res["success"]:
                    messagebox.showinfo("成功", "规则保存成功")
                    dialog.destroy()
                    refresh_callback()
                    # 重新初始化识别器规则
                    if self.summary_recognizer:
                        self.summary_recognizer._init_recognition_rules()
                else:
                    messagebox.showerror("错误", res["message"])
            except Exception as e:
                messagebox.showerror("错误", f"保存异常: {e}")

        btn_frame = ttk.Frame(dialog, padding=10)
        btn_frame.pack(side="bottom", fill="x")
        ttk.Button(btn_frame, text="保存", command=save).pack(side="right", padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side="right")


    def _edit_base_data_record(self):
        """编辑基础数据记录（支持缓存批量编辑）"""
        if not self.base_data_mgr:
            return

        selections = self.base_data_tree.selection()
        if not selections:
            messagebox.showwarning("提示", "请先选择要编辑的记录")
            return

        table_name = self.current_table.get()
        
        # === 批量编辑逻辑 (仅限智能识别缓存) ===
        if len(selections) > 1:
            if table_name == "smart_recognition_cache":
                self._batch_edit_cached_recognition_dialog(selections, self._load_base_data_table)
                return
            else:
                messagebox.showwarning("提示", "当前数据类型不支持批量编辑，请只选择一条记录。")
                return

        # === 单条编辑逻辑 ===
        item = selections[0]
        values = self.base_data_tree.item(item, "values")
        
        if table_name == "smart_recognition_cache":
            # 针对缓存表，只允许编辑科目编码
            columns = ["id", "summary", "match_items", "account_code", "created_at"]
            record_data = dict(zip(columns, values))
            self._edit_cached_recognition_dialog(record_data, self._load_base_data_table)
            return
        elif table_name == "recognition_rules":
            # 针对规则表，使用专用编辑框
            columns = self.base_data_mgr.get_table_columns(table_name)
            record_data = dict(zip(columns, values))
            self._recognition_rule_dialog(record_data, self._load_base_data_table)
            return

        columns = self.base_data_mgr.get_table_columns(table_name)
        record_data = dict(zip(columns, values))

        self._edit_record_dialog(table_name, columns, record_data, self._load_base_data_table)

    def _batch_edit_cached_recognition_dialog(self, selections, refresh_callback):
        """批量编辑智能识别缓存对话框"""
        count = len(selections)
        dialog = tk.Toplevel(self.root)
        dialog.title(f"批量编辑 ({count} 条记录)")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text=f"正在批量修改 {count} 条缓存记录", font=("", 10, "bold")).pack(pady=10)
        
        frame = ttk.Frame(dialog)
        frame.pack(pady=5)
        ttk.Label(frame, text="统一修改科目编码为:").pack(side="left")
        new_code_var = tk.StringVar()
        ttk.Entry(frame, textvariable=new_code_var, width=20).pack(side="left", padx=5)

        def save_batch():
            new_code = new_code_var.get().strip()
            if not new_code:
                messagebox.showwarning("警告", "科目编码不能为空！")
                return
            
            success_count = 0
            
            # 循环更新
            for item in selections:
                values = self.base_data_tree.item(item, "values")
                record_id = int(values[0])
                result = self.base_data_mgr.update_cached_recognition(record_id, new_code)
                if result["success"]:
                    success_count += 1
            
            messagebox.showinfo("完成", f"成功更新 {success_count} 条记录")
            dialog.destroy()
            refresh_callback()

        ttk.Button(dialog, text="保存全部", command=save_batch).pack(pady=15)

    def _delete_base_data_record(self):
        """删除基础数据记录（支持批量）"""
        if not self.base_data_mgr:
            return

        selections = self.base_data_tree.selection()
        if not selections:
            messagebox.showwarning("提示", "请先选择要删除的记录")
            return

        count = len(selections)
        table_name = self.current_table.get()
        
        # 确认删除
        confirm = messagebox.askyesno("确认删除", f"确定要删除选中的 {count} 条记录吗？\n此操作不可恢复！")
        if not confirm:
            return
        
        success_count = 0
        fail_count = 0
        
        for item in selections:
            values = self.base_data_tree.item(item, "values")
            record_id = values[0]  # 第一列是id
            
            result = None
            if table_name == "smart_recognition_cache":
                result = self.base_data_mgr.delete_cached_recognition(int(record_id))
            else:
                result = self.base_data_mgr.delete_record(table_name, int(record_id))
            
            if result and result["success"]:
                success_count += 1
            else:
                fail_count += 1

        if success_count > 0:
            if table_name == "bank_account":
                self._invalidate_bank_account_cache()
            # 如果删除的是基础数据表（非智能识别缓存），自动刷新识别缓存
            if table_name != "smart_recognition_cache" and self.summary_recognizer:
                self.summary_recognizer.refresh_cache()
                
                # 如果删除的是规则表，还需要重新加载规则
                if table_name == "recognition_rules":
                    self.summary_recognizer._init_recognition_rules()
                
                messagebox.showinfo("操作完成",
                    f"成功删除 {success_count} 条记录" + (f"，失败 {fail_count} 条" if fail_count > 0 else "") +
                    "\n\n已自动刷新智能识别缓存/规则。")
            else:
                messagebox.showinfo("操作完成", f"成功删除 {success_count} 条记录" + (f"，失败 {fail_count} 条" if fail_count > 0 else ""))
            self._load_base_data_table()  # 刷新列表
        else:
            messagebox.showerror("错误", "删除失败")

    def _edit_cached_recognition_dialog(self, record_data: Dict, refresh_callback):
        """编辑智能识别缓存对话框 (科目编码 + 映射匹配项)"""
        dialog = tk.Toplevel(self.root)
        dialog.title("编辑智能识别缓存")
        dialog.geometry("560x320")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="摘要:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        ttk.Label(dialog, text=record_data.get("summary", ""), wraplength=350, justify="left").grid(row=0, column=1, sticky="w", padx=5, pady=5)

        ttk.Label(dialog, text="当前科目编码:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        current_code_label = ttk.Label(dialog, text=record_data.get("account_code", ""), font=("", 10, "bold"))
        current_code_label.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        ttk.Label(dialog, text="新科目编码:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
        new_code_var = tk.StringVar(value=record_data.get("account_code", ""))
        new_code_entry = ttk.Entry(dialog, textvariable=new_code_var, width=30)
        new_code_entry.grid(row=2, column=1, sticky="we", padx=5, pady=5)

        ttk.Label(dialog, text="映射匹配项:").grid(row=3, column=0, sticky="ne", padx=5, pady=5)
        match_items_text = tk.Text(dialog, height=3, width=45, wrap="word")
        match_items_text.grid(row=3, column=1, columnspan=2, sticky="we", padx=5, pady=5)

        # 预填充匹配项
        existing_items = record_data.get("match_items", "")
        preset = ""
        try:
            parsed = json.loads(existing_items) if isinstance(existing_items, str) else (existing_items or [])
            if isinstance(parsed, list):
                preset = "\n".join(str(i).strip() for i in parsed if str(i).strip())
            elif parsed:
                preset = str(parsed)
        except Exception:
            preset = str(existing_items) if existing_items else ""
        if preset:
            match_items_text.insert("1.0", preset)

        def _parse_match_items() -> List[str]:
            raw = match_items_text.get("1.0", "end").strip()
            if not raw:
                return []
            parts = re.split(r"[\n,，;；、]+", raw)
            return [p.strip() for p in parts if p and p.strip()]

        # AI 重新识别功能
        def retry_ai():
            if not self.summary_recognizer:
                messagebox.showerror("错误", "智能识别器未初始化")
                return
            
            summary = record_data.get("summary", "")
            if not summary:
                return

            # 禁用按钮防止重复点击
            retry_btn.config(state="disabled", text="识别中...")
            dialog.update()

            try:
                # 调用核心 AI 逻辑（不走缓存）
                code = self.summary_recognizer.recognize_account_with_ai_core(summary)
                if code:
                    new_code_var.set(code)
                    messagebox.showinfo("识别成功", f"AI 建议科目: {code}")
                else:
                    messagebox.showwarning("无结果", "AI 未能识别出有效科目")
            except Exception as e:
                messagebox.showerror("错误", f"识别失败: {e}")
            finally:
                retry_btn.config(state="normal", text="AI 重新识别")

        retry_btn = ttk.Button(dialog, text="AI 重新识别", command=retry_ai, width=12)
        retry_btn.grid(row=2, column=2, padx=5, pady=5)

        def save():
            new_code = new_code_var.get().strip()
            if not new_code:
                messagebox.showwarning("警告", "科目编码不能为空！")
                return
            
            match_items = _parse_match_items()
            result = self.base_data_mgr.update_cached_recognition(record_data["id"], new_code, match_items)
            if result["success"]:
                messagebox.showinfo("成功", result["message"])
                dialog.destroy()
                refresh_callback()
            else:
                messagebox.showerror("错误", result["message"])

        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=4, column=0, columnspan=3, pady=10)
        ttk.Button(button_frame, text="保存", command=save, width=10).pack(side="left", padx=5)
        ttk.Button(button_frame, text="取消", command=dialog.destroy, width=10).pack(side="left", padx=5)


    def _set_as_default_base_data(self):
        """基础数据管理页已取消默认值设定"""
        messagebox.showinfo("提示", "基础数据管理页已取消“设为默认值”。\n请在转换前参数确认或“设置默认值”中维护。")

    def _set_default_partner_via_subject(self):
        """默认往来单位功能已取消"""
        messagebox.showinfo("提示", "默认往来单位功能已取消。")

    # ========== 基础数据管理菜单功能 ==========
    def import_base_data(self):
        """导入基础数据"""
        if not self.base_data_mgr:
            messagebox.showerror("错误", "基础数据管理器未初始化")
            return

        # 询问是否确认导入
        confirm = messagebox.askyesno(
            "确认导入",
            "将从 '基础数据/基础数据' 目录导入所有基础数据。\n"
            "这将覆盖现有数据，是否继续？"
        )
        if not confirm:
            return

        # 显示进度提示
        progress_window = tk.Toplevel(self.root)
        progress_window.title("导入中")
        progress_window.geometry("300x100")
        progress_label = ttk.Label(progress_window, text="正在导入基础数据，请稍候...")
        progress_label.pack(expand=True)
        progress_window.update()

        try:
            result = self.base_data_mgr.import_all_data()
            progress_window.destroy()
            self._invalidate_bank_account_cache()

            # 显示详细结果
            details_text = "\n".join([
                f"{file}: {info['message']}"
                for file, info in result["details"].items()
            ])

            messagebox.showinfo(
                "导入结果",
                f"{result['message']}\n\n详细信息：\n{details_text}"
            )

        except Exception as e:
            progress_window.destroy()
            messagebox.showerror("错误", f"导入失败：\n{e}")

    def export_base_data_templates(self):
        """导出基础数据模板"""
        if not self.base_data_mgr:
            messagebox.showerror("错误", "基础数据管理器未初始化")
            return

        out_dir = filedialog.askdirectory(title="选择模板导出目录")
        if not out_dir:
            return

        try:
            result = self.base_data_mgr.export_base_templates(out_dir)
            details_text = "\n".join([
                f"{file}: {info['message']}"
                for file, info in result["details"].items()
            ])
            messagebox.showinfo(
                "导出结果",
                f"{result['message']}\n\n详细信息：\n{details_text}"
            )
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：\n{e}")

    def import_base_data_batch(self):
        """批量导入基础数据Excel"""
        if not self.base_data_mgr:
            messagebox.showerror("错误", "基础数据管理器未初始化")
            return

        folder = filedialog.askdirectory(title="选择基础数据Excel目录")
        if not folder:
            return

        # 显示进度提示
        progress_window = tk.Toplevel(self.root)
        progress_window.title("导入中")
        progress_window.geometry("320x120")
        progress_label = ttk.Label(progress_window, text="正在批量导入基础数据，请稍候...")
        progress_label.pack(expand=True)
        progress_window.update()

        try:
            file_paths = []
            for name in os.listdir(folder):
                if name.lower().endswith((".xlsx", ".xls")):
                    file_paths.append(os.path.join(folder, name))
            if not file_paths:
                progress_window.destroy()
                messagebox.showwarning("提示", "目录内未找到Excel文件")
                return

            result = self.base_data_mgr.import_batch_files(file_paths)
            progress_window.destroy()
            self._invalidate_bank_account_cache()
            details_text = "\n".join([
                f"{file}: {info.get('message', '')}"
                for file, info in result["details"].items()
            ])
            messagebox.showinfo(
                "批量导入结果",
                f"{result['message']}\n\n详细信息：\n{details_text}"
            )
            self._load_base_data_table()
        except Exception as e:
            progress_window.destroy()
            messagebox.showerror("错误", f"批量导入失败：\n{e}")

    def import_ai_training_data(self):
        """导入历史凭证Excel到AI训练缓存"""
        if not self.base_data_mgr:
            messagebox.showerror("错误", "基础数据管理器未初始化")
            return

        path = filedialog.askopenfilename(
            title="选择历史凭证 Excel（包含摘要/科目列）",
            filetypes=[("Excel 文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
        )
        if not path:
            return

        result = self.base_data_mgr.import_training_cache(path)
        msg = result.get("message", "完成")
        messagebox.showinfo("完成", msg)

    def show_base_data_stats(self):
        """显示基础数据统计信息"""
        if not self.base_data_mgr:
            messagebox.showerror("错误", "基础数据管理器未初始化")
            return

        try:
            stats = self.base_data_mgr.get_statistics()

            # 表名中文映射
            table_names = {
                "currency": "币种",
                "department": "部门",
                "warehouse": "仓库",
                "account_subject": "科目编码",
                "product": "品目信息",
                "business_partner": "往来单位",
                "bank_account": "账户"
            }

            stats_text = "\n".join([
                f"{table_names.get(table, table)}: {count} 条"
                for table, count in stats.items()
            ])

            total = sum(stats.values())

            messagebox.showinfo(
                "基础数据统计",
                f"数据库文件: {self.base_data_mgr.db_path}\n\n"
                f"{stats_text}\n\n"
                f"总计: {total} 条记录"
            )

        except Exception as e:
            messagebox.showerror("错误", f"获取统计信息失败：\n{e}")

    def refresh_recognition_cache(self):
        """刷新智能识别缓存（删除/修改基础数据后需要刷新）"""
        if not self.summary_recognizer:
            messagebox.showerror("错误", "智能识别器未初始化")
            return

        try:
            result = self.summary_recognizer.refresh_cache()

            old = result["old"]
            new = result["new"]

            message = (
                "缓存刷新完成！\n\n"
                f"往来单位: {old['partners']} -> {new['partners']}\n"
                f"科目编码: {old['accounts']} -> {new['accounts']}\n"
                f"部门: {old['departments']} -> {new['departments']}\n\n"
                "提示：如果删除了基础数据，刷新后智能识别将不再匹配被删除的数据。"
            )

            messagebox.showinfo("刷新缓存", message)

        except Exception as e:
            messagebox.showerror("错误", f"刷新缓存失败：\n{e}")

    def query_base_data(self, table_name: str):
        """查询基础数据"""
        if not self.base_data_mgr:
            messagebox.showerror("错误", "基础数据管理器未初始化")
            return

        # 创建查询窗口
        query_window = tk.Toplevel(self.root)
        query_window.title(f"查询 - {table_name}")
        query_window.geometry("800x600")

        # 搜索框
        search_frame = ttk.Frame(query_window, padding=10)
        search_frame.pack(fill="x")

        ttk.Label(search_frame, text="搜索（名称）:").pack(side="left")
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side="left", padx=5)

        def do_search():
            keyword = search_var.get().strip()
            if keyword:
                results = self.base_data_mgr.search_by_name(table_name, keyword)
            else:
                results = self.base_data_mgr.query(table_name)

            # 清空树形视图
            for item in tree.get_children():
                tree.delete(item)

            # 显示结果
            if results:
                for row in results:
                    values = [row.get(col, "") for col in columns]
                    tree.insert("", "end", values=values)
                status_label.config(text=f"找到 {len(results)} 条记录")
            else:
                status_label.config(text="未找到记录")

        ttk.Button(search_frame, text="搜索", command=do_search).pack(side="left", padx=5)
        ttk.Button(search_frame, text="显示全部", command=do_search).pack(side="left")

        # 编辑按钮框
        def add_record():
            self._edit_record_dialog(table_name, columns, None, do_search)

        def edit_record():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("提示", "请先选择要编辑的记录")
                return

            # 获取选中行的数据
            item = selection[0]
            values = tree.item(item, "values")
            record_data = dict(zip(columns, values))

            self._edit_record_dialog(table_name, columns, record_data, do_search)

        def delete_record():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("提示", "请先选择要删除的记录")
                return

            # 获取选中行的ID
            item = selection[0]
            values = tree.item(item, "values")
            record_id = values[0]  # 第一列是id

            # 确认删除
            confirm = messagebox.askyesno("确认删除", f"确定要删除ID为 {record_id} 的记录吗？")
            if not confirm:
                return

            result = self.base_data_mgr.delete_record(table_name, int(record_id))
            if result["success"]:
                if table_name == "bank_account":
                    self._invalidate_bank_account_cache()
                # 自动刷新智能识别缓存
                if self.summary_recognizer:
                    self.summary_recognizer.refresh_cache()
                messagebox.showinfo("成功", result["message"] + "\n\n已自动刷新智能识别缓存。")
                do_search()  # 刷新列表
            else:
                messagebox.showerror("错误", result["message"])

        edit_button_frame = ttk.Frame(search_frame)
        edit_button_frame.pack(side="right")

        ttk.Button(edit_button_frame, text="新增", command=add_record).pack(side="left", padx=2)
        ttk.Button(edit_button_frame, text="编辑", command=edit_record).pack(side="left", padx=2)
        ttk.Button(edit_button_frame, text="删除", command=delete_record).pack(side="left", padx=2)

        # 结果显示区域
        result_frame = ttk.Frame(query_window)
        result_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # 创建滚动条
        scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
        scrollbar_y.pack(side="right", fill="y")

        scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")
        scrollbar_x.pack(side="bottom", fill="x")

        # 获取列信息
        sample_data = self.base_data_mgr.query(table_name)
        if sample_data:
            columns = list(sample_data[0].keys())
        else:
            columns = ["id", "code", "name"]

        # 创建树形视图
        tree = ttk.Treeview(
            result_frame,
            columns=columns,
            show="headings",
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )

        scrollbar_y.config(command=tree.yview)
        scrollbar_x.config(command=tree.xview)

        # 设置列标题
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)

        tree.pack(fill="both", expand=True)
        attach_treeview_tools(tree)

        # 状态栏
        status_label = ttk.Label(query_window, text="", relief="sunken")
        status_label.pack(fill="x", side="bottom")

        # 自动加载全部数据
        do_search()

    def _edit_record_dialog(self, table_name: str, columns: List[str], record_data: Optional[Dict], refresh_callback):
        """编辑/新增记录对话框"""
        is_new = record_data is None

        # 创建对话框
        dialog = tk.Toplevel(self.root)
        dialog.title("新增记录" if is_new else "编辑记录")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()

        # 字段中文名映射
        field_labels = {
            "id": "ID",
            "code": "编码",
            "name": "名称",
            "exchange_rate": "汇率",
            "use_type": "使用类型",
            "is_active": "是否启用",
            "type": "类型",
            "production_process": "生产流程",
            "outsource_partner": "外包单位",
            "code_name": "科目编码名称",
            "is_subject": "是否科目",
            "debit_credit_type": "借贷类型",
            "subject_type": "科目类型",
            "contra_type": "备抵类型",
            "summary": "摘要",
            "parent_subject": "上级科目",
            "display_name": "显示名称",
            "product_type": "品目类型",
            "spec_info": "规格信息",
            "unit": "单位",
            "search_keyword": "搜索关键词",
            "pack_qty": "装数",
            "unit_conversion_denominator": "单位转换分母",
            "unit_conversion_numerator": "单位转换分子",
            "specification": "规格",
            "in_price": "入库单价",
            "out_price": "出库单价",
            "price_a": "单价A",
            "price_b": "单价B",
            "price_c": "单价C",
            "length": "长度",
            "width": "宽度",
            "height": "高度",
            "volume": "体积",
            "weight": "重量",
            "color": "颜色",
            "size_range": "尺码范围",
            "contact_person": "联系人",
            "mobile": "手机",
            "phone": "电话",
            "email": "邮箱",
            "category": "分类",
            "file_management": "文件管理",
            "tax_number": "税号",
            "bank_name": "开户行",
            "bank_account": "银行账号",
            "account_subject": "会计科目",
            "foreign_currency": "外币存折",
            "match_items": "映射匹配项",
            "local_code": "当地系统编码",
        }

        # 创建滚动框架
        canvas = tk.Canvas(dialog)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 输入字段
        entry_vars = {}
        text_widgets = {}
        row = 0

        for col in columns:
            if col == "id" and is_new:
                continue  # 新增时跳过ID字段

            label_text = field_labels.get(col, col)
            ttk.Label(scrollable_frame, text=f"{label_text}:").grid(
                row=row, column=0, sticky="e", padx=5, pady=5
            )

            # 映射匹配项使用多行输入
            if col == "match_items":
                preset = ""
                if record_data and col in record_data and record_data[col]:
                    raw_val = record_data[col]
                    try:
                        parsed = json.loads(raw_val) if isinstance(raw_val, str) else (raw_val or [])
                        if isinstance(parsed, list):
                            preset = "\n".join(str(i).strip() for i in parsed if str(i).strip())
                        elif parsed:
                            preset = str(parsed)
                    except Exception:
                        preset = str(raw_val)

                text = tk.Text(scrollable_frame, height=3, width=40, wrap="word")
                text.grid(row=row, column=1, sticky="we", padx=5, pady=5)
                if preset:
                    text.insert("1.0", preset)
                text_widgets[col] = text
                row += 1
                continue

            var = tk.StringVar()
            if record_data and col in record_data:
                value = record_data[col]
                var.set(str(value) if value is not None else "")

            entry = ttk.Entry(scrollable_frame, textvariable=var, width=40)
            entry.grid(row=row, column=1, sticky="we", padx=5, pady=5)

            if col == "id" and not is_new:
                entry.config(state="readonly")  # ID字段只读

            entry_vars[col] = var
            row += 1

        scrollable_frame.grid_columnconfigure(1, weight=1)

        # 布局滚动框架
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

        # 按钮框
        button_frame = ttk.Frame(dialog)
        button_frame.pack(fill="x", padx=10, pady=10)

        def save_record():
            # 收集数据
            data = {}
            for col, var in entry_vars.items():
                value = var.get().strip()
                # 空字符串转为None，数字字段尝试转换
                if value == "":
                    data[col] = None
                else:
                    # 尝试转换数字字段
                    if col in ["exchange_rate", "pack_qty", "unit_conversion_denominator",
                               "unit_conversion_numerator", "in_price", "out_price",
                               "price_a", "price_b", "price_c", "length", "width",
                               "height", "volume", "weight"]:
                        try:
                            data[col] = float(value) if value else None
                        except ValueError:
                            data[col] = value
                    else:
                        data[col] = value

            # 处理 match_items
            for col, widget in text_widgets.items():
                raw = widget.get("1.0", "end").strip()
                if not raw:
                    data[col] = "[]"
                else:
                    parts = re.split(r"[\\n,，;；、]+", raw)
                    cleaned = [p.strip() for p in parts if p and p.strip()]
                    data[col] = json.dumps(cleaned, ensure_ascii=False)

            # 保存
            if is_new:
                result = self.base_data_mgr.add_record(table_name, data)
            else:
                if "id" in entry_vars:
                    record_id = int(entry_vars["id"].get())
                    result = self.base_data_mgr.update_record(table_name, record_id, data)
                else:
                    messagebox.showerror("错误", "无法定位记录ID，请重试。")
                    return

            if result["success"]:
                if table_name == "bank_account":
                    self._invalidate_bank_account_cache()
                messagebox.showinfo("成功", result["message"])
                dialog.destroy()
                refresh_callback()  # 刷新列表
            else:
                messagebox.showerror("错误", result["message"])

        ttk.Button(button_frame, text="保存", command=save_record).pack(side="right", padx=5)
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side="right")


    def _build_batch_merge_tab(self):
        """构建批量合并标签页"""
        if ExcelMergerGUI:
            merge_frame = ttk.Frame(self.notebook)
            self.notebook.add(merge_frame, text="批量合并")
            
            # Embed the ExcelMergerGUI into this frame
            # 修复：必须保存实例引用，防止垃圾回收导致潜在不稳定
            self.excel_merger = ExcelMergerGUI(merge_frame, open_in_converter=self._open_in_converter)
        else:
            merge_frame = ttk.Frame(self.notebook)
            self.notebook.add(merge_frame, text="批量合并")
            ttk.Label(merge_frame, text="无法加载 ExcelMergerGUI 模块，请检查 excel_merger.py 是否存在。", foreground="red").pack(padx=20, pady=20)


    def _build_shipping_tab(self):
        """构建报关清单汇总标签页"""
        shipping_frame = ttk.Frame(self.notebook)
        self.notebook.add(shipping_frame, text="报关清单汇总")
        # 修复：必须保存实例引用
        self.shipping_module = ShippingModule(shipping_frame, db_path="shipping.bd", open_in_converter=self._open_in_converter)
        self.shipping_module.pack(fill="both", expand=True)

    def _build_report_tab(self):
        """构建经营报告生成标签页"""
        report_frame = ttk.Frame(self.notebook)
        self.notebook.add(report_frame, text="经营报告")

        # Top: Settings
        top_frame = ttk.LabelFrame(report_frame, text="报告设置", padding=10)
        top_frame.pack(fill="x", padx=10, pady=10)

        # Base Data Dir
        ttk.Label(top_frame, text="基础资料目录:").grid(row=0, column=0, sticky="e", pady=5)
        self.report_base_dir_var = tk.StringVar(value=os.path.join(os.getcwd(), "基础资料"))
        ttk.Entry(top_frame, textvariable=self.report_base_dir_var, width=50).grid(row=0, column=1, columnspan=3, sticky="ew", padx=5)
        ttk.Button(top_frame, text="浏览...", command=self._select_report_base_dir).grid(row=0, column=4)

        # Template Path
        ttk.Label(top_frame, text="报告模板文件:").grid(row=1, column=0, sticky="e", pady=5)
        default_template = "11月汇总结果_整理优化美化_含仪表盘_目标预算达成异常 (1).xlsx"
        self.report_template_var = tk.StringVar(value=os.path.join(os.getcwd(), default_template))
        ttk.Entry(top_frame, textvariable=self.report_template_var, width=50).grid(row=1, column=1, columnspan=3, sticky="ew", padx=5)
        ttk.Button(top_frame, text="浏览...", command=lambda: self._select_file(self.report_template_var)).grid(row=1, column=4)

        # Target Year/Month Selection
        ttk.Label(top_frame, text="目标月份:").grid(row=2, column=0, sticky="e", pady=5)
        
        current_year = datetime.now().year
        self.report_year_var = tk.StringVar(value=str(current_year))
        self.report_year_cb = ttk.Combobox(top_frame, textvariable=self.report_year_var, values=[str(y) for y in range(current_year-2, current_year+3)], width=6, state="normal")
        self.report_year_cb.grid(row=2, column=1, sticky="w", padx=5)
        ttk.Label(top_frame, text="年").grid(row=2, column=2, sticky="w")
        
        current_month = datetime.now().month
        self.report_month_var = tk.StringVar(value=str(current_month))
        self.report_month_cb = ttk.Combobox(top_frame, textvariable=self.report_month_var, values=[str(m) for m in range(1, 13)], width=4, state="normal")
        self.report_month_cb.grid(row=2, column=3, sticky="w", padx=5)
        ttk.Label(top_frame, text="月").grid(row=2, column=4, sticky="w")

        ttk.Label(top_frame, text="年份显示:").grid(row=3, column=0, sticky="e", pady=5)
        scope_values = ["仅当前年份", "显示历年"]
        self.report_year_scope_var = tk.StringVar(value=scope_values[0])
        scope_cb = ttk.Combobox(top_frame, textvariable=self.report_year_scope_var, values=scope_values, width=12, state="readonly")
        scope_cb.grid(row=3, column=1, columnspan=3, sticky="w", padx=5)

        self.report_ai_analysis_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            top_frame,
            text="生成后AI分析并写入表格",
            variable=self.report_ai_analysis_var,
        ).grid(row=4, column=1, columnspan=3, sticky="w", padx=5, pady=5)
        self.report_ai_chart_recognition_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            top_frame,
            text="AI模式启用图表识别",
            variable=self.report_ai_chart_recognition_var,
        ).grid(row=5, column=1, columnspan=3, sticky="w", padx=5, pady=2)

        top_frame.columnconfigure(1, weight=0)
        top_frame.columnconfigure(2, weight=0)
        top_frame.columnconfigure(3, weight=0)

        self._report_years_dir = None
        self.report_base_dir_var.trace_add("write", lambda *args: self._refresh_report_year_options())
        self.report_year_cb.bind("<<ComboboxSelected>>", lambda e: self._refresh_report_month_options())
        self._refresh_report_year_options()

        # Alert Parameters
        param_frame = ttk.LabelFrame(report_frame, text="预警参数", padding=10)
        param_frame.pack(fill="x", padx=10, pady=(0, 10))

        self.report_replenish_lead_var = tk.StringVar(value="30")
        self.report_replenish_safety_var = tk.StringVar(value="20")
        self.report_replenish_window_var = tk.StringVar(value="3")
        self.report_cash_dso_var = tk.StringVar(value="90")
        self.report_cash_dio_var = tk.StringVar(value="180")
        self.report_cash_ccc_var = tk.StringVar(value="120")
        self.report_cash_cover_var = tk.StringVar(value="1.5")

        ttk.Label(param_frame, text="补货：采购周期(天)").grid(row=0, column=0, sticky="e", pady=2)
        ttk.Entry(param_frame, textvariable=self.report_replenish_lead_var, width=6).grid(row=0, column=1, sticky="w", padx=5)
        ttk.Label(param_frame, text="安全库存(天)").grid(row=0, column=2, sticky="e", pady=2)
        ttk.Entry(param_frame, textvariable=self.report_replenish_safety_var, width=6).grid(row=0, column=3, sticky="w", padx=5)
        ttk.Label(param_frame, text="销量窗口(月)").grid(row=0, column=4, sticky="e", pady=2)
        ttk.Entry(param_frame, textvariable=self.report_replenish_window_var, width=6).grid(row=0, column=5, sticky="w", padx=5)

        ttk.Label(param_frame, text="资金链：DSO阈值(天)").grid(row=1, column=0, sticky="e", pady=2)
        ttk.Entry(param_frame, textvariable=self.report_cash_dso_var, width=6).grid(row=1, column=1, sticky="w", padx=5)
        ttk.Label(param_frame, text="DIO阈值(天)").grid(row=1, column=2, sticky="e", pady=2)
        ttk.Entry(param_frame, textvariable=self.report_cash_dio_var, width=6).grid(row=1, column=3, sticky="w", padx=5)
        ttk.Label(param_frame, text="CCC阈值(天)").grid(row=1, column=4, sticky="e", pady=2)
        ttk.Entry(param_frame, textvariable=self.report_cash_ccc_var, width=6).grid(row=1, column=5, sticky="w", padx=5)

        ttk.Label(param_frame, text="现金覆盖(月)").grid(row=2, column=0, sticky="e", pady=2)
        ttk.Entry(param_frame, textvariable=self.report_cash_cover_var, width=6).grid(row=2, column=1, sticky="w", padx=5)

        for col in range(6):
            param_frame.columnconfigure(col, weight=0)

        # Middle: Actions
        btn_frame = ttk.Frame(report_frame, padding=10)
        btn_frame.pack(fill="x")

        ttk.Button(btn_frame, text="开始生成报告", command=self._generate_business_report, width=20).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="生成并转入凭证转换", command=lambda: self._generate_business_report(True), width=22).pack(side="left", padx=10)

        # Bottom: Log (Using a separate log for this tab or shared? Separate is better for context)
        log_frame = ttk.LabelFrame(report_frame, text="生成日志", padding=10)
        log_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.report_log_text = tk.Text(log_frame, height=15, width=80)
        self.report_log_text.pack(fill="both", expand=True)

    def _select_directory(self, var):
        d = filedialog.askdirectory()
        if d:
            var.set(d)

    def _select_report_base_dir(self):
        d = filedialog.askdirectory()
        if d:
            self.report_base_dir_var.set(d)
            self._refresh_report_year_options()

    def _refresh_report_year_options(self):
        if not hasattr(self, "report_year_cb"):
            return
        base_dir = self.report_base_dir_var.get().strip()
        if not base_dir or not os.path.isdir(base_dir):
            return
        if self._report_years_dir == base_dir:
            return

        self._report_years_dir = base_dir
        years = []
        generator = None
        if ReportGenerator:
            try:
                generator = ReportGenerator(base_dir)
                years = generator.list_available_years()
            except Exception:
                years = []

        if not years:
            current_year = datetime.now().year
            years = list(range(current_year - 2, current_year + 3))

        year_values = [str(y) for y in years]
        current_year_value = self.report_year_var.get().strip()
        if current_year_value and current_year_value not in year_values:
            if current_year_value.isdigit() and len(current_year_value) == 4:
                year_values.append(current_year_value)
        self.report_year_cb["values"] = year_values

        if not current_year_value and year_values:
            self.report_year_var.set(year_values[-1])

        self._refresh_report_month_options(generator=generator)

    def _refresh_report_month_options(self, generator=None):
        if not hasattr(self, "report_month_cb"):
            return
        base_dir = self.report_base_dir_var.get().strip()
        if not base_dir or not os.path.isdir(base_dir):
            return
        if generator is None and ReportGenerator:
            try:
                generator = ReportGenerator(base_dir)
            except Exception:
                generator = None
        months = []
        if generator:
            try:
                months = generator.list_available_months()
            except Exception:
                months = []

        target_year = self.report_year_var.get().strip()
        month_values = []
        if target_year and months:
            for m in months:
                if isinstance(m, str) and m.startswith(f"{target_year}-"):
                    try:
                        month_values.append(str(int(m[-2:])))
                    except Exception:
                        continue
        if not month_values:
            month_values = [str(m) for m in range(1, 13)]

        current_month_value = self.report_month_var.get().strip()
        if current_month_value and current_month_value not in month_values:
            if current_month_value.isdigit():
                month_num = int(current_month_value)
                if 1 <= month_num <= 12:
                    month_values.append(str(month_num))

        self.report_month_cb["values"] = month_values
        if not current_month_value and month_values:
            self.report_month_var.set(month_values[-1])

    def _log_report(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.report_log_text.insert("end", f"[{ts}] {msg}\n")
        self.report_log_text.see("end")
        self.root.update_idletasks()

    def _generate_business_report(self, open_in_converter=False):
        if not ReportGenerator:
            messagebox.showerror("错误", "ReportGenerator 模块加载失败，请检查 report_generator.py")
            return

        def _parse_int(value, default):
            try:
                return int(float(str(value).strip()))
            except Exception:
                return default

        def _parse_float(value, default):
            try:
                return float(str(value).strip())
            except Exception:
                return default

        base_dir = self.report_base_dir_var.get()
        template_path = self.report_template_var.get()
        target_year = self.report_year_var.get()
        target_month = self.report_month_var.get()
        scope_label = self.report_year_scope_var.get()
        year_scope = "all" if "历年" in scope_label else "current"

        replenishment_params = {
            "lead_days": _parse_int(self.report_replenish_lead_var.get(), 30),
            "safety_days": _parse_int(self.report_replenish_safety_var.get(), 20),
            "window_months": _parse_int(self.report_replenish_window_var.get(), 3),
        }
        cashflow_params = {
            "dso_threshold": _parse_int(self.report_cash_dso_var.get(), 90),
            "dio_threshold": _parse_int(self.report_cash_dio_var.get(), 180),
            "ccc_threshold": _parse_int(self.report_cash_ccc_var.get(), 120),
            "cash_coverage_threshold": _parse_float(self.report_cash_cover_var.get(), 1.5),
        }

        if not os.path.exists(base_dir):
            messagebox.showwarning("提示", f"基础资料目录不存在: {base_dir}")
            return
        if not os.path.exists(template_path):
            messagebox.showwarning("提示", f"模板文件不存在: {template_path}")
            return

        # Redirect stdout to my log
        import sys
        import io
        
        class StdoutRedirector(io.StringIO):
            def __init__(self, text_widget, tk_root):
                super().__init__()
                self.text_widget = text_widget
                self.tk_root = tk_root
            def write(self, string):
                if string.strip():
                    self.text_widget.insert("end", string + "\n")
                    self.text_widget.see("end")
                    self.tk_root.update_idletasks()
                super().write(string)

        original_stdout = sys.stdout
        sys.stdout = StdoutRedirector(self.report_log_text, self.root)

        try:
            self._log_report("开始初始化报告生成器...")
            generator = ReportGenerator(base_dir)
            
            self._log_report("正在加载基础数据...")
            generator.load_all_data()

            quality_summary = (
                generator.get_data_quality_summary()
                if hasattr(generator, "get_data_quality_summary")
                else {"ERROR": 0, "WARN": 0, "INFO": 0, "TOTAL": 0}
            )
            if quality_summary.get("ERROR", 0) > 0:
                dq_msg = (
                    f"检测到数据质量问题：ERROR={quality_summary.get('ERROR', 0)}，"
                    f"WARN={quality_summary.get('WARN', 0)}，INFO={quality_summary.get('INFO', 0)}。\n"
                    "继续生成可能导致报表口径异常。"
                )
                self._log_report(dq_msg)
                proceed = messagebox.askyesno("数据质量风险", dq_msg + "\n\n是否继续生成报告？")
                if not proceed:
                    self._log_report("用户取消操作。")
                    return
            elif quality_summary.get("WARN", 0) > 0:
                self._log_report(
                    f"数据质量检查提示：WARN={quality_summary.get('WARN', 0)}，INFO={quality_summary.get('INFO', 0)}。"
                )
             
            # Check data completeness
            self._log_report(f"正在检查 {target_year}年{target_month}月 的数据完整性...")
            missing = generator.check_data_completeness(target_year, target_month)
            
            if missing:
                msg = f"警告：{target_year}-{int(target_month):02d} 缺少以下关键数据：\n" + ", ".join([m.upper() for m in missing])
                self._log_report(msg)
                proceed = messagebox.askyesno("数据缺失", msg + "\n\n是否继续生成报告？")
                if not proceed:
                    self._log_report("用户取消操作。")
                    return
            else:
                self._log_report("✅ 数据检查通过，所需资料齐全。")

            output_filename = f"{target_year}年{int(target_month):02d}月_经营分析报告.xlsx"
            output_path = os.path.join(os.path.dirname(template_path), output_filename)
            
            self._log_report(f"正在生成报告到: {output_path}")
            success = generator.generate_report(
                template_path,
                output_path,
                target_year,
                target_month,
                year_scope=year_scope,
                replenishment_params=replenishment_params,
                cashflow_params=cashflow_params,
                include_ai_placeholders=self.report_ai_analysis_var.get(),
                fail_on_validation_error=True,
                fail_on_data_quality_error=False,
                allow_generated_report_template=False,
            )
            
            if success:
                self._log_report("✅ 报告生成成功！")
                
                if self.report_ai_analysis_var.get():
                    try:
                        self._log_report("启动 AI 分析并写入表格...")
                        try:
                            from local_llm_analyzer import LocalLLMAnalyzer
                        except ImportError:
                            self._log_report("错误: 未找到 local_llm_analyzer 模块。")
                            messagebox.showerror("错误", "未找到 AI 分析模块 (local_llm_analyzer)。")
                            sys.stdout = original_stdout
                            return

                        backend = self._normalize_ai_backend(
                            self._get_ai_backend_for_task("report_analysis", allow_legacy=True)
                        )
                        analyzer = LocalLLMAnalyzer(
                            api_base=backend["base_url"],
                            model=backend["model"],
                            api_key=backend["api_key"],
                            provider=backend["provider"],
                            base_data_dir=base_dir,
                            enable_chart_recognition=self.report_ai_chart_recognition_var.get(),
                        )
                        analysis_output = output_path.replace(".xlsx", "_AI_Analysis.md")
                        analyzer.analyze_report(output_path, analysis_output, embed_to_excel=True)
                        self._log_report("AI 分析已写入报告各表。")
                        self._log_report(f"AI 分析报告已生成: {analysis_output}")
                        messagebox.showinfo(
                            "AI 分析完成",
                            f"AI 解读已写入报告：\n{output_path}\n\n分析报告: {analysis_output}",
                        )
                    except Exception as e:
                        self._log_report(f"AI 分析失败: {e}")
                        messagebox.showerror("AI 分析失败", str(e))
                else:
                    messagebox.showinfo("成功", f"报告已生成:\n{output_path}")
                if open_in_converter:
                    self._open_in_converter(output_path)
            else:
                self._log_report("❌ 报告生成失败，请检查日志。")
                messagebox.showerror("失败", "报告生成过程中出现错误")
                
        except Exception as e:
            self._log_report(f"发生异常: {e}")
            import traceback
            self._log_report(traceback.format_exc())
            messagebox.showerror("错误", f"发生异常: {e}")
        finally:
            sys.stdout = original_stdout


def main():
    root = tk.Tk()
    app = ExcelConverterGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
