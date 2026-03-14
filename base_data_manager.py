# -*- coding: utf-8 -*-
"""
基础数据管理模块
用于导入、存储和查询基础数据（币种、部门、仓库、科目编码、品目信息、往来单位、账户）
"""

import os
import sqlite3
import pandas as pd
import difflib
import json
import hashlib
from typing import Optional, List, Dict, Any


class BaseDataManager:
    """基础数据管理器"""

    APP_DIR = os.path.dirname(os.path.abspath(__file__))
    DB_FILE = os.path.join(APP_DIR, "base_data.db")
    BASE_DATA_DIR = os.path.join(APP_DIR, "基础数据", "基础数据")

    # 定义基础数据文件和对应的表名
    DATA_FILES = {
        "币种.xlsx": "currency",
        "部门.xlsx": "department",
        "仓库xlsx.xlsx": "warehouse",
        "科目编码.xlsx": "account_subject",
        "品目信息.xlsx": "product",
        "往来单位.xlsx": "business_partner",
        "账户xlsx.xlsx": "bank_account",
    }
    ALLOWED_TABLES = set(DATA_FILES.values()) | {
        "import_log",
        "smart_recognition_cache",
        "app_config",
        "mapping_schemes",
        "auto_mapping_cache",
        "recognition_rules",
    }

    def __init__(self, db_path: Optional[str] = None):
        """初始化数据库连接"""
        self.db_path = db_path or self.DB_FILE
        self.conn = None
        # 本地缓存：标准化后的摘要 -> 科目编码（含别名）
        self._cache_lookup = {}
        self._init_database()

    def _init_database(self):
        """初始化数据库结构"""
        self.conn = sqlite3.connect(self.db_path)
        self.conn.execute("PRAGMA encoding = 'UTF-8'")
        cursor = self.conn.cursor()

        # 创建币种表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS currency (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE,
                name TEXT,
                exchange_rate REAL,
                use_type TEXT
            )
        """)

        # 创建部门表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS department (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE,
                name TEXT,
                is_active TEXT
            )
        """)

        # 创建仓库表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS warehouse (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE,
                name TEXT,
                type TEXT,
                production_process TEXT,
                outsource_partner TEXT
            )
        """)

        # 创建科目编码表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS account_subject (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code_name TEXT,
                is_subject TEXT,
                debit_credit_type TEXT,
                subject_type TEXT,
                contra_type TEXT,
                use_type TEXT,
                summary TEXT,
                parent_subject TEXT,
                display_name TEXT,
                match_items TEXT    -- 额外映射匹配项（JSON 数组，支持多对一）
            )
        """)

        # 创建品目信息表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS product (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE,
                name TEXT,
                product_type TEXT,
                spec_info TEXT,
                unit TEXT,
                search_keyword TEXT,
                pack_qty REAL,
                unit_conversion_denominator REAL,
                unit_conversion_numerator REAL,
                specification TEXT,
                in_price REAL,
                out_price REAL,
                price_a REAL,
                price_b REAL,
                price_c REAL,
                length REAL,
                width REAL,
                height REAL,
                volume REAL,
                weight REAL,
                color TEXT,
                size_range TEXT,
                match_items TEXT    -- 额外映射匹配项（JSON 数组，支持多对一）
            )
        """)

        # 创建往来单位表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS business_partner (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE,
                name TEXT,
                contact_person TEXT,
                mobile TEXT,
                phone TEXT,
                email TEXT,
                category TEXT,
                file_management TEXT,
                tax_number TEXT,
                bank_name TEXT,
                bank_account TEXT,
                local_code TEXT,    -- 当地系统编码 (用于对账)
                match_items TEXT    -- 额外映射匹配项（JSON 数组，支持多对一）
            )
        """)

        # 创建账户表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS bank_account (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE,
                name TEXT,
                account_subject TEXT,
                search_keyword TEXT,
                summary TEXT,
                foreign_currency TEXT,
                is_active TEXT
            )
        """)

        # 自定义基础数据分类与记录
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS custom_category (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE,
                display_name TEXT,
                fields_json TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS custom_record (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_name TEXT,
                code TEXT,
                name TEXT,
                data_json TEXT
            )
        """)

        # 创建导入记录表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS import_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                table_name TEXT,
                file_name TEXT,
                import_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                record_count INTEGER,
                status TEXT
            )
        """)

        # 创建智能识别缓存表 (新增)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS smart_recognition_cache (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                summary TEXT UNIQUE, -- 摘要内容作为唯一键
                account_code TEXT,   -- 识别出的科目编码
                match_items TEXT,    -- 额外映射匹配项（JSON 数组，支持多对一）
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # 创建应用配置表 (新增，用于持久化默认值)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS app_config (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)

        # 创建映射方案表 (新增)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS mapping_schemes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE,
                base_mode TEXT,
                template_path TEXT,
                mapping_json TEXT,
                composite_json TEXT DEFAULT '{}',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # 创建自动映射缓存表 (新增)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS auto_mapping_cache (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_path TEXT,
                base_mode TEXT,
                source_signature TEXT,
                source_columns_json TEXT,
                mapping_json TEXT,
                composite_json TEXT DEFAULT '{}',
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(template_path, base_mode, source_signature)
            )
        """)

        # 创建自定义识别规则表 (新增)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS recognition_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rule_type TEXT,       -- 规则类型: business(业务), account(科目), department(部门)
                name TEXT,            -- 规则名称
                keywords TEXT,        -- 关键词列表 (JSON字符串)
                account_code TEXT,    -- 对应的科目编码
                transaction_type TEXT, -- 借贷类型 (1/2)
                summary_code TEXT,    -- 摘要编码
                dept_code TEXT,       -- 部门编码
                priority INTEGER DEFAULT 0 -- 优先级
            )
        """)

        self.conn.commit()

    def get_table_columns(self, table_name: str) -> List[str]:
        """获取表字段（排除id）"""
        if self._is_custom_table(table_name):
            custom_name = self._get_custom_name(table_name)
            fields = self.get_custom_category_fields(custom_name)
            return ["code", "name"] + fields
        self._assert_valid_table(table_name)
        cursor = self.conn.cursor()
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = []
        for _, name, *_ in cursor.fetchall():
            if name == "id":
                continue
            columns.append(name)
        return columns

    def export_base_templates(self, output_dir: str) -> Dict[str, Any]:
        """导出基础数据Excel模板（含首行说明）"""
        if not output_dir:
            return {"success": False, "message": "输出目录为空"}
        os.makedirs(output_dir, exist_ok=True)
        results = {}
        for file_name, table_name in self.DATA_FILES.items():
            try:
                columns = self.get_table_columns(table_name)
                df = pd.DataFrame(columns=columns)
                file_path = os.path.join(output_dir, file_name)
                df.to_excel(file_path, index=False)
                # 插入首行说明，保证与导入逻辑一致（header=1）
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    ws.insert_rows(1)
                    ws.cell(row=1, column=1).value = "公司名称 : （可选）"
                    wb.save(file_path)
                except Exception:
                    pass
                results[file_name] = {"success": True, "message": f"已导出: {file_path}"}
            except Exception as e:
                results[file_name] = {"success": False, "message": str(e)}
        # 导出自定义分类模板
        try:
            custom_categories = self.list_custom_categories()
        except Exception:
            custom_categories = []
        for cat in custom_categories:
            try:
                display = cat.get("display_name") or cat.get("name")
                name_key = cat.get("name")
                if not display or not name_key:
                    continue
                columns = ["code", "name"] + [f.get("name") for f in cat.get("fields", []) if f.get("name")]
                df = pd.DataFrame(columns=columns)
                file_name = f"自定义_{display}.xlsx"
                file_path = os.path.join(output_dir, file_name)
                df.to_excel(file_path, index=False)
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    ws.insert_rows(1)
                    ws.cell(row=1, column=1).value = "公司名称 : （可选）"
                    wb.save(file_path)
                except Exception:
                    pass
                results[file_name] = {"success": True, "message": f"已导出: {file_path}"}
            except Exception as e:
                results[file_name] = {"success": False, "message": str(e)}
        success_count = sum(1 for r in results.values() if r["success"])
        total_count = len(results)
        return {
            "success": success_count == total_count,
            "message": f"已导出 {success_count}/{total_count} 个模板",
            "details": results,
        }

    def _match_table_by_filename(self, file_name: str) -> Optional[str]:
        """根据文件名匹配表名"""
        try:
            custom_categories = self.list_custom_categories()
        except Exception:
            custom_categories = []
        base_name = os.path.splitext(file_name)[0]
        if base_name.startswith("自定义_"):
            key = base_name.replace("自定义_", "", 1)
            for cat in custom_categories:
                if key in (cat.get("display_name"), cat.get("name")):
                    return f"custom:{cat.get('name')}"
        for name, table in self.DATA_FILES.items():
            if file_name == name:
                return table
        key_map = {
            "币种": "currency",
            "部门": "department",
            "仓库": "warehouse",
            "科目": "account_subject",
            "品目": "product",
            "往来": "business_partner",
            "账户": "bank_account",
        }
        for key, table in key_map.items():
            if key in file_name:
                return table
        for cat in custom_categories:
            display = cat.get("display_name")
            name_key = cat.get("name")
            if display and display in file_name:
                return f"custom:{name_key}"
            if name_key and name_key in file_name:
                return f"custom:{name_key}"
        return None

    def import_batch_files(self, file_paths: List[str]) -> Dict[str, Any]:
        """批量导入Excel文件（按文件名匹配表）"""
        results = {}
        for path in file_paths:
            file_name = os.path.basename(path)
            table_name = self._match_table_by_filename(file_name)
            if not table_name:
                results[file_name] = {
                    "success": False,
                    "message": "未识别到对应基础数据表，已跳过",
                    "count": 0,
                }
                continue
            result = self.import_single_file(path, table_name)
            results[file_name] = result
        success_count = sum(1 for r in results.values() if r.get("success"))
        self._ensure_mapping_scheme_composite_column()
        self._ensure_cache_extra_columns()
        self._ensure_base_match_columns()
        self._ensure_business_partner_local_code()
        self._init_default_rules()
        self._load_cache_maps()
        return {
            "success": success_count == len(results),
            "message": f"批量导入完成：{success_count}/{len(results)} 成功",
            "details": results,
        }

    def _assert_valid_table(self, table_name: str):
        """校验表名，防止 SQL 注入"""
        if table_name not in self.ALLOWED_TABLES and not self._is_custom_table(table_name):
            raise ValueError(f"不允许的表名: {table_name}")

    def _is_custom_table(self, table_name: str) -> bool:
        return isinstance(table_name, str) and table_name.startswith("custom:")

    def _get_custom_name(self, table_name: str) -> str:
        return table_name.split("custom:", 1)[1] if self._is_custom_table(table_name) else table_name

    def _ensure_business_partner_local_code(self):
        """为往来单位表补充 local_code 列"""
        cursor = self.conn.cursor()
        cursor.execute("PRAGMA table_info(business_partner)")
        columns = [row[1] for row in cursor.fetchall()]
        if "local_code" not in columns:
            try:
                cursor.execute("ALTER TABLE business_partner ADD COLUMN local_code TEXT")
                self.conn.commit()
                print("已为 business_partner 补充 local_code 列")
            except Exception as e:
                print(f"补充 local_code 列失败: {e}")

    def _ensure_mapping_scheme_composite_column(self):
        """为映射方案表补充 composite_json 列（向后兼容旧库）"""
        cursor = self.conn.cursor()
        cursor.execute("PRAGMA table_info(mapping_schemes)")
        columns = [row[1] for row in cursor.fetchall()]
        if "composite_json" not in columns:
            try:
                cursor.execute("ALTER TABLE mapping_schemes ADD COLUMN composite_json TEXT DEFAULT '{}'")
                self.conn.commit()
                print("已为 mapping_schemes 补充 composite_json 列")
            except Exception as e:
                # 迁移失败会导致后续读取映射方案异常，直接抛出让调用侧感知
                raise RuntimeError(f"补充 composite_json 列失败: {e}")

    def _build_source_signature(self, source_columns: List[str]) -> str:
        normalized = [str(col).strip() for col in source_columns]
        normalized_sorted = sorted(normalized)
        payload = "\n".join(normalized_sorted).encode("utf-8")
        return hashlib.sha1(payload).hexdigest()

    def _init_default_rules(self):
        """初始化默认的识别规则（如果表为空）"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT count(*) FROM recognition_rules")
        if cursor.fetchone()[0] > 0:
            return

        print("正在初始化默认识别规则到数据库...")
        
        # 1. 业务类型规则
        business_rules = [
            ("business", "销售", '["销售", "销货", "售出", "出售", "卖出"]', "1122", "1", "01", None),
            ("business", "销售退回", '["销售退回", "退货", "退款"]', "1122", "2", "02", None),
            ("business", "采购", '["采购", "购买", "进货", "买入"]', "2202", "2", "03", None),
            ("business", "采购退回", '["采购退回", "退货"]', "2202", "1", "04", None),
            ("business", "收款", '["收到", "收款", "回款", "收汇"]', "1002", "1", "05", None),
            ("business", "付款", '["支付", "付款", "付汇", "转账"]', "1002", "2", "06", None),
            ("business", "运费", '["运费", "运输费", "物流费"]', "6401", "1", "07", None),
            ("business", "办公费", '["办公费", "办公用品", "文具"]', "6602", "1", "08", None),
            ("business", "工资", '["工资", "薪酬", "薪资"]', None, "2", "09", None), # 这里的科目设为NULL，遵循用户刚才的修改
            ("business", "入库", '["入库", "验收入库", "采购入库"]', "1405", "1", "10", None),
            ("business", "出库", '["出库", "发货", "领用"]', "1405", "2", "11", None),
        ]

        # 2. 特定科目关键词规则 (原 keyword_rules)
        # 注意: 原逻辑中有 target_names 用于查找，这里简化为直接映射 fallback 代码，
        # 如果需要更复杂的逻辑(先找名字再用代码)，可能需要更复杂的表结构，但通常直接映射代码足够。
        account_rules = [
            ("account", "现金存", '["现金存", "现金存bac", "现金存st"]', "100102", None, None, None),
            ("account", "应收兜底", '["deposito", "depósito", "cheque", "ach"]', "1122", None, None, None),
            ("account", "房租", '["住房", "房租", "租金", "租赁"]', "660102", None, None, None),
            ("account", "刷卡手续费", '["刷卡机", "刷卡", "手续费", "POS", "刷卡手续费"]', "660301", None, None, None),
            ("account", "银行利息", '["银行利息", "利息"]', "6603", None, None, None),
            ("account", "扣款", '["转账扣款", "划扣", "扣款"]', "660301", None, None, None),
            ("account", "会费", '["会费"]', "6602", None, None, None),
            ("account", "佣金", '["佣金"]', "6601", None, None, None),
        ]

        # 3. 部门规则
        dept_rules = [
            ("department", "巴拿马", '["巴拿马"]', None, None, None, "10001"),
            ("department", "中国", '["中国"]', None, None, None, "10002"),
        ]

        all_rules = business_rules + account_rules + dept_rules
        
        try:
            cursor.executemany("""
                INSERT INTO recognition_rules (rule_type, name, keywords, account_code, transaction_type, summary_code, dept_code)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, all_rules)
            self.conn.commit()
            print("默认识别规则初始化完成")
        except Exception as e:
            print(f"初始化默认规则失败: {e}")


    # ---------- 映射方案管理 ----------
    def save_mapping_scheme(self, name: str, base_mode: str, template_path: str, mapping: Dict[str, str],
                             composite: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """保存映射方案"""
        if not name:
            return {"success": False, "message": "方案名称不能为空"}
        
        try:
            cursor = self.conn.cursor()
            mapping_json = json.dumps(mapping, ensure_ascii=False)
            composite_json = json.dumps(composite or {}, ensure_ascii=False)
            cursor.execute("""
                INSERT OR REPLACE INTO mapping_schemes (name, base_mode, template_path, mapping_json, composite_json)
                VALUES (?, ?, ?, ?, ?)
            """, (name, base_mode, template_path, mapping_json, composite_json))
            self.conn.commit()
            return {"success": True, "message": "方案保存成功"}
        except Exception as e:
            return {"success": False, "message": f"保存方案失败: {e}"}

    def get_mapping_schemes(self) -> List[Dict]:
        """获取所有映射方案"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT id, name, base_mode, template_path, mapping_json, composite_json FROM mapping_schemes ORDER BY created_at DESC")
        columns = [description[0] for description in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def delete_mapping_scheme(self, name: str) -> Dict[str, Any]:
        """删除映射方案"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("DELETE FROM mapping_schemes WHERE name = ?", (name,))
            self.conn.commit()
            if cursor.rowcount > 0:
                return {"success": True, "message": "方案删除成功"}
            return {"success": False, "message": "未找到指定方案"}
        except Exception as e:
            return {"success": False, "message": f"删除方案失败: {e}"}

    # ---------- 自动映射缓存 ----------
    def save_auto_mapping(self, template_path: str, base_mode: str, source_columns: List[str],
                          mapping: Dict[str, str], composite: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """保存自动映射缓存（按模板 + 模式 + 源表头签名）"""
        if not template_path:
            return {"success": False, "message": "模板路径不能为空"}
        if not source_columns:
            return {"success": False, "message": "源表头不能为空"}
        if not mapping:
            return {"success": False, "message": "映射为空，已跳过保存"}

        try:
            signature = self._build_source_signature(source_columns)
            mapping_json = json.dumps(mapping, ensure_ascii=False)
            columns_json = json.dumps(list(source_columns), ensure_ascii=False)
            composite_json = json.dumps(composite or {}, ensure_ascii=False)
            cursor = self.conn.cursor()
            cursor.execute("""
                INSERT OR REPLACE INTO auto_mapping_cache
                (template_path, base_mode, source_signature, source_columns_json, mapping_json, composite_json, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (template_path, base_mode, signature, columns_json, mapping_json, composite_json))
            self.conn.commit()
            return {"success": True, "message": "自动映射缓存已保存"}
        except Exception as e:
            return {"success": False, "message": f"保存自动映射失败: {e}"}

    def get_auto_mapping(self, template_path: str, base_mode: str, source_columns: List[str]) -> Optional[Dict[str, Any]]:
        """获取自动映射缓存"""
        if not template_path or not source_columns:
            return None
        try:
            signature = self._build_source_signature(source_columns)
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT mapping_json, composite_json
                FROM auto_mapping_cache
                WHERE template_path = ? AND base_mode = ? AND source_signature = ?
                ORDER BY updated_at DESC
                LIMIT 1
            """, (template_path, base_mode, signature))
            row = cursor.fetchone()
            if not row:
                return None
            mapping_json, composite_json = row
            mapping = json.loads(mapping_json) if mapping_json else {}
            composite = json.loads(composite_json) if composite_json else {}
            return {"mapping": mapping, "composite": composite}
        except Exception:
            return None

    def _ensure_cache_extra_columns(self):
        """为智能识别缓存表补充新增列（向后兼容已有数据库）"""
        cursor = self.conn.cursor()
        cursor.execute("PRAGMA table_info(smart_recognition_cache)")
        columns = [row[1] for row in cursor.fetchall()]
        if "match_items" not in columns:
            try:
                cursor.execute("ALTER TABLE smart_recognition_cache ADD COLUMN match_items TEXT")
                cursor.execute("UPDATE smart_recognition_cache SET match_items = '[]' WHERE match_items IS NULL")
                self.conn.commit()
                print("已为 smart_recognition_cache 补充 match_items 列")
            except Exception as e:
                print(f"补充 match_items 列失败: {e}")

    def _ensure_base_match_columns(self):
        """为基础数据表补充 match_items 列（向后兼容）"""
        tables = ["account_subject", "product", "business_partner"]
        cursor = self.conn.cursor()
        for table in tables:
            cursor.execute(f"PRAGMA table_info({table})")
            columns = [row[1] for row in cursor.fetchall()]
            if "match_items" not in columns:
                try:
                    cursor.execute(f"ALTER TABLE {table} ADD COLUMN match_items TEXT")
                    cursor.execute(f"UPDATE {table} SET match_items = '[]' WHERE match_items IS NULL")
                    self.conn.commit()
                    print(f"已为 {table} 补充 match_items 列")
                except Exception as e:
                    print(f"补充 {table}.match_items 列失败: {e}")

    def _load_cache_maps(self):
        """加载智能识别缓存到内存，避免高频全表扫描"""
        self._cache_lookup = {}
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT summary, match_items, account_code
                FROM smart_recognition_cache
                ORDER BY datetime(created_at) DESC, id DESC
            """)
            rows = cursor.fetchall()
            for summary, match_items_raw, account_code in rows:
                if not account_code:
                    continue
                norm_summary = self._normalize_cache_key(summary)
                if norm_summary and norm_summary not in self._cache_lookup:
                    self._cache_lookup[norm_summary] = account_code
                for alias in self._parse_match_items(match_items_raw):
                    norm_alias = self._normalize_cache_key(alias)
                    if norm_alias and norm_alias not in self._cache_lookup:
                        self._cache_lookup[norm_alias] = account_code
        except Exception as e:
            print(f"缓存索引加载失败: {e}")

    def get_config(self, key: str, default: Any = None) -> str:
        """获取配置项"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT value FROM app_config WHERE key = ?", (key,))
        row = cursor.fetchone()
        return row[0] if row else default

    def set_config(self, key: str, value: str):
        """设置配置项"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("INSERT OR REPLACE INTO app_config (key, value) VALUES (?, ?)", (key, str(value)))
            self.conn.commit()
        except Exception as e:
            print(f"配置保存失败: {e}")

    # ---------- 科目-往来单位映射（存于 app_config） ----------
    def set_partner_for_subject(self, subject_code: str, partner_code: str):
        if not subject_code:
            return
        key = f"partner_for_subject_{subject_code}"
        self.set_config(key, partner_code or "")

    def get_partner_for_subject(self, subject_code: str) -> Optional[str]:
        if not subject_code:
            return None
        key = f"partner_for_subject_{subject_code}"
        return self.get_config(key, None)

    def get_subject_partner_map(self) -> Dict[str, str]:
        cfg = self.get_all_configs()
        prefix = "partner_for_subject_"
        return {k[len(prefix):]: v for k, v in cfg.items() if k.startswith(prefix)}

    def get_all_configs(self) -> Dict[str, str]:
        """获取所有配置"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT key, value FROM app_config")
        return dict(cursor.fetchall())

    # ---------- 识别规则管理 ----------
    def get_recognition_rules(self, rule_type: Optional[str] = None) -> List[Dict]:
        """获取识别规则"""
        cursor = self.conn.cursor()
        if rule_type:
            cursor.execute("SELECT * FROM recognition_rules WHERE rule_type = ? ORDER BY priority DESC, id ASC", (rule_type,))
        else:
            cursor.execute("SELECT * FROM recognition_rules ORDER BY rule_type, priority DESC, id ASC")
        
        columns = [description[0] for description in cursor.description]
        rows = cursor.fetchall()
        return [dict(zip(columns, row)) for row in rows]

    def add_recognition_rule(self, rule_type: str, name: str, keywords: List[str], 
                           account_code: Optional[str] = None, 
                           transaction_type: Optional[str] = None, 
                           summary_code: Optional[str] = None,
                           dept_code: Optional[str] = None) -> Dict[str, Any]:
        """添加识别规则"""
        try:
            keywords_json = json.dumps(keywords, ensure_ascii=False)
            cursor = self.conn.cursor()
            cursor.execute("""
                INSERT INTO recognition_rules (rule_type, name, keywords, account_code, transaction_type, summary_code, dept_code)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (rule_type, name, keywords_json, account_code, transaction_type, summary_code, dept_code))
            self.conn.commit()
            return {"success": True, "message": "规则添加成功", "id": cursor.lastrowid}
        except Exception as e:
            return {"success": False, "message": f"规则添加失败: {e}"}

    def update_recognition_rule(self, rule_id: int, data: Dict[str, Any]) -> Dict[str, Any]:
        """更新识别规则"""
        try:
            # 如果包含 keywords，需要转为 JSON
            if "keywords" in data and isinstance(data["keywords"], list):
                data["keywords"] = json.dumps(data["keywords"], ensure_ascii=False)
            
            # 移除 id 防止更新
            data = {k: v for k, v in data.items() if k != 'id'}
            if not data:
                return {"success": False, "message": "未提供更新字段"}

            allowed_fields = {
                "rule_type", "name", "keywords", "account_code", "transaction_type",
                "summary_code", "dept_code", "priority"
            }
            invalid_fields = [k for k in data.keys() if k not in allowed_fields]
            if invalid_fields:
                return {"success": False, "message": f"不允许的字段: {', '.join(invalid_fields)}"}
            
            set_clause = ', '.join([f"{k} = ?" for k in data.keys()])
            values = list(data.values()) + [rule_id]
            
            cursor = self.conn.cursor()
            cursor.execute(f"UPDATE recognition_rules SET {set_clause} WHERE id = ?", values)
            self.conn.commit()
            return {"success": True, "message": "规则更新成功"}
        except Exception as e:
            return {"success": False, "message": f"规则更新失败: {e}"}

    def delete_recognition_rule(self, rule_id: int) -> Dict[str, Any]:
        """删除识别规则"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("DELETE FROM recognition_rules WHERE id = ?", (rule_id,))
            self.conn.commit()
            return {"success": True, "message": "规则删除成功"}
        except Exception as e:
            return {"success": False, "message": f"规则删除失败: {e}"}

    def _normalize_cache_key(self, text: str) -> str:
        """标准化缓存匹配键（轻量版，去首尾空格并小写）"""
        return str(text).strip().lower()

    def _parse_match_items(self, raw: Any) -> List[str]:
        """解析数据库中的 match_items 字段"""
        if not raw:
            return []
        if isinstance(raw, list):
            items = raw
        else:
            try:
                items = json.loads(raw)
            except Exception:
                # 兜底：按逗号/换行拆分
                if isinstance(raw, str):
                    parts = (
                        raw.replace("\n", ",")
                        .replace("；", ",")
                        .replace("、", ",")
                        .replace("，", ",")
                    ).split(",")
                    items = [p.strip() for p in parts if p and p.strip()]
                else:
                    items = []
        # 过滤空值并标准化
        cleaned = []
        for item in items:
            if item is None:
                continue
            s = str(item).strip()
            if s:
                cleaned.append(s)
        return cleaned

    def _normalize_match_items(self, match_items: Optional[Any]) -> List[str]:
        """将传入的匹配项统一整理为去重列表"""
        items = []
        if match_items is None:
            return items
        if isinstance(match_items, str):
            # 尝试解析 JSON；失败则用分隔符拆分
            try:
                parsed = json.loads(match_items)
                if isinstance(parsed, list):
                    items = parsed
                else:
                    items = [match_items]
            except Exception:
                parts = match_items.replace("\n", ",").replace("；", ",").replace("、", ",").split(",")
                items = [p.strip() for p in parts if p and p.strip()]
        elif isinstance(match_items, list):
            items = match_items
        else:
            items = [match_items]

        normalized = []
        seen = set()
        for item in items:
            if item is None:
                continue
            cleaned = str(item).strip()
            key = self._normalize_cache_key(cleaned)
            if not key or key in seen:
                continue
            seen.add(key)
            normalized.append(cleaned)
        return normalized

    def get_cached_recognition(self, summary: str) -> Optional[str]:
        """获取缓存的识别结果"""
        if not summary:
            return None

        target = self._normalize_cache_key(summary)
        return self._cache_lookup.get(target)

    def get_cached_recognition_fuzzy(self, summary: str, min_ratio: float = 0.65) -> Optional[str]:
        """
        模糊获取缓存的识别结果（通过相似度）
        - 仅使用相似度评分，避免短摘要因“包含关系”误命中错误科目
        """
        if not summary:
            return None

        norm_summary = self._normalize_cache_key(summary)
        best_code = None
        best_score = 0.0
        summary_len = len(norm_summary)

        for cand, account_code in self._cache_lookup.items():
            if not cand:
                continue
            cand_len = len(cand)
            min_len = min(summary_len, cand_len)
            max_len = max(summary_len, cand_len)
            if min_len > 0 and max_len / min_len > 2.5:
                continue
            score = difflib.SequenceMatcher(None, norm_summary, cand).ratio()

            if score > best_score and score >= min_ratio:
                best_score = score
                best_code = account_code
                if score >= 0.95:
                    break

        return best_code

    def save_cached_recognition(self, summary: str, account_code: str, match_items: Optional[Any] = None):
        """保存识别结果到缓存"""
        if not summary or not account_code:
            return
        try:
            cursor = self.conn.cursor()
            items_json = json.dumps(self._normalize_match_items(match_items), ensure_ascii=False)
            # 使用 REPLACE INTO 确保更新重复的摘要
            cursor.execute(
                "INSERT OR REPLACE INTO smart_recognition_cache (summary, account_code, match_items) VALUES (?, ?, ?)",
                (summary, account_code, items_json)
            )
            self.conn.commit()
            # 直接重建索引，避免旧别名残留或重复别名无法被最新记录覆盖
            self._load_cache_maps()
        except Exception as e:
            print(f"缓存保存失败: {e}")

    def get_all_cached_recognitions(self) -> List[Dict]:
        """获取所有缓存的识别结果"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT id, summary, match_items, account_code, created_at FROM smart_recognition_cache ORDER BY created_at DESC")
        columns = [description[0] for description in cursor.description]
        rows = cursor.fetchall()
        return [dict(zip(columns, row)) for row in rows]

    def update_cached_recognition(self, record_id: int, new_account_code: Optional[str] = None, match_items: Optional[Any] = None) -> Dict[str, Any]:
        """更新缓存的识别结果（支持科目编码和匹配项）"""
        if new_account_code is None and match_items is None:
            return {"success": False, "message": "未提供更新字段"}

        try:
            cursor = self.conn.cursor()
            set_clause = []
            params = []

            if new_account_code is not None:
                set_clause.append("account_code = ?")
                params.append(new_account_code)

            if match_items is not None:
                items_json = json.dumps(self._normalize_match_items(match_items), ensure_ascii=False)
                set_clause.append("match_items = ?")
                params.append(items_json)

            params.append(record_id)
            cursor.execute(
                f"UPDATE smart_recognition_cache SET {', '.join(set_clause)} WHERE id = ?",
                params
            )
            self.conn.commit()
            if cursor.rowcount > 0:
                self._load_cache_maps()
                return {"success": True, "message": "缓存更新成功"}
            else:
                return {"success": False, "message": "未找到要更新的缓存记录"}
        except Exception as e:
            return {"success": False, "message": f"缓存更新失败: {str(e)}"}

    def delete_cached_recognition(self, record_id: int) -> Dict[str, Any]:
        """删除缓存的识别结果"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("DELETE FROM smart_recognition_cache WHERE id = ?", (record_id,))
            self.conn.commit()
            if cursor.rowcount > 0:
                self._load_cache_maps()
                return {"success": True, "message": "缓存删除成功"}
            else:
                return {"success": False, "message": "未找到要删除的缓存记录"}
        except Exception as e:
            return {"success": False, "message": f"缓存删除失败: {str(e)}"}

    def import_all_data(self, base_data_dir: Optional[str] = None) -> Dict[str, Any]:
        """导入所有基础数据"""
        base_dir = base_data_dir or self.BASE_DATA_DIR

        if not os.path.exists(base_dir):
            return {
                "success": False,
                "message": f"基础数据目录不存在: {base_dir}",
                "details": {}
            }

        results = {}
        for file_name, table_name in self.DATA_FILES.items():
            self._assert_valid_table(table_name)
            file_path = os.path.join(base_dir, file_name)
            result = self.import_single_file(file_path, table_name)
            results[file_name] = result

        success_count = sum(1 for r in results.values() if r["success"])

        return {
            "success": success_count == len(self.DATA_FILES),
            "message": f"成功导入 {success_count}/{len(self.DATA_FILES)} 个文件",
            "details": results
        }

    def import_single_file(self, file_path: str, table_name: str) -> Dict[str, Any]:
        """导入单个文件到数据库"""
        if self._is_custom_table(table_name):
            custom_name = self._get_custom_name(table_name)
            if not os.path.exists(file_path):
                return {"success": False, "message": f"文件不存在: {file_path}", "count": 0}
            try:
                df = pd.read_excel(file_path, header=1)
                df = df.dropna(how="all")
                if df.empty:
                    return {"success": True, "message": "无数据", "count": 0}
                cols = [str(c).strip() for c in df.columns]
                if "code" not in cols or "name" not in cols:
                    return {"success": False, "message": "自定义分类必须包含 code 和 name 列", "count": 0}
                df.columns = cols
                fields = self.get_custom_category_fields(custom_name)
                cursor = self.conn.cursor()
                cursor.execute("DELETE FROM custom_record WHERE category_name = ?", (custom_name,))
                records = 0
                for _, row in df.iterrows():
                    code = str(row.get("code", "")).strip()
                    name = str(row.get("name", "")).strip()
                    if not code or not name:
                        continue
                    payload = {f: row.get(f, "") for f in fields if f in df.columns}
                    cursor.execute(
                        "INSERT INTO custom_record (category_name, code, name, data_json) VALUES (?, ?, ?, ?)",
                        (custom_name, code, name, json.dumps(payload, ensure_ascii=False))
                    )
                    records += 1
                self.conn.commit()
                return {"success": True, "message": "导入成功", "count": records}
            except Exception as exc:
                return {"success": False, "message": f"导入失败: {exc}", "count": 0}

        self._assert_valid_table(table_name)
        if not os.path.exists(file_path):
            return {
                "success": False,
                "message": f"文件不存在: {file_path}",
                "count": 0
            }

        tmp_table = None
        try:
            # 读取Excel文件（跳过第一行公司名称，从第二行开始读取表头）
            df = pd.read_excel(file_path, header=1)

            # 过滤掉时间戳行（最后一行通常是导出时间）
            if len(df) > 0:
                # 检查最后一行是否为时间戳
                last_row = df.iloc[-1]
                if pd.notna(last_row.iloc[0]) and ":" in str(last_row.iloc[0]):
                    df = df[:-1]

            # 根据表名进行特定的列映射
            df_clean = self._clean_dataframe(df, table_name)

            if df_clean is None or len(df_clean) == 0:
                return {
                    "success": False,
                    "message": f"文件无有效数据: {file_path}",
                    "count": 0
                }

            tmp_table = f"{table_name}_import_tmp"
            df_clean.to_sql(tmp_table, self.conn, if_exists='replace', index=False)
            columns = ", ".join([f"\"{c}\"" for c in df_clean.columns])

            with self.conn:
                cursor = self.conn.cursor()
                cursor.execute(f"DELETE FROM {table_name}")
                cursor.execute(
                    f"INSERT INTO {table_name} ({columns}) SELECT {columns} FROM {tmp_table}"
                )
                cursor.execute("""
                    INSERT INTO import_log (table_name, file_name, record_count, status)
                    VALUES (?, ?, ?, 'success')
                """, (table_name, os.path.basename(file_path), len(df_clean)))

            return {
                "success": True,
                "message": f"成功导入 {len(df_clean)} 条记录",
                "count": len(df_clean)
            }

        except Exception as e:
            # 记录失败日志
            try:
                cursor = self.conn.cursor()
                cursor.execute("""
                    INSERT INTO import_log (table_name, file_name, record_count, status)
                    VALUES (?, ?, 0, ?)
                """, (table_name, os.path.basename(file_path), f"failed: {str(e)}"))
                self.conn.commit()
            except:
                pass

            return {
                "success": False,
                "message": f"导入失败: {str(e)}",
                "count": 0
            }
        finally:
            if tmp_table:
                try:
                    self.conn.execute(f"DROP TABLE IF EXISTS {tmp_table}")
                    self.conn.commit()
                except Exception:
                    pass

    def import_training_cache(self, file_path: str) -> Dict[str, Any]:
        """
        导入历史凭证作为AI训练缓存
        期望至少包含：
          - 摘要列: 摘要/摘要名/摘要说明/描述/description
          - 科目列: 科目编码/科目代码/会计科目/科目
        可选：
          - 类型列: 类型/借贷/借贷标志/方向/借贷类型（用于匹配项）
        """
        if not os.path.exists(file_path):
            return {"success": False, "message": f"文件不存在: {file_path}", "imported": 0}

        try:
            df = pd.read_excel(file_path)
        except Exception as e:
            return {"success": False, "message": f"读取Excel失败: {e}", "imported": 0}

        summary_cols = ["摘要", "摘要名", "摘要说明", "描述", "description", "摘要描述"]
        account_cols = ["科目编码", "科目代码", "会计科目", "科目"]
        type_cols = ["类型", "借贷", "借贷标志", "方向", "借贷类型"]

        def pick_first(row, cols):
            for c in cols:
                if c in row and pd.notna(row[c]) and str(row[c]).strip() != "":
                    return str(row[c]).strip()
            return None

        imported = 0
        skipped = 0
        for _, row in df.iterrows():
            summary_val = pick_first(row, summary_cols)
            account_val = pick_first(row, account_cols)
            if not summary_val or not account_val:
                skipped += 1
                continue

            # 额外匹配项：原摘要、类型信息
            match_items = [summary_val]
            type_val = pick_first(row, type_cols)
            if type_val:
                match_items.append(f"类型:{type_val}")

            try:
                self.save_cached_recognition(summary_val, account_val, match_items=match_items)
                imported += 1
            except Exception:
                skipped += 1

        return {
            "success": True,
            "message": f"已导入训练样本 {imported} 条，跳过 {skipped} 条",
            "imported": imported,
            "skipped": skipped
        }

    def _clean_dataframe(self, df: pd.DataFrame, table_name: str) -> Optional[pd.DataFrame]:
        """清理和映射DataFrame列名"""

        # 币种表
        if table_name == "currency":
            if '外币编码' in df.columns:
                return df[['外币编码', '外币名', '汇率', '使用类型']].rename(columns={
                    '外币编码': 'code',
                    '外币名': 'name',
                    '汇率': 'exchange_rate',
                    '使用类型': 'use_type'
                })

        # 部门表
        elif table_name == "department":
            if '部门编码' in df.columns:
                return df[['部门编码', '部门名', '使用']].rename(columns={
                    '部门编码': 'code',
                    '部门名': 'name',
                    '使用': 'is_active'
                })

        # 仓库表
        elif table_name == "warehouse":
            if '仓库/工厂编码' in df.columns:
                df_result = df[['仓库/工厂编码', '仓库/工厂名', '类型']].copy()
                df_result['生产流程名'] = df.get('生产流程名', '')
                df_result['外包往来单位名'] = df.get('外包往来单位名', '')
                return df_result.rename(columns={
                    '仓库/工厂编码': 'code',
                    '仓库/工厂名': 'name',
                    '类型': 'type',
                    '生产流程名': 'production_process',
                    '外包往来单位名': 'outsource_partner'
                })

        # 科目编码表
        elif table_name == "account_subject":
            if '[科目编码]科目名' in df.columns:
                df_result = pd.DataFrame()
                df_result['code_name'] = df['[科目编码]科目名']
                df_result['is_subject'] = df.get('科目', '')
                df_result['debit_credit_type'] = df.get('借贷类型', '')
                df_result['subject_type'] = df.get('科目类型', '')
                df_result['contra_type'] = df.get('备抵类型', '')
                df_result['use_type'] = df.get('使用类型', '')
                df_result['summary'] = df.get('摘要', '')
                df_result['parent_subject'] = df.get('上级科目', '')
                df_result['display_name'] = df.get('显示名1', '')
                df_result['match_items'] = '[]'
                return df_result

        # 品目信息表
        elif table_name == "product":
            if '品目编码' in df.columns:
                df_result = pd.DataFrame()
                df_result['code'] = df['品目编码']
                df_result['name'] = df.get('品目名', '')
                df_result['product_type'] = df.get('品目类型', '')
                df_result['spec_info'] = df.get('规格信息', '')
                df_result['unit'] = df.get('单位', '')
                df_result['search_keyword'] = df.get('查询关键词', '')
                df_result['pack_qty'] = df.get('装数', 0)
                df_result['unit_conversion_denominator'] = df.get('单位转换比率(分母)', 0)
                df_result['unit_conversion_numerator'] = df.get('单位转换比率(分子)', 0)
                df_result['specification'] = df.get('规格', '')
                df_result['in_price'] = df.get('入库单价', 0)
                df_result['out_price'] = df.get('出库单价', 0)
                df_result['price_a'] = df.get('单价 A', 0)
                df_result['price_b'] = df.get('单价 B', 0)
                df_result['price_c'] = df.get('单价 C', 0)
                df_result['length'] = df.get('长度', 0)
                df_result['width'] = df.get('宽度', 0)
                df_result['height'] = df.get('高度', 0)
                df_result['volume'] = df.get('体积', 0)
                df_result['weight'] = df.get('单件重量', 0)
                df_result['color'] = df.get('COLOR', '')
                df_result['size_range'] = df.get('No.TAMA.', '')
                df_result['match_items'] = '[]'
                return df_result

        # 往来单位表
        elif table_name == "business_partner":
            if '往来单位编码' in df.columns:
                df_result = pd.DataFrame()
                df_result['code'] = df['往来单位编码']
                df_result['name'] = df.get('往来单位名', '')
                df_result['contact_person'] = df.get('联系人', '')
                df_result['mobile'] = df.get('手机号码', '')
                df_result['phone'] = df.get('电话（文本格式）', '')
                df_result['email'] = df.get('邮件（文本格式）', '')
                df_result['category'] = df.get('往来单位分级组合名', '')
                df_result['file_management'] = df.get('文件管理', '')
                df_result['tax_number'] = df.get('税号', '')
                df_result['bank_name'] = df.get('开户行', '')
                df_result['bank_account'] = df.get('银行账号', '')
                df_result['local_code'] = df.get('当地编码', '')
                df_result['match_items'] = '[]'
                return df_result

        # 账户表
        elif table_name == "bank_account":
            if '账号编码' in df.columns:
                df_result = pd.DataFrame()
                df_result['code'] = df['账号编码']
                df_result['name'] = df.get('账号名', '')
                df_result['account_subject'] = df.get('科目名(会计科目)', '')
                df_result['search_keyword'] = df.get('查询关键词', '')
                df_result['summary'] = df.get('摘 要', '')
                df_result['foreign_currency'] = df.get('外币存折', '')
                df_result['is_active'] = df.get('是否使用', '')
                return df_result

        return None

    def query(self, table_name: str, code: Optional[str] = None) -> List[Dict]:
        """查询数据"""
        if self._is_custom_table(table_name):
            custom_name = self._get_custom_name(table_name)
            results = self._query_custom_records(custom_name)
            if code:
                results = [r for r in results if str(r.get("code", "")).strip() == str(code)]
            return results
        self._assert_valid_table(table_name)
        cursor = self.conn.cursor()

        if code:
            cursor.execute(f"SELECT * FROM {table_name} WHERE code = ?", (code,))
        else:
            cursor.execute(f"SELECT * FROM {table_name}")

        columns = [description[0] for description in cursor.description]
        rows = cursor.fetchall()

        return [dict(zip(columns, row)) for row in rows]

    def search_by_name(self, table_name: str, keyword: str) -> List[Dict]:
        """根据名称关键词搜索"""
        if self._is_custom_table(table_name):
            custom_name = self._get_custom_name(table_name)
            results = self._query_custom_records(custom_name)
            k_lower = (keyword or "").lower()
            if k_lower:
                results = [
                    r for r in results
                    if any(k_lower in str(v).lower() for v in r.values())
                ]
            return results
        self._assert_valid_table(table_name)
        cursor = self.conn.cursor()
        # 扩展：如果表有 code 列，也一并模糊查找
        cursor.execute(f"PRAGMA table_info({table_name})")
        cols = [row[1] for row in cursor.fetchall()]
        has_code = "code" in cols
        if has_code:
            cursor.execute(
                f"SELECT * FROM {table_name} WHERE name LIKE ? OR code LIKE ?",
                (f"%{keyword}%", f"%{keyword}%")
            )
        else:
            cursor.execute(f"SELECT * FROM {table_name} WHERE name LIKE ?", (f"%{keyword}%",))

        columns = [description[0] for description in cursor.description]
        rows = cursor.fetchall()

        return [dict(zip(columns, row)) for row in rows]

    def get_statistics(self) -> Dict[str, int]:
        """获取各表记录数统计"""
        cursor = self.conn.cursor()
        stats = {}

        for table_name in self.DATA_FILES.values():
            self._assert_valid_table(table_name)
            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
            count = cursor.fetchone()[0]
            stats[table_name] = count

        try:
            for cat in self.list_custom_categories():
                name_key = cat.get("name")
                display = cat.get("display_name") or name_key
                if not name_key:
                    continue
                cursor.execute(
                    "SELECT COUNT(*) FROM custom_record WHERE category_name = ?",
                    (name_key,)
                )
                stats[f"custom:{display}"] = cursor.fetchone()[0]
        except Exception:
            pass

        return stats

    def add_record(self, table_name: str, data: Dict[str, Any]) -> Dict[str, Any]:
        """添加记录"""
        if self._is_custom_table(table_name):
            return self._add_custom_record(self._get_custom_name(table_name), data)
        self._assert_valid_table(table_name)
        try:
            # 移除id字段（如果存在）
            data = {k: v for k, v in data.items() if k != 'id'}
            if not data:
                return {"success": False, "message": "未提供要添加的字段"}

            valid_columns = set(self.get_table_columns(table_name))
            invalid_fields = [k for k in data.keys() if k not in valid_columns]
            if invalid_fields:
                return {"success": False, "message": f"非法字段: {', '.join(invalid_fields)}"}

            columns = ', '.join(data.keys())
            placeholders = ', '.join(['?' for _ in data])
            values = list(data.values())

            cursor = self.conn.cursor()
            cursor.execute(
                f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})",
                values
            )
            self.conn.commit()

            return {
                "success": True,
                "message": "添加成功",
                "id": cursor.lastrowid
            }
        except sqlite3.IntegrityError as e:
            return {
                "success": False,
                "message": f"添加失败：可能存在重复的编码 ({str(e)})"
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"添加失败：{str(e)}"
            }

    def update_record(self, table_name: str, record_id: int, data: Dict[str, Any]) -> Dict[str, Any]:
        """更新记录"""
        if self._is_custom_table(table_name):
            return self._update_custom_record(self._get_custom_name(table_name), record_id, data)
        self._assert_valid_table(table_name)
        try:
            # 移除id字段
            data = {k: v for k, v in data.items() if k != 'id'}
            if not data:
                return {"success": False, "message": "未提供更新字段"}

            valid_columns = set(self.get_table_columns(table_name))
            invalid_fields = [k for k in data.keys() if k not in valid_columns]
            if invalid_fields:
                return {"success": False, "message": f"非法字段: {', '.join(invalid_fields)}"}

            set_clause = ', '.join([f"{k} = ?" for k in data.keys()])
            values = list(data.values()) + [record_id]

            cursor = self.conn.cursor()
            cursor.execute(
                f"UPDATE {table_name} SET {set_clause} WHERE id = ?",
                values
            )
            self.conn.commit()

            if cursor.rowcount > 0:
                return {
                    "success": True,
                    "message": "更新成功"
                }
            else:
                return {
                    "success": False,
                    "message": "未找到要更新的记录"
                }
        except sqlite3.IntegrityError as e:
            return {
                "success": False,
                "message": f"更新失败：可能存在重复的编码 ({str(e)})"
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"更新失败：{str(e)}"
            }

    def delete_record(self, table_name: str, record_id: int) -> Dict[str, Any]:
        """删除记录"""
        if self._is_custom_table(table_name):
            return self._delete_custom_record(self._get_custom_name(table_name), record_id)
        self._assert_valid_table(table_name)
        try:
            cursor = self.conn.cursor()
            cursor.execute(f"DELETE FROM {table_name} WHERE id = ?", (record_id,))
            self.conn.commit()

            if cursor.rowcount > 0:
                return {
                    "success": True,
                    "message": "删除成功"
                }
            else:
                return {
                    "success": False,
                    "message": "未找到要删除的记录"
                }
        except Exception as e:
            return {
                "success": False,
                "message": f"删除失败：{str(e)}"
            }

    def get_record_by_id(self, table_name: str, record_id: int) -> Optional[Dict]:
        """根据ID获取单条记录"""
        if self._is_custom_table(table_name):
            return self._get_custom_record_by_id(self._get_custom_name(table_name), record_id)
        self._assert_valid_table(table_name)
        cursor = self.conn.cursor()
        cursor.execute(f"SELECT * FROM {table_name} WHERE id = ?", (record_id,))

        row = cursor.fetchone()
        if row:
            columns = [description[0] for description in cursor.description]
            return dict(zip(columns, row))
        return None

    def list_custom_categories(self) -> List[Dict[str, Any]]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT name, display_name, fields_json FROM custom_category ORDER BY display_name")
        rows = cursor.fetchall()
        results = []
        for row in rows:
            fields = []
            raw = row[2] if len(row) > 2 else None
            if raw:
                try:
                    fields = json.loads(raw)
                except Exception:
                    fields = []
            results.append({
                "name": row[0],
                "display_name": row[1] or row[0],
                "fields": fields
            })
        return results

    def add_custom_category(self, name: str, display_name: str, fields: List[Dict[str, Any]]) -> Dict[str, Any]:
        if not name or not display_name:
            return {"success": False, "message": "分类名称为空"}
        fields = fields or []
        payload = json.dumps(fields, ensure_ascii=False)
        try:
            cursor = self.conn.cursor()
            cursor.execute(
                "INSERT INTO custom_category (name, display_name, fields_json) VALUES (?, ?, ?)",
                (name, display_name, payload)
            )
            self.conn.commit()
            return {"success": True, "message": "分类已创建"}
        except Exception as exc:
            return {"success": False, "message": str(exc)}

    def delete_custom_category(self, name: str) -> Dict[str, Any]:
        if not name:
            return {"success": False, "message": "分类名称为空"}
        try:
            cursor = self.conn.cursor()
            cursor.execute("DELETE FROM custom_category WHERE name = ?", (name,))
            cursor.execute("DELETE FROM custom_record WHERE category_name = ?", (name,))
            self.conn.commit()
            return {"success": True, "message": "分类已删除"}
        except Exception as exc:
            return {"success": False, "message": str(exc)}

    def get_custom_category_fields(self, name: str) -> List[str]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT fields_json FROM custom_category WHERE name = ?", (name,))
        row = cursor.fetchone()
        fields = []
        if row and row[0]:
            try:
                fields = json.loads(row[0])
            except Exception:
                fields = []
        return [f.get("name") for f in fields if f.get("name")]

    def _query_custom_records(self, category_name: str) -> List[Dict[str, Any]]:
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT id, code, name, data_json FROM custom_record WHERE category_name = ?",
            (category_name,)
        )
        rows = cursor.fetchall()
        fields = self.get_custom_category_fields(category_name)
        results = []
        for row in rows:
            data = {}
            raw = row[3]
            if raw:
                try:
                    data = json.loads(raw)
                except Exception:
                    data = {}
            entry = {"id": row[0], "code": row[1], "name": row[2]}
            for fname in fields:
                entry[fname] = data.get(fname, "")
            results.append(entry)
        return results

    def _add_custom_record(self, category_name: str, data: Dict[str, Any]) -> Dict[str, Any]:
        code = str(data.get("code", "")).strip()
        name = str(data.get("name", "")).strip()
        if not code or not name:
            return {"success": False, "message": "编码或名称为空"}
        fields = self.get_custom_category_fields(category_name)
        payload = {k: data.get(k) for k in fields if k in data}
        try:
            cursor = self.conn.cursor()
            cursor.execute(
                "INSERT INTO custom_record (category_name, code, name, data_json) VALUES (?, ?, ?, ?)",
                (category_name, code, name, json.dumps(payload, ensure_ascii=False))
            )
            self.conn.commit()
            return {"success": True, "message": "新增成功"}
        except Exception as exc:
            return {"success": False, "message": str(exc)}

    def _update_custom_record(self, category_name: str, record_id: int, data: Dict[str, Any]) -> Dict[str, Any]:
        code = str(data.get("code", "")).strip()
        name = str(data.get("name", "")).strip()
        if not code or not name:
            return {"success": False, "message": "编码或名称为空"}
        fields = self.get_custom_category_fields(category_name)
        payload = {k: data.get(k) for k in fields if k in data}
        try:
            cursor = self.conn.cursor()
            cursor.execute(
                "UPDATE custom_record SET code=?, name=?, data_json=? WHERE id=? AND category_name=?",
                (code, name, json.dumps(payload, ensure_ascii=False), record_id, category_name)
            )
            self.conn.commit()
            return {"success": True, "message": "更新成功"}
        except Exception as exc:
            return {"success": False, "message": str(exc)}

    def _delete_custom_record(self, category_name: str, record_id: int) -> Dict[str, Any]:
        try:
            cursor = self.conn.cursor()
            cursor.execute(
                "DELETE FROM custom_record WHERE id=? AND category_name=?",
                (record_id, category_name)
            )
            self.conn.commit()
            return {"success": True, "message": "删除成功"}
        except Exception as exc:
            return {"success": False, "message": str(exc)}

    def _get_custom_record_by_id(self, category_name: str, record_id: int) -> Optional[Dict[str, Any]]:
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT id, code, name, data_json FROM custom_record WHERE id=? AND category_name=?",
            (record_id, category_name)
        )
        row = cursor.fetchone()
        if not row:
            return None
        data = {}
        if row[3]:
            try:
                data = json.loads(row[3])
            except Exception:
                data = {}
        entry = {"id": row[0], "code": row[1], "name": row[2]}
        fields = self.get_custom_category_fields(category_name)
        for fname in fields:
            entry[fname] = data.get(fname, "")
        return entry

    def lookup_value(self, table_name: str, key_col: str, key_val: Any, target_col: str) -> Optional[Any]:
        """
        在指定表中查找特定值
        SELECT target_col FROM table_name WHERE key_col = key_val LIMIT 1
        """
        if self._is_custom_table(table_name):
            custom_name = self._get_custom_name(table_name)
            records = self._query_custom_records(custom_name)
            for row in records:
                if str(row.get(key_col, "")).strip() == str(key_val).strip():
                    return row.get(target_col)
            return None
        self._assert_valid_table(table_name)
        # 简单的防注入检查：列名只能包含字母数字下划线
        if not (key_col.isidentifier() and target_col.isidentifier()):
             print(f"lookup_value 列名非法: {key_col}, {target_col}")
             return None

        try:
            cursor = self.conn.cursor()
            # 检查列是否存在
            cols = self.get_table_columns(table_name)
            if key_col not in cols or target_col not in cols:
                return None
                
            query = f"SELECT {target_col} FROM {table_name} WHERE {key_col} = ? LIMIT 1"
            cursor.execute(query, (key_val,))
            row = cursor.fetchone()
            return row[0] if row else None
        except Exception as e:
            print(f"lookup_value 查询失败: {e}")
            return None

    def find_best_match(self, table_name: str, query_str: str, min_score: float = 0.85) -> Optional[str]:
        """
        在指定表中查找最佳匹配的编码
        1. 精确匹配 Code
        2. 精确匹配 Name
        3. 精确匹配 Match Items
        4. 模糊匹配 Name / Match Items
        """
        if not query_str:
            return None
        
        query_norm = str(query_str).strip().lower()
        if not query_norm:
            return None

        # 简单缓存机制，避免单次批量处理时重复查询数据库
        if not hasattr(self, "_lookup_cache"):
             self._lookup_cache = {}
        
        if table_name not in self._lookup_cache:
            self._lookup_cache[table_name] = self.query(table_name)
        
        data = self._lookup_cache[table_name]
        
        # 1. Exact Match (Priority)
        for row in data:
            code = str(row.get('code', '')).strip()
            if code.lower() == query_norm:
                return code
            if str(row.get('name', '')).strip().lower() == query_norm:
                return code
            
            match_items = self._parse_match_items(row.get('match_items'))
            for item in match_items:
                if str(item).strip().lower() == query_norm:
                    return code
                    
        # 2. Fuzzy Match
        best_code = None
        best_score = 0.0
        
        for row in data:
            code = str(row.get('code', '')).strip()
            candidates = [str(row.get('name', '')).strip().lower()]
            candidates.extend([str(item).strip().lower() for item in self._parse_match_items(row.get('match_items'))])
            
            for cand in candidates:
                if not cand: continue
                
                score = 0.0
                # Check containment (Higher priority than raw fuzzy)
                if query_norm in cand or cand in query_norm:
                    score = 0.9
                else:
                    score = difflib.SequenceMatcher(None, query_norm, cand).ratio()
                
                if score > best_score and score >= min_score:
                    best_score = score
                    best_code = code
        
        return best_code

    def clear_lookup_cache(self):
        """清除查询缓存"""
        if hasattr(self, "_lookup_cache"):
            self._lookup_cache = {}

    def close(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


# 便捷函数
def get_base_data_manager() -> BaseDataManager:
    """获取基础数据管理器单例"""
    return BaseDataManager()
