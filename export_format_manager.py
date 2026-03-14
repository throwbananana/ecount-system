import json
from datetime import datetime
import os
import shutil
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog, simpledialog
from openpyxl import load_workbook
import pandas as pd
from base_data_manager import BaseDataManager

APP_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(APP_DIR, "config.json")
EXPORT_FORMATS_KEY = "EXPORT_FORMATS"
EXPORT_FORMAT_SAMPLES_KEY = "EXPORT_FORMAT_SAMPLES"
CORE_EXPORT_MODULES = (
    "main_export",
    "summary_match_export",
    "shipping_product",
    "shipping_container",
    "image_recognition",
)


def _load_config() -> dict:
    if not os.path.exists(CONFIG_FILE):
        return {}
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_config(config: dict) -> None:
    if os.path.exists(CONFIG_FILE):
        try:
            shutil.copy2(CONFIG_FILE, CONFIG_FILE + ".bak")
        except Exception:
            pass
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)


def load_export_formats() -> dict:
    config = _load_config()
    formats = config.get(EXPORT_FORMATS_KEY, {})
    if not isinstance(formats, dict):
        formats = {}
    changed = False
    for module_key in CORE_EXPORT_MODULES:
        module = formats.get(module_key)
        if not isinstance(module, dict):
            formats[module_key] = {"active": "", "formats": {}}
            changed = True
            continue
        if "active" not in module:
            module["active"] = ""
            changed = True
        if "formats" not in module or not isinstance(module.get("formats"), dict):
            module["formats"] = {}
            changed = True
    if changed:
        config[EXPORT_FORMATS_KEY] = formats
        _save_config(config)
    return formats


def save_export_formats(formats: dict) -> None:
    config = _load_config()
    config[EXPORT_FORMATS_KEY] = formats
    _save_config(config)


def load_export_format_samples() -> dict:
    return _load_config().get(EXPORT_FORMAT_SAMPLES_KEY, {})


def save_export_format_samples(samples: dict) -> None:
    config = _load_config()
    config[EXPORT_FORMAT_SAMPLES_KEY] = samples
    _save_config(config)


def _ensure_module(formats: dict, module_key: str) -> dict:
    module = formats.get(module_key)
    if not isinstance(module, dict):
        module = {}
    if "active" not in module:
        module["active"] = ""
    if "formats" not in module or not isinstance(module["formats"], dict):
        module["formats"] = {}
    formats[module_key] = module
    return module


def get_export_format_names(module_key: str) -> list:
    formats = load_export_formats()
    module = formats.get(module_key, {})
    names = list((module.get("formats") or {}).keys())
    names.sort()
    return names


def get_active_export_format_name(module_key: str) -> str:
    formats = load_export_formats()
    module = formats.get(module_key, {})
    return module.get("active", "") or ""


def set_active_export_format(module_key: str, name: str) -> None:
    formats = load_export_formats()
    module = _ensure_module(formats, module_key)
    module["active"] = name or ""
    save_export_formats(formats)


def get_active_export_mapping(module_key: str) -> list:
    formats = load_export_formats()
    module = formats.get(module_key, {})
    active = module.get("active", "")
    mapping = (module.get("formats") or {}).get(active)
    return mapping or []


def apply_export_format(module_key: str, headers: list, rows: list, base_data_mgr=None):
    formats = load_export_formats()
    module = formats.get(module_key, {})
    if module.get("use_original"):
        return headers, rows, True
    mapping = (module.get("formats") or {}).get(module.get("active", ""))
    if not mapping:
        return headers, rows, False
    mapped_headers, mapped_rows = apply_mapping(headers, rows, mapping, base_data_mgr=base_data_mgr)
    return mapped_headers, mapped_rows, True


def apply_mapping(headers: list, rows: list, mapping: list, base_data_mgr=None):
    name_to_idx = {h: i for i, h in enumerate(headers)}
    new_headers = []
    new_rows = []
    temp_mgr = None
    if base_data_mgr is None and any((item.get("source") or "").strip().startswith("BD:") for item in mapping):
        try:
            temp_mgr = BaseDataManager()
            base_data_mgr = temp_mgr
        except Exception:
            base_data_mgr = None

    for item in mapping:
        output = (item.get("output") or "").strip()
        if output:
            new_headers.append(output)

    for row in rows:
        new_row = []
        for item in mapping:
            output = (item.get("output") or "").strip()
            if not output:
                continue
            source = (item.get("source") or "").strip()
            default = item.get("default", "")
            value = None
            if source.startswith("BD:"):
                value = _resolve_base_data_value(source, name_to_idx, row, base_data_mgr)
            elif source.startswith("BDV:"):
                value = _resolve_fixed_base_data_value(source, base_data_mgr)
            else:
                idx = name_to_idx.get(source)
                if idx is not None and idx < len(row):
                    value = row[idx]
            new_row.append(_resolve_default_value(default) if _is_empty_value(value) else value)
        new_rows.append(new_row)

    if temp_mgr is not None:
        try:
            temp_mgr.close()
        except Exception:
            pass

    return new_headers, new_rows


def _resolve_default_value(default_value):
    if default_value is None:
        return ""
    text = str(default_value)
    now = datetime.now()
    replacements = {
        "{TODAY_YYYYMMDD}": now.strftime("%Y%m%d"),
        "{TODAY_YYYY-MM-DD}": now.strftime("%Y-%m-%d"),
        "{NOW_HHMMSS}": now.strftime("%H%M%S"),
        "{NOW_HH:MM:SS}": now.strftime("%H:%M:%S"),
        "{NOW_YYYYMMDD_HHMMSS}": now.strftime("%Y%m%d_%H%M%S"),
    }
    for token, value in replacements.items():
        text = text.replace(token, value)
    return text


def _is_empty_value(value):
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _resolve_base_data_value(source: str, name_to_idx: dict, row: list, base_data_mgr=None):
    parts = source[3:].split("|")
    if len(parts) < 3:
        return None
    table_name = parts[0]
    target_col = parts[1]
    key_header = parts[2]
    key_idx = name_to_idx.get(key_header)
    if key_idx is None or key_idx >= len(row):
        return None
    key_val = row[key_idx]
    if key_val is None or str(key_val).strip() == "":
        return None
    mgr = base_data_mgr
    temp_mgr = None
    if mgr is None:
        try:
            temp_mgr = BaseDataManager()
            mgr = temp_mgr
        except Exception:
            return None
    try:
        return mgr.lookup_value(table_name, "code", key_val, target_col)
    except Exception:
        return None
    finally:
        if temp_mgr is not None:
            try:
                temp_mgr.close()
            except Exception:
                pass


def _resolve_fixed_base_data_value(source: str, base_data_mgr=None):
    parts = source[4:].split("|", 3)
    if len(parts) < 4:
        return None
    table_name, target_col, key_col, key_val = parts
    if not table_name or not target_col or not key_col or key_val == "":
        return None
    mgr = base_data_mgr
    temp_mgr = None
    if mgr is None:
        try:
            temp_mgr = BaseDataManager()
            mgr = temp_mgr
        except Exception:
            return None
    try:
        return mgr.lookup_value(table_name, key_col, key_val, target_col)
    except Exception:
        return None
    finally:
        if temp_mgr is not None:
            try:
                temp_mgr.close()
            except Exception:
                pass


def _parse_mapping_text(text: str) -> list:
    mappings = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "\t" in line:
            parts = line.split("\t")
        elif "|" in line:
            parts = line.split("|")
        elif "=" in line:
            parts = line.split("=")
        else:
            parts = [line]
        parts = [p.strip() for p in parts]
        if not parts:
            continue
        output = parts[0] if len(parts) > 0 else ""
        source = parts[1] if len(parts) > 1 else ""
        default = parts[2] if len(parts) > 2 else ""
        if not output:
            continue
        mappings.append({"output": output, "source": source, "default": default})
    return mappings


def _render_mapping_text(mapping: list) -> str:
    lines = ["# 输出表头\t源字段\t默认值"]
    for item in mapping:
        output = item.get("output", "")
        source = item.get("source", "")
        default = item.get("default", "")
        lines.append(f"{output}\t{source}\t{default}")
    return "\n".join(lines)


def read_template_headers(template_path: str) -> list:
    wb = load_workbook(template_path, read_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        if cell.value is None:
            continue
        header = str(cell.value).strip()
        if header:
            headers.append(header)
    wb.close()
    return headers


def build_template_mapping(source_headers: list, template_headers: list) -> list:
    mapping = []
    source_set = {str(h).strip() for h in (source_headers or [])}
    for out_header in template_headers or []:
        name = str(out_header).strip()
        source = name if name in source_set else ""
        mapping.append({"output": name, "source": source, "default": ""})
    return mapping


class ExportFormatEditor(tk.Toplevel):
    def __init__(self, parent, module_key: str, source_headers: list, title: str = None, base_data_mgr=None):
        super().__init__(parent)
        self.module_key = module_key
        self.source_headers = source_headers or []
        self.base_data_mgr = base_data_mgr
        self.title(title or "导出格式设置")
        self.geometry("760x560")

        self.formats = load_export_formats()
        self.module = _ensure_module(self.formats, self.module_key)

        self.format_name_var = tk.StringVar(value=self.module.get("active") or "")
        self.use_original_var = tk.BooleanVar(value=bool(self.module.get("use_original")))

        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=10)
        ttk.Label(top, text="格式名称:").pack(side="left")
        self.format_combo = ttk.Combobox(
            top,
            textvariable=self.format_name_var,
            values=sorted(self.module["formats"].keys()),
            width=30
        )
        self.format_combo.pack(side="left", padx=6)
        self.format_combo.bind("<<ComboboxSelected>>", self._load_selected_format)
        ttk.Button(top, text="新建", command=self._new_format).pack(side="left", padx=4)
        ttk.Button(top, text="删除", command=self._delete_format).pack(side="left", padx=4)
        ttk.Button(top, text="导入配置", command=self._import_format_file).pack(side="left", padx=4)
        ttk.Button(top, text="导出配置", command=self._export_format_file).pack(side="left", padx=4)

        ttk.Checkbutton(
            top,
            text="原格式导出(忽略映射)",
            variable=self.use_original_var
        ).pack(side="right")

        table_frame = ttk.Frame(self)
        table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        columns = ("output", "source", "default")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=12)
        self.tree.heading("output", text="输出表头")
        self.tree.heading("source", text="源字段")
        self.tree.heading("default", text="默认值")
        self.tree.column("output", width=200, minwidth=100)
        self.tree.column("source", width=200, minwidth=100)
        self.tree.column("default", width=160, minwidth=80)
        self.tree.pack(side="left", fill="both", expand=True)

        scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        scroll.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self.tree.bind("<ButtonPress-1>", self._on_drag_start, add="+")
        self.tree.bind("<B1-Motion>", self._on_drag_motion, add="+")
        self.tree.bind("<ButtonRelease-1>", self._on_drag_end, add="+")
        self.tree.bind("<Button-3>", self._show_tree_menu, add="+")
        self._drag_item = None

        form = ttk.Frame(self)
        form.pack(fill="x", padx=10, pady=(0, 8))
        ttk.Label(form, text="输出表头:").grid(row=0, column=0, sticky="e")
        self.output_entry = ttk.Entry(form, width=30)
        self.output_entry.grid(row=0, column=1, sticky="w", padx=5)
        ttk.Label(form, text="源字段:").grid(row=0, column=2, sticky="e")
        self.base_data_placeholder = "<选择基础数据字段>"
        self.base_data_quick_options = [
            "BD:account_subject|name|科目编码",
            "BD:account_subject|code|科目编码",
            "BD:business_partner|name|往来单位编码",
            "BD:business_partner|local_code|往来单位编码",
            "BD:business_partner|bank_account|往来单位编码",
            "BD:product|name|品目编码",
            "BD:bank_account|name|默认账户",
        ]
        source_options = list(self.source_headers) + self.base_data_quick_options + [self.base_data_placeholder]
        self.source_combo = ttk.Combobox(form, values=source_options, width=28)
        self.source_combo.grid(row=0, column=3, sticky="w", padx=5)
        self.source_combo.bind("<<ComboboxSelected>>", self._on_source_combo_selected)
        self._build_tree_menu()
        ttk.Label(form, text="默认值:").grid(row=0, column=4, sticky="e")
        self.default_entry = ttk.Entry(form, width=18)
        self.default_entry.grid(row=0, column=5, sticky="w", padx=5)
        ttk.Button(form, text="常用默认值", command=self._show_default_presets).grid(
            row=0, column=8, padx=4
        )
        ttk.Button(form, text="基础数据默认...", command=self._open_base_data_default_selector).grid(
            row=0, column=7, padx=4
        )
        ttk.Button(form, text="基础数据...", command=self._open_base_data_selector).grid(
            row=0, column=6, padx=4
        )

        form_btns = ttk.Frame(self)
        form_btns.pack(fill="x", padx=10, pady=(0, 8))
        ttk.Button(form_btns, text="新增/更新", command=self._add_or_update_row).pack(side="left")
        ttk.Button(form_btns, text="删除选中", command=self._remove_selected).pack(side="left", padx=6)
        ttk.Button(form_btns, text="清空", command=self._clear_form).pack(side="left", padx=6)
        ttk.Button(form_btns, text="新增表头", command=self._prompt_add_header).pack(side="left", padx=6)
        ttk.Button(form_btns, text="从模板识别", command=self._load_from_template).pack(side="right")

        src_frame = ttk.LabelFrame(self, text="可用字段(只读)")
        src_frame.pack(fill="x", padx=10, pady=8)
        src_text = scrolledtext.ScrolledText(src_frame, height=4, wrap="word")
        src_text.pack(fill="x", expand=True, padx=6, pady=6)
        src_text.insert("1.0", ", ".join(self.source_headers))
        src_text.config(state="disabled")

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=8)
        ttk.Button(btns, text="验证映射", command=self._validate_mapping).pack(side="left")
        ttk.Button(btns, text="样本预览", command=self._preview_with_sample).pack(side="left", padx=6)
        ttk.Button(btns, text="保存并启用", command=self._save_and_close).pack(side="right")
        ttk.Button(btns, text="保存", command=self._save_only).pack(side="right", padx=6)
        ttk.Button(btns, text="取消", command=self.destroy).pack(side="right")

        self._load_selected_format()

    def _set_table_rows(self, mapping: list):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in mapping:
            self.tree.insert(
                "",
                "end",
                values=(row.get("output", ""), row.get("source", ""), row.get("default", ""))
            )

    def _get_table_rows(self) -> list:
        mapping = []
        for item in self.tree.get_children():
            output, source, default = self.tree.item(item, "values")
            output = (output or "").strip()
            if not output:
                continue
            mapping.append({"output": output, "source": source or "", "default": default or ""})
        return mapping

    def _on_tree_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        output, source, default = self.tree.item(sel[0], "values")
        self.output_entry.delete(0, "end")
        self.output_entry.insert(0, output or "")
        self.source_combo.set(source or "")
        self.default_entry.delete(0, "end")
        self.default_entry.insert(0, default or "")

    def _on_source_combo_selected(self, event=None):
        if self.source_combo.get() == self.base_data_placeholder:
            self._open_base_data_selector()

    def _show_default_presets(self):
        menu = tk.Menu(self, tearoff=0)
        presets = [
            ("系统日期(YYYYMMDD)", "{TODAY_YYYYMMDD}"),
            ("系统日期(YYYY-MM-DD)", "{TODAY_YYYY-MM-DD}"),
            ("系统时间(HHMMSS)", "{NOW_HHMMSS}"),
            ("系统时间(HH:MM:SS)", "{NOW_HH:MM:SS}"),
            ("系统日期时间(YYYYMMDD_HHMMSS)", "{NOW_YYYYMMDD_HHMMSS}"),
        ]
        for label, token in presets:
            menu.add_command(
                label=label,
                command=lambda t=token: self._insert_default_token(t)
            )
        menu.add_separator()
        menu.add_command(label="清空默认值", command=lambda: self.default_entry.delete(0, "end"))
        try:
            x = self.default_entry.winfo_rootx()
            y = self.default_entry.winfo_rooty() + self.default_entry.winfo_height()
            menu.tk_popup(x, y)
        finally:
            menu.grab_release()

    def _insert_default_token(self, token):
        self.default_entry.delete(0, "end")
        self.default_entry.insert(0, token)

    def _add_or_update_row(self):
        output = self.output_entry.get().strip()
        if not output:
            messagebox.showwarning("提示", "输出表头不能为空。")
            return
        source = self.source_combo.get().strip()
        default = self.default_entry.get().strip()
        sel = self.tree.selection()
        if sel:
            self.tree.item(sel[0], values=(output, source, default))
        else:
            self.tree.insert("", "end", values=(output, source, default))

    def _remove_selected(self):
        for item in self.tree.selection():
            self.tree.delete(item)

    def _clear_form(self):
        self.output_entry.delete(0, "end")
        self.source_combo.set("")
        self.default_entry.delete(0, "end")
        self.tree.selection_remove(self.tree.selection())

    def _prompt_add_header(self):
        name = simpledialog.askstring("新增表头", "请输入新表头名称：", parent=self)
        if not name:
            return
        header = name.strip()
        if not header:
            return
        self.tree.insert("", "end", values=(header, "", ""))

    def _build_tree_menu(self):
        self.tree_menu = tk.Menu(self, tearoff=0)
        self.tree_menu.add_command(label="设置基础数据字段...", command=self._apply_base_data_to_selected)
        self.tree_menu.add_separator()
        for option in self.base_data_quick_options:
            self.tree_menu.add_command(
                label=option,
                command=lambda opt=option: self._apply_base_data_token(opt)
            )

    def _show_tree_menu(self, event):
        row_id = self.tree.identify_row(event.y)
        if row_id:
            self.tree.selection_set(row_id)
        try:
            self.tree_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.tree_menu.grab_release()

    def _apply_base_data_to_selected(self):
        if not self.tree.selection():
            return
        self._open_base_data_selector()
        token = self.source_combo.get().strip()
        if token.startswith("BD:"):
            self._apply_base_data_token(token)

    def _apply_base_data_token(self, token: str):
        sel = self.tree.selection()
        if not sel:
            return
        output, _, default = self.tree.item(sel[0], "values")
        self.tree.item(sel[0], values=(output, token, default))

    def _validate_mapping(self):
        mapping = self._get_table_rows()
        missing = []
        for item in mapping:
            src = (item.get("source") or "").strip()
            if not src or src.startswith("BD:") or src.startswith("BDV:"):
                continue
            if src not in self.source_headers:
                missing.append(src)
        if missing:
            messagebox.showwarning("校验结果", f"以下源字段不存在：\n{', '.join(sorted(set(missing)))}")
        else:
            messagebox.showinfo("校验结果", "映射校验通过。")

    def _preview_with_sample(self):
        mapping = self._get_table_rows()
        if not mapping:
            messagebox.showwarning("提示", "映射为空，无法预览。")
            return
        file_path = filedialog.askopenfilename(
            title="选择样本Excel",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        try:
            df = pd.read_excel(file_path, header=0)
            if df.columns.size > 0 and "公司名称" in str(df.columns[0]):
                df = pd.read_excel(file_path, header=1)
        except Exception as exc:
            messagebox.showerror("错误", f"读取样本失败: {exc}")
            return
        if df.empty:
            messagebox.showwarning("提示", "样本为空。")
            return
        headers = [str(c) for c in df.columns]
        rows = df.head(20).values.tolist()
        mapped_headers, mapped_rows = apply_mapping(headers, rows, mapping, base_data_mgr=self.base_data_mgr)

        win = tk.Toplevel(self)
        win.title("样本预览")
        win.geometry("900x500")
        frame = ttk.Frame(win)
        frame.pack(fill="both", expand=True)
        tree = ttk.Treeview(frame, columns=mapped_headers, show="headings")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        for col in mapped_headers:
            tree.heading(col, text=col)
            tree.column(col, width=120)
        for row in mapped_rows:
            tree.insert("", "end", values=row)
    def _open_base_data_selector(self):
        dialog = tk.Toplevel(self)
        dialog.title("选择基础数据字段")
        dialog.geometry("520x420")
        dialog.transient(self)
        dialog.grab_set()

        table_map = {
            "科目编码": "account_subject",
            "往来单位": "business_partner",
            "品目": "product",
            "账户": "bank_account",
            "币种": "currency",
            "部门": "department",
            "仓库": "warehouse",
        }
        try:
            mgr = self.base_data_mgr or BaseDataManager()
            for cat in mgr.list_custom_categories():
                label = cat.get("display_name") or cat.get("name")
                name_key = cat.get("name")
                if label and name_key:
                    table_map[label] = f"custom:{name_key}"
            if mgr is not self.base_data_mgr:
                mgr.close()
        except Exception:
            pass
        table_var = tk.StringVar(value="科目编码")
        mode_var = tk.StringVar(value="by_source")
        target_var = tk.StringVar()
        key_var = tk.StringVar()
        search_var = tk.StringVar()
        record_list = []

        ttk.Label(dialog, text="基础数据表:").grid(row=0, column=0, padx=10, pady=8, sticky="e")
        table_combo = ttk.Combobox(dialog, values=list(table_map.keys()), textvariable=table_var, state="readonly")
        table_combo.grid(row=0, column=1, padx=6, pady=8, sticky="w")

        mode_frame = ttk.Frame(dialog)
        mode_frame.grid(row=1, column=0, columnspan=2, padx=10, sticky="w")
        ttk.Radiobutton(mode_frame, text="按源字段匹配", variable=mode_var, value="by_source").pack(side="left")
        ttk.Radiobutton(mode_frame, text="选择具体数据", variable=mode_var, value="fixed").pack(side="left", padx=8)

        ttk.Label(dialog, text="目标字段:").grid(row=2, column=0, padx=10, pady=8, sticky="e")
        target_combo = ttk.Combobox(dialog, textvariable=target_var, state="readonly", width=26)
        target_combo.grid(row=2, column=1, padx=6, pady=8, sticky="w")

        ttk.Label(dialog, text="来源字段:").grid(row=3, column=0, padx=10, pady=8, sticky="e")
        key_combo = ttk.Combobox(dialog, values=self.source_headers, textvariable=key_var, state="readonly", width=26)
        key_combo.grid(row=3, column=1, padx=6, pady=8, sticky="w")

        ttk.Label(dialog, text="搜索:").grid(row=4, column=0, padx=10, pady=6, sticky="e")
        search_entry = ttk.Entry(dialog, textvariable=search_var, width=28)
        search_entry.grid(row=4, column=1, padx=6, pady=6, sticky="w")

        list_frame = ttk.LabelFrame(dialog, text="选择具体数据")
        list_frame.grid(row=5, column=0, columnspan=2, padx=10, pady=8, sticky="nsew")
        record_tree = ttk.Treeview(list_frame, columns=("code", "name"), show="headings", height=8)
        record_tree.heading("code", text="编码")
        record_tree.heading("name", text="名称")
        record_tree.column("code", width=140, anchor="w")
        record_tree.column("name", width=240, anchor="w")
        record_tree.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        list_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=record_tree.yview)
        list_scroll.pack(side="right", fill="y")
        record_tree.configure(yscrollcommand=list_scroll.set)

        dialog.grid_columnconfigure(1, weight=1)
        dialog.grid_rowconfigure(5, weight=1)
        status_var = tk.StringVar(value="共 0 条")
        ttk.Label(dialog, textvariable=status_var, foreground="gray").grid(
            row=6, column=0, columnspan=2, sticky="w", padx=10
        )

        def load_columns():
            table_name = table_map.get(table_var.get())
            if not table_name:
                target_combo["values"] = []
                return
            mgr = self.base_data_mgr
            try:
                if mgr is None:
                    mgr = BaseDataManager()
                cols = mgr.get_table_columns(table_name)
                cols = [c for c in cols if c not in ("id", "match_items")]
                target_combo["values"] = cols
                if cols:
                    target_var.set(cols[0])
                records = mgr.query(table_name)
                if not records and hasattr(mgr, "import_all_data"):
                    try:
                        mgr.import_all_data()
                        records = mgr.query(table_name)
                    except Exception:
                        pass
                _reload_records(records)
            except Exception:
                target_combo["values"] = []
                _reload_records([])
            finally:
                if mgr and mgr is not self.base_data_mgr:
                    try:
                        mgr.close()
                    except Exception:
                        pass

        def _reload_records(records):
            nonlocal record_list
            record_list = []
            record_tree.delete(*record_tree.get_children())
            for row in records:
                code = str(row.get("code", "")).strip()
                name = str(row.get("name", "")).strip()
                record_list.append({"code": code, "name": name})
                record_tree.insert("", "end", values=(code, name))
            status_var.set(f"共 {len(record_list)} 条")

        def _filter_records(*args):
            term = search_var.get().strip().lower()
            record_tree.delete(*record_tree.get_children())
            for row in record_list:
                label = f"{row.get('code', '')} {row.get('name', '')}".lower()
                if not term or term in label:
                    record_tree.insert("", "end", values=(row.get("code", ""), row.get("name", "")))

        def _import_base_data():
            mgr = self.base_data_mgr
            try:
                if mgr is None:
                    mgr = BaseDataManager()
                if hasattr(mgr, "import_all_data"):
                    mgr.import_all_data()
            finally:
                if mgr and mgr is not self.base_data_mgr:
                    try:
                        mgr.close()
                    except Exception:
                        pass
            load_columns()

        def _toggle_mode(*args):
            if mode_var.get() == "fixed":
                key_combo.configure(state="disabled")
            else:
                key_combo.configure(state="readonly")

        def confirm():
            table_name = table_map.get(table_var.get())
            target_col = target_var.get().strip()
            if not table_name or not target_col:
                messagebox.showwarning("提示", "请选择基础数据表和目标字段。")
                return
            if mode_var.get() == "fixed":
                sel = record_tree.selection()
                if not sel:
                    messagebox.showwarning("提示", "请选择具体数据。")
                    return
                values = record_tree.item(sel[0], "values") or []
                code = values[0] if values else None
                if not code:
                    messagebox.showwarning("提示", "无法识别选中的数据。")
                    return
                token = f"BDV:{table_name}|{target_col}|code|{code}"
            else:
                key_header = key_var.get().strip()
                if not key_header:
                    messagebox.showwarning("提示", "请选择来源字段。")
                    return
                token = f"BD:{table_name}|{target_col}|{key_header}"
            self.source_combo.set(token)
            dialog.destroy()

        table_combo.bind("<<ComboboxSelected>>", lambda e: load_columns())
        search_var.trace_add("write", _filter_records)
        mode_var.trace_add("write", _toggle_mode)
        load_columns()
        _toggle_mode()

        def import_base_data():
            mgr = self.base_data_mgr
            try:
                if mgr is None:
                    mgr = BaseDataManager()
                if hasattr(mgr, "import_all_data"):
                    mgr.import_all_data()
            finally:
                if mgr and mgr is not self.base_data_mgr:
                    try:
                        mgr.close()
                    except Exception:
                        pass
            load_columns()

        btns = ttk.Frame(dialog)
        btns.grid(row=7, column=0, columnspan=2, pady=10)
        ttk.Button(btns, text="刷新", command=load_columns).pack(side="left", padx=4)
        ttk.Button(btns, text="导入基础数据", command=import_base_data).pack(side="left", padx=4)
        ttk.Button(btns, text="确定", command=confirm).pack(side="right", padx=4)
        ttk.Button(btns, text="取消", command=dialog.destroy).pack(side="right")

    def _open_base_data_default_selector(self):
        dialog = tk.Toplevel(self)
        dialog.title("选择基础数据默认值")
        dialog.geometry("520x420")
        dialog.transient(self)
        dialog.grab_set()

        table_map = {
            "科目编码": "account_subject",
            "往来单位": "business_partner",
            "品目": "product",
            "账户": "bank_account",
            "币种": "currency",
            "部门": "department",
            "仓库": "warehouse",
        }
        try:
            mgr = self.base_data_mgr or BaseDataManager()
            for cat in mgr.list_custom_categories():
                label = cat.get("display_name") or cat.get("name")
                name_key = cat.get("name")
                if label and name_key:
                    table_map[label] = f"custom:{name_key}"
            if mgr is not self.base_data_mgr:
                mgr.close()
        except Exception:
            pass
        table_var = tk.StringVar(value="科目编码")
        target_var = tk.StringVar()
        search_var = tk.StringVar()
        record_list = []

        ttk.Label(dialog, text="基础数据表:").grid(row=0, column=0, padx=10, pady=8, sticky="e")
        table_combo = ttk.Combobox(dialog, values=list(table_map.keys()), textvariable=table_var, state="readonly")
        table_combo.grid(row=0, column=1, padx=6, pady=8, sticky="w")

        ttk.Label(dialog, text="目标字段:").grid(row=1, column=0, padx=10, pady=8, sticky="e")
        target_combo = ttk.Combobox(dialog, textvariable=target_var, state="readonly", width=26)
        target_combo.grid(row=1, column=1, padx=6, pady=8, sticky="w")

        ttk.Label(dialog, text="搜索:").grid(row=2, column=0, padx=10, pady=6, sticky="e")
        search_entry = ttk.Entry(dialog, textvariable=search_var, width=28)
        search_entry.grid(row=2, column=1, padx=6, pady=6, sticky="w")

        list_frame = ttk.LabelFrame(dialog, text="选择具体数据")
        list_frame.grid(row=3, column=0, columnspan=2, padx=10, pady=8, sticky="nsew")
        record_tree = ttk.Treeview(list_frame, columns=("code", "name"), show="headings", height=10)
        record_tree.heading("code", text="编码")
        record_tree.heading("name", text="名称")
        record_tree.column("code", width=140, anchor="w")
        record_tree.column("name", width=240, anchor="w")
        record_tree.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        list_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=record_tree.yview)
        list_scroll.pack(side="right", fill="y")
        record_tree.configure(yscrollcommand=list_scroll.set)

        dialog.grid_columnconfigure(1, weight=1)
        dialog.grid_rowconfigure(3, weight=1)
        status_var = tk.StringVar(value="共 0 条")
        ttk.Label(dialog, textvariable=status_var, foreground="gray").grid(
            row=4, column=0, columnspan=2, sticky="w", padx=10
        )

        def load_columns():
            table_name = table_map.get(table_var.get())
            if not table_name:
                target_combo["values"] = []
                return
            mgr = self.base_data_mgr
            try:
                if mgr is None:
                    mgr = BaseDataManager()
                cols = mgr.get_table_columns(table_name)
                cols = [c for c in cols if c not in ("id", "match_items")]
                target_combo["values"] = cols
                if cols:
                    target_var.set(cols[0])
                records = mgr.query(table_name)
                if not records and hasattr(mgr, "import_all_data"):
                    try:
                        mgr.import_all_data()
                        records = mgr.query(table_name)
                    except Exception:
                        pass
                _reload_records(records)
            except Exception:
                target_combo["values"] = []
                _reload_records([])
            finally:
                if mgr and mgr is not self.base_data_mgr:
                    try:
                        mgr.close()
                    except Exception:
                        pass

        def _reload_records(records):
            nonlocal record_list
            record_list = []
            record_tree.delete(*record_tree.get_children())
            for row in records:
                code = str(row.get("code", "")).strip()
                name = str(row.get("name", "")).strip()
                record_list.append({"code": code, "name": name})
                record_tree.insert("", "end", values=(code, name))
            status_var.set(f"共 {len(record_list)} 条")

        def _filter_records(*args):
            term = search_var.get().strip().lower()
            record_tree.delete(*record_tree.get_children())
            for row in record_list:
                label = f"{row.get('code', '')} {row.get('name', '')}".lower()
                if not term or term in label:
                    record_tree.insert("", "end", values=(row.get("code", ""), row.get("name", "")))

        def confirm():
            table_name = table_map.get(table_var.get())
            target_col = target_var.get().strip()
            if not table_name or not target_col:
                messagebox.showwarning("提示", "请选择基础数据表和目标字段。")
                return
            sel = record_tree.selection()
            if not sel:
                messagebox.showwarning("提示", "请选择具体数据。")
                return
            values = record_tree.item(sel[0], "values") or []
            code = values[0] if values else None
            if not code:
                messagebox.showwarning("提示", "无法识别选中的数据。")
                return
            token = f"BDV:{table_name}|{target_col}|code|{code}"
            self.default_entry.delete(0, "end")
            self.default_entry.insert(0, token)
            dialog.destroy()

        table_combo.bind("<<ComboboxSelected>>", lambda e: load_columns())
        search_var.trace_add("write", _filter_records)
        load_columns()

        btns = ttk.Frame(dialog)
        btns.grid(row=5, column=0, columnspan=2, pady=12)
        ttk.Button(btns, text="刷新", command=load_columns).pack(side="left", padx=4)
        ttk.Button(btns, text="导入基础数据", command=_import_base_data).pack(side="left", padx=4)
        ttk.Button(btns, text="确定", command=confirm).pack(side="right", padx=6)
        ttk.Button(btns, text="取消", command=dialog.destroy).pack(side="right")

    def _on_drag_start(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self._drag_item = item

    def _on_drag_motion(self, event):
        if not self._drag_item:
            return
        target = self.tree.identify_row(event.y)
        if not target or target == self._drag_item:
            return
        target_index = self.tree.index(target)
        self.tree.move(self._drag_item, "", target_index)

    def _on_drag_end(self, event):
        self._drag_item = None

    def _load_selected_format(self, event=None):
        name = self.format_name_var.get().strip()
        mapping = (self.module.get("formats") or {}).get(name) or []
        self._set_table_rows(mapping)

    def _new_format(self):
        self.format_name_var.set("")
        self._set_table_rows([])

    def _delete_format(self):
        name = self.format_name_var.get().strip()
        if not name:
            return
        if not messagebox.askyesno("确认", f"删除导出格式：{name}？"):
            return
        self.module["formats"].pop(name, None)
        if self.module.get("active") == name:
            self.module["active"] = ""
        self.format_combo["values"] = sorted(self.module["formats"].keys())
        self.format_name_var.set("")
        self._set_table_rows([])
        save_export_formats(self.formats)

    def _import_format_file(self):
        file_path = filedialog.askopenfilename(
            title="选择导出格式配置文件",
            filetypes=[("JSON 配置文件", "*.json"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            
            # 兼容性检查：如果是直接的映射列表，则让用户起个名字
            if isinstance(data, list):
                name = simpledialog.askstring("导入配置", "请为导入的格式起个名称：", parent=self)
                if not name: return
                mapping = data
            elif isinstance(data, dict) and "mapping" in data:
                name = data.get("name") or simpledialog.askstring("导入配置", "请输入格式名称：", parent=self)
                if not name: return
                mapping = data["mapping"]
            else:
                messagebox.showerror("错误", "无效的配置文件格式。")
                return

            if not mapping:
                messagebox.showwarning("警告", "配置文件中的映射内容为空。")
                return

            self.module["formats"][name] = mapping
            self.format_combo["values"] = sorted(self.module["formats"].keys())
            self.format_name_var.set(name)
            self._set_table_rows(mapping)
            messagebox.showinfo("成功", f"格式 '{name}' 已导入。")
        except Exception as e:
            messagebox.showerror("导入失败", f"无法读取配置文件：{e}")

    def _export_format_file(self):
        name = self.format_name_var.get().strip()
        mapping = self._get_table_rows()
        if not mapping:
            messagebox.showwarning("提示", "当前映射为空，无法导出。")
            return
        
        default_file = f"导出格式_{name}_{self.module_key}.json" if name else f"导出格式_{self.module_key}.json"
        file_path = filedialog.asksaveasfilename(
            title="导出映射配置",
            initialfile=default_file,
            defaultextension=".json",
            filetypes=[("JSON 配置文件", "*.json")]
        )
        if not file_path:
            return
        
        export_data = {
            "name": name,
            "module": self.module_key,
            "mapping": mapping,
            "export_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(export_data, f, ensure_ascii=False, indent=4)
            messagebox.showinfo("成功", f"配置已导出至：\n{file_path}")
        except Exception as e:
            messagebox.showerror("导出失败", f"保存文件时出错：{e}")

    def _save_only(self):
        name = self.format_name_var.get().strip()
        if not name:
            messagebox.showwarning("提示", "请填写格式名称。")
            return
        mapping = self._get_table_rows()
        if not mapping:
            messagebox.showwarning("提示", "映射为空，至少需要一行。")
            return
        self.module["formats"][name] = mapping
        self.module["use_original"] = bool(self.use_original_var.get())
        self.format_combo["values"] = sorted(self.module["formats"].keys())
        save_export_formats(self.formats)

    def _save_and_close(self):
        self._save_only()
        name = self.format_name_var.get().strip()
        if name:
            self.module["active"] = name
            save_export_formats(self.formats)
        self.destroy()

    def _load_from_template(self):
        template_path = filedialog.askopenfilename(
            title="选择模板文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm"), ("所有文件", "*.*")]
        )
        if not template_path:
            return
        try:
            template_headers = read_template_headers(template_path)
        except Exception as exc:
            messagebox.showerror("错误", f"读取模板失败: {exc}")
            return
        mapping = build_template_mapping(self.source_headers, template_headers)
        if not mapping:
            messagebox.showwarning("提示", "模板未读取到有效表头。")
            return
        self._set_table_rows(mapping)


def open_export_format_editor(parent, module_key: str, source_headers: list, title: str = None, base_data_mgr=None):
    editor = ExportFormatEditor(parent, module_key, source_headers, title=title, base_data_mgr=base_data_mgr)
    editor.transient(parent)
    editor.grab_set()
    parent.wait_window(editor)


def open_export_format_center(parent, base_data_mgr=None):
    module_labels = {
        "main_export": "凭证转换",
        "summary_match_export": "摘要匹配",
        "shipping_product": "报关清单-产品",
        "shipping_container": "报关清单-货柜",
        "image_recognition": "图片识别",
    }
    win = tk.Toplevel(parent)
    win.title("导出格式中心")
    win.geometry("520x380")
    win.transient(parent)
    win.grab_set()

    tree = ttk.Treeview(win, columns=("module", "active", "original", "sample"), show="headings")
    tree.heading("module", text="模块")
    tree.heading("active", text="当前格式")
    tree.heading("original", text="原格式")
    tree.heading("sample", text="样本")
    tree.column("module", width=200)
    tree.column("active", width=180)
    tree.column("original", width=80)
    tree.column("sample", width=220)
    tree.pack(fill="both", expand=True, padx=10, pady=10)

    def refresh():
        for item in tree.get_children():
            tree.delete(item)
        formats = load_export_formats()
        samples = load_export_format_samples()
        for key, label in module_labels.items():
            module = formats.get(key, {})
            active = module.get("active", "")
            original = "是" if module.get("use_original") else ""
            sample_path = samples.get(key, "")
            tree.insert("", "end", values=(label, active, original, sample_path), tags=(key,))

    def get_selected_key():
        sel = tree.selection()
        if not sel:
            return None
        tag = tree.item(sel[0], "tags")
        return tag[0] if tag else None

    def edit_format():
        key = get_selected_key()
        if not key:
            return
        open_export_format_editor(
            win,
            key,
            source_headers=[],
            title=f"导出格式设置 - {module_labels.get(key, key)}",
            base_data_mgr=base_data_mgr
        )
        refresh()

    def _resolve_headers_from_sample(sample_path: str):
        if not sample_path:
            return []
        try:
            df = pd.read_excel(sample_path, header=0)
            if df.columns.size > 0 and "公司名称" in str(df.columns[0]):
                df = pd.read_excel(sample_path, header=1)
            return [str(c) for c in df.columns]
        except Exception:
            return []

    def edit_with_sample():
        key = get_selected_key()
        if not key:
            return
        samples = load_export_format_samples()
        file_path = samples.get(key, "")
        if not file_path:
            file_path = filedialog.askopenfilename(
                title="选择样本Excel",
                filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
            )
            if file_path:
                samples[key] = file_path
                save_export_format_samples(samples)
        headers = _resolve_headers_from_sample(file_path)
        open_export_format_editor(
            win,
            key,
            source_headers=headers,
            title=f"导出格式设置 - {module_labels.get(key, key)}",
            base_data_mgr=base_data_mgr
        )
        refresh()

    def choose_sample():
        key = get_selected_key()
        if not key:
            return
        file_path = filedialog.askopenfilename(
            title="选择样本Excel",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        samples = load_export_format_samples()
        samples[key] = file_path
        save_export_format_samples(samples)
        refresh()

    def clear_sample():
        key = get_selected_key()
        if not key:
            return
        samples = load_export_format_samples()
        samples.pop(key, None)
        save_export_format_samples(samples)
        refresh()

    btns = ttk.Frame(win)
    btns.pack(fill="x", padx=10, pady=(0, 10))
    ttk.Button(btns, text="编辑格式", command=edit_format).pack(side="left")
    ttk.Button(btns, text="样本编辑", command=edit_with_sample).pack(side="left", padx=6)
    ttk.Button(btns, text="选择样本", command=choose_sample).pack(side="left", padx=6)
    ttk.Button(btns, text="清除样本", command=clear_sample).pack(side="left", padx=6)
    ttk.Button(btns, text="刷新", command=refresh).pack(side="right")
    ttk.Button(btns, text="关闭", command=win.destroy).pack(side="right", padx=6)

    refresh()
