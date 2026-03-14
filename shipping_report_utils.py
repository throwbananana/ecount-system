# -*- coding: utf-8 -*-
"""
报关模块报表增强工具
提供图表生成和数据汇总功能。
"""

import pandas as pd
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter

def add_charts_to_product_report(file_path):
    """
    为产品导出报表增加图表：
    1. 产品货值 Top 10 柱状图
    2. 厂家供货占比饼图
    3. 体积与重量散点图
    """
    try:
        df = pd.read_excel(file_path)
        if df.empty:
            return

        # 辅助识别列名（处理由于映射导致的中文名变化）
        def find_col(keywords):
            for col in df.columns:
                if any(k in str(col) for k in keywords):
                    return col
            return None

        col_name = find_col(["品名", "名称", "name"])
        col_amount = find_col(["总金额", "金额", "amount"])
        col_factory = find_col(["厂家", "factory"])
        col_volume = find_col(["总体积", "体积", "volume"])
        col_weight = find_col(["总毛重", "毛重", "weight"])

        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 创建一个汇总数据区（放在数据右侧）
        summary_start_col = ws.max_column + 2
        
        # --- 1. 产品 Top 10 汇总 ---
        if col_name and col_amount:
            top_df = df.groupby(col_name)[col_amount].sum().sort_values(ascending=False).head(10).reset_index()
            row_idx = 2
            ws.cell(row=1, column=summary_start_col).value = "产品Top10汇总"
            ws.cell(row=1, column=summary_start_col + 1).value = "总金额"
            for _, r in top_df.iterrows():
                ws.cell(row=row_idx, column=summary_start_col).value = r[col_name]
                ws.cell(row=row_idx, column=summary_start_col + 1).value = r[col_amount]
                row_idx += 1
            
            chart1 = BarChart()
            chart1.title = "产品货值 Top 10"
            chart1.y_axis.title = "金额"
            data = Reference(ws, min_col=summary_start_col+1, min_row=1, max_row=row_idx-1)
            cats = Reference(ws, min_col=summary_start_col, min_row=2, max_row=row_idx-1)
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)
            chart1.width = 15
            chart1.height = 8
            ws.add_chart(chart1, f"{get_column_letter(summary_start_col + 3)}2")

        # --- 2. 厂家占比 ---
        if col_factory and col_amount:
            fac_df = df.groupby(col_factory)[col_amount].sum().reset_index()
            row_idx = 15 # 往下挪一点
            ws.cell(row=row_idx, column=summary_start_col).value = "厂家汇总"
            ws.cell(row=row_idx, column=summary_start_col + 1).value = "总金额"
            start_row = row_idx
            row_idx += 1
            for _, r in fac_df.iterrows():
                ws.cell(row=row_idx, column=summary_start_col).value = r[col_factory]
                ws.cell(row=row_idx, column=summary_start_col + 1).value = r[col_amount]
                row_idx += 1
            
            chart2 = PieChart()
            chart2.title = "厂家供货金额占比"
            data = Reference(ws, min_col=summary_start_col+1, min_row=start_row, max_row=row_idx-1)
            cats = Reference(ws, min_col=summary_start_col, min_row=start_row+1, max_row=row_idx-1)
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(cats)
            chart2.width = 10
            chart2.height = 8
            ws.add_chart(chart2, f"{get_column_letter(summary_start_col + 3)}18")

        # --- 3. 散点图 (体积 vs 重量) ---
        if col_volume and col_weight:
            chart3 = ScatterChart()
            chart3.title = "体积 vs 重量 分布图"
            chart3.x_axis.title = "毛重"
            chart3.y_axis.title = "体积"
            
            # 直接引用原始数据列
            col_idx_v = list(df.columns).index(col_volume) + 1
            col_idx_w = list(df.columns).index(col_weight) + 1
            
            xvalues = Reference(ws, min_col=col_idx_w, min_row=2, max_row=ws.max_row)
            yvalues = Reference(ws, min_col=col_idx_v, min_row=2, max_row=ws.max_row)
            series = Series(yvalues, xvalues, title=str(col_volume))
            chart3.series.append(series)
            chart3.width = 15
            chart3.height = 8
            ws.add_chart(chart3, f"{get_column_letter(summary_start_col + 3)}34")

        wb.save(file_path)
    except Exception as e:
        print(f"生成图表失败: {e}")

def add_charts_to_container_report(file_path):
    """
    为货柜导出报表增加图表：
    1. 费用构成占比饼图
    2. 货柜退税额对比柱状图
    """
    try:
        df = pd.read_excel(file_path)
        if df.empty:
            return

        def find_col(keywords):
            for col in df.columns:
                if any(k in str(col).lower() for k in keywords):
                    return col
            return None

        col_cn = find_col(["货柜号", "container"])
        col_tax = find_col(["退税额", "tax_refund"])
        
        # 费用列
        col_sea = find_col(["海运费", "sea_freight"])
        col_all = find_col(["包干费", "all_in"])
        col_agency = find_col(["代理费", "agency"])
        col_misc = find_col(["杂费", "misc"])
        col_exchange = find_col(["汇率", "exchange"])

        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        summary_start_col = ws.max_column + 2

        # --- 1. 费用构成 ---
        fee_sum = {
            "海运费(RMB)": 0.0,
            "包干费": df[col_all].sum() if col_all else 0.0,
            "代理费": df[col_agency].sum() if col_agency else 0.0,
            "其他杂费": df[col_misc].sum() if col_misc else 0.0
        }
        if col_sea and col_exchange:
            fee_sum["海运费(RMB)"] = (df[col_sea] * df[col_exchange]).sum()
        elif col_sea:
             fee_sum["海运费"] = df[col_sea].sum()

        row_idx = 2
        ws.cell(row=1, column=summary_start_col).value = "费用类型"
        ws.cell(row=1, column=summary_start_col + 1).value = "金额"
        for k, v in fee_sum.items():
            ws.cell(row=row_idx, column=summary_start_col).value = k
            ws.cell(row=row_idx, column=summary_start_col + 1).value = v
            row_idx += 1
        
        chart1 = PieChart()
        chart1.title = "总体费用构成分析"
        data = Reference(ws, min_col=summary_start_col+1, min_row=1, max_row=row_idx-1)
        cats = Reference(ws, min_col=summary_start_col, min_row=2, max_row=row_idx-1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.width = 12
        chart1.height = 8
        ws.add_chart(chart1, f"{get_column_letter(summary_start_col + 3)}2")

        # --- 2. 各柜退税额对比 ---
        if col_cn and col_tax:
            chart2 = BarChart()
            chart2.title = "各货柜退税额对比"
            # 直接引用原表
            col_idx_cn = list(df.columns).index(col_cn) + 1
            col_idx_tax = list(df.columns).index(col_tax) + 1
            data = Reference(ws, min_col=col_idx_tax, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=col_idx_cn, min_row=2, max_row=ws.max_row)
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(cats)
            chart2.width = 18
            chart2.height = 8
            ws.add_chart(chart2, f"{get_column_letter(summary_start_col + 3)}18")

        wb.save(file_path)
    except Exception as e:
        print(f"生成货柜图表失败: {e}")
