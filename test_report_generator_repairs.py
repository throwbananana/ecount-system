import os
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook

from report_generator import ReportGenerator


def _series_title_token(series):
    title = getattr(series, "title", None)
    if title is None:
        return None
    if getattr(title, "v", None):
        return str(title.v)

    str_ref = getattr(title, "strRef", None)
    if str_ref is None:
        tx = getattr(title, "tx", None)
        str_ref = getattr(tx, "strRef", None) if tx is not None else None
    if str_ref is None:
        return None
    return str_ref.f


def _chart_title_text(chart):
    try:
        return chart.title.tx.rich.p[0].r[0].t
    except Exception:
        return ""


def _save_blank_workbook(path):
    wb = Workbook()
    wb.active["A1"] = "test"
    wb.save(path)


def test_fill_product_summary_aggregates_rows():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {'MonthStr': '2025-12', '品目名': 'A', '数量': 2, '合计': 100, '品目编码': '001'},
        {'MonthStr': '2025-12', '品目名': 'A', '数量': 3, '合计': 180, '品目编码': '001'},
        {'MonthStr': '2025-12', '品目名': 'B', '数量': 4, '合计': 160, '品目编码': '002'},
    ])
    gen.data['cost']['2025-12'] = pd.DataFrame({
        '品目编码': ['001', '002'],
        'dummy_减少.1': [40, 20],
    })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '按产品汇总(含合计数)'
    headers = [
        '产品',
        '2025-12_销售收入',
        '2025-12_销售数量',
        '2025-12_销售成本',
        '2025-12_销售利润',
        '2025-12_毛利率',
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    ws.cell(row=2, column=1).value = 'A'
    ws.cell(row=3, column=1).value = 'B'
    ws.cell(row=4, column=1).value = '合计'

    gen._fill_product_summary(ws, '2025', '12', 'current')

    # A: revenue=280 qty=5 cost=200 profit=80 margin=80/280
    assert ws.cell(2, 2).value == 280
    assert ws.cell(2, 3).value == 5
    assert ws.cell(2, 4).value == 200
    assert ws.cell(2, 5).value == 80
    assert abs(ws.cell(2, 6).value - (80 / 280)) < 1e-12

    # Total: A + B => revenue=440 qty=9 cost=280 profit=160
    assert ws.cell(4, 2).value == 440
    assert ws.cell(4, 3).value == 9
    assert ws.cell(4, 4).value == 280
    assert ws.cell(4, 5).value == 160


def test_list_available_months_uses_core_intersection_when_loaded():
    gen = ReportGenerator('.')
    gen.data['profit'] = {'2025-12': pd.DataFrame(), '2026-01': pd.DataFrame()}
    gen.data['cost'] = {'2025-12': pd.DataFrame()}
    gen.data['asset'] = {'2025-12': pd.DataFrame()}
    gen.data['expense'] = {'2026-01': pd.DataFrame()}
    gen.data['sales'] = {'2025-12': pd.DataFrame(), '2026-01': pd.DataFrame()}

    assert gen.list_available_months() == ['2025-12']
    assert gen.list_available_years() == [2025]


def test_build_month_range_supports_cross_year_continuous_batch():
    assert ReportGenerator.build_month_range('2025', '11', '2026', '02') == [
        '2025-11',
        '2025-12',
        '2026-01',
        '2026-02',
    ]

    try:
        ReportGenerator.build_month_range('2026', '03', '2026', '02')
    except ValueError as exc:
        assert "开始月份不能晚于结束月份" in str(exc)
    else:
        raise AssertionError("Expected ValueError for reversed month range")


def test_generate_continuous_batch_reports_uses_cross_year_output_names(tmp_path):
    gen = ReportGenerator('.')
    calls = []

    def fake_generate_report(template_path, output_path, target_year, target_month, **kwargs):
        calls.append((template_path, os.path.basename(output_path), target_year, target_month))
        return True

    gen.generate_report = fake_generate_report
    summary = gen.generate_continuous_batch_reports(
        "template.xlsx",
        str(tmp_path),
        "2025",
        "12",
        "2026",
        "01",
    )

    assert calls == [
        ("template.xlsx", "2025年12月_经营分析报告.xlsx", "2025", "12"),
        ("template.xlsx", "2026年01月_经营分析报告.xlsx", "2026", "01"),
    ]
    assert summary == [
        ("2025-12", os.path.join(str(tmp_path), "2025年12月_经营分析报告.xlsx"), "成功"),
        ("2026-01", os.path.join(str(tmp_path), "2026年01月_经营分析报告.xlsx"), "成功"),
    ]
    assert list(tmp_path.glob("连续批量生成摘要_*.txt"))


def test_inspect_data_folder_marks_old_duplicate_as_cleanup_candidate(tmp_path):
    old_profit = tmp_path / "利润表2026.01-2026.01_旧.xlsx"
    new_profit = tmp_path / "利润表2026.01-2026.01.xlsx"
    cost = tmp_path / "成本合计表2026.01-2026.01.xlsx"
    for path in [old_profit, new_profit, cost]:
        _save_blank_workbook(path)
    os.utime(old_profit, (1000, 1000))
    os.utime(new_profit, (2000, 2000))
    os.utime(cost, (1500, 1500))

    gen = ReportGenerator(str(tmp_path))
    inspection = gen.inspect_data_folder(
        required_categories=["profit", "cost"],
        expected_months=["2026-01", "2026-02"],
    )

    assert len(inspection["duplicates"]) == 1
    duplicate = inspection["duplicates"][0]
    assert duplicate["category"] == "profit"
    assert duplicate["period"] == "2026-01"
    assert duplicate["latest"]["filename"] == new_profit.name

    candidates = inspection["cleanup_candidates"]
    assert [item["filename"] for item in candidates] == [old_profit.name]
    assert candidates[0]["superseded_periods"] == ["2026-01"]
    assert inspection["missing"] == [
        {
            "period": "2026-02",
            "missing_categories": ["profit", "cost"],
            "missing_labels": ["利润表", "成本合计表"],
        }
    ]

    result = gen.delete_data_files([str(old_profit)])
    assert result["failed"] == []
    assert result["deleted"] == [str(old_profit)]
    assert not old_profit.exists()


def test_inspect_data_folder_keeps_partially_overlapped_range_file_for_manual_review(tmp_path):
    old_range = tmp_path / "销售出库明细2026.01-2026.03.xlsx"
    new_feb = tmp_path / "销售出库明细2026.02-2026.02.xlsx"
    for path in [old_range, new_feb]:
        _save_blank_workbook(path)
    os.utime(old_range, (1000, 1000))
    os.utime(new_feb, (2000, 2000))

    gen = ReportGenerator(str(tmp_path))
    inspection = gen.inspect_data_folder(
        required_categories=["sales"],
        expected_months=["2026-01", "2026-02", "2026-03"],
    )

    assert len(inspection["duplicates"]) == 1
    assert inspection["cleanup_candidates"] == []
    assert len(inspection["manual_review"]) == 1
    review = inspection["manual_review"][0]
    assert review["filename"] == old_range.name
    assert review["superseded_periods"] == ["2026-02"]
    assert review["blocking_periods"] == ["2026-01", "2026-03"]


def test_inspect_data_folder_treats_profit_range_header_as_period_end_month(tmp_path):
    profit_path = tmp_path / "利润表2026.04-2026.04.xlsx"
    wb = Workbook()
    wb.active["A1"] = "2026/01/01-2026/04/30"
    wb.save(profit_path)

    gen = ReportGenerator(str(tmp_path))
    inspection = gen.inspect_data_folder(required_categories=["profit"])

    assert inspection["files"][0]["periods"] == ["2026-04"]
    assert inspection["expected_months"] == ["2026-04"]
    assert inspection["missing"] == []


def test_apply_output_sheet_filter_hides_unselected_sheets():
    gen = ReportGenerator('.')
    wb = Workbook()
    wb.active.title = '目录'
    wb.create_sheet('仪表盘')
    wb.create_sheet('经营指标')

    hidden = gen._apply_output_sheet_filter(wb, ['仪表盘', '未生成Sheet'])

    assert wb.sheetnames == ['目录', '仪表盘', '经营指标']
    assert hidden == ['目录', '经营指标']
    assert wb['目录'].sheet_state == 'hidden'
    assert wb['仪表盘'].sheet_state == 'visible'
    assert wb['经营指标'].sheet_state == 'hidden'
    assert wb.active.title == '仪表盘'


def test_check_data_completeness_includes_sales_and_ar():
    gen = ReportGenerator('.')
    key = '2025-12'
    for cat in ['profit', 'cost', 'expense', 'asset', 'sales']:
        gen.data[cat][key] = pd.DataFrame({'x': [1]})

    missing = gen.check_data_completeness('2025', '12')
    assert missing == ['ar']

    gen.ar_detail_df = pd.DataFrame({'客户': ['A']})
    assert gen.check_data_completeness('2025', '12') == []


def test_load_ar_data_groups_cross_year_detail_by_transaction_month(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "公司名称 : 浙江宙恒进出口有限公司 / 2026/01/01  ~ 2026/12/31  / 科目账簿 / 1122(应收账款)"
    headers = [
        "日期-号码", "摘要", "科目名", "科目编码", "相对科目编码名", "相对科目编码",
        "往来单位编码", "往来单位名", "外币借方金额", "外币贷方金额", "借方金额", "贷方金额", "余额",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=idx).value = header
    ws.append(["2025/12/31 -1", "年末销售", "应收账款", "1122", "主营业务收入", "6001", "C001", "客户A", 10, None, 70, None, 70])
    ws.append(["2026/01/05 -1", "新年销售", "应收账款", "1122", "主营业务收入", "6001", "C001", "客户A", 20, None, 140, None, 210])
    path = tmp_path / "应收账款2023-2026.xlsx"
    wb.save(path)

    gen = ReportGenerator(str(tmp_path))
    gen._load_ar_data(str(path), path.name)

    assert sorted(gen.data["ar"].keys()) == ["2025-12", "2026-01"]
    assert gen.ar_detail_df is not None
    assert set(gen.ar_detail_df["MonthStr"].unique()) == {"2025-12", "2026-01"}


def test_load_ar_data_recognizes_multiblock_customer_ledger(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "科目 往来单位明细账"
    headers = [
        "摘要", "相对科目编码名", "日期-号码", "对应往来单位编码", "对应往来单位名",
        "相对科目编码", "债权债务号码", "外币借方金额", "外币贷方金额", "外币余额",
        "借方金额", "贷方金额", "余额", "部门名",
    ]

    ws.append(["公司名称 : A / 应收账款 / 2026/01/01  ~ 2026/03/31  / 科目/客户/供应商明细账 / C001(客户A)"])
    ws.append(headers)
    ws.append(["期末调汇", "汇兑损益", "2026/01/31 -7", None, None, "660301", None, None, None, None, 10, None, 10, "销售部"])
    ws.append(["合计", None, None, None, None, None, None, None, None, None, 10, None, 10, None])
    ws.append(["2026/04/27 22:07:00"])
    ws.append(["公司名称 : A / 应收账款 / 2026/01/01  ~ 2026/03/31  / 科目/客户/供应商明细账 / C002(客户B)"])
    ws.append(headers)
    ws.append(["销售", "主营业务收入", "2026/02/01 -1", None, None, "6001", None, None, None, None, 20, None, 20, "销售部"])
    path = tmp_path / "应收_202601-202603.xlsx"
    wb.save(path)

    gen = ReportGenerator(str(tmp_path))
    gen._load_ar_data(str(path), path.name)

    assert sorted(gen.data["ar"].keys()) == ["2026-01", "2026-02"]
    assert "往来单位名" in gen.ar_detail_df.columns
    assert gen.data["ar"]["2026-01"]["往来单位名"].iloc[0] == "客户A"
    assert gen.data["ar"]["2026-02"]["往来单位名"].iloc[0] == "客户B"

    gen._run_data_quality_checks()
    assert not [
        issue for issue in gen.data_quality_issues
        if issue["category"] == "ar" and issue["issue_type"] == "客户/单位缺失"
    ]


def test_classify_source_file_uses_workbook_header_content(tmp_path):
    def make_book(filename, title):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = title
        ws["A2"] = "日期-号码"
        path = tmp_path / filename
        wb.save(path)
        return path

    gen = ReportGenerator(str(tmp_path))
    ap_path = make_book("ledger_a.xlsx", "公司名称 : A / 应付账款 / 2025/01/01  ~ 2025/01/31 / 科目/客户/供应商明细账")
    cash_path = make_book("ledger_b.xlsx", "公司名称 : A / 银行存款 / 2025/01/01  ~ 2025/01/31 / 科目/客户/供应商明细账")
    expense_path = make_book("ledger_c.xlsx", "公司名称 : A / 2025/01/01  ~ 2025/01/31 / 科目账簿 / 660101(办公费)")

    assert gen._classify_source_file(ap_path.name, str(ap_path)) == "ap"
    assert gen._classify_source_file(cash_path.name, str(cash_path)) == "cash"
    assert gen._classify_source_file(expense_path.name, str(expense_path)) == "expense"


def test_load_ap_and_cash_detail_groups_by_transaction_month(tmp_path):
    def make_ledger(path, title, amount_row):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = title
        headers = ["摘要", "相对科目编码名", "日期-号码", "对应往来单位名", "借方金额", "贷方金额", "余额"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=2, column=idx).value = header
        ws.append(amount_row)
        wb.save(path)

    ap_path = tmp_path / "应付明细.xlsx"
    cash_path = tmp_path / "银行明细.xlsx"
    make_ledger(
        ap_path,
        "公司名称 : A / 应付账款 / 2026/03/01  ~ 2026/03/31 / 科目/客户/供应商明细账",
        ["采购", "库存商品", "2026/03/05 -1", "供应商A", 0, 300, -300],
    )
    make_ledger(
        cash_path,
        "公司名称 : A / 银行存款 / 2026/03/01  ~ 2026/03/31 / 科目/客户/供应商明细账",
        ["收款", "主营业务收入", "2026/03/08 -1", "客户A", 500, 0, 500],
    )

    gen = ReportGenerator(str(tmp_path))
    gen._load_ap_data(str(ap_path), ap_path.name)
    gen._load_cash_data(str(cash_path), cash_path.name)

    assert sorted(gen.data["ap"].keys()) == ["2026-03"]
    assert sorted(gen.data["cash"].keys()) == ["2026-03"]
    assert gen.ap_detail_df is not None and len(gen.ap_detail_df) == 1
    assert gen.cash_detail_df is not None and len(gen.cash_detail_df) == 1


def test_load_expense_data_drops_export_footer_timestamp_rows(tmp_path):
    path = tmp_path / "费用_202601-202603.xlsx"
    df = pd.DataFrame([
        {
            "日期-号码": "2026/03/31 -1",
            "摘要": "办公用品",
            "科目名": "办公费",
            "科目编码": "660101",
            "借方金额": 100,
            "贷方金额": None,
            "余额": 100,
            "外币借方金额": None,
            "外币贷方金额": None,
        },
        {
            "日期-号码": "2026/04/27 22:09:08",
            "摘要": None,
            "科目名": None,
            "科目编码": None,
            "借方金额": None,
            "贷方金额": None,
            "余额": None,
            "外币借方金额": None,
            "外币贷方金额": None,
        },
    ])
    with pd.ExcelWriter(path) as writer:
        df.to_excel(writer, index=False)

    gen = ReportGenerator(str(tmp_path))
    gen._load_expense_data(str(path), path.name)
    gen._run_data_quality_checks()

    assert sorted(gen.data["expense"].keys()) == ["2026-03"]
    assert len(gen.data["expense"]["2026-03"]) == 1
    assert not [
        issue for issue in gen.data_quality_issues
        if issue["category"] == "expense" and issue["period"] == "2026-04"
    ]
    assert not [
        issue for issue in gen.data_quality_issues
        if issue["category"] == "expense" and issue["issue_type"] == "数值缺失" and "外币" in issue["detail"]
    ]


def test_load_expense_data_merges_same_month_from_multiple_files(tmp_path):
    first = tmp_path / "费用_A.xlsx"
    second = tmp_path / "费用_B.xlsx"
    pd.DataFrame([
        {"日期-号码": "2026/03/01 -1", "摘要": "办公用品", "科目名": "管理费用-办公费", "借方金额": 100, "贷方金额": 0},
    ]).to_excel(first, index=False)
    pd.DataFrame([
        {"日期-号码": "2026/03/02 -1", "摘要": "水费", "科目名": "管理费用-水电费", "借方金额": 200, "贷方金额": 0},
    ]).to_excel(second, index=False)

    gen = ReportGenerator(str(tmp_path))
    gen._load_expense_data(str(first), first.name)
    gen._load_expense_data(str(second), second.name)

    assert sorted(gen.data["expense"].keys()) == ["2026-03"]
    assert len(gen.data["expense"]["2026-03"]) == 2


def test_expense_behavior_totals_are_closed_with_unclassified_and_finance():
    gen = ReportGenerator('.')
    gen.data['expense']['2026-03'] = pd.DataFrame([
        {"日期": "2026-03-01", "科目名": "销售费用-运费", "借方金额": 100, "贷方金额": 0, "摘要": "运输"},
        {"日期": "2026-03-02", "科目名": "管理费用-工资", "借方金额": 200, "贷方金额": 0, "摘要": "工资"},
        {"日期": "2026-03-03", "科目名": "财务费用-汇兑损益", "借方金额": 0, "贷方金额": 30, "摘要": "期末调汇"},
        {"日期": "2026-03-04", "科目名": "管理费用-存货盘亏或盘盈", "借方金额": 900, "贷方金额": 100, "摘要": "盘点盘亏盘盈"},
        {"日期": "2026-03-05", "科目名": "销售费用-交际应酬费", "借方金额": 50, "贷方金额": 0, "摘要": "请客户吃饭"},
        {"日期": "2026-03-06", "科目名": "财务费用-其他", "借方金额": 10, "贷方金额": 0, "摘要": "u盾费用"},
        {"日期": "2026-03-07", "科目名": "管理费用-其他", "借方金额": 20, "贷方金额": 0, "摘要": "律师费用"},
        {"日期": "2026-03-04", "科目名": "管理费用-其他", "借方金额": 40, "贷方金额": 0, "摘要": "临时事项"},
    ])

    totals = gen._calculate_expense_behavior_totals('2026', '03', 'current')

    assert totals["variable"]["2026-03"] == 150
    assert totals["fixed"]["2026-03"] == 220
    assert totals["financial_adjustment"]["2026-03"] == -20
    assert totals["inventory_adjustment"]["2026-03"] == 800
    assert totals["unclassified"]["2026-03"] == 40
    assert totals["total"]["2026-03"] == 1190


def test_expense_structure_sheet_separates_inventory_adjustments_from_unclassified():
    gen = ReportGenerator('.')
    gen.data['expense']['2026-04'] = pd.DataFrame([
        {"日期": "2026-04-01", "科目名": "管理费用-存货盘亏或盘盈", "借方金额": 1000, "贷方金额": 200, "摘要": "盘点盘亏盘盈"},
        {"日期": "2026-04-02", "科目名": "销售费用-交际应酬费", "借方金额": 50, "贷方金额": 0, "摘要": "请客户吃饭"},
        {"日期": "2026-04-03", "科目名": "管理费用-其他", "借方金额": 40, "贷方金额": 0, "摘要": "临时事项"},
    ])
    wb = openpyxl.Workbook()

    gen._update_expense_structure_sheet(
        wb,
        {"2026-04": {"revenue": 2000}},
        "2026",
        "04",
        "current",
    )

    ws = wb["费用结构与弹性"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 15)]
    assert headers[:7] == [
        "月份",
        "变动费用",
        "固定费用",
        "财务/汇兑调整",
        "库存/盘点调整",
        "未分类费用",
        "费用合计",
    ]
    assert ws.cell(row=2, column=2).value == 50
    assert ws.cell(row=2, column=5).value == 800
    assert ws.cell(row=2, column=6).value == 40
    assert ws.cell(row=2, column=7).value == 890


def test_expense_behavior_infers_from_historical_subject_profile():
    gen = ReportGenerator('.')
    gen.data['expense']['2026-01'] = pd.DataFrame([
        {"日期": "2026-01-01", "科目名": "管理费用-专业服务费", "借方金额": 100, "贷方金额": 0, "摘要": "律师费用"},
        {"日期": "2026-01-02", "科目名": "管理费用-安保费", "借方金额": 100, "贷方金额": 0, "摘要": "服务合同"},
    ])
    gen.data['expense']['2026-02'] = pd.DataFrame([
        {"日期": "2026-02-01", "科目名": "管理费用-专业服务费", "借方金额": 200, "贷方金额": 0, "摘要": "年度服务"},
        {"日期": "2026-02-02", "科目名": "管理费用-安保费", "借方金额": 105, "贷方金额": 0, "摘要": "服务合同"},
    ])
    gen.data['expense']['2026-03'] = pd.DataFrame([
        {"日期": "2026-03-01", "科目名": "管理费用-安保费", "借方金额": 95, "贷方金额": 0, "摘要": "服务合同"},
        {"日期": "2026-03-02", "科目名": "管理费用-其他", "借方金额": 40, "贷方金额": 0, "摘要": "临时事项"},
    ])

    totals = gen._calculate_expense_behavior_totals('2026', '03', 'current')

    assert totals["fixed"]["2026-01"] == 200
    assert totals["fixed"]["2026-02"] == 305
    assert totals["fixed"]["2026-03"] == 95
    assert totals["unclassified"]["2026-03"] == 40


def test_expense_diagnostic_matrix_does_not_reuse_history_anomaly_for_target_month():
    gen = ReportGenerator('.')
    gen.data['expense']['2025-10'] = pd.DataFrame([
        {"日期": "2025-10-01", "科目名": "管理费用-办公费", "借方金额": 100, "贷方金额": 0, "部门名": "行政", "摘要": "常规办公"},
    ])
    gen.data['expense']['2025-11'] = pd.DataFrame([
        {"日期": "2025-11-01", "科目名": "管理费用-办公费", "借方金额": 20000, "贷方金额": 0, "部门名": "行政", "摘要": "集中采购"},
    ])
    gen.data['expense']['2025-12'] = pd.DataFrame([
        {"日期": "2025-12-01", "科目名": "管理费用-办公费", "借方金额": 20050, "贷方金额": 0, "部门名": "行政", "摘要": "常规办公"},
    ])
    metrics = {
        "2025-10": {"revenue": 100000},
        "2025-11": {"revenue": 100000},
        "2025-12": {"revenue": 100000},
    }

    wb = openpyxl.Workbook()
    wb.active.title = '费用明细环比分析'
    gen._update_expense_diagnostic_center(wb, metrics, '2025', '12', 'current')
    ws = wb['费用分析']

    matrix_header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "费用类别" and ws.cell(r, 10).value == "明细键":
            matrix_header_row = r
            break
    assert matrix_header_row is not None

    matrix_keys = []
    for r in range(matrix_header_row + 1, ws.max_row + 1):
        if str(ws.cell(r, 1).value).startswith("C. "):
            break
        key = ws.cell(r, 10).value
        if key:
            matrix_keys.append(str(key))
    assert all(key.startswith("2025-12|") for key in matrix_keys)


def test_write_table_sanitizes_formula_like_text():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active

    gen._write_table(ws, 1, 1, ["摘要"], [["=HYPERLINK(\"http://example.com\")"]])

    assert ws.cell(2, 1).value.startswith("'=")


def test_data_quality_ignores_blank_nan_keys_for_duplicate_checks():
    gen = ReportGenerator('.')
    gen.data['sales']['2026-01'] = pd.DataFrame({
        '日期-号码': ['2026/01/01 -1', '2026/01/02 -1', '2026/01/03 -1'],
        '销售订单号': [None, float('nan'), ''],
        '品目编码': ['A', 'B', 'C'],
        '数量': [1, 2, 3],
        '合计': [10, 20, 30],
        '销售金额合计': [10, 20, 30],
        '销售出库供应价合计': [5, 10, 15],
        '往来单位名': ['客户A', '客户B', '客户C'],
    })

    gen._run_data_quality_checks()

    assert not [
        issue for issue in gen.data_quality_issues
        if issue["category"] == "sales" and issue["issue_type"] == "单号重复"
    ]


def test_update_ap_and_cash_analysis_sheets():
    gen = ReportGenerator('.')
    gen.ap_detail_df = pd.DataFrame([
        {
            "ParsedDate": pd.Timestamp("2026-03-05"),
            "MonthStr": "2026-03",
            "对应往来单位名": "供应商A",
            "借方金额": 50,
            "贷方金额": 300,
            "余额": -250,
            "摘要": "采购",
        },
    ])
    gen.cash_detail_df = pd.DataFrame([
        {
            "ParsedDate": pd.Timestamp("2026-03-08"),
            "MonthStr": "2026-03",
            "相对科目编码名": "主营业务收入",
            "借方金额": 500,
            "贷方金额": 0,
            "余额": 500,
            "摘要": "收款",
        },
    ])
    metrics = {"2026-03": {"ap_balance": 250, "cash": 500}}
    wb = Workbook()
    wb.active.title = "现金流量表(估算)"
    wb.create_sheet("资金链预警")
    wb.create_sheet("资产负债表")

    gen._update_cash_balance_analysis_sheet(wb, metrics, "2026", "03", "current")
    gen._update_ap_analysis_sheet(wb, metrics, "2026", "03", "current")

    assert "货币资金分析" in wb.sheetnames
    assert "应付账款分析" in wb.sheetnames
    assert wb["货币资金分析"]["A1"].value == "月份"
    assert wb["货币资金分析"]["A2"].value == "2026/03"
    assert wb["应付账款分析"]["A1"].value == "月份"
    assert wb["应付账款分析"]["A2"].value == "2026/03"
    assert len(wb["货币资金分析"]._charts) >= 1
    assert len(wb["应付账款分析"]._charts) >= 1


def test_auto_warning_params_derive_from_inventory_and_cash_history():
    gen = ReportGenerator('.')
    gen.report_params["replenishment"].update({"manual": False})
    gen.report_params["cashflow"].update({"manual": False})
    gen.sales_df = pd.DataFrame([
        {"MonthStr": "2026-01", "品目编码": "A", "数量": 100},
        {"MonthStr": "2026-02", "品目编码": "A", "数量": 100},
        {"MonthStr": "2026-03", "品目编码": "A", "数量": 100},
        {"MonthStr": "2026-01", "品目编码": "B", "数量": 30},
        {"MonthStr": "2026-02", "品目编码": "B", "数量": 30},
        {"MonthStr": "2026-03", "品目编码": "B", "数量": 30},
    ])
    gen.data["cost"]["2026-01"] = pd.DataFrame({
        "品目编码": ["A", "B"],
        "库存_期末": [50, 120],
        "库存_期末.2": [1000, 2000],
    })
    gen.data["cost"]["2026-02"] = pd.DataFrame({
        "品目编码": ["A", "B"],
        "库存_期末": [50, 120],
        "库存_期末.2": [1500, 2500],
    })
    gen.data["cost"]["2026-03"] = pd.DataFrame({
        "品目编码": ["A", "B"],
        "库存_期末": [50, 120],
        "库存_期末.2": [2000, 3000],
    })
    metrics = {
        "2026-01": {
            "revenue": 1000,
            "cost": 1000,
            "ar_balance": 2000,
            "ap_balance": 1000,
            "cash": 300,
            "sales_expense": 100,
            "admin_expense": 100,
            "financial_expense": 100,
        },
        "2026-02": {
            "revenue": 1000,
            "cost": 1000,
            "ar_balance": 3000,
            "ap_balance": 1000,
            "cash": 200,
            "sales_expense": 100,
            "admin_expense": 100,
            "financial_expense": 100,
        },
        "2026-03": {
            "revenue": 1000,
            "cost": 1000,
            "ar_balance": 4000,
            "ap_balance": 1000,
            "cash": 100,
            "sales_expense": 100,
            "admin_expense": 100,
            "financial_expense": 100,
        },
    }

    repl_params, cash_params, notes = gen._resolve_warning_params(metrics, "2026", "03", "current")

    assert repl_params == {"lead_days": 15, "safety_days": 10, "window_months": 3}
    assert cash_params["dso_threshold"] == 126
    assert cash_params["dio_threshold"] == 162
    assert cash_params["ccc_threshold"] == 252
    assert cash_params["cash_coverage_threshold"] == 0.5
    assert "自动分析" in notes["replenishment"]
    assert "自动分析" in notes["cashflow"]


def test_replenishment_alert_uses_product_specific_auto_params():
    gen = ReportGenerator('.')
    gen.sales_df = pd.DataFrame([
        {"MonthStr": "2026-01", "品目编码": "A", "品目名": "稳定品", "数量": 30},
        {"MonthStr": "2026-02", "品目编码": "A", "品目名": "稳定品", "数量": 30},
        {"MonthStr": "2026-03", "品目编码": "A", "品目名": "稳定品", "数量": 30},
        {"MonthStr": "2026-01", "品目编码": "B", "品目名": "波动品", "数量": 10},
        {"MonthStr": "2026-02", "品目编码": "B", "品目名": "波动品", "数量": 80},
        {"MonthStr": "2026-03", "品目编码": "B", "品目名": "波动品", "数量": 10},
    ])
    gen.data["cost"]["2026-01"] = pd.DataFrame({
        "品目编码": ["A", "B"],
        "品目名": ["稳定品", "波动品"],
        "库存_期初": [30, 20],
        "库存_增加": [30, 0],
        "库存_减少": [30, 10],
        "库存_期末": [30, 10],
    })
    gen.data["cost"]["2026-02"] = pd.DataFrame({
        "品目编码": ["A", "B"],
        "品目名": ["稳定品", "波动品"],
        "库存_期初": [30, 10],
        "库存_增加": [30, 100],
        "库存_减少": [30, 80],
        "库存_期末": [30, 30],
    })
    gen.data["cost"]["2026-03"] = pd.DataFrame({
        "品目编码": ["A", "B"],
        "品目名": ["稳定品", "波动品"],
        "库存_期初": [30, 30],
        "库存_增加": [30, 0],
        "库存_减少": [55, 35],
        "库存_期末": [5, 5],
    })

    wb = Workbook()
    wb.active.title = "存货健康度"
    gen._update_replenishment_alert_sheet(
        wb,
        "2026",
        "03",
        "current",
        lead_days=20,
        safety_days=20,
        window_months=3,
        use_product_params=True,
    )

    ws = wb["补货预警"]
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    rows = {
        ws.cell(row=r, column=header_map["品目编码"]).value: r
        for r in range(2, ws.max_row + 1)
    }

    assert ws.cell(row=rows["A"], column=header_map["采购/生产周期(天)"]).value == 30
    assert ws.cell(row=rows["A"], column=header_map["安全库存天数"]).value == 10
    assert ws.cell(row=rows["B"], column=header_map["采购/生产周期(天)"]).value == 15
    assert ws.cell(row=rows["B"], column=header_map["安全库存天数"]).value == 30


def test_manual_warning_params_are_preserved_and_sanitized():
    gen = ReportGenerator('.')
    gen.report_params["replenishment"] = {
        "manual": True,
        "lead_days": "45",
        "safety_days": "25",
        "window_months": "4",
    }
    gen.report_params["cashflow"] = {
        "manual": True,
        "dso_threshold": "80",
        "dio_threshold": "160",
        "ccc_threshold": "140",
        "cash_coverage_threshold": "2.3",
    }

    repl_params, cash_params, notes = gen._resolve_warning_params({}, "2026", "03", "current")

    assert repl_params == {"lead_days": 45, "safety_days": 25, "window_months": 4}
    assert cash_params == {
        "dso_threshold": 80,
        "dio_threshold": 160,
        "ccc_threshold": 140,
        "cash_coverage_threshold": 2.3,
    }
    assert "manual" not in repl_params
    assert "manual" not in cash_params
    assert notes["replenishment"] == "参数来源：手动填入"
    assert notes["cashflow"] == "参数来源：手动填入"


def test_read_budget_targets_allows_missing_optional_target_columns():
    gen = ReportGenerator('.')
    wb = Workbook()
    ws = wb.active
    ws.title = "目标_预算"
    ws.cell(row=1, column=1).value = "月份"
    ws.cell(row=1, column=2).value = "主营业务收入目标"
    ws.cell(row=2, column=1).value = "2026/03"
    ws.cell(row=2, column=2).value = 1000

    targets, _, _ = gen._read_budget_targets(wb)

    assert targets["2026-03"]["revenue"] == 1000
    assert targets["2026-03"]["sales_rate"] is None
    assert targets["2026-03"]["financial_rate"] is None

    gen._update_budget_variance_sheet(
        wb,
        {"2026-03": {"revenue": 1200, "operating_profit": 100}},
        "2026",
        "03",
        "current",
    )
    assert wb["预算执行与偏差"].cell(row=2, column=1).value == "2026/03"


def test_fill_product_summary_total_uses_weighted_averages():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-11'] = pd.DataFrame([
        {
            'MonthStr': '2025-11',
            'ParsedDate': pd.Timestamp('2025-11-15'),
            '品目编码': '001',
            '品目名': 'A',
            '品目组合1名': '鞋类',
            '数量': 10,
            '合计': 1000,
        },
    ])
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            'MonthStr': '2025-12',
            'ParsedDate': pd.Timestamp('2025-12-15'),
            '品目编码': '001',
            '品目名': 'A',
            '品目组合1名': '鞋类',
            '数量': 10,
            '合计': 1000,
        },
        {
            'MonthStr': '2025-12',
            'ParsedDate': pd.Timestamp('2025-12-15'),
            '品目编码': '002',
            '品目名': 'B',
            '品目组合1名': '鞋类',
            '数量': 1,
            '合计': 100,
        },
    ])
    gen.data['cost']['2025-11'] = pd.DataFrame({
        '品目编码': ['001', '002'],
        '品目名规格': ['A规格', 'B规格'],
        '库存_期初': [100, 50],
        '库存_期初.2': [200, 100],
        '库存_期末': [120, 40],
        '库存_期末.2': [300, 100],
        '单价_减少.1': [50, 10],
    })
    gen.data['cost']['2025-12'] = pd.DataFrame({
        '品目编码': ['001', '002'],
        '品目名规格': ['A规格', 'B规格'],
        '库存_期初': [120, 40],
        '库存_期初.2': [200, 100],
        '库存_期末': [150, 35],
        '库存_期末.2': [400, 100],
        '单价_减少.1': [50, 10],
    })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '按产品汇总_含合计'
    headers = [
        '产品名称',
        '品目编码',
        '年销售数量合计',
        '年销售收入合计',
        '年销售成本合计',
        '年销售利润合计',
        '年初存货金额',
        '年末存货金额',
        '年平均存货',
        '存货周转率',
        '存货周转天数',
        '年销售数量平均',
        '年销售收入平均',
        '年销售成本平均',
        '年销售利润平均',
        '年毛利率平均',
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    ws.cell(row=2, column=2).value = '001'
    ws.cell(row=3, column=2).value = '002'
    ws.cell(row=4, column=1).value = '合计'

    gen._fill_product_summary_total(ws, '2025', '12', 'current')

    header_map = {
        str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value
    }
    total_row = next(
        r for r in range(2, ws.max_row + 1)
        if str(ws.cell(row=r, column=1).value).strip() == '合计'
    )

    # 加权口径（总额分母法）预期值：
    # 年销售数量合计=21，期内月数=2 => 年销售数量平均=10.5
    # 年销售收入合计=2100，年销售成本合计=1010，年销售利润合计=1090
    # 年初存货金额=300，年末存货金额=500 => 年平均存货=400
    # 存货周转率=1010/400=2.525，存货周转天数=365/2.525
    # 年毛利率平均=(1090/2)/(2100/2)=1090/2100
    assert total_row == 2
    assert abs(ws.cell(total_row, header_map['年销售数量平均']).value - 10.5) < 1e-12
    assert abs(ws.cell(total_row, header_map['年销售收入平均']).value - 1050) < 1e-12
    assert abs(ws.cell(total_row, header_map['年销售成本平均']).value - 505) < 1e-12
    assert abs(ws.cell(total_row, header_map['年销售利润平均']).value - 545) < 1e-12
    assert abs(ws.cell(total_row, header_map['年平均存货']).value - 400) < 1e-12
    assert abs(ws.cell(total_row, header_map['存货周转率']).value - 2.525) < 1e-12
    assert abs(ws.cell(total_row, header_map['存货周转天数']).value - (365 / 2.525)) < 1e-12
    assert abs(ws.cell(total_row, header_map['年毛利率平均']).value - (1090 / 2100)) < 1e-12
    assert abs(ws.cell(total_row, header_map['月销售数量平均']).value - 10.5) < 1e-12
    assert abs(ws.cell(total_row, header_map['月销售收入平均']).value - 1050) < 1e-12
    assert abs(ws.cell(total_row, header_map['月销售成本平均']).value - 505) < 1e-12
    assert abs(ws.cell(total_row, header_map['月销售利润平均']).value - 545) < 1e-12
    assert abs(ws.cell(total_row, header_map['月毛利率平均']).value - (1090 / 2100)) < 1e-12
    expected_after_margin = [
        '月销售数量平均',
        '月销售收入平均',
        '月销售成本平均',
        '月销售利润平均',
        '月毛利率平均',
    ]
    margin_col = header_map['年毛利率平均']
    actual_after_margin = [
        ws.cell(row=1, column=margin_col + offset).value
        for offset in range(1, len(expected_after_margin) + 1)
    ]
    assert actual_after_margin == expected_after_margin

    for header in ['年平均存货', '年销售数量平均', '年销售收入平均', '年销售成本平均', '年销售利润平均']:
        assert ws.column_dimensions[get_column_letter(header_map[header])].hidden is True
    assert ws.column_dimensions[get_column_letter(header_map['年毛利率平均'])].hidden is False
    for header in ['月销售数量平均', '月销售收入平均', '月销售成本平均', '月销售利润平均', '月毛利率平均']:
        assert ws.column_dimensions[get_column_letter(header_map[header])].hidden is False


def test_fill_product_summary_total_handles_total_marker_and_missing_parsed_date():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            'MonthStr': '2025-12',
            '品目编码': '001',
            '品目名': 'A',
            '品目组合1名': '鞋类',
            '数量': 10,
            '合计': 1000,
        },
    ])
    gen.data['cost']['2025-12'] = pd.DataFrame({
        '品目编码': ['001'],
        '品目名规格': ['A规格'],
        '库存_期初.2': [100],
        '库存_期末.2': [200],
        '单价_减少.1': [20],
    })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '按产品汇总_含合计'
    headers = [
        '产品名称',
        '品目编码',
        '年销售数量合计',
        '年销售收入合计',
        '年销售成本合计',
        '年销售利润合计',
        '年初存货金额',
        '年末存货金额',
        '年平均存货',
        '存货周转率',
        '存货周转天数',
        '年销售数量平均',
        '年销售收入平均',
        '年销售成本平均',
        '年销售利润平均',
        '年毛利率平均',
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    ws.cell(row=2, column=2).value = '001'
    ws.cell(row=3, column=2).value = '合计'
    ws.cell(row=3, column=10).value = 999999  # stale template value
    ws.cell(row=3, column=13).value = 999999  # stale template value

    gen._fill_product_summary_total(ws, '2025', '12', 'current')

    assert abs(ws.cell(row=3, column=13).value - 1000) < 1e-12
    assert abs(ws.cell(row=3, column=10).value - (200 / 150)) < 1e-12


def test_fill_product_summary_total_keeps_total_row_after_inserting_missing_codes():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            'MonthStr': '2025-12',
            '品目编码': '001',
            '品目名': 'A',
            '品目组合1名': '鞋类',
            '数量': 10,
            '合计': 1000,
        },
        {
            'MonthStr': '2025-12',
            '品目编码': '5501',
            '品目名': '5501',
            '品目组合1名': '配件',
            '数量': 1,
            '合计': 100,
        },
    ])
    gen.data['cost']['2025-12'] = pd.DataFrame({
        '品目编码': ['001', '5501'],
        '品目名规格': ['A规格', '5501'],
        '库存_期初.2': [100, 10],
        '库存_期末.2': [200, 20],
        '单价_减少.1': [20, 5],
    })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '按产品汇总_含合计'
    headers = [
        '产品名称',
        '品目编码',
        '年销售数量合计',
        '年销售收入合计',
        '年销售成本合计',
        '年销售利润合计',
        '年初存货金额',
        '年末存货金额',
        '年平均存货',
        '存货周转率',
        '存货周转天数',
        '年销售数量平均',
        '年销售收入平均',
        '年销售成本平均',
        '年销售利润平均',
        '年毛利率平均',
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    # 模板仅有产品001与合计，5501会被插入到合计行之前。
    ws.cell(row=2, column=2).value = '001'
    ws.cell(row=3, column=1).value = '合计'
    ws.cell(row=3, column=3).value = 999999  # stale template total to ensure overwritten

    gen._fill_product_summary_total(ws, '2025', '12', 'current')

    total_row = None
    row_5501 = None
    for r in range(2, ws.max_row + 1):
        first = ws.cell(row=r, column=1).value
        code = ws.cell(row=r, column=2).value
        if first is not None and str(first).strip() == '合计':
            total_row = r
        if code is not None and str(code).strip() == '5501':
            row_5501 = r

    assert total_row is not None
    assert row_5501 is not None
    assert total_row == 2
    assert row_5501 > total_row

    # 5501为自身数据，不应被覆盖为总计。
    assert abs(ws.cell(row=row_5501, column=3).value - 1) < 1e-12
    # 总计应为001+5501。
    assert abs(ws.cell(row=total_row, column=3).value - 11) < 1e-12


def test_fill_product_summary_total_current_inventory_matches_cost_ending_total():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            'MonthStr': '2025-12',
            '品目编码': '001',
            '品目名': 'A',
            '品目组合1名': '鞋类',
            '数量': 10,
            '合计': 1000,
        },
    ])
    gen.data['cost']['2025-12'] = pd.DataFrame({
        '品目编码': ['001', '002', '商品 合计', '累 计', '2026/05/13  21:37:23'],
        '品目名规格': ['A规格', 'B规格', None, None, None],
        '库存_期末': [10, 20, None, 30, None],
        '库存_期末.2': [200, 500, None, 700, None],
    })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '按产品汇总_含合计'
    headers = [
        '产品名称',
        '品目编码',
        '当前库存数量',
        '当前库存金额',
        '年末存货金额',
        '年销售收入合计',
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    ws.cell(row=2, column=2).value = '001'
    ws.cell(row=3, column=1).value = '合计'

    gen._fill_product_summary_total(ws, '2025', '12', 'current')

    code_rows = {
        str(ws.cell(row=r, column=2).value).strip(): r
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=2).value
    }
    total_row = next(
        r for r in range(2, ws.max_row + 1)
        if str(ws.cell(row=r, column=1).value).strip() == '合计'
    )

    assert '002' in code_rows
    assert '累 计' not in code_rows
    assert '商品 合计' not in code_rows
    assert ws.cell(row=code_rows['002'], column=1).value == 'B规格'
    assert ws.cell(row=code_rows['002'], column=4).value == 500
    assert ws.cell(row=total_row, column=3).value == 30
    assert ws.cell(row=total_row, column=4).value == 700
    assert ws.cell(row=total_row, column=5).value == 700


def test_build_cost_inventory_map_skips_cost_footer_rows():
    gen = ReportGenerator('.')
    cost_df = pd.DataFrame({
        '品目编码': ['001', '商品 合计', '累 计', '2026/05/13  21:37:23'],
        '品目名规格': ['A规格', None, None, None],
        '库存_期初': [8, None, 8, None],
        '库存_期初.2': [100, None, 100, None],
        '库存_期末': [10, None, 10, None],
        '库存_期末.2': [120, None, 120, None],
    })

    month_map = gen._build_cost_inventory_map(cost_df, '2025-12')

    assert list(month_map.keys()) == ['001']
    assert month_map['001']['qty_end'] == 10
    assert month_map['001']['amt_end'] == 120


def test_fill_product_summary_total_year_columns_ignore_prior_year_when_all_scope():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            'MonthStr': '2025-12',
            '品目编码': '001',
            '品目名': 'A',
            '品目组合1名': '鞋类',
            '数量': 1,
            '合计': 100,
        },
    ])
    gen.data['sales']['2026-01'] = pd.DataFrame([
        {
            'MonthStr': '2026-01',
            '品目编码': '001',
            '品目名': 'A',
            '品目组合1名': '鞋类',
            '数量': 3,
            '合计': 300,
        },
    ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '按产品汇总_含合计'
    headers = [
        '产品名称',
        '品目编码',
        '年销售数量合计',
        '年销售收入合计',
        '2026/01_销售数量',
        '2026/01_销售收入',
        '2025/12_销售数量',
        '2025/12_销售收入',
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    ws.cell(row=2, column=2).value = '001'
    ws.cell(row=3, column=1).value = '合计'

    gen._fill_product_summary_total(ws, '2026', '01', 'all')

    assert ws.cell(row=2, column=3).value == 3
    assert ws.cell(row=2, column=4).value == 300
    assert ws.cell(row=2, column=5).value == 3
    assert ws.cell(row=2, column=6).value == 300
    assert ws.cell(row=2, column=7).value == 1
    assert ws.cell(row=2, column=8).value == 100
    assert ws.cell(row=3, column=3).value == 3
    assert ws.cell(row=3, column=4).value == 300


def test_product_contribution_adds_inventory_risk_fields_and_charts():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            'MonthStr': '2025-12',
            '品目编码': '001',
            '品目名': 'A',
            '品目组合1名': '电器类',
            '数量': 2,
            '合计': 100,
        },
        {
            'MonthStr': '2025-12',
            '品目编码': '002',
            '品目名': 'B',
            '品目组合1名': '鞋类',
            '数量': 1,
            '合计': 50,
        },
    ])
    gen.data['cost']['2025-12'] = pd.DataFrame({
        '品目编码': ['001', '002'],
        '品目名': ['A', 'B'],
        '期末': [10, 20],
        '期末.2': [120, 500],
        '单价_减少.1': [40, 80],
    })

    wb = openpyxl.Workbook()
    wb.active.title = '目录'
    gen._update_product_contribution_sheet(wb, '2025', '12', 'current')

    ws = wb['产品贡献毛利']
    header_map = {
        str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value
    }
    assert '期末库存金额' in header_map
    assert '库存收入比' in header_map
    assert '风险标签' in header_map

    risk_by_code = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=header_map['品目编码']).value
        if code:
            risk_by_code[str(code)] = ws.cell(row=r, column=header_map['风险标签']).value
    assert risk_by_code['002'] == '负毛利'

    titles = {_chart_title_text(chart) for chart in ws._charts}
    assert any('低毛利' in title for title in titles)
    assert any('库存' in title for title in titles)
    assert any('收入 vs 毛利率' in title for title in titles)


def test_fill_expense_details_places_anomaly_section_below_main_table():
    gen = ReportGenerator('.')
    gen.data['expense']['2025-12'] = pd.DataFrame([
        {
            '日期': '2025-12-15',
            '科目名': '管理费用-办公费',
            '借方金额': 1200,
            '贷方金额': 0,
            '部门名': '行政',
            '摘要': '办公用品',
        },
    ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '费用明细'

    gen._fill_expense_details(ws, '2025', '12', 'current')

    assert ws.cell(row=1, column=1).value == '月份'
    # 主表不再并排写异常明细。
    assert ws.cell(row=1, column=9).value is None
    assert "费用异常明细" in wb.sheetnames

    detail_ws = wb["费用异常明细"]
    assert "异常项目明细" in str(detail_ws.cell(row=1, column=1).value)
    assert detail_ws.cell(row=2, column=1).value == '月份'
    # 单条样本通常不会触发异常评分，明细页应给出无可关联提示。
    assert detail_ws.cell(row=3, column=1).value == "无可关联的异常项目明细"


def test_add_chart_expense_detail_prefers_subcategory_dimension():
    gen = ReportGenerator('.')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '费用明细'
    headers = ['月份', '部门', '费用类别', '子科目', '摘要', '金额']
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h

    rows = [
        ['2025/12', '销售部', '销售费用', '工资', '工资发放', 100],
        ['2025/12', '销售部', '销售费用', '房租', '门店房租', 200],
        ['2025/12', '行政部', '管理费用', '工资', '行政工资', 50],
    ]
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = v

    start_col = ws.max_column + 2
    added = gen._add_chart_expense_detail(ws)
    assert added

    assert ws.cell(row=1, column=start_col).value == '子科目'
    assert ws.cell(row=2, column=start_col).value == '房租'
    assert abs(ws.cell(row=2, column=start_col + 1).value - 200) < 1e-12
    labels = {
        ws.cell(row=2, column=start_col).value,
        ws.cell(row=3, column=start_col).value,
    }
    assert labels == {'房租', '工资'}


def test_expense_analysis_generates_management_charts():
    gen = ReportGenerator('.')
    gen.data['expense']['2024-12'] = pd.DataFrame([
        {'MonthStr': '2024-12', '科目名': '销售费用-运费', '借方金额': 300, '贷方金额': 0, '部门名': '销售部', '摘要': '去年运输'},
        {'MonthStr': '2024-12', '科目名': '管理费用-房租', '借方金额': 70, '贷方金额': 0, '部门名': '行政部', '摘要': '去年房租'},
        {'MonthStr': '2024-12', '科目名': '财务费用-汇兑损益', '借方金额': 10, '贷方金额': 0, '部门名': '财务部', '摘要': '去年调汇'},
    ])
    gen.data['expense']['2025-11'] = pd.DataFrame([
        {'MonthStr': '2025-11', '科目名': '销售费用-运费', '借方金额': 100, '贷方金额': 0, '部门名': '销售部', '摘要': '运输'},
        {'MonthStr': '2025-11', '科目名': '管理费用-房租', '借方金额': 80, '贷方金额': 0, '部门名': '行政部', '摘要': '房租'},
        {'MonthStr': '2025-11', '科目名': '财务费用-汇兑损益', '借方金额': 20, '贷方金额': 0, '部门名': '财务部', '摘要': '调汇'},
    ])
    gen.data['expense']['2025-12'] = pd.DataFrame([
        {'MonthStr': '2025-12', '科目名': '销售费用-运费', '借方金额': 700, '贷方金额': 0, '部门名': '销售部', '摘要': '集中发货'},
        {'MonthStr': '2025-12', '科目名': '管理费用-房租', '借方金额': 90, '贷方金额': 0, '部门名': '行政部', '摘要': '房租'},
        {'MonthStr': '2025-12', '科目名': '财务费用-汇兑损益', '借方金额': 0, '贷方金额': 60, '部门名': '财务部', '摘要': '期末调汇'},
    ])

    wb = openpyxl.Workbook()
    wb.active.title = '目录'
    metrics = {
        '2025-11': {'revenue': 1000, 'sales_expense': 100, 'admin_expense': 80, 'financial_expense': 20},
        '2025-12': {'revenue': 1200, 'sales_expense': 700, 'admin_expense': 90, 'financial_expense': -60},
    }

    gen._update_expense_diagnostic_center(wb, metrics, '2025', '12', 'current')

    ws = wb['费用分析']
    titles = {_chart_title_text(chart) for chart in ws._charts}
    assert len(ws._charts) >= 7
    assert any('费用合计' in title for title in titles)
    assert any('费用结构' in title for title in titles)
    assert any('费用异常评分' in title for title in titles)
    assert any('费用科目环比变动Top' in title for title in titles)
    assert any('费用科目同比变动Top' in title for title in titles)
    assert any('部门费用' in title for title in titles)
    assert any('财务费用' in title for title in titles)


def test_ensure_report_charts_rebuilds_sales_inventory_chart_with_fallback_data():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '明细_销售与库存'

    headers = ['品目编码', '产品大类', '销售收入', '期末金额']
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h

    ws.cell(row=2, column=1).value = '001'
    ws.cell(row=2, column=2).value = '鞋类'
    ws.cell(row=2, column=4).value = 100
    ws.cell(row=3, column=1).value = '002'
    ws.cell(row=3, column=2).value = '鞋类'
    ws.cell(row=3, column=4).value = 50
    ws.cell(row=4, column=1).value = '003'
    ws.cell(row=4, column=2).value = '电器'
    ws.cell(row=4, column=4).value = 80

    # Simulate stale helper area + stale chart from previous generation.
    ws.cell(row=1, column=28).value = '品类'
    ws.cell(row=1, column=29).value = '金额'
    ws.cell(row=2, column=28).value = '旧品类'
    ws.cell(row=2, column=29).value = 999
    ws.cell(row=1, column=40).value = '旧说明'

    stale = BarChart()
    stale.add_data(Reference(ws, min_col=29, max_col=29, min_row=1, max_row=2), titles_from_data=True)
    stale.set_categories(Reference(ws, min_col=28, min_row=2, max_row=2))
    ws.add_chart(stale, 'F2')
    assert len(ws._charts) == 1

    gen._ensure_report_charts(wb)

    assert len(ws._charts) >= 4
    chart = ws._charts[0]
    series = chart.series[0]
    val_ref = series.val.numRef.f if series.val is not None and series.val.numRef is not None else None
    cat_ref = None
    if series.cat is not None:
        if series.cat.strRef is not None:
            cat_ref = series.cat.strRef.f
        elif series.cat.numRef is not None:
            cat_ref = series.cat.numRef.f

    assert val_ref == "'明细_销售与库存'!$AC$2:$AC$3"
    assert cat_ref == "'明细_销售与库存'!$AB$2:$AB$3"

    # Sales revenue column is empty, chart should fallback to ending inventory amount.
    assert ws.cell(row=2, column=28).value == '鞋类'
    assert abs(ws.cell(row=2, column=29).value - 150) < 1e-12
    assert ws.cell(row=3, column=28).value == '电器'
    assert abs(ws.cell(row=3, column=29).value - 80) < 1e-12
    titles = {_chart_title_text(chart) for chart in ws._charts}
    assert any('品类销售/库存/毛利对比' in title for title in titles)
    assert any('产品库存金额Top' in title for title in titles)
    assert any('高库存低销售风险Top' in title for title in titles)


def test_product_summary_supplements_existing_pareto_with_management_charts():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '按产品汇总_含合计'

    headers = [
        '产品名称',
        '当前库存金额',
        '年销售数量合计',
        '年销售收入合计',
        '年销售利润合计',
        '年毛利率平均',
        '存货周转天数',
        '2025/11_期末库存金额',
        '2025/11_销售收入',
        '2025/11_销售利润',
        '2025/12_期末库存金额',
        '2025/12_销售收入',
        '2025/12_销售利润',
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h

    rows = [
        ['A产品', 300, 10, 1000, 300, 0.30, 45, 250, 700, 200, 300, 1000, 300],
        ['B产品', 800, 5, 500, 20, 0.04, 220, 700, 400, 30, 800, 500, 20],
        ['C产品', 150, 8, 300, 120, 0.40, 30, 120, 200, 90, 150, 300, 120],
        ['合计', 1250, 23, 1800, 440, 0.2444, 90, 1070, 1300, 320, 1250, 1800, 440],
    ]
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = value

    existing = BarChart()
    existing.title = '产品销售帕累托分析 (ABC分析)'
    existing.add_data(Reference(ws, min_col=4, max_col=4, min_row=1, max_row=4), titles_from_data=True)
    existing.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
    ws.add_chart(existing, 'P2')

    gen._ensure_report_charts(wb)

    titles = {_chart_title_text(chart) for chart in ws._charts}
    assert any('产品销售帕累托分析' in title for title in titles)
    assert any('产品收入/毛利Top' in title for title in titles)
    assert any('产品库存占用与周转Top' in title for title in titles)
    assert any('低毛利产品Top' in title for title in titles)
    assert any('产品收入 vs 毛利率矩阵' in title for title in titles)
    assert any('产品销售/库存月度趋势' in title for title in titles)


def test_add_pareto_chart_uses_header_series_titles():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '客户贡献与回款'

    ws.cell(row=1, column=1).value = '客户'
    ws.cell(row=1, column=2).value = '销售收入'
    ws.cell(row=1, column=3).value = '累计占比'
    ws.cell(row=2, column=1).value = 'A客户'
    ws.cell(row=2, column=2).value = 100
    ws.cell(row=2, column=3).value = 0.5
    ws.cell(row=3, column=1).value = 'B客户'
    ws.cell(row=3, column=2).value = 60
    ws.cell(row=3, column=3).value = 0.8
    ws.cell(row=4, column=1).value = 'C客户'
    ws.cell(row=4, column=2).value = 40
    ws.cell(row=4, column=3).value = 1.0

    gen._add_pareto_chart(ws, 1, 2, 3, 1, 2, 4, '客户收入集中度 (Pareto)', 'E2')
    chart = ws._charts[-1]

    title_tokens = []
    for sub_chart in [chart] + list(getattr(chart, '_charts', [])):
        for series in sub_chart.series:
            token = _series_title_token(series)
            if token:
                title_tokens.append(token.replace('$', ''))

    assert any(token.endswith('!B1') for token in title_tokens)
    assert any(token.endswith('!C1') for token in title_tokens)
    assert all(token not in {'系列1', '系列2', 'Series1', 'Series2'} for token in title_tokens)


def test_add_scatter_chart_uses_y_header_as_series_title():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '销售人效分析'

    ws.cell(row=1, column=1).value = '人均收入'
    ws.cell(row=1, column=2).value = '利润率'
    ws.cell(row=2, column=1).value = 120
    ws.cell(row=2, column=2).value = 0.15
    ws.cell(row=3, column=1).value = 95
    ws.cell(row=3, column=2).value = 0.1

    gen._add_scatter_chart(ws, 1, 2, 1, 2, 3, '散点图', 'E2', x_title='人均收入', y_title='利润率')
    chart = ws._charts[-1]
    series_title = _series_title_token(chart.series[0])

    assert series_title == '利润率'
    assert series_title not in {'系列1', 'Series1'}


def test_add_doughnut_chart_uses_header_series_title():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '费用结构'

    ws.cell(row=1, column=1).value = '费用类型'
    ws.cell(row=1, column=2).value = '金额'
    ws.cell(row=2, column=1).value = '管理费用'
    ws.cell(row=2, column=2).value = 100
    ws.cell(row=3, column=1).value = '销售费用'
    ws.cell(row=3, column=2).value = 80

    gen._add_doughnut_chart(ws, 1, 2, 1, 2, 3, '费用构成', 'E2')
    chart = ws._charts[-1]
    series_title = _series_title_token(chart.series[0])

    assert series_title is not None
    assert series_title.replace('$', '').endswith('!B1')


def test_combo_and_stacked_charts_keep_month_categories():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '图表测试'
    headers = ['月份', '收入', '成本', '利润率', '短账龄', '长账龄']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    rows = [
        ['2026/04', 120, 80, 0.33, 70, 30],
        ['2026/03', 100, 60, 0.40, 65, 35],
    ]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    gen._add_combo_chart(ws, 1, [2, 3], [4], 1, 2, 3, '经营指标趋势', 'H2')
    gen._add_stacked_bar_chart(ws, 1, [5, 6], 1, 2, 3, '账龄结构', 'H18', percent=True)
    gen._add_bar_chart_by_columns(ws, 1, [2, 3], 1, 2, 3, '收入成本对比', 'H34')

    for chart in ws._charts:
        cat_formula = gen._chart_category_formula(chart)
        assert cat_formula is not None
        assert '$A$2:$A$3' in cat_formula


def test_repair_missing_chart_categories_infers_month_axis():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '经营指标'
    ws.cell(row=1, column=1).value = '月份'
    ws.cell(row=1, column=2).value = '收入'
    ws.cell(row=2, column=1).value = '2026/04'
    ws.cell(row=2, column=2).value = 120
    ws.cell(row=3, column=1).value = '2026/03'
    ws.cell(row=3, column=2).value = 100

    chart = BarChart()
    chart.title = '旧模板图表'
    chart.add_data(Reference(ws, min_col=2, max_col=2, min_row=1, max_row=3), titles_from_data=True)
    ws.add_chart(chart, 'E2')
    assert gen._chart_category_formula(ws._charts[0]) is None

    assert gen._repair_missing_chart_categories(wb) == 1
    assert '$A$2:$A$3' in gen._chart_category_formula(ws._charts[0])


def test_management_metric_chart_excludes_non_month_rows():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '经营指标'
    headers = ['月份', '部门', '主营业务收入', '主营业务成本', '营业利润', '净利润']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    rows = [
        ['2026/04', '合计', 120, 80, 30, 20],
        ['2026/03', '合计', 100, 70, 20, 10],
        ['合计', None, 220, 150, 50, 30],
    ]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    assert gen._add_chart_management_metrics(ws) is True
    assert '$A$2:$A$3' in gen._chart_category_formula(ws._charts[0])
    for series in ws._charts[0].series:
        assert '$4' not in series.val.numRef.f


def test_write_chart_note_normalizes_oversized_row_height():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.row_dimensions[5].height = 24.0
    ws.row_dimensions[6].height = 199.5
    ws.row_dimensions[7].height = 24.0

    gen._write_chart_note(ws, 1, 6, "图表说明：测试")

    assert abs(ws.row_dimensions[6].height - 24.0) < 1e-12


def test_append_chart_notes_below_keeps_note_row_height_and_disables_wrap():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '经营指标'

    ws.cell(row=1, column=1).value = '月份'
    ws.cell(row=1, column=2).value = '收入'
    ws.cell(row=2, column=1).value = '2025-11'
    ws.cell(row=2, column=2).value = 100
    ws.cell(row=3, column=1).value = '2025-12'
    ws.cell(row=3, column=2).value = 120

    chart = BarChart()
    chart.height = 10
    chart.width = 10
    data_ref = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=3)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, 'E2')

    _, _, _, row_end = gen._extract_chart_anchor_bbox(ws, chart)
    note_row = row_end + 2
    ws.row_dimensions[note_row - 1].height = 24.0
    ws.row_dimensions[note_row].height = 213.75
    ws.row_dimensions[note_row + 1].height = 24.0

    gen._append_chart_notes_below(wb, '2025', '12')

    note_cells = [
        cell for cell in ws._cells.values()
        if isinstance(cell.value, str) and '图表说明（2025年12月）' in cell.value
    ]
    assert note_cells
    note_cell = note_cells[0]
    assert note_cell.row == note_row
    assert abs(ws.row_dimensions[note_row].height - 24.0) < 1e-12
    assert note_cell.alignment is not None
    assert note_cell.alignment.wrap_text is False


def test_append_chart_notes_below_deduplicates_adjacent_same_notes():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '经营指标'

    ws.cell(row=1, column=1).value = '月份'
    ws.cell(row=1, column=2).value = '收入'
    ws.cell(row=2, column=1).value = '2025-11'
    ws.cell(row=2, column=2).value = 100
    ws.cell(row=3, column=1).value = '2025-12'
    ws.cell(row=3, column=2).value = 120

    for anchor in ('E2', 'E3'):
        chart = BarChart()
        chart.height = 10
        chart.width = 6
        data_ref = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=3)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=3)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, anchor)

    gen._append_chart_notes_below(wb, '2025', '12')

    notes = [
        cell for cell in ws._cells.values()
        if isinstance(cell.value, str) and '图表说明（2025年12月）' in cell.value
    ]
    assert len(notes) == 1


def test_data_quality_prefers_parsed_date_for_sales():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            '日期-号码': 'NO_DATE_TOKEN',
            'ParsedDate': pd.Timestamp('2025-12-15'),
            '品目编码': '001',
            '数量': 2,
            '销售金额合计': 100,
            '销售出库供应价合计': 60,
            '往来单位名': '客户A',
            '销售订单号': 'SO001',
        },
        {
            '日期-号码': 'INVALID',
            'ParsedDate': pd.Timestamp('2025-12-20'),
            '品目编码': '002',
            '数量': 3,
            '销售金额合计': 120,
            '销售出库供应价合计': 70,
            '往来单位名': '客户B',
            '销售订单号': 'SO002',
        },
    ])

    gen._run_data_quality_checks()
    sales_date_fail = [
        item for item in gen.data_quality_issues
        if item.get('category') == 'sales' and item.get('issue_type') == '日期解析失败'
    ]
    assert not sales_date_fail


def test_data_quality_skips_month_mismatch_for_multi_period_ledger():
    gen = ReportGenerator('.')
    gen.data['ar']['2026-12'] = pd.DataFrame([
        {'日期': '2025-01-10', '往来单位名': 'A', '借方金额': 1, '贷方金额': 0},
        {'日期': '2025-02-10', '往来单位名': 'B', '借方金额': 1, '贷方金额': 0},
        {'日期': '2025-03-10', '往来单位名': 'C', '借方金额': 1, '贷方金额': 0},
        {'日期': '2025-04-10', '往来单位名': 'D', '借方金额': 1, '贷方金额': 0},
    ])

    gen._run_data_quality_checks()
    ar_month_mismatch = [
        item for item in gen.data_quality_issues
        if item.get('category') == 'ar' and item.get('issue_type') == '月份不匹配'
    ]
    assert not ar_month_mismatch


def test_data_quality_sales_duplicate_order_is_info_not_warn():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            '日期': '2025-12-10',
            '品目编码': '001',
            '数量': 2,
            '销售金额合计': 100,
            '销售出库供应价合计': 60,
            '往来单位名': '客户A',
            '销售订单号': 'SO001',
        },
        {
            '日期': '2025-12-11',
            '品目编码': '002',
            '数量': 3,
            '销售金额合计': 120,
            '销售出库供应价合计': 70,
            '往来单位名': '客户A',
            '销售订单号': 'SO001',
        },
    ])

    gen._run_data_quality_checks()
    dup_items = [
        item for item in gen.data_quality_issues
        if item.get('category') == 'sales' and item.get('issue_type') == '单号重复'
    ]
    assert dup_items
    assert all((item.get('severity') or '').upper() == 'INFO' for item in dup_items)


def test_data_quality_summary_for_scope_excludes_out_of_scope_periods():
    gen = ReportGenerator('.')
    gen.data_quality_issues = [
        {'severity': 'ERROR', 'category': 'sales', 'period': '2026-02', 'issue_type': '客户/单位缺失', 'detail': 'x'},
        {'severity': 'WARN', 'category': 'sales', 'period': '2025-12', 'issue_type': '单价异常', 'detail': 'x'},
        {'severity': 'INFO', 'category': 'expense', 'period': '2025-11', 'issue_type': '金额/数量异常值', 'detail': 'x'},
    ]

    summary = gen._get_data_quality_summary_for_scope('2025', '12', 'current')
    assert summary['ERROR'] == 0
    assert summary['WARN'] == 1
    assert summary['INFO'] == 1
    assert summary['TOTAL'] == 2


def test_dashboard_template_formulas_use_prev_month_lookup_not_match_minus_one():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '仪表盘'

    # Trigger template-formula mode
    ws['A5'].value = '=OLD_FORMULA'
    ws['A7'].value = '=OLD_DELTA'
    ws['B3'].value = '2025/12'
    ws['A4'].value = '主营业务收入（元）'
    ws['E4'].value = '净利润（元）'
    ws['I4'].value = '净利润率'
    ws['M4'].value = '成本率'
    ws['Q4'].value = '营业利润（元）'
    ws['A1'].value = '较上一年：示例'

    gen._update_dashboard(wb, {}, '2025', '12')

    prev_lookup_token = 'MATCH(TEXT(DATE(LEFT($B$3,4),RIGHT($B$3,2),1)-1,"yyyy/mm")'

    assert prev_lookup_token in ws['A7'].value
    assert 'MATCH($B$3' in ws['A5'].value
    assert 'MATCH($B$3' in ws['E5'].value
    assert "'利润表'" in ws['E7'].value
    assert "'经营指标'!$G:$G" in ws['Q5'].value
    assert 'MATCH($B$3' in ws['Q5'].value
    assert prev_lookup_token in ws['I7'].value
    assert '*100' in ws['I7'].value
    assert '*100' in ws['M7'].value
    assert ws['A1'].value == '较上月：示例'


def test_ensure_month_columns_simple_backfills_intermediate_months():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '利润表'
    ws.cell(row=1, column=1).value = '指标'
    ws.cell(row=1, column=2).value = '全年汇总'
    ws.cell(row=1, column=3).value = '2025/11'
    ws.cell(row=1, column=4).value = '2025/10'
    ws.cell(row=1, column=5).value = '2025/09'

    changed = gen._ensure_month_columns_simple(ws, '2026', '2', header_row=1)

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert changed is True
    assert headers[:7] == ['指标', '全年汇总', '2026/02', '2026/01', '2025/12', '2025/11', '2025/10']


def test_ensure_month_rows_simple_backfills_intermediate_months():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '本量利分析'
    ws.cell(row=1, column=1).value = '月份'
    ws.cell(row=1, column=2).value = '部门'
    ws.cell(row=2, column=1).value = '2025/11'
    ws.cell(row=2, column=2).value = '合计'
    ws.cell(row=3, column=1).value = '2025/10'
    ws.cell(row=3, column=2).value = '合计'

    changed = gen._ensure_month_rows_simple(ws, '2026', '2', total_label='')
    gen._reorder_month_rows_desc(ws)

    labels = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert changed is True
    assert labels[:4] == ['2026/02', '2026/01', '2025/12', '2025/11']


def test_update_cvp_sheet_formats_margin_columns_and_blanks_near_zero_break_even():
    gen = ReportGenerator('.')
    gen._calculate_expense_keyword_totals = lambda *args: (
        {},
        {'2025-05': 100, '2025-06': 200},
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '本量利分析'
    ws.cell(row=1, column=1).value = '月份'
    ws.cell(row=2, column=1).value = '2025/05'
    ws.cell(row=3, column=1).value = '2025/06'
    ws.cell(row=4, column=1).value = '合计'

    metrics = {
        '2025-05': {'revenue': 1000, 'cost': 999.9},
        '2025-06': {'revenue': 1000, 'cost': 600},
    }
    gen._update_cvp_sheet(ws, metrics, '2025', '6', 'current')

    assert [ws.cell(row=r, column=1).value for r in range(2, 5)] == [
        '2025/06',
        '2025/05',
        '合计',
    ]
    assert abs(ws.cell(row=2, column=9).value - 500) < 1e-12
    assert abs(ws.cell(row=2, column=10).value - 500) < 1e-12
    assert abs(ws.cell(row=2, column=11).value - 0.5) < 1e-12

    assert ws.cell(row=3, column=9).value is None
    assert ws.cell(row=3, column=10).value is None
    assert ws.cell(row=3, column=11).value is None

    for row in (2, 3, 4):
        assert ws.cell(row=row, column=10).number_format == '#,##0.00'
        assert ws.cell(row=row, column=11).number_format == '0.0%'
    assert ws.column_dimensions['J'].width >= 14


def test_add_chart_cvp_excludes_total_row_and_sets_margin_categories():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '本量利分析'
    headers = [
        '月份',
        '合计',
        '销售收入',
        '变动成本',
        '贡献毛利',
        '贡献毛利率',
        '固定成本',
        '总成本',
        '盈亏平衡点',
        '安全边际',
        '安全边际率',
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    rows = [
        ['2026/02', '合计', 1200, 700, 500, 0.4167, 200, 900, 480, 720, 0.6],
        ['2026/01', '合计', 1000, 600, 400, 0.4, 180, 780, 450, 550, 0.55],
        ['合计', None, 2200, 1300, 900, 0.4091, 380, 1680, 929, 1271, 0.5777],
    ]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    added = gen._add_chart_cvp(ws)

    assert added is True
    assert len(ws._charts) == 3
    trend_chart, margin_chart, rate_chart = ws._charts
    assert '本量利金额趋势' in _chart_title_text(trend_chart)
    assert '盈亏平衡与安全边际' in _chart_title_text(margin_chart)
    assert '贡献毛利率/安全边际率趋势' in _chart_title_text(rate_chart)
    assert "$A$2:$A$3" in gen._chart_category_formula(trend_chart)
    assert "$A$4" not in gen._chart_category_formula(trend_chart)
    assert "$A$2:$A$3" in gen._chart_category_formula(margin_chart)
    assert "$A$2:$A$3" in gen._chart_category_formula(rate_chart)

    trend_value_ranges = [series.val.numRef.f for series in trend_chart.series]
    margin_value_ranges = [series.val.numRef.f for series in margin_chart.series]
    rate_value_ranges = [series.val.numRef.f for series in rate_chart.series]
    assert all("$4" not in formula for formula in trend_value_ranges)
    assert all("$4" not in formula for formula in margin_value_ranges)
    assert all("$4" not in formula for formula in rate_value_ranges)
    assert any("$J$2:$J$3" in formula for formula in margin_value_ranges)
    assert any("$K$2:$K$3" in formula for formula in rate_value_ranges)


def test_ensure_report_charts_rebuilds_stale_cvp_charts():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '本量利分析'
    headers = [
        '月份',
        '合计',
        '销售收入',
        '变动成本',
        '贡献毛利',
        '贡献毛利率',
        '固定成本',
        '总成本',
        '盈亏平衡点',
        '安全边际',
        '安全边际率',
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    rows = [
        ['2026/02', '合计', 1200, 700, 500, 0.4167, 200, 900, 480, 720, 0.6],
        ['2026/01', '合计', 1000, 600, 400, 0.4, 180, 780, 450, 550, 0.55],
        ['合计', None, 2200, 1300, 900, 0.4091, 380, 1680, 929, 1271, 0.5777],
    ]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    stale = BarChart()
    stale.title = '安全边际分析 (实际销售 vs 盈亏平衡)'
    stale.add_data(Reference(ws, min_col=3, max_col=3, min_row=1, max_row=4), titles_from_data=True)
    ws.add_chart(stale, 'M2')
    assert len(ws._charts) == 1

    gen._ensure_report_charts(wb)

    assert len(ws._charts) == 3
    titles = {_chart_title_text(chart) for chart in ws._charts}
    assert any('本量利金额趋势' in title for title in titles)
    assert any('盈亏平衡与安全边际' in title for title in titles)
    assert any('贡献毛利率/安全边际率趋势' in title for title in titles)
    for chart in ws._charts:
        assert "$A$2:$A$3" in gen._chart_category_formula(chart)
        for series in chart.series:
            assert "$4" not in series.val.numRef.f


def test_ensure_month_columns_grouped_by_suffix_backfills_intermediate_months():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '按品类汇总(按月)'
    headers = [
        '产品大类',
        '2025/11_毛利润',
        '2025/10_毛利润',
        '2025/11_销售成本',
        '2025/10_销售成本',
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    changed = gen._ensure_month_columns_grouped_by_suffix(ws, '2026', '2')

    out = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert changed is True
    assert '2026/02_毛利润' in out
    assert '2026/01_毛利润' in out
    assert '2025/12_毛利润' in out
    assert '2026/02_销售成本' in out
    assert '2026/01_销售成本' in out
    assert '2025/12_销售成本' in out


def test_load_ar_data_appends_multiple_files(tmp_path):
    gen = ReportGenerator('.')

    path1 = tmp_path / '2023-2025应收账款.xlsx'
    path2 = tmp_path / '2026应收账款.xlsx'

    df1 = pd.DataFrame({
        '日期': ['2025-12-15'],
        '往来单位名': ['客户A'],
        '借方金额': [100],
        '贷方金额': [0],
    })
    df2 = pd.DataFrame({
        '日期': ['2026-01-15'],
        '往来单位名': ['客户A'],
        '借方金额': [50],
        '贷方金额': [0],
    })

    with pd.ExcelWriter(path1) as writer:
        pd.DataFrame([['dummy']]).to_excel(writer, index=False, header=False)
        df1.to_excel(writer, index=False, startrow=1)
    with pd.ExcelWriter(path2) as writer:
        pd.DataFrame([['dummy']]).to_excel(writer, index=False, header=False)
        df2.to_excel(writer, index=False, startrow=1)

    gen._load_ar_data(str(path1), path1.name)
    gen._load_ar_data(str(path2), path2.name)

    assert set(gen.ar_detail_df['MonthStr'].unique()) == {'2025-12', '2026-01'}
    assert '2025-12' in gen.data['ar']
    assert '2026-01' in gen.data['ar']


def test_apply_current_scope_visibility_hides_out_of_scope_months():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '利润表'
    ws.cell(row=1, column=1).value = '指标'
    ws.cell(row=1, column=2).value = '2026/02'
    ws.cell(row=1, column=3).value = '2026/01'
    ws.cell(row=1, column=4).value = '2025/12'

    gen._apply_current_scope_visibility(wb, '2026', '2', 'current')

    assert ws.column_dimensions['B'].hidden is False
    assert ws.column_dimensions['C'].hidden is False
    assert ws.column_dimensions['D'].hidden is True


def test_apply_current_scope_visibility_hides_expense_compare_leading_gap():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '费用对比'
    ws.cell(row=1, column=1).value = '月份'
    ws.cell(row=1, column=2).value = '部门'
    ws.cell(row=5, column=1).value = '2026/02'
    ws.cell(row=5, column=2).value = '合计'

    gen._apply_current_scope_visibility(wb, '2026', '2', 'current')

    assert ws.row_dimensions[2].hidden is True
    assert ws.row_dimensions[3].hidden is True
    assert ws.row_dimensions[4].hidden is True
    assert ws.row_dimensions[5].hidden is False


def test_hide_leading_blank_rows_hides_gap_before_first_content():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = '月份'
    ws.cell(row=5, column=1).value = '2026/02'

    gen._hide_leading_blank_rows(ws, header_row=1)

    assert ws.row_dimensions[2].hidden is True
    assert ws.row_dimensions[3].hidden is True
    assert ws.row_dimensions[4].hidden is True


def test_hide_rows_before_first_month_hides_gap_before_month_data():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = '月份'
    ws.cell(row=2, column=2).value = '=IF(1=0,\"\",\"\")'
    ws.cell(row=5, column=1).value = '2026/02'

    gen._hide_rows_before_first_month(ws, header_keyword='月份', month_col=1)

    assert ws.row_dimensions[2].hidden is True
    assert ws.row_dimensions[3].hidden is True
    assert ws.row_dimensions[4].hidden is True


def test_validate_report_accepts_equivalent_b3_dropdown_formula():
    gen = ReportGenerator('.')

    wb = openpyxl.Workbook()

    ws_profit = wb.active
    ws_profit.title = '利润表'
    ws_profit['A1'].value = '项目'
    ws_profit['B1'].value = '2025/12'

    ws_asset = wb.create_sheet('资产负债表')
    ws_asset['A1'].value = '项目'
    ws_asset['B1'].value = '2025/12'

    ws_metrics = wb.create_sheet('经营指标')
    ws_metrics['A1'].value = '月份'
    ws_metrics['A2'].value = '2025/12'

    ws_dashboard = wb.create_sheet('仪表盘')
    ws_dashboard['B3'].value = '2025/12'
    dv = DataValidation(type='list', formula1='经营指标!$A$2:$A$2', allow_blank=True)
    ws_dashboard.add_data_validation(dv)
    dv.add('B3')

    tmp_path = '__tmp_validate_b3_equivalent.xlsx'
    wb.save(tmp_path)
    try:
        issues = gen.validate_report_file(tmp_path, '2025', '12', 'current')
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    assert not any('B3下拉范围未匹配' in str(item.get('message') or '') for item in issues)


def test_update_dashboard_controls_updates_unquoted_metric_chart_ranges():
    gen = ReportGenerator('.')

    wb = openpyxl.Workbook()
    ws_metric = wb.active
    ws_metric.title = '经营指标'
    ws_metric['A1'].value = '月份'
    ws_metric['A13'].value = '2026/02'
    ws_metric['A14'].value = '2026/01'
    ws_metric['C1'].value = '主营业务收入'
    ws_metric['C13'].value = 100
    ws_metric['C14'].value = 90

    ws_dash = wb.create_sheet('仪表盘')
    chart = openpyxl.chart.LineChart()
    data = Reference(ws_metric, min_col=3, min_row=1, max_row=12)
    cats = Reference(ws_metric, min_col=1, min_row=2, max_row=12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_dash.add_chart(chart, 'A1')

    gen._update_dashboard_controls(wb, '2026', '2', 'current')

    series = ws_dash._charts[0].series[0]
    assert series.val.numRef.f == '经营指标!$C$13:$C$14'
    assert series.cat.numRef.f == '经营指标!$A$13:$A$14'


def test_dashboard_includes_financial_expense_rate_in_table_and_core_chart():
    gen = ReportGenerator('.')

    wb = openpyxl.Workbook()
    ws_metric = wb.active
    ws_metric.title = '经营指标'
    metric_headers = {
        1: '月份',
        11: '成本率',
        12: '销售费用率',
        13: '管理费用率',
        14: '营业利润率',
        17: '财务费用率',
    }
    for col, header in metric_headers.items():
        ws_metric.cell(row=1, column=col).value = header
    ws_metric.cell(row=2, column=1).value = '2025/12'
    ws_metric.cell(row=3, column=1).value = '2025/11'
    for row, values in {
        2: {11: 0.60, 12: 0.10, 13: 0.05, 14: 0.20, 17: 0.03},
        3: {11: 0.55, 12: 0.11, 13: 0.04, 14: 0.18, 17: 0.02},
    }.items():
        for col, value in values.items():
            ws_metric.cell(row=row, column=col).value = value

    chart = openpyxl.chart.LineChart()
    chart.title = '核心费率/利润率趋势'
    chart.add_data(Reference(ws_metric, min_col=11, max_col=14, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(ws_metric, min_col=1, min_row=2, max_row=3))

    ws_dash = wb.create_sheet('仪表盘')
    ws_dash.add_chart(chart, 'A32')
    ws_dash.cell(row=34, column=1).value = '指标'
    labels = [
        '主营业务收入（元）',
        '净利润（元）',
        '净利润率',
        '成本率',
        '销售费用率',
        '管理费用率',
        '应收账款余额（元）',
        '存货期末余额（元）',
        '存货周转天数（天）',
    ]
    for idx, label in enumerate(labels, start=35):
        ws_dash.cell(row=idx, column=1).value = label

    ws_budget = wb.create_sheet('目标_预算')
    budget_headers = [
        '月份',
        '主营业务收入目标',
        '营业利润目标',
        '营业利润率目标',
        '成本率目标',
        '销售费用率目标',
        '管理费用率目标',
        '应收账款余额目标',
        '存货期末余额目标',
        '存货周转天数目标',
    ]
    for col, header in enumerate(budget_headers, start=1):
        ws_budget.cell(row=10, column=col).value = header
    ws_budget.cell(row=11, column=1).value = '2025/12'
    ws_budget.cell(row=12, column=1).value = '2025/11'

    gen._update_budget_sheet(
        ws_budget,
        {
            '2025-12': {'revenue': 100, 'cost': 60, 'sales_expense': 10, 'admin_expense': 5, 'financial_expense': 3, 'operating_profit': 20},
            '2025-11': {'revenue': 90, 'cost': 50, 'sales_expense': 9, 'admin_expense': 4, 'financial_expense': 2, 'operating_profit': 18},
        },
        '2025',
        '12',
        'current',
    )
    gen._update_dashboard_controls(wb, '2025', '12', 'current')
    gen._update_dashboard_controls(wb, '2025', '12', 'current')

    budget_header_map = {
        str(ws_budget.cell(row=10, column=c).value).strip(): c
        for c in range(1, ws_budget.max_column + 1)
        if ws_budget.cell(row=10, column=c).value
    }
    assert '财务费用率目标' in budget_header_map
    assert abs(ws_budget.cell(row=11, column=budget_header_map['财务费用率目标']).value - 0.03) < 1e-12

    financial_rows = [
        r for r in range(35, ws_dash.max_row + 1)
        if ws_dash.cell(row=r, column=1).value == '财务费用率'
    ]
    assert financial_rows == [41]
    finance_row = financial_rows[0]
    assert 'MATCH("财务费用率"' in ws_dash.cell(row=finance_row, column=2).value
    assert 'MATCH("财务费用率目标"' in ws_dash.cell(row=finance_row, column=3).value
    assert 'TEXT(DATE(LEFT($B$3,4),RIGHT($B$3,2),1)-1,"yyyy/mm")' in ws_dash.cell(row=finance_row, column=6).value

    core_chart = ws_dash._charts[0]
    value_formulas = [
        series.val.numRef.f
        for series in core_chart.series
        if series.val and series.val.numRef
    ]
    assert any('$Q$' in formula for formula in value_formulas)
    assert len([formula for formula in value_formulas if '$Q$' in formula]) == 1


def test_trim_chart_data_ranges_shrinks_to_active_rows():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '经营指标'
    ws['A1'].value = '月份'
    ws['C1'].value = '主营业务收入'
    ws['A2'].value = None
    ws['A3'].value = '2026/02'
    ws['A4'].value = '2026/01'
    ws['A5'].value = None
    ws['C3'].value = 100
    ws['C4'].value = 90

    chart = openpyxl.chart.LineChart()
    data = Reference(ws, min_col=3, min_row=1, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, 'E1')

    changed = gen._trim_single_chart_data_range(wb, ws._charts[0])

    series = ws._charts[0].series[0]
    assert changed is True
    assert series.val.numRef.f == "'经营指标'!$C$3:$C$4" or series.val.numRef.f == '经营指标!$C$3:$C$4'
    assert series.cat.numRef.f == "'经营指标'!$A$3:$A$4" or series.cat.numRef.f == '经营指标!$A$3:$A$4'


def test_history_scope_chart_month_categories_use_year_month_text():
    gen = ReportGenerator('.')
    gen.year_scope = 'all'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '经营指标'
    ws.append(['月份', '主营业务收入'])
    ws.append(['2025/12', 100])
    ws.append(['2026/01', 120])
    ws.append(['2026/02', 130])

    added = gen._add_line_chart_by_columns(ws, 1, [2], 1, 2, 4, '收入趋势', 'D1')

    assert added is True
    assert [ws.cell(row=r, column=1).value for r in range(2, 5)] == [
        '2025/12',
        '2026/01',
        '2026/02',
    ]
    parsed = gen._parse_chart_range_formula(gen._chart_category_formula(ws._charts[0]))
    assert parsed is not None
    _, cat_col, cat_start, _, cat_end = parsed
    assert cat_col != 1
    assert [ws.cell(row=r, column=cat_col).value for r in range(cat_start, cat_end + 1)] == [
        '2025.12月',
        '2026.1月',
        '2026.2月',
    ]
    assert gen._month_label_exact_to_key('2026.2月') == '2026-02'
    assert gen._label_to_month_key('2026.2月') == '2026-02'

    assert gen._reorder_month_rows_desc(ws) is True
    assert [ws.cell(row=r, column=1).value for r in range(2, 5)] == [
        '2026/02',
        '2026/01',
        '2025/12',
    ]
    assert [ws.cell(row=r, column=cat_col).value for r in range(2, 5)] == [
        '2026.2月',
        '2026.1月',
        '2025.12月',
    ]


def test_current_scope_chart_month_categories_keep_existing_label():
    gen = ReportGenerator('.')
    gen.year_scope = 'current'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['月份', '主营业务收入'])
    ws.append(['2026/01', 120])
    ws.append(['2026/02', 130])

    gen._add_line_chart_by_columns(ws, 1, [2], 1, 2, 3, '收入趋势', 'D1')

    assert [ws.cell(row=r, column=1).value for r in range(2, 4)] == ['2026/01', '2026/02']
    assert ws.max_column == 2


def test_existing_history_chart_categories_are_repointed_to_year_month_helper():
    gen = ReportGenerator('.')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '经营指标'
    ws.append(['月份', '主营业务收入'])
    ws.append(['2025/12', 100])
    ws.append(['2026/01', 120])

    chart = openpyxl.chart.LineChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, 'D1')

    changed = gen._rewrite_chart_month_categories_for_history(wb, 'all')

    assert changed == 1
    assert [ws.cell(row=r, column=1).value for r in range(2, 4)] == ['2025/12', '2026/01']
    parsed = gen._parse_chart_range_formula(gen._chart_category_formula(ws._charts[0]))
    assert parsed is not None
    _, cat_col, cat_start, _, cat_end = parsed
    assert cat_col != 1
    assert [ws.cell(row=r, column=cat_col).value for r in range(cat_start, cat_end + 1)] == [
        '2025.12月',
        '2026.1月',
    ]
    assert gen._rewrite_chart_month_categories_for_history(wb, 'all') == 0


def test_management_metrics_sheet_includes_yoy_and_mom_columns():
    gen = ReportGenerator('.')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '经营指标'
    headers = ['月份', '部门', '主营业务收入', '主营业务成本', '销售费用', '管理费用', '营业利润']
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h
    ws.cell(row=2, column=1).value = '2025/12'
    ws.cell(row=3, column=1).value = '2025/11'

    scoped = {
        '2025-11': {
            'revenue': 100,
            'cost': 60,
            'sales_expense': 10,
            'admin_expense': 5,
            'financial_expense': 2,
            'financial_expense_rate': 0.02,
            'operating_profit': 25,
        },
        '2025-12': {
            'revenue': 120,
            'cost': 70,
            'sales_expense': 12,
            'admin_expense': 6,
            'financial_expense': 3,
            'financial_expense_rate': 0.025,
            'operating_profit': 32,
        },
    }
    all_metrics = {
        **scoped,
        '2024-12': {
            'revenue': 80,
            'cost': 50,
            'sales_expense': 8,
            'admin_expense': 4,
            'financial_expense': 1,
            'operating_profit': 18,
        },
    }

    gen._update_management_metrics_sheet(
        ws,
        scoped,
        '2025',
        '12',
        'current',
        metrics_by_month_all=all_metrics,
    )

    header_map = {str(ws.cell(row=1, column=c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
    row_2025_12 = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value).strip() == '2025/12':
            row_2025_12 = r
            break
    assert row_2025_12 is not None

    assert abs(ws.cell(row_2025_12, column=header_map['主营业务收入_同比增量']).value - 40) < 1e-12
    assert abs(ws.cell(row_2025_12, column=header_map['主营业务收入_同比增速']).value - 0.5) < 1e-12
    assert abs(ws.cell(row_2025_12, column=header_map['主营业务收入_环比增量']).value - 20) < 1e-12
    assert abs(ws.cell(row_2025_12, column=header_map['主营业务收入_环比增速']).value - 0.2) < 1e-12
    assert ws.cell(row_2025_12, column=header_map['财务费用']).value == 3
    assert abs(ws.cell(row_2025_12, column=header_map['财务费用率']).value - 0.025) < 1e-12
    assert abs(ws.cell(row_2025_12, column=header_map['财务费用_同比增量']).value - 2) < 1e-12
    assert abs(ws.cell(row_2025_12, column=header_map['财务费用_环比增量']).value - 1) < 1e-12


def test_expense_compare_sheet_includes_financial_expense_columns_and_totals():
    gen = ReportGenerator('.')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '费用对比'
    headers = ['月份', '部门', '主营业务收入', '主营业务成本', '销售费用', '管理费用', '营业利润', '销售费用占比', '管理费用占比']
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h
    ws.cell(row=2, column=1).value = '2025/11'
    ws.cell(row=3, column=1).value = '2025/12'
    ws.cell(row=4, column=1).value = '合计'

    metrics = {
        '2025-11': {
            'revenue': 100,
            'cost': 60,
            'sales_expense': 10,
            'admin_expense': 5,
            'financial_expense': 2,
            'operating_profit': 23,
        },
        '2025-12': {
            'revenue': 200,
            'cost': 120,
            'sales_expense': 20,
            'admin_expense': 10,
            'financial_expense': 4,
            'operating_profit': 46,
        },
    }

    gen._update_expense_compare_sheet(ws, metrics, '2025', '12', 'current')

    header_map = {str(ws.cell(row=1, column=c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
    assert '财务费用' in header_map
    assert '财务费用占比' in header_map
    assert ws.cell(row=3, column=header_map['财务费用']).value == 4
    assert abs(ws.cell(row=3, column=header_map['财务费用占比']).value - 0.02) < 1e-12
    assert ws.cell(row=4, column=header_map['财务费用']).value == 6
    assert abs(ws.cell(row=4, column=header_map['财务费用占比']).value - 0.02) < 1e-12


def test_compare_sheet_adds_financial_expense_indicator_rows():
    gen = ReportGenerator('.')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '同比经营分析'
    headers = ['月份', '部门', '指标', '本期值', '同比增量', '同比增速', '环比增量', '环比增速']
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h
    ws.append(['2025/12', '合计', '销售费用', None, None, None, None, None])

    metrics = {
        '2025-12': {'financial_expense': 30},
        '2024-12': {'financial_expense': 10},
    }

    gen._update_compare_sheet(ws, metrics, 'yoy', '2025', '12', 'current')

    financial_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == '2025/12' and ws.cell(row=r, column=3).value == '财务费用':
            financial_row = r
            break

    assert financial_row is not None
    assert ws.cell(row=financial_row, column=4).value == 30
    assert ws.cell(row=financial_row, column=5).value == 20
    assert ws.cell(row=financial_row, column=6).value == 2


def test_annual_metrics_sheet_includes_financial_expense_and_rate():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    wb.active.title = '经营指标'

    metrics = {
        '2025-11': {
            'revenue': 100,
            'cost': 60,
            'financial_expense': 2,
            'financial_expense_rate': 0.02,
        },
        '2025-12': {
            'revenue': 200,
            'cost': 120,
            'financial_expense': 6,
            'financial_expense_rate': 0.03,
        },
    }

    gen._update_annual_metrics_sheet(wb, metrics, '2025')

    ws = wb['年度经营指标']
    header_map = {str(ws.cell(row=1, column=c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
    assert '财务费用' in header_map
    assert '财务费用率' in header_map
    total_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == '全年合计':
            total_row = r
            break
    assert total_row is not None
    assert ws.cell(row=total_row, column=header_map['财务费用']).value == 8
    assert abs(ws.cell(row=total_row, column=header_map['财务费用率']).value - (8 / 300)) < 1e-12


def test_category_month_sheet_labels_show_revenue_share_and_remain_idempotent():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            'MonthStr': '2025-12',
            'ParsedDate': pd.Timestamp('2025-12-15'),
            '品目编码': '001',
            '品目组合1名': '电器类',
            '数量': 3,
            '合计': 300,
        },
        {
            'MonthStr': '2025-12',
            'ParsedDate': pd.Timestamp('2025-12-20'),
            '品目编码': '002',
            '品目组合1名': '鞋类',
            '数量': 1,
            '合计': 100,
        },
    ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '按品类汇总(按月)'
    headers = ['产品大类', '年销售数量', '年销售收入', '年销售成本', '年毛利润', '2025/12_毛利润']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    ws.cell(row=2, column=1).value = '电器类'
    ws.cell(row=3, column=1).value = '鞋类'
    ws.cell(row=4, column=1).value = '合计'

    gen._update_category_month_sheet(ws, '2025', '12', 'current')

    assert ws.cell(row=2, column=1).value == '电器类'
    assert ws.cell(row=3, column=1).value == '电器类占比'
    assert ws.cell(row=4, column=1).value == '鞋类'
    assert ws.cell(row=5, column=1).value == '鞋类占比'
    assert abs(ws.cell(row=3, column=3).value - 0.75) < 1e-12
    assert abs(ws.cell(row=5, column=3).value - 0.25) < 1e-12

    # 再次执行不应叠加“占比”文本，且仍可正确匹配并回写。
    gen._update_category_month_sheet(ws, '2025', '12', 'current')
    assert ws.cell(row=2, column=1).value == '电器类'
    assert ws.cell(row=3, column=1).value == '电器类占比'
    assert ws.cell(row=4, column=1).value == '鞋类'
    assert ws.cell(row=5, column=1).value == '鞋类占比'


def test_category_month_sheet_year_columns_ignore_prior_year_when_all_scope():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            'MonthStr': '2025-12',
            'ParsedDate': pd.Timestamp('2025-12-15'),
            '品目编码': '001',
            '品目组合1名': '鞋类',
            '数量': 1,
            '合计': 100,
        },
    ])
    gen.data['sales']['2026-01'] = pd.DataFrame([
        {
            'MonthStr': '2026-01',
            'ParsedDate': pd.Timestamp('2026-01-15'),
            '品目编码': '001',
            '品目组合1名': '鞋类',
            '数量': 3,
            '合计': 300,
        },
    ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '按品类汇总(按月)'
    headers = [
        '产品大类',
        '年销售数量',
        '年销售收入',
        '2026/01_销售数量',
        '2026/01_销售收入',
        '2025/12_销售数量',
        '2025/12_销售收入',
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    ws.cell(row=2, column=1).value = '鞋类'
    ws.cell(row=3, column=1).value = '合计'

    gen._update_category_month_sheet(ws, '2026', '01', 'all')

    assert ws.cell(row=2, column=2).value == 3
    assert ws.cell(row=2, column=3).value == 300
    assert ws.cell(row=2, column=4).value == 3
    assert ws.cell(row=2, column=5).value == 300
    assert ws.cell(row=2, column=6).value == 1
    assert ws.cell(row=2, column=7).value == 100
    assert ws.cell(row=4, column=2).value == 3
    assert ws.cell(row=4, column=3).value == 300


def test_category_contribution_is_merged_and_old_sheet_redirects():
    gen = ReportGenerator('.')
    gen.data['sales']['2025-12'] = pd.DataFrame([
        {
            'MonthStr': '2025-12',
            '品目编码': '001',
            '品目名': 'A',
            '品目组合1名': '鞋类',
            '数量': 2,
            '合计': 100,
        },
        {
            'MonthStr': '2025-12',
            '品目编码': '002',
            '品目名': 'B',
            '品目组合1名': '电器类',
            '数量': 1,
            '合计': 80,
        },
    ])
    gen.data['cost']['2025-12'] = pd.DataFrame({
        '品目编码': ['001', '002'],
        '单价_减少.1': [30, 20],
    })

    wb = openpyxl.Workbook()
    ws_month = wb.active
    ws_month.title = '按品类汇总(按月)'
    ws_month.cell(row=1, column=1).value = '产品大类'
    ws_month.cell(row=1, column=2).value = '2025/12_毛利润'
    ws_month.cell(row=2, column=1).value = '鞋类'
    ws_month.cell(row=2, column=2).value = 40
    ws_month.cell(row=3, column=1).value = '电器类'
    ws_month.cell(row=3, column=2).value = 60
    wb.create_sheet('品类贡献毛利')

    gen._update_category_contribution_sheet(wb, '2025', '12', 'current')

    merge_col = None
    for c in range(1, ws_month.max_column + 1):
        if str(ws_month.cell(row=1, column=c).value).strip() == '品类贡献分析(合并视图)':
            merge_col = c
            break
    assert merge_col is not None
    assert ws_month.cell(row=4, column=merge_col).value == '品类'
    titles = {_chart_title_text(chart) for chart in ws_month._charts}
    assert any('品类毛利润Top' in title for title in titles)
    assert any('品类收入占比' in title for title in titles)
    gen._ensure_report_charts(wb)
    titles = {_chart_title_text(chart) for chart in ws_month._charts}
    assert any('品类毛利润趋势' in title for title in titles)


def test_delete_merged_sheets_keeps_expense_details():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    wb.active.title = '经营指标'
    for name in [
        '同比经营分析',
        '环比经营分析',
        '品类贡献毛利',
        '费用明细环比分析',
        '费用结构与弹性',
        '异常预警',
        '年度费用异常Top',
        '费用明细',
    ]:
        wb.create_sheet(name)

    gen._delete_sheets_if_exist(
        wb,
        [
            '同比经营分析',
            '环比经营分析',
            '品类贡献毛利',
            '费用明细环比分析',
            '费用结构与弹性',
            '异常预警',
            '年度费用异常Top',
        ],
    )

    for removed in [
        '同比经营分析',
        '环比经营分析',
        '品类贡献毛利',
        '费用明细环比分析',
        '费用结构与弹性',
        '异常预警',
        '年度费用异常Top',
    ]:
        assert removed not in wb.sheetnames
    assert '费用明细' in wb.sheetnames


def test_expense_diagnostic_center_generates_and_links_details():
    gen = ReportGenerator('.')
    gen.data['expense']['2025-11'] = pd.DataFrame([
        {
            '日期': '2025-11-15',
            '科目名': '管理费用-办公费',
            '借方金额': 100,
            '贷方金额': 0,
            '部门名': '行政',
            '摘要': '办公用品',
        },
    ])
    gen.data['expense']['2025-12'] = pd.DataFrame([
        {
            '日期': '2025-12-15',
            '科目名': '管理费用-办公费',
            '借方金额': 12000,
            '贷方金额': 0,
            '部门名': '行政',
            '摘要': '年末集中采购',
        },
    ])

    metrics = {
        '2025-11': {
            'revenue': 100000,
            'cost': 70000,
            'sales_expense': 3000,
            'admin_expense': 2000,
            'operating_profit': 15000,
            'inventory_start': 50000,
            'inventory_end': 52000,
            'cost_rate': 0.7,
            'ar_balance': 20000,
        },
        '2025-12': {
            'revenue': 105000,
            'cost': 73000,
            'sales_expense': 3200,
            'admin_expense': 12500,
            'operating_profit': 12000,
            'inventory_start': 52000,
            'inventory_end': 58000,
            'cost_rate': 73000 / 105000,
            'ar_balance': 26000,
        },
    }

    wb = openpyxl.Workbook()
    wb.active.title = '费用明细环比分析'
    gen._update_expense_diagnostic_center(wb, metrics, '2025', '12', 'current', anomaly_top_n=20, matrix_top_n=20, detail_lines_per_key=2)

    assert '费用分析' in wb.sheetnames
    ws = wb['费用分析']
    assert '费用分析' in str(ws.cell(row=1, column=1).value)

    has_anomaly_section = any(
        str(ws.cell(row=r, column=1).value).startswith('C. 异常Top')
        for r in range(1, min(ws.max_row, 200) + 1)
        if ws.cell(row=r, column=1).value is not None
    )
    assert has_anomaly_section

    has_internal_link = False
    for r in range(1, min(ws.max_row, 400) + 1):
        for c in range(1, min(ws.max_column, 40) + 1):
            link = ws.cell(row=r, column=c).hyperlink
            if link and link.location and "费用分析'!A" in link.location:
                has_internal_link = True
                break
        if has_internal_link:
            break
    assert has_internal_link


def test_drilldown_links_point_to_expense_diagnostic_center():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    wb.active.title = '经营指标'
    wb.create_sheet('费用对比')
    wb.create_sheet('费用分析')
    wb.create_sheet('利润表')

    gen._add_drilldown_links(wb)

    ws_metric = wb['经营指标']
    metric_links = [
        ws_metric.cell(row=r, column=ws_metric.max_column).hyperlink.location
        for r in range(1, ws_metric.max_row + 1)
        if ws_metric.cell(row=r, column=ws_metric.max_column).hyperlink is not None
    ]
    assert "'费用分析'!A1" in metric_links

    ws_exp = wb['费用对比']
    exp_links = [
        ws_exp.cell(row=r, column=ws_exp.max_column).hyperlink.location
        for r in range(1, ws_exp.max_row + 1)
        if ws_exp.cell(row=r, column=ws_exp.max_column).hyperlink is not None
    ]
    assert exp_links == ["'费用分析'!A1"]


def test_fill_profit_sheet_refreshes_annual_total_column():
    gen = ReportGenerator('.')
    gen.data['profit']['2025-11'] = pd.DataFrame([
        {'项目': '四、净利润', '2025/11': 100},
    ])
    gen.data['profit']['2025-12'] = pd.DataFrame([
        {'项目': '四、净利润', '2025/12': 200},
    ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '利润表'
    ws.cell(row=1, column=1).value = '指标'
    ws.cell(row=1, column=2).value = '全年汇总'
    ws.cell(row=1, column=3).value = '2025/12'
    ws.cell(row=1, column=4).value = '2025/11'
    ws.cell(row=2, column=1).value = '四、净利润'
    ws.cell(row=2, column=2).value = 999999  # stale template value

    gen._fill_profit_sheet(ws, '2025', '12', 'current')

    assert ws.cell(row=2, column=3).value == 200
    assert ws.cell(row=2, column=4).value == 100
    assert ws.cell(row=2, column=2).value == 300


def test_fill_profit_sheet_annual_total_uses_target_year_when_all_scope():
    gen = ReportGenerator('.')
    gen.data['profit']['2025-12'] = pd.DataFrame([
        {'项目': '四、净利润', '2025/12': 100},
    ])
    gen.data['profit']['2026-01'] = pd.DataFrame([
        {'项目': '四、净利润', '2026/01': 200},
    ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '利润表'
    ws.cell(row=1, column=1).value = '指标'
    ws.cell(row=1, column=2).value = '全年汇总'
    ws.cell(row=1, column=3).value = '2026/01'
    ws.cell(row=1, column=4).value = '2025/12'
    ws.cell(row=2, column=1).value = '四、净利润'
    ws.cell(row=2, column=2).value = 999999

    gen._fill_profit_sheet(ws, '2026', '01', 'all')

    assert ws.cell(row=2, column=3).value == 200
    assert ws.cell(row=2, column=4).value == 100
    assert ws.cell(row=2, column=2).value == 200


def test_profit_sheet_highlights_large_expense_mom_and_links_to_expense_details():
    gen = ReportGenerator('.')
    gen.data['expense']['2025-11'] = pd.DataFrame([
        {
            'MonthStr': '2025-11',
            '科目名': '管理费用-房租',
            '借方金额': 10000,
            '贷方金额': 0,
            '部门名': '行政',
            '摘要': '11月房租',
        },
    ])
    gen.data['expense']['2025-12'] = pd.DataFrame([
        {
            'MonthStr': '2025-12',
            '科目名': '管理费用-房租',
            '借方金额': 30000,
            '贷方金额': 0,
            '部门名': '行政',
            '摘要': '12月房租',
        },
    ])

    wb = openpyxl.Workbook()
    ws_profit = wb.active
    ws_profit.title = '利润表'
    ws_profit.cell(row=1, column=1).value = '项目'
    ws_profit.cell(row=1, column=2).value = '2025/11'
    ws_profit.cell(row=1, column=3).value = '2025/12'
    ws_profit.cell(row=2, column=1).value = '管理费用-房租'
    ws_profit.cell(row=2, column=2).value = 10000
    ws_profit.cell(row=2, column=3).value = 30000
    ws_profit.cell(row=3, column=1).value = '管理费用-办公费'
    ws_profit.cell(row=3, column=2).value = 5000
    ws_profit.cell(row=3, column=3).value = 5200

    ws_expense = wb.create_sheet('费用明细')
    headers = ["月份", "部门", "费用类别", "子科目", "摘要", "金额", "异常标签", "月份键"]
    for c, h in enumerate(headers, start=1):
        ws_expense.cell(row=1, column=c).value = h
    ws_expense.cell(row=2, column=1).value = '2025/11'
    ws_expense.cell(row=2, column=2).value = '行政'
    ws_expense.cell(row=2, column=3).value = '管理费用'
    ws_expense.cell(row=2, column=4).value = '房租'
    ws_expense.cell(row=2, column=8).value = '2025-11'
    ws_expense.cell(row=3, column=1).value = '2025/12'
    ws_expense.cell(row=3, column=2).value = '行政'
    ws_expense.cell(row=3, column=3).value = '管理费用'
    ws_expense.cell(row=3, column=4).value = '房租'
    ws_expense.cell(row=3, column=8).value = '2025-12'

    gen._highlight_profit_expense_anomalies(wb, '2025', '12', 'current')

    flagged = ws_profit.cell(row=2, column=3)
    assert flagged.hyperlink is not None
    assert "'费用明细'!A3" in str(flagged.hyperlink.location)
    assert flagged.font is not None
    assert flagged.font.color is not None
    assert (flagged.font.color.rgb or '').upper() in ('00FF0000', 'FFFF0000')

    # Non-anomalous row should remain without hyperlink.
    assert ws_profit.cell(row=3, column=3).hyperlink is None


def test_update_directory_sheet_realigns_links_and_clears_missing_targets():
    gen = ReportGenerator('.')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '目录'
    ws.cell(row=1, column=1).value = '目标_预算'
    ws.cell(row=1, column=1).hyperlink = "#'明细_销售与库存'!A1"  # stale wrong link
    ws.cell(row=2, column=1).value = '同比经营分析'  # missing sheet
    ws.cell(row=2, column=1).hyperlink = "#'同比经营分析'!A1"
    wb.create_sheet('目标_预算')
    wb.create_sheet('明细_销售与库存')

    gen._update_directory_sheet(wb)

    assert ws.cell(row=1, column=1).hyperlink is not None
    assert ws.cell(row=1, column=1).hyperlink.location == "'目标_预算'!A1"
    assert ws.cell(row=2, column=1).hyperlink is None


if __name__ == '__main__':
    test_fill_product_summary_aggregates_rows()
    test_list_available_months_uses_core_intersection_when_loaded()
    test_check_data_completeness_includes_sales_and_ar()
    test_update_ap_and_cash_analysis_sheets()
    test_expense_behavior_totals_are_closed_with_unclassified_and_finance()
    test_expense_diagnostic_matrix_does_not_reuse_history_anomaly_for_target_month()
    test_write_table_sanitizes_formula_like_text()
    test_fill_product_summary_total_uses_weighted_averages()
    test_fill_product_summary_total_handles_total_marker_and_missing_parsed_date()
    test_fill_product_summary_total_keeps_total_row_after_inserting_missing_codes()
    test_product_contribution_adds_inventory_risk_fields_and_charts()
    test_fill_expense_details_places_anomaly_section_below_main_table()
    test_add_chart_expense_detail_prefers_subcategory_dimension()
    test_expense_analysis_generates_management_charts()
    test_ensure_report_charts_rebuilds_sales_inventory_chart_with_fallback_data()
    test_add_pareto_chart_uses_header_series_titles()
    test_add_scatter_chart_uses_y_header_as_series_title()
    test_add_doughnut_chart_uses_header_series_title()
    test_combo_and_stacked_charts_keep_month_categories()
    test_repair_missing_chart_categories_infers_month_axis()
    test_management_metric_chart_excludes_non_month_rows()
    test_write_chart_note_normalizes_oversized_row_height()
    test_append_chart_notes_below_keeps_note_row_height_and_disables_wrap()
    test_append_chart_notes_below_deduplicates_adjacent_same_notes()
    test_data_quality_prefers_parsed_date_for_sales()
    test_data_quality_skips_month_mismatch_for_multi_period_ledger()
    test_data_quality_sales_duplicate_order_is_info_not_warn()
    test_data_quality_summary_for_scope_excludes_out_of_scope_periods()
    test_dashboard_template_formulas_use_prev_month_lookup_not_match_minus_one()
    test_validate_report_accepts_equivalent_b3_dropdown_formula()
    test_dashboard_includes_financial_expense_rate_in_table_and_core_chart()
    test_update_cvp_sheet_formats_margin_columns_and_blanks_near_zero_break_even()
    test_add_chart_cvp_excludes_total_row_and_sets_margin_categories()
    test_ensure_report_charts_rebuilds_stale_cvp_charts()
    test_management_metrics_sheet_includes_yoy_and_mom_columns()
    test_category_month_sheet_labels_show_revenue_share_and_remain_idempotent()
    test_category_contribution_is_merged_and_old_sheet_redirects()
    test_delete_merged_sheets_keeps_expense_details()
    test_expense_diagnostic_center_generates_and_links_details()
    test_drilldown_links_point_to_expense_diagnostic_center()
    test_fill_profit_sheet_refreshes_annual_total_column()
    test_profit_sheet_highlights_large_expense_mom_and_links_to_expense_details()
    test_update_directory_sheet_realigns_links_and_clears_missing_targets()
    print('PASS: test_report_generator_repairs')
